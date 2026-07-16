import json
import re
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from django.contrib.staticfiles import finders
from django.test import Client, TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from PIL import Image

from references.models import (
    Dormitory,
    DormitoryBlock,
    DormitorySection,
    Equipment,
    EquipmentModel,
    EquipmentType,
)
from shifts.models import EmployeeShift
from users.models import DriverPrimaryRegistration, Employee, EmployeeAccess, Role

from .models import AssignmentStatus, CrewPlanSlot, CrewPlanStatus, EquipmentAssignment, WorkShiftType
from .services import get_active_equipment_assignment, get_or_create_crew_draft, set_active_equipment_assignment


class DeputyPlanningViewTests(TestCase):
    def setUp(self):
        self.deputy_role, _created = Role.objects.update_or_create(
            code='deputy_mining_manager',
            defaults={
                'name': 'Зам. начальника горного участка',
                'is_active': True,
            },
        )
        self.driver_role = Role.objects.create(code='driver', name='Водитель самосвала')
        self.excavator_role = Role.objects.create(
            code='excavator_operator',
            name='Машинист экскаватора',
        )
        self.deputy, self.deputy_access = self.create_employee_with_access(
            'Заместитель начальника участка',
            self.deputy_role,
            phone='+79000000010',
            access_code='610001',
        )
        self.driver, self.driver_access = self.create_employee_with_access(
            'Иванов Сергей Петрович',
            self.driver_role,
            phone='+79000000001',
            access_code='210001',
        )

        self.truck_type = EquipmentType.objects.create(name='Самосвал')
        self.truck_model = EquipmentModel.objects.create(
            equipment_type=self.truck_type,
            name='БелАЗ тестовый',
            fuel_capacity_limit_l='2000',
        )
        self.truck_1 = Equipment.objects.create(
            equipment_type=self.truck_type,
            model=self.truck_model,
            garage_number='Т-01',
        )
        self.truck_2 = Equipment.objects.create(
            equipment_type=self.truck_type,
            model=self.truck_model,
            garage_number='Т-02',
        )
        self.original_assignment, _created = set_active_equipment_assignment(
            employee=self.driver,
            role=self.driver_role,
            equipment=self.truck_1,
            shift_type=WorkShiftType.SHIFT_1,
            assigned_by=self.deputy,
        )

        dormitory = Dormitory.objects.create(number='Тест')
        block = DormitoryBlock.objects.create(dormitory=dormitory, name='Блок')
        section = DormitorySection.objects.create(block=block, name='Секция')
        DriverPrimaryRegistration.objects.create(
            employee=self.driver,
            dormitory_section=section,
        )

        self.authenticate(self.client, self.deputy_access)

    def create_employee_with_access(self, full_name, role, *, phone, access_code):
        employee = Employee.objects.create(
            full_name=full_name,
            phone=phone,
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        access = EmployeeAccess.objects.create(
            employee=employee,
            role=role,
            access_code=access_code,
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        return employee, access

    def authenticate(self, client, access):
        session = client.session
        session['employee_access_id'] = access.id
        session.save()

    def create_draft(self):
        plan, _created = get_or_create_crew_draft(
            role=self.driver_role,
            actor=self.deputy,
        )
        return plan

    def autosave_slot(self, plan, *, equipment, shift_type, employee):
        response = self.client.post(
            reverse('deputy_mining_manager_slot'),
            data=json.dumps({
                'plan_id': plan.id,
                'expected_version': plan.version,
                'equipment_id': equipment.id,
                'shift_type': shift_type,
                'employee_id': employee.id if employee else None,
            }),
            content_type='application/json',
            HTTP_HOST='localhost',
        )
        self.assertEqual(response.status_code, 200, response.content)
        plan.refresh_from_db()
        return response, plan

    def autosave_driver_on_second_truck(self, plan):
        return self.autosave_slot(
            plan,
            equipment=self.truck_2,
            shift_type=WorkShiftType.SHIFT_1,
            employee=self.driver,
        )

    def publish(self, plan):
        response = self.client.post(
            reverse('deputy_mining_manager_publish'),
            data=json.dumps({
                'plan_id': plan.id,
                'expected_version': plan.version,
            }),
            content_type='application/json',
            HTTP_HOST='localhost',
        )
        self.assertEqual(response.status_code, 200, response.content)
        return response

    def test_profile_is_available_only_to_deputy_and_role_home_routes_to_board(self):
        response = self.client.get(reverse('role_home'), HTTP_HOST='localhost')
        self.assertRedirects(
            response,
            reverse('deputy_mining_manager_placement'),
            fetch_redirect_response=False,
        )

        driver_client = Client()
        self.authenticate(driver_client, self.driver_access)
        for view_name in (
            'deputy_mining_manager_placement',
            'deputy_mining_manager_reports',
        ):
            forbidden_page = driver_client.get(reverse(view_name), HTTP_HOST='localhost')
            self.assertRedirects(
                forbidden_page,
                reverse('role_home'),
                fetch_redirect_response=False,
            )

        plan = self.create_draft()
        forbidden_slot = driver_client.post(
            reverse('deputy_mining_manager_slot'),
            data=json.dumps({
                'plan_id': plan.id,
                'expected_version': plan.version,
                'equipment_id': self.truck_2.id,
                'shift_type': WorkShiftType.SHIFT_1,
                'employee_id': self.driver.id,
            }),
            content_type='application/json',
            HTTP_HOST='localhost',
        )
        forbidden_publish = driver_client.post(
            reverse('deputy_mining_manager_publish'),
            data=json.dumps({'plan_id': plan.id, 'expected_version': plan.version}),
            content_type='application/json',
            HTTP_HOST='localhost',
        )
        self.assertEqual(forbidden_slot.status_code, 403)
        self.assertEqual(forbidden_publish.status_code, 403)

    def test_deputy_pages_expose_shared_pwa_metadata(self):
        for view_name in (
            'deputy_mining_manager_placement',
            'deputy_mining_manager_reports',
        ):
            response = self.client.get(reverse(view_name), HTTP_HOST='localhost')

            self.assertEqual(response.status_code, 200)
            self.assertContains(response, reverse('deputy_mining_manager_manifest'))
            self.assertContains(response, 'rel="manifest"')
            self.assertContains(response, 'name="theme-color" content="#2E7D52"')
            self.assertContains(response, 'name="apple-mobile-web-app-capable" content="yes"')
            self.assertContains(response, 'deputy-mining-manager-180.png')
            self.assertContains(response, 'deputy-mining-manager-pwa-v1.js')

    def test_deputy_manifest_is_installable_landscape_pwa(self):
        response = Client().get(
            reverse('deputy_mining_manager_manifest'),
            HTTP_HOST='localhost',
        )
        manifest = json.loads(response.content.decode('utf-8'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/manifest+json; charset=utf-8')
        self.assertEqual(response['Cache-Control'], 'no-cache')
        self.assertEqual(response['X-Content-Type-Options'], 'nosniff')
        self.assertEqual(manifest['id'], reverse('deputy_mining_manager_placement'))
        self.assertEqual(manifest['start_url'], reverse('deputy_mining_manager_placement'))
        self.assertEqual(manifest['scope'], '/deputy-mining-manager/')
        self.assertEqual(manifest['display'], 'standalone')
        self.assertEqual(manifest['orientation'], 'landscape')
        self.assertTrue(any(icon.get('sizes') == '192x192' for icon in manifest['icons']))
        self.assertTrue(any(icon.get('sizes') == '512x512' for icon in manifest['icons']))
        self.assertTrue(any(icon.get('purpose') == 'maskable' for icon in manifest['icons']))
        self.assertEqual(
            {shortcut['url'] for shortcut in manifest['shortcuts']},
            {
                reverse('deputy_mining_manager_placement'),
                reverse('deputy_mining_manager_reports'),
            },
        )

    def test_deputy_pwa_icons_have_declared_dimensions(self):
        icon_root = Path(__file__).resolve().parents[1] / 'static' / 'img' / 'pwa'
        expected_sizes = {
            'deputy-mining-manager-180.png': (180, 180),
            'deputy-mining-manager-192.png': (192, 192),
            'deputy-mining-manager-512.png': (512, 512),
            'deputy-mining-manager-maskable-512.png': (512, 512),
        }

        for filename, expected_size in expected_sizes.items():
            with self.subTest(filename=filename):
                with Image.open(icon_root / filename) as icon:
                    self.assertEqual(icon.size, expected_size)
                    self.assertEqual(icon.format, 'PNG')

        manifest_response = Client().get(
            reverse('deputy_mining_manager_manifest'),
            HTTP_HOST='localhost',
        )
        manifest = json.loads(manifest_response.content.decode('utf-8'))
        for icon in manifest['icons']:
            with self.subTest(manifest_icon=icon['src']):
                static_path = icon['src'].removeprefix('/static/')
                resolved_path = finders.find(static_path)
                self.assertIsNotNone(resolved_path)
                declared_size = tuple(int(value) for value in icon['sizes'].split('x'))
                with Image.open(resolved_path) as image:
                    self.assertEqual(image.size, declared_size)

    def test_deputy_service_worker_keeps_planning_data_network_only(self):
        response = Client().get(
            reverse('deputy_mining_manager_service_worker'),
            HTTP_HOST='localhost',
        )
        script = response.content.decode('utf-8')
        core_assets = script.split('const CORE_ASSETS = [', 1)[1].split('];', 1)[0]

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/javascript; charset=utf-8')
        self.assertEqual(response['Cache-Control'], 'no-cache')
        self.assertEqual(response['Service-Worker-Allowed'], '/deputy-mining-manager/')
        self.assertEqual(response['X-Content-Type-Options'], 'nosniff')
        self.assertIn('deputy-mining-manager-desktop-shell-', script)
        self.assertIn('`${CACHE_PREFIX}v5`', script)
        self.assertIn('key.startsWith(CACHE_PREFIX) && key !== CACHE_NAME', script)
        self.assertIn('removeCachedPlanningDocuments()', script)
        self.assertIn('LEGACY_ROOT_FALLBACK_URL', script)
        self.assertIn('cacheName.startsWith(LEGACY_CACHE_PREFIX)', script)
        self.assertIn('cachedHtml.includes("data-deputy-planning-root")', script)
        self.assertIn('if (request.method !== "GET") return;', script)
        self.assertIn('networkOnlyNavigation(request)', script)
        self.assertIn('STATIC_ASSET_PATHS.has(url.pathname)', script)
        self.assertIn('networkFirstStatic(request)', script)
        self.assertIn('const cacheKey = canonicalStaticRequest(request);', script)
        self.assertIn('fetch(request, { cache: "no-store" })', script)
        self.assertNotIn('cacheFirst(request)', script)
        self.assertIn('Для работы с расстановкой требуется сеть.', script)
        self.assertIn('SKIP_WAITING', script)
        self.assertNotIn('GET_VERSION', script)
        self.assertIn('/static/css/deputy-mining-manager-v3.css', core_assets)
        self.assertIn('/static/js/deputy-mining-manager-v3.js', core_assets)
        self.assertIn('/static/js/deputy-mining-manager-pwa-v1.js', core_assets)
        self.assertIn('/static/img/equipment/truck-green.png', core_assets)
        self.assertIn('/static/img/equipment/excavator-green.png', core_assets)
        self.assertNotIn('"/deputy-mining-manager/"', core_assets)
        self.assertNotIn('/reports/', core_assets)
        self.assertNotIn('/slot/', core_assets)
        self.assertNotIn('/publish/', core_assets)
        self.assertNotIn('/media/', core_assets)

        for static_url in re.findall(r'"(/static/[^"?]+)"', core_assets):
            with self.subTest(core_asset=static_url):
                self.assertIsNotNone(finders.find(static_url.removeprefix('/static/')))

    def test_deputy_pwa_registration_updates_without_forced_page_reload(self):
        registration_script = (
            Path(__file__).resolve().parents[1]
            / 'static'
            / 'js'
            / 'deputy-mining-manager-pwa-v1.js'
        ).read_text(encoding='utf-8')

        self.assertIn('navigator.serviceWorker.register(WORKER_URL, { scope: PWA_SCOPE })', registration_script)
        self.assertIn(
            f'var WORKER_URL = "{reverse("deputy_mining_manager_service_worker")}";',
            registration_script,
        )
        self.assertIn(
            f'var APP_PATH_PREFIX = "{reverse("deputy_mining_manager_placement")}";',
            registration_script,
        )
        self.assertIn('meta[name="role-app-scope"]', registration_script)
        self.assertIn('registrationRef.update()', registration_script)
        self.assertIn('visibilitychange', registration_script)
        self.assertIn('window.addEventListener("focus"', registration_script)
        self.assertIn('window.addEventListener("online"', registration_script)
        self.assertIn('SKIP_WAITING', registration_script)
        self.assertIn('removeCachedPlanningDocuments()', registration_script)
        self.assertNotIn('/mining-master/assignments/', registration_script)
        self.assertNotIn('controllerchange', registration_script)
        self.assertNotIn('location.reload', registration_script)

    def test_deputy_layout_uses_full_scaled_viewport_and_balanced_columns(self):
        stylesheet = (
            Path(__file__).resolve().parents[1]
            / 'static'
            / 'css'
            / 'deputy-mining-manager-v3.css'
        ).read_text(encoding='utf-8')

        self.assertIn('--admin-console-header-height: 76px;', stylesheet)
        self.assertIn(
            'height: calc((100dvh / var(--admin-interface-scale)) - var(--admin-console-header-height));',
            stylesheet,
        )
        self.assertIn(
            'grid-template-columns: clamp(360px, 21vw, 400px) minmax(0, 1fr);',
            stylesheet,
        )
        self.assertRegex(
            stylesheet,
            r'\.deputy-assignment-table th:first-child\s*\{\s*width:\s*24%;',
        )
        self.assertRegex(
            stylesheet,
            r'\.deputy-assignment-table th:not\(:first-child\)\s*\{\s*width:\s*38%;',
        )

    def test_get_board_builds_driver_plan_with_day_and_night_slots(self):
        response = self.client.get(
            reverse('deputy_mining_manager_placement'),
            HTTP_HOST='localhost',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'assignments/deputy_mining_manager_placement.html')
        payload = response.context['planning_payload']
        self.assertEqual(payload['role']['code'], 'driver')
        self.assertEqual(payload['summary']['equipment_total'], 2)
        self.assertEqual(payload['summary']['assigned_count'], 1)
        self.assertEqual(payload['summary']['unfilled_count'], 3)
        self.assertEqual(len(payload['rows']), 2)
        self.assertTrue(all(len(row['slots']) == 2 for row in payload['rows']))
        self.assertEqual(
            payload['endpoints']['export'],
            reverse('deputy_mining_manager_export', args=[payload['plan']['id']]),
        )
        self.assertContains(response, 'data-export-excel', count=1)

    def test_board_and_excel_use_natural_equipment_number_order(self):
        self.truck_1.garage_number = '10'
        self.truck_1.save(update_fields=['garage_number'])
        self.truck_2.garage_number = '2'
        self.truck_2.save(update_fields=['garage_number'])

        response = self.client.get(
            reverse('deputy_mining_manager_placement'),
            HTTP_HOST='localhost',
        )
        payload = response.context['planning_payload']
        self.assertEqual(
            [row['equipment']['label'] for row in payload['rows']],
            ['2', '10'],
        )

        export_response = self.client.get(
            payload['endpoints']['export'],
            HTTP_HOST='localhost',
        )
        sheet = load_workbook(BytesIO(export_response.content))['Расстановка']
        self.assertEqual(
            [sheet.cell(row_index, 1).value for row_index in range(8, sheet.max_row + 1)],
            ['2', '10'],
        )

    def test_excel_export_contains_current_draft_and_is_ready_for_print(self):
        self.driver.personnel_number = 'Т-001'
        self.driver.rotation = 'Бригада 1'
        self.driver.save(update_fields=['personnel_number', 'rotation'])
        plan = self.create_draft()
        _response, plan = self.autosave_slot(
            plan,
            equipment=self.truck_2,
            shift_type=WorkShiftType.SHIFT_2,
            employee=self.driver,
        )

        response = self.client.get(
            reverse('deputy_mining_manager_export', args=[plan.id]),
            HTTP_HOST='localhost',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('attachment; filename="crew_plan_driver_', response['Content-Disposition'])
        self.assertIn("filename*=UTF-8''", response['Content-Disposition'])
        self.assertEqual(response['Cache-Control'], 'private, no-store, max-age=0')
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook['Расстановка']
        self.assertEqual(sheet['A1'].value, 'COPPER RESOURCES · РАССТАНОВКА ПО ТЕХНИКЕ')
        self.assertIn('Статус: Черновик', sheet['C3'].value)
        rows_by_equipment = {
            sheet.cell(row_index, 1).value: row_index
            for row_index in range(8, sheet.max_row + 1)
        }
        truck_2_row = rows_by_equipment['Т-02']
        self.assertEqual(sheet.cell(truck_2_row, 3).value, 'Не назначен')
        self.assertIn('Иванов Сергей Петрович', sheet.cell(truck_2_row, 4).value)
        self.assertIn('Таб. № Т-001', sheet.cell(truck_2_row, 4).value)
        self.assertIn('Бригада 1', sheet.cell(truck_2_row, 4).value)
        self.assertEqual(sheet.page_setup.orientation, 'landscape')
        self.assertEqual(sheet.page_setup.paperSize, 9)
        self.assertEqual(sheet.page_setup.fitToWidth, 1)
        self.assertEqual(sheet.page_setup.fitToHeight, 0)
        self.assertEqual(sheet.freeze_panes, 'C8')
        self.assertEqual(sheet.print_title_rows, '$7:$7')
        self.assertIn('$A$1:$E$', str(sheet.print_area))
        self.assertEqual(sheet.oddFooter.center.text, 'Страница &P из &N')

    def test_excel_export_never_turns_employee_text_into_formula(self):
        self.driver.full_name = '=HYPERLINK("https://example.invalid";"Открыть")'
        self.driver.save(update_fields=['full_name'])
        plan = self.create_draft()

        response = self.client.get(
            reverse('deputy_mining_manager_export', args=[plan.id]),
            HTTP_HOST='localhost',
        )

        workbook = load_workbook(BytesIO(response.content), data_only=False)
        sheet = workbook['Расстановка']
        employee_cell = next(
            sheet.cell(row_index, 3)
            for row_index in range(8, sheet.max_row + 1)
            if sheet.cell(row_index, 1).value == 'Т-01'
        )
        self.assertTrue(employee_cell.value.startswith("'="))
        self.assertEqual(employee_cell.data_type, 's')

    def test_excel_export_is_restricted_to_deputy_and_known_plan(self):
        plan = self.create_draft()
        driver_client = Client()
        self.authenticate(driver_client, self.driver_access)

        forbidden_response = driver_client.get(
            reverse('deputy_mining_manager_export', args=[plan.id]),
            HTTP_HOST='localhost',
        )
        missing_response = self.client.get(
            reverse('deputy_mining_manager_export', args=[plan.id + 9999]),
            HTTP_HOST='localhost',
        )

        self.assertRedirects(
            forbidden_response,
            reverse('role_home'),
            fetch_redirect_response=False,
        )
        self.assertEqual(missing_response.status_code, 404)

    def test_board_payload_exposes_brigade_and_record_details_without_new_schema(self):
        self.driver.rotation = 'Бригада № 1'
        self.driver.position = 'Водитель самосвала'
        self.driver.personnel_number = 'Т-001'
        self.driver.save(update_fields=['rotation', 'position', 'personnel_number'])
        second_driver, _second_access = self.create_employee_with_access(
            'Петров Алексей Иванович',
            self.driver_role,
            phone='+79000000002',
            access_code='210002',
        )
        second_driver.rotation = 'Вахта 2'
        second_driver.save(update_fields=['rotation'])
        unclassified_driver, _unclassified_access = self.create_employee_with_access(
            'Сидоров Николай Петрович',
            self.driver_role,
            phone='+79000000003',
            access_code='210003',
        )
        unclassified_driver.rotation = 'Вахта 15/15'
        unclassified_driver.save(update_fields=['rotation'])

        response = self.client.get(
            reverse('deputy_mining_manager_placement'),
            HTTP_HOST='localhost',
        )

        payload = response.context['planning_payload']
        assigned_employee = next(
            slot['employee']
            for row in payload['rows']
            for slot in row['slots']
            if slot['employee'] and slot['employee']['id'] == self.driver.id
        )
        free_employee = next(item for item in payload['employees'] if item['id'] == second_driver.id)
        unclassified_employee = next(item for item in payload['employees'] if item['id'] == unclassified_driver.id)
        equipment = payload['rows'][0]['equipment']
        self.assertEqual(assigned_employee['brigade_code'], '1')
        self.assertEqual(assigned_employee['brigade_label'], 'Бригада 1')
        self.assertEqual(assigned_employee['phone'], '+79000000001')
        self.assertEqual(free_employee['brigade_code'], '2')
        self.assertEqual(unclassified_employee['brigade_code'], '')
        self.assertEqual(equipment['type_label'], 'Самосвал')
        self.assertEqual(equipment['ownership_label'], 'Собственная')
        self.assertEqual(equipment['status_label'], '')
        self.assertContains(response, 'data-brigade-filter="1"')
        self.assertContains(response, 'data-brigade-filter="2"')
        self.assertContains(response, 'data-record-dialog')

    def test_autosave_changes_draft_but_not_equipment_assignment(self):
        plan = self.create_draft()

        response, plan = self.autosave_driver_on_second_truck(plan)

        self.assertTrue(response.json()['ok'])
        self.original_assignment.refresh_from_db()
        self.assertIsNone(self.original_assignment.ended_at)
        self.assertEqual(
            get_active_equipment_assignment(self.driver, 'driver').equipment,
            self.truck_1,
        )
        self.assertEqual(
            plan.slots.get(
                equipment=self.truck_2,
                shift_type=WorkShiftType.SHIFT_1,
            ).employee,
            self.driver,
        )

    def test_autosave_moves_employee_between_equipment_and_shift(self):
        plan = self.create_draft()

        response, plan = self.autosave_slot(
            plan,
            equipment=self.truck_2,
            shift_type=WorkShiftType.SHIFT_2,
            employee=self.driver,
        )

        self.assertTrue(response.json()['ok'])
        self.assertIsNone(
            plan.slots.get(
                equipment=self.truck_1,
                shift_type=WorkShiftType.SHIFT_1,
            ).employee,
        )
        self.assertEqual(
            plan.slots.get(
                equipment=self.truck_2,
                shift_type=WorkShiftType.SHIFT_2,
            ).employee,
            self.driver,
        )
        active_assignment = get_active_equipment_assignment(self.driver, 'driver')
        self.assertEqual(active_assignment.equipment, self.truck_1)
        self.assertEqual(active_assignment.shift_type, WorkShiftType.SHIFT_1)

    def test_autosave_clears_slot_and_returns_employee_to_free_pool(self):
        plan = self.create_draft()

        response, plan = self.autosave_slot(
            plan,
            equipment=self.truck_1,
            shift_type=WorkShiftType.SHIFT_1,
            employee=None,
        )

        self.assertIsNone(
            plan.slots.get(
                equipment=self.truck_1,
                shift_type=WorkShiftType.SHIFT_1,
            ).employee,
        )
        free_employee_ids = {
            item['id'] for item in response.json()['payload']['employees']
        }
        self.assertIn(self.driver.id, free_employee_ids)
        self.assertEqual(
            get_active_equipment_assignment(self.driver, 'driver').equipment,
            self.truck_1,
        )

    def test_autosave_locks_only_slot_table_with_nullable_employee_join(self):
        plan = self.create_draft()

        with patch.object(
            CrewPlanSlot.objects,
            'select_for_update',
            wraps=CrewPlanSlot.objects.select_for_update,
        ) as slot_lock:
            self.autosave_slot(
                plan,
                equipment=self.truck_1,
                shift_type=WorkShiftType.SHIFT_1,
                employee=None,
            )

        self.assertTrue(any(
            call.kwargs.get('of') == ('self',)
            for call in slot_lock.call_args_list
        ))

    def test_publish_replaces_base_equipment_assignment(self):
        plan = self.create_draft()
        _response, plan = self.autosave_driver_on_second_truck(plan)

        response = self.publish(plan)

        self.assertTrue(response.json()['published'])
        plan.refresh_from_db()
        self.assertEqual(plan.status, CrewPlanStatus.PUBLISHED)
        published_export_url = response.json()['payload']['endpoints']['export']
        self.assertEqual(
            published_export_url,
            reverse('deputy_mining_manager_export', args=[plan.id]),
        )
        exported_sheet = load_workbook(BytesIO(self.client.get(
            published_export_url,
            HTTP_HOST='localhost',
        ).content))['Расстановка']
        self.assertIn('Статус: Опубликован', exported_sheet['C3'].value)
        self.original_assignment.refresh_from_db()
        self.assertIsNotNone(self.original_assignment.ended_at)
        self.assertEqual(self.original_assignment.ended_by, self.deputy)
        active_assignment = get_active_equipment_assignment(self.driver, 'driver')
        self.assertEqual(active_assignment.equipment, self.truck_2)
        self.assertEqual(active_assignment.shift_type, WorkShiftType.SHIFT_1)
        self.assertEqual(active_assignment.assigned_by, self.deputy)
        self.assertEqual(
            EquipmentAssignment.objects.filter(
                employee=self.driver,
                role=self.driver_role,
                status=AssignmentStatus.ACCEPTED,
                ended_at__isnull=True,
            ).count(),
            1,
        )

    def test_publish_locks_only_slot_table_with_nullable_employee_joins(self):
        plan = self.create_draft()

        with patch.object(
            CrewPlanSlot.objects,
            'select_for_update',
            wraps=CrewPlanSlot.objects.select_for_update,
        ) as slot_lock:
            self.publish(plan)

        self.assertTrue(any(
            call.kwargs.get('of') == ('self',)
            for call in slot_lock.call_args_list
        ))

    def test_published_move_is_visible_in_admin_employee_card(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        _admin, admin_access = self.create_employee_with_access(
            'Администратор тестовый',
            admin_role,
            phone='+79000000099',
            access_code='110099',
        )
        admin_client = Client()
        self.authenticate(admin_client, admin_access)
        plan = self.create_draft()
        _response, plan = self.autosave_slot(
            plan,
            equipment=self.truck_2,
            shift_type=WorkShiftType.SHIFT_2,
            employee=self.driver,
        )

        self.publish(plan)
        response = admin_client.get(
            reverse('system_admin_employee_detail', args=[self.driver.id]),
            HTTP_HOST='localhost',
        )

        self.assertEqual(response.status_code, 200)
        active_assignment = response.context['active_equipment_assignment']
        self.assertEqual(active_assignment.equipment, self.truck_2)
        self.assertEqual(active_assignment.shift_type, WorkShiftType.SHIFT_2)

    def test_published_removal_is_visible_in_admin_employee_card(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        _admin, admin_access = self.create_employee_with_access(
            'Администратор тестовый',
            admin_role,
            phone='+79000000099',
            access_code='110099',
        )
        admin_client = Client()
        self.authenticate(admin_client, admin_access)
        plan = self.create_draft()
        _response, plan = self.autosave_slot(
            plan,
            equipment=self.truck_1,
            shift_type=WorkShiftType.SHIFT_1,
            employee=None,
        )

        self.publish(plan)
        response = admin_client.get(
            reverse('system_admin_employee_detail', args=[self.driver.id]),
            HTTP_HOST='localhost',
        )

        self.assertEqual(response.status_code, 200)
        self.assertIsNone(response.context['active_equipment_assignment'])
        self.assertContains(response, 'Рабочее назначение не задано')

    def test_publish_does_not_change_snapshot_of_already_open_shift(self):
        open_shift = EmployeeShift.objects.create(
            employee=self.driver,
            equipment=self.truck_1,
            shift_type=WorkShiftType.SHIFT_1,
            opened_at=timezone.now(),
            opened_by=self.driver,
        )
        plan = self.create_draft()
        _response, plan = self.autosave_driver_on_second_truck(plan)

        self.publish(plan)

        open_shift.refresh_from_db()
        self.assertEqual(open_shift.equipment, self.truck_1)
        self.assertEqual(open_shift.shift_type, WorkShiftType.SHIFT_1)
        self.assertIsNone(open_shift.closed_at)
        self.assertEqual(
            get_active_equipment_assignment(self.driver, 'driver').equipment,
            self.truck_2,
        )

    def test_driver_login_sees_new_equipment_after_publication(self):
        plan = self.create_draft()
        _response, plan = self.autosave_driver_on_second_truck(plan)
        self.publish(plan)
        driver_client = Client()

        response = driver_client.post(
            reverse('login'),
            {
                'phone': '+7 (900) 000-00-01',
                'access_code': '210001',
            },
            follow=True,
            HTTP_HOST='localhost',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(driver_client.session['employee_access_id'], self.driver_access.id)
        self.assertEqual(response.context['work_assignment_state'], 'assigned')
        self.assertEqual(response.context['work_assignment_equipment'], self.truck_2)
        self.assertEqual(response.context['work_assignment'].shift_type, WorkShiftType.SHIFT_1)

    def test_reports_page_lists_published_crew_plan(self):
        plan = self.create_draft()
        self.publish(plan)

        response = self.client.get(
            reverse('deputy_mining_manager_reports'),
            HTTP_HOST='localhost',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'reports/deputy_mining_manager_reports.html')
        publications = response.context['publications']
        self.assertEqual(len(publications), 1)
        self.assertEqual(publications[0].id, plan.id)
        self.assertEqual(publications[0].slot_count, 4)
        self.assertEqual(publications[0].assigned_count, 1)
        self.assertIn(f'plan={plan.id}', publications[0].url)

        detail_response = self.client.get(publications[0].url, HTTP_HOST='localhost')
        self.assertEqual(detail_response.status_code, 200)
        detail_payload = detail_response.context['planning_payload']
        self.assertEqual(detail_payload['plan']['id'], plan.id)
        self.assertEqual(detail_payload['plan']['status'], CrewPlanStatus.PUBLISHED)
        self.assertFalse(detail_payload['plan']['editable'])
        self.assertEqual(detail_payload['employees'], [])

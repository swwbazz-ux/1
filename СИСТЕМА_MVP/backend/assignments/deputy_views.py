import json
import re
from datetime import date
from urllib.parse import quote, urlencode

from django.core.exceptions import ValidationError
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.templatetags.static import static
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_GET, require_POST
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins

from references.models import Equipment
from users.models import Employee, EmployeeAccess, Role
from users.role_apps import role_app_manifest_response, role_app_service_worker_response

from .models import (
    AssignmentStatus,
    CrewPlan,
    CrewPlanStatus,
    EquipmentAssignment,
    WorkShiftType,
)
from .services import (
    WORK_ASSIGNMENT_ROLE_EQUIPMENT_TYPES,
    get_or_create_crew_draft,
    production_work_date,
    publish_crew_plan,
    update_crew_draft_slot,
)


DEPUTY_ROLE_CODE = 'deputy_mining_manager'
TARGET_ROLE_LABELS = {
    'driver': 'Самосвалы',
    'excavator_operator': 'Экскаваторы',
}

DEPUTY_MANIFEST = {
    'id': '/deputy-mining-manager/',
    'name': 'Заместитель начальника горного участка',
    'short_name': 'Расстановка',
    'description': 'Рабочее место для расстановки сотрудников по технике и контроля опубликованных назначений.',
    'start_url': '/deputy-mining-manager/',
    'scope': '/deputy-mining-manager/',
    'display': 'standalone',
    'display_override': ['standalone', 'fullscreen'],
    'orientation': 'landscape',
    'background_color': '#dce4e7',
    'theme_color': '#198e55',
    'lang': 'ru',
    'categories': ['business', 'productivity'],
    'prefer_related_applications': False,
    'icons': [
        {
            'src': '/static/img/pwa/deputy-mining-manager-192.png',
            'sizes': '192x192',
            'type': 'image/png',
            'purpose': 'any',
        },
        {
            'src': '/static/img/pwa/deputy-mining-manager-512.png',
            'sizes': '512x512',
            'type': 'image/png',
            'purpose': 'any',
        },
        {
            'src': '/static/img/pwa/deputy-mining-manager-maskable-512.png',
            'sizes': '512x512',
            'type': 'image/png',
            'purpose': 'maskable',
        },
    ],
    'shortcuts': [
        {
            'name': 'Расстановка',
            'short_name': 'Расстановка',
            'url': '/deputy-mining-manager/',
            'description': 'Открыть расстановку сотрудников по технике.',
        },
        {
            'name': 'Отчёты',
            'short_name': 'Отчёты',
            'url': '/deputy-mining-manager/reports/',
            'description': 'Открыть историю опубликованных расстановок.',
        },
    ],
}

DEPUTY_SERVICE_WORKER_JS = r"""
const CACHE_PREFIX = "deputy-mining-manager-desktop-shell-";
const CACHE_NAME = `${CACHE_PREFIX}v6`;
const APP_SCOPE = "/deputy-mining-manager/";
const MANIFEST_URL = "/deputy-mining-manager.webmanifest";
const LEGACY_ROOT_FALLBACK_URL = "/mining-master/assignments/";
const LEGACY_CACHE_PREFIX = "mining-master-mobile-shell-";
const CORE_ASSETS = [
  MANIFEST_URL,
  "/static/css/app.css",
  "/static/css/deputy-mining-manager-v3.css",
  "/static/js/deputy-mining-manager-v3.js",
  "/static/js/deputy-mining-manager-pwa-v1.js",
  "/static/favicon.ico",
  "/static/img/pwa/deputy-mining-manager-180.png",
  "/static/img/pwa/deputy-mining-manager-192.png",
  "/static/img/pwa/deputy-mining-manager-512.png",
  "/static/img/pwa/deputy-mining-manager-maskable-512.png",
  "/static/img/equipment/truck-green.png",
  "/static/img/equipment/truck-gray.png",
  "/static/img/equipment/excavator-green.png",
  "/static/img/equipment/excavator-gray.png"
];
const STATIC_ASSET_PATHS = new Set(
  CORE_ASSETS.filter(url => url.startsWith("/static/"))
);

async function removeCachedPlanningDocuments() {
  const cacheNames = await caches.keys();
  await Promise.all(cacheNames.map(async cacheName => {
    const cache = await caches.open(cacheName);
    const requests = await cache.keys();
    await Promise.all(requests.map(async request => {
      const url = new URL(request.url);
      if (url.origin !== self.location.origin) return;
      if (url.pathname.startsWith(APP_SCOPE)) {
        await cache.delete(request);
        return;
      }
      if (
        !cacheName.startsWith(LEGACY_CACHE_PREFIX) ||
        url.pathname !== LEGACY_ROOT_FALLBACK_URL
      ) return;
      const cachedResponse = await cache.match(request);
      if (!cachedResponse) return;
      const contentType = cachedResponse.headers.get("Content-Type") || "";
      if (!contentType.includes("text/html")) return;
      const cachedHtml = await cachedResponse.clone().text();
      if (
        cachedHtml.includes("deputy-mining-manager-screen") ||
        cachedHtml.includes("data-deputy-planning-root")
      ) {
        await cache.delete(request);
      }
    }));
  }));
}

self.addEventListener("install", event => {
  event.waitUntil(
    caches.open(CACHE_NAME)
      .then(cache => Promise.allSettled(
        CORE_ASSETS.map(url => cache.add(new Request(url, { cache: "reload" })))
      ))
      .then(() => self.skipWaiting())
  );
});

self.addEventListener("activate", event => {
  event.waitUntil(
    caches.keys()
      .then(keys => Promise.all(
        keys
          .filter(key => key.startsWith(CACHE_PREFIX) && key !== CACHE_NAME)
          .map(key => caches.delete(key))
      ))
      .then(() => removeCachedPlanningDocuments())
      .then(() => self.clients.claim())
  );
});

function offlineDocument() {
  return new Response(`<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <meta name="theme-color" content="#198e55">
  <title>Нет подключения</title>
  <style>
    *{box-sizing:border-box}body{margin:0;min-height:100vh;display:grid;place-items:center;padding:24px;background:#dce4e7;color:#162022;font-family:Arial,sans-serif}.offline{width:min(460px,100%);padding:28px;border:1px solid #cdd7da;border-top:5px solid #198e55;border-radius:10px;background:#f7faf9;box-shadow:0 18px 44px rgba(53,70,78,.18)}h1{margin:0 0 10px;font-size:24px}p{margin:0;color:#647277;font-size:16px;line-height:1.5}
  </style>
</head>
<body><main class="offline"><h1>Нет подключения к серверу</h1><p>Для работы с расстановкой требуется сеть. Подключитесь к интернету и откройте приложение ещё раз.</p></main></body>
</html>`, {
    status: 503,
    headers: {
      "Content-Type": "text/html; charset=utf-8",
      "Cache-Control": "no-store"
    }
  });
}

async function networkOnlyNavigation(request) {
  try {
    return await fetch(request, { cache: "no-store" });
  } catch (error) {
    return offlineDocument();
  }
}

async function networkFirst(request) {
  const cache = await caches.open(CACHE_NAME);
  try {
    const response = await fetch(request, { cache: "no-store" });
    if (response && response.ok) {
      cache.put(request, response.clone()).catch(() => undefined);
    }
    return response;
  } catch (error) {
    return (await cache.match(request, { ignoreSearch: true })) ||
      new Response("Ресурс недоступен без сети.", {
        status: 503,
        headers: { "Content-Type": "text/plain; charset=utf-8" }
      });
  }
}

function canonicalStaticRequest(request) {
  const url = new URL(request.url);
  return new Request(`${url.origin}${url.pathname}`, {
    method: "GET",
    credentials: "same-origin"
  });
}

async function networkFirstStatic(request) {
  const cache = await caches.open(CACHE_NAME);
  const cacheKey = canonicalStaticRequest(request);
  try {
    const response = await fetch(request, { cache: "no-store" });
    if (response && response.ok) {
      cache.put(cacheKey, response.clone()).catch(() => undefined);
    }
    return response;
  } catch (error) {
    return (await cache.match(cacheKey)) ||
      new Response("Ресурс недоступен без сети.", {
        status: 503,
        headers: { "Content-Type": "text/plain; charset=utf-8" }
      });
  }
}

self.addEventListener("fetch", event => {
  const request = event.request;
  if (request.method !== "GET") return;
  const url = new URL(request.url);
  if (url.origin !== self.location.origin) return;

  if (request.mode === "navigate" && url.pathname.startsWith(APP_SCOPE)) {
    event.respondWith(networkOnlyNavigation(request));
    return;
  }
  if (url.pathname === MANIFEST_URL) {
    event.respondWith(networkFirst(request));
    return;
  }
  if (STATIC_ASSET_PATHS.has(url.pathname)) {
    event.respondWith(networkFirstStatic(request));
  }
});

self.addEventListener("message", event => {
  if (!event.data) return;
  if (event.data.type === "SKIP_WAITING") {
    self.skipWaiting();
  }
});
"""


def deputy_access_from_request(request):
    access_id = request.session.get('employee_access_id')
    if not access_id:
        return None
    return (
        EmployeeAccess.objects
        .select_related('employee', 'role')
        .filter(
            id=access_id,
            is_active=True,
            employee__is_active=True,
            role__is_active=True,
            role__code=DEPUTY_ROLE_CODE,
        )
        .exclude(status__in=[EmployeeAccess.Status.BLOCKED, EmployeeAccess.Status.DEACTIVATED])
        .first()
    )


def _employee_short_name(employee):
    parts = (employee.full_name or '').split()
    if not parts:
        return ''
    initials = ''.join(f'{part[0].upper()}.' for part in parts[1:3] if part)
    return f'{parts[0]} {initials}'.strip()


def _employee_brigade_code(employee):
    rotation = (employee.rotation or '').strip().casefold()
    if not rotation:
        return ''
    compact = re.sub(r'\s+', ' ', rotation.replace('№', ' ')).strip(' .,-')
    aliases = {
        '1': '1',
        '1-я': '1',
        '1 я': '1',
        'первая': '1',
        'первая бригада': '1',
        'бригада 1': '1',
        '1 бригада': '1',
        'вахта 1': '1',
        '1 вахта': '1',
        '2': '2',
        '2-я': '2',
        '2 я': '2',
        'вторая': '2',
        'вторая бригада': '2',
        'бригада 2': '2',
        '2 бригада': '2',
        'вахта 2': '2',
        '2 вахта': '2',
    }
    if compact in aliases:
        return aliases[compact]
    brigade_match = re.search(
        r'(?:бригада|вахта)\s*[-:]?\s*([12])(?:-?я)?(?=$|[\s,;.(])',
        compact,
    ) or re.search(
        r'^([12])(?:-?я)?\s*(?:бригада|вахта)(?=$|[\s,;.(])',
        compact,
    )
    return brigade_match.group(1) if brigade_match else ''


def _employee_payload(employee):
    if not employee:
        return None
    photo_url = ''
    if employee.photo:
        try:
            photo_url = employee.photo.url
        except ValueError:
            photo_url = ''
    initials = ''.join(part[0] for part in (employee.full_name or '').split()[:2]).upper() or '—'
    brigade_code = _employee_brigade_code(employee)
    return {
        'id': employee.id,
        'full_name': employee.full_name or '',
        'short_name': _employee_short_name(employee),
        'position_label': employee.position or '',
        'personnel_number': employee.personnel_number or '',
        'phone': employee.phone or '',
        'photo_url': photo_url,
        'initials': initials,
        'status_label': employee.get_status_display(),
        'rotation_label': employee.rotation or '',
        'brigade_code': brigade_code,
        'brigade_label': f'Бригада {brigade_code}' if brigade_code else 'Не указана',
        'search': f'{employee.full_name} {employee.personnel_number}'.strip().lower(),
    }


def _role_from_request(request):
    role_code = request.GET.get('role') or request.POST.get('role') or 'driver'
    if role_code not in WORK_ASSIGNMENT_ROLE_EQUIPMENT_TYPES:
        role_code = 'driver'
    return get_object_or_404(Role, code=role_code, is_active=True)


def _work_date_from_request(request):
    current_date = production_work_date()
    raw_value = request.GET.get('date', '').strip()
    if not raw_value:
        return current_date
    try:
        selected_date = date.fromisoformat(raw_value)
    except ValueError:
        return current_date
    return min(selected_date, current_date)


def _slot_issue(slot, eligible_employee_ids, other_role_assignment_employee_ids):
    employee = slot.employee
    if not employee:
        return ''
    if not employee.is_active or employee.status != Employee.Status.ACTIVE:
        return 'Сотрудник неактивен'
    if employee.id not in eligible_employee_ids:
        return 'Не соответствует рабочей категории'
    if employee.id in other_role_assignment_employee_ids:
        return 'Назначен по другой роли'
    if not slot.equipment.is_active:
        return 'Техника недоступна'
    return ''


def _natural_equipment_sort_key(equipment):
    label = equipment.garage_number or str(equipment)
    return tuple(
        int(part) if part.isdigit() else part.casefold()
        for part in re.split(r'(\d+)', label)
    )


def build_crew_plan_payload(plan, *, request=None):
    if not plan:
        return {
            'plan': None,
            'role': None,
            'categories': [],
            'endpoints': {},
            'summary': {
                'equipment_total': 0,
                'assigned_count': 0,
                'unfilled_count': 0,
                'conflict_count': 0,
                'changed_count': 0,
            },
            'employees': [],
            'rows': [],
        }

    slots = list(
        plan.slots
        .select_related('equipment', 'equipment__equipment_type', 'equipment__model', 'employee', 'baseline_employee')
        .order_by('equipment__garage_number', 'shift_type')
    )
    assigned_employee_ids = {slot.employee_id for slot in slots if slot.employee_id}
    activated_employee_ids = set(
        EmployeeAccess.objects.filter(
            role=plan.role,
            is_active=True,
            status=EmployeeAccess.Status.ACTIVATED,
            employee__is_active=True,
            employee__status=Employee.Status.ACTIVE,
        ).values_list('employee_id', flat=True)
    )
    eligible_employee_ids = set(
        Employee.objects.filter(
            Q(work_category=plan.role.code)
            | Q(work_category=Employee.WorkCategory.OTHER, id__in=activated_employee_ids),
            is_active=True,
            status=Employee.Status.ACTIVE,
        ).values_list('id', flat=True)
    )
    other_role_assignment_employee_ids = set(
        EquipmentAssignment.objects.filter(
            employee_id__in=eligible_employee_ids,
            status=AssignmentStatus.ACCEPTED,
            ended_at__isnull=True,
            shift__isnull=True,
            role__isnull=False,
            shift_type__in=WorkShiftType.values,
        )
        .exclude(role=plan.role)
        .values_list('employee_id', flat=True)
    )
    editable = plan.status == CrewPlanStatus.DRAFT and plan.work_date == production_work_date()
    eligible_employees = []
    if editable:
        eligible_employees = list(
            Employee.objects.filter(id__in=eligible_employee_ids)
            .exclude(id__in=assigned_employee_ids)
            .exclude(id__in=other_role_assignment_employee_ids)
            .order_by('full_name')
        )

    slot_map = {}
    for slot in slots:
        slot_map[(slot.equipment_id, slot.shift_type)] = slot

    equipment_items = []
    seen_equipment_ids = set()
    for slot in slots:
        if slot.equipment_id in seen_equipment_ids:
            continue
        seen_equipment_ids.add(slot.equipment_id)
        equipment_items.append(slot.equipment)
    equipment_items.sort(key=_natural_equipment_sort_key)

    rows = []
    assigned_count = 0
    unfilled_count = 0
    conflict_count = 0
    changed_count = 0
    icon_prefix = 'truck' if plan.role.code == 'driver' else 'excavator'
    for equipment in equipment_items:
        row_slots = []
        row_attention = False
        row_conflict = False
        row_changed = False
        for shift_type, shift_label in WorkShiftType.choices:
            slot = slot_map.get((equipment.id, shift_type))
            if not slot:
                continue
            issue = _slot_issue(
                slot,
                eligible_employee_ids,
                other_role_assignment_employee_ids,
            )
            changed = slot.employee_id != slot.baseline_employee_id
            if slot.employee_id:
                assigned_count += 1
            else:
                unfilled_count += 1
                row_attention = True
            if issue:
                conflict_count += 1
                row_attention = True
                row_conflict = True
            if changed:
                changed_count += 1
                row_changed = True
            row_slots.append({
                'shift_type': shift_type,
                'label': 'День' if shift_type == WorkShiftType.SHIFT_1 else 'Ночь',
                'time_label': '07:00–19:00' if shift_type == WorkShiftType.SHIFT_1 else '19:00–07:00',
                'employee': _employee_payload(slot.employee),
                'changed': changed,
                'conflict': bool(issue),
                'issue': issue,
            })
        status_label = '' if equipment.is_active else 'Недоступна'
        rows.append({
            'equipment': {
                'id': equipment.id,
                'label': equipment.garage_number or str(equipment),
                'type_label': equipment.equipment_type.name,
                'model_label': equipment.model.name if equipment.model_id else '',
                'icon_url': static(f'img/equipment/{icon_prefix}-{"green" if equipment.is_active else "gray"}.png'),
                'is_active': equipment.is_active,
                'status_label': status_label,
                'serial_number': equipment.vin or '',
                'ownership_label': 'Собственная' if equipment.is_own else 'Подрядная',
            },
            'attention': row_attention,
            'conflict': row_conflict,
            'changed': row_changed,
            'search': ' '.join([
                equipment.garage_number or '',
                equipment.model.name if equipment.model_id else '',
                *[item['employee']['full_name'] for item in row_slots if item['employee']],
            ]).lower(),
            'slots': row_slots,
        })

    categories = []
    selected_date = plan.work_date.isoformat()
    for role_code, equipment_type_name in WORK_ASSIGNMENT_ROLE_EQUIPMENT_TYPES.items():
        query = urlencode({'role': role_code, 'date': selected_date})
        categories.append({
            'code': role_code,
            'label': TARGET_ROLE_LABELS.get(role_code, equipment_type_name),
            'url': f'{reverse("deputy_mining_manager_placement")}?{query}',
        })

    return {
        'plan': {
            'id': plan.id,
            'version': plan.version,
            'editable': editable,
            'status': plan.status,
            'work_date': plan.work_date.isoformat(),
            'work_date_label': plan.work_date.strftime('%d.%m.%Y'),
            'updated_at_label': timezone.localtime(plan.updated_at).strftime('%H:%M'),
        },
        'role': {
            'code': plan.role.code,
            'label': plan.role.name,
            'category_label': TARGET_ROLE_LABELS.get(plan.role.code, plan.role.name),
        },
        'categories': categories,
        'endpoints': {
            'slot': reverse('deputy_mining_manager_slot'),
            'publish': reverse('deputy_mining_manager_publish'),
            'export': reverse('deputy_mining_manager_export', args=[plan.id]),
        },
        'summary': {
            'equipment_total': len(rows),
            'assigned_count': assigned_count,
            'unfilled_count': unfilled_count,
            'conflict_count': conflict_count,
            'changed_count': changed_count,
        },
        'employees': [_employee_payload(employee) for employee in eligible_employees],
        'rows': rows,
    }


DEPUTY_XLSX_CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
DEPUTY_XLSX_TABLE_HEADER_ROW = 7


def _excel_text(value):
    text = '' if value is None else str(value)
    if text.startswith(('=', '+', '-', '@')):
        return f"'{text}"
    return text


def _employee_excel_value(employee):
    if not employee:
        return 'Не назначен'
    details = []
    if employee.get('personnel_number'):
        details.append(f"Таб. № {employee['personnel_number']}")
    brigade_label = employee.get('brigade_label') or ''
    if brigade_label and brigade_label != 'Не указана':
        details.append(brigade_label)
    lines = [_excel_text(employee.get('full_name') or 'Сотрудник')]
    if details:
        lines.append(_excel_text(' · '.join(details)))
    return '\n'.join(lines)


def _style_merged_range(sheet, cell_range, *, fill, font, alignment, border):
    anchor_reference = cell_range.split(':', 1)[0]
    if ':' in cell_range and cell_range.split(':', 1)[1] != anchor_reference:
        sheet.merge_cells(cell_range)
    selected_cells = sheet[cell_range]
    if not isinstance(selected_cells, tuple):
        selected_cells = ((selected_cells,),)
    elif selected_cells and not isinstance(selected_cells[0], tuple):
        selected_cells = (selected_cells,)
    for row in selected_cells:
        for cell in row:
            cell.fill = fill
            cell.border = border
            cell.alignment = alignment
    anchor = sheet[anchor_reference]
    anchor.font = font
    return anchor


def build_deputy_crew_plan_workbook(plan, *, actor):
    payload = build_crew_plan_payload(plan)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Расстановка'
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 90

    colors = {
        'ink': '162022',
        'muted': '647277',
        'line': 'CDD7DA',
        'line_soft': 'DDE5E7',
        'panel': 'F7FAF9',
        'panel_soft': 'E9F0EE',
        'green': '198E55',
        'green_soft': 'DFF5EA',
        'yellow_soft': 'FFF0C9',
        'red': 'D64A42',
        'red_soft': 'FFE2DF',
        'white': 'FFFFFF',
    }
    thin = Side(style='thin', color=colors['line'])
    table_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    metadata_border = Border(bottom=Side(style='thin', color=colors['line_soft']))
    title_fill = PatternFill('solid', fgColor=colors['ink'])
    green_fill = PatternFill('solid', fgColor=colors['green'])
    green_soft_fill = PatternFill('solid', fgColor=colors['green_soft'])
    panel_fill = PatternFill('solid', fgColor=colors['panel'])
    panel_soft_fill = PatternFill('solid', fgColor=colors['panel_soft'])
    yellow_fill = PatternFill('solid', fgColor=colors['yellow_soft'])
    red_fill = PatternFill('solid', fgColor=colors['red_soft'])

    title = _style_merged_range(
        sheet,
        'A1:E1',
        fill=title_fill,
        font=Font(name='Arial', size=18, bold=True, color=colors['white']),
        alignment=Alignment(horizontal='left', vertical='center'),
        border=Border(),
    )
    title.value = 'COPPER RESOURCES · РАССТАНОВКА ПО ТЕХНИКЕ'
    sheet.row_dimensions[1].height = 32

    subtitle = _style_merged_range(
        sheet,
        'A2:E2',
        fill=PatternFill('solid', fgColor=colors['ink']),
        font=Font(name='Arial', size=11, bold=True, color='DFF5EA'),
        alignment=Alignment(horizontal='left', vertical='center'),
        border=Border(bottom=Side(style='medium', color=colors['green'])),
    )
    subtitle.value = _excel_text(
        f"{payload['role']['category_label']} · дата расстановки {plan.work_date:%d.%m.%Y}"
    )
    sheet.row_dimensions[2].height = 22

    status_fill, status_color = {
        CrewPlanStatus.DRAFT: (yellow_fill, '8A5B00'),
        CrewPlanStatus.PUBLISHED: (green_soft_fill, colors['green']),
        CrewPlanStatus.SUPERSEDED: (panel_soft_fill, colors['muted']),
    }.get(plan.status, (panel_soft_fill, colors['muted']))
    metadata_alignment = Alignment(horizontal='left', vertical='center')
    metadata_font = Font(name='Arial', size=10, bold=True, color=colors['ink'])
    metadata_ranges = (
        ('A3:B3', panel_fill, metadata_font, f"Категория: {payload['role']['category_label']}"),
        ('C3:D3', status_fill, Font(name='Arial', size=10, bold=True, color=status_color), f"Статус: {plan.get_status_display()}"),
        ('E3', panel_fill, metadata_font, f"Ревизия: {plan.revision} · версия: {plan.version}"),
        (
            'A4:E4',
            panel_fill,
            metadata_font,
            f"Сформировал: {actor.full_name or 'Пользователь'}   ·   "
            f"Сформировано: {timezone.localtime():%d.%m.%Y %H:%M}",
        ),
    )
    for cell_range, fill, font, value in metadata_ranges:
        anchor = _style_merged_range(
            sheet,
            cell_range,
            fill=fill,
            font=font,
            alignment=metadata_alignment,
            border=metadata_border,
        )
        anchor.value = _excel_text(value)
    sheet.row_dimensions[3].height = 23
    sheet.row_dimensions[4].height = 23

    summary = payload['summary']
    summary_cell = _style_merged_range(
        sheet,
        'A5:E5',
        fill=green_soft_fill,
        font=Font(name='Arial', size=10, bold=True, color=colors['green']),
        alignment=Alignment(horizontal='left', vertical='center'),
        border=Border(
            top=Side(style='thin', color=colors['green']),
            bottom=Side(style='thin', color=colors['green']),
        ),
    )
    summary_cell.value = (
        f"Техника: {summary['equipment_total']}    "
        f"Назначено: {summary['assigned_count']}    "
        f"Не заполнено: {summary['unfilled_count']}    "
        f"Конфликты: {summary['conflict_count']}    "
        f"Изменено: {summary['changed_count']}"
    )
    sheet.row_dimensions[5].height = 24
    sheet.row_dimensions[6].height = 8

    headers = ('Техника', 'Модель', 'День · 07:00–19:00', 'Ночь · 19:00–07:00', 'Примечание')
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(DEPUTY_XLSX_TABLE_HEADER_ROW, column, header)
        cell.fill = green_fill
        cell.font = Font(name='Arial', size=10, bold=True, color=colors['white'])
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = table_border
    sheet.row_dimensions[DEPUTY_XLSX_TABLE_HEADER_ROW].height = 28

    data_start_row = DEPUTY_XLSX_TABLE_HEADER_ROW + 1
    for row_index, row in enumerate(payload['rows'], start=data_start_row):
        equipment = row['equipment']
        slots = {slot['shift_type']: slot for slot in row['slots']}
        day_slot = slots.get(WorkShiftType.SHIFT_1, {})
        night_slot = slots.get(WorkShiftType.SHIFT_2, {})
        issues = []
        if not equipment.get('is_active'):
            issues.append('Техника недоступна')
        for slot in (day_slot, night_slot):
            if slot.get('issue'):
                issues.append(f"{slot.get('label') or 'Смена'}: {slot['issue']}")
        values = (
            _excel_text(equipment.get('label')),
            _excel_text(equipment.get('model_label') or equipment.get('type_label')),
            _employee_excel_value(day_slot.get('employee')),
            _employee_excel_value(night_slot.get('employee')),
            _excel_text('\n'.join(issues)),
        )
        base_fill = panel_fill if (row_index - data_start_row) % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column, value)
            cell.fill = base_fill
            cell.font = Font(name='Arial', size=9, color=colors['ink'], bold=column == 1)
            cell.alignment = Alignment(
                horizontal='left',
                vertical='center',
                wrap_text=True,
            )
            cell.border = table_border
        for column, slot in ((3, day_slot), (4, night_slot)):
            cell = sheet.cell(row_index, column)
            if slot.get('conflict'):
                cell.fill = red_fill
                cell.font = Font(name='Arial', size=9, bold=True, color=colors['red'])
            elif not slot.get('employee'):
                cell.fill = yellow_fill
                cell.font = Font(name='Arial', size=9, italic=True, color='8A5B00')
        if issues:
            sheet.cell(row_index, 5).fill = red_fill
            sheet.cell(row_index, 5).font = Font(name='Arial', size=9, bold=True, color=colors['red'])
        sheet.row_dimensions[row_index].height = 34

    last_row = max(DEPUTY_XLSX_TABLE_HEADER_ROW, sheet.max_row)
    widths = {'A': 16, 'B': 23, 'C': 34, 'D': 34, 'E': 24}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = f'C{data_start_row}'
    sheet.auto_filter.ref = f'A{DEPUTY_XLSX_TABLE_HEADER_ROW}:E{last_row}'
    sheet.print_area = f'A1:E{last_row}'
    sheet.print_title_rows = f'{DEPUTY_XLSX_TABLE_HEADER_ROW}:{DEPUTY_XLSX_TABLE_HEADER_ROW}'
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_options.horizontalCentered = True
    sheet.page_margins = PageMargins(left=.25, right=.25, top=.35, bottom=.45, header=.15, footer=.2)
    sheet.oddFooter.center.text = 'Страница &P из &N'
    sheet.oddFooter.center.size = 9
    sheet.oddFooter.center.color = colors['muted']
    sheet.oddFooter.right.text = 'Copper Resources'
    sheet.oddFooter.right.size = 9
    sheet.oddFooter.right.color = colors['muted']
    return workbook


def _deputy_crew_plan_filename(plan):
    category = TARGET_ROLE_LABELS.get(plan.role.code, plan.role.name).replace(' ', '_')
    status = {
        CrewPlanStatus.DRAFT: 'черновик',
        CrewPlanStatus.PUBLISHED: 'опубликовано',
        CrewPlanStatus.SUPERSEDED: 'архив',
    }.get(plan.status, plan.status)
    localized = f'Расстановка_{category}_{plan.work_date:%Y-%m-%d}_{status}.xlsx'
    ascii_name = f'crew_plan_{plan.role.code}_{plan.work_date:%Y-%m-%d}_r{plan.revision}.xlsx'
    return ascii_name, localized


def deputy_mining_manager_placement_view(request):
    access = deputy_access_from_request(request)
    if not access:
        return redirect('role_home')
    role = _role_from_request(request)
    selected_date = _work_date_from_request(request)
    requested_plan_id = request.GET.get('plan', '').strip()
    if requested_plan_id:
        plan = get_object_or_404(
            CrewPlan.objects.filter(
                role=role,
                status__in=[CrewPlanStatus.PUBLISHED, CrewPlanStatus.SUPERSEDED],
            ),
            id=requested_plan_id,
        )
        selected_date = plan.work_date
    elif selected_date == production_work_date():
        plan, _ = get_or_create_crew_draft(work_date=selected_date, role=role, actor=access.employee)
    else:
        plan = (
            CrewPlan.objects
            .filter(work_date=selected_date, role=role, status=CrewPlanStatus.PUBLISHED)
            .order_by('-revision')
            .first()
        )
    payload = build_crew_plan_payload(plan, request=request)
    return render(
        request,
        'assignments/deputy_mining_manager_placement.html',
        {
            'access': access,
            'planning_payload': payload,
            'selected_role': role,
            'selected_date': selected_date,
            'current_production_date': production_work_date(),
        },
    )


@require_GET
def deputy_mining_manager_export_view(request, plan_id):
    access = deputy_access_from_request(request)
    if not access:
        return redirect('role_home')
    plan = get_object_or_404(
        CrewPlan.objects
        .select_related('role', 'created_by', 'updated_by', 'published_by')
        .filter(role__code__in=WORK_ASSIGNMENT_ROLE_EQUIPMENT_TYPES),
        id=plan_id,
    )
    workbook = build_deputy_crew_plan_workbook(plan, actor=access.employee)
    ascii_name, localized_name = _deputy_crew_plan_filename(plan)
    response = HttpResponse(content_type=DEPUTY_XLSX_CONTENT_TYPE)
    response['Content-Disposition'] = (
        f'attachment; filename="{ascii_name}"; '
        f"filename*=UTF-8''{quote(localized_name)}"
    )
    response['Cache-Control'] = 'private, no-store, max-age=0'
    response['X-Content-Type-Options'] = 'nosniff'
    workbook.save(response)
    return response


def deputy_mining_manager_manifest_view(request):
    return role_app_manifest_response(request, 'deputy_mining_manager')


def deputy_mining_manager_service_worker_view(request):
    return role_app_service_worker_response(
        request,
        'deputy_mining_manager',
        DEPUTY_SERVICE_WORKER_JS,
    )


def _json_payload(request):
    try:
        return json.loads(request.body.decode('utf-8') or '{}')
    except (UnicodeDecodeError, json.JSONDecodeError):
        raise ValidationError('Переданы некорректные данные.')


def _validation_response(error):
    code = getattr(error, 'code', '') or ''
    status = 409 if code in {'stale_version', 'stale_baseline', 'plan_work_date_closed'} else 400
    messages = getattr(error, 'messages', None) or [str(error)]
    return JsonResponse({'ok': False, 'error': messages[0], 'code': code}, status=status)


@require_POST
def deputy_mining_manager_slot_view(request):
    access = deputy_access_from_request(request)
    if not access:
        return JsonResponse({'ok': False, 'error': 'Недостаточно прав.'}, status=403)
    try:
        payload = _json_payload(request)
        plan = get_object_or_404(CrewPlan, id=payload.get('plan_id'))
        equipment = get_object_or_404(Equipment, id=payload.get('equipment_id'))
        employee_id = payload.get('employee_id')
        employee = get_object_or_404(Employee, id=employee_id) if employee_id else None
        plan = update_crew_draft_slot(
            plan=plan,
            equipment=equipment,
            shift_type=payload.get('shift_type'),
            employee=employee,
            expected_version=int(payload.get('expected_version')),
            actor=access.employee,
        )
    except (TypeError, ValueError, ValidationError) as error:
        if isinstance(error, ValidationError):
            return _validation_response(error)
        return JsonResponse({'ok': False, 'error': 'Обновите страницу и повторите действие.'}, status=400)
    return JsonResponse({'ok': True, 'payload': build_crew_plan_payload(plan, request=request)})


@require_POST
def deputy_mining_manager_publish_view(request):
    access = deputy_access_from_request(request)
    if not access:
        return JsonResponse({'ok': False, 'error': 'Недостаточно прав.'}, status=403)
    try:
        payload = _json_payload(request)
        plan = get_object_or_404(CrewPlan, id=payload.get('plan_id'))
        published_plan = publish_crew_plan(
            plan=plan,
            expected_version=int(payload.get('expected_version')),
            actor=access.employee,
        )
        next_draft, _ = get_or_create_crew_draft(
            work_date=published_plan.work_date,
            role=published_plan.role,
            actor=access.employee,
        )
    except (TypeError, ValueError, ValidationError) as error:
        if isinstance(error, ValidationError):
            return _validation_response(error)
        return JsonResponse({'ok': False, 'error': 'Обновите страницу и повторите действие.'}, status=400)
    next_payload = build_crew_plan_payload(next_draft, request=request)
    # Сразу после публикации печатается именно утвержденная ревизия. Как только
    # пользователь изменит новый черновик, обычный slot-ответ вернет его URL.
    next_payload['endpoints']['export'] = reverse(
        'deputy_mining_manager_export',
        args=[published_plan.id],
    )
    return JsonResponse({
        'ok': True,
        'published': True,
        'payload': next_payload,
    })


def deputy_mining_manager_reports_view(request):
    access = deputy_access_from_request(request)
    if not access:
        return redirect('role_home')
    publications = list(
        CrewPlan.objects
        .filter(status__in=[CrewPlanStatus.PUBLISHED, CrewPlanStatus.SUPERSEDED])
        .select_related('role', 'published_by')
        .annotate(
            slot_count=Count('slots'),
            assigned_count=Count('slots', filter=Q(slots__employee__isnull=False)),
        )
        .order_by('-work_date', '-revision')[:30]
    )
    for publication in publications:
        publication.work_date_label = publication.work_date.strftime('%d.%m.%Y')
        publication.published_at_label = (
            timezone.localtime(publication.published_at).strftime('%d.%m.%Y %H:%M')
            if publication.published_at
            else '—'
        )
        publication.published_by_label = (
            publication.published_by.full_name
            if publication.published_by
            else 'Система'
        )
        publication.status_label = publication.get_status_display()
        publication.url = (
            f'{reverse("deputy_mining_manager_placement")}?' +
            urlencode({
                'role': publication.role.code,
                'date': publication.work_date.isoformat(),
                'plan': publication.id,
            })
        )
    return render(
        request,
        'reports/deputy_mining_manager_reports.html',
        {'access': access, 'publications': publications},
    )

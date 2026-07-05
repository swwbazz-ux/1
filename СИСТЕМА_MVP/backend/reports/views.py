from collections import defaultdict
from datetime import datetime, timedelta
from decimal import Decimal

from django.contrib import messages
from django.db.models import Count, Sum
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from downtimes.models import DowntimeEvent, DowntimeReason
from references.models import DumpPoint, Equipment, RockType
from shifts.models import EmployeeShift
from shifts.services import shift_plan_totals
from trips.dispatcher_header import build_dispatcher_header_context
from trips.models import DispatcherActionLog, DispatcherActionType, Trip, TripStatus
from users.models import EmployeeAccess

from .forms import PilotFeedbackForm
from .models import PilotFeedback, ReportTemplate, ReportType
from .shift_analytics import build_excavator_dynamics, build_shift_analytics


STATUS_COLOR_GROUPS = {'gray', 'yellow', 'green', 'blue', 'orange', 'red'}
STATUS_COLOR_SEVERITY = {'red': 0, 'orange': 1, 'yellow': 2, 'blue': 3, 'green': 4, 'gray': 5}
RED_DOWNTIME_STATE_CODES = {'downtime', 'breakdown', 'conflict'}


def downtime_reason_color_group(reason):
    color_group = getattr(reason, 'effective_color_group', '') or 'yellow'
    return color_group if color_group in STATUS_COLOR_GROUPS else 'yellow'


def downtime_reason_state_code(reason):
    return getattr(reason, 'effective_equipment_state_code', '') or 'waiting'


def is_red_downtime_reason(reason):
    return (
        downtime_reason_color_group(reason) == 'red'
        or downtime_reason_state_code(reason) in RED_DOWNTIME_STATE_CODES
    )


def highest_priority_color(color_counts):
    colors = [color for color in color_counts.keys() if color in STATUS_COLOR_GROUPS]
    if not colors:
        return 'yellow'
    return min(colors, key=lambda color: STATUS_COLOR_SEVERITY.get(color, 99))


def calculate_trip_deviation(trip):
    if trip.planned_volume_m3 is None and trip.volume_m3 is None:
        return ''
    return (trip.volume_m3 or 0) - (trip.planned_volume_m3 or 0)


def calculate_trip_plan_completion_percent(trip):
    if not trip.planned_volume_m3 or trip.volume_m3 is None:
        return ''
    return ((trip.volume_m3 / trip.planned_volume_m3) * Decimal('100')).quantize(Decimal('0.01'))


VOLUME_REPORT_COLUMNS = {
    'truck': ('Самосвал', lambda trip: str(trip.truck)),
    'excavator': ('Экскаватор', lambda trip: str(trip.excavator)),
    'rock_type': ('Порода', lambda trip: str(trip.rock_type)),
    'dump_point': ('Точка разгрузки', lambda trip: str(trip.dump_point)),
    'planned_volume_m3': ('План, м3', lambda trip: trip.planned_volume_m3 or ''),
    'volume_m3': ('Объем, м3', lambda trip: trip.volume_m3 or ''),
    'deviation_m3': ('Отклонение, м3', calculate_trip_deviation),
    'plan_completion_percent': ('Выполнение, %', calculate_trip_plan_completion_percent),
    'tonnage': ('Тоннаж', lambda trip: trip.tonnage or ''),
    'loading_horizon': ('Горизонт', lambda trip: trip.loading_horizon),
    'loading_block': ('Блок', lambda trip: trip.loading_block),
    'transport_distance_km': ('Плечо, км', lambda trip: trip.transport_distance_km or ''),
    'downtime_text': ('Простои', lambda trip: trip.downtime_text),
    'note': ('Примечание', lambda trip: trip.note),
    'loading_shift': ('Смена загрузки', lambda trip: trip.loading_shift.get_shift_type_display() if trip.loading_shift else ''),
    'unloading_shift': ('Смена разгрузки', lambda trip: trip.unloading_shift.get_shift_type_display() if trip.unloading_shift else ''),
    'is_carryover': ('Переходящий рейс', lambda trip: 'Да' if trip.is_carryover else 'Нет'),
    'completed_at': ('Выполнен', lambda trip: trip.completed_at.strftime('%d.%m.%Y %H:%M') if trip.completed_at else ''),
}

DEFAULT_VOLUME_REPORT_COLUMNS = [
    'truck',
    'excavator',
    'rock_type',
    'dump_point',
    'planned_volume_m3',
    'volume_m3',
    'deviation_m3',
    'plan_completion_percent',
    'tonnage',
    'loading_horizon',
    'loading_block',
    'transport_distance_km',
    'loading_shift',
    'unloading_shift',
    'is_carryover',
    'completed_at',
]

REPORT_TEMPLATE_FILTER_FIELDS = [
    'date_from',
    'date_to',
    'loading_shift_type',
    'unloading_shift_type',
    'carryover',
    'truck',
    'excavator',
    'rock_type',
    'dump_point',
]

VOLUME_REPORT_FILTER_LABELS = {
    'date_from': 'Дата с',
    'date_to': 'Дата по',
    'loading_shift_type': 'Смена загрузки',
    'unloading_shift_type': 'Смена разгрузки',
    'carryover': 'Переходящий рейс',
    'truck': 'Самосвал',
    'excavator': 'Экскаватор',
    'rock_type': 'Порода/груз',
    'dump_point': 'Точка разгрузки',
}

SHIFT_TYPE_LABELS = {
    'day': 'Дневная',
    'night': 'Ночная',
}

UNLOADING_WAITING_REASONS = {
    'ожидание разгрузки ккд': 'ККД',
    'ожидание разгрузки скдр': 'СКДР',
}

CARRYOVER_LABELS = {
    'yes': 'Да',
    'no': 'Нет',
}

VOLUME_REPORT_GROUPS = {
    'truck': ('Самосвал', lambda trip: str(trip.truck)),
    'excavator': ('Экскаватор', lambda trip: str(trip.excavator)),
    'rock_type': ('Порода/груз', lambda trip: str(trip.rock_type)),
    'dump_point': ('Точка разгрузки', lambda trip: str(trip.dump_point)),
    'completed_hour': (
        'Час выполнения рейса',
        lambda trip: timezone.localtime(trip.completed_at).strftime('%H:00') if trip.completed_at else 'не задано',
    ),
    'loading_shift': ('Смена загрузки', lambda trip: trip.loading_shift.get_shift_type_display() if trip.loading_shift else 'не задано'),
    'unloading_shift': ('Смена разгрузки', lambda trip: trip.unloading_shift.get_shift_type_display() if trip.unloading_shift else 'не задано'),
}


def get_object_label(model, object_id):
    if not str(object_id).isdigit():
        return str(object_id)
    instance = model.objects.filter(id=object_id).first()
    return str(instance) if instance else str(object_id)


def get_filter_display_value(key, value):
    value = str(value or '').strip()
    if not value:
        return ''
    if key in {'loading_shift_type', 'unloading_shift_type'}:
        return SHIFT_TYPE_LABELS.get(value, value)
    if key == 'carryover':
        return CARRYOVER_LABELS.get(value, value)
    if key == 'truck':
        return get_object_label(Equipment, value)
    if key == 'excavator':
        return get_object_label(Equipment, value)
    if key == 'rock_type':
        return get_object_label(RockType, value)
    if key == 'dump_point':
        return get_object_label(DumpPoint, value)
    return value


def get_active_filter_rows(filters):
    rows = []
    for key in REPORT_TEMPLATE_FILTER_FIELDS:
        display_value = get_filter_display_value(key, filters.get(key, ''))
        if display_value:
            rows.append([VOLUME_REPORT_FILTER_LABELS[key], display_value])
    return rows


def get_group_by_label(group_by):
    if group_by in VOLUME_REPORT_GROUPS:
        return VOLUME_REPORT_GROUPS[group_by][0]
    return 'Без группировки'


def get_volume_report_templates():
    return ReportTemplate.objects.filter(is_active=True).order_by('name')


def get_selected_report_template(request):
    template_id = request.GET.get('template', '').strip()
    if not template_id:
        return None
    return ReportTemplate.objects.filter(id=template_id, is_active=True).first()


def get_selected_columns(template):
    if not template:
        return DEFAULT_VOLUME_REPORT_COLUMNS
    columns = [column for column in template.columns if column in VOLUME_REPORT_COLUMNS]
    return columns or DEFAULT_VOLUME_REPORT_COLUMNS


def get_template_column_labels(template):
    if not template:
        return {}
    return template.column_labels or {}


def get_column_label(column, column_labels=None):
    default_label = VOLUME_REPORT_COLUMNS[column][0]
    if not column_labels:
        return default_label
    custom_label = str(column_labels.get(column, '')).strip()
    return custom_label or default_label


def report_template_column_options(selected_columns, column_labels=None):
    return [
        {
            'code': code,
            'label': label,
            'custom_label': str((column_labels or {}).get(code, '')).strip(),
            'checked': code in selected_columns,
        }
        for code, (label, _getter) in VOLUME_REPORT_COLUMNS.items()
    ]


def report_template_group_options(selected_group_by):
    return [
        {
            'code': code,
            'label': label,
            'selected': code == selected_group_by,
        }
        for code, (label, _getter) in VOLUME_REPORT_GROUPS.items()
    ]


def get_template_filters(template):
    if not template:
        return {}
    return template.filters or {}


def get_effective_filter_value(request, template, key):
    request_value = request.GET.get(key, '').strip()
    if request_value:
        return request_value
    return str(get_template_filters(template).get(key, '')).strip()


def get_selected_group_by(request, template):
    request_value = request.GET.get('group_by')
    if request_value is not None:
        request_value = request_value.strip()
        if request_value == 'none':
            return ''
        return request_value if request_value in VOLUME_REPORT_GROUPS else ''
    if template and template.group_by in VOLUME_REPORT_GROUPS:
        return template.group_by
    return ''


def get_volume_report_filters(request, template=None):
    return {
        key: get_effective_filter_value(request, template, key)
        for key in REPORT_TEMPLATE_FILTER_FIELDS
    }


def parse_filter_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return None


def get_report_access(request):
    access_id = request.session.get('employee_access_id')
    if not access_id:
        return None
    return EmployeeAccess.objects.select_related('employee', 'role').filter(id=access_id, is_active=True).first()


def require_dispatcher_report_access(request):
    access = get_report_access(request)
    if not access:
        return None, redirect('login')
    if access.role.code not in {'dispatcher', 'admin', 'manager'}:
        return None, redirect('role_home')
    return access, None


def decimal_total(values):
    total = Decimal('0')
    for value in values:
        if value is not None:
            total += value
    return total


def percent(part, total):
    if not total:
        return Decimal('0')
    return ((part / total) * Decimal('100')).quantize(Decimal('0.1'))


def format_volume(value):
    value = value or Decimal('0')
    return f'{int(value):,}'.replace(',', ' ')


def dashboard_status_by_percent(value):
    if value < Decimal('70'):
        return 'danger'
    if value < Decimal('90'):
        return 'risk'
    return 'ok'


def dashboard_status_by_deviation(value):
    if value < Decimal('0'):
        return 'danger'
    if value == Decimal('0'):
        return 'risk'
    return 'ok'


def dispatcher_mining_filters(request):
    selected_date = parse_filter_date(request.GET.get('date')) or timezone.localdate()
    shift_type = request.GET.get('shift_type', '').strip()
    if shift_type not in {'', 'day', 'night'}:
        shift_type = ''
    return {
        'date': selected_date,
        'date_value': selected_date.strftime('%Y-%m-%d'),
        'shift_type': shift_type,
        'query_string': request.GET.urlencode(),
    }


def dispatcher_mining_trip_queryset(filters):
    trips = Trip.objects.filter(status=TripStatus.COMPLETED).select_related(
        'truck',
        'excavator',
        'rock_type',
        'dump_point',
        'loading_shift',
        'unloading_shift',
    )
    trips = trips.filter(completed_at__date=filters['date'])
    if filters['shift_type']:
        trips = trips.filter(loading_shift__shift_type=filters['shift_type'])
    return trips.order_by('-completed_at')


def aggregate_dispatcher_rows(trips, key_getter, label_getter):
    grouped = {}
    for trip in trips:
        key = key_getter(trip) or 'not-set'
        if key not in grouped:
            grouped[key] = {
                'label': label_getter(trip) or 'Не указано',
                'volume': Decimal('0'),
                'tonnage': Decimal('0'),
                'planned': Decimal('0'),
                'trips': 0,
            }
        grouped[key]['volume'] += trip.volume_m3 or Decimal('0')
        grouped[key]['tonnage'] += trip.tonnage or Decimal('0')
        grouped[key]['planned'] += trip.planned_volume_m3 or Decimal('0')
        grouped[key]['trips'] += 1

    rows = sorted(grouped.values(), key=lambda item: item['volume'], reverse=True)
    max_volume = rows[0]['volume'] if rows else Decimal('0')
    total_volume = decimal_total(row['volume'] for row in rows)
    for row in rows:
        row['volume_display'] = format_volume(row['volume'])
        row['tonnage_display'] = format_volume(row['tonnage'])
        row['share'] = percent(row['volume'], total_volume)
        row['bar'] = int(percent(row['volume'], max_volume)) if max_volume else 0
        row['deviation'] = row['volume'] - row['planned']
        row['status'] = dashboard_status_by_deviation(row['deviation'])
    return rows


def dispatcher_hourly_rows(trips):
    grouped = defaultdict(lambda: {'label': '', 'volume': Decimal('0'), 'trips': 0})
    for trip in trips:
        hour = timezone.localtime(trip.completed_at).replace(minute=0, second=0, microsecond=0) if trip.completed_at else None
        key = hour.strftime('%H:00') if hour else 'Не указано'
        grouped[key]['label'] = key
        grouped[key]['volume'] += trip.volume_m3 or Decimal('0')
        grouped[key]['trips'] += 1
    rows = sorted(grouped.values(), key=lambda item: item['label'])
    max_volume = max((row['volume'] for row in rows), default=Decimal('0'))
    for row in rows:
        row['volume_display'] = format_volume(row['volume'])
        row['bar'] = int(percent(row['volume'], max_volume)) if max_volume else 0
    return rows


def dispatcher_complex_rows(trips):
    grouped = {}
    for trip in trips:
        key = trip.excavator_id or 'not-set'
        if key not in grouped:
            grouped[key] = {
                'label': f'К-{trip.excavator.garage_number}' if trip.excavator else 'Без экскаватора',
                'excavator': str(trip.excavator) if trip.excavator else 'Не указан',
                'volume': Decimal('0'),
                'planned': Decimal('0'),
                'trips': 0,
                'trucks': set(),
                'dump_points': defaultdict(Decimal),
                'faces': set(),
            }
        row = grouped[key]
        row['volume'] += trip.volume_m3 or Decimal('0')
        row['planned'] += trip.planned_volume_m3 or Decimal('0')
        row['trips'] += 1
        if trip.truck:
            row['trucks'].add(trip.truck.garage_number)
        if trip.dump_point:
            row['dump_points'][trip.dump_point.name] += trip.volume_m3 or Decimal('0')
        face = ' / '.join(part for part in [trip.loading_horizon, trip.loading_block] if part)
        if face:
            row['faces'].add(face)

    rows = sorted(grouped.values(), key=lambda item: item['volume'], reverse=True)
    for row in rows:
        row['volume_display'] = format_volume(row['volume'])
        row['completion'] = percent(row['volume'], row['planned'])
        row['status'] = dashboard_status_by_percent(row['completion']) if row['planned'] else 'risk'
        row['trucks_count'] = len(row['trucks'])
        row['trucks_display'] = ', '.join(sorted(row['trucks'])) or 'Нет'
        row['faces_display'] = ', '.join(sorted(row['faces'])) or 'Не указан'
        dump_total = decimal_total(row['dump_points'].values())
        row['dump_chips'] = [
            {
                'name': name,
                'share': percent(volume, dump_total),
                'volume_display': format_volume(volume),
            }
            for name, volume in sorted(row['dump_points'].items(), key=lambda item: item[1], reverse=True)
        ]
    return rows


def dispatcher_mining_context(request, access):
    filters = dispatcher_mining_filters(request)
    trips = list(dispatcher_mining_trip_queryset(filters))
    month_start = filters['date'].replace(day=1)
    month_trips = list(
        Trip.objects.filter(
            status=TripStatus.COMPLETED,
            completed_at__date__gte=month_start,
            completed_at__date__lte=filters['date'],
        )
    )

    plan_total = decimal_total(trip.planned_volume_m3 for trip in trips)
    volume_total = decimal_total(trip.volume_m3 for trip in trips)
    tonnage_total = decimal_total(trip.tonnage for trip in trips)
    completion = percent(volume_total, plan_total)
    deviation = volume_total - plan_total
    month_volume = decimal_total(trip.volume_m3 for trip in month_trips)

    dump_rows = aggregate_dispatcher_rows(
        trips,
        lambda trip: trip.dump_point_id,
        lambda trip: trip.dump_point.name if trip.dump_point else '',
    )
    rock_rows = aggregate_dispatcher_rows(
        trips,
        lambda trip: trip.rock_type_id,
        lambda trip: trip.rock_type.name if trip.rock_type else '',
    )
    face_rows = aggregate_dispatcher_rows(
        trips,
        lambda trip: '|'.join(part for part in [trip.loading_horizon, trip.loading_block] if part),
        lambda trip: ' / '.join(part for part in [trip.loading_horizon, trip.loading_block] if part),
    )

    return {
        'access': access,
        'dispatcher_header': build_dispatcher_header_context(access),
        'filters': filters,
        'current_time': timezone.localtime(timezone.now()).strftime('%H:%M'),
        'current_date': timezone.localdate().strftime('%d.%m.%Y'),
        'shift_label': 'Дневная' if filters['shift_type'] == 'day' else 'Ночная' if filters['shift_type'] == 'night' else 'Все смены',
        'kpis': {
            'plan': format_volume(plan_total),
            'volume': format_volume(volume_total),
            'tonnage': format_volume(tonnage_total),
            'trips': len(trips),
            'completion': completion,
            'deviation': format_volume(abs(deviation)),
            'deviation_negative': deviation < 0,
            'month_volume': format_volume(month_volume),
            'complexes': len({trip.excavator_id for trip in trips if trip.excavator_id}),
            'trucks': len({trip.truck_id for trip in trips if trip.truck_id}),
        },
        'completion_status': dashboard_status_by_percent(completion),
        'dump_rows': dump_rows[:6],
        'rock_rows': rock_rows[:5],
        'face_rows': face_rows[:6],
        'hourly_rows': dispatcher_hourly_rows(trips),
        'complex_rows': dispatcher_complex_rows(trips)[:8],
        'query_string': request.GET.urlencode(),
    }


def dispatcher_mining_volumes_view(request):
    access, response = require_dispatcher_report_access(request)
    if response:
        return response
    return render(request, 'reports/dispatcher_mining_volumes.html', dispatcher_mining_context(request, access))


def write_dispatcher_mining_sheet(sheet, context):
    sheet.title = 'Горные объемы'
    sheet['A1'] = 'Горные объемы диспетчера'
    sheet['A1'].font = Font(size=16, bold=True, color='12232E')
    sheet.append(['Дата', context['filters']['date'].strftime('%d.%m.%Y')])
    sheet.append(['Смена', context['shift_label']])
    sheet.append(['Сформирован', f"{context['current_date']} {context['current_time']}"])
    sheet.append([])
    sheet.append(['План', context['kpis']['plan']])
    sheet.append(['Факт', context['kpis']['volume']])
    sheet.append(['Тоннаж', context['kpis']['tonnage']])
    sheet.append(['Рейсы', context['kpis']['trips']])
    sheet.append(['Комплексы', context['kpis']['complexes']])
    sheet.append([])

    sections = [
        ('По точкам разгрузки', context['dump_rows'], ['label', 'volume_display', 'share', 'trips']),
        ('По породе', context['rock_rows'], ['label', 'volume_display', 'share', 'trips']),
        ('По забоям', context['face_rows'], ['label', 'volume_display', 'share', 'trips']),
        ('По комплексам', context['complex_rows'], ['label', 'volume_display', 'completion', 'trucks_count', 'faces_display']),
    ]
    headers_by_key = {
        'label': 'Разрез',
        'volume_display': 'Объем',
        'share': 'Доля, %',
        'trips': 'Рейсы',
        'completion': 'Выполнение, %',
        'trucks_count': 'Самосвалы',
        'faces_display': 'Забой',
    }
    for title, rows, keys in sections:
        sheet.append([title])
        title_row = sheet.max_row
        sheet.cell(row=title_row, column=1).font = Font(bold=True, color='12232E')
        sheet.append([headers_by_key[key] for key in keys])
        header_row = sheet.max_row
        for cell in sheet[header_row]:
            cell.fill = PatternFill('solid', fgColor='12232E')
            cell.font = Font(color='FFFFFF', bold=True)
        for row in rows:
            sheet.append([row.get(key, '') for key in keys])
        sheet.append([])

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    for column_index in range(1, 8):
        sheet.column_dimensions[get_column_letter(column_index)].width = 22


def dispatcher_mining_volumes_export_view(request):
    access, response = require_dispatcher_report_access(request)
    if response:
        return response
    workbook = Workbook()
    write_dispatcher_mining_sheet(workbook.active, dispatcher_mining_context(request, access))
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="dispatcher_mining_volumes.xlsx"'
    workbook.save(response)
    return response


def format_decimal_value(value, places=1):
    if value is None:
        return '-'
    quant = Decimal('1') if places == 0 else Decimal('0.' + ('0' * (places - 1)) + '1')
    return str(value.quantize(quant)).replace('.', ',')


def dispatcher_transport_filters(request):
    selected_date = parse_filter_date(request.GET.get('date')) or timezone.localdate()
    shift_type = request.GET.get('shift_type', '').strip()
    if shift_type not in {'', 'day', 'night'}:
        shift_type = ''
    return {
        'date': selected_date,
        'date_value': selected_date.strftime('%Y-%m-%d'),
        'shift_type': shift_type,
        'query_string': request.GET.urlencode(),
    }


def dispatcher_transport_shift_queryset(filters):
    shifts = EmployeeShift.objects.select_related(
        'employee',
        'equipment',
        'equipment__equipment_type',
        'equipment__model',
    ).filter(
        equipment__equipment_type__name__icontains='Самосвал',
        opened_at__date=filters['date'],
    )
    if filters['shift_type']:
        shifts = shifts.filter(shift_type=filters['shift_type'])
    return shifts.order_by('equipment__garage_number', 'opened_at')


def dispatcher_transport_trip_stats(filters):
    trips = Trip.objects.filter(
        status=TripStatus.COMPLETED,
        completed_at__date=filters['date'],
        truck_id__isnull=False,
    )
    grouped = defaultdict(lambda: {
        'trips': 0,
        'volume': Decimal('0'),
        'tonnage': Decimal('0'),
    })
    for trip in trips:
        row = grouped[trip.truck_id]
        row['trips'] += 1
        row['volume'] += trip.volume_m3 or Decimal('0')
        row['tonnage'] += trip.tonnage or Decimal('0')
    return grouped


def value_delta(start, end, reverse=False):
    if start is None or end is None:
        return None
    return (start - end) if reverse else (end - start)


def transport_row_status(row):
    if row['has_negative_delta']:
        return 'danger'
    if row['missing_end']:
        return 'risk'
    return 'ok'


def dispatcher_transport_rows(shifts, trip_stats):
    rows = []
    for shift in shifts:
        fuel_delta = value_delta(shift.start_fuel, shift.end_fuel, reverse=True)
        mileage_delta = value_delta(shift.start_mileage, shift.end_mileage)
        hours_delta = value_delta(shift.start_engine_hours, shift.end_engine_hours)
        stats = trip_stats.get(shift.equipment_id, {'trips': 0, 'volume': Decimal('0'), 'tonnage': Decimal('0')})
        row = {
            'equipment': shift.equipment,
            'equipment_label': shift.equipment.garage_number if shift.equipment else '-',
            'model': shift.equipment.model.name if shift.equipment and shift.equipment.model else 'не указана',
            'driver': shift.employee.full_name,
            'shift_type': shift.get_shift_type_display(),
            'opened_at': timezone.localtime(shift.opened_at).strftime('%H:%M') if shift.opened_at else '-',
            'closed_at': timezone.localtime(shift.closed_at).strftime('%H:%M') if shift.closed_at else 'открыта',
            'start_fuel': shift.start_fuel,
            'end_fuel': shift.end_fuel,
            'fuel_delta': fuel_delta,
            'start_mileage': shift.start_mileage,
            'end_mileage': shift.end_mileage,
            'mileage_delta': mileage_delta,
            'start_engine_hours': shift.start_engine_hours,
            'end_engine_hours': shift.end_engine_hours,
            'hours_delta': hours_delta,
            'missing_end': any(value is None for value in [shift.end_fuel, shift.end_mileage, shift.end_engine_hours]),
            'has_negative_delta': any(value is not None and value < 0 for value in [fuel_delta, mileage_delta, hours_delta]),
            'trips': stats['trips'],
            'volume': stats['volume'],
            'tonnage': stats['tonnage'],
        }
        row['status'] = transport_row_status(row)
        row['fuel_display'] = format_decimal_value(fuel_delta)
        row['mileage_display'] = format_decimal_value(mileage_delta)
        row['hours_display'] = format_decimal_value(hours_delta)
        row['fuel_per_km'] = format_decimal_value((fuel_delta / mileage_delta), 2) if fuel_delta is not None and mileage_delta else '-'
        row['fuel_per_hour'] = format_decimal_value((fuel_delta / hours_delta), 2) if fuel_delta is not None and hours_delta else '-'
        row['volume_display'] = format_volume(row['volume'])
        rows.append(row)
    return rows


def dispatcher_transport_context(request, access):
    filters = dispatcher_transport_filters(request)
    shifts = list(dispatcher_transport_shift_queryset(filters))
    rows = dispatcher_transport_rows(shifts, dispatcher_transport_trip_stats(filters))
    fuel_total = decimal_total(row['fuel_delta'] for row in rows if row['fuel_delta'] is not None and row['fuel_delta'] >= 0)
    mileage_total = decimal_total(row['mileage_delta'] for row in rows if row['mileage_delta'] is not None and row['mileage_delta'] >= 0)
    hours_total = decimal_total(row['hours_delta'] for row in rows if row['hours_delta'] is not None and row['hours_delta'] >= 0)
    trip_total = sum(row['trips'] for row in rows)
    missing_count = sum(1 for row in rows if row['missing_end'])
    anomaly_count = sum(1 for row in rows if row['has_negative_delta'])
    volume_total = decimal_total(row['volume'] for row in rows)

    sorted_by_fuel = sorted(
        [row for row in rows if row['fuel_delta'] is not None and row['fuel_delta'] >= 0],
        key=lambda row: row['fuel_delta'],
        reverse=True,
    )
    max_fuel = sorted_by_fuel[0]['fuel_delta'] if sorted_by_fuel else Decimal('0')
    for row in rows:
        row['fuel_bar'] = int(percent(row['fuel_delta'], max_fuel)) if row['fuel_delta'] is not None and row['fuel_delta'] >= 0 and max_fuel else 0

    return {
        'access': access,
        'dispatcher_header': build_dispatcher_header_context(access),
        'filters': filters,
        'current_time': timezone.localtime(timezone.now()).strftime('%H:%M'),
        'current_date': timezone.localdate().strftime('%d.%m.%Y'),
        'shift_label': 'Дневная' if filters['shift_type'] == 'day' else 'Ночная' if filters['shift_type'] == 'night' else 'Все смены',
        'kpis': {
            'shifts': len(rows),
            'closed': sum(1 for row in rows if not row['missing_end']),
            'missing': missing_count,
            'anomalies': anomaly_count,
            'fuel': format_decimal_value(fuel_total),
            'mileage': format_decimal_value(mileage_total),
            'hours': format_decimal_value(hours_total),
            'trips': trip_total,
            'volume': format_volume(volume_total),
            'fuel_per_km': format_decimal_value((fuel_total / mileage_total), 2) if mileage_total else '-',
            'fuel_per_hour': format_decimal_value((fuel_total / hours_total), 2) if hours_total else '-',
        },
        'rows': rows,
        'fuel_leaders': sorted_by_fuel[:6],
        'attention_rows': [row for row in rows if row['status'] != 'ok'][:8],
        'query_string': request.GET.urlencode(),
    }


def dispatcher_transport_view(request):
    access, response = require_dispatcher_report_access(request)
    if response:
        return response
    return render(request, 'reports/dispatcher_transport.html', dispatcher_transport_context(request, access))


def write_dispatcher_transport_sheet(sheet, context):
    sheet.title = 'Автотранспорт'
    sheet['A1'] = 'Автотранспорт диспетчера'
    sheet['A1'].font = Font(size=16, bold=True, color='12232E')
    sheet.append(['Дата', context['filters']['date'].strftime('%d.%m.%Y')])
    sheet.append(['Смена', context['shift_label']])
    sheet.append(['Сформирован', f"{context['current_date']} {context['current_time']}"])
    sheet.append([])
    for label, key in [
        ('Смены', 'shifts'),
        ('Закрыто без пропусков', 'closed'),
        ('Требуют внимания', 'missing'),
        ('Аномалии', 'anomalies'),
        ('Топливо', 'fuel'),
        ('Пробег', 'mileage'),
        ('Моточасы', 'hours'),
        ('Рейсы', 'trips'),
        ('Объем', 'volume'),
    ]:
        sheet.append([label, context['kpis'][key]])
    sheet.append([])
    headers = [
        'Самосвал',
        'Водитель',
        'Смена',
        'Топливо начало',
        'Топливо конец',
        'Расход',
        'Пробег',
        'Моточасы',
        'Рейсы',
        'Объем',
        'Статус',
    ]
    sheet.append(headers)
    header_row = sheet.max_row
    for cell in sheet[header_row]:
        cell.fill = PatternFill('solid', fgColor='12232E')
        cell.font = Font(color='FFFFFF', bold=True)
    for row in context['rows']:
        sheet.append([
            row['equipment_label'],
            row['driver'],
            row['shift_type'],
            row['start_fuel'],
            row['end_fuel'],
            row['fuel_delta'],
            row['mileage_delta'],
            row['hours_delta'],
            row['trips'],
            row['volume'],
            row['status'],
        ])
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    for column_index in range(1, 12):
        sheet.column_dimensions[get_column_letter(column_index)].width = 18


def dispatcher_transport_export_view(request):
    access, response = require_dispatcher_report_access(request)
    if response:
        return response
    workbook = Workbook()
    write_dispatcher_transport_sheet(workbook.active, dispatcher_transport_context(request, access))
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="dispatcher_transport.xlsx"'
    workbook.save(response)
    return response


def dispatcher_downtime_filters(request):
    selected_date = parse_filter_date(request.GET.get('date')) or timezone.localdate()
    status = request.GET.get('status', '').strip()
    if status not in {'', 'open', 'critical', 'closed'}:
        status = ''
    return {
        'date': selected_date,
        'date_value': selected_date.strftime('%Y-%m-%d'),
        'status': status,
        'query_string': request.GET.urlencode(),
    }


def dispatcher_downtime_queryset(filters):
    events = DowntimeEvent.objects.select_related(
        'equipment',
        'equipment__equipment_type',
        'reason',
        'reason__equipment_state',
        'employee',
    ).filter(started_at__date=filters['date'])
    if filters['status'] == 'open':
        events = events.filter(ended_at__isnull=True)
    elif filters['status'] == 'closed':
        events = events.filter(ended_at__isnull=False)
    return events.order_by('-started_at')


def dispatcher_downtime_row(event):
    duration_hours = downtime_duration_hours(event)
    color_group = downtime_reason_color_group(event.reason)
    state_code = downtime_reason_state_code(event.reason)
    is_critical = is_red_downtime_reason(event.reason)
    return {
        'started_at': event.started_at,
        'ended_at': event.ended_at,
        'started_at_display': timezone.localtime(event.started_at).strftime('%H:%M'),
        'ended_at_display': timezone.localtime(event.ended_at).strftime('%H:%M') if event.ended_at else 'открыт',
        'equipment': str(event.equipment),
        'equipment_number': event.equipment.garage_number if event.equipment else '-',
        'equipment_type': event.equipment.equipment_type.name if event.equipment and event.equipment.equipment_type else 'не указан',
        'reason': event.reason.name,
        'is_critical': is_critical,
        'status': color_group,
        'color_group': color_group,
        'equipment_state_code': state_code,
        'status_label': event.reason.equipment_state.label if event.reason.equipment_state_id and event.reason.equipment_state else event.reason.name,
        'duration_hours': duration_hours,
        'duration_display': format_decimal_value(duration_hours, 2),
        'employee': event.employee.full_name if event.employee else '-',
        'comment': event.comment or '-',
    }


def dispatcher_downtime_summary(rows, key):
    grouped = {}
    for row in rows:
        name = row[key] or 'не указано'
        if name not in grouped:
            grouped[name] = {
                'name': name,
                'count': 0,
                'open_count': 0,
                'critical_count': 0,
                'duration_hours': Decimal('0'),
                'color_counts': defaultdict(int),
            }
        grouped[name]['count'] += 1
        grouped[name]['duration_hours'] += row['duration_hours']
        grouped[name]['color_counts'][row.get('color_group') or row.get('status') or 'yellow'] += 1
        if row['ended_at'] is None:
            grouped[name]['open_count'] += 1
        if row['is_critical']:
            grouped[name]['critical_count'] += 1
    result = sorted(grouped.values(), key=lambda item: (item['open_count'], item['critical_count'], item['duration_hours']), reverse=True)
    max_duration = result[0]['duration_hours'] if result else Decimal('0')
    for item in result:
        item['duration_display'] = format_decimal_value(item['duration_hours'], 2)
        item['bar'] = int(percent(item['duration_hours'], max_duration)) if max_duration else 0
        item['status'] = highest_priority_color(item['color_counts'])
    return result


def dispatcher_downtime_context(request, access):
    filters = dispatcher_downtime_filters(request)
    rows = [dispatcher_downtime_row(event) for event in dispatcher_downtime_queryset(filters)]
    if filters['status'] == 'critical':
        rows = [row for row in rows if row['is_critical']]
    total_duration_hours = decimal_total(row['duration_hours'] for row in rows)
    open_rows = [row for row in rows if row['ended_at'] is None]
    critical_rows = [row for row in rows if row['is_critical']]
    closed_rows = [row for row in rows if row['ended_at'] is not None]
    avg_duration = (total_duration_hours / Decimal(len(rows))).quantize(Decimal('0.01')) if rows else Decimal('0')

    reason_rows = dispatcher_downtime_summary(rows, 'reason')
    equipment_rows = dispatcher_downtime_summary(rows, 'equipment')
    equipment_type_rows = dispatcher_downtime_summary(rows, 'equipment_type')

    return {
        'access': access,
        'dispatcher_header': build_dispatcher_header_context(access),
        'filters': filters,
        'current_time': timezone.localtime(timezone.now()).strftime('%H:%M'),
        'current_date': timezone.localdate().strftime('%d.%m.%Y'),
        'shift_label': filters['date'].strftime('%d.%m.%Y'),
        'kpis': {
            'events': len(rows),
            'open': len(open_rows),
            'closed': len(closed_rows),
            'critical': len(critical_rows),
            'hours': format_decimal_value(total_duration_hours, 2),
            'average': format_decimal_value(avg_duration, 2),
            'equipment': len({row['equipment'] for row in rows}),
        },
        'status_choices': [
            {'code': '', 'label': 'Все'},
            {'code': 'open', 'label': 'Открытые'},
            {'code': 'critical', 'label': 'Критические'},
            {'code': 'closed', 'label': 'Закрытые'},
        ],
        'reason_rows': reason_rows[:8],
        'equipment_rows': equipment_rows[:8],
        'equipment_type_rows': equipment_type_rows[:5],
        'open_rows': open_rows[:8],
        'critical_rows': critical_rows[:8],
        'rows': rows[:120],
        'query_string': request.GET.urlencode(),
    }


def dispatcher_downtimes_view(request):
    access, response = require_dispatcher_report_access(request)
    if response:
        return response
    return render(request, 'reports/dispatcher_downtimes.html', dispatcher_downtime_context(request, access))


def write_dispatcher_downtimes_sheet(sheet, context):
    sheet.title = 'Простои'
    sheet['A1'] = 'Простои и отклонения диспетчера'
    sheet['A1'].font = Font(size=16, bold=True, color='12232E')
    sheet.append(['Дата', context['filters']['date'].strftime('%d.%m.%Y')])
    sheet.append(['Статус', next((item['label'] for item in context['status_choices'] if item['code'] == context['filters']['status']), 'Все')])
    sheet.append(['Сформирован', f"{context['current_date']} {context['current_time']}"])
    sheet.append([])
    for label, key in [
        ('События', 'events'),
        ('Открытые', 'open'),
        ('Закрытые', 'closed'),
        ('Критические', 'critical'),
        ('Техника', 'equipment'),
        ('Потери, ч', 'hours'),
        ('Средний простой, ч', 'average'),
    ]:
        sheet.append([label, context['kpis'][key]])
    sheet.append([])

    def write_summary(title, rows):
        sheet.append([title])
        title_row = sheet.max_row
        sheet.cell(row=title_row, column=1).font = Font(bold=True, color='12232E')
        sheet.append(['Разрез', 'События', 'Открытые', 'Критические', 'Потери, ч'])
        header_row = sheet.max_row
        for cell in sheet[header_row]:
            cell.fill = PatternFill('solid', fgColor='12232E')
            cell.font = Font(color='FFFFFF', bold=True)
        for row in rows:
            sheet.append([row['name'], row['count'], row['open_count'], row['critical_count'], row['duration_display']])
        sheet.append([])

    write_summary('По причинам', context['reason_rows'])
    write_summary('По технике', context['equipment_rows'])
    write_summary('По видам техники', context['equipment_type_rows'])

    sheet.append(['Начало', 'Окончание', 'Статус', 'Техника', 'Вид', 'Причина', 'Потери, ч', 'Кто зафиксировал', 'Комментарий'])
    header_row = sheet.max_row
    for cell in sheet[header_row]:
        cell.fill = PatternFill('solid', fgColor='12232E')
        cell.font = Font(color='FFFFFF', bold=True)
    for row in context['rows']:
        sheet.append([
            timezone.localtime(row['started_at']).strftime('%d.%m.%Y %H:%M'),
            timezone.localtime(row['ended_at']).strftime('%d.%m.%Y %H:%M') if row['ended_at'] else 'открыт',
            row['status_label'],
            row['equipment'],
            row['equipment_type'],
            row['reason'],
            row['duration_display'],
            row['employee'],
            row['comment'],
        ])

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    for column_index in range(1, 10):
        sheet.column_dimensions[get_column_letter(column_index)].width = 20


def dispatcher_downtimes_export_view(request):
    access, response = require_dispatcher_report_access(request)
    if response:
        return response
    workbook = Workbook()
    write_dispatcher_downtimes_sheet(workbook.active, dispatcher_downtime_context(request, access))
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="dispatcher_downtimes.xlsx"'
    workbook.save(response)
    return response


def dispatcher_shift_log_filters(request):
    selected_date = parse_filter_date(request.GET.get('date')) or timezone.localdate()
    event_type = request.GET.get('event_type', '').strip()
    if event_type not in {'', 'dispatcher', 'downtime', 'trip'}:
        event_type = ''
    return {
        'date': selected_date,
        'date_value': selected_date.strftime('%Y-%m-%d'),
        'event_type': event_type,
        'query_string': request.GET.urlencode(),
    }


def dispatcher_action_status(action):
    if action.action_type == DispatcherActionType.CANCEL_TRIP:
        return 'danger'
    if action.action_type in {DispatcherActionType.CANCEL_ASSIGNMENT, DispatcherActionType.SERVICE_CLOSE_SHIFT}:
        return 'risk'
    return 'ok'


def dispatcher_trip_time(trip):
    return trip.completed_at or trip.created_at


def dispatcher_trip_status(trip):
    if trip.status == TripStatus.CANCELLED:
        return 'danger'
    if trip.status == TripStatus.ACTIVE:
        return 'info'
    return 'ok'


def dispatcher_trip_title(trip):
    if trip.status == TripStatus.CANCELLED:
        return 'Рейс отменен'
    if trip.status == TripStatus.ACTIVE:
        return 'Рейс открыт'
    return 'Рейс выполнен'


def dispatcher_trip_route(trip):
    loading = ' / '.join(part for part in [trip.loading_horizon, trip.loading_block] if part) or '-'
    return f'{loading} -> {trip.dump_point}'


def dispatcher_shift_action_rows(filters):
    actions = DispatcherActionLog.objects.select_related(
        'actor',
        'trip',
        'trip__truck',
        'trip__excavator',
        'trip__dump_point',
        'trip__rock_type',
    ).filter(created_at__date=filters['date']).order_by('-created_at')
    return [
        {
            'time': action.created_at,
            'time_display': timezone.localtime(action.created_at).strftime('%H:%M'),
            'type': 'dispatcher',
            'type_label': 'действие',
            'status': dispatcher_action_status(action),
            'title': action.get_action_type_display(),
            'summary': action.target_summary,
            'reason': action.reason or '-',
            'equipment': str(action.trip.truck) if action.trip else '-',
            'actor': action.actor.full_name if action.actor else '-',
            'route': dispatcher_trip_route(action.trip) if action.trip else '-',
            'volume': format_volume(action.trip.volume_m3) + ' м3' if action.trip and action.trip.volume_m3 else '-',
        }
        for action in actions
    ]


def dispatcher_shift_downtime_rows(filters):
    events = DowntimeEvent.objects.select_related(
        'equipment',
        'equipment__equipment_type',
        'reason',
        'reason__equipment_state',
        'employee',
    ).filter(started_at__date=filters['date']).order_by('-started_at')
    rows = []
    for event in events:
        duration = downtime_duration_hours(event)
        status = downtime_reason_color_group(event.reason)
        rows.append({
            'time': event.started_at,
            'time_display': timezone.localtime(event.started_at).strftime('%H:%M'),
            'type': 'downtime',
            'type_label': 'простой',
            'status': status,
            'title': event.reason.name,
            'summary': event.comment or ('Открытый простой' if event.ended_at is None else 'Простой закрыт'),
            'reason': event.reason.name,
            'equipment': str(event.equipment),
            'actor': event.employee.full_name if event.employee else '-',
            'route': event.equipment.equipment_type.name if event.equipment and event.equipment.equipment_type else '-',
            'volume': f'{format_decimal_value(duration, 2)} ч',
        })
    return rows


def dispatcher_shift_trip_rows(filters):
    trips = Trip.objects.select_related(
        'truck',
        'excavator',
        'rock_type',
        'dump_point',
        'driver',
        'excavator_operator',
    ).filter(completed_at__date=filters['date']).order_by('-completed_at')
    rows = []
    for trip in trips:
        rows.append({
            'time': dispatcher_trip_time(trip),
            'time_display': timezone.localtime(dispatcher_trip_time(trip)).strftime('%H:%M'),
            'type': 'trip',
            'type_label': 'рейс',
            'status': dispatcher_trip_status(trip),
            'title': dispatcher_trip_title(trip),
            'summary': f'{trip.truck} из-под {trip.excavator} на {trip.dump_point}',
            'reason': str(trip.rock_type),
            'equipment': f'{trip.truck} / {trip.excavator}',
            'actor': trip.driver.full_name if trip.driver else trip.excavator_operator.full_name if trip.excavator_operator else '-',
            'route': dispatcher_trip_route(trip),
            'volume': format_volume(trip.volume_m3) + ' м3' if trip.volume_m3 else '-',
        })
    return rows


def dispatcher_shift_log_context(request, access):
    filters = dispatcher_shift_log_filters(request)
    action_rows = dispatcher_shift_action_rows(filters)
    downtime_rows = dispatcher_shift_downtime_rows(filters)
    trip_rows = dispatcher_shift_trip_rows(filters)
    rows = action_rows + downtime_rows + trip_rows
    if filters['event_type']:
        rows = [row for row in rows if row['type'] == filters['event_type']]
    rows = sorted(rows, key=lambda row: row['time'], reverse=True)
    critical_downtimes = [row for row in downtime_rows if row['status'] == 'red']
    service_actions = [row for row in action_rows if row['status'] in {'danger', 'risk'}]
    return {
        'access': access,
        'dispatcher_header': build_dispatcher_header_context(access),
        'filters': filters,
        'current_time': timezone.localtime(timezone.now()).strftime('%H:%M'),
        'current_date': timezone.localdate().strftime('%d.%m.%Y'),
        'shift_label': filters['date'].strftime('%d.%m.%Y'),
        'kpis': {
            'events': len(rows),
            'actions': len(action_rows),
            'downtimes': len(downtime_rows),
            'critical_downtimes': len(critical_downtimes),
            'trips': len(trip_rows),
            'service_actions': len(service_actions),
        },
        'event_type_choices': [
            {'code': '', 'label': 'Все события'},
            {'code': 'dispatcher', 'label': 'Действия'},
            {'code': 'downtime', 'label': 'Простои'},
            {'code': 'trip', 'label': 'Рейсы'},
        ],
        'rows': rows[:180],
        'action_rows': action_rows[:12],
        'downtime_rows': downtime_rows[:12],
        'trip_rows': trip_rows[:12],
        'query_string': request.GET.urlencode(),
    }


def dispatcher_shift_log_view(request):
    access, response = require_dispatcher_report_access(request)
    if response:
        return response
    return render(request, 'reports/dispatcher_shift_log.html', dispatcher_shift_log_context(request, access))


def write_dispatcher_shift_log_sheet(sheet, context):
    sheet.title = 'Журнал смены'
    sheet['A1'] = 'Журнал смены диспетчера'
    sheet['A1'].font = Font(size=16, bold=True, color='12232E')
    sheet.append(['Дата', context['filters']['date'].strftime('%d.%m.%Y')])
    sheet.append(['Тип событий', next((item['label'] for item in context['event_type_choices'] if item['code'] == context['filters']['event_type']), 'Все события')])
    sheet.append(['Сформирован', f"{context['current_date']} {context['current_time']}"])
    sheet.append([])
    for label, key in [
        ('События', 'events'),
        ('Действия', 'actions'),
        ('Простои', 'downtimes'),
        ('Критические простои', 'critical_downtimes'),
        ('Рейсы', 'trips'),
        ('Служебные действия', 'service_actions'),
    ]:
        sheet.append([label, context['kpis'][key]])
    sheet.append([])
    sheet.append(['Время', 'Тип', 'Статус', 'Событие', 'Описание', 'Объект', 'Причина/порода', 'Участник', 'Маршрут', 'Объем/потери'])
    header_row = sheet.max_row
    for cell in sheet[header_row]:
        cell.fill = PatternFill('solid', fgColor='12232E')
        cell.font = Font(color='FFFFFF', bold=True)
    for row in context['rows']:
        sheet.append([
            timezone.localtime(row['time']).strftime('%d.%m.%Y %H:%M'),
            row['type_label'],
            row['status'],
            row['title'],
            row['summary'],
            row['equipment'],
            row['reason'],
            row['actor'],
            row['route'],
            row['volume'],
        ])
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    for column_index in range(1, 11):
        sheet.column_dimensions[get_column_letter(column_index)].width = 20


def dispatcher_shift_log_export_view(request):
    access, response = require_dispatcher_report_access(request)
    if response:
        return response
    workbook = Workbook()
    write_dispatcher_shift_log_sheet(workbook.active, dispatcher_shift_log_context(request, access))
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="dispatcher_shift_log.xlsx"'
    workbook.save(response)
    return response


def dispatcher_reports_context(request, access):
    filters = dispatcher_mining_filters(request)
    trips = list(dispatcher_mining_trip_queryset(filters))
    transport_rows = dispatcher_transport_rows(
        list(dispatcher_transport_shift_queryset(filters)),
        dispatcher_transport_trip_stats(filters),
    )
    downtime_rows = [dispatcher_downtime_row(event) for event in dispatcher_downtime_queryset({'date': filters['date'], 'status': ''})]
    shift_actions_count = DispatcherActionLog.objects.filter(created_at__date=filters['date']).count()
    active_templates = list(ReportTemplate.objects.filter(is_active=True).order_by('name')[:8])

    plan_total = decimal_total(trip.planned_volume_m3 for trip in trips)
    volume_total = decimal_total(trip.volume_m3 for trip in trips)
    tonnage_total = decimal_total(trip.tonnage for trip in trips)
    completion = percent(volume_total, plan_total)
    missing_transport_count = sum(1 for row in transport_rows if row['missing_end'])
    anomaly_transport_count = sum(1 for row in transport_rows if row['has_negative_delta'])
    open_downtime_count = sum(1 for row in downtime_rows if row['ended_at'] is None)
    critical_downtime_count = sum(1 for row in downtime_rows if row['is_critical'])
    shift_log_events_count = len(trips) + len(downtime_rows) + shift_actions_count

    def report_status(kind):
        if kind == 'mining':
            if not trips:
                return 'risk'
            return dashboard_status_by_percent(completion)
        if kind == 'transport':
            if anomaly_transport_count:
                return 'danger'
            if missing_transport_count or not transport_rows:
                return 'risk'
            return 'ok'
        if kind == 'downtimes':
            if critical_downtime_count:
                return 'danger'
            if open_downtime_count:
                return 'risk'
            return 'ok'
        if kind == 'log':
            return 'ok' if shift_log_events_count else 'risk'
        if kind == 'templates':
            return 'ok' if active_templates else 'risk'
        return 'ok'

    status_labels = {
        'ok': 'готов',
        'risk': 'проверить',
        'danger': 'критично',
    }

    report_tiles = [
        {
            'title': 'Сменные объемы',
            'kind': 'mining',
            'status': report_status('mining'),
            'primary': f'{format_volume(volume_total)} м3',
            'secondary': f'{len(trips)} рейс. / {len({trip.truck_id for trip in trips if trip.truck_id})} самосв.',
            'readiness': f'План {format_volume(plan_total)} м3 / {completion}%',
            'view_url': reverse('dispatcher_mining_volumes'),
            'export_url': reverse('dispatcher_mining_volumes_export'),
        },
        {
            'title': 'Автотранспорт',
            'kind': 'transport',
            'status': report_status('transport'),
            'primary': f'{len(transport_rows)} смен',
            'secondary': f'{missing_transport_count} без закрытия / {anomaly_transport_count} аном.',
            'readiness': f'{sum(1 for row in transport_rows if not row["missing_end"])} закрыто',
            'view_url': reverse('dispatcher_transport'),
            'export_url': reverse('dispatcher_transport_export'),
        },
        {
            'title': 'Простои',
            'kind': 'downtimes',
            'status': report_status('downtimes'),
            'primary': f'{len(downtime_rows)} событий',
            'secondary': f'{open_downtime_count} открыто / {critical_downtime_count} крит.',
            'readiness': f'{format_decimal_value(decimal_total(row["duration_hours"] for row in downtime_rows), 2)} ч потерь',
            'view_url': reverse('dispatcher_downtimes'),
            'export_url': reverse('dispatcher_downtimes_export'),
        },
        {
            'title': 'Журнал смены',
            'kind': 'log',
            'status': report_status('log'),
            'primary': f'{shift_log_events_count} событий',
            'secondary': f'{shift_actions_count} действий / {len(trips)} рейс.',
            'readiness': 'Хронология смены',
            'view_url': reverse('dispatcher_shift_log'),
            'export_url': reverse('dispatcher_shift_log_export'),
        },
        {
            'title': 'Суточный заказчику',
            'kind': 'customer',
            'status': report_status('mining'),
            'primary': f'{format_volume(tonnage_total)} т',
            'secondary': 'Официальная выгрузка',
            'readiness': 'По завершенным рейсам',
            'view_url': reverse('customer_daily_report'),
            'export_url': reverse('customer_daily_report_export'),
        },
        {
            'title': 'Конструктор',
            'kind': 'templates',
            'status': report_status('templates'),
            'primary': f'{len(active_templates)} шабл.',
            'secondary': 'Настройка колонок и фильтров',
            'readiness': 'Шаблоны отчетов',
            'view_url': reverse('report_template_builder'),
            'export_url': '',
        },
    ]
    for tile in report_tiles:
        tile['status_label'] = status_labels[tile['status']]

    query_suffix = f'?{request.GET.urlencode()}' if request.GET.urlencode() else ''
    return {
        'access': access,
        'dispatcher_header': build_dispatcher_header_context(access),
        'filters': filters,
        'current_time': timezone.localtime(timezone.now()).strftime('%H:%M'),
        'current_date': timezone.localdate().strftime('%d.%m.%Y'),
        'shift_label': 'Дневная' if filters['shift_type'] == 'day' else 'Ночная' if filters['shift_type'] == 'night' else 'Все смены',
        'query_string': request.GET.urlencode(),
        'query_suffix': query_suffix,
        'kpis': {
            'reports': len(report_tiles),
            'ready': sum(1 for tile in report_tiles if tile['status'] == 'ok'),
            'risk': sum(1 for tile in report_tiles if tile['status'] == 'risk'),
            'danger': sum(1 for tile in report_tiles if tile['status'] == 'danger'),
            'templates': len(active_templates),
            'volume': format_volume(volume_total),
            'trips': len(trips),
            'transport_missing': missing_transport_count,
        },
        'report_tiles': report_tiles,
        'active_templates': active_templates,
    }


def dispatcher_reports_view(request):
    access, response = require_dispatcher_report_access(request)
    if response:
        return response
    return render(request, 'reports/dispatcher_reports.html', dispatcher_reports_context(request, access))


def write_dispatcher_reports_sheet(sheet, context):
    sheet.title = 'Отчеты'
    sheet['A1'] = 'Отчеты диспетчерской'
    sheet['A1'].font = Font(size=16, bold=True, color='12232E')
    sheet.append(['Дата', context['filters']['date'].strftime('%d.%m.%Y')])
    sheet.append(['Смена', context['shift_label']])
    sheet.append(['Сформирован', f"{context['current_date']} {context['current_time']}"])
    sheet.append([])
    sheet.append(['Отчет', 'Статус', 'Готовность', 'Показатель', 'Сводка'])
    header_row = sheet.max_row
    for cell in sheet[header_row]:
        cell.fill = PatternFill('solid', fgColor='12232E')
        cell.font = Font(color='FFFFFF', bold=True)
    for tile in context['report_tiles']:
        sheet.append([tile['title'], tile['status_label'], tile['readiness'], tile['primary'], tile['secondary']])
    sheet.append([])
    sheet.append(['Активные шаблоны'])
    for template in context['active_templates']:
        sheet.append([template.name, template.get_report_type_display(), template.group_by or 'без группировки'])
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    for column_index in range(1, 7):
        sheet.column_dimensions[get_column_letter(column_index)].width = 24


def dispatcher_reports_export_view(request):
    access, response = require_dispatcher_report_access(request)
    if response:
        return response
    workbook = Workbook()
    write_dispatcher_reports_sheet(workbook.active, dispatcher_reports_context(request, access))
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="dispatcher_reports.xlsx"'
    workbook.save(response)
    return response


def dispatcher_management_attention_rows(downtime_rows, transport_rows):
    rows = []
    for row in downtime_rows:
        if row['status'] == 'ok':
            continue
        rows.append({
            'status': row['status'],
            'title': row['equipment_number'],
            'summary': row['reason'],
            'detail': f"{row['status_label']} / {row['duration_display']} ч",
        })
    for row in transport_rows:
        if row['status'] == 'ok':
            continue
        rows.append({
            'status': row['status'],
            'title': row['equipment_label'],
            'summary': 'Показания автотранспорта',
            'detail': 'нет закрытия' if row['missing_end'] else 'аномалия пробега/топлива',
        })
    status_order = {'danger': 0, 'red': 0, 'orange': 1, 'risk': 2, 'yellow': 2, 'blue': 3, 'ok': 4, 'green': 4, 'gray': 5}
    return sorted(rows, key=lambda item: status_order.get(item['status'], 3))[:8]


def dispatcher_management_context(request, access):
    filters = dispatcher_mining_filters(request)
    trips = list(dispatcher_mining_trip_queryset(filters))
    month_start = filters['date'].replace(day=1)
    month_trips = list(
        Trip.objects.filter(
            status=TripStatus.COMPLETED,
            completed_at__date__gte=month_start,
            completed_at__date__lte=filters['date'],
        )
    )
    if filters['shift_type']:
        month_trips = [trip for trip in month_trips if trip.loading_shift and trip.loading_shift.shift_type == filters['shift_type']]

    transport_rows = dispatcher_transport_rows(
        list(dispatcher_transport_shift_queryset(filters)),
        dispatcher_transport_trip_stats(filters),
    )
    downtime_rows = [dispatcher_downtime_row(event) for event in dispatcher_downtime_queryset({'date': filters['date'], 'status': ''})]
    downtime_reason_rows = dispatcher_downtime_summary(downtime_rows, 'reason')
    complex_rows = dispatcher_complex_rows(trips)
    dump_rows = aggregate_dispatcher_rows(
        trips,
        lambda trip: trip.dump_point_id,
        lambda trip: trip.dump_point.name if trip.dump_point else '',
    )
    rock_rows = aggregate_dispatcher_rows(
        trips,
        lambda trip: trip.rock_type_id,
        lambda trip: trip.rock_type.name if trip.rock_type else '',
    )
    hourly_rows = dispatcher_hourly_rows(trips)

    plan_total = decimal_total(trip.planned_volume_m3 for trip in trips)
    volume_total = decimal_total(trip.volume_m3 for trip in trips)
    tonnage_total = decimal_total(trip.tonnage for trip in trips)
    completion = percent(volume_total, plan_total)
    deviation = volume_total - plan_total
    month_volume = decimal_total(trip.volume_m3 for trip in month_trips)
    month_tonnage = decimal_total(trip.tonnage for trip in month_trips)
    open_downtime_count = sum(1 for row in downtime_rows if row['ended_at'] is None)
    critical_downtime_count = sum(1 for row in downtime_rows if row['is_critical'])
    missing_transport_count = sum(1 for row in transport_rows if row['missing_end'])
    anomaly_transport_count = sum(1 for row in transport_rows if row['has_negative_delta'])
    action_count = DispatcherActionLog.objects.filter(created_at__date=filters['date']).count()
    attention_rows = dispatcher_management_attention_rows(downtime_rows, transport_rows)

    overall_status = dashboard_status_by_percent(completion) if trips else 'risk'
    if critical_downtime_count or anomaly_transport_count:
        overall_status = 'danger'
    elif open_downtime_count or missing_transport_count:
        overall_status = 'risk'

    return {
        'access': access,
        'dispatcher_header': build_dispatcher_header_context(access),
        'filters': filters,
        'current_time': timezone.localtime(timezone.now()).strftime('%H:%M'),
        'current_date': timezone.localdate().strftime('%d.%m.%Y'),
        'shift_label': 'Дневная' if filters['shift_type'] == 'day' else 'Ночная' if filters['shift_type'] == 'night' else 'Все смены',
        'query_string': request.GET.urlencode(),
        'query_suffix': f"?{request.GET.urlencode()}" if request.GET.urlencode() else '',
        'overall_status': overall_status,
        'kpis': {
            'plan': format_volume(plan_total),
            'volume': format_volume(volume_total),
            'tonnage': format_volume(tonnage_total),
            'completion': completion,
            'deviation': format_volume(abs(deviation)),
            'deviation_negative': deviation < 0,
            'trips': len(trips),
            'complexes': len({trip.excavator_id for trip in trips if trip.excavator_id}),
            'trucks': len({trip.truck_id for trip in trips if trip.truck_id}),
            'month_volume': format_volume(month_volume),
            'month_tonnage': format_volume(month_tonnage),
            'month_trips': len(month_trips),
            'transport_rows': len(transport_rows),
            'transport_missing': missing_transport_count,
            'transport_anomaly': anomaly_transport_count,
            'downtimes': len(downtime_rows),
            'open_downtimes': open_downtime_count,
            'critical_downtimes': critical_downtime_count,
            'downtime_hours': format_decimal_value(decimal_total(row['duration_hours'] for row in downtime_rows), 2),
            'actions': action_count,
            'attention': len(attention_rows),
        },
        'complex_rows': complex_rows[:6],
        'dump_rows': dump_rows[:5],
        'rock_rows': rock_rows[:5],
        'hourly_rows': hourly_rows,
        'downtime_reason_rows': downtime_reason_rows[:5],
        'attention_rows': attention_rows,
        'export_url': reverse('dispatcher_management_export'),
    }


def dispatcher_management_view(request):
    access, response = require_dispatcher_report_access(request)
    if response:
        return response
    return render(request, 'reports/dispatcher_management.html', dispatcher_management_context(request, access))


def write_dispatcher_management_sheet(sheet, context):
    sheet.title = 'Витрина'
    sheet['A1'] = 'Витрина диспетчерской'
    sheet['A1'].font = Font(size=16, bold=True, color='12232E')
    sheet.append(['Дата', context['filters']['date'].strftime('%d.%m.%Y')])
    sheet.append(['Смена', context['shift_label']])
    sheet.append(['Сформирован', f"{context['current_date']} {context['current_time']}"])
    sheet.append([])
    sheet.append(['Итог смены'])
    sheet.append(['План', context['kpis']['plan']])
    sheet.append(['Факт', context['kpis']['volume']])
    sheet.append(['Выполнение', f"{context['kpis']['completion']}%"])
    sheet.append(['Отклонение', ('-' if context['kpis']['deviation_negative'] else '+') + context['kpis']['deviation']])
    sheet.append(['Рейсы', context['kpis']['trips']])
    sheet.append(['Самосвалы', context['kpis']['trucks']])
    sheet.append(['Комплексы', context['kpis']['complexes']])
    sheet.append([])
    sheet.append(['Комплексы', 'Экскаватор', 'Объем', 'Выполнение', 'Самосвалы', 'Забои'])
    header_row = sheet.max_row
    for cell in sheet[header_row]:
        cell.fill = PatternFill('solid', fgColor='12232E')
        cell.font = Font(color='FFFFFF', bold=True)
    for row in context['complex_rows']:
        sheet.append([row['label'], row['excavator'], row['volume_display'], f"{row['completion']}%", row['trucks_display'], row['faces_display']])
    sheet.append([])
    sheet.append(['Разгрузки', 'Объем', 'Доля', 'Рейсы'])
    for row in context['dump_rows']:
        sheet.append([row['label'], row['volume_display'], f"{row['share']}%", row['trips']])
    sheet.append([])
    sheet.append(['Внимание', 'Статус', 'Сводка', 'Деталь'])
    for row in context['attention_rows']:
        sheet.append([row['title'], row['status'], row['summary'], row['detail']])
    sheet.append([])
    sheet.append(['Простои по причинам', 'Событий', 'Открыто', 'Критично', 'Часы'])
    for row in context['downtime_reason_rows']:
        sheet.append([row['name'], row['count'], row['open_count'], row['critical_count'], row['duration_display']])
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    for column_index in range(1, 8):
        sheet.column_dimensions[get_column_letter(column_index)].width = 24


def dispatcher_management_export_view(request):
    access, response = require_dispatcher_report_access(request)
    if response:
        return response
    workbook = Workbook()
    write_dispatcher_management_sheet(workbook.active, dispatcher_management_context(request, access))
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="dispatcher_management.xlsx"'
    workbook.save(response)
    return response


def get_downtime_report_filters(request):
    return {
        'date_from': request.GET.get('date_from', '').strip(),
        'date_to': request.GET.get('date_to', '').strip(),
        'status': request.GET.get('status', '').strip(),
        'critical': request.GET.get('critical', '').strip(),
        'equipment': request.GET.get('equipment', '').strip(),
        'reason': request.GET.get('reason', '').strip(),
        'query_string': request.GET.urlencode(),
    }


def apply_downtime_report_filters(queryset, filters):
    date_from = parse_filter_date(filters.get('date_from'))
    date_to = parse_filter_date(filters.get('date_to'))
    status = filters.get('status')
    critical = filters.get('critical')
    equipment = filters.get('equipment', '')
    reason = filters.get('reason', '')

    if date_from:
        queryset = queryset.filter(started_at__date__gte=date_from)
    if date_to:
        queryset = queryset.filter(started_at__date__lte=date_to)
    if status == 'open':
        queryset = queryset.filter(ended_at__isnull=True)
    elif status == 'closed':
        queryset = queryset.filter(ended_at__isnull=False)
    if equipment.isdigit():
        queryset = queryset.filter(equipment_id=equipment)
    if reason.isdigit():
        queryset = queryset.filter(reason_id=reason)
    return queryset


def downtime_duration_hours(event):
    end_time = event.ended_at or timezone.now()
    seconds = max((end_time - event.started_at).total_seconds(), 0)
    return (Decimal(str(seconds)) / Decimal('3600')).quantize(Decimal('0.01'))


def downtime_status_label(event):
    return 'Открыт' if event.ended_at is None else 'Закрыт'


def downtime_report_rows(events):
    rows = []
    for event in events:
        is_critical = is_red_downtime_reason(event.reason)
        rows.append({
            'started_at': event.started_at,
            'ended_at': event.ended_at,
            'equipment': event.equipment,
            'reason': event.reason,
            'is_critical': is_critical,
            'color_group': downtime_reason_color_group(event.reason),
            'equipment_state_code': downtime_reason_state_code(event.reason),
            'status': downtime_status_label(event),
            'duration_hours': downtime_duration_hours(event),
            'employee': event.employee,
            'comment': event.comment,
        })
    return rows


def downtime_summary_by(rows, key):
    summary = {}
    for row in rows:
        name = str(row[key]) if row[key] else '-'
        if name not in summary:
            summary[name] = {
                'name': name,
                'count': 0,
                'open_count': 0,
                'critical_count': 0,
                'duration_hours': Decimal('0'),
            }
        summary[name]['count'] += 1
        if row['ended_at'] is None:
            summary[name]['open_count'] += 1
        if row['is_critical']:
            summary[name]['critical_count'] += 1
        summary[name]['duration_hours'] += row['duration_hours']
    result = []
    for item in summary.values():
        item['duration_hours'] = item['duration_hours'].quantize(Decimal('0.01'))
        result.append(item)
    return sorted(result, key=lambda item: (item['open_count'], item['duration_hours'], item['count']), reverse=True)


def downtime_daily_summary(rows):
    summary = {}
    for row in rows:
        day = timezone.localtime(row['started_at']).date()
        key = day.isoformat()
        if key not in summary:
            summary[key] = {
                'date': day,
                'count': 0,
                'open_count': 0,
                'critical_count': 0,
                'duration_hours': Decimal('0'),
            }
        summary[key]['count'] += 1
        if row['ended_at'] is None:
            summary[key]['open_count'] += 1
        if row['is_critical']:
            summary[key]['critical_count'] += 1
        summary[key]['duration_hours'] += row['duration_hours']
    result = []
    for item in summary.values():
        item['duration_hours'] = item['duration_hours'].quantize(Decimal('0.01'))
        result.append(item)
    return sorted(result, key=lambda item: item['date'], reverse=True)


def get_unloading_waiting_destination(reason):
    normalized = str(reason or '').strip().lower()
    for marker, destination in UNLOADING_WAITING_REASONS.items():
        if marker in normalized:
            return destination
    return ''


def downtime_unloading_waiting_summary(rows):
    summary = {}
    for row in rows:
        destination = get_unloading_waiting_destination(row['reason'])
        if not destination:
            continue
        if destination not in summary:
            summary[destination] = {
                'destination': destination,
                'reason': f'Ожидание разгрузки {destination}',
                'event_count': 0,
                'open_count': 0,
                'equipment_names': set(),
                'duration_hours': Decimal('0'),
            }
        summary[destination]['event_count'] += 1
        if row['ended_at'] is None:
            summary[destination]['open_count'] += 1
        summary[destination]['equipment_names'].add(str(row['equipment']))
        summary[destination]['duration_hours'] += row['duration_hours']

    result = []
    for item in summary.values():
        duration_hours = item['duration_hours'].quantize(Decimal('0.01'))
        duration_minutes = (item['duration_hours'] * Decimal('60')).quantize(Decimal('0.01'))
        event_count = item['event_count']
        item['duration_hours'] = duration_hours
        item['duration_minutes'] = duration_minutes
        item['equipment_count'] = len(item['equipment_names'])
        item['avg_minutes_per_event'] = (
            duration_minutes / Decimal(event_count)
        ).quantize(Decimal('0.01')) if event_count else Decimal('0')
        del item['equipment_names']
        result.append(item)
    return sorted(result, key=lambda item: item['destination'])


def build_report_table(trips, selected_columns, column_labels=None):
    headers = [get_column_label(column, column_labels) for column in selected_columns]
    rows = [
        [VOLUME_REPORT_COLUMNS[column][1](trip) for column in selected_columns]
        for trip in trips
    ]
    return headers, rows


def build_grouped_report_table(trips, group_by, column_labels=None):
    group_label, group_getter = VOLUME_REPORT_GROUPS[group_by]
    if group_by in VOLUME_REPORT_COLUMNS:
        group_label = get_column_label(group_by, column_labels)
    volume_label = get_column_label('volume_m3', column_labels)
    tonnage_label = get_column_label('tonnage', column_labels)
    grouped = defaultdict(lambda: {
        'volume_m3': 0,
        'tonnage': 0,
        'trip_count': 0,
    })

    for trip in trips:
        key = group_getter(trip) or 'не задано'
        grouped[key]['volume_m3'] += trip.volume_m3 or 0
        grouped[key]['tonnage'] += trip.tonnage or 0
        grouped[key]['trip_count'] += 1

    headers = [group_label, volume_label, tonnage_label, 'Рейсы']
    rows = [
        [group_name, values['volume_m3'], values['tonnage'], values['trip_count']]
        for group_name, values in grouped.items()
    ]
    rows.sort(key=lambda row: str(row[0]))
    return headers, rows


def is_number(value):
    return isinstance(value, (int, float)) or hasattr(value, 'as_tuple')


def append_total_row(sheet, headers, rows, row_index, total_column_indexes):
    if not rows:
        return row_index

    total_row = [''] * len(headers)
    total_row[0] = 'Итого'
    for column_index in total_column_indexes:
        total = sum(row[column_index] for row in rows if is_number(row[column_index]))
        total_row[column_index] = total

    sheet.append(total_row)
    for cell in sheet[row_index]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='EAF4FF')
        cell.border = Border(top=Side(style='thin', color='9DBAD5'))
    return row_index + 1


def style_volume_report_workbook(sheet, table_header_row, total_columns):
    header_fill = PatternFill('solid', fgColor='12232E')
    header_font = Font(color='FFFFFF', bold=True)
    thin_border = Border(bottom=Side(style='thin', color='D9E2EC'))

    for cell in sheet[table_header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for row in sheet.iter_rows(min_row=table_header_row + 1):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border
            if is_number(cell.value):
                cell.number_format = '#,##0.00'

    for row in range(1, table_header_row):
        sheet.row_dimensions[row].height = 22

    sheet.freeze_panes = f'A{table_header_row + 1}'
    sheet.auto_filter.ref = f'A{table_header_row}:{get_column_letter(total_columns)}{sheet.max_row}'

    for column_index in range(1, total_columns + 1):
        column_letter = get_column_letter(column_index)
        max_length = 0
        for cell in sheet[column_letter]:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 42)


def write_volume_report_excel(sheet, headers, rows, selected_template, selected_group_by, filters, access, total_column_indexes):
    total_columns = max(len(headers), 4)
    generated_at = timezone.localtime(timezone.now()).strftime('%d.%m.%Y %H:%M')
    template_name = selected_template.name if selected_template else 'Стандартный отчет'
    employee_name = access.employee.full_name if access and access.employee else ''
    active_filter_rows = get_active_filter_rows(filters)

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    sheet['A1'] = 'Отчет по объемам'
    sheet['A1'].font = Font(size=16, bold=True, color='12232E')
    sheet['A1'].alignment = Alignment(horizontal='left')

    metadata_rows = [
        ['Сформирован', generated_at],
        ['Шаблон', template_name],
        ['Группировка', get_group_by_label(selected_group_by)],
    ]
    if employee_name:
        metadata_rows.append(['Сформировал', employee_name])

    for metadata_row in metadata_rows:
        sheet.append(metadata_row)

    sheet.append([])
    sheet.append(['Активные фильтры'])
    sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True)
    if active_filter_rows:
        for filter_row in active_filter_rows:
            sheet.append(filter_row)
    else:
        sheet.append(['Нет'])

    sheet.append([])
    table_header_row = sheet.max_row + 1
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    append_total_row(sheet, headers, rows, sheet.max_row + 1, total_column_indexes)

    for row in range(2, table_header_row):
        sheet.cell(row=row, column=1).font = Font(bold=True)

    style_volume_report_workbook(sheet, table_header_row, total_columns)


def apply_volume_report_filters(queryset, filters):
    date_from = parse_filter_date(filters.get('date_from'))
    date_to = parse_filter_date(filters.get('date_to'))
    loading_shift_type = filters.get('loading_shift_type', '')
    unloading_shift_type = filters.get('unloading_shift_type', '')
    carryover = filters.get('carryover', '')
    truck = filters.get('truck', '')
    excavator = filters.get('excavator', '')
    rock_type = filters.get('rock_type', '')
    dump_point = filters.get('dump_point', '')

    if date_from:
        queryset = queryset.filter(completed_at__date__gte=date_from)
    if date_to:
        queryset = queryset.filter(completed_at__date__lte=date_to)
    if loading_shift_type:
        queryset = queryset.filter(loading_shift__shift_type=loading_shift_type)
    if unloading_shift_type:
        queryset = queryset.filter(unloading_shift__shift_type=unloading_shift_type)
    if carryover == 'yes':
        queryset = queryset.filter(is_carryover=True)
    elif carryover == 'no':
        queryset = queryset.filter(is_carryover=False)
    if truck.isdigit():
        queryset = queryset.filter(truck_id=truck)
    if excavator.isdigit():
        queryset = queryset.filter(excavator_id=excavator)
    if rock_type.isdigit():
        queryset = queryset.filter(rock_type_id=rock_type)
    if dump_point.isdigit():
        queryset = queryset.filter(dump_point_id=dump_point)
    return queryset


def volume_report_filter_context(request, selected_template):
    filters = get_volume_report_filters(request, selected_template)
    return {
        **filters,
        'template': request.GET.get('template', '').strip(),
        'group_by': request.GET.get('group_by', '').strip(),
        'query_string': request.GET.urlencode(),
    }


def report_filter_choices():
    return {
        'trucks': Equipment.objects.filter(is_active=True, equipment_type__name__icontains='Самосвал').order_by('garage_number'),
        'excavators': Equipment.objects.filter(is_active=True, equipment_type__name__icontains='Экскаватор').order_by('garage_number'),
        'rock_types': RockType.objects.filter(is_active=True).order_by('name'),
        'dump_points': DumpPoint.objects.filter(is_active=True).order_by('name'),
    }


def get_detail_total_column_indexes(selected_columns):
    total_columns = {'planned_volume_m3', 'volume_m3', 'deviation_m3', 'tonnage'}
    return [
        index
        for index, column in enumerate(selected_columns)
        if column in total_columns
    ]


def get_grouped_total_column_indexes():
    return [1, 2, 3]


def volume_report_view(request):
    access_id = request.session.get('employee_access_id')
    if not access_id:
        return redirect('login')
    access = EmployeeAccess.objects.select_related('employee', 'role').filter(id=access_id, is_active=True).first()
    if not access or access.role.code not in {'dispatcher', 'admin', 'manager'}:
        return redirect('role_home')

    trips = Trip.objects.filter(status=TripStatus.COMPLETED).select_related(
        'truck',
        'excavator',
        'rock_type',
        'dump_point',
        'loading_shift',
        'unloading_shift',
    ).order_by('-completed_at')
    selected_template = get_selected_report_template(request)
    filters = get_volume_report_filters(request, selected_template)
    trips = apply_volume_report_filters(trips, filters)
    selected_columns = get_selected_columns(selected_template)
    column_labels = get_template_column_labels(selected_template)
    selected_group_by = get_selected_group_by(request, selected_template)
    if selected_group_by:
        headers, rows = build_grouped_report_table(trips, selected_group_by, column_labels)
    else:
        headers, rows = build_report_table(trips[:100], selected_columns, column_labels)
    total_volume = trips.aggregate(total=Sum('volume_m3'))['total'] or 0
    total_tonnage = trips.aggregate(total=Sum('tonnage'))['total'] or 0
    return render(
        request,
        'reports/volume_report.html',
        {
            'access': access,
            'report_headers': headers,
            'report_rows': rows,
            'total_volume': total_volume,
            'total_tonnage': total_tonnage,
            'filters': volume_report_filter_context(request, selected_template),
            'report_templates': get_volume_report_templates(),
            'selected_template': selected_template,
            'selected_group_by': selected_group_by,
            'group_options': report_template_group_options(selected_group_by),
            **report_filter_choices(),
        },
    )


def volume_report_export_view(request):
    access_id = request.session.get('employee_access_id')
    if not access_id:
        return redirect('login')
    access = EmployeeAccess.objects.select_related('employee', 'role').filter(id=access_id, is_active=True).first()
    if not access or access.role.code not in {'dispatcher', 'admin', 'manager'}:
        return redirect('role_home')

    trips = Trip.objects.filter(status=TripStatus.COMPLETED).select_related(
        'truck',
        'excavator',
        'rock_type',
        'dump_point',
        'loading_shift',
        'unloading_shift',
    ).order_by('-completed_at')
    selected_template = get_selected_report_template(request)
    filters = get_volume_report_filters(request, selected_template)
    trips = apply_volume_report_filters(trips, filters)
    selected_columns = get_selected_columns(selected_template)
    column_labels = get_template_column_labels(selected_template)
    selected_group_by = get_selected_group_by(request, selected_template)
    if selected_group_by:
        headers, rows = build_grouped_report_table(trips, selected_group_by, column_labels)
        total_column_indexes = get_grouped_total_column_indexes()
    else:
        headers, rows = build_report_table(trips, selected_columns, column_labels)
        total_column_indexes = get_detail_total_column_indexes(selected_columns)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Объемы'
    write_volume_report_excel(sheet, headers, rows, selected_template, selected_group_by, filters, access, total_column_indexes)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=\"volume_report.xlsx\"'
    workbook.save(response)
    return response


def shift_analytics_report_context(request):
    selected_date = parse_filter_date(request.GET.get('date')) or timezone.localdate()
    shift_type = request.GET.get('shift_type', '').strip()
    if shift_type not in {'', 'day', 'night'}:
        shift_type = ''
    analytics = build_shift_analytics(selected_date, shift_type)
    return {
        **analytics,
        'date_value': selected_date.strftime('%Y-%m-%d'),
        'query_string': request.GET.urlencode(),
        'export_url': f"{reverse('shift_analytics_report_export')}?{request.GET.urlencode()}" if request.GET.urlencode() else reverse('shift_analytics_report_export'),
        'current_time': timezone.localtime(timezone.now()).strftime('%H:%M'),
        'current_date': timezone.localdate().strftime('%d.%m.%Y'),
    }


def shift_analytics_report_view(request):
    access = get_reports_access(request, {'dispatcher', 'admin', 'manager'})
    if not access:
        return redirect('login' if not request.session.get('employee_access_id') else 'role_home')
    return render(
        request,
        'reports/shift_analytics.html',
        {
            'access': access,
            **shift_analytics_report_context(request),
        },
    )


def management_dynamics_report_context(request):
    today = timezone.localdate()
    date_to = parse_filter_date(request.GET.get('date_to')) or today
    date_from = parse_filter_date(request.GET.get('date_from')) or (date_to - timedelta(days=6))
    if date_from > date_to:
        date_from, date_to = date_to, date_from

    granularity = request.GET.get('granularity', 'day').strip()
    if granularity not in {'hour', 'shift', 'day', 'month'}:
        granularity = 'day'

    selected_excavator_ids = [
        int(item)
        for item in request.GET.getlist('excavators')
        if str(item).isdigit()
    ]
    excavator_choices = (
        Equipment.objects
        .filter(is_active=True, equipment_type__name__icontains='Экскаватор')
        .order_by('garage_number', 'id')
    )
    dynamics = build_excavator_dynamics(date_from, date_to, granularity, selected_excavator_ids)
    return {
        'dynamics': dynamics,
        'date_from_value': date_from.strftime('%Y-%m-%d'),
        'date_to_value': date_to.strftime('%Y-%m-%d'),
        'granularity': granularity,
        'excavator_choices': excavator_choices,
        'selected_excavator_ids': selected_excavator_ids,
        'query_string': request.GET.urlencode(),
        'current_time': timezone.localtime(timezone.now()).strftime('%H:%M'),
        'current_date': timezone.localdate().strftime('%d.%m.%Y'),
    }


def management_dynamics_view(request):
    access = get_reports_access(request, {'dispatcher', 'admin', 'manager'})
    if not access:
        return redirect('login' if not request.session.get('employee_access_id') else 'role_home')
    return render(
        request,
        'reports/management_dynamics.html',
        {
            'access': access,
            **management_dynamics_report_context(request),
        },
    )


def append_shift_analytics_rows(sheet, title, rows, headers, getters):
    sheet.append([])
    sheet.append([title])
    sheet.cell(sheet.max_row, 1).font = Font(bold=True, size=13)
    sheet.append(headers)
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='E8F3F1')
    if not rows:
        sheet.append(['Нет данных'])
        return
    for row in rows:
        sheet.append([getter(row) for getter in getters])


def shift_analytics_report_export_view(request):
    access = get_reports_access(request, {'dispatcher', 'admin', 'manager'})
    if not access:
        return redirect('login' if not request.session.get('employee_access_id') else 'role_home')
    context = shift_analytics_report_context(request)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Сменная аналитика'
    sheet['A1'] = 'Сменная аналитика производственного контура'
    sheet['A1'].font = Font(size=16, bold=True, color='12232E')
    sheet.append(['Дата', context['selected_date'].strftime('%d.%m.%Y')])
    sheet.append(['Смена', context['shift_label']])
    sheet.append(['Сформирован', f"{context['current_date']} {context['current_time']}"])
    sheet.append([])
    sheet.append(['Показатель', 'Значение'])
    for label, key in [
        ('Отгружено экскаваторами, рейсов', 'loaded_trip_count'),
        ('Разгружено самосвалами, рейсов', 'unloaded_trip_count'),
        ('Объем, м3', 'volume_display'),
        ('Тоннаж, т', 'tonnage_display'),
        ('Переходящие рейсы', 'carryover_count'),
        ('Открытые рейсы', 'open_trip_count'),
        ('Простои, событий', 'downtime_count'),
        ('Простои, часов', 'downtime_hours_display'),
    ]:
        sheet.append([label, context['totals'][key]])

    common_headers = ['Наименование', 'Рейсы', 'Объем, м3', 'Тоннаж, т', 'Породы', 'Разгрузка', 'Горизонт/блок']
    common_getters = [
        lambda row: row['label'],
        lambda row: row['trip_count'],
        lambda row: row['volume_display'],
        lambda row: row['tonnage_display'],
        lambda row: row['rocks_display'],
        lambda row: row['dump_points_display'],
        lambda row: row['faces_display'],
    ]
    append_shift_analytics_rows(sheet, 'Экскаваторы', context['excavator_rows'], common_headers, common_getters)
    append_shift_analytics_rows(sheet, 'Самосвалы', context['truck_rows'], common_headers, common_getters)
    append_shift_analytics_rows(
        sheet,
        'Сотрудники',
        context['employee_rows'],
        ['Сотрудник', 'Роль', 'Отгружено', 'Разгружено', 'Объем, м3', 'Тоннаж, т'],
        [
            lambda row: row['label'],
            lambda row: row['role'],
            lambda row: row['loaded_count'],
            lambda row: row['unloaded_count'],
            lambda row: row['volume_display'],
            lambda row: row['tonnage_display'],
        ],
    )
    append_shift_analytics_rows(sheet, 'Породы', context['rock_rows'], common_headers, common_getters)
    append_shift_analytics_rows(sheet, 'Точки разгрузки', context['dump_point_rows'], common_headers, common_getters)
    append_shift_analytics_rows(sheet, 'Горизонты и блоки', context['face_rows'], common_headers, common_getters)
    append_shift_analytics_rows(
        sheet,
        'Простои',
        context['downtime_reason_rows'],
        ['Причина', 'Событий', 'Открыто', 'Часов', 'Техника', 'Сотрудники'],
        [
            lambda row: row['label'],
            lambda row: row['count'],
            lambda row: row['open_count'],
            lambda row: row['duration_display'],
            lambda row: row['equipment_display'],
            lambda row: row['employees_display'],
        ],
    )

    for column_index in range(1, 8):
        sheet.column_dimensions[get_column_letter(column_index)].width = 22
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=\"shift_analytics.xlsx\"'
    workbook.save(response)
    return response


def downtime_report_context(request):
    filters = get_downtime_report_filters(request)
    selected_single_date = filters['date_from'] if filters['date_from'] and filters['date_from'] == filters['date_to'] else ''
    events = (
        DowntimeEvent.objects
        .select_related('equipment', 'equipment__equipment_type', 'reason', 'reason__equipment_state', 'employee')
        .order_by('-started_at')
    )
    events = apply_downtime_report_filters(events, filters)
    all_rows = downtime_report_rows(events)
    if filters.get('critical') == 'yes':
        all_rows = [row for row in all_rows if row['is_critical']]
    elif filters.get('critical') == 'no':
        all_rows = [row for row in all_rows if not row['is_critical']]
    visible_rows = all_rows[:200]
    total_duration_hours = sum((row['duration_hours'] for row in all_rows), Decimal('0')).quantize(Decimal('0.01'))
    open_count = sum(1 for row in all_rows if row['ended_at'] is None)
    closed_count = len(all_rows) - open_count
    critical_count = sum(1 for row in all_rows if row['is_critical'])
    unloading_waiting_summary = downtime_unloading_waiting_summary(all_rows)
    unloading_waiting_total_minutes = sum(
        (item['duration_minutes'] for item in unloading_waiting_summary),
        Decimal('0'),
    ).quantize(Decimal('0.01'))
    unloading_waiting_event_count = sum(item['event_count'] for item in unloading_waiting_summary)
    unloading_waiting_equipment_count = sum(item['equipment_count'] for item in unloading_waiting_summary)
    return {
        'filters': filters,
        'selected_single_date': selected_single_date,
        'rows': visible_rows,
        'export_rows': all_rows,
        'visible_count': len(visible_rows),
        'total_count': len(all_rows),
        'open_count': open_count,
        'closed_count': closed_count,
        'critical_count': critical_count,
        'total_duration_hours': total_duration_hours,
        'daily_summary': downtime_daily_summary(all_rows),
        'equipment_summary': downtime_summary_by(all_rows, 'equipment'),
        'reason_summary': downtime_summary_by(all_rows, 'reason'),
        'unloading_waiting_summary': unloading_waiting_summary,
        'unloading_waiting_total_minutes': unloading_waiting_total_minutes,
        'unloading_waiting_event_count': unloading_waiting_event_count,
        'unloading_waiting_equipment_count': unloading_waiting_equipment_count,
        'unloading_waiting_reconciliation': UNLOADING_WAITING_EXCEL_RECONCILIATION,
        'equipment_choices': Equipment.objects.filter(is_active=True).order_by('equipment_type__name', 'garage_number'),
        'reason_choices': DowntimeReason.objects.filter(is_active=True).order_by('name'),
        'status_choices': [
            {'code': '', 'label': 'Все'},
            {'code': 'open', 'label': 'Открытые'},
            {'code': 'closed', 'label': 'Закрытые'},
        ],
        'critical_choices': [
            {'code': '', 'label': 'Все'},
            {'code': 'yes', 'label': 'Только критические'},
            {'code': 'no', 'label': 'Только обычные'},
        ],
    }


def downtime_report_view(request):
    access = get_reports_access(request, {'admin', 'dispatcher', 'manager', 'mechanic'})
    if not access:
        return redirect('login' if not request.session.get('employee_access_id') else 'role_home')
    return render(
        request,
        'reports/downtime_report.html',
        {
            'access': access,
            **downtime_report_context(request),
        },
    )


def downtime_report_export_view(request):
    access = get_reports_access(request, {'admin', 'dispatcher', 'manager', 'mechanic'})
    if not access:
        return redirect('login' if not request.session.get('employee_access_id') else 'role_home')
    context = downtime_report_context(request)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Простои'
    sheet['A1'] = 'Отчет по простоям техники'
    sheet['A1'].font = Font(bold=True, size=14)
    sheet['A2'] = f"Сформировал: {access.employee.full_name}"
    sheet['A3'] = f"Дата формирования: {timezone.localtime(timezone.now()):%d.%m.%Y %H:%M}"
    summary_rows = [
        ['Показатель', 'Значение'],
        ['Всего событий', context['total_count']],
        ['Открытые', context['open_count']],
        ['Закрытые', context['closed_count']],
        ['Критические', context['critical_count']],
        ['Длительность, ч', context['total_duration_hours']],
    ]
    for row_index, row in enumerate(summary_rows, start=5):
        for column_index, value in enumerate(row, start=1):
            cell = sheet.cell(row_index, column_index, value)
            if row_index == 5:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', fgColor='17232E')

    def write_downtime_summary(title, rows, start_row, start_col):
        sheet.cell(start_row, start_col, title)
        sheet.cell(start_row, start_col).font = Font(bold=True, size=12)
        headers = ['Название', 'События', 'Открытые', 'Критические', 'Длительность, ч']
        for offset, header in enumerate(headers):
            cell = sheet.cell(start_row + 1, start_col + offset, header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='17232E')
        for row_index, item in enumerate(rows[:10], start=start_row + 2):
            values = [item['name'], item['count'], item['open_count'], item['critical_count'], item['duration_hours']]
            for offset, value in enumerate(values):
                sheet.cell(row_index, start_col + offset, value)

    write_downtime_summary('Сводка по технике', context['equipment_summary'], 5, 4)
    write_downtime_summary('Сводка по причинам', context['reason_summary'], 5, 10)
    write_downtime_summary(
        'Сводка по датам',
        [
            {
                **item,
                'name': item['date'].strftime('%d.%m.%Y'),
            }
            for item in context['daily_summary']
        ],
        12,
        4,
    )

    headers = ['Начало', 'Окончание', 'Статус', 'Техника', 'Причина', 'Критичность', 'Длительность, ч', 'Кто зафиксировал', 'Комментарий']
    table_start = 25
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(table_start, column_index, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='17232E')
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    for row_index, row in enumerate(context['export_rows'], start=table_start + 1):
        values = [
            timezone.localtime(row['started_at']).strftime('%d.%m.%Y %H:%M'),
            timezone.localtime(row['ended_at']).strftime('%d.%m.%Y %H:%M') if row['ended_at'] else 'открыт',
            row['status'],
            str(row['equipment']),
            str(row['reason']),
            'Критический' if row['is_critical'] else 'Обычный',
            row['duration_hours'],
            str(row['employee']) if row['employee'] else '-',
            row['comment'] or '-',
        ]
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column_index, value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    for column_index in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = 18
    sheet.column_dimensions['I'].width = 36
    sheet.auto_filter.ref = f'A{table_start}:I{table_start + max(len(context["export_rows"]), 1)}'
    sheet.freeze_panes = f'A{table_start + 1}'

    unloading_sheet = workbook.create_sheet('ОР ККД СКДР')
    unloading_sheet['A1'] = 'Сверка ожидания разгрузки ККД/СКДР'
    unloading_sheet['A1'].font = Font(bold=True, size=14)
    unloading_sheet['A2'] = 'Источник старой формы: ОР ККД СКДР март.xlsx'
    unloading_sheet['A3'] = 'Принцип MVP: ожидание разгрузки фиксируется как событие простоя самосвала.'
    reconciliation_headers = ['Старый блок', 'Старые поля', 'Блок MVP', 'Статус', 'Примечание']
    for column_index, header in enumerate(reconciliation_headers, start=1):
        cell = unloading_sheet.cell(5, column_index, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='17232E')
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    for row_index, item in enumerate(context['unloading_waiting_reconciliation'], start=6):
        values = [item['old_block'], item['old_fields'], item['mvp_block'], item['status'], item['note']]
        for column_index, value in enumerate(values, start=1):
            cell = unloading_sheet.cell(row_index, column_index, value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    summary_start = 6 + len(context['unloading_waiting_reconciliation']) + 2
    unloading_sheet.cell(summary_start, 1, 'Сводка по выбранным фильтрам')
    unloading_sheet.cell(summary_start, 1).font = Font(bold=True, size=12)
    summary_headers = ['Направление', 'Причина', 'События', 'Открытые', 'Техника', 'Минуты', 'Часы', 'Среднее мин/событие']
    for column_index, header in enumerate(summary_headers, start=1):
        cell = unloading_sheet.cell(summary_start + 1, column_index, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='17232E')
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    for row_index, item in enumerate(context['unloading_waiting_summary'], start=summary_start + 2):
        values = [
            item['destination'],
            item['reason'],
            item['event_count'],
            item['open_count'],
            item['equipment_count'],
            item['duration_minutes'],
            item['duration_hours'],
            item['avg_minutes_per_event'],
        ]
        for column_index, value in enumerate(values, start=1):
            unloading_sheet.cell(row_index, column_index, value)
    if not context['unloading_waiting_summary']:
        unloading_sheet.cell(summary_start + 2, 1, 'По выбранным фильтрам ожидания разгрузки ККД/СКДР нет.')

    for column_index in range(1, len(summary_headers) + 1):
        unloading_sheet.column_dimensions[get_column_letter(column_index)].width = 22
    unloading_sheet.column_dimensions['B'].width = 30
    unloading_sheet.column_dimensions['E'].width = 40
    unloading_sheet.freeze_panes = 'A6'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=\"downtime_report.xlsx\"'
    workbook.save(response)
    return response


def get_reports_access(request, allowed_roles):
    access_id = request.session.get('employee_access_id')
    if not access_id:
        return None
    access = EmployeeAccess.objects.select_related('employee', 'role').filter(id=access_id, is_active=True).first()
    if not access or access.role.code not in allowed_roles:
        return None
    return access


PILOT_REPORT_CHECKLIST_SECTIONS = [
    {
        'title': '1. Витрина руководства',
        'items': [
            {
                'text': 'Открыть управленческую витрину и проверить суточные KPI, выполнение плана, день/ночь и динамику за 7 дней.',
                'url': '/reports/management/',
                'url_text': 'Открыть витрину',
            },
            {
                'text': 'Выгрузить витрину в Excel и сверить листы: Сводка, Динамика 7 дней, День ночь.',
                'url': '/reports/management/export/',
                'url_text': 'Выгрузить Excel',
            },
        ],
    },
    {
        'title': '2. Диспетчерский контроль',
        'items': [
            {
                'text': 'Проверить активные рейсы, неподтвержденные назначения, открытые механические простои и последние завершенные рейсы.',
                'url': '/dispatcher/control/',
                'url_text': 'Открыть пульт',
            },
        ],
    },
    {
        'title': '3. Отчет по объемам',
        'items': [
            {
                'text': 'Проверить фильтры, группировки, переходящие рейсы и выбранный шаблон отчета.',
                'url': '/reports/volume/',
                'url_text': 'Открыть отчет',
            },
            {
                'text': 'Проверить Excel-выгрузку отчета по объемам с теми же фильтрами и столбцами.',
                'url': '/reports/volume/export/',
                'url_text': 'Выгрузить Excel',
            },
            {
                'text': 'Проверить конструктор шаблонов: состав столбцов, подписи, фильтры и группировку.',
                'url': '/reports/templates/',
                'url_text': 'Открыть конструктор',
            },
        ],
    },
    {
        'title': '4. Суточный отчет заказчику',
        'items': [
            {
                'text': 'Открыть суточный отчет и проверить смены, объемы, месячную сводку и механические простои.',
                'url': '/reports/customer-daily/',
                'url_text': 'Открыть отчет',
            },
            {
                'text': 'Выгрузить суточный отчет в Excel и сверить структуру с текущей формой заказчика.',
                'url': '/reports/customer-daily/export/',
                'url_text': 'Выгрузить Excel',
            },
        ],
    },
    {
        'title': '5. Механические простои',
        'items': [
            {
                'text': 'Проверить отчет по простоям: открытые/закрытые события, критичность, причины, техника и длительность.',
                'url': '/reports/downtimes/',
                'url_text': 'Открыть отчет',
            },
            {
                'text': 'Выгрузить отчет по простоям в Excel и проверить, что полная история попадает в файл.',
                'url': '/reports/downtimes/export/',
                'url_text': 'Выгрузить Excel',
            },
        ],
    },
    {
        'title': '6. Вопросы перед пилотом',
        'items': [
            {
                'text': 'Сверить, какие текущие Excel-формы диспетчерская обязана сдавать каждый день.',
                'url': '',
                'url_text': '',
            },
            {
                'text': 'Отметить расхождения между системой и действующими отчетами: недостающие столбцы, названия, формулы и порядок строк.',
                'url': '',
                'url_text': '',
            },
            {
                'text': 'Решить, что исправляем до пилота, а что фиксируем как ограничение первой версии.',
                'url': '',
                'url_text': '',
            },
        ],
    },
]


PILOT_LAUNCH_SCENARIO_STEPS = [
    {
        'title': '0. Админка и доступы',
        'role': 'Администратор',
        'access_code': '1000',
        'url': '/system-admin/',
        'checks': [
            'Открыть админку MVP.',
            'Проверить список сотрудников.',
            'Проверить справочники админки.',
            'Убедиться, что есть доступ к конфликтам и журналу действий.',
            'При необходимости создать сотрудника и выдать первичный пинкод.',
        ],
        'expected_result': 'Администратор понимает, где подготовить сотрудников, доступы и справочники перед проходом пилота.',
    },
    {
        'title': '1. Вход и карта интерфейсов',
        'role': 'Диспетчер / руководство',
        'access_code': '5000 / 6000',
        'url': '/interfaces/',
        'checks': [
            'Открыть карту интерфейсов.',
            'Проверить ссылки на рабочие экраны и отчеты.',
            'Убедиться, что вход по коду открывает интерфейс по роли.',
        ],
        'expected_result': 'Пользователь быстро находит нужный экран и не ищет адреса вручную.',
    },
    {
        'title': '2. Расстановка техники',
        'role': 'Горный мастер',
        'access_code': '4000',
        'url': '/mining-master/assignments/',
        'checks': [
            'Выбрать экскаватор.',
            'Выбрать самосвал.',
            'Создать назначение.',
            'Проверить, что назначение видно водителю и диспетчеру.',
        ],
        'expected_result': 'Рабочая связка экскаватор - самосвал появляется в системе.',
    },
    {
        'title': '3. Работа водителя',
        'role': 'Водитель самосвала',
        'access_code': '2000',
        'url': '/driver/shift/',
        'checks': [
            'Если смена не открыта - открыть смену и проверить стартовые показатели техники.',
            'Если смена уже открыта - проверить активный рейс и текущее назначение.',
            'Подтвердить назначение кнопкой Принял.',
            'Завершить активный рейс кнопкой Выполнено.',
        ],
        'expected_result': 'Водитель выполняет минимум действий, а рейс попадает в отчет.',
    },
    {
        'title': '4. Создание рейса',
        'role': 'Машинист экскаватора',
        'access_code': '3000',
        'url': '/excavator/work/',
        'checks': [
            'Выбрать назначенный самосвал.',
            'Выбрать породу или груз.',
            'Выбрать точку разгрузки.',
            'Создать рейс.',
        ],
        'expected_result': 'Система считает объем и тоннаж, рейс становится активным у водителя и диспетчера.',
    },
    {
        'title': '5. Диспетчерский контроль',
        'role': 'Диспетчер',
        'access_code': '5000',
        'url': '/dispatcher/control/',
        'checks': [
            'Проверить активные рейсы.',
            'Проверить назначения без подтверждения.',
            'Проверить незакрытые смены.',
            'Проверить открытые простои и журнал действий.',
        ],
        'expected_result': 'Диспетчер видит текущую смену в процессе, а не собирает ее вручную после факта.',
    },
    {
        'title': '6. Простои техники',
        'role': 'Механик',
        'access_code': '7000',
        'url': '/mechanic/downtimes/',
        'checks': [
            'Проверить открытые простои.',
            'Если есть техника в блоке "Техника, где уже виден простой" - открыть простой из этой строки.',
            'Если есть открытый простой - закрыть простой кнопкой Закрыть простой.',
            'Открыть отчет по простоям.',
            'Проверить сводку ОР ККД/СКДР и Excel-выгрузку.',
        ],
        'expected_result': 'Простои фиксируются как события техники и попадают в отчетность.',
    },
    {
        'title': '7. Отчеты и витрина',
        'role': 'Диспетчер / руководство',
        'access_code': '5000 / 6000',
        'url': '/reports/pilot-checklist/',
        'checks': [
            'Открыть отчет по объемам и группировку по часу.',
            'Открыть суточный отчет заказчику.',
            'Открыть витрину руководства.',
            'Выгрузить основные Excel-файлы.',
        ],
        'expected_result': 'Отчеты формируются из данных системы и готовы к сверке со старыми Excel-формами.',
    },
]


PILOT_FEEDBACK_QUESTIONS = [
    'Достаточно ли группировки по часу вместо старой почасовой матрицы?',
    'Кто должен фиксировать ожидание разгрузки ККД/СКДР?',
    'Как считать среднее мин на 1 а/с?',
    'Как связать укрупненные грузы с точными породами заказчика?',
    'Какие действия в интерфейсах лишние или неудобные?',
    'Какие данные невозможно вводить стабильно в карьере?',
]


PILOT_REPORT_EXCEL_COVERAGE = [
    {
        'file': 'Отчет_Коппер. Рисорсез_Март.xlsx',
        'purpose': 'Суточный отчет заказчику',
        'coverage': 'частично покрыт',
        'system_link': '/reports/customer-daily/',
        'next_step': 'Сверить форму, порядок блоков и обязательные итоговые строки.',
    },
    {
        'file': 'почасовой Март.xlsx',
        'purpose': 'Почасовой отчет диспетчерской',
        'coverage': 'частично покрыт группировкой по часу',
        'system_link': '/reports/volume/?group_by=completed_hour',
        'next_step': 'Проверить с диспетчерской, нужна ли точная старая почасовая матрица после первого пилота.',
    },
    {
        'file': 'ОР ККД СКДР март.xlsx',
        'purpose': 'Ожидание разгрузки ККД/СКДР',
        'coverage': 'частично покрыт отдельной сводкой простоев',
        'system_link': '/reports/downtimes/',
        'next_step': 'На пилоте определить, кто фиксирует ожидание разгрузки: водитель, диспетчер или автоматический контроль рейса.',
    },
    {
        'file': 'СВОД Простоев на ККД Март.xlsx',
        'purpose': 'Свод простоев и невыполненного объема',
        'coverage': 'частично покрыт',
        'system_link': '/reports/downtimes/',
        'next_step': 'Решить, нужен ли расчет невыполненного объема в MVP.',
    },
    {
        'file': 'Работа экс Март (1).xlsx',
        'purpose': 'Работа экскаваторов',
        'coverage': 'частично покрыто',
        'system_link': '/reports/management/',
        'next_step': 'Проверить группировку по экскаваторам, объемам, рейсам и плечу.',
    },
    {
        'file': 'Работа экс Март ПЕРЕГОНЫ ПРИЧИНЫ.xlsx',
        'purpose': 'Перегоны и смена фронта работ',
        'coverage': 'не покрыт',
        'system_link': '',
        'next_step': 'Нужен модуль статусов и перегонов экскаватора.',
    },
    {
        'file': 'КИП/КТГ и КИО/КТГ',
        'purpose': 'Показатели использования и технической готовности',
        'coverage': 'не покрыт',
        'system_link': '',
        'next_step': 'Определить формулы, источники рабочего времени, простоев и ремонтов.',
    },
    {
        'file': 'График ТО.xlsx',
        'purpose': 'Планирование технического обслуживания',
        'coverage': 'не покрыт',
        'system_link': '',
        'next_step': 'Отнести к развитию механического модуля после первого пилота.',
    },
    {
        'file': 'удельный_веса_руд_и_пород_Малмыжского_местородения.xlsx',
        'purpose': 'Справочник пород, плотностей и коэффициентов разрыхления',
        'coverage': 'структурно сверено',
        'system_link': '/admin/references/rocktype/',
        'next_step': 'На пилоте уточнить соответствие укрупненных рабочих названий точным породам заказчика.',
    },
]


CUSTOMER_DAILY_EXCEL_RECONCILIATION = [
    {
        'old_block': 'Заголовок отчета заказчику',
        'old_fields': 'Дата, заказчик, подрядчик, дневная и ночная смена',
        'mvp_block': 'Шапка страницы и Excel-лист "Суточный отчет"',
        'status': 'покрыто',
        'note': 'Название отчета и выбранная дата выводятся на экран и в Excel-выгрузку.',
    },
    {
        'old_block': 'Работа выемочного оборудования',
        'old_fields': 'Тип грунта, экскаватор, план, факт, горизонт, блок, место разгрузки, плечо, простои, примечание',
        'mvp_block': 'Таблицы I смена и II смена',
        'status': 'покрыто частично',
        'note': 'Базовые поля уже есть. Перед пилотом нужно сверить точные названия пород и порядок строк с действующей формой.',
    },
    {
        'old_block': 'Суточная сводка',
        'old_fields': 'План, факт, отклонение, день, ночь, сутки',
        'mvp_block': 'Блок "Суточная сводка"',
        'status': 'покрыто',
        'note': 'План/факт/отклонение считаются по рейсам и плановым заданиям, внесенным в систему.',
    },
    {
        'old_block': 'С начала месяца',
        'old_fields': 'План с начала месяца, факт с начала месяца, отклонение',
        'mvp_block': 'Блок "С начала месяца"',
        'status': 'покрыто частично',
        'note': 'Факт берется из рейсов с начала месяца. Отдельную модель месячного плана нужно уточнить перед промышленным запуском.',
    },
    {
        'old_block': 'Итоги по породам',
        'old_fields': 'Горная масса, добыча руды, сульфидная, переходная, вскрыша, окисленная, рыхлая, скальная, ПСП, ППСП',
        'mvp_block': 'Блок "Итоги по породам"',
        'status': 'покрыто частично',
        'note': 'Система группирует по справочнику пород. Нужно сверить справочник пород с действующими названиями заказчика.',
    },
    {
        'old_block': 'Средневзвешенное плечо',
        'old_fields': 'Среднее плечо по породам, сменам, суткам и с начала месяца',
        'mvp_block': 'Плечо в строках сменных таблиц',
        'status': 'требует доработки',
        'note': 'В рейсах хранится плечо, но отдельный расчет средневзвешенного плеча пока не вынесен в сводку.',
    },
    {
        'old_block': 'Расчет выполненных работ по самосвалам',
        'old_fields': '№ самосвала, рейсы, км, объем, итог, м3*км, простои',
        'mvp_block': 'Отчет по объемам и конструктор отчетов',
        'status': 'покрыто частично',
        'note': 'Рейсы, объем, тоннаж, самосвал и плечо есть в данных. М3*км и группировку точно как в старой расчетной вкладке нужно добавить отдельным шаблоном.',
    },
    {
        'old_block': 'Простои и примечания',
        'old_fields': 'Простои в сменных строках и комментарии по технике',
        'mvp_block': 'Простои в рейсах и механические простои',
        'status': 'покрыто частично',
        'note': 'Механические простои уже выделены отдельно. Производственные простои экскаватора нужно довести через справочник статусов экскаватора.',
    },
]


UNLOADING_WAITING_EXCEL_RECONCILIATION = [
    {
        'old_block': 'Дневные листы по датам',
        'old_fields': 'Дата, смена, направление ККД/СКДР',
        'mvp_block': 'Фильтры отчета по простоям и сводка ожидания разгрузки',
        'status': 'покрыто частично',
        'note': 'Дата берется по началу события простоя. Разделение по сменам нужно уточнить после решения, кто фиксирует событие.',
    },
    {
        'old_block': 'Строки по самосвалам',
        'old_fields': 'Номер самосвала, причина ожидания, минуты ожидания',
        'mvp_block': 'События простоев по технике',
        'status': 'покрыто',
        'note': 'Каждое ожидание разгрузки хранится как событие простоя самосвала с началом, окончанием, причиной и длительностью.',
    },
    {
        'old_block': 'Итоги по направлению',
        'old_fields': 'Количество машин/событий, всего минут, среднее минут на 1 а/с',
        'mvp_block': 'Сводка ОР ККД/СКДР',
        'status': 'покрыто частично',
        'note': 'Система считает события, технику, минуты, часы и среднее минут на событие. Формулировку среднего нужно сверить с диспетчерской.',
    },
    {
        'old_block': 'Форма Excel для передачи',
        'old_fields': 'Отдельная таблица ожидания разгрузки ККД/СКДР',
        'mvp_block': 'Лист "ОР ККД СКДР" в Excel-выгрузке простоев',
        'status': 'покрыто частично',
        'note': 'Выгрузка уже содержит отдельный лист сверки. Точную старую раскладку по листам дат можно делать после пилотной проверки.',
    },
]


def pilot_report_checklist_view(request):
    access = get_reports_access(request, {'admin', 'dispatcher', 'manager'})
    if not access:
        return redirect('login' if not request.session.get('employee_access_id') else 'role_home')

    return render(
        request,
        'reports/pilot_checklist.html',
        {
            'access': access,
            'sections': PILOT_REPORT_CHECKLIST_SECTIONS,
            'excel_coverage': PILOT_REPORT_EXCEL_COVERAGE,
            'progress_stage': '9 из 10',
            'progress_percent': 99,
            'remaining_stages': 1,
        },
    )


def pilot_launch_scenario_view(request):
    access = get_reports_access(request, {'admin', 'dispatcher', 'manager'})
    if not access:
        return redirect('login' if not request.session.get('employee_access_id') else 'role_home')

    return render(
        request,
        'reports/pilot_scenario.html',
        {
            'access': access,
            'steps': PILOT_LAUNCH_SCENARIO_STEPS,
            'feedback_questions': PILOT_FEEDBACK_QUESTIONS,
            'progress_stage': '9 из 10',
            'progress_percent': 99,
            'remaining_stages': 1,
        },
    )


def pilot_feedback_view(request):
    access = get_reports_access(request, {'admin', 'dispatcher', 'manager'})
    if not access:
        return redirect('login' if not request.session.get('employee_access_id') else 'role_home')

    if request.method == 'POST':
        action = request.POST.get('action', '').strip()
        feedback_id = request.POST.get('feedback_id', '').strip()
        next_status = request.POST.get('status', '').strip()
        if action == 'change_status' and feedback_id and next_status in {'in_work', 'decided', 'rejected'}:
            feedback = PilotFeedback.objects.filter(id=feedback_id).first()
            if feedback:
                feedback.status = next_status
                feedback.save(update_fields=['status', 'updated_at'])
                messages.success(request, 'Статус замечания обновлен.')
            return redirect('pilot_feedback')
        else:
            form = PilotFeedbackForm(request.POST)
            if form.is_valid():
                feedback = form.save(commit=False)
                feedback.created_by = access.employee
                feedback.save()
                messages.success(request, 'Замечание пилота зафиксировано.')
                return redirect('pilot_feedback')
    else:
        form = PilotFeedbackForm()

    feedback_items = (
        PilotFeedback.objects
        .select_related('created_by')
        .order_by('priority', '-created_at')
    )
    feedback_summary = {
        'total': feedback_items.count(),
        'p0': feedback_items.filter(priority='p0').count(),
        'p1': feedback_items.filter(priority='p1').count(),
        'open': feedback_items.exclude(status__in=['decided', 'rejected']).count(),
    }
    return render(
        request,
        'reports/pilot_feedback.html',
        {
            'access': access,
            'form': form,
            'feedback_items': feedback_items,
            'feedback_summary': feedback_summary,
            'progress_stage': '9 из 10',
            'progress_percent': 99,
            'remaining_stages': 1,
        },
    )


def pilot_feedback_export_view(request):
    access = get_reports_access(request, {'admin', 'dispatcher', 'manager'})
    if not access:
        return redirect('login' if not request.session.get('employee_access_id') else 'role_home')

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Замечания пилота'
    sheet.append(['Журнал замечаний пилотного запуска'])
    sheet.append([f'Сформировал: {access.employee.full_name}'])
    sheet.append([])
    headers = [
        'Дата',
        'Приоритет',
        'Статус',
        'Категория',
        'Экран или процесс',
        'Краткое замечание',
        'Описание',
        'Решение',
        'Кто зафиксировал',
    ]
    sheet.append(headers)
    for cell in sheet[4]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='17232E')
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    feedback_items = PilotFeedback.objects.select_related('created_by').order_by('priority', '-created_at')
    for feedback in feedback_items:
        sheet.append([
            timezone.localtime(feedback.created_at).strftime('%d.%m.%Y %H:%M'),
            feedback.get_priority_display(),
            feedback.get_status_display(),
            feedback.get_category_display(),
            feedback.screen,
            feedback.title,
            feedback.description,
            feedback.decision,
            feedback.created_by.full_name,
        ])

    widths = [18, 24, 20, 22, 28, 42, 56, 56, 28]
    for column_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    sheet.freeze_panes = 'A5'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=\"pilot_feedback.xlsx\"'
    workbook.save(response)
    return response


def report_template_builder_view(request):
    access = get_reports_access(request, {'admin', 'dispatcher'})
    if not access:
        return redirect('login' if not request.session.get('employee_access_id') else 'role_home')

    edit_template = None
    edit_template_id = request.GET.get('edit', '').strip()
    if edit_template_id:
        edit_template = ReportTemplate.objects.filter(id=edit_template_id).first()

    if request.method == 'POST':
        template_id = request.POST.get('template_id', '').strip()
        name = request.POST.get('name', '').strip()
        columns = [
            column
            for column in request.POST.getlist('columns')
            if column in VOLUME_REPORT_COLUMNS
        ]
        column_labels = {}
        for column in VOLUME_REPORT_COLUMNS:
            custom_label = request.POST.get(f'column_label_{column}', '').strip()
            default_label = VOLUME_REPORT_COLUMNS[column][0]
            if custom_label and custom_label != default_label:
                column_labels[column] = custom_label
        filters = {
            key: request.POST.get(key, '').strip()
            for key in REPORT_TEMPLATE_FILTER_FIELDS
            if request.POST.get(key, '').strip()
        }
        group_by = request.POST.get('group_by', '').strip()
        if group_by not in VOLUME_REPORT_GROUPS:
            group_by = ''
        is_active = request.POST.get('is_active') == 'on'

        template = None
        if template_id:
            template = ReportTemplate.objects.filter(id=template_id).first()
            if not template:
                messages.error(request, 'Шаблон отчета не найден.')
                return redirect('report_template_builder')

        duplicate_query = ReportTemplate.objects.filter(name=name)
        if template:
            duplicate_query = duplicate_query.exclude(id=template.id)

        if not name:
            messages.error(request, 'Укажи название шаблона отчета.')
        elif not columns:
            messages.error(request, 'Выбери хотя бы один столбец отчета.')
        elif duplicate_query.exists():
            messages.error(request, 'Шаблон с таким названием уже существует.')
        else:
            if not template:
                template = ReportTemplate(created_by=access.employee)
            template.name = name
            template.report_type = ReportType.SHIFT_VOLUME
            template.columns = columns
            template.column_labels = column_labels
            template.filters = filters
            template.group_by = group_by
            template.is_active = is_active
            template.updated_by = access.employee
            template.save()
            messages.success(request, 'Шаблон отчета сохранен.')
            return redirect('report_template_builder')

        edit_template = template

    selected_columns = get_selected_columns(edit_template) if edit_template else DEFAULT_VOLUME_REPORT_COLUMNS
    column_labels = get_template_column_labels(edit_template)
    template_filters = get_template_filters(edit_template)
    selected_group_by = edit_template.group_by if edit_template and edit_template.group_by in VOLUME_REPORT_GROUPS else ''
    return render(
        request,
        'reports/report_template_builder.html',
        {
            'access': access,
            'report_templates': ReportTemplate.objects.order_by('name'),
            'edit_template': edit_template,
            'column_options': report_template_column_options(selected_columns, column_labels),
            'template_filters': template_filters,
            'selected_group_by': selected_group_by,
            'group_options': report_template_group_options(selected_group_by),
            **report_filter_choices(),
        },
    )


def parse_customer_report_date(request):
    date_value = request.GET.get('date', '').strip()
    if date_value:
        try:
            return datetime.strptime(date_value, '%Y-%m-%d').date()
        except ValueError:
            pass

    latest_trip = (
        Trip.objects
        .filter(status=TripStatus.COMPLETED)
        .order_by('-completed_at', '-created_at')
        .first()
    )
    if latest_trip and latest_trip.completed_at:
        return timezone.localtime(latest_trip.completed_at).date()
    if latest_trip:
        return timezone.localtime(latest_trip.created_at).date()
    return timezone.localdate()


def trip_report_date(trip):
    if trip.loading_shift:
        return timezone.localtime(trip.loading_shift.opened_at).date()
    if trip.completed_at:
        return timezone.localtime(trip.completed_at).date()
    return timezone.localtime(trip.created_at).date()


def trip_shift_type(trip):
    if trip.loading_shift:
        return trip.loading_shift.shift_type
    if trip.unloading_shift:
        return trip.unloading_shift.shift_type
    return 'day'


def calculate_plan_completion_percent(volume, plan):
    if not plan:
        return None
    return (((volume or Decimal('0')) / plan) * Decimal('100')).quantize(Decimal('0.1'))


def build_management_daily_trend(trips, selected_date):
    trend_start = selected_date - timedelta(days=6)
    trend_by_date = {}
    for day_offset in range(7):
        report_date = trend_start + timedelta(days=day_offset)
        manual_plan_volume = shift_plan_totals(report_date)['volume_m3']
        trend_by_date[report_date] = {
            'date': report_date,
            'volume': Decimal('0'),
            'plan': manual_plan_volume,
            'has_manual_plan': bool(manual_plan_volume),
            'tonnage': Decimal('0'),
            'trip_count': 0,
        }

    for trip in trips:
        report_date = trip_report_date(trip)
        if report_date not in trend_by_date:
            continue
        trend_by_date[report_date]['volume'] += trip.volume_m3 or 0
        if not trend_by_date[report_date]['has_manual_plan']:
            trend_by_date[report_date]['plan'] += trip.planned_volume_m3 or 0
        trend_by_date[report_date]['tonnage'] += trip.tonnage or 0
        trend_by_date[report_date]['trip_count'] += 1

    daily_trend = []
    for values in trend_by_date.values():
        completion_percent = calculate_plan_completion_percent(values['volume'], values['plan'])
        daily_trend.append({
            **values,
            'deviation': values['volume'] - values['plan'],
            'completion_percent': completion_percent,
            'has_plan': completion_percent is not None,
        })
    return daily_trend


def customer_report_group_key(trip, include_report_date=False):
    key = (
        trip_shift_type(trip),
        str(trip.rock_type),
        str(trip.excavator),
        str(trip.dump_point),
        trip.loading_horizon,
        trip.loading_block,
        trip.transport_distance_km,
        trip.planned_volume_m3,
    )
    if include_report_date:
        return (trip_report_date(trip), *key)
    return key


def calculate_customer_accumulated_totals(trips):
    grouped = defaultdict(lambda: {
        'planned_volume': 0,
        'volume_m3': 0,
        'tonnage': 0,
        'trip_count': 0,
    })

    for trip in trips:
        key = customer_report_group_key(trip, include_report_date=True)
        grouped[key]['planned_volume'] = trip.planned_volume_m3 or 0
        grouped[key]['volume_m3'] += trip.volume_m3 or 0
        grouped[key]['tonnage'] += trip.tonnage or 0
        grouped[key]['trip_count'] += 1

    total_plan = sum(values['planned_volume'] for values in grouped.values())
    total_volume = sum(values['volume_m3'] for values in grouped.values())
    total_tonnage = sum(values['tonnage'] for values in grouped.values())
    total_trip_count = sum(values['trip_count'] for values in grouped.values())

    return {
        'plan': total_plan,
        'volume': total_volume,
        'deviation': total_volume - total_plan,
        'tonnage': total_tonnage,
        'trip_count': total_trip_count,
    }


def build_mechanic_downtime_rows(selected_date):
    events = (
        DowntimeEvent.objects
        .filter(started_at__date=selected_date)
        .select_related('equipment', 'reason', 'employee')
        .order_by('started_at')
    )
    rows = []
    for event in events:
        end_time = event.ended_at or timezone.now()
        duration_minutes = max(int((end_time - event.started_at).total_seconds() // 60), 0)
        rows.append({
            'started_at': event.started_at,
            'ended_at': event.ended_at,
            'equipment': event.equipment,
            'reason': event.reason,
            'employee': event.employee,
            'comment': event.comment,
            'duration_minutes': duration_minutes,
            'duration_hours': Decimal(duration_minutes) / Decimal('60'),
            'is_open': event.ended_at is None,
        })
    return rows


def build_customer_daily_report(selected_date):
    trips = Trip.objects.filter(status=TripStatus.COMPLETED).select_related(
        'truck',
        'excavator',
        'rock_type',
        'dump_point',
        'loading_shift',
        'unloading_shift',
    )
    all_trips = list(trips)
    trips = [trip for trip in all_trips if trip_report_date(trip) == selected_date]
    month_start = selected_date.replace(day=1)
    month_trips = [
        trip
        for trip in all_trips
        if month_start <= trip_report_date(trip) <= selected_date
    ]
    month_totals = calculate_customer_accumulated_totals(month_trips)

    grouped = defaultdict(lambda: {
        'volume_m3': 0,
        'tonnage': 0,
        'trip_count': 0,
        'carryover_count': 0,
        'trucks': set(),
        'downtimes': set(),
        'notes': set(),
    })

    for trip in trips:
        key = customer_report_group_key(trip)
        grouped[key]['volume_m3'] += trip.volume_m3 or 0
        grouped[key]['tonnage'] += trip.tonnage or 0
        grouped[key]['trip_count'] += 1
        grouped[key]['carryover_count'] += 1 if trip.is_carryover else 0
        grouped[key]['trucks'].add(str(trip.truck))
        if trip.downtime_text:
            grouped[key]['downtimes'].add(trip.downtime_text)
        if trip.note:
            grouped[key]['notes'].add(trip.note)

    rows_by_shift = {'day': [], 'night': []}
    for (
        shift_type,
        rock_type,
        excavator,
        dump_point,
        loading_horizon,
        loading_block,
        transport_distance_km,
        planned_volume_m3,
    ), values in grouped.items():
        note_parts = []
        if values['carryover_count']:
            note_parts.append(f"переходящих рейсов: {values['carryover_count']}")
        if values['trucks']:
            note_parts.append('самосвалы: ' + ', '.join(sorted(values['trucks'])))
        if values['notes']:
            note_parts.extend(sorted(values['notes']))
        rows_by_shift[shift_type].append({
            'rock_type': rock_type,
            'excavator': excavator,
            'planned_volume': planned_volume_m3,
            'volume_m3': values['volume_m3'],
            'volume_deviation': values['volume_m3'] - planned_volume_m3 if planned_volume_m3 is not None else None,
            'horizon': loading_horizon,
            'block': loading_block,
            'dump_point': dump_point,
            'distance_km': transport_distance_km,
            'downtime': '; '.join(sorted(values['downtimes'])),
            'note': '; '.join(note_parts),
            'tonnage': values['tonnage'],
            'trip_count': values['trip_count'],
        })

    for shift_rows in rows_by_shift.values():
        shift_rows.sort(key=lambda row: (row['excavator'], row['rock_type'], row['dump_point']))

    day_total = sum(row['volume_m3'] for row in rows_by_shift['day'])
    night_total = sum(row['volume_m3'] for row in rows_by_shift['night'])
    day_plan_total = sum(row['planned_volume'] or 0 for row in rows_by_shift['day'])
    night_plan_total = sum(row['planned_volume'] or 0 for row in rows_by_shift['night'])
    day_tonnage = sum(row['tonnage'] for row in rows_by_shift['day'])
    night_tonnage = sum(row['tonnage'] for row in rows_by_shift['night'])
    day_trip_count = sum(row['trip_count'] for row in rows_by_shift['day'])
    night_trip_count = sum(row['trip_count'] for row in rows_by_shift['night'])

    rock_summary = (
        Trip.objects
        .filter(id__in=[trip.id for trip in trips])
        .values('rock_type__name')
        .annotate(total_volume=Sum('volume_m3'), total_tonnage=Sum('tonnage'), trip_count=Count('id'))
        .order_by('-total_volume')
    )
    mechanic_downtime_rows = build_mechanic_downtime_rows(selected_date)

    return {
        'rows_by_shift': rows_by_shift,
        'day_total': day_total,
        'night_total': night_total,
        'total_volume': day_total + night_total,
        'day_plan_total': day_plan_total,
        'night_plan_total': night_plan_total,
        'total_plan': day_plan_total + night_plan_total,
        'day_deviation': day_total - day_plan_total,
        'night_deviation': night_total - night_plan_total,
        'total_deviation': (day_total + night_total) - (day_plan_total + night_plan_total),
        'day_tonnage': day_tonnage,
        'night_tonnage': night_tonnage,
        'total_tonnage': day_tonnage + night_tonnage,
        'day_trip_count': day_trip_count,
        'night_trip_count': night_trip_count,
        'total_trip_count': day_trip_count + night_trip_count,
        'month_start': month_start,
        'month_plan_total': month_totals['plan'],
        'month_total_volume': month_totals['volume'],
        'month_total_deviation': month_totals['deviation'],
        'month_total_tonnage': month_totals['tonnage'],
        'month_total_trip_count': month_totals['trip_count'],
        'rock_summary': rock_summary,
        'mechanic_downtime_rows': mechanic_downtime_rows,
        'mechanic_downtime_count': len(mechanic_downtime_rows),
        'mechanic_open_downtime_count': sum(1 for row in mechanic_downtime_rows if row['is_open']),
        'mechanic_downtime_hours': sum((row['duration_hours'] for row in mechanic_downtime_rows), Decimal('0')).quantize(Decimal('0.01')),
    }


def customer_daily_report_context(request):
    selected_date = parse_customer_report_date(request)
    report = build_customer_daily_report(selected_date)
    return {
        **report,
        'selected_date': selected_date,
        'date_input': selected_date.strftime('%Y-%m-%d'),
        'excel_reconciliation': CUSTOMER_DAILY_EXCEL_RECONCILIATION,
    }


def customer_daily_report_view(request):
    access = get_reports_access(request, {'dispatcher', 'admin', 'manager'})
    if not access:
        return redirect('login' if not request.session.get('employee_access_id') else 'role_home')

    return render(
        request,
        'reports/customer_daily_report.html',
        {
            'access': access,
            **customer_daily_report_context(request),
        },
    )


def append_customer_shift_table(sheet, title, start_row, start_col, rows):
    headers = [
        'Тип грунта',
        'Экскаватор',
        'План, м3',
        'Факт, м3',
        'Отклонение, м3',
        'Горизонт',
        'Блок',
        'Место разгрузки',
        'Плечо, км',
        'Простои, час',
        'Примечание',
    ]
    sheet.cell(start_row, start_col, title)
    sheet.cell(start_row, start_col).font = Font(bold=True, size=13)
    for offset, header in enumerate(headers):
        cell = sheet.cell(start_row + 1, start_col + offset, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='17232E')
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    for row_index, row in enumerate(rows, start=start_row + 2):
        values = [
            row['rock_type'],
            row['excavator'],
            row['planned_volume'] or 'не задано',
            row['volume_m3'],
            row['volume_deviation'] if row['volume_deviation'] is not None else '-',
            row['horizon'] or 'не задано',
            row['block'] or 'не задано',
            row['dump_point'],
            row['distance_km'] or 'не задано',
            row['downtime'] or 'не задано',
            row['note'],
        ]
        for offset, value in enumerate(values):
            cell = sheet.cell(row_index, start_col + offset, value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    return start_row + max(len(rows), 1) + 3


def append_customer_reconciliation_sheet(workbook, reconciliation_rows):
    sheet = workbook.create_sheet('Сверка с Excel')
    sheet['A1'] = 'Сверка суточного отчета с действующей Excel-формой заказчика'
    sheet['A1'].font = Font(bold=True, size=14)
    sheet['A2'] = 'Эталон для сверки: Отчет_Коппер. Рисорсез_Март.xlsx'
    sheet['A2'].alignment = Alignment(wrap_text=True)

    headers = ['Блок старой формы', 'Поля старой формы', 'Где в MVP', 'Статус', 'Комментарий']
    for offset, header in enumerate(headers, start=1):
        cell = sheet.cell(4, offset, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='17232E')
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    for row_index, row in enumerate(reconciliation_rows, start=5):
        values = [
            row['old_block'],
            row['old_fields'],
            row['mvp_block'],
            row['status'],
            row['note'],
        ]
        for offset, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, offset, value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    widths = [28, 42, 34, 20, 58]
    for column_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    return sheet


def customer_daily_report_export_view(request):
    access = get_reports_access(request, {'dispatcher', 'admin', 'manager'})
    if not access:
        return redirect('login' if not request.session.get('employee_access_id') else 'role_home')

    context = customer_daily_report_context(request)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Суточный отчет'
    sheet['A1'] = 'Отчет о работе ООО "Коппер Рисорсез" на ООО "Амур Минералс"'
    sheet['A1'].font = Font(bold=True, size=14)
    sheet['A2'] = f"Дата отчета: {context['selected_date']:%d.%m.%Y}"
    sheet['A4'] = 'Суточная сводка'
    sheet['A4'].font = Font(bold=True)
    summary_rows = [
        ['Показатель', 'День', 'Ночь', 'Сутки'],
        ['План, м3', context['day_plan_total'], context['night_plan_total'], context['total_plan']],
        ['Факт, м3', context['day_total'], context['night_total'], context['total_volume']],
        ['Отклонение, м3', context['day_deviation'], context['night_deviation'], context['total_deviation']],
        ['Тоннаж', context['day_tonnage'], context['night_tonnage'], context['total_tonnage']],
        ['Рейсы', context['day_trip_count'], context['night_trip_count'], context['total_trip_count']],
        ['Механические простои', '', '', context['mechanic_downtime_count']],
        ['Открытые механические простои', '', '', context['mechanic_open_downtime_count']],
        ['Механические простои, ч', '', '', context['mechanic_downtime_hours']],
    ]
    for row in summary_rows:
        sheet.append(row)

    sheet['F4'] = f"С начала месяца ({context['month_start']:%d.%m.%Y} - {context['selected_date']:%d.%m.%Y})"
    sheet['F4'].font = Font(bold=True)
    month_rows = [
        ['Показатель', 'Значение'],
        ['План, м3', context['month_plan_total']],
        ['Факт, м3', context['month_total_volume']],
        ['Отклонение, м3', context['month_total_deviation']],
        ['Тоннаж', context['month_total_tonnage']],
        ['Рейсы', context['month_total_trip_count']],
    ]
    for row_index, row in enumerate(month_rows, start=5):
        for column_offset, value in enumerate(row, start=6):
            cell = sheet.cell(row_index, column_offset, value)
            if row_index == 5:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', fgColor='17232E')

    day_end_row = append_customer_shift_table(sheet, 'I смена (дневная 08:00 - 20:00)', 13, 1, context['rows_by_shift']['day'])
    night_end_row = append_customer_shift_table(sheet, 'II смена (ночная 20:00 - 08:00)', 13, 12, context['rows_by_shift']['night'])

    downtime_start_row = max(day_end_row, night_end_row) + 1
    sheet.cell(downtime_start_row, 1, 'Механические простои за дату')
    sheet.cell(downtime_start_row, 1).font = Font(bold=True, size=13)
    downtime_headers = ['Начало', 'Окончание', 'Техника', 'Причина', 'Длительность, ч', 'Кто зафиксировал', 'Комментарий']
    for offset, header in enumerate(downtime_headers, start=1):
        cell = sheet.cell(downtime_start_row + 1, offset, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='17232E')
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    if context['mechanic_downtime_rows']:
        for row_index, downtime in enumerate(context['mechanic_downtime_rows'], start=downtime_start_row + 2):
            values = [
                timezone.localtime(downtime['started_at']).strftime('%d.%m.%Y %H:%M'),
                timezone.localtime(downtime['ended_at']).strftime('%d.%m.%Y %H:%M') if downtime['ended_at'] else 'открыт',
                str(downtime['equipment']),
                str(downtime['reason']),
                downtime['duration_hours'],
                str(downtime['employee']) if downtime['employee'] else '-',
                downtime['comment'] or '-',
            ]
            for offset, value in enumerate(values, start=1):
                cell = sheet.cell(row_index, offset, value)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    else:
        sheet.cell(downtime_start_row + 2, 1, 'Механических простоев за выбранную дату нет.')

    for column_index in range(1, 23):
        sheet.column_dimensions[get_column_letter(column_index)].width = 16
    sheet.column_dimensions['K'].width = 36
    sheet.column_dimensions['V'].width = 36
    append_customer_reconciliation_sheet(workbook, context['excel_reconciliation'])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=\"customer_daily_report.xlsx\"'
    workbook.save(response)
    return response


def management_dashboard_context(request, access):
    completed_trips = Trip.objects.filter(status=TripStatus.COMPLETED).select_related(
        'truck',
        'excavator',
        'rock_type',
        'dump_point',
        'loading_shift',
        'unloading_shift',
    )
    active_trips = Trip.objects.filter(status=TripStatus.ACTIVE)
    open_mechanic_downtimes = (
        DowntimeEvent.objects
        .filter(ended_at__isnull=True)
        .select_related('equipment', 'reason', 'employee')
        .order_by('started_at')[:8]
    )
    selected_date = parse_customer_report_date(request)
    shift_analytics = build_shift_analytics(selected_date, '')
    shift_analytics_day = build_shift_analytics(selected_date, 'day')
    shift_analytics_night = build_shift_analytics(selected_date, 'night')
    shift_analytics_shift_cards = [
        {
            'label': 'День',
            'totals': shift_analytics_day['totals'],
            'url': f"{reverse('shift_analytics_report')}?date={selected_date:%Y-%m-%d}&shift_type=day",
        },
        {
            'label': 'Ночь',
            'totals': shift_analytics_night['totals'],
            'url': f"{reverse('shift_analytics_report')}?date={selected_date:%Y-%m-%d}&shift_type=night",
        },
    ]
    completed_trip_list = list(completed_trips)
    daily_trips = [
        trip
        for trip in completed_trip_list
        if trip_report_date(trip) == selected_date
    ]
    daily_total_volume = sum((trip.volume_m3 or 0) for trip in daily_trips)
    daily_total_tonnage = sum((trip.tonnage or 0) for trip in daily_trips)
    manual_plan_totals = shift_plan_totals(selected_date)
    manual_plan_by_shift = manual_plan_totals['by_shift']
    daily_plan_source = 'из сменных планов админки' if manual_plan_totals['volume_m3'] else 'по заданиям в рейсах'
    daily_shift_totals = {
        'day': {
            'label': 'Дневная смена',
            'css_class': 'day',
            'volume': Decimal('0'),
            'plan': manual_plan_by_shift['day']['volume_m3'],
            'tonnage': Decimal('0'),
            'trip_count': 0,
        },
        'night': {
            'label': 'Ночная смена',
            'css_class': 'night',
            'volume': Decimal('0'),
            'plan': manual_plan_by_shift['night']['volume_m3'],
            'tonnage': Decimal('0'),
            'trip_count': 0,
        },
        'unknown': {
            'label': 'Смена не указана',
            'css_class': 'unknown',
            'volume': Decimal('0'),
            'plan': manual_plan_by_shift['unknown']['volume_m3'],
            'tonnage': Decimal('0'),
            'trip_count': 0,
        },
    }
    manual_plan_present_by_shift = {
        key: bool(value['volume_m3'])
        for key, value in manual_plan_by_shift.items()
    }
    for trip in daily_trips:
        shift_type = trip_shift_type(trip) or 'unknown'
        if shift_type not in daily_shift_totals:
            shift_type = 'unknown'
        daily_shift_totals[shift_type]['volume'] += trip.volume_m3 or 0
        if not manual_plan_present_by_shift.get(shift_type):
            daily_shift_totals[shift_type]['plan'] += trip.planned_volume_m3 or 0
        daily_shift_totals[shift_type]['tonnage'] += trip.tonnage or 0
        daily_shift_totals[shift_type]['trip_count'] += 1
    daily_plan_total = sum((item['plan'] for item in daily_shift_totals.values()), Decimal('0'))
    daily_deviation = daily_total_volume - daily_plan_total
    daily_plan_completion_percent = calculate_plan_completion_percent(daily_total_volume, daily_plan_total)
    daily_plan_completion_class = (
        'success'
        if daily_plan_completion_percent is not None and daily_plan_completion_percent >= Decimal('100')
        else 'danger'
    )
    daily_shift_comparison = []
    max_daily_shift_volume = max((item['volume'] for item in daily_shift_totals.values()), default=Decimal('0'))
    max_daily_shift_plan = max((item['plan'] for item in daily_shift_totals.values()), default=Decimal('0'))
    for key in ('day', 'night', 'unknown'):
        item = daily_shift_totals[key]
        if key == 'unknown' and item['trip_count'] == 0:
            continue
        daily_shift_comparison.append({
            **item,
            'deviation': item['volume'] - item['plan'],
        })
    daily_trend = build_management_daily_trend(completed_trip_list, selected_date)
    max_daily_trend_volume = max((item['volume'] for item in daily_trend), default=Decimal('0'))
    max_daily_trend_plan = max((item['plan'] for item in daily_trend), default=Decimal('0'))
    trend_total_volume = sum((item['volume'] for item in daily_trend), Decimal('0'))
    trend_total_plan = sum((item['plan'] for item in daily_trend), Decimal('0'))
    trend_total_deviation = trend_total_volume - trend_total_plan
    trend_trip_count = sum(item['trip_count'] for item in daily_trend)
    trend_completion_percent = calculate_plan_completion_percent(trend_total_volume, trend_total_plan)
    trend_best_day = max(daily_trend, key=lambda item: item['volume']) if trend_trip_count else None
    trend_worst_day = min(
        (item for item in daily_trend if item['has_plan'] or item['trip_count']),
        key=lambda item: item['deviation'],
        default=None,
    )
    completed_summary = completed_trips.aggregate(
        total_volume=Sum('volume_m3'),
        total_tonnage=Sum('tonnage'),
        trip_count=Count('id'),
    )
    top_excavators = (
        completed_trips
        .values('excavator__garage_number')
        .annotate(total_volume=Sum('volume_m3'), trip_count=Count('id'))
        .order_by('-total_volume')[:5]
    )
    top_rocks = (
        completed_trips
        .values('rock_type__name')
        .annotate(total_volume=Sum('volume_m3'), trip_count=Count('id'))
        .order_by('-total_volume')[:5]
    )
    daily_excavator_totals = defaultdict(lambda: {'volume': 0, 'trip_count': 0})
    daily_rock_totals = defaultdict(lambda: {'volume': 0, 'trip_count': 0})
    for trip in daily_trips:
        excavator_name = f'Экскаватор {trip.excavator.garage_number}' if trip.excavator else '-'
        rock_name = str(trip.rock_type) if trip.rock_type else '-'
        daily_excavator_totals[excavator_name]['volume'] += trip.volume_m3 or 0
        daily_excavator_totals[excavator_name]['trip_count'] += 1
        daily_rock_totals[rock_name]['volume'] += trip.volume_m3 or 0
        daily_rock_totals[rock_name]['trip_count'] += 1
    daily_top_excavators = sorted(
        [
            {'name': name, **values}
            for name, values in daily_excavator_totals.items()
        ],
        key=lambda item: item['volume'],
        reverse=True,
    )[:5]
    daily_top_rocks = sorted(
        [
            {'name': name, **values}
            for name, values in daily_rock_totals.items()
        ],
        key=lambda item: item['volume'],
        reverse=True,
    )[:5]
    recent_completed_trips = completed_trips.order_by('-completed_at')[:8]

    max_excavator_volume = max((item['total_volume'] or 0 for item in top_excavators), default=0)
    max_rock_volume = max((item['total_volume'] or 0 for item in top_rocks), default=0)
    daily_max_excavator_volume = max((item['volume'] or 0 for item in daily_top_excavators), default=0)
    daily_max_rock_volume = max((item['volume'] or 0 for item in daily_top_rocks), default=0)

    return {
        'access': access,
        'current_time': timezone.localtime(timezone.now()).strftime('%H:%M'),
        'current_date': timezone.localdate().strftime('%d.%m.%Y'),
        'management_header_shift_label': f'Срез на {selected_date:%d.%m.%Y}',
        'management_nav_items': [
            {'label': 'Витрина', 'href': reverse('management_dashboard'), 'active': True},
            {'label': 'Сутки', 'href': f"{reverse('customer_daily_report')}?date={selected_date:%Y-%m-%d}"},
            {'label': 'Объемы', 'href': reverse('volume_report')},
            {'label': 'Простои', 'href': reverse('downtime_report')},
            {'label': 'Механики', 'href': reverse('mechanic_dashboard')},
            {'label': 'Пилот', 'href': reverse('pilot_report_checklist')},
        ],
        'selected_date': selected_date,
        'daily_plan_total': daily_plan_total,
        'daily_plan_source': daily_plan_source,
        'daily_total_volume': daily_total_volume,
        'daily_deviation': daily_deviation,
        'daily_plan_completion_percent': daily_plan_completion_percent,
        'daily_plan_completion_css': str(daily_plan_completion_percent or 0),
        'daily_plan_completion_class': daily_plan_completion_class,
        'daily_total_tonnage': daily_total_tonnage,
        'daily_trip_count': len(daily_trips),
        'daily_top_excavators': daily_top_excavators,
        'daily_top_rocks': daily_top_rocks,
        'shift_analytics': shift_analytics,
        'shift_analytics_totals': shift_analytics['totals'],
        'shift_analytics_excavator_rows': shift_analytics['excavator_rows'][:4],
        'shift_analytics_truck_rows': shift_analytics['truck_rows'][:4],
        'shift_analytics_downtime_reason_rows': shift_analytics['downtime_reason_rows'][:4],
        'shift_analytics_shift_cards': shift_analytics_shift_cards,
        'daily_max_excavator_volume': daily_max_excavator_volume,
        'daily_max_rock_volume': daily_max_rock_volume,
        'daily_shift_comparison': daily_shift_comparison,
        'max_daily_shift_volume': max_daily_shift_volume,
        'max_daily_shift_plan': max_daily_shift_plan,
        'daily_trend': daily_trend,
        'max_daily_trend_volume': max_daily_trend_volume,
        'max_daily_trend_plan': max_daily_trend_plan,
        'trend_total_volume': trend_total_volume,
        'trend_total_plan': trend_total_plan,
        'trend_total_deviation': trend_total_deviation,
        'trend_trip_count': trend_trip_count,
        'trend_completion_percent': trend_completion_percent,
        'trend_best_day': trend_best_day,
        'trend_worst_day': trend_worst_day,
        'total_volume': completed_summary['total_volume'] or 0,
        'total_tonnage': completed_summary['total_tonnage'] or 0,
        'completed_trip_count': completed_summary['trip_count'] or 0,
        'active_trip_count': active_trips.count(),
        'open_mechanic_downtime_count': DowntimeEvent.objects.filter(ended_at__isnull=True).count(),
        'carryover_trip_count': completed_trips.filter(is_carryover=True).count(),
        'mechanic_downtime_count': len(build_mechanic_downtime_rows(selected_date)),
        'open_mechanic_downtimes': open_mechanic_downtimes,
        'top_excavators': top_excavators,
        'top_rocks': top_rocks,
        'max_excavator_volume': max_excavator_volume,
        'max_rock_volume': max_rock_volume,
        'recent_completed_trips': recent_completed_trips,
    }


def management_dashboard_view(request):
    access = get_reports_access(request, {'manager', 'admin', 'dispatcher'})
    if not access:
        return redirect('login' if not request.session.get('employee_access_id') else 'role_home')
    return render(request, 'reports/management_dashboard.html', management_dashboard_context(request, access))


def write_key_value_rows(sheet, start_row, rows):
    for offset, (label, value) in enumerate(rows):
        row = start_row + offset
        sheet.cell(row=row, column=1, value=label)
        sheet.cell(row=row, column=2, value=value)
    return start_row + len(rows)


def style_management_export_sheet(sheet):
    header_fill = PatternFill('solid', fgColor='12232E')
    header_font = Font(color='FFFFFF', bold=True)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if cell.row == 1:
                cell.font = Font(bold=True, size=14)
    for row in sheet.iter_rows():
        if row[0].value and all(cell.value for cell in row[:2]):
            if row[0].row > 1 and str(row[0].value).startswith(('Показатель', 'Дата', 'Смена')):
                for cell in row:
                    cell.fill = header_fill
                    cell.font = header_font
    for column in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 22


def management_dashboard_export_view(request):
    access = get_reports_access(request, {'manager', 'admin', 'dispatcher'})
    if not access:
        return redirect('login' if not request.session.get('employee_access_id') else 'role_home')

    context = management_dashboard_context(request, access)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = 'Сводка'
    summary_sheet['A1'] = 'Витрина руководства'
    summary_sheet['A2'] = f"Дата среза: {context['selected_date']:%d.%m.%Y}"
    summary_sheet['A3'] = f"Сформировал: {access.employee.full_name}"

    rows = [
        ('Факт за сутки, м3', context['daily_total_volume']),
        ('План за сутки, м3', context['daily_plan_total']),
        ('Выполнение плана, %', context['daily_plan_completion_percent'] or 'Нет плана'),
        ('Отклонение за сутки, м3', context['daily_deviation']),
        ('Тоннаж за сутки, т', context['daily_total_tonnage']),
        ('Рейсы за сутки', context['daily_trip_count']),
        ('Отгружено экскаваторами за дату, рейсов', context['shift_analytics_totals']['loaded_trip_count']),
        ('Разгружено самосвалами за дату, рейсов', context['shift_analytics_totals']['unloaded_trip_count']),
        ('Открытые груженые рейсы', context['shift_analytics_totals']['open_trip_count']),
        ('Сменная аналитика, объем м3', context['shift_analytics_totals']['volume_display']),
        ('Сменная аналитика, тоннаж т', context['shift_analytics_totals']['tonnage_display']),
        ('Простои за дату, событий', context['shift_analytics_totals']['downtime_count']),
        ('Простои за дату, часов', context['shift_analytics_totals']['downtime_hours_display']),
        ('Активные рейсы', context['active_trip_count']),
        ('Открытые механические простои', context['open_mechanic_downtime_count']),
        ('Переходящие рейсы', context['carryover_trip_count']),
        ('Факт за 7 дней, м3', context['trend_total_volume']),
        ('План за 7 дней, м3', context['trend_total_plan']),
        ('Выполнение за неделю, %', context['trend_completion_percent'] or 'Нет плана'),
        ('Отклонение за неделю, м3', context['trend_total_deviation']),
        ('Рейсы за 7 дней', context['trend_trip_count']),
    ]
    summary_sheet.append([])
    summary_sheet.append(['Показатель', 'Значение'])
    write_key_value_rows(summary_sheet, 6, rows)

    trend_sheet = workbook.create_sheet('Динамика 7 дней')
    trend_sheet.append(['Дата', 'Факт, м3', 'План, м3', 'Выполнение, %', 'Отклонение, м3', 'Рейсы', 'Тоннаж, т'])
    for item in context['daily_trend']:
        trend_sheet.append([
            item['date'].strftime('%d.%m.%Y'),
            item['volume'],
            item['plan'],
            item['completion_percent'] if item['has_plan'] else 'Нет плана',
            item['deviation'],
            item['trip_count'],
            item['tonnage'],
        ])

    shifts_sheet = workbook.create_sheet('День ночь')
    shifts_sheet.append(['Смена', 'Факт, м3', 'План, м3', 'Отклонение, м3', 'Рейсы', 'Тоннаж, т'])
    for item in context['daily_shift_comparison']:
        shifts_sheet.append([
            item['label'],
            item['volume'],
            item['plan'],
            item['deviation'],
            item['trip_count'],
            item['tonnage'],
        ])

    shift_analytics_sheet = workbook.create_sheet('Сменная аналитика')
    shift_analytics_sheet.append(['Сменная аналитика производственного контура'])
    shift_analytics_sheet.append(['Дата', context['selected_date'].strftime('%d.%m.%Y')])
    shift_analytics_sheet.append(['Смена', 'Все смены'])
    shift_analytics_sheet.append([])
    shift_analytics_sheet.append(['Показатель', 'Значение'])
    for label, key in [
        ('Отгружено экскаваторами, рейсов', 'loaded_trip_count'),
        ('Разгружено самосвалами, рейсов', 'unloaded_trip_count'),
        ('Объем, м3', 'volume_display'),
        ('Тоннаж, т', 'tonnage_display'),
        ('Переходящие рейсы', 'carryover_count'),
        ('Открытые рейсы', 'open_trip_count'),
        ('Простои, событий', 'downtime_count'),
        ('Простои, часов', 'downtime_hours_display'),
    ]:
        shift_analytics_sheet.append([label, context['shift_analytics_totals'][key]])

    common_headers = ['Наименование', 'Рейсы', 'Объем, м3', 'Тоннаж, т', 'Породы', 'Разгрузка', 'Горизонт/блок']
    common_getters = [
        lambda row: row['label'],
        lambda row: row['trip_count'],
        lambda row: row['volume_display'],
        lambda row: row['tonnage_display'],
        lambda row: row['rocks_display'],
        lambda row: row['dump_points_display'],
        lambda row: row['faces_display'],
    ]
    append_shift_analytics_rows(
        shift_analytics_sheet,
        'Экскаваторы',
        context['shift_analytics']['excavator_rows'],
        common_headers,
        common_getters,
    )
    append_shift_analytics_rows(
        shift_analytics_sheet,
        'Самосвалы',
        context['shift_analytics']['truck_rows'],
        common_headers,
        common_getters,
    )
    append_shift_analytics_rows(
        shift_analytics_sheet,
        'Простои',
        context['shift_analytics']['downtime_reason_rows'],
        ['Причина', 'Событий', 'Открыто', 'Часов', 'Техника', 'Сотрудники'],
        [
            lambda row: row['label'],
            lambda row: row['count'],
            lambda row: row['open_count'],
            lambda row: row['duration_display'],
            lambda row: row['equipment_display'],
            lambda row: row['employees_display'],
        ],
    )

    for sheet in workbook.worksheets:
        style_management_export_sheet(sheet)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="management_dashboard.xlsx"'
    workbook.save(response)
    return response

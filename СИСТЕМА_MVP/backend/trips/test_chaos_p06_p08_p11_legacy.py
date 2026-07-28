import json
from datetime import datetime, timezone as dt_timezone
from decimal import Decimal
from io import BytesIO

from django.db import connection
from django.test import Client, TestCase
from django.test.utils import CaptureQueriesContext
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from assignments.models import AssignmentStatus, EquipmentAssignment, HaulAssignment
from core.production_time import production_work_date
from downtimes.models import DowntimeEvent, DowntimeReason
from references.equipment_states import upsert_default_equipment_states
from references.models import (
    Dormitory,
    DormitoryBlock,
    DormitorySection,
    DumpPoint,
    Equipment,
    EquipmentModel,
    EquipmentType,
    RockType,
    TruckCapacityRule,
)
from reports.shift_analytics import build_shift_analytics
from shifts.models import EmployeeShift
from trips.models import Trip, TripClientAction, TripStatus
from trips.views import EXCAVATOR_AUTO_DOWNTIME_COMMENT, finalize_trip_unloaded
from users.models import DriverPrimaryRegistration, Employee, EmployeeAccess, Role


class ChaosP06P08LoadingParityRegressionTests(TestCase):
    """Regression contract for QA-CHAOS-P1-P06-015-P08-006."""

    def setUp(self):
        upsert_default_equipment_states()
        self.operator_role = Role.objects.create(
            code='excavator_operator',
            name='Машинист экскаватора',
        )
        self.driver_role = Role.objects.create(code='driver', name='Водитель')
        self.operator = Employee.objects.create(
            full_name='Машинист P06/P08',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        self.operator_access = EmployeeAccess.objects.create(
            employee=self.operator,
            role=self.operator_role,
            access_code='P06008',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )

        self.excavator_type = EquipmentType.objects.create(name='Экскаватор')
        self.truck_type = EquipmentType.objects.create(name='Самосвал')
        self.excavator_model = EquipmentModel.objects.create(
            equipment_type=self.excavator_type,
            name='Экскаватор P06/P08',
        )
        self.truck_model = EquipmentModel.objects.create(
            equipment_type=self.truck_type,
            name='Самосвал P06/P08',
            body_volume_m3=Decimal('40.00'),
        )
        self.excavator = Equipment.objects.create(
            equipment_type=self.excavator_type,
            model=self.excavator_model,
            garage_number='P06-EXC',
        )
        self.trucks = [
            Equipment.objects.create(
                equipment_type=self.truck_type,
                model=self.truck_model,
                garage_number=f'P06-TRUCK-{index}',
            )
            for index in (1, 2)
        ]
        self.rock = RockType.objects.create(
            name='Руда P06/P08',
            density=Decimal('2.0000'),
        )
        self.dump_point = DumpPoint.objects.create(name='ККД P06/P08')
        self.capacity_rule = TruckCapacityRule.objects.create(
            equipment_model=self.truck_model,
            rock_type=self.rock,
            volume_m3=Decimal('40.00'),
        )

        self.operator_shift = EmployeeShift.objects.create(
            employee=self.operator,
            shift_type='day',
            workplace_code='excavator_operator',
            equipment=self.excavator,
            opened_at=timezone.now(),
            opened_by=self.operator,
        )
        EquipmentAssignment.objects.create(
            employee=self.operator,
            role=self.operator_role,
            equipment=self.excavator,
            shift_type='day',
            assigned_by=self.operator,
            status=AssignmentStatus.ACCEPTED,
            accepted_at=timezone.now(),
        )

        self.drivers = []
        self.driver_accesses = []
        self.driver_shifts = []
        self.haul_assignments = []
        for index, truck in enumerate(self.trucks, start=1):
            driver, access, shift = self._create_driver(index, truck)
            self.drivers.append(driver)
            self.driver_accesses.append(access)
            self.driver_shifts.append(shift)
            self.haul_assignments.append(
                HaulAssignment.objects.create(
                    excavator=self.excavator,
                    truck=truck,
                    assigned_by=self.operator,
                    status=AssignmentStatus.ACCEPTED,
                    accepted_at=timezone.now(),
                )
            )

        self.waiting_reason, _ = DowntimeReason.objects.get_or_create(
            name='Ожидание самосвалов',
            defaults={
                'equipment_type': self.excavator_type,
                'show_for_excavator_operator': True,
            },
        )
        self.waiting_reason.equipment_type = self.excavator_type
        self.waiting_reason.show_for_excavator_operator = True
        self.waiting_reason.is_active = True
        self.waiting_reason.save(
            update_fields=[
                'equipment_type',
                'show_for_excavator_operator',
                'is_active',
            ]
        )

        session = self.client.session
        session['employee_access_id'] = self.operator_access.id
        session.save()

    def _create_driver(self, index, truck):
        driver = Employee.objects.create(
            full_name=f'Водитель P06/P08 {index}',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        access = EmployeeAccess.objects.create(
            employee=driver,
            role=self.driver_role,
            access_code=f'P06DRV{index}',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        dormitory, _ = Dormitory.objects.get_or_create(number='P06')
        block, _ = DormitoryBlock.objects.get_or_create(
            dormitory=dormitory,
            name='Блок P06',
        )
        section, _ = DormitorySection.objects.get_or_create(
            block=block,
            name=str(index),
        )
        DriverPrimaryRegistration.objects.create(
            employee=driver,
            dormitory_section=section,
        )
        shift = EmployeeShift.objects.create(
            employee=driver,
            shift_type='day',
            workplace_code='driver',
            equipment=truck,
            opened_at=timezone.now(),
            opened_by=driver,
        )
        EquipmentAssignment.objects.create(
            employee=driver,
            role=self.driver_role,
            equipment=truck,
            shift=shift,
            assigned_by=self.operator,
            status=AssignmentStatus.ACCEPTED,
            accepted_at=timezone.now(),
        )
        return driver, access, shift

    def _post_json_load(self, truck_index=0, action_id='p06-json-load'):
        response = self.client.post(
            reverse('excavator_truck_loaded'),
            data=json.dumps(
                {
                    'client_action_id': action_id,
                    'truck_id': self.trucks[truck_index].id,
                    'excavator_id': self.excavator.id,
                    'dump_point_id': self.dump_point.id,
                    'rock_type_id': self.rock.id,
                    'loading_horizon': '125',
                    'loading_block': '4',
                }
            ),
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 200, response.content)
        return response.json(), Trip.objects.get(pk=response.json()['trip_id'])

    def _post_html_load(self, truck_index=1, action_id='p06-html-load'):
        response = self.client.post(
            reverse('excavator_work'),
            data={
                'client_action_id': action_id,
                'assignment': self.haul_assignments[truck_index].id,
                'rock_type': self.rock.id,
                'dump_point': self.dump_point.id,
                'loading_horizon': '125',
                'loading_block': '4',
            },
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )
        self.assertEqual(response.status_code, 200, response.content)
        return response.json(), Trip.objects.get(pk=response.json()['trip_id'])

    @staticmethod
    def _pending_state(trip):
        return (
            trip.status,
            trip.volume_m3,
            trip.tonnage,
            trip.loading_horizon,
            trip.loading_block,
            trip.dump_point_id,
            trip.assigned_dump_point_id,
            trip.actual_dump_point_id,
        )

    @staticmethod
    def _client_for_access(access):
        client = Client()
        session = client.session
        session['employee_access_id'] = access.id
        session.save()
        return client

    def test_json_and_html_create_the_same_pending_server_state(self):
        _json_payload, json_trip = self._post_json_load()
        _html_payload, html_trip = self._post_html_load()
        json_trip.refresh_from_db()
        html_trip.refresh_from_db()

        self.assertEqual(
            self._pending_state(json_trip),
            self._pending_state(html_trip),
        )
        for trip in (json_trip, html_trip):
            with self.subTest(trip=trip.pk):
                self.assertEqual(trip.status, TripStatus.LOADED_WAITING_UNLOAD)
                self.assertEqual(trip.volume_m3, Decimal('40.00'))
                self.assertEqual(trip.tonnage, Decimal('80.00'))

    def test_both_routes_are_visible_to_driver_and_loading_analytics(self):
        _json_payload, json_trip = self._post_json_load()
        _html_payload, html_trip = self._post_html_load()

        for access, expected_trip in zip(
            self.driver_accesses,
            (json_trip, html_trip),
        ):
            with self.subTest(access=access.pk):
                response = self._client_for_access(access).get(reverse('driver_work'))
                self.assertEqual(response.status_code, 200)
                self.assertEqual(response.context['active_trip'].pk, expected_trip.pk)

        work_date = production_work_date(self.operator_shift.opened_at)
        analytics = build_shift_analytics(work_date, 'day')
        self.assertEqual(analytics['totals']['loaded_trip_count'], 2)
        self.assertEqual(analytics['totals']['open_trip_count'], 2)

    def test_both_routes_apply_current_references_at_unload_and_then_freeze_fact(self):
        _json_payload, json_trip = self._post_json_load()
        _html_payload, html_trip = self._post_html_load()

        self.capacity_rule.volume_m3 = Decimal('60.00')
        self.capacity_rule.save(update_fields=['volume_m3'])
        self.rock.density = Decimal('3.0000')
        self.rock.save(update_fields=['density'])

        finalize_trip_unloaded(
            json_trip,
            driver=self.drivers[0],
            unloading_shift=self.driver_shifts[0],
        )
        finalize_trip_unloaded(
            html_trip,
            driver=self.drivers[1],
            unloading_shift=self.driver_shifts[1],
        )

        for trip in (json_trip, html_trip):
            trip.refresh_from_db()
            with self.subTest(trip=trip.pk):
                self.assertEqual(trip.status, TripStatus.COMPLETED)
                self.assertEqual(trip.volume_m3, Decimal('60.00'))
                self.assertEqual(trip.tonnage, Decimal('180.00'))

        self.capacity_rule.volume_m3 = Decimal('70.00')
        self.capacity_rule.save(update_fields=['volume_m3'])
        self.rock.density = Decimal('4.0000')
        self.rock.save(update_fields=['density'])

        analytics = build_shift_analytics(
            production_work_date(self.operator_shift.opened_at),
            'day',
        )
        self.assertEqual(analytics['totals']['volume_m3'], Decimal('120.00'))
        self.assertEqual(analytics['totals']['tonnage'], Decimal('360.00'))
        for trip in (json_trip, html_trip):
            trip.refresh_from_db()
            self.assertEqual(trip.volume_m3, Decimal('60.00'))
            self.assertEqual(trip.tonnage, Decimal('180.00'))

    def test_same_client_action_id_is_idempotent_for_both_routes(self):
        first_json, json_trip = self._post_json_load(action_id='same-json-load')
        second_json, repeated_json_trip = self._post_json_load(
            action_id='same-json-load'
        )
        first_html, html_trip = self._post_html_load(action_id='same-html-load')
        second_html, repeated_html_trip = self._post_html_load(
            action_id='same-html-load'
        )

        self.assertEqual(first_json['trip_id'], second_json['trip_id'])
        self.assertEqual(first_html['trip_id'], second_html['trip_id'])
        self.assertEqual(json_trip.pk, repeated_json_trip.pk)
        self.assertEqual(html_trip.pk, repeated_html_trip.pk)
        self.assertTrue(second_json['deduplicated'])
        self.assertTrue(second_html['deduplicated'])
        self.assertEqual(Trip.objects.count(), 2)
        self.assertEqual(
            TripClientAction.objects.filter(action_type='truck_loaded').count(),
            2,
        )

    def test_legacy_active_trip_completes_once_and_stays_report_visible(self):
        legacy_trip = Trip.objects.create(
            excavator=self.excavator,
            truck=self.trucks[0],
            excavator_operator=self.operator,
            loading_shift=self.operator_shift,
            rock_type=self.rock,
            dump_point=self.dump_point,
            assigned_dump_point=self.dump_point,
            actual_dump_point=self.dump_point,
            status=TripStatus.ACTIVE,
        )
        self.capacity_rule.volume_m3 = Decimal('60.00')
        self.capacity_rule.save(update_fields=['volume_m3'])
        self.rock.density = Decimal('3.0000')
        self.rock.save(update_fields=['density'])
        driver_client = self._client_for_access(self.driver_accesses[0])

        first = driver_client.post(
            reverse('driver_complete_trip', args=[legacy_trip.pk]),
            data={'client_action_id': 'legacy-active-unload'},
        )
        legacy_trip.refresh_from_db()
        completed_at = legacy_trip.completed_at
        second = driver_client.post(
            reverse('driver_complete_trip', args=[legacy_trip.pk]),
            data={'client_action_id': 'legacy-active-unload'},
        )

        self.assertEqual(first.status_code, 302)
        self.assertEqual(second.status_code, 302)
        legacy_trip.refresh_from_db()
        self.assertEqual(legacy_trip.status, TripStatus.COMPLETED)
        self.assertEqual(legacy_trip.completed_at, completed_at)
        self.assertEqual(legacy_trip.volume_m3, Decimal('60.00'))
        self.assertEqual(legacy_trip.tonnage, Decimal('180.00'))
        self.assertEqual(
            TripClientAction.objects.filter(
                action_type='trip_unloaded',
                client_action_id='legacy-active-unload',
                trip=legacy_trip,
            ).count(),
            1,
        )
        analytics = build_shift_analytics(
            production_work_date(self.operator_shift.opened_at),
            'day',
        )
        self.assertEqual(analytics['totals']['loaded_trip_count'], 1)
        self.assertEqual(analytics['totals']['unloaded_trip_count'], 1)
        self.assertEqual(analytics['totals']['volume_m3'], Decimal('60.00'))

    def test_html_retry_closes_current_downtime_once_and_opens_one_waiting(self):
        self.haul_assignments[1].ended_at = timezone.now()
        self.haul_assignments[1].save(update_fields=['ended_at'])
        work_reason = DowntimeReason.objects.create(
            name='Текущий простой P06/P08',
            equipment_type=self.excavator_type,
            show_for_excavator_operator=True,
        )
        current_downtime = DowntimeEvent.objects.create(
            equipment=self.excavator,
            employee=self.operator,
            reason=work_reason,
            started_at=timezone.now(),
        )

        first_payload, _trip = self._post_html_load(
            truck_index=0,
            action_id='html-downtime-once',
        )
        current_downtime.refresh_from_db()
        first_ended_at = current_downtime.ended_at
        second_payload, repeated_trip = self._post_html_load(
            truck_index=0,
            action_id='html-downtime-once',
        )

        self.assertIsNotNone(first_ended_at)
        self.assertEqual(first_payload['trip_id'], second_payload['trip_id'])
        self.assertTrue(second_payload['deduplicated'])
        current_downtime.refresh_from_db()
        self.assertEqual(current_downtime.ended_at, first_ended_at)
        self.assertEqual(
            DowntimeEvent.objects.filter(
                equipment=self.excavator,
                reason=self.waiting_reason,
                ended_at__isnull=True,
                comment=EXCAVATOR_AUTO_DOWNTIME_COMMENT,
            ).count(),
            1,
        )
        self.assertEqual(
            TripClientAction.objects.filter(
                action_type='truck_loaded',
                client_action_id='html-downtime-once',
                trip=repeated_trip,
            ).count(),
            1,
        )


class ChaosP11LegacyReportFallbackRegressionTests(TestCase):
    """Read-only report contract for the P11-012 extension."""

    production_date = datetime(2026, 7, 23).date()
    # 24.07.2026 01:30 Asia/Vladivostok: night shift of production date 23.07.
    completed_at = datetime(2026, 7, 23, 15, 30, tzinfo=dt_timezone.utc)
    # Intentionally a different stage context: 23.07.2026 10:00 Vladivostok/day.
    created_at = datetime(2026, 7, 23, 0, 0, tzinfo=dt_timezone.utc)

    def setUp(self):
        self.admin_role = Role.objects.create(code='admin', name='Администратор')
        self.admin = Employee.objects.create(
            full_name='Администратор P11',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        self.admin_access = EmployeeAccess.objects.create(
            employee=self.admin,
            role=self.admin_role,
            access_code='P11012',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        session = self.client.session
        session['employee_access_id'] = self.admin_access.id
        session.save()

        self.excavator_type = EquipmentType.objects.create(name='Экскаватор')
        self.truck_type = EquipmentType.objects.create(name='Самосвал')
        self.excavator_model = EquipmentModel.objects.create(
            equipment_type=self.excavator_type,
            name='Экскаватор legacy P11',
        )
        self.truck_model = EquipmentModel.objects.create(
            equipment_type=self.truck_type,
            name='Самосвал legacy P11',
            body_volume_m3=Decimal('40.00'),
        )
        self.excavator = Equipment.objects.create(
            equipment_type=self.excavator_type,
            model=self.excavator_model,
            garage_number='P11-EXC',
        )
        self.truck = Equipment.objects.create(
            equipment_type=self.truck_type,
            model=self.truck_model,
            garage_number='P11-TRUCK',
        )
        self.rock = RockType.objects.create(
            name='Руда legacy P11',
            density=Decimal('2.0000'),
        )
        self.dump_point = DumpPoint.objects.create(name='ККД legacy P11')
        self.capacity_rule = TruckCapacityRule.objects.create(
            equipment_model=self.truck_model,
            rock_type=self.rock,
            volume_m3=Decimal('40.00'),
        )
        self.operator = Employee.objects.create(
            full_name='Legacy машинист',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        self.driver = Employee.objects.create(
            full_name='Legacy водитель',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        self.trip = Trip.objects.create(
            excavator=self.excavator,
            truck=self.truck,
            excavator_operator=self.operator,
            driver=self.driver,
            loading_shift=None,
            unloading_shift=None,
            rock_type=self.rock,
            dump_point=self.dump_point,
            assigned_dump_point=self.dump_point,
            actual_dump_point=self.dump_point,
            volume_m3=None,
            tonnage=None,
            loading_horizon='125',
            loading_block='4',
            status=TripStatus.COMPLETED,
            completed_at=self.completed_at,
        )
        Trip.objects.filter(pk=self.trip.pk).update(created_at=self.created_at)
        self.trip.refresh_from_db()

    @property
    def date_value(self):
        return self.production_date.isoformat()

    @staticmethod
    def _workbook(response):
        return load_workbook(BytesIO(response.content), data_only=True)

    @staticmethod
    def _key_values(sheet):
        result = {}
        for row in sheet.iter_rows(values_only=True):
            if len(row) >= 2 and row[0] is not None:
                result[str(row[0])] = row[1]
        return result

    @staticmethod
    def _row_starting_with(sheet, value):
        for row in sheet.iter_rows(values_only=True):
            if row and row[0] == value:
                return row
        raise AssertionError(f'В Excel не найдена строка {value!r}')

    @staticmethod
    def _as_decimal(value):
        if isinstance(value, Decimal):
            return value
        if value in (None, ''):
            return Decimal('0')
        return Decimal(str(value).replace(' ', '').replace(',', '.'))

    def _get(self, url_name, **params):
        response = self.client.get(reverse(url_name), params)
        self.assertEqual(response.status_code, 200, response.content[:500])
        return response

    def _trip_fingerprint(self):
        return Trip.objects.filter(pk=self.trip.pk).values(
            'status',
            'created_at',
            'completed_at',
            'loading_shift_id',
            'unloading_shift_id',
            'volume_m3',
            'tonnage',
        ).get()

    def _night_semantic_snapshot(self):
        params = {'date': self.date_value, 'shift_type': 'night'}
        transport = self._get('dispatcher_transport', **params)
        daily = self._get('customer_daily_report', date=self.date_value)
        analytics = self._get('shift_analytics_report', **params)
        dispatcher_management = self._get('dispatcher_management', **params)
        management = self._get('management_dashboard', date=self.date_value)

        transport_row = next(
            row
            for row in transport.context['rows']
            if row['equipment_label'] == self.truck.garage_number
        )
        daily_night = daily.context['rows_by_shift']['night']
        management_night = next(
            row
            for row in management.context['daily_shift_comparison']
            if 'ноч' in row['label'].lower()
        )

        transport_book = self._workbook(
            self._get('dispatcher_transport_export', **params)
        )
        transport_excel_row = self._row_starting_with(
            transport_book['Автотранспорт'],
            self.truck.garage_number,
        )

        daily_book = self._workbook(
            self._get('customer_daily_report_export', date=self.date_value)
        )
        daily_fact_row = self._row_starting_with(
            daily_book['Суточный отчет'],
            'Факт, м3',
        )
        daily_tonnage_row = self._row_starting_with(
            daily_book['Суточный отчет'],
            'Тоннаж',
        )
        daily_trip_row = self._row_starting_with(
            daily_book['Суточный отчет'],
            'Рейсы',
        )

        analytics_book = self._workbook(
            self._get('shift_analytics_report_export', **params)
        )
        analytics_values = self._key_values(analytics_book['Сменная аналитика'])

        dispatcher_management_book = self._workbook(
            self._get('dispatcher_management_export', **params)
        )
        dispatcher_management_values = self._key_values(
            dispatcher_management_book['Витрина']
        )

        management_book = self._workbook(
            self._get('management_dashboard_export', date=self.date_value)
        )
        management_values = self._key_values(management_book['Сводка'])
        management_night_excel = next(
            row
            for row in management_book['День ночь'].iter_rows(values_only=True)
            if row and isinstance(row[0], str) and 'ноч' in row[0].lower()
        )

        return {
            'transport_html': (
                transport.context['kpis']['trips'],
                transport.context['kpis']['volume'],
                transport_row['shift_type'],
                transport_row['trips'],
                transport_row['volume'],
            ),
            'transport_excel': (
                transport_excel_row[2],
                transport_excel_row[8],
                transport_excel_row[9],
            ),
            'daily_html': (
                daily.context['total_trip_count'],
                daily.context['night_trip_count'],
                daily.context['total_volume'],
                daily.context['total_tonnage'],
                len(daily_night),
            ),
            'daily_excel': (
                daily_trip_row[2],
                daily_trip_row[3],
                daily_fact_row[2],
                daily_fact_row[3],
                daily_tonnage_row[2],
                daily_tonnage_row[3],
            ),
            'analytics_html': (
                analytics.context['totals']['loaded_trip_count'],
                analytics.context['totals']['unloaded_trip_count'],
                analytics.context['totals']['volume_m3'],
                analytics.context['totals']['tonnage'],
            ),
            'analytics_excel': (
                analytics_values['Отгружено экскаваторами, рейсов'],
                analytics_values['Разгружено самосвалами, рейсов'],
                analytics_values['Объем, м3'],
                analytics_values['Тоннаж, т'],
            ),
            'dispatcher_management_html': (
                dispatcher_management.context['kpis']['trips'],
                dispatcher_management.context['kpis']['volume'],
                dispatcher_management.context['kpis']['tonnage'],
                dispatcher_management.context['shift_label'],
            ),
            'dispatcher_management_excel': (
                dispatcher_management_values['Факт'],
                dispatcher_management_values['Рейсы'],
            ),
            'management_html': (
                management.context['daily_trip_count'],
                management.context['daily_total_volume'],
                management.context['daily_total_tonnage'],
                management_night['trip_count'],
                management_night['volume'],
                management.context['shift_analytics_totals']['loaded_trip_count'],
                management.context['shift_analytics_totals']['unloaded_trip_count'],
                management.context['shift_analytics_totals']['volume_m3'],
                management.context['shift_analytics_totals']['tonnage'],
            ),
            'management_excel': (
                management_values['Рейсы за сутки'],
                management_values['Факт за сутки, м3'],
                management_values['Тоннаж за сутки, т'],
                management_values['Отгружено экскаваторами за дату, рейсов'],
                management_values['Разгружено самосвалами за дату, рейсов'],
                management_values['Сменная аналитика, объем м3'],
                management_values['Сменная аналитика, тоннаж т'],
                management_night_excel[1],
                management_night_excel[4],
                management_night_excel[5],
            ),
        }

    def test_completed_legacy_null_fact_is_stable_zero_before_and_after_reference_change(self):
        before = self._night_semantic_snapshot()

        self.capacity_rule.volume_m3 = Decimal('60.00')
        self.capacity_rule.save(update_fields=['volume_m3'])
        self.rock.density = Decimal('3.0000')
        self.rock.save(update_fields=['density'])

        after = self._night_semantic_snapshot()

        self.assertEqual(after, before)
        self.assertEqual(before['transport_html'][0:2], (1, '0'))
        self.assertEqual(before['transport_html'][3:], (1, Decimal('0')))
        self.assertEqual(before['daily_html'][0:4], (1, 1, 0, 0))
        self.assertEqual(
            before['analytics_html'],
            (1, 1, Decimal('0'), Decimal('0')),
        )
        self.assertEqual(
            before['dispatcher_management_html'][0:3],
            (1, '0', '0'),
        )
        self.assertEqual(
            before['management_html'],
            (
                1,
                0,
                0,
                1,
                0,
                1,
                1,
                Decimal('0'),
                Decimal('0'),
            ),
        )

    def test_legacy_trip_has_one_night_date_and_shift_in_html_and_excel(self):
        snapshot = self._night_semantic_snapshot()

        self.assertIn('ноч', snapshot['transport_html'][2].lower())
        self.assertIn('ноч', snapshot['transport_excel'][0].lower())
        self.assertEqual(snapshot['daily_html'][0:2], (1, 1))
        self.assertEqual(snapshot['daily_excel'][0:2], (1, 1))
        self.assertEqual(snapshot['analytics_html'][0:2], (1, 1))
        self.assertEqual(snapshot['analytics_excel'][0:2], (1, 1))
        self.assertEqual(snapshot['dispatcher_management_html'][0], 1)
        self.assertIn('ноч', snapshot['dispatcher_management_html'][3].lower())
        self.assertEqual(snapshot['management_html'][0], 1)
        self.assertEqual(snapshot['management_html'][3], 1)
        self.assertEqual(snapshot['management_excel'][0], 1)
        self.assertEqual(snapshot['management_excel'][8], 1)

    def test_all_equals_day_plus_night_for_legacy_trip(self):
        transport = {}
        analytics = {}
        dispatcher_management = {}
        for shift_type in ('', 'day', 'night'):
            suffix = shift_type or 'all'
            params = {'date': self.date_value}
            if shift_type:
                params['shift_type'] = shift_type
            transport[suffix] = self._get(
                'dispatcher_transport',
                **params,
            ).context['kpis']
            analytics[suffix] = self._get(
                'shift_analytics_report',
                **params,
            ).context['totals']
            dispatcher_management[suffix] = self._get(
                'dispatcher_management',
                **params,
            ).context['kpis']

        self.assertEqual(
            transport['all']['trips'],
            transport['day']['trips'] + transport['night']['trips'],
        )
        self.assertEqual(
            self._as_decimal(transport['all']['volume']),
            self._as_decimal(transport['day']['volume'])
            + self._as_decimal(transport['night']['volume']),
        )
        for key in (
            'loaded_trip_count',
            'unloaded_trip_count',
            'volume_m3',
            'tonnage',
        ):
            self.assertEqual(
                analytics['all'][key],
                analytics['day'][key] + analytics['night'][key],
                key,
            )
        self.assertEqual(
            dispatcher_management['all']['trips'],
            dispatcher_management['day']['trips']
            + dispatcher_management['night']['trips'],
        )
        self.assertEqual(
            self._as_decimal(dispatcher_management['all']['volume']),
            self._as_decimal(dispatcher_management['day']['volume'])
            + self._as_decimal(dispatcher_management['night']['volume']),
        )

        daily = self._get(
            'customer_daily_report',
            date=self.date_value,
        ).context
        self.assertEqual(
            daily['total_trip_count'],
            daily['day_trip_count'] + daily['night_trip_count'],
        )
        self.assertEqual(
            daily['total_volume'],
            daily['day_total'] + daily['night_total'],
        )
        management = self._get(
            'management_dashboard',
            date=self.date_value,
        ).context
        self.assertEqual(
            management['daily_trip_count'],
            sum(row['trip_count'] for row in management['daily_shift_comparison']),
        )
        self.assertEqual(
            management['daily_total_volume'],
            sum(
                (row['volume'] for row in management['daily_shift_comparison']),
                Decimal('0'),
            ),
        )

    def test_management_dynamics_uses_night_for_every_granularity(self):
        for granularity in ('hour', 'day', 'shift', 'month'):
            with self.subTest(granularity=granularity):
                night = self._get(
                    'management_dynamics',
                    date_from=self.date_value,
                    date_to=self.date_value,
                    shift_type='night',
                    granularity=granularity,
                ).context['dynamics']
                day = self._get(
                    'management_dynamics',
                    date_from=self.date_value,
                    date_to=self.date_value,
                    shift_type='day',
                    granularity=granularity,
                ).context['dynamics']
                self.assertEqual(night['trip_count'], 1)
                self.assertEqual(night['total_volume'], Decimal('0'))
                self.assertEqual(day['trip_count'], 0)
                self.assertEqual(day['total_volume'], Decimal('0'))

    def test_report_reads_do_not_mutate_legacy_trip_or_database(self):
        before = self._trip_fingerprint()
        trip_count = Trip.objects.count()

        with CaptureQueriesContext(connection) as captured:
            first = self._night_semantic_snapshot()
            second = self._night_semantic_snapshot()

        mutating_sql = [
            query['sql']
            for query in captured.captured_queries
            if query['sql'].lstrip().upper().startswith(
                ('INSERT ', 'UPDATE ', 'DELETE ', 'REPLACE ')
            )
        ]
        self.assertEqual(mutating_sql, [])
        self.assertEqual(first, second)
        self.assertEqual(Trip.objects.count(), trip_count)
        self.assertEqual(self._trip_fingerprint(), before)
        self.trip.refresh_from_db()
        self.assertIsNone(self.trip.volume_m3)
        self.assertIsNone(self.trip.tonnage)

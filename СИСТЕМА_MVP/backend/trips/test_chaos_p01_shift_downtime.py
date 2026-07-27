import json
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

from django.core.exceptions import ValidationError
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from assignments.models import (
    AssignmentStatus,
    EquipmentAssignment,
    ExcavatorPlacement,
    HaulAssignment,
)
from core.production_time import production_day_bounds, production_work_date
from downtimes.models import DowntimeEvent, DowntimeReason
from references.equipment_states import upsert_default_equipment_states
from references.models import (
    Dormitory,
    DormitoryBlock,
    DormitorySection,
    DumpPoint,
    Equipment,
    EquipmentModel,
    EquipmentType,
    RockType,
)
from reports.shift_analytics import build_shift_analytics
from shifts.models import EmployeeShift, ShiftClientAction
from shifts.services import (
    calculate_truck_shift_progress,
    close_driver_shift,
    close_excavator_shift,
    open_driver_shift,
    open_excavator_shift,
)
from trips.models import (
    DispatcherActionLog,
    DispatcherActionType,
    Trip,
    TripClientAction,
    TripStatus,
)
from users.models import DriverPrimaryRegistration, Employee, EmployeeAccess, Role
from users.views import driver_shift_downtime_seconds
from trips.views import equipment_shift_downtime_seconds


class ChaosDriverShiftHandoffRegressionTests(TestCase):
    def setUp(self):
        self.driver_role = Role.objects.create(code='driver', name='Водитель')
        self.truck_type = EquipmentType.objects.create(name='Самосвал')
        self.excavator_type = EquipmentType.objects.create(name='Экскаватор')
        self.truck_model = EquipmentModel.objects.create(
            equipment_type=self.truck_type,
            name='БелАЗ CHAOS P01',
            fuel_capacity_limit_l=Decimal('2000'),
        )
        self.truck = Equipment.objects.create(
            equipment_type=self.truck_type,
            model=self.truck_model,
            garage_number='CHAOS-TRUCK-002',
        )
        self.excavator = Equipment.objects.create(
            equipment_type=self.excavator_type,
            garage_number='CHAOS-EXC-002',
        )
        self.rock = RockType.objects.create(name='Руда CHAOS P01')
        self.dump_point = DumpPoint.objects.create(name='ККД CHAOS P01')
        self.driver = self.create_driver(
            'Водитель исходной смены',
            'CHAOS-DRIVER-OLD',
            section_name='QA-OLD',
        )
        self.replacement = self.create_driver(
            'Водитель сменщик',
            'CHAOS-DRIVER-NEW',
            section_name='QA-NEW',
        )
        self.assertEqual(
            set(
                DriverPrimaryRegistration.objects
                .filter(employee__in=(self.driver, self.replacement))
                .values_list('dormitory_section__name', flat=True)
            ),
            {'QA-OLD', 'QA-NEW'},
        )
        self.work_assignment = EquipmentAssignment.objects.create(
            employee=self.driver,
            role=self.driver_role,
            equipment=self.truck,
            shift_type='day',
            assigned_by=self.driver,
            status=AssignmentStatus.ACCEPTED,
            accepted_at=timezone.now(),
        )
        self.replacement_work_assignment = EquipmentAssignment.objects.create(
            employee=self.replacement,
            role=self.driver_role,
            equipment=self.truck,
            shift_type='night',
            assigned_by=self.driver,
            status=AssignmentStatus.ACCEPTED,
            accepted_at=timezone.now(),
        )
        self.downtime_reason = DowntimeReason.objects.create(
            name='Простой CHAOS P01 водитель',
            equipment_type=self.truck_type,
            show_for_truck_driver=True,
        )

    def create_driver(self, full_name, access_code, *, section_name):
        employee = Employee.objects.create(
            full_name=full_name,
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        access = EmployeeAccess.objects.create(
            employee=employee,
            role=self.driver_role,
            access_code=access_code,
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        dormitory, _ = Dormitory.objects.get_or_create(number='CHAOS-P01')
        block, _ = DormitoryBlock.objects.get_or_create(
            dormitory=dormitory,
            name='Блок CHAOS',
        )
        section, _ = DormitorySection.objects.get_or_create(
            block=block,
            name=section_name,
        )
        DriverPrimaryRegistration.objects.create(
            employee=employee,
            dormitory_section=section,
        )
        return employee

    @staticmethod
    def opening_readings():
        return {
            'start_fuel': Decimal('1000'),
            'start_mileage': Decimal('10000'),
            'start_engine_hours': Decimal('1000'),
        }

    @staticmethod
    def closing_readings():
        return {
            'end_fuel': Decimal('900'),
            'end_mileage': Decimal('10100'),
            'end_engine_hours': Decimal('1010'),
        }

    def open_driver_shift(self, employee, action_id):
        return open_driver_shift(
            employee=employee,
            work_assignment=self.work_assignment,
            readings=self.opening_readings(),
            client_action_id=action_id,
        )[0]

    def manually_create_handoff(self):
        old_opened_at = timezone.now() - timedelta(hours=1)
        old_closed_at = timezone.now() - timedelta(minutes=10)
        old_shift = EmployeeShift.objects.create(
            employee=self.driver,
            shift_type='day',
            workplace_code='driver',
            equipment=self.truck,
            start_fuel=Decimal('1000'),
            start_mileage=Decimal('10000'),
            start_engine_hours=Decimal('1000'),
            end_fuel=Decimal('900'),
            end_mileage=Decimal('10100'),
            end_engine_hours=Decimal('1010'),
            opened_at=old_opened_at,
            closed_at=old_closed_at,
            opened_by=self.driver,
            closed_by=self.driver,
        )
        event = DowntimeEvent.objects.create(
            equipment=self.truck,
            employee=self.driver,
            reason=self.downtime_reason,
            started_at=timezone.now() - timedelta(minutes=30),
            comment='Исходный простой предшественника',
        )
        replacement_shift = EmployeeShift.objects.create(
            employee=self.replacement,
            shift_type='day',
            workplace_code='driver',
            equipment=self.truck,
            start_fuel=Decimal('900'),
            start_mileage=Decimal('10100'),
            start_engine_hours=Decimal('1010'),
            opened_at=timezone.now() - timedelta(minutes=5),
            opened_by=self.replacement,
        )
        EquipmentAssignment.objects.create(
            employee=self.replacement,
            role=self.driver_role,
            equipment=self.truck,
            shift=replacement_shift,
            assigned_by=self.replacement,
            status=AssignmentStatus.ACCEPTED,
            accepted_at=timezone.now(),
        )
        replacement_access = EmployeeAccess.objects.get(
            employee=self.replacement,
            role=self.driver_role,
        )
        session = self.client.session
        session['employee_access_id'] = replacement_access.id
        session.save()
        return old_shift, replacement_shift, event

    def test_driver_shift_can_close_with_loaded_trip_and_preserves_one_open_trip(self):
        shift = self.open_driver_shift(self.driver, 'chaos-open-loaded-trip')
        trip = Trip.objects.create(
            excavator=self.excavator,
            truck=self.truck,
            excavator_operator=None,
            driver=None,
            rock_type=self.rock,
            dump_point=self.dump_point,
            status=TripStatus.LOADED_WAITING_UNLOAD,
        )

        try:
            closed_shift, created = close_driver_shift(
                shift=shift,
                employee=self.driver,
                readings=self.closing_readings(),
                client_action_id='chaos-close-loaded-trip',
            )
        except ValidationError as error:
            self.fail(f'Загруженный рейс не должен блокировать пересменку: {error}')

        closed_shift.refresh_from_db()
        trip.refresh_from_db()
        self.assertTrue(created)
        self.assertIsNotNone(closed_shift.closed_at)
        self.assertEqual(trip.status, TripStatus.LOADED_WAITING_UNLOAD)
        self.assertIsNone(trip.completed_at)
        self.assertIsNone(trip.driver_id)
        self.assertTrue(trip.is_carryover)
        self.assertEqual(
            Trip.objects.filter(
                truck=self.truck,
                status=TripStatus.LOADED_WAITING_UNLOAD,
            ).count(),
            1,
        )

    def test_driver_http_open_retry_returns_saved_shift_result(self):
        access = EmployeeAccess.objects.get(
            employee=self.driver,
            role=self.driver_role,
        )
        session = self.client.session
        session['employee_access_id'] = access.id
        session.save()
        payload = {
            'start_fuel': '1000',
            'start_mileage': '10000',
            'start_engine_hours': '1000',
            'client_action_id': 'chaos-driver-http-open-retry',
        }

        first = self.client.post(reverse('driver_work'), data=payload)
        second = self.client.post(reverse('driver_work'), data=payload)

        self.assertEqual(first.status_code, 302)
        self.assertEqual(second.status_code, 302)
        self.assertEqual(
            EmployeeShift.objects.filter(
                employee=self.driver,
                workplace_code='driver',
                closed_at__isnull=True,
            ).count(),
            1,
        )
        self.assertEqual(
            ShiftClientAction.objects.filter(
                action_type='driver_shift_opened',
                client_action_id='chaos-driver-http-open-retry',
            ).count(),
            1,
        )

    def test_driver_open_rechecks_stale_equipment_assignment_under_lock(self):
        self.work_assignment.status = AssignmentStatus.CANCELLED
        self.work_assignment.ended_at = timezone.now()
        self.work_assignment.save(update_fields=['status', 'ended_at'])

        with self.assertRaisesMessage(
            ValidationError,
            'Назначение изменилось. Обновите экран перед началом смены.',
        ):
            open_driver_shift(
                employee=self.driver,
                work_assignment=self.work_assignment,
                readings=self.opening_readings(),
                client_action_id='chaos-open-stale-assignment',
            )

        self.assertFalse(
            EmployeeShift.objects.filter(
                employee=self.driver,
                workplace_code='driver',
                closed_at__isnull=True,
            ).exists(),
        )

    def test_replacement_driver_opens_shift_and_unloads_the_same_trip(self):
        original_shift = self.open_driver_shift(self.driver, 'chaos-open-before-handover')
        trip = Trip.objects.create(
            excavator=self.excavator,
            truck=self.truck,
            rock_type=self.rock,
            dump_point=self.dump_point,
            status=TripStatus.LOADED_WAITING_UNLOAD,
        )
        close_driver_shift(
            shift=original_shift,
            employee=self.driver,
            readings=self.closing_readings(),
            client_action_id='chaos-close-before-handover',
        )
        replacement_shift = open_driver_shift(
            employee=self.replacement,
            work_assignment=self.replacement_work_assignment,
            readings={
                'start_fuel': Decimal('900'),
                'start_mileage': Decimal('10100'),
                'start_engine_hours': Decimal('1010'),
            },
            client_action_id='chaos-open-replacement',
        )[0]
        replacement_access = EmployeeAccess.objects.get(
            employee=self.replacement,
            role=self.driver_role,
        )
        session = self.client.session
        session['employee_access_id'] = replacement_access.id
        session.save()

        response = self.client.post(
            reverse('driver_complete_trip', args=[trip.id]),
            {'client_action_id': 'chaos-unload-after-handover'},
        )

        self.assertEqual(response.status_code, 302)
        trip.refresh_from_db()
        self.assertEqual(trip.status, TripStatus.COMPLETED)
        self.assertEqual(trip.driver, self.replacement)
        self.assertEqual(trip.unloading_shift, replacement_shift)
        self.assertTrue(trip.is_carryover)
        self.assertEqual(Trip.objects.filter(pk=trip.pk).count(), 1)

    def test_p05_002_loaded_trip_survives_driver_handoff_and_reports_once(self):
        upsert_default_equipment_states()
        self.truck_model.body_volume_m3 = Decimal('38.00')
        self.truck_model.save(update_fields=['body_volume_m3'])
        self.rock.density = Decimal('2.0000')
        self.rock.save(update_fields=['density'])
        waiting_reason, _ = DowntimeReason.objects.get_or_create(
            name='Ожидание самосвалов',
            defaults={
                'equipment_type': self.excavator_type,
                'show_for_excavator_operator': True,
            },
        )
        waiting_reason.equipment_type = self.excavator_type
        waiting_reason.show_for_excavator_operator = True
        waiting_reason.is_active = True
        waiting_reason.save(
            update_fields=[
                'equipment_type',
                'show_for_excavator_operator',
                'is_active',
            ],
        )

        work_date = production_work_date()
        production_start, _ = production_day_bounds(work_date)
        original_opened_at = production_start + timedelta(hours=1)
        original_closed_at = production_start + timedelta(hours=11, minutes=55)
        replacement_opened_at = production_start + timedelta(hours=12, minutes=5)

        with patch('shifts.services.timezone.now', return_value=original_opened_at):
            original_shift = self.open_driver_shift(
                self.driver,
                'p05-002-open-original-driver',
            )

        operator_role = Role.objects.create(
            code='excavator_operator',
            name='Машинист экскаватора',
        )
        operator = Employee.objects.create(
            full_name='Машинист P05-002',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        operator_access = EmployeeAccess.objects.create(
            employee=operator,
            role=operator_role,
            access_code='P05-002-OPERATOR',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        operator_shift = EmployeeShift.objects.create(
            employee=operator,
            shift_type='day',
            workplace_code='excavator_operator',
            equipment=self.excavator,
            start_fuel=Decimal('1000'),
            start_engine_hours=Decimal('5000'),
            opened_at=original_opened_at,
            opened_by=operator,
        )
        EquipmentAssignment.objects.create(
            employee=operator,
            role=operator_role,
            equipment=self.excavator,
            shift=operator_shift,
            assigned_by=operator,
            status=AssignmentStatus.ACCEPTED,
            accepted_at=original_opened_at,
        )
        HaulAssignment.objects.create(
            truck=self.truck,
            excavator=self.excavator,
            status=AssignmentStatus.ACCEPTED,
            accepted_at=original_opened_at,
        )
        operator_client = self.client_class()
        operator_session = operator_client.session
        operator_session['employee_access_id'] = operator_access.id
        operator_session.save()

        load_response = operator_client.post(
            reverse('excavator_truck_loaded'),
            data=json.dumps({
                'client_action_id': 'p05-002-truck-loaded',
                'truck_id': self.truck.id,
                'excavator_id': self.excavator.id,
                'dump_point_id': self.dump_point.id,
                'rock_type': self.rock.id,
                'loading_horizon': '125',
                'loading_block': '4',
            }),
            content_type='application/json',
        )

        self.assertEqual(load_response.status_code, 200)
        trip = Trip.objects.get(truck=self.truck, excavator=self.excavator)
        trip_id = trip.id
        self.assertEqual(trip.status, TripStatus.LOADED_WAITING_UNLOAD)
        self.assertEqual(trip.loading_shift, operator_shift)
        self.assertIsNone(trip.unloading_shift_id)
        self.assertIsNone(trip.driver_id)

        with patch('shifts.services.timezone.now', return_value=original_closed_at):
            closed_shift, close_created = close_driver_shift(
                shift=original_shift,
                employee=self.driver,
                readings=self.closing_readings(),
                client_action_id='p05-002-close-original-driver',
            )

        closed_shift.refresh_from_db()
        trip.refresh_from_db()
        self.assertTrue(close_created)
        self.assertEqual(closed_shift.closed_at, original_closed_at)
        self.assertEqual(trip.id, trip_id)
        self.assertEqual(trip.status, TripStatus.LOADED_WAITING_UNLOAD)
        self.assertEqual(trip.loading_shift, operator_shift)
        self.assertIsNone(trip.unloading_shift_id)
        self.assertTrue(trip.is_carryover)

        with patch('shifts.services.timezone.now', return_value=replacement_opened_at):
            replacement_shift, replacement_created = open_driver_shift(
                employee=self.replacement,
                work_assignment=self.replacement_work_assignment,
                readings={
                    'start_fuel': Decimal('900'),
                    'start_mileage': Decimal('10100'),
                    'start_engine_hours': Decimal('1010'),
                },
                client_action_id='p05-002-open-replacement-driver',
            )

        self.assertTrue(replacement_created)
        self.assertEqual(replacement_shift.equipment, self.truck)
        replacement_access = EmployeeAccess.objects.get(
            employee=self.replacement,
            role=self.driver_role,
        )
        replacement_session = self.client.session
        replacement_session['employee_access_id'] = replacement_access.id
        replacement_session.save()
        unload_url = reverse('driver_complete_trip', args=[trip.id])
        unload_payload = {'client_action_id': 'p05-002-trip-unloaded'}

        first_unload = self.client.post(unload_url, unload_payload)
        repeated_unload = self.client.post(unload_url, unload_payload)

        self.assertEqual(first_unload.status_code, 302)
        self.assertEqual(repeated_unload.status_code, 302)
        trip.refresh_from_db()
        self.assertEqual(trip.id, trip_id)
        self.assertEqual(trip.status, TripStatus.COMPLETED)
        self.assertEqual(trip.loading_shift, operator_shift)
        self.assertEqual(trip.unloading_shift, replacement_shift)
        self.assertEqual(trip.driver, self.replacement)
        self.assertEqual(trip.volume_m3, Decimal('38.00'))
        self.assertEqual(trip.tonnage, Decimal('76.00'))
        self.assertTrue(trip.is_carryover)
        self.assertEqual(
            Trip.objects.filter(
                truck=self.truck,
                excavator=self.excavator,
            ).count(),
            1,
        )
        self.assertFalse(
            Trip.objects.filter(
                truck=self.truck,
                status__in=(TripStatus.ACTIVE, TripStatus.LOADED_WAITING_UNLOAD),
            ).exists(),
        )
        self.assertEqual(
            TripClientAction.objects.filter(
                action_type='truck_loaded',
                client_action_id='p05-002-truck-loaded',
                trip=trip,
            ).count(),
            1,
        )
        self.assertEqual(
            TripClientAction.objects.filter(
                action_type='trip_unloaded',
                client_action_id='p05-002-trip-unloaded',
                trip=trip,
            ).count(),
            1,
        )

        day_analytics = build_shift_analytics(work_date, 'day')
        night_analytics = build_shift_analytics(work_date, 'night')
        all_analytics = build_shift_analytics(work_date)
        self.assertEqual(day_analytics['totals']['loaded_trip_count'], 1)
        self.assertEqual(day_analytics['totals']['unloaded_trip_count'], 0)
        self.assertEqual(night_analytics['totals']['loaded_trip_count'], 0)
        self.assertEqual(night_analytics['totals']['unloaded_trip_count'], 1)
        self.assertEqual(night_analytics['totals']['carryover_count'], 1)
        self.assertEqual(all_analytics['totals']['loaded_trip_count'], 1)
        self.assertEqual(all_analytics['totals']['unloaded_trip_count'], 1)
        self.assertEqual(
            [loaded_trip.id for loaded_trip in all_analytics['loading_trips']],
            [trip.id],
        )
        self.assertEqual(
            [unloaded_trip.id for unloaded_trip in all_analytics['unloading_trips']],
            [trip.id],
        )
        self.assertEqual(
            calculate_truck_shift_progress(
                self.truck,
                reference_shift=original_shift,
            )['trip_count'],
            0,
        )
        self.assertEqual(
            calculate_truck_shift_progress(
                self.truck,
                reference_shift=replacement_shift,
            )['trip_count'],
            1,
        )

    def test_driver_shift_can_close_with_active_downtime_without_mutating_event(self):
        shift = self.open_driver_shift(self.driver, 'chaos-open-downtime')
        started_at = timezone.now() - timedelta(minutes=20)
        event = DowntimeEvent.objects.create(
            equipment=self.truck,
            employee=self.driver,
            reason=self.downtime_reason,
            started_at=started_at,
            comment='Непрерывный простой',
        )

        try:
            closed_shift, created = close_driver_shift(
                shift=shift,
                employee=self.driver,
                readings=self.closing_readings(),
                client_action_id='chaos-close-downtime',
            )
        except ValidationError as error:
            self.fail(f'Активный простой не должен блокировать пересменку: {error}')

        closed_shift.refresh_from_db()
        event.refresh_from_db()
        self.assertTrue(created)
        self.assertIsNotNone(closed_shift.closed_at)
        self.assertIsNone(event.ended_at)
        self.assertEqual(event.employee, self.driver)
        self.assertEqual(event.reason, self.downtime_reason)
        self.assertEqual(event.started_at, started_at)
        self.assertEqual(event.comment, 'Непрерывный простой')

    def test_driver_downtime_retry_after_lost_response_keeps_one_event(self):
        self.open_driver_shift(self.driver, 'chaos-open-downtime-retry')
        access = EmployeeAccess.objects.get(
            employee=self.driver,
            role=self.driver_role,
        )
        session = self.client.session
        session['employee_access_id'] = access.id
        session.save()
        payload = json.dumps({
            'action': 'start',
            'reason_id': self.downtime_reason.id,
            'client_action_id': 'chaos-driver-downtime-retry',
        })

        first = self.client.post(
            reverse('driver_downtime_action'),
            data=payload,
            content_type='application/json',
            HTTP_ACCEPT='application/json',
        )
        second = self.client.post(
            reverse('driver_downtime_action'),
            data=payload,
            content_type='application/json',
            HTTP_ACCEPT='application/json',
        )

        self.assertEqual(first.status_code, 200)
        self.assertEqual(second.status_code, 200)
        events = DowntimeEvent.objects.filter(
            equipment=self.truck,
            ended_at__isnull=True,
        )
        self.assertEqual(events.count(), 1)
        event = events.get()
        self.assertEqual(event.employee, self.driver)
        self.assertEqual(event.reason, self.downtime_reason)
        self.assertEqual(event.comment, 'Зафиксировано водителем самосвала')

    def test_driver_downtime_rechecks_active_role_inside_transaction(self):
        self.open_driver_shift(self.driver, 'chaos-open-inactive-downtime')
        access = EmployeeAccess.objects.get(
            employee=self.driver,
            role=self.driver_role,
        )
        session = self.client.session
        session['employee_access_id'] = access.id
        session.save()

        with patch(
            'users.views.role_session_state',
            return_value={'is_active': False},
        ):
            response = self.client.post(
                reverse('driver_downtime_action'),
                data=json.dumps({
                    'action': 'start',
                    'reason_id': self.downtime_reason.id,
                    'client_action_id': 'chaos-inactive-driver-downtime',
                }),
                content_type='application/json',
                HTTP_ACCEPT='application/json',
            )

        self.assertEqual(response.status_code, 409)
        self.assertEqual(response.json()['code'], 'inactive_role')
        self.assertFalse(
            DowntimeEvent.objects.filter(
                equipment=self.truck,
                ended_at__isnull=True,
            ).exists(),
        )

    def test_driver_assignment_accept_rechecks_active_role_inside_transaction(self):
        self.open_driver_shift(self.driver, 'chaos-open-inactive-assignment')
        pending = HaulAssignment.objects.create(
            truck=self.truck,
            excavator=self.excavator,
            status=AssignmentStatus.PENDING,
        )
        access = EmployeeAccess.objects.get(
            employee=self.driver,
            role=self.driver_role,
        )
        session = self.client.session
        session['employee_access_id'] = access.id
        session.save()

        with patch(
            'users.views.role_session_state',
            return_value={'is_active': False},
        ):
            response = self.client.post(
                reverse('driver_accept_assignment', args=[pending.id]),
                HTTP_X_REQUESTED_WITH='XMLHttpRequest',
            )

        self.assertEqual(response.status_code, 409)
        self.assertEqual(response.json()['code'], 'inactive_role')
        pending.refresh_from_db()
        self.assertEqual(pending.status, AssignmentStatus.PENDING)
        self.assertIsNone(pending.accepted_at)

    def test_replacement_driver_sees_transferred_downtime(self):
        _, replacement_shift, event = self.manually_create_handoff()

        response = self.client.get(reverse('driver_work'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['active_downtime'], event)
        self.assertEqual(response.context['open_shift'], replacement_shift)

    def test_replacement_driver_closes_transferred_downtime_without_rewriting_author(self):
        _, _, event = self.manually_create_handoff()
        original_started_at = event.started_at
        original_comment = event.comment

        response = self.client.post(
            reverse('driver_downtime_action'),
            data=json.dumps({'action': 'close'}),
            content_type='application/json',
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )

        self.assertEqual(response.status_code, 200)
        event.refresh_from_db()
        self.assertIsNotNone(event.ended_at)
        self.assertEqual(event.employee, self.driver)
        self.assertEqual(event.reason, self.downtime_reason)
        self.assertEqual(event.started_at, original_started_at)
        self.assertEqual(event.comment, original_comment)
        self.assertEqual(DowntimeEvent.objects.filter(equipment=self.truck).count(), 1)

    def test_cross_employee_active_downtime_does_not_create_second_event(self):
        _, _, event = self.manually_create_handoff()
        alternate_reason = DowntimeReason.objects.create(
            name='Новая причина сменщика CHAOS P01',
            equipment_type=self.truck_type,
            show_for_truck_driver=True,
        )

        response = self.client.post(
            reverse('driver_downtime_action'),
            data=json.dumps({
                'action': 'start',
                'reason_id': alternate_reason.id,
            }),
            content_type='application/json',
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )

        self.assertEqual(response.status_code, 409)
        event.refresh_from_db()
        self.assertEqual(DowntimeEvent.objects.filter(equipment=self.truck, ended_at__isnull=True).count(), 1)
        self.assertEqual(event.employee, self.driver)
        self.assertEqual(event.reason, self.downtime_reason)
        self.assertEqual(event.comment, 'Исходный простой предшественника')

    def test_transferred_downtime_duration_stays_with_shift_that_started_it(self):
        old_shift, replacement_shift, event = self.manually_create_handoff()
        ended_at = timezone.now() + timedelta(minutes=10)
        event.ended_at = ended_at
        event.save(update_fields=['ended_at'])

        expected_old_seconds = int((ended_at - event.started_at).total_seconds())

        self.assertEqual(
            driver_shift_downtime_seconds(self.truck, old_shift, until=ended_at),
            expected_old_seconds,
        )
        self.assertEqual(
            driver_shift_downtime_seconds(self.truck, replacement_shift, until=ended_at),
            0,
        )

    def test_chaos_p01_005_transferred_downtime_keeps_source_attribution_in_reports_and_excel(self):
        production_date = date(2026, 7, 23)
        production_start, _ = production_day_bounds(production_date)
        started_at = production_start + timedelta(hours=11, minutes=55)
        handoff_at = started_at + timedelta(minutes=5)
        ended_at = started_at + timedelta(minutes=10)

        source_shift = EmployeeShift.objects.create(
            employee=self.driver,
            shift_type='day',
            workplace_code='driver',
            equipment=self.truck,
            start_fuel=Decimal('1000'),
            start_mileage=Decimal('10000'),
            start_engine_hours=Decimal('1000'),
            opened_at=production_start,
            opened_by=self.driver,
        )
        event = DowntimeEvent.objects.create(
            equipment=self.truck,
            employee=self.driver,
            reason=self.downtime_reason,
            started_at=started_at,
            comment='Непрерывный простой через пересменку CHAOS P01-005',
        )

        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(
            full_name='Диспетчер отчёта CHAOS P01-005',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        dispatcher_access = EmployeeAccess.objects.create(
            employee=dispatcher,
            role=dispatcher_role,
            access_code='CHAOS-P01-005-DISPATCHER',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        EmployeeShift.objects.create(
            employee=dispatcher,
            shift_type='day',
            workplace_code='dispatcher',
            opened_at=production_start,
            opened_by=dispatcher,
        )
        session = self.client.session
        session['employee_access_id'] = dispatcher_access.id
        session.save()

        with patch('trips.views.timezone.now', return_value=handoff_at):
            service_close_response = self.client.post(
                reverse('dispatcher_service_close_shift', args=[source_shift.id]),
                {
                    'reason': 'Пересменка с непрерывным простоем CHAOS P01-005',
                    'end_fuel': '900',
                    'end_mileage': '10100',
                    'end_engine_hours': '1010',
                },
            )

        self.assertEqual(service_close_response.status_code, 302)
        source_shift.refresh_from_db()
        event.refresh_from_db()
        self.assertEqual(source_shift.closed_at, handoff_at)
        self.assertTrue(source_shift.is_service_closed)
        self.assertIsNone(event.ended_at)
        self.assertEqual(event.employee, self.driver)
        self.assertEqual(event.started_at, started_at)

        replacement_shift = EmployeeShift.objects.create(
            employee=self.replacement,
            shift_type='night',
            workplace_code='driver',
            equipment=self.truck,
            start_fuel=Decimal('900'),
            start_mileage=Decimal('10100'),
            start_engine_hours=Decimal('1010'),
            opened_at=handoff_at,
            opened_by=self.replacement,
        )
        self.replacement_work_assignment.shift = replacement_shift
        self.replacement_work_assignment.save(update_fields=['shift'])
        replacement_access = EmployeeAccess.objects.get(
            employee=self.replacement,
            role=self.driver_role,
        )
        session = self.client.session
        session['employee_access_id'] = replacement_access.id
        session.save()

        with patch('users.views.timezone.now', return_value=ended_at):
            downtime_close_response = self.client.post(
                reverse('driver_downtime_action'),
                data=json.dumps({'action': 'close'}),
                content_type='application/json',
                HTTP_X_REQUESTED_WITH='XMLHttpRequest',
            )

        self.assertEqual(downtime_close_response.status_code, 200)
        event.refresh_from_db()
        self.assertEqual(event.ended_at, ended_at)
        self.assertEqual(event.employee, self.driver)
        self.assertEqual(event.started_at, started_at)
        self.assertEqual(DowntimeEvent.objects.filter(equipment=self.truck).count(), 1)

        session = self.client.session
        session['employee_access_id'] = dispatcher_access.id
        session.save()
        dispatcher_page = self.client.get(
            reverse('dispatcher_downtimes'),
            {'date': production_date.isoformat()},
        )
        dispatcher_export = self.client.get(
            reverse('dispatcher_downtimes_export'),
            {'date': production_date.isoformat()},
        )

        self.assertEqual(dispatcher_page.status_code, 200)
        self.assertEqual(dispatcher_export.status_code, 200)
        self.assertEqual(len(dispatcher_page.context['rows']), 1)
        dispatcher_row = dispatcher_page.context['rows'][0]
        self.assertEqual(dispatcher_row['started_at'], started_at)
        self.assertEqual(dispatcher_row['ended_at'], ended_at)
        self.assertEqual(dispatcher_row['employee'], self.driver.full_name)
        self.assertEqual(dispatcher_row['duration_hours'], Decimal('0.17'))

        dispatcher_sheet_rows = list(
            load_workbook(BytesIO(dispatcher_export.content), data_only=True)
            .active
            .iter_rows(values_only=True)
        )
        dispatcher_detail_rows = [
            row
            for row in dispatcher_sheet_rows
            if len(row) >= 8
            and row[3] == str(self.truck)
            and row[7] == self.driver.full_name
        ]
        self.assertEqual(len(dispatcher_detail_rows), 1)
        self.assertEqual(dispatcher_detail_rows[0][0], '23.07.2026 18:55')
        self.assertEqual(dispatcher_detail_rows[0][1], '23.07.2026 19:05')
        self.assertEqual(dispatcher_detail_rows[0][6], '0,17')
        self.assertNotIn(
            self.replacement.full_name,
            [value for row in dispatcher_sheet_rows for value in row],
        )

        manager_role = Role.objects.create(code='manager', name='Руководитель')
        manager = Employee.objects.create(
            full_name='Руководитель отчёта CHAOS P01-005',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        manager_access = EmployeeAccess.objects.create(
            employee=manager,
            role=manager_role,
            access_code='CHAOS-P01-005-MANAGER',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        session = self.client.session
        session['employee_access_id'] = manager_access.id
        session.save()
        manager_query = {
            'date_from': production_date.isoformat(),
            'date_to': production_date.isoformat(),
        }
        manager_page = self.client.get(reverse('downtime_report'), manager_query)
        manager_export = self.client.get(reverse('downtime_report_export'), manager_query)

        self.assertEqual(manager_page.status_code, 200)
        self.assertEqual(manager_export.status_code, 200)
        self.assertEqual(manager_page.context['total_count'], 1)
        self.assertEqual(len(manager_page.context['export_rows']), 1)
        manager_row = manager_page.context['export_rows'][0]
        self.assertEqual(manager_row['started_at'], started_at)
        self.assertEqual(manager_row['ended_at'], ended_at)
        self.assertEqual(manager_row['employee'], self.driver)
        self.assertEqual(manager_row['duration_hours'], Decimal('0.17'))
        self.assertEqual(manager_page.context['daily_summary'][0]['date'], production_date)

        manager_sheet_rows = list(
            load_workbook(BytesIO(manager_export.content), data_only=True)
            .active
            .iter_rows(values_only=True)
        )
        manager_detail_rows = [
            row
            for row in manager_sheet_rows
            if len(row) >= 8
            and row[3] == str(self.truck)
            and row[7] == self.driver.full_name
        ]
        self.assertEqual(len(manager_detail_rows), 1)
        self.assertEqual(manager_detail_rows[0][0], '23.07.2026 18:55')
        self.assertEqual(manager_detail_rows[0][1], '23.07.2026 19:05')
        self.assertAlmostEqual(manager_detail_rows[0][6], 0.17, places=2)
        self.assertNotIn(
            self.replacement.full_name,
            [value for row in manager_sheet_rows for value in row],
        )

        day_analytics = build_shift_analytics(production_date, 'day')
        night_analytics = build_shift_analytics(production_date, 'night')
        self.assertEqual(day_analytics['totals']['downtime_count'], 1)
        self.assertEqual(day_analytics['totals']['downtime_hours'], Decimal('0.17'))
        self.assertEqual(
            day_analytics['downtime_reason_rows'][0]['employees_display'],
            self.driver.full_name,
        )
        self.assertEqual(night_analytics['totals']['downtime_count'], 0)
        self.assertEqual(night_analytics['totals']['downtime_hours'], Decimal('0'))


class ChaosExcavatorLoadingDowntimeRegressionTests(TestCase):
    def setUp(self):
        upsert_default_equipment_states()
        self.operator_role = Role.objects.create(
            code='excavator_operator',
            name='Машинист экскаватора',
        )
        self.driver_role = Role.objects.create(code='driver', name='Водитель')
        self.operator = Employee.objects.create(
            full_name='Машинист CHAOS P01',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        self.operator_access = EmployeeAccess.objects.create(
            employee=self.operator,
            role=self.operator_role,
            access_code='CHAOS-OPERATOR-004',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        self.truck_type = EquipmentType.objects.create(name='Самосвал')
        self.excavator_type = EquipmentType.objects.create(name='Экскаватор')
        self.excavator_model = EquipmentModel.objects.create(
            equipment_type=self.excavator_type,
            name='Экскаватор CHAOS P01',
            fuel_capacity_limit_l=Decimal('7000'),
        )
        self.truck_model = EquipmentModel.objects.create(
            equipment_type=self.truck_type,
            name='Самосвал CHAOS P01',
            fuel_capacity_limit_l=Decimal('2000'),
        )
        self.excavator = Equipment.objects.create(
            equipment_type=self.excavator_type,
            model=self.excavator_model,
            garage_number='CHAOS-EXC-004',
        )
        self.truck = Equipment.objects.create(
            equipment_type=self.truck_type,
            model=self.truck_model,
            garage_number='CHAOS-TRUCK-004',
        )
        self.rock = RockType.objects.create(name='Руда CHAOS P01 погрузка')
        self.dump_point = DumpPoint.objects.create(name='ККД CHAOS P01 погрузка')
        self.waiting_reason, _ = DowntimeReason.objects.get_or_create(
            name='Ожидание самосвалов',
            defaults={
                'equipment_type': self.excavator_type,
                'show_for_excavator_operator': True,
            },
        )
        self.waiting_reason.equipment_type = self.excavator_type
        self.waiting_reason.show_for_excavator_operator = True
        self.waiting_reason.is_active = True
        self.waiting_reason.save(
            update_fields=[
                'equipment_type',
                'show_for_excavator_operator',
                'is_active',
            ],
        )
        self.manual_reason = DowntimeReason.objects.create(
            name='Ручной простой CHAOS P01 погрузка',
            equipment_type=self.excavator_type,
            show_for_excavator_operator=True,
        )
        self.operator_shift = EmployeeShift.objects.create(
            employee=self.operator,
            shift_type='day',
            workplace_code='excavator_operator',
            equipment=self.excavator,
            start_fuel=Decimal('100'),
            start_engine_hours=Decimal('1200'),
            opened_at=timezone.now() - timedelta(hours=1),
            opened_by=self.operator,
        )
        EquipmentAssignment.objects.create(
            employee=self.operator,
            role=self.operator_role,
            equipment=self.excavator,
            shift=self.operator_shift,
            assigned_by=self.operator,
            status=AssignmentStatus.ACCEPTED,
            accepted_at=timezone.now(),
        )
        self.create_loadable_truck(self.truck, 'Водитель CHAOS 004-1', 'CHAOS-DRIVER-004-1')
        session = self.client.session
        session['employee_access_id'] = self.operator_access.id
        session.save()

    def create_loadable_truck(self, truck, full_name, access_code):
        driver = Employee.objects.create(
            full_name=full_name,
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        EmployeeAccess.objects.create(
            employee=driver,
            role=self.driver_role,
            access_code=access_code,
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        shift = EmployeeShift.objects.create(
            employee=driver,
            shift_type='day',
            workplace_code='driver',
            equipment=truck,
            opened_at=timezone.now() - timedelta(minutes=30),
            opened_by=driver,
        )
        EquipmentAssignment.objects.create(
            employee=driver,
            role=self.driver_role,
            equipment=truck,
            shift=shift,
            assigned_by=self.operator,
            status=AssignmentStatus.ACCEPTED,
            accepted_at=timezone.now(),
        )
        HaulAssignment.objects.create(
            truck=truck,
            excavator=self.excavator,
            status=AssignmentStatus.ACCEPTED,
            accepted_at=timezone.now(),
        )
        return driver, shift

    def add_second_loadable_truck(self):
        second_truck = Equipment.objects.create(
            equipment_type=self.truck_type,
            model=self.truck_model,
            garage_number='CHAOS-TRUCK-004-2',
        )
        self.create_loadable_truck(
            second_truck,
            'Водитель CHAOS 004-2',
            'CHAOS-DRIVER-004-2',
        )
        return second_truck

    def post_loaded(self, action_id):
        return self.client.post(
            reverse('excavator_truck_loaded'),
            data=json.dumps({
                'client_action_id': action_id,
                'truck_id': self.truck.id,
                'excavator_id': self.excavator.id,
                'dump_point_id': self.dump_point.id,
                'rock_type': self.rock.id,
                'loading_horizon': '125',
                'loading_block': '4',
            }),
            content_type='application/json',
        )

    def test_excavator_settings_rechecks_active_role_inside_transaction(self):
        with patch(
            'trips.views.role_session_state',
            return_value={'is_active': False},
        ):
            response = self.client.post(
                reverse('excavator_work_settings'),
                data=json.dumps({
                    'client_action_id': 'chaos-inactive-settings',
                    'rock_type_id': self.rock.id,
                    'dump_point_ids': [self.dump_point.id],
                    'loading_horizon': '125',
                    'loading_block': '4',
                }),
                content_type='application/json',
            )

        self.assertEqual(response.status_code, 409)
        self.assertEqual(response.json()['code'], 'inactive_role')
        self.assertFalse(
            ExcavatorPlacement.objects.filter(excavator=self.excavator).exists(),
        )
        self.assertFalse(
            DowntimeEvent.objects.filter(
                equipment=self.excavator,
                ended_at__isnull=True,
            ).exists(),
        )

    def test_excavator_downtime_rechecks_active_role_inside_transaction(self):
        with patch(
            'trips.views.role_session_state',
            return_value={'is_active': False},
        ):
            response = self.client.post(
                reverse('excavator_downtime_action'),
                data=json.dumps({
                    'action': 'start',
                    'reason_id': self.manual_reason.id,
                    'client_action_id': 'chaos-inactive-excavator-downtime',
                }),
                content_type='application/json',
            )

        self.assertEqual(response.status_code, 409)
        self.assertEqual(response.json()['code'], 'inactive_role')
        self.assertFalse(
            DowntimeEvent.objects.filter(
                equipment=self.excavator,
                ended_at__isnull=True,
            ).exists(),
        )

    def test_loading_closes_manual_downtime_and_starts_new_waiting_when_no_truck_remains(self):
        old_event = DowntimeEvent.objects.create(
            equipment=self.excavator,
            employee=self.operator,
            reason=self.manual_reason,
            started_at=timezone.now() - timedelta(minutes=20),
            comment='Ручной простой до фактической погрузки',
        )

        response = self.post_loaded('chaos-manual-last-truck')

        self.assertEqual(response.status_code, 200)
        old_event.refresh_from_db()
        self.assertIsNotNone(old_event.ended_at)
        active_events = DowntimeEvent.objects.filter(
            equipment=self.excavator,
            ended_at__isnull=True,
        )
        self.assertEqual(active_events.count(), 1)
        waiting = active_events.get()
        self.assertNotEqual(waiting.id, old_event.id)
        self.assertEqual(waiting.reason, self.waiting_reason)
        self.assertGreaterEqual(waiting.started_at, old_event.ended_at)
        self.assertEqual(old_event.employee, self.operator)
        self.assertEqual(old_event.comment, 'Ручной простой до фактической погрузки')

    def test_loading_closes_manual_downtime_without_waiting_when_another_truck_is_available(self):
        self.add_second_loadable_truck()
        old_event = DowntimeEvent.objects.create(
            equipment=self.excavator,
            employee=self.operator,
            reason=self.manual_reason,
            started_at=timezone.now() - timedelta(minutes=20),
            comment='Ручной простой при наличии второй машины',
        )

        response = self.post_loaded('chaos-manual-another-truck')

        self.assertEqual(response.status_code, 200)
        old_event.refresh_from_db()
        self.assertIsNotNone(old_event.ended_at)
        self.assertFalse(
            DowntimeEvent.objects.filter(
                equipment=self.excavator,
                ended_at__isnull=True,
            ).exists(),
        )

    def test_loading_restarts_existing_waiting_with_a_new_event_boundary(self):
        old_waiting = DowntimeEvent.objects.create(
            equipment=self.excavator,
            employee=self.operator,
            reason=self.waiting_reason,
            started_at=timezone.now() - timedelta(minutes=20),
            comment='Автоматически по производственному событию',
        )

        response = self.post_loaded('chaos-restart-waiting')

        self.assertEqual(response.status_code, 200)
        old_waiting.refresh_from_db()
        self.assertIsNotNone(old_waiting.ended_at)
        new_waiting = DowntimeEvent.objects.get(
            equipment=self.excavator,
            reason=self.waiting_reason,
            ended_at__isnull=True,
        )
        self.assertNotEqual(new_waiting.id, old_waiting.id)
        self.assertGreaterEqual(new_waiting.started_at, old_waiting.ended_at)

    def test_available_accepted_assignment_closes_waiting_on_reconcile(self):
        self.assertEqual(
            self.post_loaded('chaos-waiting-before-new-assignment').status_code,
            200,
        )
        waiting = DowntimeEvent.objects.get(
            equipment=self.excavator,
            reason=self.waiting_reason,
            ended_at__isnull=True,
        )
        self.add_second_loadable_truck()

        response = self.client.get(reverse('excavator_work'))

        self.assertEqual(response.status_code, 200)
        waiting.refresh_from_db()
        self.assertIsNotNone(waiting.ended_at)
        self.assertFalse(
            DowntimeEvent.objects.filter(
                equipment=self.excavator,
                ended_at__isnull=True,
            ).exists(),
        )

    def test_loading_closes_duplicate_downtimes_and_leaves_exactly_one_new_waiting(self):
        first_event = DowntimeEvent.objects.create(
            equipment=self.excavator,
            employee=self.operator,
            reason=self.manual_reason,
            started_at=timezone.now() - timedelta(minutes=30),
            comment='Первый ошибочный дубль',
        )
        second_reason = DowntimeReason.objects.create(
            name='Второй ручной простой CHAOS P01',
            equipment_type=self.excavator_type,
            show_for_excavator_operator=True,
        )
        second_event = DowntimeEvent.objects.create(
            equipment=self.excavator,
            employee=self.operator,
            reason=second_reason,
            started_at=timezone.now() - timedelta(minutes=25),
            comment='Второй ошибочный дубль',
        )

        response = self.post_loaded('chaos-close-duplicate-downtimes')

        self.assertEqual(response.status_code, 200)
        first_event.refresh_from_db()
        second_event.refresh_from_db()
        self.assertIsNotNone(first_event.ended_at)
        self.assertIsNotNone(second_event.ended_at)
        active_events = DowntimeEvent.objects.filter(
            equipment=self.excavator,
            ended_at__isnull=True,
        )
        self.assertEqual(active_events.count(), 1)
        self.assertEqual(active_events.get().reason, self.waiting_reason)

    def test_legacy_html_loading_path_uses_the_same_downtime_boundary(self):
        old_event = DowntimeEvent.objects.create(
            equipment=self.excavator,
            employee=self.operator,
            reason=self.manual_reason,
            started_at=timezone.now() - timedelta(minutes=20),
            comment='Ручной простой перед HTML-погрузкой',
        )
        assignment = HaulAssignment.objects.get(
            truck=self.truck,
            excavator=self.excavator,
            ended_at__isnull=True,
        )

        response = self.client.post(
            reverse('excavator_work'),
            data={
                'assignment': assignment.id,
                'rock_type': self.rock.id,
                'dump_point': self.dump_point.id,
                'loading_horizon': '125',
                'loading_block': '4',
                'client_action_id': 'chaos-legacy-html-boundary',
            },
        )

        self.assertEqual(response.status_code, 302)
        old_event.refresh_from_db()
        self.assertIsNotNone(old_event.ended_at)
        active_events = DowntimeEvent.objects.filter(
            equipment=self.excavator,
            ended_at__isnull=True,
        )
        self.assertEqual(active_events.count(), 1)
        self.assertEqual(active_events.get().reason, self.waiting_reason)

    def test_legacy_html_loading_retry_uses_saved_client_action(self):
        assignment = HaulAssignment.objects.get(
            truck=self.truck,
            excavator=self.excavator,
            ended_at__isnull=True,
        )
        payload = {
            'assignment': assignment.id,
            'rock_type': self.rock.id,
            'dump_point': self.dump_point.id,
            'loading_horizon': '125',
            'loading_block': '4',
            'client_action_id': 'chaos-legacy-html-load',
        }

        first = self.client.post(reverse('excavator_work'), data=payload)
        second = self.client.post(
            reverse('excavator_work'),
            data={
                **payload,
                'assignment': 999999,
                'rock_type': 999999,
                'dump_point': 999999,
            },
        )

        self.assertEqual(first.status_code, 302)
        self.assertEqual(second.status_code, 302)
        self.assertEqual(
            Trip.objects.filter(
                truck=self.truck,
                excavator=self.excavator,
            ).count(),
            1,
        )
        self.assertEqual(
            TripClientAction.objects.filter(
                action_type='truck_loaded',
                client_action_id='chaos-legacy-html-load',
            ).count(),
            1,
        )

    def test_replacement_excavator_operator_can_only_close_transferred_downtime(self):
        event = DowntimeEvent.objects.create(
            equipment=self.excavator,
            employee=self.operator,
            reason=self.manual_reason,
            started_at=timezone.now() - timedelta(minutes=30),
            comment='Простой исходного машиниста',
        )
        close_payload = close_excavator_shift(
            employee=self.operator,
            fuel_value='90',
            engine_hours_value='1205',
            client_action_id='chaos-operator-handover-close',
        )
        self.assertTrue(close_payload['ok'])
        self.assertFalse(close_payload['shift_open'])
        replacement = Employee.objects.create(
            full_name='Машинист сменщик CHAOS P01',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        replacement_access = EmployeeAccess.objects.create(
            employee=replacement,
            role=self.operator_role,
            access_code='CHAOS-OPERATOR-REPLACEMENT',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        EmployeeShift.objects.create(
            employee=replacement,
            shift_type='day',
            workplace_code='excavator_operator',
            equipment=self.excavator,
            start_fuel=Decimal('90'),
            start_engine_hours=Decimal('1205'),
            opened_at=timezone.now() - timedelta(minutes=4),
            opened_by=replacement,
        )
        session = self.client.session
        session['employee_access_id'] = replacement_access.id
        session.save()

        update_response = self.client.post(
            reverse('excavator_downtime_action'),
            data=json.dumps({
                'action': 'start',
                'reason_id': self.waiting_reason.id,
                'comment': 'Попытка сменить автора',
            }),
            content_type='application/json',
        )
        event.refresh_from_db()
        self.assertEqual(update_response.status_code, 409)
        self.assertEqual(event.employee, self.operator)
        self.assertEqual(event.reason, self.manual_reason)
        self.assertEqual(event.comment, 'Простой исходного машиниста')
        self.assertIsNone(event.ended_at)

        close_response = self.client.post(
            reverse('excavator_downtime_action'),
            data=json.dumps({'action': 'close'}),
            content_type='application/json',
        )
        event.refresh_from_db()
        self.assertEqual(close_response.status_code, 200)
        self.assertIsNotNone(event.ended_at)
        self.assertEqual(event.employee, self.operator)
        self.assertEqual(event.reason, self.manual_reason)
        self.assertEqual(event.comment, 'Простой исходного машиниста')

    def test_excavator_transferred_downtime_duration_stays_with_source_shift(self):
        started_at = timezone.now() - timedelta(minutes=30)
        event = DowntimeEvent.objects.create(
            equipment=self.excavator,
            employee=self.operator,
            reason=self.manual_reason,
            started_at=started_at,
            comment='Непрерывный простой через пересменку',
        )
        close_excavator_shift(
            employee=self.operator,
            fuel_value='90',
            engine_hours_value='1205',
            client_action_id='chaos-operator-duration-close',
        )
        self.operator_shift.refresh_from_db()
        replacement = Employee.objects.create(
            full_name='Машинист сменщик для длительности',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        replacement_opened_at = self.operator_shift.closed_at + timedelta(seconds=1)
        replacement_shift = EmployeeShift.objects.create(
            employee=replacement,
            shift_type='day',
            workplace_code='excavator_operator',
            equipment=self.excavator,
            start_fuel=Decimal('90'),
            start_engine_hours=Decimal('1205'),
            opened_at=replacement_opened_at,
            opened_by=replacement,
        )
        ended_at = replacement_opened_at + timedelta(minutes=10)
        event.ended_at = ended_at
        event.save(update_fields=['ended_at'])
        expected_seconds = int((ended_at - started_at).total_seconds())

        self.assertEqual(
            equipment_shift_downtime_seconds(
                self.excavator,
                self.operator_shift,
                until=ended_at,
            ),
            expected_seconds,
        )
        self.assertEqual(
            equipment_shift_downtime_seconds(
                self.excavator,
                replacement_shift,
                until=ended_at,
            ),
            0,
        )

    def test_qa_reg_p3_007_excavator_downtime_handoff_keeps_source_attribution_in_reports_and_excel(self):
        production_date = date(2026, 7, 23)
        production_start, _ = production_day_bounds(production_date)
        started_at = production_start + timedelta(hours=11, minutes=55)
        handoff_at = started_at + timedelta(minutes=5)
        ended_at = started_at + timedelta(minutes=10)

        self.operator_shift.opened_at = production_start
        self.operator_shift.save(update_fields=['opened_at'])
        event = DowntimeEvent.objects.create(
            equipment=self.excavator,
            employee=self.operator,
            reason=self.manual_reason,
            started_at=started_at,
            comment='Непрерывный простой машиниста через пересменку QA-REG-P3-007',
        )

        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(
            full_name='Диспетчер отчёта QA-REG-P3-007',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        dispatcher_access = EmployeeAccess.objects.create(
            employee=dispatcher,
            role=dispatcher_role,
            access_code='QA-REG-P3-007-DISPATCHER',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        EmployeeShift.objects.create(
            employee=dispatcher,
            shift_type='day',
            workplace_code='dispatcher',
            opened_at=production_start,
            opened_by=dispatcher,
        )
        session = self.client.session
        session['employee_access_id'] = dispatcher_access.id
        session.save()

        with patch('trips.views.timezone.now', return_value=handoff_at):
            service_close_response = self.client.post(
                reverse('dispatcher_service_close_shift', args=[self.operator_shift.id]),
                {
                    'reason': 'Пересменка с непрерывным простоем QA-REG-P3-007',
                    'end_fuel': '90',
                    'end_engine_hours': '1205',
                },
            )

        self.assertEqual(service_close_response.status_code, 302)
        self.operator_shift.refresh_from_db()
        event.refresh_from_db()
        self.assertEqual(self.operator_shift.closed_at, handoff_at)
        self.assertTrue(self.operator_shift.is_service_closed)
        self.assertIsNone(event.ended_at)
        self.assertEqual(event.employee, self.operator)
        self.assertEqual(event.started_at, started_at)

        replacement = Employee.objects.create(
            full_name='Машинист сменщик QA-REG-P3-007',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        replacement_access = EmployeeAccess.objects.create(
            employee=replacement,
            role=self.operator_role,
            access_code='QA-REG-P3-007-OPERATOR',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        replacement_assignment = EquipmentAssignment.objects.create(
            employee=replacement,
            role=self.operator_role,
            equipment=self.excavator,
            shift_type='night',
            assigned_by=dispatcher,
            status=AssignmentStatus.ACCEPTED,
            accepted_at=handoff_at,
        )
        with patch('shifts.services.timezone.now', return_value=handoff_at):
            replacement_shift_payload = open_excavator_shift(
                employee=replacement,
                equipment=self.excavator,
                shift_type='night',
                fuel_value='90',
                engine_hours_value='1205',
                client_action_id='qa-reg-p3-007-replacement-open',
            )
        self.assertTrue(replacement_shift_payload['ok'])
        self.assertTrue(replacement_shift_payload['shift_open'])
        replacement_shift = EmployeeShift.objects.get(
            pk=replacement_shift_payload['shift_id'],
        )
        replacement_assignment.shift = replacement_shift
        replacement_assignment.save(update_fields=['shift'])

        session = self.client.session
        session['employee_access_id'] = replacement_access.id
        session.save()
        replacement_page = self.client.get(reverse('excavator_work'))
        self.assertEqual(replacement_page.status_code, 200)
        self.assertEqual(replacement_page.context['active_downtime'], event)
        self.assertEqual(replacement_page.context['open_shift'], replacement_shift)

        with patch('trips.views.timezone.now', return_value=ended_at):
            downtime_close_response = self.client.post(
                reverse('excavator_downtime_action'),
                data=json.dumps({'action': 'close'}),
                content_type='application/json',
            )

        self.assertEqual(downtime_close_response.status_code, 200)
        event.refresh_from_db()
        self.assertEqual(event.ended_at, ended_at)
        self.assertEqual(event.employee, self.operator)
        self.assertEqual(event.started_at, started_at)
        self.assertEqual(event.ended_at - event.started_at, timedelta(minutes=10))
        self.assertEqual(
            DowntimeEvent.objects.filter(equipment=self.excavator).count(),
            1,
        )

        session = self.client.session
        session['employee_access_id'] = dispatcher_access.id
        session.save()
        dispatcher_page = self.client.get(
            reverse('dispatcher_downtimes'),
            {'date': production_date.isoformat()},
        )
        dispatcher_export = self.client.get(
            reverse('dispatcher_downtimes_export'),
            {'date': production_date.isoformat()},
        )

        self.assertEqual(dispatcher_page.status_code, 200)
        self.assertEqual(dispatcher_export.status_code, 200)
        self.assertEqual(len(dispatcher_page.context['rows']), 1)
        dispatcher_row = dispatcher_page.context['rows'][0]
        self.assertEqual(dispatcher_row['started_at'], started_at)
        self.assertEqual(dispatcher_row['ended_at'], ended_at)
        self.assertEqual(dispatcher_row['employee'], self.operator.full_name)
        self.assertEqual(dispatcher_row['duration_hours'], Decimal('0.17'))

        dispatcher_sheet_rows = list(
            load_workbook(BytesIO(dispatcher_export.content), data_only=True)
            .active
            .iter_rows(values_only=True)
        )
        dispatcher_detail_rows = [
            row
            for row in dispatcher_sheet_rows
            if len(row) >= 8
            and row[3] == str(self.excavator)
            and row[7] == self.operator.full_name
        ]
        self.assertEqual(len(dispatcher_detail_rows), 1)
        self.assertEqual(dispatcher_detail_rows[0][0], '23.07.2026 18:55')
        self.assertEqual(dispatcher_detail_rows[0][1], '23.07.2026 19:05')
        self.assertEqual(dispatcher_detail_rows[0][6], '0,17')
        self.assertNotIn(
            replacement.full_name,
            [value for row in dispatcher_sheet_rows for value in row],
        )

        manager_role = Role.objects.create(code='manager', name='Руководитель')
        manager = Employee.objects.create(
            full_name='Руководитель отчёта QA-REG-P3-007',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        manager_access = EmployeeAccess.objects.create(
            employee=manager,
            role=manager_role,
            access_code='QA-REG-P3-007-MANAGER',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        session = self.client.session
        session['employee_access_id'] = manager_access.id
        session.save()
        manager_query = {
            'date_from': production_date.isoformat(),
            'date_to': production_date.isoformat(),
        }
        manager_page = self.client.get(reverse('downtime_report'), manager_query)
        manager_export = self.client.get(reverse('downtime_report_export'), manager_query)

        self.assertEqual(manager_page.status_code, 200)
        self.assertEqual(manager_export.status_code, 200)
        self.assertEqual(manager_page.context['total_count'], 1)
        self.assertEqual(len(manager_page.context['export_rows']), 1)
        manager_row = manager_page.context['export_rows'][0]
        self.assertEqual(manager_row['started_at'], started_at)
        self.assertEqual(manager_row['ended_at'], ended_at)
        self.assertEqual(manager_row['employee'], self.operator)
        self.assertEqual(manager_row['duration_hours'], Decimal('0.17'))
        self.assertEqual(manager_page.context['daily_summary'][0]['date'], production_date)

        manager_sheet_rows = list(
            load_workbook(BytesIO(manager_export.content), data_only=True)
            .active
            .iter_rows(values_only=True)
        )
        manager_detail_rows = [
            row
            for row in manager_sheet_rows
            if len(row) >= 8
            and row[3] == str(self.excavator)
            and row[7] == self.operator.full_name
        ]
        self.assertEqual(len(manager_detail_rows), 1)
        self.assertEqual(manager_detail_rows[0][0], '23.07.2026 18:55')
        self.assertEqual(manager_detail_rows[0][1], '23.07.2026 19:05')
        self.assertAlmostEqual(manager_detail_rows[0][6], 0.17, places=2)
        self.assertNotIn(
            replacement.full_name,
            [value for row in manager_sheet_rows for value in row],
        )
        self.assertEqual(
            DowntimeEvent.objects.filter(equipment=self.excavator).count(),
            1,
        )


class ChaosServiceShiftHandoffRegressionTests(TestCase):
    def test_dispatcher_service_close_preserves_readings_trip_downtime_and_audit(self):
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        driver_role = Role.objects.create(code='driver', name='Водитель')
        dispatcher = Employee.objects.create(
            full_name='Диспетчер CHAOS P01',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        dispatcher_access = EmployeeAccess.objects.create(
            employee=dispatcher,
            role=dispatcher_role,
            access_code='CHAOS-DISPATCHER-002',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        EmployeeShift.objects.create(
            employee=dispatcher,
            shift_type='day',
            workplace_code='dispatcher',
            opened_at=timezone.now() - timedelta(hours=1),
            opened_by=dispatcher,
        )
        driver = Employee.objects.create(
            full_name='Водитель служебной пересменки',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        EmployeeAccess.objects.create(
            employee=driver,
            role=driver_role,
            access_code='CHAOS-SERVICE-DRIVER-002',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck_model = EquipmentModel.objects.create(
            equipment_type=truck_type,
            name='Самосвал служебной пересменки',
            fuel_capacity_limit_l=Decimal('2000'),
        )
        truck = Equipment.objects.create(
            equipment_type=truck_type,
            model=truck_model,
            garage_number='CHAOS-SERVICE-TRUCK-002',
        )
        excavator = Equipment.objects.create(
            equipment_type=excavator_type,
            garage_number='CHAOS-SERVICE-EXC-002',
        )
        shift = EmployeeShift.objects.create(
            employee=driver,
            shift_type='day',
            workplace_code='driver',
            equipment=truck,
            start_fuel=Decimal('1000'),
            start_mileage=Decimal('10000'),
            start_engine_hours=Decimal('1000'),
            opened_at=timezone.now() - timedelta(hours=2),
            opened_by=driver,
        )
        rock = RockType.objects.create(name='Руда служебной пересменки')
        dump_point = DumpPoint.objects.create(name='ККД служебной пересменки')
        trip = Trip.objects.create(
            excavator=excavator,
            truck=truck,
            rock_type=rock,
            dump_point=dump_point,
            status=TripStatus.LOADED_WAITING_UNLOAD,
        )
        reason = DowntimeReason.objects.create(
            name='Простой служебной пересменки',
            equipment_type=truck_type,
        )
        event = DowntimeEvent.objects.create(
            equipment=truck,
            employee=driver,
            reason=reason,
            started_at=timezone.now() - timedelta(minutes=20),
            comment='Непрерывный простой',
        )
        session = self.client.session
        session['employee_access_id'] = dispatcher_access.id
        session.save()

        response = self.client.post(
            reverse('dispatcher_service_close_shift', args=[shift.id]),
            {
                'reason': 'Сотрудник недоступен, показания переданы диспетчеру',
                'end_fuel': '900',
                'end_mileage': '10100',
                'end_engine_hours': '1010',
            },
        )

        self.assertEqual(response.status_code, 302)
        shift.refresh_from_db()
        trip.refresh_from_db()
        event.refresh_from_db()
        self.assertEqual(shift.end_fuel, Decimal('900'))
        self.assertEqual(shift.end_mileage, Decimal('10100'))
        self.assertEqual(shift.end_engine_hours, Decimal('1010'))
        self.assertTrue(shift.is_service_closed)
        self.assertTrue(trip.is_carryover)
        self.assertEqual(trip.status, TripStatus.LOADED_WAITING_UNLOAD)
        self.assertIsNone(event.ended_at)
        self.assertEqual(event.employee, driver)
        self.assertTrue(
            DispatcherActionLog.objects.filter(
                action_type=DispatcherActionType.SERVICE_CLOSE_SHIFT,
                shift=shift,
                reason='Сотрудник недоступен, показания переданы диспетчеру',
            ).exists(),
        )

    def test_dispatcher_service_close_preserves_excavator_downtime_for_replacement(self):
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        operator_role = Role.objects.create(
            code='excavator_operator',
            name='Машинист экскаватора',
        )
        dispatcher = Employee.objects.create(
            full_name='Диспетчер служебной передачи экскаватора',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        dispatcher_access = EmployeeAccess.objects.create(
            employee=dispatcher,
            role=dispatcher_role,
            access_code='CHAOS-SERVICE-DISPATCHER-EXC',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        EmployeeShift.objects.create(
            employee=dispatcher,
            shift_type='day',
            workplace_code='dispatcher',
            opened_at=timezone.now() - timedelta(hours=1),
            opened_by=dispatcher,
        )
        operator = Employee.objects.create(
            full_name='Машинист служебной передачи',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        EmployeeAccess.objects.create(
            employee=operator,
            role=operator_role,
            access_code='CHAOS-SERVICE-OPERATOR',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        excavator_model = EquipmentModel.objects.create(
            equipment_type=excavator_type,
            name='Экскаватор служебной передачи',
            fuel_capacity_limit_l=Decimal('7000'),
        )
        excavator = Equipment.objects.create(
            equipment_type=excavator_type,
            model=excavator_model,
            garage_number='CHAOS-SERVICE-EXC-HANDOFF',
        )
        shift = EmployeeShift.objects.create(
            employee=operator,
            shift_type='day',
            workplace_code='excavator_operator',
            equipment=excavator,
            start_fuel=Decimal('1000'),
            start_engine_hours=Decimal('5000'),
            opened_at=timezone.now() - timedelta(hours=2),
            opened_by=operator,
        )
        reason = DowntimeReason.objects.create(
            name='Простой экскаватора служебной передачи',
            equipment_type=excavator_type,
            show_for_excavator_operator=True,
        )
        event = DowntimeEvent.objects.create(
            equipment=excavator,
            employee=operator,
            reason=reason,
            started_at=timezone.now() - timedelta(minutes=20),
            comment='Исходный простой машиниста',
        )
        session = self.client.session
        session['employee_access_id'] = dispatcher_access.id
        session.save()

        response = self.client.post(
            reverse('dispatcher_service_close_shift', args=[shift.id]),
            {
                'reason': 'Машинист недоступен, показания переданы диспетчеру',
                'end_fuel': '900',
                'end_engine_hours': '5008',
            },
        )

        self.assertEqual(response.status_code, 302)
        shift.refresh_from_db()
        event.refresh_from_db()
        self.assertEqual(shift.end_fuel, Decimal('900'))
        self.assertEqual(shift.end_engine_hours, Decimal('5008'))
        self.assertTrue(shift.is_service_closed)
        self.assertIsNone(event.ended_at)
        self.assertEqual(event.employee, operator)
        self.assertEqual(event.reason, reason)
        self.assertEqual(event.comment, 'Исходный простой машиниста')
        self.assertTrue(
            DispatcherActionLog.objects.filter(
                action_type=DispatcherActionType.SERVICE_CLOSE_SHIFT,
                shift=shift,
                reason='Машинист недоступен, показания переданы диспетчеру',
            ).exists(),
        )

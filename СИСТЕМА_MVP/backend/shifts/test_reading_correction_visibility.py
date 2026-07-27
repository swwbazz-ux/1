from datetime import timedelta
from decimal import Decimal
from io import BytesIO

from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from core.production_time import production_work_date
from references.models import Equipment, EquipmentType
from users.models import Employee, EmployeeAccess, Role

from .models import EmployeeShift, ShiftReadingCorrection


class ShiftReadingCorrectionVisibilityTests(TestCase):
    def setUp(self):
        equipment_type = EquipmentType.objects.create(name='Самосвал')
        self.equipment = Equipment.objects.create(
            equipment_type=equipment_type,
            garage_number='QA-R1-77',
        )
        self.previous_employee = Employee.objects.create(
            full_name='Предыдущий водитель QA-R1',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        self.current_employee = Employee.objects.create(
            full_name='Новый водитель QA-R1',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        now = timezone.now()
        self.previous_shift = EmployeeShift.objects.create(
            employee=self.previous_employee,
            shift_type='day',
            workplace_code='driver',
            equipment=self.equipment,
            start_fuel=Decimal('500.00'),
            start_mileage=Decimal('9000.00'),
            start_engine_hours=Decimal('700.00'),
            end_fuel=Decimal('432.10'),
            end_mileage=Decimal('9100.00'),
            end_engine_hours=Decimal('710.00'),
            opened_at=now - timedelta(hours=16),
            closed_at=now - timedelta(hours=4),
            closed_by=self.previous_employee,
        )
        self.new_shift = EmployeeShift.objects.create(
            employee=self.current_employee,
            shift_type='night',
            workplace_code='driver',
            equipment=self.equipment,
            start_fuel=Decimal('876.54'),
            start_mileage=Decimal('9100.00'),
            start_engine_hours=Decimal('710.00'),
            opened_at=now - timedelta(hours=3),
            opened_by=self.current_employee,
        )
        self.correction = ShiftReadingCorrection.objects.create(
            equipment=self.equipment,
            previous_shift=self.previous_shift,
            new_shift=self.new_shift,
            metric=ShiftReadingCorrection.Metric.FUEL,
            transferred_value=Decimal('432.10'),
            actual_value=Decimal('876.54'),
            employee=self.current_employee,
        )
        self.work_date = production_work_date(self.correction.corrected_at)

        self.dispatcher_access = self.create_access(
            role_code='dispatcher',
            role_name='Диспетчер',
            employee_name='Диспетчер QA-R1',
            access_code='502201',
        )
        self.admin_access = self.create_access(
            role_code='admin',
            role_name='Системный администратор',
            employee_name='Администратор QA-R1',
            access_code='102201',
        )
        self.manager_access = self.create_access(
            role_code='manager',
            role_name='Руководитель',
            employee_name='Руководитель QA-R1',
            access_code='602201',
        )

    @staticmethod
    def create_access(*, role_code, role_name, employee_name, access_code):
        role = Role.objects.create(code=role_code, name=role_name)
        employee = Employee.objects.create(
            full_name=employee_name,
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        return EmployeeAccess.objects.create(
            employee=employee,
            role=role,
            access_code=access_code,
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )

    def authenticate(self, access):
        session = self.client.session
        session['employee_access_id'] = access.pk
        session.save()

    def assert_correction_visible(self, response, *, template_name):
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, template_name)
        self.assertTemplateUsed(response, 'includes/shift_reading_corrections.html')
        self.assertEqual(
            [item.pk for item in response.context['shift_reading_corrections']],
            [self.correction.pk],
        )
        self.assertContains(
            response,
            f'data-shift-reading-correction-id="{self.correction.pk}"',
        )
        for expected in (
            'Расхождения начальных показаний',
            str(self.equipment),
            f'№{self.previous_shift.pk}',
            f'№{self.new_shift.pk}',
            self.previous_shift.get_shift_type_display(),
            self.new_shift.get_shift_type_display(),
            self.correction.get_metric_display(),
            '432,10',
            '876,54',
            self.current_employee.full_name,
            timezone.localtime(self.correction.corrected_at).strftime('%d.%m.%Y %H:%M'),
        ):
            self.assertContains(response, expected)

    def closed_shift_snapshot(self):
        self.previous_shift.refresh_from_db()
        return {
            'end_fuel': self.previous_shift.end_fuel,
            'end_mileage': self.previous_shift.end_mileage,
            'end_engine_hours': self.previous_shift.end_engine_hours,
            'closed_at': self.previous_shift.closed_at,
            'closed_by_id': self.previous_shift.closed_by_id,
        }

    def assert_correction_exported(self, response):
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        values = [
            cell
            for sheet in workbook.worksheets
            for row in sheet.iter_rows(values_only=True)
            for cell in row
            if cell not in {None, ''}
        ]
        text_values = [str(value) for value in values]

        for header in (
            'Зафиксировано',
            'Техника',
            'Предыдущая смена',
            'Новая смена',
            'Показатель',
            'Переданное значение',
            'Фактическое значение',
            'Сотрудник',
        ):
            self.assertIn(header, text_values)
        for expected in (
            str(self.equipment),
            f'№{self.previous_shift.pk}',
            f'№{self.new_shift.pk}',
            self.previous_shift.get_shift_type_display(),
            self.new_shift.get_shift_type_display(),
            self.correction.get_metric_display(),
            self.current_employee.full_name,
            timezone.localtime(self.correction.corrected_at).strftime('%d.%m.%Y %H:%M'),
        ):
            self.assertTrue(
                any(expected in value for value in text_values),
                f'В Excel отсутствует значение коррекции: {expected}',
            )
        numeric_values = {
            Decimal(str(value))
            for value in values
            if isinstance(value, (int, float, Decimal))
        }
        self.assertIn(self.correction.transferred_value, numeric_values)
        self.assertIn(self.correction.actual_value, numeric_values)

    def test_dispatcher_shift_log_excel_exports_reading_correction_without_rewriting_shift(self):
        self.authenticate(self.dispatcher_access)
        snapshot_before = self.closed_shift_snapshot()

        response = self.client.get(
            reverse('dispatcher_shift_log_export'),
            {'date': self.work_date.isoformat()},
        )

        self.assert_correction_exported(response)
        self.assertEqual(self.closed_shift_snapshot(), snapshot_before)
        self.assertTrue(
            ShiftReadingCorrection.objects.filter(pk=self.correction.pk).exists()
        )

    def test_management_excel_exports_reading_correction_without_rewriting_shift(self):
        self.authenticate(self.manager_access)
        snapshot_before = self.closed_shift_snapshot()

        response = self.client.get(
            reverse('management_dashboard_export'),
            {'date': self.work_date.isoformat()},
        )

        self.assert_correction_exported(response)
        self.assertEqual(self.closed_shift_snapshot(), snapshot_before)
        self.assertTrue(
            ShiftReadingCorrection.objects.filter(pk=self.correction.pk).exists()
        )

    def test_dispatcher_shift_log_shows_reading_correction(self):
        self.authenticate(self.dispatcher_access)

        response = self.client.get(
            reverse('dispatcher_shift_log'),
            {'date': self.work_date.isoformat()},
        )

        self.assert_correction_visible(
            response,
            template_name='reports/dispatcher_shift_log.html',
        )

    def test_system_admin_dashboard_shows_reading_correction(self):
        self.authenticate(self.admin_access)

        response = self.client.get(reverse('system_admin_dashboard'))

        self.assert_correction_visible(
            response,
            template_name='users/system_admin_dashboard.html',
        )

    def test_management_dashboard_shows_reading_correction(self):
        self.authenticate(self.manager_access)

        response = self.client.get(
            reverse('management_dashboard'),
            {'date': self.work_date.isoformat()},
        )

        self.assert_correction_visible(
            response,
            template_name='reports/management_dashboard.html',
        )

    def test_dispatcher_shift_log_does_not_hide_corrections_after_twentieth_row(self):
        extra_corrections = []
        now = timezone.now()
        for pair_index in range(8):
            previous_shift = EmployeeShift.objects.create(
                employee=self.previous_employee,
                shift_type='day',
                workplace_code='driver',
                equipment=self.equipment,
                opened_at=now - timedelta(days=2, hours=pair_index + 2),
                closed_at=now - timedelta(days=2, hours=pair_index + 1),
            )
            new_shift = EmployeeShift.objects.create(
                employee=self.current_employee,
                shift_type='night',
                workplace_code='driver',
                equipment=self.equipment,
                opened_at=now - timedelta(days=2, minutes=pair_index + 2),
                closed_at=now - timedelta(days=2, minutes=pair_index + 1),
            )
            for metric_index, metric in enumerate(
                (
                    ShiftReadingCorrection.Metric.FUEL,
                    ShiftReadingCorrection.Metric.MILEAGE,
                    ShiftReadingCorrection.Metric.ENGINE_HOURS,
                )
            ):
                extra_corrections.append(
                    ShiftReadingCorrection.objects.create(
                        equipment=self.equipment,
                        previous_shift=previous_shift,
                        new_shift=new_shift,
                        metric=metric,
                        transferred_value=Decimal('432.10'),
                        actual_value=Decimal(
                            f'{877 + pair_index}.{50 + metric_index}'
                        ),
                        employee=self.current_employee,
                    )
                )
        self.authenticate(self.dispatcher_access)

        response = self.client.get(
            reverse('dispatcher_shift_log'),
            {'date': self.work_date.isoformat()},
        )

        visible_ids = [
            item.pk for item in response.context['shift_reading_corrections']
        ]
        self.assertEqual(len(visible_ids), 25)
        self.assertIn(self.correction.pk, visible_ids)
        self.assertIn(extra_corrections[-1].pk, visible_ids)
        self.assertContains(
            response,
            f'data-shift-reading-correction-id="{self.correction.pk}"',
        )

    def test_read_only_views_do_not_rewrite_closed_previous_shift(self):
        closed_snapshot = {
            'end_fuel': self.previous_shift.end_fuel,
            'end_mileage': self.previous_shift.end_mileage,
            'end_engine_hours': self.previous_shift.end_engine_hours,
            'closed_at': self.previous_shift.closed_at,
            'closed_by_id': self.previous_shift.closed_by_id,
        }

        for access, url, query in (
            (
                self.dispatcher_access,
                reverse('dispatcher_shift_log'),
                {'date': self.work_date.isoformat()},
            ),
            (self.admin_access, reverse('system_admin_dashboard'), {}),
            (
                self.manager_access,
                reverse('management_dashboard'),
                {'date': self.work_date.isoformat()},
            ),
        ):
            with self.subTest(role=access.role.code):
                self.authenticate(access)
                response = self.client.get(url, query)
                self.assertEqual(response.status_code, 200)

        self.previous_shift.refresh_from_db()
        self.assertEqual(
            {
                'end_fuel': self.previous_shift.end_fuel,
                'end_mileage': self.previous_shift.end_mileage,
                'end_engine_hours': self.previous_shift.end_engine_hours,
                'closed_at': self.previous_shift.closed_at,
                'closed_by_id': self.previous_shift.closed_by_id,
            },
            closed_snapshot,
        )

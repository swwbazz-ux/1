import json
from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO
from tempfile import TemporaryDirectory
from unittest.mock import patch
from zoneinfo import ZoneInfo

from django.core.files.base import ContentFile
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from PIL import Image

from assignments.models import AssignmentStatus, EquipmentAssignment, ExcavatorPlacement, HaulAssignment, WorkShiftType
from core.models import OperationalStateEvent
from core.production_time import production_work_date
from downtimes.models import DowntimeEvent, DowntimeReason
from references.models import (
    Dormitory,
    DormitoryBlock,
    DormitorySection,
    DumpPoint,
    Equipment,
    EquipmentModel,
    EquipmentType,
    RockType,
    TruckCapacityRule,
)
from reports.models import PilotFeedback, ReportTemplate, ReportType
from shifts.models import AchievementPrize, EmployeeShift, EquipmentPlanGroup, EquipmentShiftPlan, PlanAssignmentStatus, PlanCalculationMode, ShiftPlan
from trips.models import DispatcherActionLog, DispatcherActionType, Trip, TripClientAction, TripStatus

from .forms import AdminEmployeeEditForm
from .models import AdminActionLog, AdminConflict, DriverPrimaryRegistration, Employee, EmployeeAccess, Role


class AccessLoginTests(TestCase):
    def setUp(self):
        self.role = Role.objects.create(code='driver', name='Водитель самосвала')
        self.employee = Employee.objects.create(full_name='Тестовый водитель')
        self.access = EmployeeAccess.objects.create(
            employee=self.employee,
            role=self.role,
            access_code='2000',
        )

    def assign_driver_work(self, truck, *, shift_type='day'):
        if not truck.model_id:
            model, _ = EquipmentModel.objects.get_or_create(
                equipment_type=truck.equipment_type,
                name='БелАЗ тестовый',
                defaults={'fuel_capacity_limit_l': Decimal('2000')},
            )
            if model.fuel_capacity_limit_l is None:
                model.fuel_capacity_limit_l = Decimal('2000')
                model.save(update_fields=['fuel_capacity_limit_l'])
            truck.model = model
            truck.save(update_fields=['model'])
        group, _ = EquipmentPlanGroup.objects.get_or_create(
            code='driver-work-assignment-tests',
            defaults={
                'name': 'Самосвалы для тестов назначений',
                'calculation_mode': PlanCalculationMode.TRIPS,
                'plan_value': '18.00',
                'is_active': True,
            },
        )
        group.equipment.add(truck)
        return EquipmentAssignment.objects.create(
            employee=self.employee,
            role=self.role,
            equipment=truck,
            shift_type=shift_type,
            assigned_by=self.employee,
            status=AssignmentStatus.ACCEPTED,
            accepted_at=timezone.now(),
        )

    def create_registered_driver_shift(self, truck=None):
        self.employee.status = Employee.Status.ACTIVE
        self.employee.is_active = True
        self.employee.save(update_fields=['status', 'is_active'])
        self.access.status = EmployeeAccess.Status.ACTIVATED
        self.access.is_active = True
        self.access.save(update_fields=['status', 'is_active'])
        truck_type = truck.equipment_type if truck else EquipmentType.objects.create(name='Самосвал')
        truck = truck or Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        dormitory = Dormitory.objects.create(number='5')
        block = DormitoryBlock.objects.create(dormitory=dormitory, name='Блок 1')
        section = DormitorySection.objects.create(block=block, name='А')
        DriverPrimaryRegistration.objects.create(
            employee=self.employee,
            dormitory_section=section,
        )
        EmployeeShift.objects.create(
            employee=self.employee,
            shift_type='day',
            equipment=truck,
            start_fuel='1000.00',
            start_mileage='12560.00',
            start_engine_hours='1354.00',
            opened_at=timezone.now(),
            opened_by=self.employee,
        )
        session = self.client.session
        session['employee_access_id'] = self.access.id
        session.save()
        return truck

    def test_registered_driver_opens_shift_screen(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        dormitory = Dormitory.objects.create(number='5')
        block = DormitoryBlock.objects.create(dormitory=dormitory, name='Блок 1')
        section = DormitorySection.objects.create(block=block, name='А')
        DriverPrimaryRegistration.objects.create(
            employee=self.employee,
            dormitory_section=section,
        )
        self.assign_driver_work(truck)

        response = self.client.post('/', {'access_code': '2000'}, follow=True, HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Работа водителя')
        self.assertContains(response, '>Начать смену<')
        self.assertContains(response, 'data-mobile-shift-role="driver"')
        self.assertContains(response, 'data-mobile-shift-state="closed"')
        self.assertContains(response, 'Показатели техники')
        self.assertContains(response, 'Итог смены')
        self.assertContains(response, 'Смена закрыта')
        self.assertContains(response, '0 шт.')
        self.assertContains(response, 'mobile-shift__actions')
        self.assertContains(response, 'aria-label="Удерживайте 1 секунду, чтобы начать смену"')
        self.assertContains(response, 'function bindDriverShiftHoldAction(form, button, options)')
        self.assertContains(response, '/static/js/mobile-shift-unified-v1.js')
        self.assertNotContains(response, 'class="driver-form-card"')
        self.assertEqual(self.client.session.get('employee_access_id'), self.access.id)

    def test_driver_active_shift_uses_shared_mobile_shift_component(self):
        self.create_registered_driver_shift()

        response = self.client.get(f"{reverse('driver_shift')}?tab=shift", HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'class="mobile-shift"')
        self.assertContains(response, 'data-mobile-shift-role="driver"')
        self.assertContains(response, 'data-mobile-shift-state="open"')
        self.assertContains(response, 'Показатели техники')
        self.assertContains(response, 'Итог смены')
        self.assertContains(response, 'data-driver-shift-close-button')
        self.assertContains(response, 'aria-label="Удерживайте 2 секунды, чтобы закрыть смену"')
        self.assertContains(response, 'holdMs: 2000')
        self.assertContains(response, 'data-driver-shift-inputs')
        self.assertContains(response, 'data-driver-shift-scroll')
        self.assertContains(response, "css/mobile-shift-unified-v1.css")
        self.assertNotContains(response, 'mobile-shift__assignment')
        self.assertContains(response, 'mobile-shift__readings')
        self.assertContains(response, 'mobile-shift__summary')
        self.assertContains(response, 'mobile-shift__actions')
        self.assertContains(response, 'mobile-shift-nav-icon')
        self.assertNotContains(response, 'data-driver-shift-inputs hidden')
        self.assertContains(response, 'data-mobile-shift-field="fuel"')
        self.assertContains(response, 'data-mobile-shift-field="mileage"')
        self.assertContains(response, 'data-mobile-shift-field="engine_hours"')
        self.assertContains(response, '>Закрыть смену<')
        self.assertContains(response, 'Начало')
        self.assertContains(response, '12560')
        self.assertContains(response, '1354')
        self.assertContains(response, '>Выйти<')
        self.assertContains(response, '/static/js/mobile-shift-unified-v1.js')
        self.assertContains(response, 'data-mobile-shift-label')
        self.assertContains(response, 'if (logoutLabel) logoutLabel.textContent = "Выходим";')
        self.assertNotContains(response, '>Закрытие смены<')
        self.assertNotContains(response, 'Проверить показания')
        self.assertNotContains(response, 'data-eo-shift-review')
        self.assertNotContains(response, 'data-driver-shift-review')
        self.assertNotContains(response, 'driverShiftCloseHoldGuard')
        self.assertContains(response, 'data-driver-manifest-view-open="report"')
        self.assertContains(response, 'data-driver-manifest-view-open="timeline"')
        self.assertContains(response, 'data-driver-report-copy')
        self.assertContains(response, 'data-driver-report-share')
        self.assertContains(response, 'buildDriverShiftReportText')
        self.assertContains(response, 'Отчёт водителя за смену')
        self.assertContains(response, 'Событий в текущей смене пока нет')

    def test_driver_screen_includes_own_pwa_install_metadata(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        dormitory = Dormitory.objects.create(number='5')
        block = DormitoryBlock.objects.create(dormitory=dormitory, name='Блок 1')
        section = DormitorySection.objects.create(block=block, name='А')
        DriverPrimaryRegistration.objects.create(
            employee=self.employee,
            dormitory_section=section,
        )
        session = self.client.session
        session['employee_access_id'] = self.access.id
        session.save()

        response = self.client.get(reverse('driver_work'), HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse('driver_manifest'))
        self.assertContains(response, 'rel="manifest"')
        self.assertContains(response, '/driver-sw.js')
        self.assertContains(response, 'driver-mobile-shell-v177')
        self.assertContains(response, 'data-driver-pwa-update-modal')
        self.assertContains(response, 'data-driver-pwa-update-badge')
        self.assertContains(
            response,
            'mode: "custom", path: "^/driver/(?:shift(?:/close)?/?)?$"',
        )
        self.assertContains(response, 'window.applyOperationalStateRefresh')
        self.assertContains(response, 'window.bindDriverMobileShell')
        self.assertNotContains(response, '?version_check')
        self.assertNotContains(response, 'navigator.serviceWorker.register("/driver-sw.js"')
        self.assertNotContains(response, 'window.' + 'alert')
        self.assertContains(response, '--driver-nav-total-h')
        self.assertContains(response, '--driver-edge: clamp(16px, 5vw, 34px)')
        self.assertContains(response, '--driver-header-h: 54px')
        self.assertContains(response, '--driver-header-total-h')
        self.assertContains(response, '--driver-nav-content-h: 78px')
        self.assertContains(response, 'height: var(--driver-header-total-h)')
        self.assertContains(response, 'height: var(--driver-nav-total-h, 78px)')
        self.assertContains(response, 'grid-template-columns: minmax(0, 1fr) auto')
        self.assertContains(response, 'class="driver-header-id')
        self.assertContains(response, 'Самосвал')
        self.assertContains(response, 'white-space: nowrap')
        self.assertContains(response, 'text-overflow: clip')
        self.assertContains(response, 'body.driver-mobile-screen .driver-online::after')
        self.assertContains(response, 'display: none !important')
        self.assertNotContains(response, '>Активная смена<')
        self.assertContains(response, '--driver-dial-size: clamp(320px, min(92vw, 62dvh), 520px)')
        self.assertContains(response, 'grid-template-areas:')
        self.assertContains(response, '"context"')
        self.assertContains(response, '"dial"')
        self.assertContains(response, '"assign"')
        self.assertContains(response, 'gap: var(--driver-work-gap)')
        self.assertContains(response, 'class="driver-work-context-card"')
        self.assertContains(response, '--driver-work-column: min(100%, clamp(320px, 92vw, 720px))')
        self.assertContains(response, 'class="driver-work-context-heading"')
        self.assertContains(response, 'class="driver-work-context-machine"')
        self.assertContains(response, 'class="driver-work-context-geology"')
        self.assertContains(response, 'class="driver-work-context-geology-values"')
        self.assertContains(response, 'Место погрузки · порода')
        self.assertContains(response, 'class="driver-work-context-location"')
        self.assertContains(response, 'class="driver-work-context-rock"')
        self.assertContains(response, 'Комплекс')
        self.assertNotContains(response, 'function enhanceDriverContextLine()')
        self.assertContains(response, 'class="driver-work-ticks"')
        self.assertContains(response, 'data-driver-dial-label')
        self.assertContains(response, 'className = "driver-work-label-line"')
        self.assertContains(response, 'function splitDriverDialLabel(text)')
        self.assertContains(response, 'function preferredDriverDialFontSize(coreWidth, lineCount, textLength)')
        self.assertContains(response, 'function minimumDriverDialFontSize(lineCount, textLength)')
        self.assertContains(response, 'function fitDriverDialLabel(label)')
        self.assertContains(response, 'Math.min(60, Math.max(50, coreWidth * 0.23))')
        self.assertContains(response, 'Math.min(54, Math.max(42, coreWidth * 0.2))')
        self.assertContains(response, 'Math.min(48, Math.max(35, coreWidth * 0.18))')
        self.assertContains(response, 'Math.min(40, Math.max(30, coreWidth * 0.15))')
        self.assertContains(response, 'return 25')
        self.assertContains(response, 'has-multiline-label')
        self.assertContains(response, 'row-gap: 6px')
        self.assertContains(response, 'row-gap: 4px')
        self.assertContains(response, 'line-height: 1.04')
        self.assertContains(response, 'font-size: 22px')
        self.assertContains(response, 'font-size: 15px')
        self.assertContains(response, 'linear-gradient(180deg, rgba(153,255,104,0.08)')
        self.assertContains(response, 'linear-gradient(90deg, transparent, rgba(20,54,35,0.34)')
        self.assertContains(response, 'body.driver-mobile-screen .driver-work-dial-button.is-loaded .driver-work-percent')
        self.assertContains(response, 'body.driver-mobile-screen .driver-work-dial-core::before')
        self.assertContains(response, 'isolation: isolate')
        self.assertContains(response, '-webkit-background-clip: text')
        self.assertContains(response, '-webkit-text-fill-color: transparent')
        self.assertContains(response, 'drop-shadow(0 3px 2px rgba(0,0,0,0.46))')
        self.assertContains(response, 'body.driver-mobile-screen .driver-work-dial-button.is-holding .driver-work-label')
        self.assertContains(response, 'transform: translateY(2px)')
        self.assertContains(response, 'body.driver-mobile-screen .driver-work-ticks::before')
        self.assertContains(response, 'body.driver-mobile-screen .driver-work-ticks::after')
        self.assertContains(response, 'repeating-conic-gradient(from -0.35deg')
        self.assertContains(response, 'stroke-linecap: round')
        self.assertContains(response, 'class="driver-work-over-progress"')
        self.assertContains(response, 'class="driver-work-hold-bar is-core"')
        self.assertContains(response, 'class="driver-work-hold-bar is-outer"')
        self.assertContains(response, 'body.driver-mobile-screen .driver-work-hold-bar.is-core')
        self.assertContains(response, 'body.driver-mobile-screen .driver-work-hold-bar.is-outer')
        self.assertContains(response, 'conic-gradient(')
        self.assertContains(response, 'var(--driver-hold-angle)')
        self.assertContains(response, 'transform: scaleX(-1)')
        self.assertContains(response, 'inset: 13%')
        self.assertContains(response, 'inset: 1.5%')
        self.assertNotContains(response, 'class="driver-work-hold-svg"')
        self.assertContains(response, 'width: 74%')
        self.assertContains(response, 'width: max-content')
        self.assertNotContains(response, '--driver-dial-size: clamp(260px, 76vw, 380px)')
        self.assertContains(response, 'var holdMs = 2000')
        self.assertContains(
            response,
            'holdButton.style.setProperty("--driver-hold", percent.toFixed(2))',
        )
        self.assertContains(response, '((percent / 100) * 360).toFixed(2) + "deg"')
        self.assertContains(response, 'window.requestAnimationFrame(function ()')
        self.assertContains(response, 'data-driver-progress=')
        self.assertContains(response, 'function syncDriverDialProgress()')
        self.assertContains(response, '--driver-progress-capped')
        self.assertContains(response, '--driver-over-progress')
        self.assertContains(response, 'is-over-plan')
        self.assertContains(response, 'body.driver-mobile-screen .driver-work-dial-button.is-pending .driver-work-label')
        self.assertContains(response, 'max-width: 88%')
        self.assertContains(response, 'body.driver-mobile-screen .driver-work-dial-button.is-pending .driver-work-percent')
        self.assertContains(response, 'body.driver-mobile-screen .driver-work-assignment')
        self.assertContains(response, 'width: min(var(--driver-dial-size), 100%)')
        self.assertContains(response, 'max-width: 520px')
        self.assertContains(response, '--driver-dial-size: clamp(320px, min(42vw, 58dvh), 520px)')
        self.assertContains(response, 'max-width: 520px')
        self.assertNotContains(response, 'width: min(var(--driver-dial-size), 57dvh, 100%)')
        self.assertNotContains(response, '"context dial"')
        self.assertNotContains(response, '"assign dial"')
        self.assertContains(response, 'grid-template-rows: auto minmax(0, 1fr) auto')
        self.assertContains(response, '[data-driver-tab-panel="downtimes"].is-active')
        self.assertContains(response, 'max-height: 100%')
        self.assertContains(response, 'padding: 0 0 max(10px, var(--driver-safe-bottom, 0px))')
        self.assertContains(response, 'width: 100%')
        self.assertContains(response, 'max-width: none')
        self.assertContains(response, 'min-width: 320px')
        self.assertNotContains(response, 'width: min(100vw, 430px)')
        self.assertNotContains(response, 'max-width: 430px')
        self.assertNotContains(response, 'margin: 0 auto')
        self.assertContains(response, 'position: fixed')
        self.assertContains(response, 'right: 0')
        self.assertContains(response, 'body.driver-mobile-screen .driver-bottom-nav .mm-mobile-nav-item.is-active::after')
        self.assertContains(response, 'body.driver-mobile-screen .driver-bottom-nav .mm-mobile-nav-item:focus-visible')
        self.assertContains(response, 'body.driver-mobile-screen .driver-nav-truck')
        self.assertNotContains(response, 'driver-work-' + 'bottom-nav')

    def test_browser_driver_keeps_pwa_update_ui(self):
        self.create_registered_driver_shift()

        response = self.client.get(reverse('driver_work'), HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<small class="driver-shell-version" aria-label="Версия приложения">')
        self.assertContains(response, '<div class="driver-mobile-update-modal" data-driver-pwa-update-modal')
        self.assertContains(response, '<span class="driver-mobile-update-badge" data-driver-pwa-update-badge')
        self.assertContains(response, 'data-driver-tab-open="manifest" data-driver-pwa-update-nav-target')
        self.assertContains(response, 'data-driver-shift-update')

    def test_native_driver_hides_pwa_update_ui_but_keeps_operational_refresh(self):
        self.create_registered_driver_shift()

        response = self.client.get(
            reverse('driver_work'),
            HTTP_HOST='localhost',
            HTTP_USER_AGENT='Mozilla/5.0 CopperResourcesNative/driver',
        )

        self.assertEqual(response.status_code, 200)
        # Метка без версии — старая сборка. Версию оболочки показывать нельзя:
        # её примут за версию приложения. Лучше не показывать ничего.
        self.assertNotContains(response, '<small class="driver-shell-version"')
        self.assertContains(response, 'class="driver-native-app-version native-app-version"', count=1)
        self.assertContains(response, 'data-native-app-version=""', count=1)
        self.assertNotContains(response, '<div class="driver-mobile-update-modal" data-driver-pwa-update-modal')
        self.assertNotContains(response, '<span class="driver-mobile-update-badge" data-driver-pwa-update-badge')
        self.assertNotContains(response, 'data-driver-tab-open="manifest" data-driver-pwa-update-nav-target')

    def test_native_driver_shows_real_app_version_when_reported(self):
        """Приложение сообщает свою версию — показываем именно её.

        Раньше версии в приложении не было вообще: версию веб-оболочки
        спрятали как вводящую в заблуждение, а своей ещё не было. Теперь
        сборка дописывает её в User-Agent.
        """
        self.create_registered_driver_shift()

        response = self.client.get(
            reverse('driver_work'),
            HTTP_HOST='localhost',
            HTTP_USER_AGENT='Mozilla/5.0 (Linux; Android 13; wv) CopperResourcesNative/driver/0.1.3 Mobile Safari/537.36',
        )

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, '<small class="driver-shell-version"')
        self.assertContains(response, 'class="driver-native-app-version native-app-version"', count=1)
        self.assertContains(response, 'data-native-app-version="0.1.3"', count=1)
        self.assertContains(response, '>Версия 0.1.3</span>', count=1)
        # Проверяем именно сам элемент версии: строка «driver-mobile-shell-»
        # есть на странице и в других местах (service worker, метаданные),
        # поэтому голый поиск подстроки ложно падает.
        self.assertNotContains(
            response,
            'aria-label="Текущая версия приложения">Версия driver-mobile-shell-',
        )

    def test_driver_manifest_is_installable_pwa_manifest(self):
        response = self.client.get(reverse('driver_manifest'), HTTP_HOST='localhost')
        payload = response.json()

        self.assertEqual(response.status_code, 200)
        self.assertIn('application/manifest+json', response['Content-Type'])
        self.assertEqual(payload['name'], 'Водитель самосвала')
        self.assertEqual(payload['start_url'], '/driver/')
        self.assertEqual(payload['scope'], '/driver/')
        self.assertEqual(payload['display'], 'standalone')
        self.assertEqual(payload['orientation'], 'portrait')
        self.assertTrue(payload['icons'])

    def test_driver_service_worker_caches_driver_assets_and_reports_version(self):
        response = self.client.get(reverse('driver_service_worker'), HTTP_HOST='localhost')
        script = response.content.decode('utf-8')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Service-Worker-Allowed'], '/driver/')
        self.assertIn('driver-mobile-shell-v177', script)
        self.assertIn('/driver/', script)
        self.assertIn('/driver/shift/', script)
        self.assertIn('/driver.webmanifest', script)
        self.assertIn('/static/css/app.css', script)
        self.assertIn('/static/css/native-app-update-v1.css', script)
        self.assertIn('/static/js/role-readonly.js', script)
        self.assertIn('/static/css/mobile-shift-unified-v1.css', script)
        self.assertIn('/static/js/mobile-shift-unified-v1.js', script)
        self.assertNotIn('ignoreSearch: true', script)
        self.assertIn('networkFirstStatic(request)', script)
        self.assertIn('GET_VERSION', script)
        self.assertIn('SKIP_WAITING', script)
        self.assertIn('skipWaiting', script)
        install_handler = script.split('const APP_CONTRACT_VERSION', maxsplit=1)[0]
        self.assertNotIn('skipWaiting', install_handler)
        self.assertIn('clients.claim', script)
        self.assertIn('key.startsWith(CACHE_PREFIX)', script)

    def test_admin_opens_system_admin_dashboard(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        admin_employee = Employee.objects.create(full_name='Администратор MVP', status=Employee.Status.ACTIVE)
        EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='1000',
            status=EmployeeAccess.Status.ACTIVATED,
        )

        login_response = self.client.post('/', {'access_code': '1000'}, follow=True, HTTP_HOST='localhost')
        dashboard_response = self.client.get('/system-admin/', HTTP_HOST='localhost')

        self.assertRedirects(login_response, '/system-admin/', target_status_code=200)
        self.assertEqual(dashboard_response.status_code, 200)
        self.assertContains(dashboard_response, 'Админка MVP')
        self.assertContains(dashboard_response, 'Создать сотрудника')
        self.assertContains(dashboard_response, 'Справочники')
        self.assertContains(dashboard_response, 'href="/system-admin/employees/"')
        self.assertContains(dashboard_response, 'href="/system-admin/employees/?status=active"')
        self.assertContains(dashboard_response, 'href="/system-admin/employees/?access_status=not_activated"')
        self.assertContains(dashboard_response, 'href="/system-admin/employees/?access_status=blocked"')
        self.assertContains(dashboard_response, 'href="/system-admin/employees/?access_status=deactivated"')
        self.assertContains(dashboard_response, 'Журнал действий')

    def test_admin_can_reset_shift_test_data_without_deleting_base_data(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        driver_role = Role.objects.create(code='driver_reset_test', name='Водитель самосвала')
        admin_employee = Employee.objects.create(full_name='Администратор MVP', status=Employee.Status.ACTIVE)
        driver = Employee.objects.create(full_name='Тестовый водитель', status=Employee.Status.ACTIVE)
        admin_access = EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='1000',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        EmployeeAccess.objects.create(
            employee=driver,
            role=driver_role,
            access_code='2000',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        equipment_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=equipment_type, garage_number='10')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='8')
        rock_type = RockType.objects.create(name='Руда', density=Decimal('2.6'))
        dump_point = DumpPoint.objects.create(name='ККД')
        shift = EmployeeShift.objects.create(
            employee=driver,
            shift_type='day',
            equipment=truck,
            opened_at=timezone.now(),
            opened_by=driver,
        )
        plan = ShiftPlan.objects.create(name='Дневной план', plan_volume_m3=Decimal('1000'))
        downtime_reason, _ = DowntimeReason.objects.get_or_create(
            name='Ожидание разгрузки',
            defaults={'show_for_truck_driver': True},
        )
        trip = Trip.objects.create(
            excavator=excavator,
            truck=truck,
            driver=driver,
            unloading_shift=shift,
            rock_type=rock_type,
            dump_point=dump_point,
            volume_m3=Decimal('49.4'),
            tonnage=Decimal('128.44'),
            status=TripStatus.COMPLETED,
            completed_at=timezone.now(),
        )
        TripClientAction.objects.create(
            action_type='trip_unloaded',
            client_action_id='reset-test-action',
            trip=trip,
            actor=driver,
        )
        DispatcherActionLog.objects.create(
            actor=admin_employee,
            action_type=DispatcherActionType.COMPLETE_TRIP,
            trip=trip,
            target_summary='Тестовый рейс',
        )
        DowntimeEvent.objects.create(
            equipment=truck,
            employee=driver,
            reason=downtime_reason,
            started_at=timezone.now() - timedelta(minutes=20),
            ended_at=timezone.now(),
        )
        OperationalStateEvent.objects.create(
            key='production',
            version=1,
            event_type='trip_changed',
            object_type='Trip',
            object_id=str(trip.id),
        )

        session = self.client.session
        session['employee_access_id'] = admin_access.id
        session.save()
        response = self.client.post(reverse('system_admin_reset_shift_test_data'), follow=True, HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(Trip.objects.count(), 0)
        self.assertEqual(TripClientAction.objects.count(), 0)
        self.assertEqual(DispatcherActionLog.objects.count(), 0)
        self.assertEqual(DowntimeEvent.objects.count(), 0)
        self.assertTrue(Employee.objects.filter(id=admin_employee.id).exists())
        self.assertTrue(Employee.objects.filter(id=driver.id).exists())
        self.assertTrue(Equipment.objects.filter(id=truck.id).exists())
        self.assertTrue(Equipment.objects.filter(id=excavator.id).exists())
        self.assertTrue(RockType.objects.filter(id=rock_type.id).exists())
        self.assertTrue(DumpPoint.objects.filter(id=dump_point.id).exists())
        self.assertTrue(EmployeeShift.objects.filter(id=shift.id).exists())
        self.assertTrue(ShiftPlan.objects.filter(id=plan.id).exists())
        self.assertTrue(OperationalStateEvent.objects.filter(event_type='test_shift_data_reset').exists())
        self.assertTrue(AdminActionLog.objects.filter(action='Сброшены тестовые показатели смены').exists())

    def test_admin_employee_list_can_filter_by_access_status(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        driver_role = Role.objects.create(code='driver_access_filter', name='Водитель самосвала')
        admin_employee = Employee.objects.create(full_name='Администратор MVP', status=Employee.Status.ACTIVE)
        waiting_employee = Employee.objects.create(full_name='Ожидает активации', status=Employee.Status.NOT_ACTIVATED)
        active_employee = Employee.objects.create(full_name='Активный водитель', status=Employee.Status.ACTIVE)
        EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='1000',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        EmployeeAccess.objects.create(
            employee=waiting_employee,
            role=driver_role,
            access_code='200001',
            status=EmployeeAccess.Status.NOT_ACTIVATED,
            is_active=True,
        )
        EmployeeAccess.objects.create(
            employee=active_employee,
            role=driver_role,
            access_code='200002',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )

        self.client.post('/', {'access_code': '1000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get('/system-admin/employees/?access_status=not_activated', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Ожидает активации')
        self.assertNotContains(response, 'Активный водитель')
        self.assertContains(response, 'name="access_status"')

    def test_admin_cannot_block_own_access(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        admin_employee = Employee.objects.create(full_name='Администратор MVP', status=Employee.Status.ACTIVE)
        admin_access = EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='1000',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )

        self.client.post('/', {'access_code': '1000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.post(
            f'/system-admin/accesses/{admin_access.id}/block/',
            {'reason': 'Случайная самоблокировка'},
            follow=True,
            HTTP_HOST='localhost',
        )
        admin_access.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(admin_access.status, EmployeeAccess.Status.ACTIVATED)
        self.assertTrue(admin_access.is_active)

    def test_admin_cannot_deactivate_own_employee_card(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        admin_employee = Employee.objects.create(full_name='Администратор MVP', status=Employee.Status.ACTIVE)
        EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='1000',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )

        self.client.post('/', {'access_code': '1000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.post(
            f'/system-admin/employees/{admin_employee.id}/deactivate/',
            follow=True,
            HTTP_HOST='localhost',
        )
        admin_employee.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(admin_employee.status, Employee.Status.ACTIVE)
        self.assertTrue(admin_employee.is_active)

    def test_admin_employee_card_has_photo_upload_block(self):
        admin_role = Role.objects.create(code='admin', name='Admin')
        admin_employee = Employee.objects.create(full_name='Admin MVP', status=Employee.Status.ACTIVE)
        employee = Employee.objects.create(full_name='Employee With Photo')
        EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='1000',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )

        self.client.post('/', {'access_code': '1000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get(f'/system-admin/employees/{employee.id}/', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'employee-card-profile')
        self.assertContains(response, 'employee-card-photo')
        self.assertContains(response, 'employee-card-photo-actions')
        self.assertContains(response, 'name="position"')
        self.assertContains(response, 'Кадровая должность')
        self.assertContains(response, 'type="file"')

    def test_admin_employee_card_keeps_selected_role_and_primary_pin_status(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        driver_role = Role.objects.create(code='driver_primary_pin', name='Водитель самосвала')
        admin_employee = Employee.objects.create(full_name='Администратор MVP', status=Employee.Status.ACTIVE)
        employee = Employee.objects.create(
            full_name='Водитель с доступом',
            phone='+79990000000',
            status=Employee.Status.ACTIVE,
        )
        employee_access = EmployeeAccess.objects.create(
            employee=employee,
            role=driver_role,
            access_code='246824',
            status=EmployeeAccess.Status.NOT_ACTIVATED,
            primary_code_issued_at=timezone.now(),
            is_active=True,
        )
        EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='1000',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )

        self.client.post('/', {'access_code': '1000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get(f'/system-admin/employees/{employee.id}/', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, f'<option value="{driver_role.id}" selected>', html=False)
        self.assertContains(response, '2468')
        self.assertContains(response, 'ожидает первого входа')
        self.assertContains(response, 'employee-login-share')
        self.assertContains(response, 'https://driverform.ru')
        self.assertContains(response, 'Телефон: +79990000000')
        self.assertContains(response, 'Пин код: 246824')
        self.assertContains(response, 'Скопировать')
        self.assertContains(response, 'Отправить')

        employee_access.access_code = '8642'
        employee_access.status = EmployeeAccess.Status.ACTIVATED
        employee_access.activated_at = timezone.now()
        employee_access.last_login_at = timezone.now()
        employee_access.save(update_fields=['access_code', 'status', 'activated_at', 'last_login_at'])

        activated_response = self.client.get(f'/system-admin/employees/{employee.id}/', HTTP_HOST='localhost')

        self.assertContains(activated_response, f'<option value="{driver_role.id}" selected>', html=False)
        self.assertContains(activated_response, 'Пинкод активирован')
        self.assertNotContains(activated_response, 'employee-login-share')
        self.assertNotContains(activated_response, '8642')

        reset_response = self.client.post(
            f'/system-admin/employees/{employee.id}/generate-access/',
            {'role': driver_role.id},
            follow=True,
            HTTP_HOST='localhost',
        )
        employee_access.refresh_from_db()

        self.assertEqual(reset_response.status_code, 200)
        self.assertEqual(employee_access.status, EmployeeAccess.Status.NOT_ACTIVATED)
        self.assertIsNone(employee_access.activated_at)
        self.assertNotEqual(employee_access.access_code, '8642')
        self.assertContains(reset_response, employee_access.access_code)
        self.assertContains(reset_response, 'Телефон: +79990000000')
        self.assertContains(reset_response, f'Пин код: {employee_access.access_code}')
        self.assertContains(reset_response, 'ожидает первого входа')
        self.assertNotContains(reset_response, '8642')

    def test_admin_can_change_employee_role_without_resetting_credentials(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        manager_role = Role.objects.create(code='manager_role_change', name='Руководство')
        driver_role = Role.objects.create(code='driver_role_change', name='Водитель самосвала')
        admin_employee = Employee.objects.create(full_name='Администратор MVP', status=Employee.Status.ACTIVE)
        employee = Employee.objects.create(full_name='Сотрудник для смены роли', status=Employee.Status.ACTIVE)
        admin_access = EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='100000',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        activated_at = timezone.now() - timedelta(days=10)
        last_login_at = timezone.now() - timedelta(hours=2)
        employee_access = EmployeeAccess.objects.create(
            employee=employee,
            role=manager_role,
            access_code='654321',
            status=EmployeeAccess.Status.ACTIVATED,
            activated_at=activated_at,
            last_login_at=last_login_at,
            is_active=True,
        )
        session = self.client.session
        session['employee_access_id'] = admin_access.id
        session.save()

        lock_order = []
        employee_select_for_update = Employee.objects.select_for_update
        access_select_for_update = EmployeeAccess.objects.select_for_update

        def lock_employee(*args, **kwargs):
            lock_order.append('employee')
            return employee_select_for_update(*args, **kwargs)

        def lock_access(*args, **kwargs):
            lock_order.append('access')
            return access_select_for_update(*args, **kwargs)

        with (
            patch.object(Employee.objects, 'select_for_update', side_effect=lock_employee),
            patch.object(EmployeeAccess.objects, 'select_for_update', side_effect=lock_access),
        ):
            response = self.client.post(
                reverse('system_admin_change_access_role', args=[employee_access.id]),
                {'role': driver_role.id},
                follow=True,
                HTTP_HOST='localhost',
            )
        employee_access.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(employee_access.role, driver_role)
        self.assertEqual(employee_access.access_code, '654321')
        self.assertEqual(employee_access.status, EmployeeAccess.Status.ACTIVATED)
        self.assertEqual(employee_access.activated_at, activated_at)
        self.assertEqual(employee_access.last_login_at, last_login_at)
        self.assertTrue(employee_access.is_active)
        self.assertEqual(lock_order[:2], ['employee', 'access'])
        self.assertEqual(EmployeeAccess.objects.filter(employee=employee).count(), 1)
        self.assertContains(response, f'<option value="{driver_role.id}" selected>', html=False)
        self.assertTrue(
            AdminActionLog.objects.filter(
                action='Изменена роль доступа сотрудника',
                old_value=manager_role.name,
                new_value=driver_role.name,
            ).exists()
        )

    def test_employee_card_has_separate_role_change_and_pin_reset_actions(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        manager_role = Role.objects.create(code='manager_role_ui', name='Руководство')
        admin_employee = Employee.objects.create(full_name='Администратор MVP', status=Employee.Status.ACTIVE)
        employee = Employee.objects.create(full_name='Сотрудник с ролью', status=Employee.Status.ACTIVE)
        admin_access = EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='100000',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        employee_access = EmployeeAccess.objects.create(
            employee=employee,
            role=manager_role,
            access_code='654321',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        session = self.client.session
        session['employee_access_id'] = admin_access.id
        session.save()

        response = self.client.get(
            reverse('system_admin_employee_detail', args=[employee.id]),
            HTTP_HOST='localhost',
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse('system_admin_change_access_role', args=[employee_access.id]))
        self.assertContains(response, 'Сохранить роль')
        self.assertContains(response, 'PIN, пароль и статус активации останутся прежними.')
        self.assertContains(response, 'Сбросить PIN')
        self.assertContains(response, f'name="role" value="{manager_role.id}"', html=False)

    def test_role_change_clears_incompatible_permanent_equipment_assignment(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        driver_role = self.role
        manager_role = Role.objects.create(code='manager_role_assignment', name='Руководство')
        admin_employee = Employee.objects.create(full_name='Администратор MVP', status=Employee.Status.ACTIVE)
        employee = Employee.objects.create(full_name='Водитель со старым назначением', status=Employee.Status.ACTIVE)
        admin_access = EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='100000',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        employee_access = EmployeeAccess.objects.create(
            employee=employee,
            role=driver_role,
            access_code='654321',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        truck_type = EquipmentType.objects.create(name='Самосвал')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        assignment = EquipmentAssignment.objects.create(
            employee=employee,
            role=driver_role,
            equipment=truck,
            shift_type='day',
            assigned_by=admin_employee,
            status=AssignmentStatus.ACCEPTED,
            accepted_at=timezone.now(),
        )
        session = self.client.session
        session['employee_access_id'] = admin_access.id
        session.save()

        response = self.client.post(
            reverse('system_admin_change_access_role', args=[employee_access.id]),
            {'role': manager_role.id},
            follow=True,
            HTTP_HOST='localhost',
        )
        assignment.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertIsNotNone(assignment.ended_at)
        self.assertEqual(assignment.ended_by, admin_employee)
        employee_access.refresh_from_db()
        self.assertEqual(employee_access.role, manager_role)

    def test_role_change_is_blocked_while_employee_shift_is_open(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        manager_role = Role.objects.create(code='manager_role_open_shift', name='Руководство')
        driver_role = Role.objects.create(code='driver_role_open_shift', name='Водитель самосвала')
        admin_employee = Employee.objects.create(full_name='Администратор MVP', status=Employee.Status.ACTIVE)
        employee = Employee.objects.create(full_name='Сотрудник в открытой смене', status=Employee.Status.ACTIVE)
        admin_access = EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='100000',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        employee_access = EmployeeAccess.objects.create(
            employee=employee,
            role=manager_role,
            access_code='654321',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        EmployeeShift.objects.create(
            employee=employee,
            shift_type='day',
            opened_at=timezone.now(),
            opened_by=employee,
        )
        session = self.client.session
        session['employee_access_id'] = admin_access.id
        session.save()

        response = self.client.post(
            reverse('system_admin_change_access_role', args=[employee_access.id]),
            {'role': driver_role.id},
            follow=True,
            HTTP_HOST='localhost',
        )
        employee_access.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(employee_access.role, manager_role)
        self.assertEqual(employee_access.access_code, '654321')
        self.assertFalse(AdminActionLog.objects.filter(action='Изменена роль доступа сотрудника').exists())

    def test_role_change_does_not_create_duplicate_employee_role_access(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        manager_role = Role.objects.create(code='manager_role_duplicate', name='Руководство')
        driver_role = Role.objects.create(code='driver_role_duplicate', name='Водитель самосвала')
        admin_employee = Employee.objects.create(full_name='Администратор MVP', status=Employee.Status.ACTIVE)
        employee = Employee.objects.create(full_name='Сотрудник с двумя доступами', status=Employee.Status.ACTIVE)
        admin_access = EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='100000',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        manager_access = EmployeeAccess.objects.create(
            employee=employee,
            role=manager_role,
            access_code='654321',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        EmployeeAccess.objects.create(
            employee=employee,
            role=driver_role,
            access_code='123456',
            status=EmployeeAccess.Status.DEACTIVATED,
            is_active=False,
        )
        session = self.client.session
        session['employee_access_id'] = admin_access.id
        session.save()

        response = self.client.post(
            reverse('system_admin_change_access_role', args=[manager_access.id]),
            {'role': driver_role.id},
            follow=True,
            HTTP_HOST='localhost',
        )
        manager_access.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(manager_access.role, manager_role)
        self.assertEqual(manager_access.access_code, '654321')
        self.assertEqual(EmployeeAccess.objects.filter(employee=employee).count(), 2)

    def test_admin_cannot_change_own_access_role(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        manager_role = Role.objects.create(code='manager_role_self_change', name='Руководство')
        admin_employee = Employee.objects.create(full_name='Администратор MVP', status=Employee.Status.ACTIVE)
        admin_access = EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='100000',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        session = self.client.session
        session['employee_access_id'] = admin_access.id
        session.save()

        response = self.client.post(
            reverse('system_admin_change_access_role', args=[admin_access.id]),
            {'role': manager_role.id},
            follow=True,
            HTTP_HOST='localhost',
        )
        admin_access.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(admin_access.role, admin_role)
        self.assertEqual(admin_access.access_code, '100000')
        self.assertTrue(admin_access.is_active)

    def test_admin_employee_card_with_photo_has_remove_confirmation_and_modal(self):
        with TemporaryDirectory() as media_root:
            with override_settings(MEDIA_ROOT=media_root):
                admin_role = Role.objects.create(code='admin', name='Admin')
                admin_employee = Employee.objects.create(full_name='Admin MVP', status=Employee.Status.ACTIVE)
                employee = Employee.objects.create(full_name='Employee With Photo', status=Employee.Status.ACTIVE)
                employee.photo.save('employee_photos/current.jpg', ContentFile(b'photo'), save=True)
                EmployeeAccess.objects.create(
                    employee=admin_employee,
                    role=admin_role,
                    access_code='1000',
                    status=EmployeeAccess.Status.ACTIVATED,
                    is_active=True,
                )

                self.client.post('/', {'access_code': '1000'}, follow=True, HTTP_HOST='localhost')
                response = self.client.get(f'/system-admin/employees/{employee.id}/', HTTP_HOST='localhost')

                self.assertEqual(response.status_code, 200)
                self.assertContains(response, 'employee-card-photo-action is-add')
                self.assertContains(response, 'employee-card-photo-action is-remove')
                self.assertContains(response, 'data-confirm="Удалить фото сотрудника?"')
                self.assertContains(response, 'app-confirm-modal')
                self.assertContains(response, 'data-confirm-accept')
                self.assertContains(response, 'data-confirm-cancel')
                self.assertNotContains(response, 'onclick="return window.confirm')
                self.assertContains(response, 'employee-card-photo-modal')

    def test_employee_photo_rejects_non_image_file(self):
        employee = Employee.objects.create(full_name='Employee Photo Validation')
        upload = SimpleUploadedFile('note.txt', b'not-image', content_type='text/plain')

        form = AdminEmployeeEditForm(
            data={'full_name': employee.full_name, 'status': Employee.Status.ACTIVE},
            files={'photo': upload},
            instance=employee,
        )

        self.assertFalse(form.is_valid())
        self.assertIn('photo', form.errors)

    def test_employee_photo_upload_is_converted_to_jpeg(self):
        employee = Employee.objects.create(full_name='Employee Photo Compression')
        image_buffer = BytesIO()
        image = Image.new('RGB', (900, 700), color=(40, 180, 150))
        image.save(image_buffer, format='PNG')
        upload = SimpleUploadedFile('photo.png', image_buffer.getvalue(), content_type='image/png')

        form = AdminEmployeeEditForm(
            data={'full_name': employee.full_name, 'status': Employee.Status.ACTIVE},
            files={'photo': upload},
            instance=employee,
        )

        self.assertTrue(form.is_valid(), form.errors)
        photo = form.cleaned_data['photo']
        self.assertTrue(photo.name.endswith('.jpg'))
        self.assertLessEqual(photo.size, 5 * 1024 * 1024)

    def test_admin_can_replace_existing_employee_photo(self):
        with TemporaryDirectory() as media_root:
            with override_settings(MEDIA_ROOT=media_root):
                admin_role = Role.objects.create(code='admin', name='Admin')
                admin_employee = Employee.objects.create(full_name='Admin MVP', status=Employee.Status.ACTIVE)
                employee = Employee.objects.create(full_name='Employee Replace Photo', status=Employee.Status.ACTIVE)
                employee.photo.save('employee_photos/old.jpg', ContentFile(b'old-photo'), save=True)
                old_photo_name = employee.photo.name
                EmployeeAccess.objects.create(
                    employee=admin_employee,
                    role=admin_role,
                    access_code='1000',
                    status=EmployeeAccess.Status.ACTIVATED,
                    is_active=True,
                )
                image_buffer = BytesIO()
                Image.new('RGB', (700, 700), color=(120, 80, 40)).save(image_buffer, format='PNG')
                upload = SimpleUploadedFile('new-photo.png', image_buffer.getvalue(), content_type='image/png')

                self.client.post('/', {'access_code': '1000'}, follow=True, HTTP_HOST='localhost')
                response = self.client.post(
                    f'/system-admin/employees/{employee.id}/',
                    {'full_name': employee.full_name, 'status': Employee.Status.ACTIVE, 'photo': upload},
                    follow=True,
                    HTTP_HOST='localhost',
                )
                employee.refresh_from_db()

                self.assertEqual(response.status_code, 200)
                self.assertNotEqual(employee.photo.name, old_photo_name)
                self.assertTrue(employee.photo.name.endswith('.jpg'))
                self.assertFalse(employee.photo.storage.exists(old_photo_name))
                self.assertTrue(employee.photo.storage.exists(employee.photo.name))

    def test_admin_can_remove_existing_employee_photo(self):
        with TemporaryDirectory() as media_root:
            with override_settings(MEDIA_ROOT=media_root):
                admin_role = Role.objects.create(code='admin', name='Admin')
                admin_employee = Employee.objects.create(full_name='Admin MVP', status=Employee.Status.ACTIVE)
                employee = Employee.objects.create(full_name='Employee Remove Photo', status=Employee.Status.ACTIVE)
                employee.photo.save('employee_photos/remove.jpg', ContentFile(b'old-photo'), save=True)
                old_photo_name = employee.photo.name
                EmployeeAccess.objects.create(
                    employee=admin_employee,
                    role=admin_role,
                    access_code='1000',
                    status=EmployeeAccess.Status.ACTIVATED,
                    is_active=True,
                )

                self.client.post('/', {'access_code': '1000'}, follow=True, HTTP_HOST='localhost')
                response = self.client.post(
                    f'/system-admin/employees/{employee.id}/',
                    {'remove_photo': '1'},
                    follow=True,
                    HTTP_HOST='localhost',
                )
                employee.refresh_from_db()

                self.assertEqual(response.status_code, 200)
                self.assertFalse(employee.photo)
                self.assertFalse(employee.photo.storage.exists(old_photo_name))
                self.assertTrue(
                    AdminActionLog.objects.filter(
                        object_repr=str(employee),
                        action='Удалено фото сотрудника',
                    ).exists()
                )

    def test_admin_opens_references_registry(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        admin_employee = Employee.objects.create(full_name='Администратор MVP', status=Employee.Status.ACTIVE)
        EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='1000',
            status=EmployeeAccess.Status.ACTIVATED,
        )
        EquipmentType.objects.create(name='Самосвал')
        RockType.objects.create(name='Руда')
        DumpPoint.objects.create(name='ККД')

        self.client.post('/', {'access_code': '1000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get('/system-admin/references/', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Справочники админки')
        self.assertContains(response, 'Виды техники')
        self.assertContains(response, 'Породы')
        self.assertContains(response, 'Точки разгрузки')
        self.assertContains(response, 'Ежесменные планы техники')
        self.assertContains(response, 'Приз за 100% плана')
        self.assertContains(response, 'Сменные планы (история)')
        self.assertContains(response, 'Планы техники (история)')
        self.assertContains(response, '/admin/references/equipmenttype/')
        self.assertContains(response, '/system-admin/references/equipment/')
        self.assertContains(response, '/system-admin/references/equipment-plan-groups/')
        self.assertContains(response, '/system-admin/references/achievement-prizes/')
        self.assertContains(response, '/system-admin/references/shift-plans/')
        self.assertContains(response, '/system-admin/references/equipment-shift-plans/')

        detail_response = self.client.get('/system-admin/references/equipment/', HTTP_HOST='localhost')

        self.assertEqual(detail_response.status_code, 200)
        self.assertContains(detail_response, 'reference-detail-page')
        self.assertContains(detail_response, '/admin/references/equipment/')

    def test_admin_saves_active_achievement_prize_image_from_reference_screen(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        admin_employee = Employee.objects.create(full_name='Администратор MVP', status=Employee.Status.ACTIVE)
        EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='1000',
            status=EmployeeAccess.Status.ACTIVATED,
        )
        self.client.post('/', {'access_code': '1000'}, follow=True, HTTP_HOST='localhost')

        with TemporaryDirectory() as media_root, override_settings(MEDIA_ROOT=media_root):
            page = self.client.get('/system-admin/references/achievement-prizes/', HTTP_HOST='localhost')
            self.assertContains(page, 'Приз за 100% плана')
            self.assertContains(page, 'Призовая картинка')
            image_bytes = BytesIO()
            Image.new('RGB', (8, 8), color='green').save(image_bytes, format='PNG')
            image_bytes.seek(0)

            response = self.client.post(
                '/system-admin/references/achievement-prizes/',
                {
                    'title': 'План выполнен',
                    'image': SimpleUploadedFile('prize.png', image_bytes.read(), content_type='image/png'),
                    'is_active': 'on',
                },
                HTTP_HOST='localhost',
            )

            self.assertEqual(response.status_code, 302)
            prize = AchievementPrize.objects.get(title='План выполнен')
            self.assertTrue(prize.is_active)
            self.assertTrue(prize.image.name.startswith('achievement_prizes/'))

    def test_admin_saves_shift_plans_from_reference_screen(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        admin_employee = Employee.objects.create(full_name='Администратор MVP', status=Employee.Status.ACTIVE)
        EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='1000',
            status=EmployeeAccess.Status.ACTIVATED,
        )
        equipment_type = EquipmentType.objects.create(name='Самосвал')
        equipment = Equipment.objects.create(equipment_type=equipment_type, garage_number='25', is_active=True)

        self.client.post('/', {'access_code': '1000'}, follow=True, HTTP_HOST='localhost')
        group_plan_page = self.client.get('/system-admin/references/equipment-plan-groups/', HTTP_HOST='localhost')
        self.assertContains(group_plan_page, 'Ежесменные планы техники')
        self.assertContains(group_plan_page, 'Группа техники')
        self.assertContains(group_plan_page, 'Тип расчета')
        self.assertContains(group_plan_page, 'Значение плана')
        self.assertContains(group_plan_page, 'Техника в группе')
        self.assertContains(group_plan_page, 'Дата начала действия')
        self.assertNotContains(group_plan_page, 'Расчетная смена')

        group_plan_response = self.client.post(
            '/system-admin/references/equipment-plan-groups/',
            {
                'name': 'Самосвалы БелАЗ админ',
                'code': 'belaz-admin-test',
                'calculation_mode': PlanCalculationMode.TRIPS,
                'plan_value': '18.00',
                'equipment': [str(equipment.id)],
                'is_active': 'on',
                'active_from': timezone.localdate().isoformat(),
                'comment': 'Ежесменный план без даты и day/night',
            },
            HTTP_HOST='localhost',
        )
        group_plan = EquipmentPlanGroup.objects.get(code='belaz-admin-test')

        self.assertEqual(group_plan_response.status_code, 302)
        self.assertEqual(group_plan.calculation_mode, PlanCalculationMode.TRIPS)
        self.assertEqual(group_plan.plan_value, Decimal('18.00'))
        self.assertEqual(group_plan.updated_by, admin_employee)
        self.assertEqual(list(group_plan.equipment.all()), [equipment])

        shift_plan_page = self.client.get('/system-admin/references/shift-plans/', HTTP_HOST='localhost')
        self.assertContains(shift_plan_page, 'Тип плана')
        self.assertContains(shift_plan_page, 'План объема, м3')
        self.assertNotContains(shift_plan_page, 'Дата смены')
        self.assertNotContains(shift_plan_page, 'План тоннажа')
        self.assertNotContains(shift_plan_page, 'План рейсов')

        shift_plan_response = self.client.post(
            '/system-admin/references/shift-plans/',
            {
                'plan_scope': 'day_shift',
                'name': 'План дневной смены',
                'plan_volume_m3': '2500.00',
                'is_active': 'on',
                'comment': 'Ручной план администратора',
            },
            HTTP_HOST='localhost',
        )
        shift_plan = ShiftPlan.objects.get(name='План дневной смены')

        self.assertEqual(shift_plan_response.status_code, 302)
        self.assertEqual(shift_plan.plan_scope, 'day_shift')
        self.assertEqual(shift_plan.plan_volume_m3, Decimal('2500.00'))
        self.assertIsNone(shift_plan.plan_trips)
        self.assertIsNone(shift_plan.plan_tonnage)

        equipment_plan_response = self.client.post(
            '/system-admin/references/equipment-shift-plans/',
            {
                'shift_plan': str(shift_plan.id),
                'equipment': str(equipment.id),
                'employee': '',
                'plan_trips': '20',
                'plan_volume_m3': '500.00',
                'calculation_mode': 'trips',
                'is_active': 'on',
                'comment': 'План самосвала',
            },
            HTTP_HOST='localhost',
        )
        equipment_plan = EquipmentShiftPlan.objects.get(shift_plan=shift_plan, equipment=equipment)

        self.assertEqual(equipment_plan_response.status_code, 302)
        self.assertEqual(equipment_plan.plan_trips, 20)
        self.assertIsNone(equipment_plan.plan_tonnage)
        self.assertEqual(equipment_plan.calculation_mode, 'trips')

    def test_admin_plan_group_editor_uses_checkboxes_and_validates_equipment(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        admin_employee = Employee.objects.create(full_name='Администратор MVP', status=Employee.Status.ACTIVE)
        EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='1000',
            status=EmployeeAccess.Status.ACTIVATED,
        )
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        belaz_model = EquipmentModel.objects.create(equipment_type=truck_type, name='БелАЗ 7513D')
        truck = Equipment.objects.create(equipment_type=truck_type, model=belaz_model, garage_number='25', is_active=True)
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='2', is_active=True)
        group = EquipmentPlanGroup.objects.get(code='excavators_3000')
        group.plan_value = '3000.00'
        group.is_active = True
        group.save(update_fields=['plan_value', 'is_active'])

        self.client.post('/', {'access_code': '1000'}, follow=True, HTTP_HOST='localhost')
        page = self.client.get(
            f'/system-admin/references/equipment-plan-groups/?edit={group.id}',
            HTTP_HOST='localhost',
        )
        self.assertContains(page, 'reference-checkbox-panel')
        self.assertContains(page, 'type="checkbox"')
        self.assertContains(page, 'Самосвал 25')
        self.assertContains(page, 'Экскаватор 2')

        bad_response = self.client.post(
            f'/system-admin/references/equipment-plan-groups/?edit={group.id}',
            {
                'record_id': str(group.id),
                'name': 'Экскаваторы 3000',
                'code': 'excavators_3000',
                'calculation_mode': PlanCalculationMode.VOLUME,
                'plan_value': '3000.00',
                'equipment': [str(truck.id)],
                'is_active': 'on',
                'active_from': timezone.localdate().isoformat(),
                'comment': '',
            },
            HTTP_HOST='localhost',
        )
        group.refresh_from_db()
        self.assertEqual(bad_response.status_code, 200)
        self.assertContains(bad_response, 'нельзя сохранить')
        self.assertEqual(group.equipment.count(), 0)

        good_response = self.client.post(
            f'/system-admin/references/equipment-plan-groups/?edit={group.id}',
            {
                'record_id': str(group.id),
                'name': 'Экскаваторы 3000',
                'code': 'excavators_3000',
                'calculation_mode': PlanCalculationMode.VOLUME,
                'plan_value': '3000.00',
                'equipment': [str(excavator.id)],
                'is_active': 'on',
                'active_from': timezone.localdate().isoformat(),
                'comment': '',
            },
            HTTP_HOST='localhost',
        )
        group.refresh_from_db()
        reload_response = self.client.get(
            f'/system-admin/references/equipment-plan-groups/?edit={group.id}',
            HTTP_HOST='localhost',
        )

        self.assertEqual(good_response.status_code, 302)
        self.assertEqual(list(group.equipment.all()), [excavator])
        self.assertContains(reload_response, 'Экскаватор 2')

    def test_reference_detail_save_keeps_selected_record_and_filters(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        admin_employee = Employee.objects.create(full_name='Администратор MVP', status=Employee.Status.ACTIVE)
        EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='1000',
            status=EmployeeAccess.Status.ACTIVATED,
        )
        equipment_type = EquipmentType.objects.create(name='Самосвал')
        equipment = Equipment.objects.create(equipment_type=equipment_type, garage_number='A-101', is_active=True)

        self.client.post('/', {'access_code': '1000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.post(
            f'/system-admin/references/equipment/?q=A&status=active&edit={equipment.id}',
            {
                'action': 'save',
                'record_id': str(equipment.id),
                'equipment_type': str(equipment_type.id),
                'model': '',
                'garage_number': 'A-102',
                'vin': '',
                'is_own': 'on',
                'is_active': 'on',
            },
            HTTP_HOST='localhost',
        )
        equipment.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(equipment.garage_number, 'A-102')
        self.assertIn(f'/system-admin/references/equipment/?q=A&status=active&edit={equipment.id}', response['Location'])

    def test_admin_opens_conflicts_registry(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        admin_employee = Employee.objects.create(full_name='Администратор MVP', status=Employee.Status.ACTIVE)
        EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='1000',
            status=EmployeeAccess.Status.ACTIVATED,
        )
        employee = Employee.objects.create(full_name='Сотрудник с конфликтом')
        AdminConflict.objects.create(
            employee=employee,
            role=admin_role,
            conflict_type='Тестовый конфликт',
            process='Админка MVP',
            description='Проверка журнала конфликтов',
        )

        self.client.post('/', {'access_code': '1000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get('/system-admin/conflicts/', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Конфликты админки')
        self.assertContains(response, 'Тестовый конфликт')
        self.assertContains(response, 'Сотрудник с конфликтом')
        self.assertContains(response, 'Excel')

    def test_admin_updates_conflict_status(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        admin_employee = Employee.objects.create(full_name='Администратор MVP', status=Employee.Status.ACTIVE)
        EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='1000',
            status=EmployeeAccess.Status.ACTIVATED,
        )
        conflict = AdminConflict.objects.create(
            employee=admin_employee,
            role=admin_role,
            conflict_type='Проверка статуса',
            process='Админка MVP',
            description='Проверка смены статуса конфликта',
        )

        self.client.post('/', {'access_code': '1000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.post(
            f'/system-admin/conflicts/{conflict.id}/in-progress/',
            follow=True,
            HTTP_HOST='localhost',
        )
        conflict.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(conflict.status, AdminConflict.Status.IN_PROGRESS)
        self.assertEqual(conflict.resolved_by, admin_employee)
        self.assertIsNotNone(conflict.resolved_at)
        self.assertTrue(AdminActionLog.objects.filter(action='Изменен статус административного конфликта').exists())

    def test_admin_opens_action_log_registry(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        admin_employee = Employee.objects.create(full_name='Администратор MVP', status=Employee.Status.ACTIVE)
        EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='1000',
            status=EmployeeAccess.Status.ACTIVATED,
        )
        AdminActionLog.objects.create(
            actor=admin_employee,
            action='Тестовое действие',
            object_type='Employee',
            object_repr='Тестовый объект',
            comment='Проверка журнала действий',
        )

        self.client.post('/', {'access_code': '1000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get('/system-admin/logs/?q=Тестовое', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Журнал действий админки')
        self.assertContains(response, 'Тестовое действие')
        self.assertContains(response, 'Тестовый объект')
        self.assertContains(response, 'Excel')

    def test_admin_creates_employee_with_primary_pin_and_exports_accesses(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        driver_role = self.role
        admin_employee = Employee.objects.create(full_name='Администратор MVP', status=Employee.Status.ACTIVE)
        EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='1000',
            status=EmployeeAccess.Status.ACTIVATED,
        )

        self.client.post('/', {'access_code': '1000'}, follow=True, HTTP_HOST='localhost')
        create_response = self.client.post(
            '/system-admin/employees/create/',
            {
                'full_name': 'Новый водитель',
                'personnel_number': '001',
                'phone': '+79990000000',
                'status': Employee.Status.NOT_ACTIVATED,
                'comment': 'Первичная загрузка',
                'role': driver_role.id,
                'generate_access': 'on',
            },
            follow=True,
            HTTP_HOST='localhost',
        )

        employee = Employee.objects.get(full_name='Новый водитель')
        access = EmployeeAccess.objects.get(employee=employee)

        self.assertEqual(create_response.status_code, 200)
        self.assertEqual(access.role, driver_role)
        self.assertEqual(access.status, EmployeeAccess.Status.NOT_ACTIVATED)
        self.assertEqual(len(access.access_code), 6)
        self.assertTrue(access.access_code.isdigit())
        self.assertTrue(AdminActionLog.objects.filter(action='Создан сотрудник и выдан первичный пинкод').exists())
        self.assertNotContains(create_response, 'Сотрудник создан.')
        self.assertContains(create_response, 'Пин код:')

        save_response = self.client.post(
            f'/system-admin/employees/{employee.id}/',
            {
                'full_name': employee.full_name,
                'position': 'Водитель',
                'personnel_number': employee.personnel_number,
                'phone': employee.phone,
                'status': employee.status,
                'comment': 'Данные дополнены',
                'hired_at': '',
                'dismissed_at': '',
                'rotation': '',
                'residence_text': '',
                'hr_data': '',
            },
            follow=True,
            HTTP_HOST='localhost',
        )
        employee.refresh_from_db()

        self.assertEqual(save_response.status_code, 200)
        self.assertEqual(employee.comment, 'Данные дополнены')
        self.assertNotContains(save_response, 'Карточка сотрудника сохранена.')

        block_response = self.client.post(
            f'/system-admin/accesses/{access.id}/block/',
            {'reason': 'Проверка блокировки'},
            follow=True,
            HTTP_HOST='localhost',
        )
        access.refresh_from_db()

        self.assertEqual(block_response.status_code, 200)
        self.assertEqual(access.status, EmployeeAccess.Status.BLOCKED)
        self.assertFalse(access.is_active)

        export_response = self.client.get('/system-admin/export/accesses/', HTTP_HOST='localhost')
        self.assertEqual(export_response.status_code, 200)
        workbook = load_workbook(BytesIO(export_response.content))
        self.assertIn('Доступы', workbook.sheetnames)
        values = [cell.value for row in workbook['Доступы'].iter_rows() for cell in row]
        self.assertIn('Новый водитель', values)
        self.assertIn('Водитель самосвала', values)

    def test_employee_create_form_contains_position_shift_and_role_filtered_equipment(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        excavator_role = Role.objects.create(code='excavator_operator', name='Машинист экскаватора')
        admin_employee = Employee.objects.create(full_name='Администратор MVP', status=Employee.Status.ACTIVE)
        EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='1000',
            status=EmployeeAccess.Status.ACTIVATED,
        )
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10', is_active=True)
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1', is_active=True)

        self.client.post('/', {'access_code': '1000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get('/system-admin/employees/create/', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="position"', html=False)
        self.assertContains(response, 'name="role"', html=False)
        self.assertContains(response, 'name="assignment_shift_type"', html=False)
        self.assertContains(response, 'name="assignment_equipment"', html=False)
        self.assertContains(response, f'value="{self.role.id}" data-work-role="driver"', html=False)
        self.assertContains(response, f'value="{excavator_role.id}" data-work-role="excavator_operator"', html=False)
        self.assertContains(response, f'value="{truck.id}" data-base-label="Самосвал 10" data-work-role="driver"', html=False)
        self.assertContains(response, f'value="{excavator.id}" data-base-label="Экскаватор 1" data-work-role="excavator_operator"', html=False)
        self.assertContains(response, 'employee-work-assignment-meta')
        html = response.content.decode('utf-8')
        expected_order = [
            'name="full_name"',
            'name="birth_date"',
            'name="phone"',
            'name="personnel_position"',
            'name="personnel_department"',
            'name="base_specialization"',
            'id="employee-status-readonly"',
            'name="hired_at"',
            'name="dismissed_at"',
            'name="work_schedule"',
            'name="brigade_number"',
            'name="residence_text"',
            'name="role"',
            'name="assignment_shift_type"',
            'name="assignment_equipment"',
            'name="comment"',
            'name="hr_data"',
        ]
        positions = [html.index(marker) for marker in expected_order]
        self.assertEqual(positions, sorted(positions))

    def test_admin_creates_employee_with_work_shift_and_equipment_assignment(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        admin_employee = Employee.objects.create(full_name='Администратор MVP', status=Employee.Status.ACTIVE)
        EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='1000',
            status=EmployeeAccess.Status.ACTIVATED,
        )
        truck_type = EquipmentType.objects.create(name='Самосвал')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10', is_active=True)

        self.client.post('/', {'access_code': '1000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.post(
            '/system-admin/employees/create/',
            {
                'full_name': 'Водитель с назначением',
                'position': 'Водитель автомобиля',
                'phone': '+79990000001',
                'status': Employee.Status.NOT_ACTIVATED,
                'role': self.role.id,
                'assignment_shift_type': WorkShiftType.SHIFT_1,
                'assignment_equipment': truck.id,
            },
            follow=True,
            HTTP_HOST='localhost',
        )
        employee = Employee.objects.get(full_name='Водитель с назначением')
        access = EmployeeAccess.objects.get(employee=employee)
        assignment = EquipmentAssignment.objects.get(employee=employee, ended_at__isnull=True)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(employee.position, 'Водитель автомобиля')
        self.assertEqual(access.role, self.role)
        self.assertEqual(access.access_code, '')
        self.assertEqual(access.status, EmployeeAccess.Status.NOT_ACTIVATED)
        self.assertIsNone(access.primary_code_issued_at)
        self.assertEqual(assignment.role, self.role)
        self.assertEqual(assignment.shift_type, WorkShiftType.SHIFT_1)
        self.assertEqual(assignment.equipment, truck)
        self.assertEqual(assignment.status, AssignmentStatus.ACCEPTED)
        self.assertContains(response, 'Смена 1')
        self.assertContains(response, 'Самосвал 10')

    def test_employee_creation_rolls_back_when_equipment_becomes_occupied(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        admin_employee = Employee.objects.create(full_name='Администратор MVP', status=Employee.Status.ACTIVE)
        occupied_employee = Employee.objects.create(full_name='Занявший технику', status=Employee.Status.ACTIVE)
        EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='1000',
            status=EmployeeAccess.Status.ACTIVATED,
        )
        EmployeeAccess.objects.create(
            employee=occupied_employee,
            role=self.role,
            access_code='2000',
            status=EmployeeAccess.Status.ACTIVATED,
        )
        truck_type = EquipmentType.objects.create(name='Самосвал')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10', is_active=True)
        EquipmentAssignment.objects.create(
            employee=occupied_employee,
            role=self.role,
            equipment=truck,
            shift_type=WorkShiftType.SHIFT_1,
            assigned_by=admin_employee,
            status=AssignmentStatus.ACCEPTED,
            accepted_at=timezone.now(),
        )

        self.client.post('/', {'access_code': '1000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.post(
            '/system-admin/employees/create/',
            {
                'full_name': 'Конфликтующий водитель',
                'position': 'Водитель автомобиля',
                'phone': '+79990000002',
                'status': Employee.Status.NOT_ACTIVATED,
                'role': self.role.id,
                'assignment_shift_type': WorkShiftType.SHIFT_1,
                'assignment_equipment': truck.id,
                'generate_access': 'on',
            },
            HTTP_HOST='localhost',
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Employee.objects.filter(full_name='Конфликтующий водитель').exists())
        self.assertContains(response, 'Эта техника уже назначена другому сотруднику в выбранной смене.')
        self.assertContains(response, 'data-busy-day="Занявший технику"', html=False)

    def test_primary_pin_requires_activation_and_becomes_invalid(self):
        driver_role = self.role
        employee = Employee.objects.create(full_name='Водитель с первичным пинкодом', phone='+79000001111')
        access = EmployeeAccess.objects.create(
            employee=employee,
            role=driver_role,
            access_code='246824',
            status=EmployeeAccess.Status.NOT_ACTIVATED,
            primary_code_issued_at=timezone.now(),
        )

        login_response = self.client.post(
            '/',
            {'phone': '+79000001111', 'access_code': '246824'},
            follow=True,
            HTTP_HOST='localhost',
        )
        self.assertRedirects(login_response, '/activate-access/', target_status_code=200)
        self.assertContains(login_response, 'Это вы?')
        self.assertContains(
            login_response,
            'придумайте пинкод, с ним будете заходить дальше',
        )
        self.assertContains(login_response, '+7 ••• •••-11-11')
        self.assertNotContains(login_response, 'name="phone"')
        self.assertContains(login_response, 'name="new_access_code"')
        self.assertContains(login_response, 'maxlength="6"', count=2)

        activation_response = self.client.post(
            '/activate-access/',
            {'new_access_code': '864286', 'confirm_access_code': '864286'},
            follow=True,
            HTTP_HOST='localhost',
        )
        access.refresh_from_db()
        employee.refresh_from_db()

        self.assertEqual(activation_response.status_code, 200)
        self.assertEqual(access.access_code, '864286')
        self.assertEqual(access.status, EmployeeAccess.Status.ACTIVATED)
        self.assertEqual(employee.status, Employee.Status.ACTIVE)
        self.assertIsNone(EmployeeAccess.objects.filter(access_code='246824').first())

        self.client.get('/logout/', follow=True, HTTP_HOST='localhost')
        old_code_response = self.client.post('/', {'access_code': '246824'}, follow=True, HTTP_HOST='localhost')
        self.assertContains(old_code_response, 'Телефон или пинкод указаны неверно.')

        new_code_response = self.client.post(
            '/',
            {'phone': '+79000001111', 'access_code': access.access_code},
            follow=True,
            HTTP_HOST='localhost',
        )
        self.assertEqual(new_code_response.status_code, 200)
        self.assertEqual(self.client.session.get('employee_access_id'), access.id)

    def test_activation_allows_same_pin_for_different_phone_numbers(self):
        driver_role = self.role
        first_employee = Employee.objects.create(full_name='Водитель с постоянным пинкодом', phone='+79000001111')
        EmployeeAccess.objects.create(
            employee=first_employee,
            role=driver_role,
            access_code='864286',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        second_employee = Employee.objects.create(full_name='Водитель с первичным пинкодом', phone='+79000002222')
        EmployeeAccess.objects.create(
            employee=second_employee,
            role=driver_role,
            access_code='246824',
            status=EmployeeAccess.Status.NOT_ACTIVATED,
            primary_code_issued_at=timezone.now(),
            is_active=True,
        )

        self.client.post('/', {'phone': '+79000002222', 'access_code': '246824'}, follow=True, HTTP_HOST='localhost')
        response = self.client.post(
            '/activate-access/',
            {'new_access_code': '864286', 'confirm_access_code': '864286'},
            follow=True,
            HTTP_HOST='localhost',
        )
        second_access = EmployeeAccess.objects.get(employee=second_employee)

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'Этот пинкод нельзя использовать')
        self.assertNotContains(response, 'Такой пинкод уже используется')
        self.assertEqual(second_access.access_code, '864286')
        self.assertEqual(second_access.status, EmployeeAccess.Status.ACTIVATED)

    def test_activation_rejects_temporary_and_weak_pin(self):
        employee = Employee.objects.create(full_name='Водитель с временным PIN', phone='+79000003333')
        access = EmployeeAccess.objects.create(
            employee=employee,
            role=self.role,
            access_code='246824',
            status=EmployeeAccess.Status.NOT_ACTIVATED,
            primary_code_issued_at=timezone.now(),
        )

        self.client.post('/', {'phone': '+79000003333', 'access_code': '246824'}, HTTP_HOST='localhost')

        same_pin_response = self.client.post(
            '/activate-access/',
            {'new_access_code': '246824', 'confirm_access_code': '246824'},
            HTTP_HOST='localhost',
        )
        self.assertEqual(same_pin_response.status_code, 200)
        self.assertContains(same_pin_response, 'Новый PIN не должен совпадать с временным.')

        weak_pin_response = self.client.post(
            '/activate-access/',
            {'new_access_code': '111111', 'confirm_access_code': '111111'},
            HTTP_HOST='localhost',
        )
        self.assertEqual(weak_pin_response.status_code, 200)
        self.assertContains(weak_pin_response, 'Этот PIN слишком простой.')

        access.refresh_from_db()
        self.assertEqual(access.access_code, '246824')
        self.assertEqual(access.status, EmployeeAccess.Status.NOT_ACTIVATED)

    def test_activation_rejects_invalid_pin_length_and_mismatch(self):
        employee = Employee.objects.create(full_name='Водитель с ошибкой PIN', phone='+79000004444')
        access = EmployeeAccess.objects.create(
            employee=employee,
            role=self.role,
            access_code='246824',
            status=EmployeeAccess.Status.NOT_ACTIVATED,
            primary_code_issued_at=timezone.now(),
        )

        self.client.post('/', {'phone': '+79000004444', 'access_code': '246824'}, HTTP_HOST='localhost')
        invalid_response = self.client.post(
            '/activate-access/',
            {'new_access_code': '8642867', 'confirm_access_code': '864286'},
            HTTP_HOST='localhost',
        )
        self.assertEqual(invalid_response.status_code, 200)
        self.assertContains(invalid_response, 'Убедитесь, что это значение содержит не более 6 символов')

        mismatch_response = self.client.post(
            '/activate-access/',
            {'new_access_code': '864286', 'confirm_access_code': '864287'},
            HTTP_HOST='localhost',
        )
        self.assertEqual(mismatch_response.status_code, 200)
        self.assertContains(mismatch_response, 'PIN-коды не совпадают.')

        access.refresh_from_db()
        self.assertEqual(access.status, EmployeeAccess.Status.NOT_ACTIVATED)

    def test_admin_can_delete_employee_without_production_history(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        admin_employee = Employee.objects.create(full_name='Администратор MVP', status=Employee.Status.ACTIVE)
        EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='1000',
            status=EmployeeAccess.Status.ACTIVATED,
        )
        employee = Employee.objects.create(full_name='Сотрудник без истории')

        self.client.post('/', {'access_code': '1000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.post(
            f'/system-admin/employees/{employee.id}/delete/',
            follow=True,
            HTTP_HOST='localhost',
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Employee.objects.filter(id=employee.id).exists())
        self.assertContains(response, 'удален')

    def test_admin_cannot_delete_employee_with_production_history(self):
        admin_role = Role.objects.create(code='admin', name='Администратор')
        admin_employee = Employee.objects.create(full_name='Администратор MVP', status=Employee.Status.ACTIVE)
        EmployeeAccess.objects.create(
            employee=admin_employee,
            role=admin_role,
            access_code='1000',
            status=EmployeeAccess.Status.ACTIVATED,
        )
        employee = Employee.objects.create(full_name='Сотрудник с историей')
        EmployeeShift.objects.create(
            employee=employee,
            shift_type='day',
            opened_at=timezone.now() - timedelta(hours=1),
            closed_at=timezone.now(),
        )

        self.client.post('/', {'access_code': '1000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.post(
            f'/system-admin/employees/{employee.id}/delete/',
            follow=True,
            HTTP_HOST='localhost',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(Employee.objects.filter(id=employee.id).exists())
        self.assertTrue(AdminConflict.objects.filter(employee=employee, conflict_type='Попытка удаления сотрудника с историей').exists())
        self.assertContains(response, 'Удаление запрещено')

    def test_wrong_access_code_stays_on_login(self):
        response = self.client.post('/', {'access_code': 'wrong'}, follow=True, HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Телефон или пинкод указаны неверно.')
        self.assertIsNone(self.client.session.get('employee_access_id'))
        self.assertContains(response, 'login-page')

    def test_interface_map_opens_without_login(self):
        response = self.client.get('/interfaces/', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Назад')
        self.assertContains(response, 'Главная')
        self.assertContains(response, 'Карта интерфейсов')
        self.assertContains(response, 'Карта интерфейсов MVP')
        self.assertContains(response, '/reports/volume/')
        self.assertContains(response, '/reports/templates/')
        self.assertContains(response, '/reports/management/')
        self.assertContains(response, '/reports/management/export/')
        self.assertContains(response, '/reports/pilot-checklist/')
        self.assertContains(response, '/reports/pilot-scenario/')
        self.assertContains(response, '/reports/pilot-feedback/')
        self.assertContains(response, '/system-admin/employees/')
        self.assertContains(response, '/system-admin/references/')
        self.assertContains(response, '/system-admin/conflicts/')
        self.assertContains(response, '/system-admin/logs/')
        self.assertContains(response, 'Excel-выгрузка витрины руководства')
        self.assertContains(response, 'Чеклист пилотной проверки отчетов')
        self.assertContains(response, 'Сценарий пилотного запуска')
        self.assertContains(response, 'Журнал замечаний пилота')
        self.assertContains(response, '6000')

    def test_manager_can_open_pilot_report_checklist(self):
        manager_role = Role.objects.create(code='manager', name='Руководство')
        manager = Employee.objects.create(full_name='Тестовое руководство')
        EmployeeAccess.objects.create(employee=manager, role=manager_role, access_code='6000')

        self.client.post('/', {'access_code': '6000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get('/reports/pilot-checklist/', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Чеклист пилотной проверки отчетов')
        self.assertContains(response, '9 из 10')
        self.assertContains(response, '99%')
        self.assertContains(response, 'Сверка со старыми Excel-формами')
        self.assertContains(response, 'Отчет_Коппер. Рисорсез_Март.xlsx')
        self.assertContains(response, 'почасовой Март.xlsx')
        self.assertContains(response, '/reports/volume/?group_by=completed_hour')
        self.assertContains(response, 'ОР ККД СКДР март.xlsx')
        self.assertContains(response, 'удельный_веса_руд_и_пород_Малмыжского_местородения.xlsx')
        self.assertContains(response, '/admin/references/rocktype/')
        self.assertContains(response, 'КИП/КТГ и КИО/КТГ')
        self.assertContains(response, '/reports/management/')
        self.assertContains(response, '/reports/management/export/')
        self.assertContains(response, '/dispatcher/control/')
        self.assertContains(response, '/reports/volume/')
        self.assertContains(response, '/reports/volume/export/')
        self.assertContains(response, '/reports/templates/')
        self.assertContains(response, '/reports/customer-daily/')
        self.assertContains(response, '/reports/customer-daily/export/')
        self.assertContains(response, '/reports/downtimes/')
        self.assertContains(response, '/reports/downtimes/export/')
        self.assertContains(response, '/reports/pilot-scenario/')
        self.assertContains(response, '/reports/pilot-feedback/')

    def test_manager_can_open_pilot_launch_scenario(self):
        manager_role = Role.objects.create(code='manager', name='Руководство')
        manager = Employee.objects.create(full_name='Тестовое руководство')
        EmployeeAccess.objects.create(employee=manager, role=manager_role, access_code='6000')

        self.client.post('/', {'access_code': '6000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get('/reports/pilot-scenario/', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Сценарий пилотного запуска')
        self.assertContains(response, '9 из 10')
        self.assertContains(response, '99%')
        self.assertContains(response, 'Расстановка техники')
        self.assertContains(response, 'Работа водителя')
        self.assertContains(response, 'Диспетчерский контроль')
        self.assertContains(response, 'Вопросы для фиксации во время пилота')
        self.assertContains(response, '/reports/pilot-feedback/')
        self.assertContains(response, '31_ЖУРНАЛ_ЗАМЕЧАНИЙ_ПИЛОТА.md')

    def test_manager_can_create_pilot_feedback_and_export_it(self):
        manager_role = Role.objects.create(code='manager', name='Руководство')
        manager = Employee.objects.create(full_name='Тестовое руководство')
        EmployeeAccess.objects.create(employee=manager, role=manager_role, access_code='6000')

        self.client.post('/', {'access_code': '6000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.post(
            '/reports/pilot-feedback/',
            {
                'title': 'Не хватает столбца для сверки',
                'category': 'report',
                'priority': 'p1',
                'status': 'new',
                'screen': 'Суточный отчет',
                'description': 'На пилоте нужно сверить старую форму заказчика.',
                'decision': 'Добавить в список доработок после проверки.',
            },
            follow=True,
            HTTP_HOST='localhost',
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Журнал замечаний пилота')
        self.assertContains(response, 'Не хватает столбца для сверки')
        self.assertContains(response, 'P1 - исправить до запуска')
        self.assertContains(response, 'В работу')
        self.assertContains(response, 'Решено')
        self.assertContains(response, 'Отклонено')
        self.assertEqual(PilotFeedback.objects.count(), 1)
        feedback = PilotFeedback.objects.first()
        self.assertEqual(feedback.created_by, manager)

        status_response = self.client.post(
            '/reports/pilot-feedback/',
            {
                'action': 'change_status',
                'feedback_id': str(feedback.id),
                'status': 'decided',
            },
            follow=True,
            HTTP_HOST='localhost',
        )

        self.assertEqual(status_response.status_code, 200)
        feedback.refresh_from_db()
        self.assertEqual(feedback.status, 'decided')
        self.assertContains(status_response, 'Решение принято')

        export_response = self.client.get('/reports/pilot-feedback/export/', HTTP_HOST='localhost')

        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(
            export_response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        workbook = load_workbook(BytesIO(export_response.content))
        self.assertIn('Замечания пилота', workbook.sheetnames)
        sheet = workbook['Замечания пилота']
        self.assertEqual(sheet['A1'].value, 'Журнал замечаний пилотного запуска')
        self.assertEqual(sheet['F5'].value, 'Не хватает столбца для сверки')

    def test_driver_goes_straight_to_work_without_housing_registration(self):
        """Экран «Первичная регистрация водителя» больше не заслоняет работу.

        Он спрашивал секцию проживания, а её в системе никто не читал: данные
        писались один раз и нигде не использовались. Расселение будет вести
        делопроизводитель отдельно, так что держать полсотни водителей на этом
        экране незачем.
        """
        truck_type = EquipmentType.objects.create(name='Самосвал')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        self.assign_driver_work(truck)

        login_response = self.client.post('/', {'access_code': '2000'}, follow=True, HTTP_HOST='localhost')

        self.assertEqual(login_response.status_code, 200)
        self.assertNotContains(login_response, 'Первичная регистрация водителя')
        self.assertContains(login_response, '>Начать смену<')

    def test_driver_can_open_shift_after_registration(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        truck_model = EquipmentModel.objects.create(
            equipment_type=truck_type,
            name='БелАЗ тестовый',
            fuel_capacity_limit_l='2000',
        )
        truck = Equipment.objects.create(equipment_type=truck_type, model=truck_model, garage_number='10')
        group = EquipmentPlanGroup.objects.create(
            name='Самосвалы БелАЗ водитель',
            code='belaz-driver-start-test',
            calculation_mode=PlanCalculationMode.TRIPS,
            plan_value='18.00',
            is_active=True,
            active_from=production_work_date(),
        )
        group.equipment.add(truck)
        EquipmentAssignment.objects.create(
            employee=self.employee,
            role=self.role,
            equipment=truck,
            shift_type='day',
            assigned_by=self.employee,
            status=AssignmentStatus.ACCEPTED,
            accepted_at=timezone.now(),
        )
        dormitory = Dormitory.objects.create(number='5')
        block = DormitoryBlock.objects.create(dormitory=dormitory, name='Блок 1')
        section = DormitorySection.objects.create(block=block, name='А')

        self.client.post('/', {'access_code': '2000'}, follow=True, HTTP_HOST='localhost')
        self.client.post(
            '/driver/registration/',
            {'dormitory_section': section.id},
            follow=True,
            HTTP_HOST='localhost',
        )
        response = self.client.post(
            '/driver/shift/',
            {'shift_type': 'day', 'truck': truck.id, 'start_fuel': '100', 'start_mileage': '2500', 'start_engine_hours': '700'},
            follow=True,
            HTTP_HOST='localhost',
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Смена открыта')
        shift = self.employee.employeeshift_set.get(closed_at__isnull=True)
        self.assertEqual(shift.plan_group_name, 'Самосвалы БелАЗ водитель')
        self.assertEqual(shift.plan_status, PlanAssignmentStatus.ASSIGNED)
        self.assertEqual(shift.plan_calculation_mode, PlanCalculationMode.TRIPS)
        self.assertEqual(shift.plan_value, Decimal('18.00'))
        self.assertFalse(ShiftPlan.objects.exists())

    def test_driver_can_close_shift_and_next_opening_uses_last_end_values(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        dormitory = Dormitory.objects.create(number='5')
        block = DormitoryBlock.objects.create(dormitory=dormitory, name='Блок 1')
        section = DormitorySection.objects.create(block=block, name='А')
        self.assign_driver_work(truck)

        self.client.post('/', {'access_code': '2000'}, follow=True, HTTP_HOST='localhost')
        self.client.post(
            '/driver/registration/',
            {'dormitory_section': section.id},
            follow=True,
            HTTP_HOST='localhost',
        )
        self.client.post(
            '/driver/shift/',
            {'shift_type': 'day', 'truck': truck.id, 'start_fuel': '100', 'start_mileage': '2500', 'start_engine_hours': '700'},
            follow=True,
            HTTP_HOST='localhost',
        )

        close_response = self.client.post(
            '/driver/shift/close/',
            {'end_fuel': '90', 'end_mileage': '2600', 'end_engine_hours': '712', 'client_action_id': 'close-handoff'},
            follow=True,
            HTTP_HOST='localhost',
        )
        shift = EmployeeShift.objects.get(employee=self.employee)

        self.assertEqual(close_response.status_code, 200)
        self.assertContains(close_response, 'Смена закрыта')
        self.assertTrue(close_response.redirect_chain)
        self.assertEqual(close_response.redirect_chain[-1][0], '/driver/?tab=manifest')
        self.assertContains(close_response, 'data-driver-tab-panel="manifest"')
        self.assertContains(close_response, 'driver-panel driver-tab is-active')
        self.assertIsNotNone(shift.closed_at)
        self.assertEqual(shift.end_fuel, 90)
        self.assertEqual(shift.end_mileage, 2600)
        self.assertEqual(shift.end_engine_hours, 712)

        next_open_response = self.client.get(f'/driver/shift/?truck={truck.id}', HTTP_HOST='localhost')
        self.assertContains(next_open_response, 'value="90')
        self.assertContains(next_open_response, 'value="2600')
        self.assertContains(next_open_response, 'value="712')

    def test_driver_manifest_keeps_last_closed_shift_report(self):
        truck = self.create_registered_driver_shift()
        shift = EmployeeShift.objects.get(employee=self.employee, closed_at__isnull=True)
        now = timezone.now()
        shift.opened_at = now - timedelta(hours=2)
        shift.closed_at = now
        shift.end_fuel = Decimal('90.00')
        shift.end_mileage = Decimal('2600.00')
        shift.end_engine_hours = Decimal('712.00')
        shift.closed_by = self.employee
        shift.save(update_fields=[
            'opened_at',
            'closed_at',
            'end_fuel',
            'end_mileage',
            'end_engine_hours',
            'closed_by',
        ])

        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='7')
        rock = RockType.objects.create(name='Скальная масса')
        dump_point = DumpPoint.objects.create(name='Отвал 3')
        trip = Trip.objects.create(
            excavator=excavator,
            truck=truck,
            driver=self.employee,
            loading_shift=shift,
            unloading_shift=shift,
            rock_type=rock,
            dump_point=dump_point,
            status=TripStatus.COMPLETED,
            completed_at=now - timedelta(minutes=30),
        )
        Trip.objects.filter(pk=trip.pk).update(created_at=now - timedelta(hours=1))
        reason = DowntimeReason.objects.create(
            name='Тест заправка после закрытия',
            short_label='Заправка',
            show_for_truck_driver=True,
        )
        DowntimeEvent.objects.create(
            equipment=truck,
            employee=self.employee,
            reason=reason,
            started_at=now - timedelta(minutes=25),
            ended_at=now - timedelta(minutes=15),
        )

        response = self.client.get('/driver/?tab=manifest', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-driver-tab-panel="manifest"')
        self.assertContains(response, 'data-driver-report-truck="10"')
        self.assertContains(response, 'data-driver-report-trip-total="1"')
        self.assertContains(response, 'data-driver-report-downtime-total="10 мин."')
        self.assertContains(response, 'data-driver-report-end-fuel="90')
        self.assertContains(response, 'data-driver-report-end-mileage="2600')
        self.assertContains(response, 'data-driver-report-end-engine-hours="712')
        self.assertContains(response, 'data-excavator="7" data-dump-point="Отвал 3"')
        self.assertContains(response, 'Заправка')

    def test_driver_sees_assigned_excavator_without_accept_action(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        dormitory = Dormitory.objects.create(number='5')
        block = DormitoryBlock.objects.create(dormitory=dormitory, name='Блок 1')
        section = DormitorySection.objects.create(block=block, name='А')
        self.assign_driver_work(truck)

        self.client.post('/', {'access_code': '2000'}, follow=True, HTTP_HOST='localhost')
        self.client.post(
            '/driver/registration/',
            {'dormitory_section': section.id},
            follow=True,
            HTTP_HOST='localhost',
        )
        self.client.post(
            '/driver/shift/',
            {'shift_type': 'day', 'truck': truck.id, 'start_fuel': '100', 'start_mileage': '2500', 'start_engine_hours': '700'},
            follow=True,
            HTTP_HOST='localhost',
        )
        assignment = HaulAssignment.objects.create(truck=truck, excavator=excavator)

        shift_response = self.client.get('/driver/shift/', HTTP_HOST='localhost')
        self.assertContains(shift_response, 'driver-work-context-card')
        self.assertContains(shift_response, 'ВЫ НАЗНАЧЕНЫ НА ЭКС-1')
        self.assertContains(shift_response, 'ПРИНЯТЬ')
        self.assertContains(shift_response, 'НА ЗАГРУЗКУ')
        self.assertContains(shift_response, '1')
        self.assertContains(shift_response, 'driver-work-dial-button is-empty')
        forbidden_driver_labels = (
            'НОВОЕ ' + 'НАЗНАЧЕНИЕ',
            'принять ' + 'назначение',
        )
        for label in forbidden_driver_labels:
            self.assertNotContains(shift_response, label)

        accept_response = self.client.post(
            f'/driver/assignment/{assignment.id}/accept/',
            HTTP_HOST='localhost',
        )
        assignment.refresh_from_db()

        self.assertEqual(accept_response.status_code, 302)
        self.assertEqual(assignment.status, AssignmentStatus.ACCEPTED)
        self.assertIsNotNone(assignment.accepted_at)
        accepted_response = self.client.get('/driver/shift/', HTTP_HOST='localhost')
        self.assertContains(accepted_response, 'К-1')
        self.assertContains(accepted_response, 'ЭКС-1')
        self.assertNotContains(accepted_response, 'ВЫ НАЗНАЧЕНЫ НА ЭКС-1')

    def test_driver_receives_dispatcher_assignment_excavator_context_and_loaded_trip_chain(self):
        self.employee.full_name = 'Петров Петр Петрович'
        self.employee.save(update_fields=['full_name'])
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck_model = EquipmentModel.objects.create(
            equipment_type=truck_type,
            name='БелАЗ 7513D',
        )
        truck = Equipment.objects.create(
            equipment_type=truck_type,
            model=truck_model,
            garage_number='54',
        )
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        rock = RockType.objects.create(name='Руда', density=Decimal('2.6000'))
        dump_point = DumpPoint.objects.create(name='ККД')
        truck = self.create_registered_driver_shift(truck=truck)
        TruckCapacityRule.objects.create(
            equipment_model=truck.model,
            rock_type=rock,
            volume_m3=Decimal('49.40'),
        )

        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(
            full_name='Тестовый диспетчер',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        dispatcher_access = EmployeeAccess.objects.create(
            employee=dispatcher,
            role=dispatcher_role,
            access_code='5000',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        EmployeeShift.objects.create(
            employee=dispatcher,
            shift_type='day',
            opened_at=timezone.now(),
            opened_by=dispatcher,
        )
        dispatcher_client = self.client_class(HTTP_HOST='localhost')
        dispatcher_session = dispatcher_client.session
        dispatcher_session['employee_access_id'] = dispatcher_access.id
        dispatcher_session.save()

        assign_response = dispatcher_client.post(
            '/dispatcher/control/truck/assign/',
            data=json.dumps({
                'action': 'assign',
                'truck_id': truck.id,
                'excavator_id': excavator.id,
            }),
            content_type='application/json',
            HTTP_HOST='localhost',
        )
        assignment = HaulAssignment.objects.get(truck=truck, excavator=excavator)
        initial_driver_response = self.client.get('/driver/', HTTP_HOST='localhost')

        self.assertEqual(assign_response.status_code, 200)
        self.assertTrue(assign_response.json()['ok'])
        self.assertEqual(assignment.status, AssignmentStatus.PENDING)
        self.assertEqual(assignment.truck, truck)
        self.assertContains(initial_driver_response, 'Самосвал 54 · Петров П.П.')
        self.assertContains(initial_driver_response, 'ВЫ НАЗНАЧЕНЫ НА ЭКС-1')
        self.assertContains(initial_driver_response, 'НА ЗАГРУЗКУ')
        self.assertContains(initial_driver_response, 'ПРИНЯТЬ')

        accept_response = self.client.post(
            reverse('driver_accept_assignment', args=[assignment.id]),
            HTTP_HOST='localhost',
        )
        assignment.refresh_from_db()
        self.assertEqual(accept_response.status_code, 302)
        self.assertEqual(assignment.status, AssignmentStatus.ACCEPTED)

        excavator_role = Role.objects.create(code='excavator_operator', name='Машинист экскаватора')
        excavator_operator = Employee.objects.create(
            full_name='Тестовый машинист',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        excavator_access = EmployeeAccess.objects.create(
            employee=excavator_operator,
            role=excavator_role,
            access_code='3000',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        EmployeeShift.objects.create(
            employee=excavator_operator,
            shift_type='day',
            equipment=excavator,
            opened_at=timezone.now(),
            opened_by=excavator_operator,
        )
        operator_client = self.client_class(HTTP_HOST='localhost')
        operator_session = operator_client.session
        operator_session['employee_access_id'] = excavator_access.id
        operator_session.save()

        settings_response = operator_client.post(
            reverse('excavator_work_settings'),
            data=json.dumps({
                'client_action_id': 'driver-chain-settings-1',
                'rock_type_id': rock.id,
                'dump_point_ids': [dump_point.id],
                'loading_horizon': '75',
                'loading_block': '52',
            }),
            content_type='application/json',
            HTTP_HOST='localhost',
        )
        context_driver_response = self.client.get('/driver/', HTTP_HOST='localhost')

        self.assertEqual(settings_response.status_code, 200)
        self.assertContains(context_driver_response, 'ЭКС-1')
        self.assertContains(context_driver_response, 'Комплекс К-1')
        self.assertContains(context_driver_response, 'Горизонт 75')
        self.assertContains(context_driver_response, 'Блок 52')
        self.assertContains(context_driver_response, 'Руда')
        self.assertContains(context_driver_response, 'driver-work-dial-button is-empty')

        load_response = operator_client.post(
            reverse('excavator_truck_loaded'),
            data=json.dumps({
                'client_action_id': 'driver-chain-truck-loaded-1',
                'truck_id': truck.id,
                'excavator_id': excavator.id,
                'dump_point_id': dump_point.id,
                'rock_type_id': rock.id,
                'loading_horizon': '75',
                'loading_block': '52',
            }),
            content_type='application/json',
            HTTP_HOST='localhost',
        )
        assignment.refresh_from_db()
        trip = Trip.objects.get(truck=truck, excavator=excavator)
        loaded_driver_response = self.client.get('/driver/', HTTP_HOST='localhost')

        self.assertEqual(load_response.status_code, 200)
        self.assertEqual(load_response.json()['action'], 'truck_loaded')
        self.assertEqual(assignment.status, AssignmentStatus.ACCEPTED)
        self.assertEqual(trip.status, TripStatus.LOADED_WAITING_UNLOAD)
        self.assertEqual(trip.truck, truck)
        self.assertContains(loaded_driver_response, 'ККД')
        self.assertContains(loaded_driver_response, 'ТОЧКА РАЗГРУЗКИ')
        self.assertContains(loaded_driver_response, 'driver-work-dial-button is-loaded')
        self.assertContains(loaded_driver_response, 'ЭКС-1')
        self.assertContains(loaded_driver_response, 'Комплекс К-1')
        self.assertContains(loaded_driver_response, 'Горизонт 75')
        self.assertContains(loaded_driver_response, 'Блок 52')
        self.assertContains(loaded_driver_response, 'Руда')

        complete_response = self.client.post(
            f'/driver/trip/{trip.id}/complete/',
            {'client_action_id': 'driver-chain-trip-unloaded-1'},
            follow=True,
            HTTP_HOST='localhost',
        )
        trip.refresh_from_db()
        empty_driver_response = self.client.get('/driver/', HTTP_HOST='localhost')

        self.assertEqual(complete_response.status_code, 200)
        self.assertEqual(trip.status, TripStatus.COMPLETED)
        self.assertEqual(trip.driver, self.employee)
        self.assertContains(empty_driver_response, 'ЭКС-1')
        self.assertContains(empty_driver_response, 'НА ЗАГРУЗКУ')
        self.assertContains(empty_driver_response, 'driver-work-dial-button is-empty')
        self.assertContains(empty_driver_response, 'ЭКС-1')
        self.assertContains(empty_driver_response, 'Комплекс К-1')
        self.assertContains(empty_driver_response, 'Горизонт 75')
        self.assertContains(empty_driver_response, 'Блок 52')
        self.assertContains(empty_driver_response, 'Руда')
        self.assertTrue(
            OperationalStateEvent.objects.filter(
                event_type='assignment_changed',
                object_type='HaulAssignment',
                object_id=str(assignment.id),
            ).exists()
        )
        self.assertTrue(
            OperationalStateEvent.objects.filter(
                event_type='trip_changed',
                object_type='Trip',
                object_id=str(trip.id),
                payload__action='truck_loaded',
            ).exists()
        )

    def test_driver_shows_reassignment_button_only_for_other_pending_complex(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        excavator_one = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        excavator_two = Equipment.objects.create(equipment_type=excavator_type, garage_number='2')
        dormitory = Dormitory.objects.create(number='5')
        block = DormitoryBlock.objects.create(dormitory=dormitory, name='Блок 1')
        section = DormitorySection.objects.create(block=block, name='А')
        self.assign_driver_work(truck)

        self.client.post('/', {'access_code': '2000'}, follow=True, HTTP_HOST='localhost')
        self.client.post(
            '/driver/registration/',
            {'dormitory_section': section.id},
            follow=True,
            HTTP_HOST='localhost',
        )
        self.client.post(
            '/driver/shift/',
            {'shift_type': 'day', 'truck': truck.id, 'start_fuel': '100', 'start_mileage': '2500', 'start_engine_hours': '700'},
            follow=True,
            HTTP_HOST='localhost',
        )
        HaulAssignment.objects.create(
            truck=truck,
            excavator=excavator_one,
            status=AssignmentStatus.ACCEPTED,
            accepted_at=timezone.now() - timedelta(minutes=20),
        )
        pending = HaulAssignment.objects.create(
            truck=truck,
            excavator=excavator_two,
            status=AssignmentStatus.PENDING,
        )
        HaulAssignment.objects.filter(id=pending.id).update(assigned_at=timezone.now() - timedelta(seconds=1))

        shift_response = self.client.get('/driver/shift/', HTTP_HOST='localhost')

        self.assertContains(shift_response, 'К-1')
        self.assertContains(shift_response, 'ЭКС-1')
        self.assertContains(shift_response, 'ВЫ НАЗНАЧЕНЫ НА ЭКС-2')
        self.assertContains(shift_response, 'ПРИНЯТЬ')
        self.assertNotContains(shift_response, '668:')

    def test_excavator_creates_trip_and_driver_completes_it(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        rock = RockType.objects.create(name='Руда', density=Decimal('2.6000'))
        dump_point = DumpPoint.objects.create(name='ККД')
        skdr_point = DumpPoint.objects.create(name='СКДР')
        dormitory = Dormitory.objects.create(number='5')
        block = DormitoryBlock.objects.create(dormitory=dormitory, name='Блок 1')
        section = DormitorySection.objects.create(block=block, name='А')
        excavator_role = Role.objects.create(code='excavator_operator', name='Машинист экскаватора')
        excavator_operator = Employee.objects.create(full_name='Тестовый машинист')
        EmployeeAccess.objects.create(employee=excavator_operator, role=excavator_role, access_code='3000')
        operator_shift = EmployeeShift.objects.create(
            employee=excavator_operator,
            shift_type='day',
            equipment=excavator,
            opened_at=timezone.now(),
        )
        self.assign_driver_work(truck)
        TruckCapacityRule.objects.create(
            equipment_model=truck.model,
            rock_type=rock,
            volume_m3=Decimal('49.40'),
        )

        driver_client = self.client
        driver_client.post('/', {'access_code': '2000'}, follow=True, HTTP_HOST='localhost')
        driver_client.post(
            '/driver/registration/',
            {'dormitory_section': section.id},
            follow=True,
            HTTP_HOST='localhost',
        )
        driver_client.post(
            '/driver/shift/',
            {'shift_type': 'day', 'truck': truck.id, 'start_fuel': '100', 'start_mileage': '2500', 'start_engine_hours': '700'},
            follow=True,
            HTTP_HOST='localhost',
        )
        assignment = HaulAssignment.objects.create(truck=truck, excavator=excavator, status=AssignmentStatus.ACCEPTED)

        operator_client = self.client_class(HTTP_HOST='localhost')
        operator_client.post('/', {'access_code': '3000'}, follow=True, HTTP_HOST='localhost')
        trip_response = operator_client.post(
            '/excavator/work/',
            {
                'client_action_id': 'legacy-trip-create-1',
                'assignment': assignment.id,
                'rock_type': rock.id,
                'dump_point': dump_point.id,
                'planned_volume_m3': '7000',
                'loading_horizon': '75',
                'loading_block': '52',
                'transport_distance_km': '3.10',
                'downtime_text': 'зачистка забоя',
                'note': 'проверка параметров отчета',
            },
            follow=True,
            HTTP_HOST='localhost',
        )
        self.assertEqual(trip_response.status_code, 200)
        trip = Trip.objects.get()
        self.assertEqual(trip.status, TripStatus.LOADED_WAITING_UNLOAD)
        self.assertEqual(trip.volume_m3, Decimal('49.40'))
        self.assertEqual(trip.tonnage, Decimal('128.44'))
        self.assertEqual(trip.loading_shift, operator_shift)
        self.assertEqual(trip.planned_volume_m3, Decimal('7000.00'))
        self.assertEqual(trip.loading_horizon, '75')
        self.assertEqual(trip.loading_block, '52')
        self.assertEqual(trip.transport_distance_km, Decimal('3.10'))
        self.assertEqual(trip.downtime_text, 'зачистка забоя')
        self.assertEqual(trip.note, 'проверка параметров отчета')

        next_trip_form_response = operator_client.get('/excavator/work/', HTTP_HOST='localhost')
        self.assertContains(next_trip_form_response, 'value="7000')
        self.assertContains(next_trip_form_response, 'value="75"')
        self.assertContains(next_trip_form_response, 'value="52"')
        self.assertContains(next_trip_form_response, 'value="3.10"')

        driver_shift_response = driver_client.get('/driver/shift/', HTTP_HOST='localhost')
        self.assertContains(driver_shift_response, 'ККД')
        self.assertContains(driver_shift_response, 'Горизонт 75')
        self.assertContains(driver_shift_response, 'Блок 52')
        self.assertContains(driver_shift_response, 'ТОЧКА РАЗГРУЗКИ')
        self.assertContains(driver_shift_response, 'Выбор точки')
        self.assertNotContains(driver_shift_response, 'Активный рейс')
        self.assertNotContains(driver_shift_response, 'Разгрузился')

        change_point_response = driver_client.post(
            f'/driver/trip/{trip.id}/change-unload-point/',
            {'client_action_id': 'driver-change-point-1', 'dump_point': skdr_point.id},
            follow=True,
            HTTP_HOST='localhost',
        )
        trip.refresh_from_db()

        self.assertEqual(change_point_response.status_code, 200)
        self.assertEqual(trip.assigned_dump_point, dump_point)
        self.assertEqual(trip.actual_dump_point, skdr_point)
        self.assertEqual(trip.dump_point, skdr_point)
        self.assertTrue(
            TripClientAction.objects.filter(
                action_type='change_actual_unload_point',
                client_action_id='driver-change-point-1',
                trip=trip,
                actor=self.employee,
            ).exists()
        )

        complete_response = driver_client.post(
            f'/driver/trip/{trip.id}/complete/',
            {'client_action_id': 'driver-unload-test-1'},
            follow=True,
            HTTP_HOST='localhost',
        )
        trip.refresh_from_db()

        self.assertEqual(complete_response.status_code, 200)
        self.assertEqual(trip.status, TripStatus.COMPLETED)
        self.assertEqual(trip.driver, self.employee)
        self.assertIsNotNone(trip.completed_at)
        self.assertFalse(trip.is_carryover)
        self.assertTrue(
            TripClientAction.objects.filter(
                action_type='trip_unloaded',
                client_action_id='driver-unload-test-1',
                trip=trip,
                actor=self.employee,
            ).exists()
        )

    def test_driver_sees_truck_loaded_event_from_excavator_realtime_shell(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        rock = RockType.objects.create(name='Руда', density=Decimal('2.6000'))
        dump_point = DumpPoint.objects.create(name='ККД')
        dormitory = Dormitory.objects.create(number='5')
        block = DormitoryBlock.objects.create(dormitory=dormitory, name='Блок 1')
        section = DormitorySection.objects.create(block=block, name='А')
        self.assign_driver_work(truck)
        TruckCapacityRule.objects.create(
            equipment_model=truck.model,
            rock_type=rock,
            volume_m3=Decimal('49.40'),
        )
        excavator_role = Role.objects.create(code='excavator_operator', name='Машинист экскаватора')
        excavator_operator = Employee.objects.create(
            full_name='Тестовый машинист',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        EmployeeAccess.objects.create(employee=excavator_operator, role=excavator_role, access_code='3000')
        driver = Employee.objects.create(full_name='Тестовый водитель')
        EmployeeShift.objects.create(
            employee=driver,
            shift_type='day',
            equipment=truck,
            opened_at=timezone.now(),
        )
        EmployeeShift.objects.create(
            employee=excavator_operator,
            shift_type='day',
            equipment=excavator,
            opened_at=timezone.now(),
        )

        driver_client = self.client
        driver_client.post('/', {'access_code': '2000'}, follow=True, HTTP_HOST='localhost')
        driver_client.post(
            '/driver/registration/',
            {'dormitory_section': section.id},
            follow=True,
            HTTP_HOST='localhost',
        )
        driver_client.post(
            '/driver/shift/',
            {
                'shift_type': 'day',
                'truck': truck.id,
                'start_fuel': '100',
                'start_mileage': '2500',
                'start_engine_hours': '700',
            },
            follow=True,
            HTTP_HOST='localhost',
        )
        assignment = HaulAssignment.objects.create(
            truck=truck,
            excavator=excavator,
            status=AssignmentStatus.ACCEPTED,
        )

        operator_client = self.client_class(HTTP_HOST='localhost')
        operator_client.post('/', {'access_code': '3000'}, follow=True, HTTP_HOST='localhost')
        response = operator_client.post(
            '/excavator/truck-loaded/',
            data=json.dumps({
                'client_action_id': 'truck-loaded-driver-sync-1',
                'truck_id': truck.id,
                'excavator_id': excavator.id,
                'dump_point_id': dump_point.id,
                'rock_type_id': rock.id,
            }),
            content_type='application/json',
            HTTP_HOST='localhost',
        )
        trip = Trip.objects.get()
        driver_shift_response = driver_client.get('/driver/shift/', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['action'], 'truck_loaded')
        self.assertEqual(trip.status, TripStatus.LOADED_WAITING_UNLOAD)
        self.assertEqual(trip.assigned_dump_point, dump_point)
        self.assertEqual(assignment.truck, truck)
        self.assertContains(driver_shift_response, 'ККД')
        self.assertContains(driver_shift_response, 'window.applyOperationalStateRefresh')
        self.assertContains(driver_shift_response, 'data-realtime-mode="custom"')
        self.assertContains(driver_shift_response, 'driver-mobile-shell-v177')

    def test_driver_downtime_buttons_are_rendered_from_server_reference(self):
        truck = self.create_registered_driver_shift()
        DowntimeReason.objects.all().update(show_for_truck_driver=False)
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        waiting_reason = DowntimeReason.objects.create(
            name='Тест водитель фронт работ',
            short_label='Фронт',
            show_for_truck_driver=True,
            sort_order=10,
        )
        truck_reason = DowntimeReason.objects.create(
            name='Тест водитель чистка кузова',
            short_label='Кузов',
            equipment_type=truck.equipment_type,
            show_for_truck_driver=True,
            sort_order=20,
        )
        hidden_reason = DowntimeReason.objects.create(
            name='Тест водитель скрытый простой',
            short_label='Скрытый',
            show_for_truck_driver=False,
        )
        excavator_reason = DowntimeReason.objects.create(
            name='Тест водитель чужой экскаватор',
            short_label='Забой',
            equipment_type=excavator_type,
            show_for_truck_driver=True,
        )

        response = self.client.get('/driver/?tab=downtimes', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'class="driver-downtime-title-row"')
        self.assertContains(response, '>Простои<')
        self.assertContains(response, 'data-driver-downtime-refresh')
        self.assertContains(response, 'driver-downtime-state-icon-play')
        self.assertContains(response, 'driver-downtime-state-icon-pause')
        self.assertContains(response, 'class="driver-downtime-list"')
        self.assertContains(response, 'aria-label="Начать простой: Фронт"')
        self.assertContains(response, 'registerDriverDowntimeAction')
        self.assertNotContains(response, 'registerDriverDowntimeHold')
        self.assertNotContains(response, 'Удерживайте полсекунды')
        self.assertNotContains(response, 'driverShiftOpenConfirmed')
        self.assertContains(response, f'data-driver-downtime-reason-id="{waiting_reason.id}"')
        self.assertContains(response, f'name="reason_id" value="{waiting_reason.id}"')
        self.assertContains(response, 'Фронт')
        self.assertContains(response, f'data-driver-downtime-reason-id="{truck_reason.id}"')
        self.assertContains(response, 'Кузов')
        self.assertNotContains(response, hidden_reason.button_label)
        self.assertNotContains(response, excavator_reason.button_label)

    def test_driver_downtime_empty_state_uses_reference_message(self):
        self.create_registered_driver_shift()
        DowntimeReason.objects.all().update(show_for_truck_driver=False)

        response = self.client.get('/driver/?tab=downtimes', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Активные причины простоев для самосвалов не найдены')

    def test_driver_inactive_downtime_timer_shows_shift_total(self):
        truck = self.create_registered_driver_shift()
        shift = EmployeeShift.objects.get(employee=self.employee, closed_at__isnull=True)
        shift.opened_at = timezone.now() - timedelta(hours=1)
        shift.save(update_fields=['opened_at'])
        reason = DowntimeReason.objects.create(name='Тест итог простоя', show_for_truck_driver=True)
        started_at = timezone.now() - timedelta(minutes=25)
        DowntimeEvent.objects.create(
            equipment=truck,
            employee=self.employee,
            reason=reason,
            started_at=started_at,
            ended_at=started_at + timedelta(minutes=20),
        )

        response = self.client.get('/driver/?tab=downtimes', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-driver-shift-downtime-seconds="1200"')
        self.assertContains(response, '>00:20:00</b>')
        self.assertContains(response, 'data-duration="20 мин."')
        self.assertContains(response, 'data-driver-report-downtime-total="20 мин."')
        self.assertNotContains(response, 'driver-downtime-card status-yellow is-active')

    def test_driver_downtime_action_validates_reason_by_workplace_and_equipment_type(self):
        truck = self.create_registered_driver_shift()
        DowntimeReason.objects.all().update(show_for_truck_driver=False)
        allowed_reason = DowntimeReason.objects.create(
            name='Тест водитель ожидание разгрузки',
            short_label='Разгрузка',
            show_for_truck_driver=True,
        )
        forbidden_reason = DowntimeReason.objects.create(
            name='Тест водитель диагностика механика',
            short_label='Диагностика',
            equipment_type=truck.equipment_type,
            show_for_truck_driver=False,
            show_for_mechanic=True,
        )

        forbidden_response = self.client.post(
            reverse('driver_downtime_action'),
            {'reason_id': forbidden_reason.id},
            follow=True,
            HTTP_HOST='localhost',
        )
        allowed_response = self.client.post(
            reverse('driver_downtime_action'),
            {'reason_id': allowed_reason.id},
            follow=True,
            HTTP_HOST='localhost',
        )

        self.assertEqual(forbidden_response.status_code, 200)
        self.assertContains(forbidden_response, 'Причина простоя не найдена')
        self.assertEqual(DowntimeEvent.objects.count(), 1)
        event = DowntimeEvent.objects.get()
        self.assertEqual(event.reason, allowed_reason)
        self.assertEqual(event.equipment, truck)
        self.assertEqual(event.employee, self.employee)

    def test_driver_downtime_reference_change_is_visible_after_server_refresh(self):
        self.create_registered_driver_shift()
        DowntimeReason.objects.all().update(show_for_truck_driver=False)
        DowntimeReason.objects.create(
            name='Тест водитель ожидание погрузки',
            short_label='Погрузка',
            show_for_truck_driver=True,
        )

        initial_response = self.client.get('/driver/?tab=downtimes', HTTP_HOST='localhost')
        new_reason = DowntimeReason.objects.create(
            name='Тест водитель климатические условия',
            short_label='Погода',
            show_for_truck_driver=True,
            sort_order=30,
        )
        refreshed_response = self.client.get(
            '/driver/?tab=downtimes&_driver_refresh=1',
            HTTP_HOST='localhost',
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )

        self.assertContains(initial_response, 'Погрузка')
        self.assertNotContains(initial_response, 'Погода')
        self.assertContains(refreshed_response, f'data-driver-downtime-reason-id="{new_reason.id}"')
        self.assertContains(refreshed_response, 'Погода')
        self.assertContains(refreshed_response, 'window.applyOperationalStateRefresh')
        self.assertTrue(
            OperationalStateEvent.objects.filter(
                event_type='reference_changed',
                object_type='DowntimeReason',
                object_id=str(new_reason.id),
            ).exists()
        )

    def test_trip_becomes_carryover_when_loading_and_unloading_shift_types_differ(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        rock = RockType.objects.create(name='Руда')
        dump_point = DumpPoint.objects.create(name='ККД')
        dormitory = Dormitory.objects.create(number='5')
        block = DormitoryBlock.objects.create(dormitory=dormitory, name='Блок 1')
        section = DormitorySection.objects.create(block=block, name='А')
        self.assign_driver_work(truck, shift_type='night')
        excavator_operator = Employee.objects.create(full_name='Тестовый машинист')
        loading_shift = EmployeeShift.objects.create(
            employee=excavator_operator,
            shift_type='day',
            equipment=excavator,
            opened_at=timezone.now(),
            closed_at=timezone.now(),
        )

        self.client.post('/', {'access_code': '2000'}, follow=True, HTTP_HOST='localhost')
        self.client.post(
            '/driver/registration/',
            {'dormitory_section': section.id},
            follow=True,
            HTTP_HOST='localhost',
        )
        self.client.post(
            '/driver/shift/',
            {'shift_type': 'night', 'truck': truck.id, 'start_fuel': '100', 'start_mileage': '2500', 'start_engine_hours': '700'},
            follow=True,
            HTTP_HOST='localhost',
        )
        trip = Trip.objects.create(
            excavator=excavator,
            truck=truck,
            excavator_operator=excavator_operator,
            loading_shift=loading_shift,
            rock_type=rock,
            dump_point=dump_point,
            status=TripStatus.ACTIVE,
            volume_m3='57.00',
        )

        self.client.post(
            f'/driver/trip/{trip.id}/complete/',
            {'client_action_id': 'driver-unload-carryover-1'},
            follow=True,
            HTTP_HOST='localhost',
        )
        trip.refresh_from_db()

        self.assertEqual(trip.status, TripStatus.COMPLETED)
        self.assertTrue(trip.is_carryover)
        self.assertEqual(trip.unloading_shift.shift_type, 'night')

    def test_trip_volume_and_tonnage_are_calculated_from_capacity_rule_and_density(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck_model = EquipmentModel.objects.create(
            equipment_type=truck_type,
            name='БЕЛАЗ тест',
            body_volume_m3='40.00',
        )
        truck = Equipment.objects.create(equipment_type=truck_type, model=truck_model, garage_number='10')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        rock = RockType.objects.create(name='Руда', density='2.50')
        dump_point = DumpPoint.objects.create(name='ККД')
        excavator_role = Role.objects.create(code='excavator_operator', name='Машинист экскаватора')
        excavator_operator = Employee.objects.create(
            full_name='Тестовый машинист',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        excavator_access = EmployeeAccess.objects.create(
            employee=excavator_operator,
            role=excavator_role,
            access_code='3000',
            is_active=True,
            status=EmployeeAccess.Status.ACTIVATED,
        )
        driver = Employee.objects.create(full_name='Тестовый водитель')
        driver_shift = EmployeeShift.objects.create(
            employee=driver,
            shift_type='day',
            equipment=truck,
            opened_at=timezone.now(),
        )
        EmployeeShift.objects.create(
            employee=excavator_operator,
            shift_type='day',
            equipment=excavator,
            opened_at=timezone.now(),
        )
        assignment = HaulAssignment.objects.create(truck=truck, excavator=excavator, status=AssignmentStatus.ACCEPTED)
        TruckCapacityRule.objects.create(equipment_model=truck_model, rock_type=rock, volume_m3='38.00')

        operator_client = self.client_class(HTTP_HOST='localhost')
        operator_session = operator_client.session
        operator_session['employee_access_id'] = excavator_access.id
        operator_session.save()
        operator_client.post(
            '/excavator/work/',
            {
                'client_action_id': 'legacy-capacity-trip-create-1',
                'assignment': assignment.id,
                'rock_type': rock.id,
                'dump_point': dump_point.id,
            },
            follow=True,
            HTTP_HOST='localhost',
        )
        trip = Trip.objects.get()

        self.assertEqual(trip.status, TripStatus.LOADED_WAITING_UNLOAD)
        self.assertEqual(trip.volume_m3, Decimal('38.00'))
        self.assertEqual(trip.tonnage, Decimal('95.00'))

        from trips.views import finalize_trip_unloaded

        finalize_trip_unloaded(
            trip,
            driver=driver,
            unloading_shift=driver_shift,
        )
        trip.refresh_from_db()

        self.assertEqual(trip.status, TripStatus.COMPLETED)
        self.assertEqual(trip.volume_m3, Decimal('38.00'))
        self.assertEqual(trip.tonnage, Decimal('95.00'))

    def test_dispatcher_can_see_volume_report(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        rock = RockType.objects.create(name='Руда')
        dump_point = DumpPoint.objects.create(name='ККД')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(full_name='Тестовый диспетчер')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5000')
        Trip.objects.create(
            excavator=excavator,
            truck=truck,
            rock_type=rock,
            dump_point=dump_point,
            status=TripStatus.COMPLETED,
            volume_m3='57.00',
            completed_at=timezone.now(),
        )

        response = self.client.post('/', {'access_code': '5000'}, follow=True, HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Диспетчерский пульт')

        report_response = self.client.get('/reports/volume/', HTTP_HOST='localhost')
        self.assertEqual(report_response.status_code, 200)
        self.assertContains(report_response, 'Отчет по объемам')
        self.assertContains(report_response, '57')

        export_response = self.client.get('/reports/volume/export/', HTTP_HOST='localhost')
        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(
            export_response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_dispatcher_can_open_mining_volumes_dashboard_and_export_it(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        rock = RockType.objects.create(name='Руда')
        dump_point = DumpPoint.objects.create(name='ККД')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(full_name='Тестовый диспетчер')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5000')
        EmployeeShift.objects.create(employee=dispatcher, shift_type='day', opened_at=timezone.now(), opened_by=dispatcher)
        Trip.objects.create(
            excavator=excavator,
            truck=truck,
            rock_type=rock,
            dump_point=dump_point,
            loading_horizon='245',
            loading_block='7',
            status=TripStatus.COMPLETED,
            planned_volume_m3='100.00',
            volume_m3='57.00',
            tonnage='142.50',
            completed_at=timezone.now(),
        )

        self.client.post('/', {'access_code': '5000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get('/dispatcher/mining-volumes/', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Горные объемы')
        self.assertContains(response, '57')
        self.assertContains(response, 'ККД')

        export_response = self.client.get('/dispatcher/mining-volumes/export/', HTTP_HOST='localhost')
        workbook = load_workbook(BytesIO(export_response.content))
        values = [
            cell
            for row in workbook.active.iter_rows(values_only=True)
            for cell in row
            if cell not in {None, ''}
        ]

        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(
            export_response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('Горные объемы диспетчера', values)
        self.assertIn('ККД', values)

    def test_dispatcher_can_open_transport_dashboard_and_export_it(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='15')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        rock = RockType.objects.create(name='Руда')
        dump_point = DumpPoint.objects.create(name='ККД')
        driver = Employee.objects.create(full_name='Водитель автотранспорта MVP')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(full_name='Тестовый диспетчер')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5010')
        selected_date = timezone.localdate()
        opened_at = timezone.make_aware(datetime.combine(selected_date, datetime.min.time().replace(hour=8)))
        closed_at = timezone.make_aware(datetime.combine(selected_date, datetime.min.time().replace(hour=12)))
        EmployeeShift.objects.create(
            employee=driver,
            shift_type='day',
            equipment=truck,
            start_fuel=Decimal('100.00'),
            end_fuel=Decimal('80.00'),
            start_mileage=Decimal('2500.00'),
            end_mileage=Decimal('2600.00'),
            start_engine_hours=Decimal('700.00'),
            end_engine_hours=Decimal('712.00'),
            opened_at=opened_at,
            closed_at=closed_at,
        )
        Trip.objects.create(
            excavator=excavator,
            truck=truck,
            rock_type=rock,
            dump_point=dump_point,
            status=TripStatus.COMPLETED,
            volume_m3='57.00',
            tonnage='142.50',
            completed_at=closed_at,
        )

        self.client.post('/', {'access_code': '5010'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get(
            '/dispatcher/transport/',
            {'date': selected_date.isoformat()},
            HTTP_HOST='localhost',
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Автотранспорт')
        self.assertContains(response, '15')
        self.assertContains(response, 'Водитель автотранспорта MVP')
        self.assertContains(response, '20,0')

        export_response = self.client.get(
            '/dispatcher/transport/export/',
            {'date': selected_date.isoformat()},
            HTTP_HOST='localhost',
        )
        workbook = load_workbook(BytesIO(export_response.content))
        values = [
            cell
            for row in workbook.active.iter_rows(values_only=True)
            for cell in row
            if cell not in {None, ''}
        ]

        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(
            export_response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('Автотранспорт диспетчера', values)
        self.assertIn('15', values)
        self.assertIn('Водитель автотранспорта MVP', values)

    def test_dispatcher_transport_legacy_night_trip_survives_all_night_filters_on_page_and_excel(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='LEGACY-15')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='LEGACY-1')
        rock = RockType.objects.create(name='Руда legacy')
        dump_point = DumpPoint.objects.create(name='ККД legacy')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(full_name='Диспетчер legacy')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5011')
        production_date = datetime(2026, 7, 23).date()
        completed_at = datetime(2026, 7, 24, 1, 30, tzinfo=ZoneInfo('Asia/Vladivostok'))
        Trip.objects.create(
            excavator=excavator,
            truck=truck,
            rock_type=rock,
            dump_point=dump_point,
            status=TripStatus.COMPLETED,
            volume_m3='41.00',
            tonnage='102.50',
            completed_at=completed_at,
            unloading_shift=None,
        )
        self.client.post('/', {'access_code': '5011'}, follow=True, HTTP_HOST='localhost')

        def page_trip_count(shift_type):
            response = self.client.get(
                '/dispatcher/transport/',
                {'date': production_date.isoformat(), 'shift_type': shift_type},
                HTTP_HOST='localhost',
            )
            self.assertEqual(response.status_code, 200)
            return response.context['kpis']['trips']

        def excel_trip_count(shift_type):
            response = self.client.get(
                '/dispatcher/transport/export/',
                {'date': production_date.isoformat(), 'shift_type': shift_type},
                HTTP_HOST='localhost',
            )
            self.assertEqual(response.status_code, 200)
            workbook = load_workbook(BytesIO(response.content), data_only=True)
            values = {
                row[0]: row[1]
                for row in workbook.active.iter_rows(min_row=1, max_col=2, values_only=True)
                if row[0]
            }
            return values['Рейсы']

        self.assertEqual(page_trip_count(''), 1)
        self.assertEqual(page_trip_count('day'), 0)
        self.assertEqual(page_trip_count('night'), 1)
        self.assertEqual(excel_trip_count(''), 1)
        self.assertEqual(excel_trip_count('day'), 0)
        self.assertEqual(excel_trip_count('night'), 1)

    def test_control_instant_appears_on_production_date_pages_and_excel(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='КОНТРОЛЬ-15')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='КОНТРОЛЬ-1')
        rock = RockType.objects.create(name='Руда контроль 01:30')
        dump_point = DumpPoint.objects.create(name='ККД контроль 01:30')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(full_name='Диспетчер контроль 01:30')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5012')
        control_instant = datetime(2026, 7, 24, 1, 30, tzinfo=ZoneInfo('Asia/Vladivostok'))
        production_date = datetime(2026, 7, 23).date()
        trip = Trip.objects.create(
            excavator=excavator,
            truck=truck,
            rock_type=rock,
            dump_point=dump_point,
            status=TripStatus.COMPLETED,
            volume_m3='43.00',
            completed_at=control_instant,
        )
        reason = DowntimeReason.objects.create(
            name='Простой контроль 01:30',
            equipment_type=excavator_type,
        )
        DowntimeEvent.objects.create(
            equipment=excavator,
            employee=dispatcher,
            reason=reason,
            started_at=control_instant,
        )
        action = DispatcherActionLog.objects.create(
            actor=dispatcher,
            action_type=DispatcherActionType.COMPLETE_TRIP,
            trip=trip,
            target_summary='Действие контроль 01:30',
        )
        DispatcherActionLog.objects.filter(pk=action.pk).update(created_at=control_instant)
        self.client.post('/', {'access_code': '5012'}, follow=True, HTTP_HOST='localhost')

        query = {'date': production_date.isoformat()}
        mining_page = self.client.get(
            '/dispatcher/mining-volumes/',
            {**query, 'shift_type': 'night'},
            HTTP_HOST='localhost',
        )
        downtime_page = self.client.get('/dispatcher/downtimes/', query, HTTP_HOST='localhost')
        shift_log_page = self.client.get('/dispatcher/shift-log/', query, HTTP_HOST='localhost')

        self.assertEqual(mining_page.context['kpis']['trips'], 1)
        self.assertEqual(downtime_page.context['kpis']['events'], 1)
        self.assertEqual(shift_log_page.context['kpis']['trips'], 1)
        self.assertEqual(shift_log_page.context['kpis']['downtimes'], 1)
        self.assertEqual(shift_log_page.context['kpis']['actions'], 1)

        for url, marker in (
            ('/dispatcher/mining-volumes/export/', 'ККД контроль 01:30'),
            ('/dispatcher/downtimes/export/', 'Простой контроль 01:30'),
            ('/dispatcher/shift-log/export/', 'Действие контроль 01:30'),
        ):
            export = self.client.get(url, query, HTTP_HOST='localhost')
            self.assertEqual(export.status_code, 200)
            workbook = load_workbook(BytesIO(export.content), data_only=True)
            values = {
                cell
                for row in workbook.active.iter_rows(values_only=True)
                for cell in row
                if cell not in {None, ''}
            }
            self.assertIn(marker, values)

    def test_dispatcher_transport_separates_same_truck_by_unloading_shift(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='15')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        rock = RockType.objects.create(name='Руда')
        dump_point = DumpPoint.objects.create(name='ККД')
        day_driver = Employee.objects.create(full_name='Водитель день')
        night_driver = Employee.objects.create(full_name='Водитель ночь')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(full_name='Тестовый диспетчер')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5011')
        selected_date = timezone.localdate()
        day_opened = timezone.make_aware(datetime.combine(selected_date, datetime.min.time().replace(hour=7)))
        night_opened = timezone.make_aware(datetime.combine(selected_date, datetime.min.time().replace(hour=19)))
        day_shift = EmployeeShift.objects.create(
            employee=day_driver,
            shift_type='day',
            equipment=truck,
            opened_at=day_opened,
            closed_at=day_opened + timedelta(hours=12),
        )
        night_shift = EmployeeShift.objects.create(
            employee=night_driver,
            shift_type='night',
            equipment=truck,
            opened_at=night_opened,
            closed_at=night_opened + timedelta(hours=12),
        )
        for shift, volumes in ((day_shift, ('47.00', '47.00')), (night_shift, ('47.00', '57.00'))):
            for index, volume in enumerate(volumes):
                completed_at = shift.opened_at + timedelta(minutes=index + 1)
                if shift == night_shift and index == 1:
                    completed_at = night_shift.opened_at + timedelta(hours=6, minutes=30)
                Trip.objects.create(
                    excavator=excavator,
                    truck=truck,
                    rock_type=rock,
                    dump_point=dump_point,
                    unloading_shift=shift,
                    status=TripStatus.COMPLETED,
                    volume_m3=volume,
                    tonnage='100.00',
                    completed_at=completed_at,
                )

        self.client.post('/', {'access_code': '5011'}, follow=True, HTTP_HOST='localhost')
        all_response = self.client.get(
            f'/dispatcher/transport/?date={selected_date:%Y-%m-%d}',
            HTTP_HOST='localhost',
        )
        day_response = self.client.get(
            f'/dispatcher/transport/?date={selected_date:%Y-%m-%d}&shift_type=day',
            HTTP_HOST='localhost',
        )
        night_response = self.client.get(
            f'/dispatcher/transport/?date={selected_date:%Y-%m-%d}&shift_type=night',
            HTTP_HOST='localhost',
        )
        next_date_response = self.client.get(
            f'/dispatcher/transport/?date={selected_date + timedelta(days=1):%Y-%m-%d}',
            HTTP_HOST='localhost',
        )

        self.assertEqual(all_response.context['kpis']['trips'], 4)
        self.assertEqual(all_response.context['kpis']['volume'], '198')
        self.assertEqual([row['trips'] for row in all_response.context['rows']], [2, 2])
        self.assertEqual(day_response.context['kpis']['trips'], 2)
        self.assertEqual(day_response.context['kpis']['volume'], '94')
        self.assertEqual(night_response.context['kpis']['trips'], 2)
        self.assertEqual(night_response.context['kpis']['volume'], '104')
        self.assertEqual(next_date_response.context['kpis']['trips'], 0)
        self.assertEqual(next_date_response.context['kpis']['volume'], '0')

        export_response = self.client.get(
            f'/dispatcher/transport/export/?date={selected_date:%Y-%m-%d}',
            HTTP_HOST='localhost',
        )
        workbook = load_workbook(BytesIO(export_response.content), data_only=True)
        sheet_values = list(workbook.active.iter_rows(values_only=True))
        self.assertIn(('Рейсы', 4), [row[:2] for row in sheet_values])
        self.assertIn(('Объем', '198'), [row[:2] for row in sheet_values])

    def test_dispatcher_can_open_downtimes_dashboard_and_export_it(self):
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='6')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='34')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(full_name='Тестовый диспетчер')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5020')
        critical_reason = DowntimeReason.objects.create(
            name='Аварийная поломка',
            equipment_type=excavator_type,
            is_critical=True,
        )
        waiting_reason = DowntimeReason.objects.get(name='Ожидание разгрузки')
        waiting_reason.equipment_type = truck_type
        waiting_reason.save(update_fields=['equipment_type'])
        selected_date = timezone.localdate()
        started_at = timezone.make_aware(datetime.combine(selected_date, datetime.min.time().replace(hour=9)))
        DowntimeEvent.objects.create(
            equipment=excavator,
            employee=dispatcher,
            reason=critical_reason,
            started_at=started_at,
            comment='Демо диагностика',
        )
        DowntimeEvent.objects.create(
            equipment=truck,
            employee=dispatcher,
            reason=waiting_reason,
            started_at=started_at,
            ended_at=started_at + timedelta(minutes=45),
        )

        self.client.post('/', {'access_code': '5020'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get(
            '/dispatcher/downtimes/',
            {'date': selected_date.isoformat()},
            HTTP_HOST='localhost',
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Простои и отклонения')
        self.assertContains(response, 'Аварийная поломка')
        self.assertContains(response, '6')
        self.assertContains(response, '34')

        export_response = self.client.get(
            '/dispatcher/downtimes/export/',
            {'date': selected_date.isoformat()},
            HTTP_HOST='localhost',
        )
        workbook = load_workbook(BytesIO(export_response.content))
        values = [
            cell
            for row in workbook.active.iter_rows(values_only=True)
            for cell in row
            if cell not in {None, ''}
        ]

        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(
            export_response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('Простои и отклонения диспетчера', values)
        self.assertIn('Аварийная поломка', values)
        self.assertIn('Экскаватор 6', values)

    def test_dispatcher_can_open_shift_log_and_export_it(self):
        excavator_type = EquipmentType.objects.create(name='Excavator')
        truck_type = EquipmentType.objects.create(name='Truck')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        rock = RockType.objects.create(name='Oxide')
        dump_point = DumpPoint.objects.create(name='KKD')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Dispatcher')
        dispatcher = Employee.objects.create(full_name='Shift log dispatcher')
        driver = Employee.objects.create(full_name='Shift log driver')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5030')
        event_time = timezone.now()
        trip = Trip.objects.create(
            excavator=excavator,
            truck=truck,
            driver=driver,
            rock_type=rock,
            dump_point=dump_point,
            loading_horizon='245',
            loading_block='7',
            status=TripStatus.COMPLETED,
            volume_m3='57.00',
            tonnage='142.50',
            completed_at=event_time,
        )
        DispatcherActionLog.objects.create(
            actor=dispatcher,
            action_type=DispatcherActionType.COMPLETE_TRIP,
            trip=trip,
            target_summary='Trip 10 completed manually',
            reason='Shift reconciliation demo',
        )
        reason = DowntimeReason.objects.create(name='Shift log diagnostics', equipment_type=excavator_type, is_critical=True)
        DowntimeEvent.objects.create(
            equipment=excavator,
            employee=dispatcher,
            reason=reason,
            started_at=event_time,
            comment='Shift timeline check',
        )

        self.client.post('/', {'access_code': '5030'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get('/dispatcher/shift-log/', HTTP_HOST='localhost')
        export_response = self.client.get('/dispatcher/shift-log/export/', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Trip 10 completed manually')
        self.assertContains(response, 'Shift log diagnostics')
        self.assertContains(response, 'Truck 10')
        self.assertEqual(
            export_response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        workbook = load_workbook(BytesIO(export_response.content))
        values = [
            cell
            for row in workbook.active.iter_rows(values_only=True)
            for cell in row
            if cell not in {None, ''}
        ]
        self.assertIn('Журнал смены диспетчера', values)
        self.assertIn('Trip 10 completed manually', values)
        self.assertIn('Shift log diagnostics', values)

    def test_dispatcher_can_open_reports_center_and_export_it(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='15')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        rock = RockType.objects.create(name='Руда')
        dump_point = DumpPoint.objects.create(name='ККД')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(full_name='Диспетчер отчетов')
        driver = Employee.objects.create(full_name='Водитель отчетов')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5040')
        selected_date = timezone.localdate()
        opened_at = timezone.make_aware(datetime.combine(selected_date, datetime.min.time().replace(hour=8)))
        closed_at = timezone.make_aware(datetime.combine(selected_date, datetime.min.time().replace(hour=12)))
        EmployeeShift.objects.create(
            employee=driver,
            shift_type='day',
            equipment=truck,
            start_fuel=Decimal('100.00'),
            end_fuel=Decimal('80.00'),
            start_mileage=Decimal('2500.00'),
            end_mileage=Decimal('2600.00'),
            start_engine_hours=Decimal('700.00'),
            end_engine_hours=Decimal('712.00'),
            opened_at=opened_at,
            closed_at=closed_at,
        )
        trip = Trip.objects.create(
            excavator=excavator,
            truck=truck,
            driver=driver,
            rock_type=rock,
            dump_point=dump_point,
            status=TripStatus.COMPLETED,
            planned_volume_m3='100.00',
            volume_m3='57.00',
            tonnage='142.50',
            completed_at=closed_at,
        )
        DispatcherActionLog.objects.create(
            actor=dispatcher,
            action_type=DispatcherActionType.COMPLETE_TRIP,
            trip=trip,
            target_summary='Отчетный рейс закрыт',
        )
        ReportTemplate.objects.create(
            name='Демо шаблон диспетчера',
            report_type=ReportType.SHIFT_VOLUME,
            columns=['truck', 'volume_m3'],
            created_by=dispatcher,
        )

        self.client.post('/', {'access_code': '5040'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get('/dispatcher/reports/', HTTP_HOST='localhost')
        export_response = self.client.get('/dispatcher/reports/export/', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Отчеты диспетчера')
        self.assertContains(response, 'Сменные объемы')
        self.assertContains(response, 'Автотранспорт')
        self.assertContains(response, 'Конструктор')
        self.assertEqual(
            export_response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        workbook = load_workbook(BytesIO(export_response.content))
        values = [
            cell
            for row in workbook.active.iter_rows(values_only=True)
            for cell in row
            if cell not in {None, ''}
        ]
        self.assertIn('Отчеты диспетчерской', values)
        self.assertIn('Сменные объемы', values)
        self.assertIn('Демо шаблон диспетчера', values)

    def test_dispatcher_can_open_management_showcase_and_export_it(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='15')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        rock = RockType.objects.create(name='Руда')
        dump_point = DumpPoint.objects.create(name='ККД')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(full_name='Диспетчер витрины')
        driver = Employee.objects.create(full_name='Водитель витрины')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5050')
        selected_date = timezone.localdate()
        opened_at = timezone.make_aware(datetime.combine(selected_date, datetime.min.time().replace(hour=8)))
        closed_at = timezone.make_aware(datetime.combine(selected_date, datetime.min.time().replace(hour=12)))
        EmployeeShift.objects.create(
            employee=driver,
            shift_type='day',
            equipment=truck,
            start_fuel=Decimal('100.00'),
            end_fuel=None,
            start_mileage=Decimal('2500.00'),
            end_mileage=None,
            start_engine_hours=Decimal('700.00'),
            end_engine_hours=None,
            opened_at=opened_at,
        )
        trip = Trip.objects.create(
            excavator=excavator,
            truck=truck,
            driver=driver,
            rock_type=rock,
            dump_point=dump_point,
            status=TripStatus.COMPLETED,
            planned_volume_m3='100.00',
            volume_m3='96.00',
            tonnage='240.00',
            loading_horizon='245',
            loading_block='7',
            completed_at=closed_at,
        )
        DispatcherActionLog.objects.create(
            actor=dispatcher,
            action_type=DispatcherActionType.COMPLETE_TRIP,
            trip=trip,
            target_summary='Рейс витрины закрыт',
        )
        reason = DowntimeReason.objects.create(name='Диагностика витрины', equipment_type=truck_type, is_critical=True)
        DowntimeEvent.objects.create(
            equipment=truck,
            employee=dispatcher,
            reason=reason,
            started_at=opened_at,
            comment='Проверка витрины',
        )

        self.client.post('/', {'access_code': '5050'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get(
            '/dispatcher/management/',
            {'date': selected_date.isoformat()},
            HTTP_HOST='localhost',
        )
        export_response = self.client.get(
            '/dispatcher/management/export/',
            {'date': selected_date.isoformat()},
            HTTP_HOST='localhost',
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Витрина диспетчерской')
        self.assertContains(response, 'Итог смены')
        self.assertContains(response, 'Комплексы')
        self.assertContains(response, 'ККД')
        self.assertEqual(
            export_response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        workbook = load_workbook(BytesIO(export_response.content))
        values = [
            cell
            for row in workbook.active.iter_rows(values_only=True)
            for cell in row
            if cell not in {None, ''}
        ]
        self.assertIn('Витрина диспетчерской', values)
        self.assertIn('Итог смены', values)
        self.assertIn('К-1', values)
        self.assertIn('ККД', values)

    def test_dispatcher_opens_control_panel_with_active_trips_and_pending_assignments(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        second_truck = Equipment.objects.create(equipment_type=truck_type, garage_number='11')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        rock = RockType.objects.create(name='Руда')
        dump_point = DumpPoint.objects.create(name='ККД')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(full_name='Тестовый диспетчер')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5000')
        EmployeeShift.objects.create(employee=dispatcher, shift_type='day', opened_at=timezone.now(), opened_by=dispatcher)
        Trip.objects.create(
            excavator=excavator,
            truck=truck,
            rock_type=rock,
            dump_point=dump_point,
            status=TripStatus.ACTIVE,
            volume_m3='57.00',
        )
        HaulAssignment.objects.create(
            truck=second_truck,
            excavator=excavator,
            status=AssignmentStatus.PENDING,
        )
        HaulAssignment.objects.create(
            truck=truck,
            excavator=excavator,
            status=AssignmentStatus.ACCEPTED,
            accepted_at=timezone.now(),
        )

        response = self.client.post('/', {'access_code': '5000'}, follow=True, HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Диспетчерский пульт')
        self.assertContains(response, 'Активные рейсы')
        self.assertContains(response, 'Назначения ждут подтверждения')
        self.assertContains(response, 'Принятые назначения в работе')
        self.assertContains(response, '57')
        self.assertContains(response, 'Открыть отчет по объемам')
        self.assertContains(response, 'Суточный отчет заказчику')

    def test_dispatcher_control_panel_can_filter_by_truck(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        second_truck = Equipment.objects.create(equipment_type=truck_type, garage_number='11')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        rock = RockType.objects.create(name='Руда')
        dump_point = DumpPoint.objects.create(name='ККД')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(full_name='Тестовый диспетчер')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5000')
        Trip.objects.create(
            excavator=excavator,
            truck=truck,
            rock_type=rock,
            dump_point=dump_point,
            status=TripStatus.ACTIVE,
            volume_m3='11.00',
        )
        Trip.objects.create(
            excavator=excavator,
            truck=second_truck,
            rock_type=rock,
            dump_point=dump_point,
            status=TripStatus.ACTIVE,
            volume_m3='22.00',
        )

        self.client.post('/', {'access_code': '5000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get(f'/dispatcher/control/?truck={truck.id}', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '11')
        self.assertNotContains(response, '22,00')

    def test_dispatcher_control_panel_can_hide_accepted_assignments(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(full_name='Тестовый диспетчер')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5000')
        EmployeeShift.objects.create(employee=dispatcher, shift_type='day', opened_at=timezone.now(), opened_by=dispatcher)
        HaulAssignment.objects.create(
            truck=truck,
            excavator=excavator,
            status=AssignmentStatus.ACCEPTED,
            accepted_at=timezone.now(),
        )

        self.client.post('/', {'access_code': '5000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get('/dispatcher/control/?show_accepted_assignments=0', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Принятых назначений в работе сейчас нет.')

    def test_dispatcher_cannot_change_complexes_without_open_shift(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(
            full_name='Тестовый диспетчер',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        access = EmployeeAccess.objects.create(
            employee=dispatcher,
            role=dispatcher_role,
            access_code='5000',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        ExcavatorPlacement.objects.create(excavator=excavator, zone=ExcavatorPlacement.Zone.INACTIVE)
        session = self.client.session
        session['employee_access_id'] = access.id
        session.save()

        response = self.client.post(
            '/dispatcher/control/truck/assign/',
            data=json.dumps({
                'action': 'assign',
                'truck_id': truck.id,
                'excavator_id': excavator.id,
            }),
            content_type='application/json',
            HTTP_HOST='localhost',
        )

        self.assertEqual(response.status_code, 409)
        self.assertFalse(HaulAssignment.objects.filter(truck=truck, excavator=excavator).exists())
        placement = ExcavatorPlacement.objects.get(excavator=excavator)
        self.assertEqual(placement.zone, ExcavatorPlacement.Zone.INACTIVE)

    def test_dispatcher_shift_metrics_start_from_open_shift(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        rock = RockType.objects.create(name='Руда')
        dump_point = DumpPoint.objects.create(name='ККД')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(full_name='Тестовый диспетчер')
        access = EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5000')
        old_trip = Trip.objects.create(
            excavator=excavator,
            truck=truck,
            rock_type=rock,
            dump_point=dump_point,
            status=TripStatus.ACTIVE,
            volume_m3=Decimal('57.00'),
        )
        old_trip.created_at = timezone.now() - timedelta(hours=2)
        old_trip.save(update_fields=['created_at'])
        EmployeeShift.objects.create(
            employee=dispatcher,
            shift_type='day',
            workplace_code='dispatcher',
            opened_at=timezone.now() - timedelta(hours=1),
            opened_by=dispatcher,
        )
        Trip.objects.create(
            excavator=excavator,
            truck=truck,
            rock_type=rock,
            dump_point=dump_point,
            status=TripStatus.ACTIVE,
            volume_m3=Decimal('22.00'),
        )
        session = self.client.session
        session['employee_access_id'] = access.id
        session.save()

        response = self.client.get('/dispatcher/control/', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['dispatcher_dashboard']['dispatcher_kpis']['fact_tons'], '22')

    def test_dispatcher_service_action_preserves_current_filters(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        second_truck = Equipment.objects.create(equipment_type=truck_type, garage_number='11')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        rock = RockType.objects.create(name='Руда')
        dump_point = DumpPoint.objects.create(name='ККД')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(full_name='Тестовый диспетчер')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5000')
        EmployeeShift.objects.create(employee=dispatcher, shift_type='day', opened_at=timezone.now(), opened_by=dispatcher)
        target_trip = Trip.objects.create(
            excavator=excavator,
            truck=truck,
            rock_type=rock,
            dump_point=dump_point,
            status=TripStatus.ACTIVE,
            volume_m3='11.00',
        )
        Trip.objects.create(
            excavator=excavator,
            truck=second_truck,
            rock_type=rock,
            dump_point=dump_point,
            status=TripStatus.ACTIVE,
            volume_m3='22.00',
        )

        self.client.post('/', {'access_code': '5000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.post(
            f'/dispatcher/trips/{target_trip.id}/cancel/',
            {
                'reason': 'Фильтр должен сохраниться',
                'truck': str(truck.id),
                'show_active_trips': '1',
                'show_pending_assignments': '0',
                'show_accepted_assignments': '0',
            },
            follow=True,
            HTTP_HOST='localhost',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.redirect_chain)
        self.assertTrue(
            response.redirect_chain[-1][0].endswith(
                f'/dispatcher/control/?truck={truck.id}&show_active_trips=1&show_pending_assignments=0&show_accepted_assignments=0'
            )
        )
        self.assertNotContains(response, '22,00')

    def test_dispatcher_sees_open_shifts_and_can_service_close_driver_shift(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        truck_model = EquipmentModel.objects.create(
            equipment_type=truck_type,
            name='БЕЛАЗ для служебного закрытия',
            fuel_capacity_limit_l=2000,
        )
        truck = Equipment.objects.create(
            equipment_type=truck_type,
            model=truck_model,
            garage_number='10',
        )
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        driver_role, _ = Role.objects.get_or_create(code='driver', defaults={'name': 'Водитель самосвала'})
        dispatcher = Employee.objects.create(full_name='Тестовый диспетчер')
        driver = Employee.objects.create(full_name='Тестовый водитель')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5000')
        EmployeeAccess.objects.create(employee=driver, role=driver_role, access_code='2100')
        EmployeeShift.objects.create(
            employee=dispatcher,
            shift_type='day',
            workplace_code='dispatcher',
            opened_at=timezone.now(),
            opened_by=dispatcher,
        )
        shift = EmployeeShift.objects.create(
            employee=driver,
            shift_type='day',
            workplace_code='driver',
            equipment=truck,
            start_fuel='100.00',
            start_mileage='1000.00',
            start_engine_hours='100.00',
            opened_at=timezone.now(),
            opened_by=driver,
        )

        self.client.post('/', {'access_code': '5000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get('/dispatcher/control/', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Незакрытые смены')
        self.assertContains(response, 'Тестовый водитель')
        self.assertContains(response, 'Водитель самосвала')
        self.assertContains(response, 'Закрыть служебно')

        close_response = self.client.post(
            f'/dispatcher/shifts/{shift.id}/service-close/',
            {
                'reason': 'Сотрудник не смог закрыть смену',
                'end_fuel': '90.00',
                'end_mileage': '1001.00',
                'end_engine_hours': '101.00',
            },
            follow=True,
            HTTP_HOST='localhost',
        )
        shift.refresh_from_db()

        self.assertEqual(close_response.status_code, 200)
        self.assertIsNotNone(shift.closed_at)
        self.assertTrue(shift.is_service_closed)
        self.assertEqual(shift.closed_by, dispatcher)
        self.assertContains(close_response, 'Открытых смен сейчас нет.')
        action = DispatcherActionLog.objects.get()
        self.assertEqual(action.actor, dispatcher)
        self.assertEqual(action.action_type, DispatcherActionType.SERVICE_CLOSE_SHIFT)
        self.assertEqual(action.reason, 'Сотрудник не смог закрыть смену')

    def test_manager_cannot_service_close_shift(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        manager_role = Role.objects.create(code='manager', name='Руководство')
        driver_role, _ = Role.objects.get_or_create(code='driver', defaults={'name': 'Водитель самосвала'})
        manager = Employee.objects.create(full_name='Тестовый руководитель')
        driver = Employee.objects.create(full_name='Тестовый водитель')
        EmployeeAccess.objects.create(employee=manager, role=manager_role, access_code='6000')
        EmployeeAccess.objects.create(employee=driver, role=driver_role, access_code='2101')
        shift = EmployeeShift.objects.create(
            employee=driver,
            shift_type='day',
            equipment=truck,
            opened_at=timezone.now(),
            opened_by=driver,
        )

        self.client.post('/', {'access_code': '6000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.post(
            f'/dispatcher/shifts/{shift.id}/service-close/',
            follow=True,
            HTTP_HOST='localhost',
        )
        shift.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertIsNone(shift.closed_at)
        self.assertFalse(shift.is_service_closed)

    def test_dispatcher_can_cancel_pending_assignment_from_control_panel(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(full_name='Тестовый диспетчер')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5000')
        EmployeeShift.objects.create(employee=dispatcher, shift_type='day', opened_at=timezone.now(), opened_by=dispatcher)
        assignment = HaulAssignment.objects.create(
            truck=truck,
            excavator=excavator,
            status=AssignmentStatus.PENDING,
        )

        self.client.post('/', {'access_code': '5000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.post(
            f'/dispatcher/assignments/{assignment.id}/cancel/',
            {'reason': 'Переназначение техники'},
            follow=True,
            HTTP_HOST='localhost',
        )
        assignment.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(assignment.status, AssignmentStatus.CANCELLED)
        self.assertIsNotNone(assignment.ended_at)
        self.assertContains(response, 'Ожидающих подтверждения назначений нет.')
        action = DispatcherActionLog.objects.get()
        self.assertEqual(action.action_type, DispatcherActionType.CANCEL_ASSIGNMENT)
        self.assertEqual(action.reason, 'Переназначение техники')

    def test_dispatcher_can_cancel_accepted_assignment_from_control_panel(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(full_name='Тестовый диспетчер')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5000')
        EmployeeShift.objects.create(employee=dispatcher, shift_type='day', opened_at=timezone.now(), opened_by=dispatcher)
        assignment = HaulAssignment.objects.create(
            truck=truck,
            excavator=excavator,
            status=AssignmentStatus.ACCEPTED,
            accepted_at=timezone.now(),
        )

        self.client.post('/', {'access_code': '5000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.post(
            f'/dispatcher/assignments/{assignment.id}/cancel/',
            follow=True,
            HTTP_HOST='localhost',
        )
        assignment.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(assignment.status, AssignmentStatus.ACCEPTED)
        self.assertIsNone(assignment.ended_at)
        self.assertTrue(
            HaulAssignment.objects.filter(
                truck=truck,
                status=AssignmentStatus.PENDING,
                action='release',
                ended_at__isnull=True,
            ).exists()
        )

    def test_dispatcher_can_service_complete_active_trip_from_control_panel(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        rock = RockType.objects.create(name='Руда')
        dump_point = DumpPoint.objects.create(name='ККД')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        driver_role, _ = Role.objects.get_or_create(code='driver', defaults={'name': 'Водитель самосвала'})
        dispatcher = Employee.objects.create(full_name='Тестовый диспетчер')
        driver = Employee.objects.create(full_name='Тестовый водитель')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5000')
        EmployeeAccess.objects.create(employee=driver, role=driver_role, access_code='2102')
        EmployeeShift.objects.create(employee=dispatcher, shift_type='day', opened_at=timezone.now(), opened_by=dispatcher)
        unloading_shift = EmployeeShift.objects.create(
            employee=driver,
            shift_type='day',
            equipment=truck,
            opened_at=timezone.now(),
            opened_by=driver,
        )
        trip = Trip.objects.create(
            excavator=excavator,
            truck=truck,
            rock_type=rock,
            dump_point=dump_point,
            status=TripStatus.ACTIVE,
            volume_m3='11.00',
        )

        self.client.post('/', {'access_code': '5000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.post(
            f'/dispatcher/trips/{trip.id}/complete/',
            {'reason': 'Водитель потерял связь'},
            follow=True,
            HTTP_HOST='localhost',
        )
        trip.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(trip.status, TripStatus.COMPLETED)
        self.assertEqual(trip.driver, driver)
        self.assertEqual(trip.unloading_shift, unloading_shift)
        self.assertIsNotNone(trip.completed_at)
        self.assertContains(response, 'Выполненных рейсов пока нет.', count=0)
        action = DispatcherActionLog.objects.get()
        self.assertEqual(action.action_type, DispatcherActionType.COMPLETE_TRIP)
        self.assertEqual(action.reason, 'Водитель потерял связь')

    def test_dispatcher_cannot_service_complete_active_trip_without_open_shift(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        rock = RockType.objects.create(name='Руда')
        dump_point = DumpPoint.objects.create(name='ККД')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(full_name='Тестовый диспетчер')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5000')
        EmployeeShift.objects.create(employee=dispatcher, shift_type='day', opened_at=timezone.now(), opened_by=dispatcher)
        trip = Trip.objects.create(
            excavator=excavator,
            truck=truck,
            rock_type=rock,
            dump_point=dump_point,
            status=TripStatus.ACTIVE,
            volume_m3='11.00',
        )

        self.client.post('/', {'access_code': '5000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.post(
            f'/dispatcher/trips/{trip.id}/complete/',
            follow=True,
            HTTP_HOST='localhost',
        )
        trip.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(trip.status, TripStatus.ACTIVE)
        self.assertIsNone(trip.completed_at)

    def test_dispatcher_can_cancel_active_trip_from_control_panel(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        rock = RockType.objects.create(name='Руда')
        dump_point = DumpPoint.objects.create(name='ККД')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(full_name='Тестовый диспетчер')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5000')
        EmployeeShift.objects.create(employee=dispatcher, shift_type='day', opened_at=timezone.now(), opened_by=dispatcher)
        trip = Trip.objects.create(
            excavator=excavator,
            truck=truck,
            rock_type=rock,
            dump_point=dump_point,
            status=TripStatus.ACTIVE,
            volume_m3='11.00',
        )

        self.client.post('/', {'access_code': '5000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.post(
            f'/dispatcher/trips/{trip.id}/cancel/',
            {'reason': 'Ошибочно созданный рейс'},
            follow=True,
            HTTP_HOST='localhost',
        )
        trip.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(trip.status, TripStatus.CANCELLED)
        self.assertContains(response, 'Активных рейсов сейчас нет.')
        action = DispatcherActionLog.objects.get()
        self.assertEqual(action.action_type, DispatcherActionType.CANCEL_TRIP)
        self.assertEqual(action.reason, 'Ошибочно созданный рейс')

    def test_volume_report_can_filter_by_loading_shift_type(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        rock = RockType.objects.create(name='Руда')
        dump_point = DumpPoint.objects.create(name='ККД')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(full_name='Тестовый диспетчер')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5000')
        operator = Employee.objects.create(full_name='Машинист')
        closed_at = timezone.now()
        day_shift = EmployeeShift.objects.create(
            employee=operator,
            shift_type='day',
            opened_at=closed_at - timedelta(days=1),
            closed_at=closed_at - timedelta(hours=12),
        )
        night_shift = EmployeeShift.objects.create(
            employee=operator,
            shift_type='night',
            opened_at=closed_at - timedelta(hours=11),
            closed_at=closed_at,
        )
        Trip.objects.create(
            excavator=excavator,
            truck=truck,
            rock_type=rock,
            dump_point=dump_point,
            loading_shift=day_shift,
            status=TripStatus.COMPLETED,
            volume_m3='11.00',
            completed_at=timezone.now(),
        )
        Trip.objects.create(
            excavator=excavator,
            truck=truck,
            rock_type=rock,
            dump_point=dump_point,
            loading_shift=night_shift,
            status=TripStatus.COMPLETED,
            volume_m3='22.00',
            completed_at=timezone.now(),
        )

        self.client.post('/', {'access_code': '5000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get('/reports/volume/?loading_shift_type=day', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '11')
        self.assertNotContains(response, '22,00')

    def test_volume_report_can_group_by_completed_hour_and_export_excel(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        rock = RockType.objects.create(name='Руда')
        dump_point = DumpPoint.objects.create(name='ККД')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(full_name='Тестовый диспетчер')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5000')
        first_hour = timezone.make_aware(datetime(2026, 6, 17, 10, 15))
        second_hour = timezone.make_aware(datetime(2026, 6, 17, 11, 20))
        Trip.objects.create(
            excavator=excavator,
            truck=truck,
            rock_type=rock,
            dump_point=dump_point,
            status=TripStatus.COMPLETED,
            volume_m3='11.00',
            tonnage='25.00',
            completed_at=first_hour,
        )
        Trip.objects.create(
            excavator=excavator,
            truck=truck,
            rock_type=rock,
            dump_point=dump_point,
            status=TripStatus.COMPLETED,
            volume_m3='22.00',
            tonnage='50.00',
            completed_at=second_hour,
        )

        self.client.post('/', {'access_code': '5000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get('/reports/volume/?group_by=completed_hour', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<th>Час выполнения рейса</th>', html=True)
        self.assertContains(response, 'первая MVP-замена старой формы')
        self.assertContains(response, '10:00')
        self.assertContains(response, '11:00')
        self.assertContains(response, '<th>Рейсы</th>', html=True)

        export_response = self.client.get('/reports/volume/export/?group_by=completed_hour', HTTP_HOST='localhost')
        workbook = load_workbook(BytesIO(export_response.content))
        values = [
            cell
            for row in workbook.active.iter_rows(values_only=True)
            for cell in row
            if cell not in {None, ''}
        ]

        self.assertEqual(export_response.status_code, 200)
        self.assertIn('Час выполнения рейса', values)
        self.assertIn('10:00', values)
        self.assertIn('11:00', values)
        self.assertIn('Итого', values)

    def test_volume_report_uses_selected_report_template_columns(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        rock = RockType.objects.create(name='Руда')
        dump_point = DumpPoint.objects.create(name='ККД')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(full_name='Тестовый диспетчер')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5000')
        template = ReportTemplate.objects.create(
            name='Короткий отчет',
            report_type=ReportType.SHIFT_VOLUME,
            columns=['truck', 'planned_volume_m3', 'volume_m3', 'deviation_m3', 'plan_completion_percent'],
        )
        Trip.objects.create(
            excavator=excavator,
            truck=truck,
            rock_type=rock,
            dump_point=dump_point,
            status=TripStatus.COMPLETED,
            planned_volume_m3='10.00',
            volume_m3='11.00',
            completed_at=timezone.now(),
        )

        self.client.post('/', {'access_code': '5000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get(f'/reports/volume/?template={template.id}', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<th>Самосвал</th>', html=True)
        self.assertContains(response, '<th>План, м3</th>', html=True)
        self.assertContains(response, '<th>Объем, м3</th>', html=True)
        self.assertContains(response, '<th>Отклонение, м3</th>', html=True)
        self.assertContains(response, '<th>Выполнение, %</th>', html=True)
        self.assertNotContains(response, '<th>Экскаватор</th>', html=True)
        self.assertContains(response, '11')
        self.assertContains(response, '1,00')
        self.assertContains(response, '110,00')

        export_response = self.client.get(f'/reports/volume/export/?template={template.id}', HTTP_HOST='localhost')
        workbook = load_workbook(BytesIO(export_response.content))
        values = [
            cell
            for row in workbook.active.iter_rows(values_only=True)
            for cell in row
            if cell not in {None, ''}
        ]

        self.assertEqual(export_response.status_code, 200)
        self.assertIn('Отклонение, м3', values)
        self.assertIn('Выполнение, %', values)
        self.assertIn(Decimal('1.00'), values)
        self.assertIn(Decimal('110.00'), values)

    def test_dispatcher_can_create_report_template_in_builder(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        second_truck = Equipment.objects.create(equipment_type=truck_type, garage_number='11')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        rock = RockType.objects.create(name='Руда')
        dump_point = DumpPoint.objects.create(name='ККД')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(full_name='Тестовый диспетчер')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5000')
        Trip.objects.create(
            excavator=excavator,
            truck=truck,
            rock_type=rock,
            dump_point=dump_point,
            status=TripStatus.COMPLETED,
            volume_m3='11.00',
            completed_at=timezone.now(),
        )
        Trip.objects.create(
            excavator=excavator,
            truck=second_truck,
            rock_type=rock,
            dump_point=dump_point,
            status=TripStatus.COMPLETED,
            volume_m3='22.00',
            completed_at=timezone.now(),
        )

        self.client.post('/', {'access_code': '5000'}, follow=True, HTTP_HOST='localhost')
        builder_response = self.client.get('/reports/templates/', HTTP_HOST='localhost')

        self.assertEqual(builder_response.status_code, 200)
        self.assertContains(builder_response, 'Конструктор шаблонов отчетов')

        create_response = self.client.post(
            '/reports/templates/',
            {
                'name': 'Шаблон для заказчика',
                'columns': ['truck', 'volume_m3'],
                'column_label_truck': 'БелАЗ',
                'column_label_volume_m3': 'Факт, м3',
                'group_by': 'truck',
                'truck': str(truck.id),
                'is_active': 'on',
            },
            follow=True,
            HTTP_HOST='localhost',
        )
        template = ReportTemplate.objects.get(name='Шаблон для заказчика')

        self.assertEqual(create_response.status_code, 200)
        self.assertEqual(template.report_type, ReportType.SHIFT_VOLUME)
        self.assertEqual(template.columns, ['truck', 'volume_m3'])
        self.assertEqual(template.column_labels, {'truck': 'БелАЗ', 'volume_m3': 'Факт, м3'})
        self.assertEqual(template.filters, {'truck': str(truck.id)})
        self.assertEqual(template.group_by, 'truck')
        self.assertEqual(template.created_by, dispatcher)
        self.assertEqual(template.updated_by, dispatcher)

        report_response = self.client.get(f'/reports/volume/?template={template.id}', HTTP_HOST='localhost')

        self.assertEqual(report_response.status_code, 200)
        self.assertContains(report_response, '<th>БелАЗ</th>', html=True)
        self.assertContains(report_response, '<th>Факт, м3</th>', html=True)
        self.assertContains(report_response, '<th>Тоннаж</th>', html=True)
        self.assertContains(report_response, '<th>Рейсы</th>', html=True)
        self.assertNotContains(report_response, '<th>Экскаватор</th>', html=True)
        self.assertContains(report_response, '11')
        self.assertNotContains(report_response, '22,00')

        export_response = self.client.get(f'/reports/volume/export/?template={template.id}', HTTP_HOST='localhost')
        workbook = load_workbook(BytesIO(export_response.content))
        sheet = workbook.active
        values = [
            cell
            for row in sheet.iter_rows(values_only=True)
            for cell in row
            if cell not in {None, ''}
        ]

        self.assertEqual(export_response.status_code, 200)
        self.assertIn('Отчет по объемам', values)
        self.assertIn('Шаблон для заказчика', values)
        self.assertIn('БелАЗ', values)
        self.assertIn('Факт, м3', values)
        self.assertIn('Итого', values)
        self.assertNotIn(Decimal('22.00'), values)

    def test_dispatcher_can_open_customer_daily_report_and_export_it(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        rock = RockType.objects.create(name='Первичная сульфидная')
        dump_point = DumpPoint.objects.create(name='ККД')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(full_name='Тестовый диспетчер')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5000')
        operator = Employee.objects.create(full_name='Машинист')
        report_date = datetime(2026, 6, 17, 10, 0)
        report_datetime = timezone.make_aware(report_date)
        previous_datetime = report_datetime - timedelta(days=1)
        loading_shift = EmployeeShift.objects.create(
            employee=operator,
            shift_type='day',
            equipment=excavator,
            opened_at=report_datetime,
            closed_at=report_datetime + timedelta(hours=1),
        )
        Trip.objects.create(
            excavator=excavator,
            truck=truck,
            rock_type=rock,
            dump_point=dump_point,
            loading_shift=loading_shift,
            status=TripStatus.COMPLETED,
            planned_volume_m3='7000.00',
            volume_m3='57.00',
            tonnage='142.50',
            loading_horizon='75',
            loading_block='52',
            transport_distance_km='3.10',
            downtime_text='зачистка забоя',
            note='ожидание разгрузки',
            completed_at=report_datetime,
        )
        previous_shift = EmployeeShift.objects.create(
            employee=operator,
            shift_type='day',
            equipment=excavator,
            opened_at=previous_datetime,
            closed_at=previous_datetime + timedelta(hours=1),
        )
        Trip.objects.create(
            excavator=excavator,
            truck=truck,
            rock_type=rock,
            dump_point=dump_point,
            loading_shift=previous_shift,
            status=TripStatus.COMPLETED,
            planned_volume_m3='1000.00',
            volume_m3='100.00',
            tonnage='250.00',
            loading_horizon='75',
            loading_block='52',
            transport_distance_km='3.10',
            completed_at=previous_datetime,
        )

        self.client.post('/', {'access_code': '5000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get(
            f'/reports/customer-daily/?date={report_datetime:%Y-%m-%d}',
            HTTP_HOST='localhost',
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Суточный отчет заказчику')
        self.assertContains(response, 'Первичная сульфидная')
        self.assertContains(response, 'ККД')
        self.assertContains(response, '57')
        self.assertContains(response, '7000')
        self.assertContains(response, 'Отклонение')
        self.assertContains(response, '-6943')
        self.assertContains(response, '75')
        self.assertContains(response, '52')
        self.assertContains(response, '3,10')
        self.assertContains(response, 'зачистка забоя')
        self.assertContains(response, 'ожидание разгрузки')
        self.assertContains(response, 'С начала месяца')
        self.assertContains(response, '8000')
        self.assertContains(response, '157')
        self.assertContains(response, '-7843')
        self.assertContains(response, 'Сверка со старой Excel-формой заказчика')
        self.assertContains(response, 'Работа выемочного оборудования')
        self.assertContains(response, 'Средневзвешенное плечо')
        self.assertContains(response, 'Расчет выполненных работ по самосвалам')

        export_response = self.client.get(
            f'/reports/customer-daily/export/?date={report_datetime:%Y-%m-%d}',
            HTTP_HOST='localhost',
        )
        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(
            export_response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        workbook = load_workbook(BytesIO(export_response.content))
        self.assertIn('Сверка с Excel', workbook.sheetnames)
        reconciliation_values = [
            cell.value
            for row in workbook['Сверка с Excel'].iter_rows()
            for cell in row
        ]
        self.assertIn('Эталон для сверки: Отчет_Коппер. Рисорсез_Март.xlsx', reconciliation_values)
        self.assertIn('Работа выемочного оборудования', reconciliation_values)
        self.assertIn('Средневзвешенное плечо', reconciliation_values)

    def test_seed_demo_scenario_command_creates_ready_demo_data(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck_model = EquipmentModel.objects.create(equipment_type=truck_type, name='БЕЛАЗ тест')
        excavator_model = EquipmentModel.objects.create(equipment_type=excavator_type, name='Экскаватор тест')
        for garage_number in ('10', '11', '12'):
            Equipment.objects.create(
                equipment_type=truck_type,
                model=truck_model,
                garage_number=garage_number,
                is_active=True,
            )
        Equipment.objects.create(
            equipment_type=excavator_type,
            model=excavator_model,
            garage_number='1',
            is_active=True,
        )

        call_command('seed_demo_scenario')

        self.assertTrue(EmployeeAccess.objects.filter(access_code='200000', is_active=True).exists())
        self.assertTrue(EmployeeAccess.objects.filter(access_code='500000', is_active=True).exists())
        self.assertTrue(EmployeeAccess.objects.filter(access_code='600000', is_active=True).exists())
        self.assertTrue(EmployeeAccess.objects.filter(access_code='700000', is_active=True).exists())
        self.assertTrue(DriverPrimaryRegistration.objects.exists())
        self.assertTrue(EmployeeShift.objects.filter(closed_at__isnull=True).exists())
        self.assertTrue(HaulAssignment.objects.filter(status=AssignmentStatus.ACCEPTED).exists())
        self.assertTrue(Trip.objects.filter(status=TripStatus.LOADED_WAITING_UNLOAD).exists())
        self.assertTrue(DowntimeEvent.objects.filter(ended_at__isnull=True).exists())
        self.assertTrue(DowntimeEvent.objects.filter(ended_at__isnull=False).exists())
        self.assertTrue(DowntimeEvent.objects.filter(reason__is_critical=True, ended_at__isnull=True).exists())
        self.assertTrue(DowntimeEvent.objects.filter(reason__is_critical=False, ended_at__isnull=False).exists())
        self.assertTrue(ReportTemplate.objects.filter(name='Демо отчет по объемам', is_active=True).exists())
        self.assertTrue(PilotFeedback.objects.filter(title__startswith='Демо-замечание').exists())
        self.assertFalse(Equipment.objects.filter(garage_number__startswith='ДЕМО').exists())

    def test_seed_demo_scenario_reuses_reference_trucks_without_demo_trucks(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck_model = EquipmentModel.objects.create(equipment_type=truck_type, name='БЕЛАЗ тест')
        excavator_model = EquipmentModel.objects.create(equipment_type=excavator_type, name='Экскаватор тест')
        for garage_number in ('10', '11', '12'):
            Equipment.objects.create(
                equipment_type=truck_type,
                model=truck_model,
                garage_number=garage_number,
                is_active=True,
            )
        Equipment.objects.create(
            equipment_type=excavator_type,
            model=excavator_model,
            garage_number='1',
            is_active=True,
        )

        call_command('seed_demo_scenario')

        self.assertFalse(Equipment.objects.filter(garage_number__startswith='ДЕМО').exists())
        self.assertTrue(HaulAssignment.objects.filter(truck__garage_number='11', status=AssignmentStatus.PENDING).exists())
        self.assertTrue(Trip.objects.filter(truck__garage_number='12', status=TripStatus.COMPLETED).exists())

    def test_mechanic_opens_dashboard_and_creates_downtime_event(self):
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        rock = RockType.objects.create(name='Руда')
        dump_point = DumpPoint.objects.create(name='ККД')
        mechanic_role = Role.objects.create(code='mechanic', name='Механик')
        operator_role = Role.objects.create(code='excavator_operator', name='Машинист экскаватора')
        mechanic = Employee.objects.create(full_name='Тестовый механик')
        operator = Employee.objects.create(full_name='Тестовый машинист')
        EmployeeAccess.objects.create(employee=mechanic, role=mechanic_role, access_code='7000')
        EmployeeAccess.objects.create(employee=operator, role=operator_role, access_code='3000')
        trip = Trip.objects.create(
            excavator=excavator,
            truck=truck,
            rock_type=rock,
            dump_point=dump_point,
            excavator_operator=operator,
            status=TripStatus.ACTIVE,
            downtime_text='ожидание механика',
        )
        reason = DowntimeReason.objects.create(name='Тестовая диагностика механика', equipment_type=excavator_type, show_for_mechanic=True)

        login_response = self.client.post('/', {'access_code': '7000'}, follow=True, HTTP_HOST='localhost')
        dashboard_response = self.client.get('/mechanic/downtimes/', HTTP_HOST='localhost')
        create_response = self.client.post(
            f'/mechanic/downtimes/create/{trip.id}/',
            {
                f'trip_{trip.id}-reason': str(reason.id),
                f'trip_{trip.id}-comment': 'Выехали на диагностику',
            },
            follow=True,
            HTTP_HOST='localhost',
        )

        self.assertRedirects(login_response, '/mechanic/downtimes/', target_status_code=200)
        self.assertEqual(dashboard_response.status_code, 200)
        self.assertContains(dashboard_response, f'/mechanic/downtimes/create/{trip.id}/')
        self.assertEqual(create_response.status_code, 200)
        self.assertTrue(DowntimeEvent.objects.filter(equipment=excavator, ended_at__isnull=True).exists())
        event = DowntimeEvent.objects.get(equipment=excavator, ended_at__isnull=True)
        self.assertEqual(event.reason, reason)
        self.assertEqual(event.employee, mechanic)

    def test_mechanic_cannot_open_second_active_downtime_for_same_equipment(self):
        excavator_type = EquipmentType.objects.create(name='Excavator')
        truck_type = EquipmentType.objects.create(name='Truck')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        rock = RockType.objects.create(name='Rock')
        dump_point = DumpPoint.objects.create(name='Dump')
        mechanic_role = Role.objects.create(code='mechanic', name='Mechanic')
        operator_role = Role.objects.create(code='excavator_operator', name='Operator')
        mechanic = Employee.objects.create(full_name='Mechanic MVP')
        operator = Employee.objects.create(full_name='Operator MVP')
        EmployeeAccess.objects.create(employee=mechanic, role=mechanic_role, access_code='7000')
        reason = DowntimeReason.objects.create(name='Diagnostics', equipment_type=excavator_type, show_for_mechanic=True)
        existing_reason = DowntimeReason.objects.create(name='Engine', equipment_type=excavator_type)
        DowntimeEvent.objects.create(
            equipment=excavator,
            employee=mechanic,
            reason=existing_reason,
            started_at=timezone.now() - timedelta(minutes=20),
        )
        trip = Trip.objects.create(
            excavator=excavator,
            truck=truck,
            rock_type=rock,
            dump_point=dump_point,
            excavator_operator=operator,
            status=TripStatus.ACTIVE,
            downtime_text='waiting mechanic',
        )

        self.client.post('/', {'access_code': '7000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.post(
            f'/mechanic/downtimes/create/{trip.id}/',
            {
                f'trip_{trip.id}-reason': str(reason.id),
                f'trip_{trip.id}-comment': 'second event',
            },
            follow=True,
            HTTP_HOST='localhost',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(DowntimeEvent.objects.filter(equipment=excavator, ended_at__isnull=True).count(), 1)

    def test_mechanic_can_close_open_downtime_event(self):
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        mechanic_role = Role.objects.create(code='mechanic', name='Механик')
        mechanic = Employee.objects.create(full_name='Тестовый механик')
        EmployeeAccess.objects.create(employee=mechanic, role=mechanic_role, access_code='7000')
        reason = DowntimeReason.objects.create(name='Тестовый текущий ремонт', equipment_type=excavator_type)
        event = DowntimeEvent.objects.create(
            equipment=excavator,
            employee=mechanic,
            reason=reason,
            started_at=timezone.now() - timedelta(minutes=25),
            comment='Проверка',
        )

        self.client.post('/', {'access_code': '7000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.post(
            f'/mechanic/downtimes/{event.id}/close/',
            follow=True,
            HTTP_HOST='localhost',
        )
        event.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertIsNotNone(event.ended_at)
        self.assertFalse(DowntimeEvent.objects.filter(id=event.id, ended_at__isnull=True).exists())

    def test_mechanic_dashboard_shows_recent_closed_downtimes_with_duration(self):
        excavator_type = EquipmentType.objects.create(name='Excavator')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        mechanic_role = Role.objects.create(code='mechanic', name='Mechanic')
        mechanic = Employee.objects.create(full_name='Mechanic MVP')
        EmployeeAccess.objects.create(employee=mechanic, role=mechanic_role, access_code='7000')
        reason = DowntimeReason.objects.create(
            name='Hydraulics',
            equipment_type=excavator_type,
            show_for_mechanic=True,
        )
        DowntimeEvent.objects.create(
            equipment=excavator,
            employee=mechanic,
            reason=reason,
            started_at=timezone.now() - timedelta(hours=1, minutes=30),
            ended_at=timezone.now(),
            comment='Closed downtime',
        )

        self.client.post('/', {'access_code': '7000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get('/mechanic/downtimes/', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Hydraulics')
        self.assertContains(response, 'Closed downtime')
        self.assertContains(response, '1 ч 30 мин')

    def test_downtime_daily_summary_page_and_excel_use_production_date_after_midnight(self):
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='СВОДКА-23')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(full_name='Диспетчер сводки')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='7099')
        reason = DowntimeReason.objects.create(
            name='Сводка после полуночи',
            equipment_type=excavator_type,
        )
        started_at = datetime(2026, 7, 24, 1, 30, tzinfo=ZoneInfo('Asia/Vladivostok'))
        DowntimeEvent.objects.create(
            equipment=excavator,
            employee=dispatcher,
            reason=reason,
            started_at=started_at,
            ended_at=started_at + timedelta(hours=1),
        )
        self.client.post('/', {'access_code': '7099'}, follow=True, HTTP_HOST='localhost')
        query = {'date_from': '2026-07-23', 'date_to': '2026-07-23'}

        page = self.client.get('/reports/downtimes/', query, HTTP_HOST='localhost')
        export = self.client.get('/reports/downtimes/export/', query, HTTP_HOST='localhost')

        self.assertEqual(page.status_code, 200)
        self.assertEqual(page.context['daily_summary'][0]['date'], datetime(2026, 7, 23).date())
        workbook = load_workbook(BytesIO(export.content), data_only=True)
        summary_dates = [
            workbook.active.cell(row=row, column=4).value
            for row in range(14, 25)
        ]
        self.assertIn('23.07.2026', summary_dates)
        self.assertNotIn('24.07.2026', summary_dates)

    def test_downtime_report_filters_open_events_and_exports_excel(self):
        excavator_type = EquipmentType.objects.create(name='Excavator')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        mechanic_role = Role.objects.create(code='mechanic', name='Mechanic')
        mechanic = Employee.objects.create(full_name='Mechanic MVP')
        EmployeeAccess.objects.create(employee=mechanic, role=mechanic_role, access_code='7000')
        open_reason = DowntimeReason.objects.create(name='Open diagnostics', equipment_type=excavator_type)
        closed_reason = DowntimeReason.objects.create(name='Closed repair', equipment_type=excavator_type)
        started_at = timezone.make_aware(datetime(2026, 6, 17, 9, 0))
        DowntimeEvent.objects.create(
            equipment=excavator,
            employee=mechanic,
            reason=open_reason,
            started_at=started_at,
            comment='Hidden old open event',
        )
        for index in range(200):
            DowntimeEvent.objects.create(
                equipment=excavator,
                employee=mechanic,
                reason=open_reason,
                started_at=started_at + timedelta(minutes=index + 1),
                comment=f'Visible open event {index}',
            )
        DowntimeEvent.objects.create(
            equipment=excavator,
            employee=mechanic,
            reason=closed_reason,
            started_at=started_at,
            ended_at=started_at + timedelta(hours=1),
            comment='Closed event',
        )

        self.client.post('/', {'access_code': '7000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get('/reports/downtimes/?status=open&date_from=2026-06-17&date_to=2026-06-17', HTTP_HOST='localhost')
        export_response = self.client.get('/reports/downtimes/export/?status=open&date_from=2026-06-17&date_to=2026-06-17', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '/reports/customer-daily/?date=2026-06-17')
        self.assertContains(response, 'Open diagnostics')
        self.assertContains(response, 'Visible open event 199')
        self.assertContains(response, '200 из 201')
        self.assertNotContains(response, 'Hidden old open event')
        self.assertContains(response, '17.06.2026')
        self.assertNotContains(response, 'Closed event')
        self.assertEqual(
            export_response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        workbook = load_workbook(BytesIO(export_response.content))
        values = [cell.value for row in workbook.active.iter_rows() for cell in row]
        self.assertIn('Сводка по датам', values)
        self.assertIn('17.06.2026', values)
        self.assertIn('Open diagnostics', values)
        self.assertIn('Hidden old open event', values)
        self.assertIn('Visible open event 199', values)
        self.assertNotIn('Closed repair', values)

        critical_reason = DowntimeReason.objects.create(name='Critical engine', equipment_type=excavator_type, is_critical=True)
        DowntimeEvent.objects.create(
            equipment=excavator,
            employee=mechanic,
            reason=critical_reason,
            started_at=started_at + timedelta(hours=5),
            comment='Critical only',
        )
        critical_response = self.client.get('/reports/downtimes/?critical=yes', HTTP_HOST='localhost')
        self.assertContains(critical_response, 'Critical engine')
        self.assertContains(critical_response, 'Critical only')
        self.assertNotContains(critical_response, 'Visible open event 199')

    def test_downtime_report_shows_unloading_waiting_reconciliation(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        truck_10 = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        truck_25 = Equipment.objects.create(equipment_type=truck_type, garage_number='25')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Диспетчер')
        dispatcher = Employee.objects.create(full_name='Тестовый диспетчер')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5000')
        kkd_reason = DowntimeReason.objects.get(name='Ожидание разгрузки ККД')
        skdr_reason = DowntimeReason.objects.get(name='Ожидание разгрузки СКДР')
        kkd_reason.equipment_type = truck_type
        skdr_reason.equipment_type = truck_type
        kkd_reason.save(update_fields=['equipment_type'])
        skdr_reason.save(update_fields=['equipment_type'])
        started_at = timezone.make_aware(datetime(2026, 6, 17, 9, 0))
        DowntimeEvent.objects.create(
            equipment=truck_10,
            employee=dispatcher,
            reason=kkd_reason,
            started_at=started_at,
            ended_at=started_at + timedelta(minutes=45),
            comment='Очередь на ККД',
        )
        DowntimeEvent.objects.create(
            equipment=truck_25,
            employee=dispatcher,
            reason=skdr_reason,
            started_at=started_at + timedelta(hours=1),
            ended_at=started_at + timedelta(hours=1, minutes=30),
            comment='Очередь на СКДР',
        )

        self.client.post('/', {'access_code': '5000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get('/reports/downtimes/?date_from=2026-06-17&date_to=2026-06-17', HTTP_HOST='localhost')
        export_response = self.client.get('/reports/downtimes/export/?date_from=2026-06-17&date_to=2026-06-17', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Сверка ОР ККД/СКДР')
        self.assertContains(response, 'Покрытие старой формы ОР ККД/СКДР')
        self.assertContains(response, 'Ожидание разгрузки ККД')
        self.assertContains(response, 'Ожидание разгрузки СКДР')
        self.assertContains(response, '75,00 мин')
        workbook = load_workbook(BytesIO(export_response.content))
        self.assertIn('ОР ККД СКДР', workbook.sheetnames)
        values = [cell.value for row in workbook['ОР ККД СКДР'].iter_rows() for cell in row]
        self.assertIn('Сверка ожидания разгрузки ККД/СКДР', values)
        self.assertIn('Ожидание разгрузки ККД', values)
        self.assertIn('Ожидание разгрузки СКДР', values)
        self.assertIn('Источник старой формы: ОР ККД СКДР март.xlsx', values)

    def test_dispatcher_control_shows_open_mechanic_downtimes(self):
        excavator_type = EquipmentType.objects.create(name='Excavator')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Dispatcher')
        mechanic_role = Role.objects.create(code='mechanic', name='Mechanic')
        dispatcher = Employee.objects.create(full_name='Dispatcher MVP')
        mechanic = Employee.objects.create(full_name='Mechanic MVP')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5000')
        reason = DowntimeReason.objects.create(name='Diagnostics', equipment_type=excavator_type)
        DowntimeEvent.objects.create(
            equipment=excavator,
            employee=mechanic,
            reason=reason,
            started_at=timezone.now() - timedelta(minutes=10),
            comment='Check hydraulics',
        )

        self.client.post('/', {'access_code': '5000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get('/dispatcher/control/', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '/mechanic/downtimes/')
        self.assertContains(response, 'Diagnostics')
        self.assertContains(response, 'Check hydraulics')

    def test_customer_daily_report_shows_mechanic_downtimes_and_exports_them(self):
        excavator_type = EquipmentType.objects.create(name='Excavator')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        dispatcher_role = Role.objects.create(code='dispatcher', name='Dispatcher')
        mechanic_role = Role.objects.create(code='mechanic', name='Mechanic')
        dispatcher = Employee.objects.create(full_name='Dispatcher MVP')
        mechanic = Employee.objects.create(full_name='Mechanic MVP')
        EmployeeAccess.objects.create(employee=dispatcher, role=dispatcher_role, access_code='5000')
        reason = DowntimeReason.objects.create(
            name='Hydraulics',
            equipment_type=excavator_type,
            show_for_mechanic=True,
        )
        started_at = timezone.make_aware(datetime(2026, 6, 17, 9, 0))
        ended_at = timezone.make_aware(datetime(2026, 6, 17, 10, 30))
        DowntimeEvent.objects.create(
            equipment=excavator,
            employee=mechanic,
            reason=reason,
            started_at=started_at,
            ended_at=ended_at,
            comment='Replace hose',
        )

        self.client.post('/', {'access_code': '5000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get('/reports/customer-daily/?date=2026-06-17', HTTP_HOST='localhost')
        export_response = self.client.get('/reports/customer-daily/export/?date=2026-06-17', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Hydraulics')
        self.assertContains(response, 'Replace hose')
        workbook = load_workbook(BytesIO(export_response.content))
        sheet = workbook.active
        values = [cell.value for row in sheet.iter_rows() for cell in row]
        self.assertIn('Hydraulics', values)
        self.assertIn('Replace hose', values)
        self.assertIn('I смена (дневная 07:00 - 19:00)', values)
        self.assertIn('II смена (ночная 19:00 - 07:00)', values)
        self.assertNotIn('I смена (дневная 08:00 - 20:00)', values)

    def test_management_dashboard_shows_open_mechanic_downtimes(self):
        excavator_type = EquipmentType.objects.create(name='Excavator')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        manager_role = Role.objects.create(code='manager', name='Manager')
        mechanic_role = Role.objects.create(code='mechanic', name='Mechanic')
        manager = Employee.objects.create(full_name='Manager MVP')
        mechanic = Employee.objects.create(full_name='Mechanic MVP')
        EmployeeAccess.objects.create(employee=manager, role=manager_role, access_code='6000')
        reason = DowntimeReason.objects.create(
            name='Engine',
            equipment_type=excavator_type,
            show_for_mechanic=True,
        )
        DowntimeEvent.objects.create(
            equipment=excavator,
            employee=mechanic,
            reason=reason,
            started_at=timezone.now() - timedelta(minutes=40),
            comment='No start',
        )

        self.client.post('/', {'access_code': '6000'}, follow=True, HTTP_HOST='localhost')
        response = self.client.get('/reports/management/', HTTP_HOST='localhost')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Engine')
        self.assertContains(response, 'No start')
        self.assertContains(response, '/mechanic/downtimes/')

    def test_manager_opens_management_dashboard(self):
        truck_type = EquipmentType.objects.create(name='Самосвал')
        excavator_type = EquipmentType.objects.create(name='Экскаватор')
        truck = Equipment.objects.create(equipment_type=truck_type, garage_number='10')
        excavator = Equipment.objects.create(equipment_type=excavator_type, garage_number='1')
        rock = RockType.objects.create(name='Руда')
        dump_point = DumpPoint.objects.create(name='ККД')
        manager_role = Role.objects.create(code='manager', name='Руководство')
        manager = Employee.objects.create(full_name='Тестовое руководство')
        operator = Employee.objects.create(full_name='Тестовый машинист')
        EmployeeAccess.objects.create(employee=manager, role=manager_role, access_code='6000')
        report_datetime = timezone.make_aware(datetime(2026, 6, 17, 10, 0))
        previous_datetime = report_datetime - timedelta(days=1)
        day_shift = EmployeeShift.objects.create(
            employee=operator,
            shift_type='day',
            equipment=excavator,
            opened_at=report_datetime,
        )
        Trip.objects.create(
            excavator=excavator,
            truck=truck,
            rock_type=rock,
            dump_point=dump_point,
            status=TripStatus.COMPLETED,
            loading_shift=day_shift,
            planned_volume_m3='60.00',
            volume_m3='57.00',
            tonnage='142.50',
            completed_at=report_datetime,
        )
        Trip.objects.create(
            excavator=excavator,
            truck=truck,
            rock_type=rock,
            dump_point=dump_point,
            status=TripStatus.COMPLETED,
            planned_volume_m3='20.00',
            volume_m3='22.00',
            tonnage='55.00',
            completed_at=previous_datetime,
        )

        login_response = self.client.post('/', {'access_code': '6000'}, follow=True, HTTP_HOST='localhost')
        dashboard_response = self.client.get('/reports/management/?date=2026-06-17', HTTP_HOST='localhost')
        export_response = self.client.get('/reports/management/export/?date=2026-06-17', HTTP_HOST='localhost')

        self.assertRedirects(login_response, '/reports/management/', target_status_code=200)
        self.assertEqual(dashboard_response.status_code, 200)
        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(
            export_response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertContains(dashboard_response, 'Витрина руководства')
        self.assertContains(dashboard_response, 'Выгрузить витрину в Excel')
        self.assertContains(dashboard_response, 'Чеклист пилотной проверки')
        self.assertContains(dashboard_response, '/reports/pilot-checklist/')
        self.assertContains(dashboard_response, 'Журнал замечаний пилота')
        self.assertContains(dashboard_response, '/reports/pilot-feedback/')
        self.assertContains(dashboard_response, 'Факт за сутки')
        self.assertContains(dashboard_response, 'План за сутки')
        self.assertContains(dashboard_response, 'Выполнение плана')
        self.assertContains(dashboard_response, 'Отклонение за сутки')
        self.assertContains(dashboard_response, 'День против ночи')
        self.assertContains(dashboard_response, 'Дневная смена')
        self.assertContains(dashboard_response, 'Ночная смена')
        self.assertContains(dashboard_response, 'Динамика за 7 дней')
        self.assertContains(dashboard_response, 'Итог за 7 дней')
        self.assertContains(dashboard_response, 'План за 7 дней')
        self.assertContains(dashboard_response, 'Выполнение за неделю')
        self.assertContains(dashboard_response, 'Лучший день')
        self.assertContains(dashboard_response, 'Самая сильная просадка')
        self.assertContains(dashboard_response, '16.06')
        self.assertContains(dashboard_response, '17.06')
        self.assertContains(dashboard_response, '57,00')
        self.assertContains(dashboard_response, '60,00')
        self.assertContains(dashboard_response, '95,0%')
        self.assertContains(dashboard_response, '-3,00')
        self.assertContains(dashboard_response, '22,00')
        self.assertContains(dashboard_response, '110,0%')
        self.assertContains(dashboard_response, 'Рейсы за сутки')
        self.assertContains(dashboard_response, 'Экскаваторы за сутки')
        self.assertContains(dashboard_response, 'Породы и грузы за сутки')
        self.assertContains(dashboard_response, 'Общая накопленная картина')
        self.assertContains(dashboard_response, '79 м3')
        self.assertContains(dashboard_response, '80,00 м3')
        self.assertContains(dashboard_response, '98,8%')
        self.assertContains(dashboard_response, '57,00')
        self.assertContains(dashboard_response, '142,50')
        workbook = load_workbook(BytesIO(export_response.content))
        self.assertIn('Сводка', workbook.sheetnames)
        self.assertIn('Динамика 7 дней', workbook.sheetnames)
        self.assertIn('День ночь', workbook.sheetnames)
        values = [cell.value for sheet in workbook.worksheets for row in sheet.iter_rows() for cell in row]
        self.assertIn('Витрина руководства', values)
        self.assertIn('Факт за 7 дней, м3', values)
        self.assertIn('Выполнение за неделю, %', values)
        self.assertIn('Дневная смена', values)
        self.assertIn(Decimal('79.00'), values)
        self.assertIn(98.8, values)

# Create your tests here.


import secrets
import json
from datetime import datetime, timedelta
from contextlib import nullcontext
from decimal import Decimal
from io import BytesIO
from urllib.parse import urlencode

from django.conf import settings
from django.contrib import messages
from django.core.exceptions import FieldDoesNotExist, ValidationError
from django.forms import modelform_factory
from django.forms.models import construct_instance
from django.db import transaction
from django.db.models import Count, Q
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404
from django.shortcuts import redirect, render
from django.templatetags.static import static
from django.urls import reverse
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.http import require_GET, require_POST
from openpyxl import Workbook

from assignments.models import AssignmentStatus, ExcavatorPlacement, HaulAssignment, HaulAssignmentAction
from assignments.services import (
    WORK_ASSIGNMENT_ROLE_EQUIPMENT_TYPES,
    apply_pending_haul_assignment,
    clear_active_equipment_assignment,
    get_active_equipment_assignment,
    reconcile_due_haul_assignments,
    work_assignment_state,
)
from core.models import OperationalStateEvent, OperationalStateVersion, bump_operational_state
from core.operational_fragments import operational_fragment_response
from downtimes.driver_workflow import (
    DRIVER_DOWNTIME_FLOW_WAITING_LOADING,
    DRIVER_DOWNTIME_FLOW_WAITING_UNLOAD,
    driver_downtime_flow,
    driver_downtime_opens_work,
    driver_downtime_requires_empty_truck,
    driver_downtime_requires_loaded_trip,
)
from downtimes.models import DowntimeEvent, DowntimeReason
from references.models import Dormitory, DormitorySection, DumpPoint, Equipment, EquipmentState, EquipmentType, RockType
from reports.forms import RatingPeriodReferenceForm
from reports.models import RatingPeriod, ReportTemplate
from reports.rating_period_generation import inspect_rating_period_calendar
from shifts.forms import EquipmentPlanGroupForm
from shifts.models import (
    AchievementPrize,
    EmployeeShift,
    EquipmentPlanGroup,
    EquipmentShiftPlan,
    PlanAssignmentStatus,
    PlanCalculationMode,
    ShiftClientAction,
    ShiftPlan,
    ShiftPlanScope,
)
from shifts.services import (
    calculate_truck_shift_progress,
    close_driver_shift,
    open_driver_shift,
    open_shift_conflict_message,
    plan_status_label,
    plan_unit_label,
    progress_cycle_visual_context,
    recent_shift_reading_corrections,
)
from trips.models import DispatcherActionLog, OPEN_TRIP_STATUSES, Trip, TripClientAction, TripStatus

from .access_auth import (
    format_phone_for_display,
    find_employee_access_by_credentials,
    find_unactivated_accesses_by_phone,
)
from .app_catalog import (
    APP_CATALOG_ROLE_CODES,
    app_catalog_public_url,
    app_catalog_items,
    role_app_qr_asset_path,
)
from .active_role import (
    ACTIVE_ROLE_CODE_SESSION_KEY,
    ACTIVE_ROLE_GENERATION_SESSION_KEY,
    ACTIVE_ROLE_SESSION_KEY,
    activate_role_session,
    role_session_state,
)
from .protected_cards import (
    PROTECTED_WRITE_CODE,
    allow_protected_card_write,
)
from .privacy_consent import (
    PRIVACY_CONSENT_FIELD,
    PRIVACY_CONSENT_REQUIRED_MESSAGE,
    accept_current_privacy_policy,
    privacy_consent_matches_access,
    privacy_consent_submission_is_current,
)
from .forms import (
    AdminAccessBlockForm,
    AccessActivationForm,
    AdminAccessRoleForm,
    AdminEmployeeEditForm,
    AdminEmployeeForm,
    DriverCloseShiftForm,
    DriverOpenShiftForm,
    DriverPrimaryRegistrationForm,
    is_valid_russian_mobile_phone,
    normalize_phone,
    PersonnelPositionReferenceForm,
)
from .models import (
    AdminActionLog,
    AdminConflict,
    DriverPrimaryRegistration,
    Employee,
    EmployeeAccess,
    PersonnelDepartment,
    PersonnelPosition,
    ProductionSpecialization,
    Role,
    TemporaryWorkTransfer,
    WatchComposition,
    WorkSchedule,
)
from .oup_undo import (
    get_oup_action_undo_state,
    undo_oup_action,
)
from .role_apps import (
    APP_CONTRACT_VERSION,
    get_role_app,
    get_role_app_for_request,
    role_app_manifest_response,
    role_app_service_worker_response,
)
from .session_device import (
    detect_session_device_kind,
    get_session_device_kind,
    set_session_device_kind,
)
from .work_profiles import (
    cancel_temporary_work_transfer,
    effective_specialization,
    employee_has_effective_access_role,
    expire_due_temporary_work_transfers,
    production_access_is_out_of_sync,
    sync_employee_production_access,
)


ROLE_INTERFACE_NAMES = {
    'admin': 'Админка',
    'driver': 'Интерфейс водителя самосвала',
    'excavator_operator': 'Интерфейс машиниста экскаватора',
    'mining_master': 'Интерфейс горного мастера',
    'deputy_mining_manager': 'Планирование смены зам. начальника горного участка',
    'oup': 'Рабочее место ОУП',
    'timekeeper': 'Рабочее место табельщика',
    'site_manager': 'Согласование продлений начальником участка',
    'employee_portal': 'Личные запросы сотрудника',
    'dispatcher': 'Диспетчерский экран',
    'mechanic': 'Интерфейс механика',
    'manager': 'Витрина руководства',
    'settlement_clerk': 'Рабочее место делопроизводителя',
}


ADMIN_RESTORABLE_EMPLOYEE_STATUSES = {
    Employee.Status.DEACTIVATED,
    Employee.Status.ARCHIVED,
    Employee.Status.DISMISSED,
    Employee.Status.DELETED,
}


INTERFACE_MAP = [
    {
        'section': 'Вход и администрирование',
        'items': [
            {'title': 'Единый вход', 'url': '/', 'code': 'любой демо-код', 'note': 'Открывает интерфейс по роли'},
            {'title': 'Карта интерфейсов', 'url': '/interfaces/', 'code': '-', 'note': 'Все готовые экраны MVP в одном месте'},
            {'title': 'Админка MVP', 'url': '/system-admin/', 'code': '1000', 'note': 'Сотрудники, доступы, справочники, конфликты и выгрузки'},
            {'title': 'Сотрудники админки', 'url': '/system-admin/employees/', 'code': '1000', 'note': 'Список сотрудников, фильтр по статусу, карточки и Excel'},
            {'title': 'Справочники админки', 'url': '/system-admin/references/', 'code': '1000', 'note': 'Единый реестр справочников первого этапа'},
            {'title': 'Конфликты админки', 'url': '/system-admin/conflicts/', 'code': '1000', 'note': 'Заблокированные рискованные действия и причины'},
            {'title': 'Журнал действий админки', 'url': '/system-admin/logs/', 'code': '1000', 'note': 'История важных административных действий'},
            {'title': 'Django-админка', 'url': '/admin/', 'code': 'администратор Django', 'note': 'Техническое управление справочниками и данными'},
        ],
    },
    {
        'section': 'Рабочие интерфейсы',
        'items': [
            {'title': 'Работа водителя самосвала', 'url': '/driver/', 'code': '2000', 'note': 'Главный PWA-экран Работа, смена, простои и путевка'},
            {'title': 'Машинист экскаватора', 'url': '/excavator/work/', 'code': '3000', 'note': 'Создание рейса и параметры для отчета заказчику'},
            {'title': 'Горный мастер', 'url': '/mining-master/assignments/', 'code': '4000', 'note': 'Назначение самосвалов под экскаваторы'},
            {'title': 'Зам. начальника горного участка', 'url': '/deputy-mining-manager/', 'code': 'роль зам. начальника', 'note': 'Расстановка сотрудников по технике на две смены'},
            {'title': 'Отдел управления персоналом', 'url': '/oup/', 'code': '800000 / роль ОУП', 'note': 'Создание, ведение и увольнение сотрудников'},
            {'title': 'Табельщик', 'url': '/timekeeper/', 'code': '900000 / роль табельщика', 'note': 'Сбор данных перевахты, Excel и оформление одобренных продлений'},
            {'title': 'Начальник участка', 'url': '/site-manager/extensions/', 'code': '910000 / роль начальника участка', 'note': 'Согласование заявок на продление вахты'},
            {'title': 'Ответ сотрудника по перевахте', 'url': '/my/rotation/', 'code': '920000 / любой активный доступ', 'note': 'Маршрут, даты, смена или запрос на продление'},
            {'title': 'Диспетчерский пульт', 'url': '/dispatcher/control/', 'code': '5000', 'note': 'Контроль активных рейсов и назначений'},
            {'title': 'Механическая служба', 'url': '/mechanic/downtimes/', 'code': '7000 / роль механика', 'note': 'Открытие и закрытие механических простоев по технике'},
        ],
    },
    {
        'section': 'Отчеты и руководство',
        'items': [
            {'title': 'Отчет по объемам', 'url': '/reports/volume/', 'code': '5000 / 6000', 'note': 'Фильтры, шаблоны, группировки и Excel'},
            {'title': 'Конструктор шаблонов отчетов', 'url': '/reports/templates/', 'code': '5000 / 1000', 'note': 'Столбцы, названия, фильтры, группировки, расчетные поля'},
            {'title': 'Суточный отчет заказчику', 'url': '/reports/customer-daily/', 'code': '5000 / 6000', 'note': 'Суточный отчет к 08:00 и Excel-выгрузка'},
            {'title': 'Отчет по механическим простоям', 'url': '/reports/downtimes/', 'code': '5000 / 6000 / 7000', 'note': 'Фильтры по датам, технике, причине, статусу и Excel'},
            {'title': 'Витрина руководства', 'url': '/reports/management/', 'code': '6000', 'note': 'Суточный срез, накопленная картина и показатели'},
            {'title': 'Excel-выгрузка витрины руководства', 'url': '/reports/management/export/', 'code': '6000', 'note': 'Сводка, динамика за 7 дней и сравнение день/ночь в Excel'},
            {'title': 'Чеклист пилотной проверки отчетов', 'url': '/reports/pilot-checklist/', 'code': '5000 / 6000 / 1000', 'note': 'Рабочая навигация перед пилотом: экраны, Excel-выгрузки и вопросы для сверки с текущими отчетами'},
            {'title': 'Сценарий пилотного запуска', 'url': '/reports/pilot-scenario/', 'code': '5000 / 6000 / 1000', 'note': 'Пошаговая проверка пилота по ролям: от расстановки и рейса до отчетов и витрины'},
            {'title': 'Журнал замечаний пилота', 'url': '/reports/pilot-feedback/', 'code': '5000 / 6000 / 1000', 'note': 'Фиксация замечаний, приоритетов, решений и переносов во время пилотной проверки'},
        ],
    },
]


DEMO_ACCESS_CODES = [
    ('+79000000001', '100000', 'Администратор'),
    ('+79000000002', '200000', 'Водитель самосвала'),
    ('+79000000003', '300000', 'Машинист экскаватора'),
    ('+79000000004', '400000', 'Горный мастер'),
    ('+79000000005', '500000', 'Диспетчер'),
    ('+79000000007', '700000', 'Механик'),
    ('+79000000006', '600000', 'Руководство'),
    ('+79000000008', '800000', 'Специалист ОУП'),
    ('+79000000009', '900000', 'Табельщик'),
    ('+79000000010', '910000', 'Начальник участка'),
    ('+79000000011', '920000', 'Сотрудник'),
]


DRIVER_SHELL_VERSION = 'driver-mobile-shell-v191'

DRIVER_MANIFEST = {
    'id': '/driver/',
    'name': 'Водитель самосвала',
    'short_name': 'Водитель',
    'description': 'Мобильное рабочее место водителя самосвала: работа, смена, простои и путевка.',
    'start_url': '/driver/',
    'scope': '/driver/',
    'display': 'standalone',
    'display_override': ['standalone', 'fullscreen'],
    'orientation': 'portrait',
    'background_color': '#030708',
    'theme_color': '#030708',
    'categories': ['business', 'productivity'],
    'icons': [
        {
            'src': '/static/img/pwa/driver-180.png',
            'sizes': '180x180',
            'type': 'image/png',
        },
        {
            'src': '/static/img/pwa/driver-192.png',
            'sizes': '192x192',
            'type': 'image/png',
        },
        {
            'src': '/static/img/pwa/driver-512.png',
            'sizes': '512x512',
            'type': 'image/png',
        },
        {
            'src': '/static/img/pwa/driver-maskable-512.png',
            'sizes': '512x512',
            'type': 'image/png',
            'purpose': 'maskable',
        },
    ],
}

DRIVER_SERVICE_WORKER_JS = f"""
const APP_CONTRACT_VERSION = {json.dumps(APP_CONTRACT_VERSION)};
const ROLE_CODE = "driver";
const CACHE_NAME = "{DRIVER_SHELL_VERSION}";
const CACHE_PREFIX = "driver-mobile-shell-";
const APP_SHELL_URL = "/driver/";
const LEGACY_SHELL_URL = "/driver/shift/";
const MANIFEST_URL = "/driver.webmanifest";
const CORE_ASSETS = [
    APP_SHELL_URL,
    LEGACY_SHELL_URL,
    MANIFEST_URL,
    "/company/privacy/",
    "/static/portal/css/portal-shell-v5.css?v=6",
    "/static/portal/js/portal-shell-v5.js",
    "/static/css/app.css",
    "/static/css/mobile-role-login-v1.css",
    "/static/css/mobile-shift-unified-v1.css",
    "/static/js/mobile-shift-unified-v1.js",
    "/static/js/mobile-operational-sounds-v1.js",
    "/static/css/native-app-update-v1.css",
  "/static/js/realtime-client.js",
  "/static/js/role-readonly.js",
    "/static/favicon.ico",
    "/static/img/pwa/driver-180.png",
    "/static/img/pwa/driver-192.png",
    "/static/img/pwa/driver-512.png",
    "/static/img/pwa/driver-maskable-512.png",
    "/static/img/start/start-hero-v1.webp",
    "/static/img/start/start-hero-v1.jpg",
    "/static/audio/driver/driver_truck_assigned.wav",
    "/static/audio/driver/driver_action_ok.wav",
    "/static/audio/driver/driver_action_error.wav",
    "/static/audio/driver/driver_connection_lost.wav",
    "/static/audio/driver/driver_connection_restored.wav",
    "/static/audio/driver/driver_shift_start.wav",
    "/static/audio/driver/driver_shift_end.wav"
];

self.addEventListener("install", (event) => {{
    event.waitUntil(
        caches.open(CACHE_NAME)
            .then((cache) => cache.addAll(CORE_ASSETS))
            .then(() => self.skipWaiting())
    );
}});

self.addEventListener("activate", (event) => {{
    event.waitUntil(
        caches.keys().then((keys) => Promise.all(
            keys
                .filter((key) => key.startsWith(CACHE_PREFIX) && key !== CACHE_NAME)
                .map((key) => caches.delete(key))
        )).then(() => self.clients.claim())
    );
}});

async function networkFirst(request, fallbackUrl) {{
    const cache = await caches.open(CACHE_NAME);
    try {{
        const freshRequest = new Request(request, {{ cache: "no-store" }});
        const response = await fetch(freshRequest);
        if (response && response.ok) {{
            cache.put(request, response.clone());
        }}
        return response;
    }} catch (error) {{
        return (await cache.match(request)) || (fallbackUrl ? cache.match(fallbackUrl) : undefined) || Response.error();
    }}
}}

async function networkFirstStatic(request) {{
    const cache = await caches.open(CACHE_NAME);
    try {{
        const response = await fetch(request, {{ cache: "no-store" }});
        if (response && response.ok) {{
            await cache.put(request, response.clone());
        }}
        return response;
    }} catch (error) {{
        return (await cache.match(request)) || new Response(
            "Ресурс недоступен без сети.",
            {{ status: 503, headers: {{ "Content-Type": "text/plain; charset=utf-8" }} }}
        );
    }}
}}

self.addEventListener("fetch", (event) => {{
    const request = event.request;
    if (request.method !== "GET") {{
        return;
    }}
    const url = new URL(request.url);
    if (url.origin !== self.location.origin) {{
        return;
    }}
    if (request.headers.get("x-requested-with") === "XMLHttpRequest") {{
        event.respondWith(fetch(request));
        return;
    }}
    if (request.mode === "navigate" || url.pathname === APP_SHELL_URL || url.pathname === LEGACY_SHELL_URL) {{
        event.respondWith(networkFirst(request, APP_SHELL_URL));
        return;
    }}
    if (url.pathname === MANIFEST_URL) {{
        event.respondWith(networkFirst(request, MANIFEST_URL));
        return;
    }}
    if (url.pathname.startsWith("/static/")) {{
        event.respondWith(networkFirstStatic(request));
    }}
}});

self.addEventListener("message", (event) => {{
    if (!event.data || !event.data.type) {{
        return;
    }}
    if (event.data.type === "SKIP_WAITING") {{
        self.skipWaiting();
    }}
    if (event.data.type === "GET_VERSION" && event.ports && event.ports[0]) {{
        event.ports[0].postMessage({{
            version: CACHE_NAME,
            appContractVersion: APP_CONTRACT_VERSION,
            shellVersion: CACHE_NAME,
            roleCode: ROLE_CODE
        }});
    }}
}});
""".strip()


def get_current_access(request):
    access_id = request.session.get('employee_access_id')
    if not access_id:
        return None
    expire_due_temporary_work_transfers()
    access_queryset = (
        EmployeeAccess.objects
        .select_related('employee', 'role')
        .filter(
            id=access_id,
            is_active=True,
            status=EmployeeAccess.Status.ACTIVATED,
            employee__is_active=True,
            employee__status=Employee.Status.ACTIVE,
            role__is_active=True,
        )
    )
    role_app = get_role_app_for_request(request)
    if role_app:
        access_queryset = access_queryset.filter(role__code=role_app.role_code)
    access = access_queryset.first()
    if access and not employee_has_effective_access_role(access.employee, access.role.code):
        return None
    return access


def require_admin_access(request):
    access = get_current_access(request)
    if not access:
        return None
    if access.role.code != 'admin':
        return None
    return access


def generate_unique_access_code():
    while True:
        code = ''.join(str(secrets.randbelow(10)) for _ in range(6))
        if not EmployeeAccess.objects.filter(access_code=code).exists():
            return code


def log_admin_action(actor, action, obj=None, old_value='', new_value='', comment=''):
    AdminActionLog.objects.create(
        actor=actor,
        action=action,
        object_type=obj.__class__.__name__ if obj else '',
        object_id=str(obj.pk) if obj and obj.pk else '',
        object_repr=str(obj) if obj else '',
        old_value=old_value,
        new_value=new_value,
        comment=comment,
    )



def redirect_after_admin_action(request, fallback_view, **kwargs):
    next_url = request.POST.get('next', '')
    if next_url and url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return redirect(next_url)
    return redirect(fallback_view, **kwargs)


def _validated_next_url(request, value):
    value = (value or '').strip()
    if value and url_has_allowed_host_and_scheme(
        value,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return value
    return ''


def _role_landing_url(access):
    """Return the role's real workplace instead of the redirect-only /home/."""
    app = get_role_app(access.role.code) if access else None
    return app.start_url if app else reverse('role_home')


def _login_redirect_response(request, target_url):
    """Let the fetch-driven login perform exactly one final navigation."""
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'ok': True, 'redirect_url': target_url})
    return redirect(target_url)


def _active_role_session_matches_access(request, access):
    state = role_session_state(request, access)
    active_access = state.get('active_access')
    return bool(
        state['authenticated']
        and state['is_active']
        and active_access is not None
        and active_access.id == access.id
        and active_access.last_login_at is not None
        and request.session.get(ACTIVE_ROLE_SESSION_KEY) == access.id
        and request.session.get(ACTIVE_ROLE_GENERATION_SESSION_KEY)
        and request.session.get(ACTIVE_ROLE_CODE_SESSION_KEY)
        == access.role.code
    )

def build_workbook_response(workbook, filename):
    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def excel_value(value):
    if isinstance(value, datetime) and timezone.is_aware(value):
        return timezone.localtime(value).replace(tzinfo=None)
    return value


def driver_manifest_view(request):
    return role_app_manifest_response(request, 'driver')


def driver_service_worker_view(request):
    return role_app_service_worker_response(request, 'driver', DRIVER_SERVICE_WORKER_JS)


def interface_map_view(request):
    return render(
        request,
        'users/interface_map.html',
        {
            'interface_sections': INTERFACE_MAP,
            'demo_access_codes': DEMO_ACCESS_CODES,
        },
    )


@require_GET
def app_catalog_view(request):
    if get_role_app_for_request(request):
        return redirect(app_catalog_public_url(request))
    catalog_apps = app_catalog_items(request)
    selected_role_code = (request.GET.get('app') or '').strip()
    selected_app = next(
        (item for item in catalog_apps if item['role_code'] == selected_role_code),
        None,
    )
    response = render(
        request,
        'users/app_catalog.html',
        {
            'catalog_apps': catalog_apps,
            'selected_app': selected_app,
        },
    )
    response['Cache-Control'] = 'no-cache'
    return response


@require_GET
def app_catalog_qr_view(request, role_code):
    if role_code not in APP_CATALOG_ROLE_CODES:
        raise Http404
    response = redirect(static(role_app_qr_asset_path(role_code)))
    response['Cache-Control'] = 'public, max-age=86400'
    response['X-Content-Type-Options'] = 'nosniff'
    return response


def _masked_activation_phone(value):
    digits = normalize_phone(value)
    if len(digits) == 11 and digits.startswith('7'):
        return f'+7 ••• •••-{digits[-4:-2]}-{digits[-2:]}'
    return 'Телефон подтвержден'


def login_view(
    request,
    *,
    allowed_role_codes=None,
    target_role_app=None,
    forced_next_url='',
):
    role_app = get_role_app_for_request(request)
    login_role_app = target_role_app or role_app
    combined_mobile_login = bool(
        role_app
        and role_app.role_code in {'driver', 'excavator_operator'}
        and target_role_app is None
    )
    allowed_role_codes = tuple(allowed_role_codes or ())
    next_url = forced_next_url or _validated_next_url(
        request,
        (
            request.POST.get('next')
            if request.method == 'POST'
            else request.GET.get('next')
        ),
    )
    current_access = get_current_access(request) if request.method == 'GET' else None
    rating_tv_reauthentication_required = bool(
        current_access
        and next_url == reverse('driver_rating_tv')
        and not _active_role_session_matches_access(request, current_access)
    )
    targeted_role_reauthentication_required = bool(
        current_access
        and allowed_role_codes
        and (
            current_access.role.code not in allowed_role_codes
            or not _active_role_session_matches_access(request, current_access)
        )
    )
    if (
        request.method == 'GET'
        and current_access
        and not rating_tv_reauthentication_required
        and not targeted_role_reauthentication_required
    ):
        return redirect(next_url or _role_landing_url(current_access))
    if (
        request.method == 'GET'
        and role_app
        and target_role_app is None
        and request.session.get('employee_access_id')
    ):
        request.session.flush()
    selected_device_kind = request.POST.get('device_kind') if request.method == 'POST' else detect_session_device_kind(request)
    if selected_device_kind not in {'personal', 'shared'}:
        selected_device_kind = detect_session_device_kind(request)
    # Номер возвращаем в форму: при неверном пинкоде набирать его заново незачем.
    submitted_phone = (
        request.POST.get('phone', '').strip() if request.method == 'POST' else ''
    )
    # Со /start/ номер уже известен — ссылка на приложение несёт его в query.
    # Незачем набирать его снова: не путать с submitted_phone, который отвечает
    # ещё и за то, показывать ли поле пинкода (первый неудачный вход).
    prefilled_phone = (
        request.GET.get('phone', '').strip() if request.method == 'GET' else ''
    )
    submitted_action = request.POST.get('action', '') if request.method == 'POST' else ''
    login_action = submitted_action if submitted_action in {'register', 'continue'} else 'login'
    submitted_privacy_consent = (
        request.POST.get(PRIVACY_CONSENT_FIELD, '')
        if request.method == 'POST'
        else ''
    )
    if (
        request.method == 'POST'
        and combined_mobile_login
        and login_action in {'login', 'register', 'continue'}
        and not privacy_consent_submission_is_current(submitted_privacy_consent)
    ):
        messages.error(request, PRIVACY_CONSENT_REQUIRED_MESSAGE)
        return render(
            request,
            'users/login.html',
            {
                'selected_device_kind': selected_device_kind,
                'next_url': next_url,
                'login_role_app': login_role_app,
                'combined_mobile_login': True,
                'submitted_phone': submitted_phone or prefilled_phone,
                'login_step': 'pin' if submitted_phone else '',
                'privacy_consent_error': True,
            },
        )

    if request.method == 'POST' and login_action in {'register', 'continue'}:
        # Первый вход. Раньше сюда пускал только выданный вручную временный код,
        # и раздавать его приходилось каждому — при текучке в двадцать человек за
        # вахту это неподъёмно. Теперь ключ — номер телефона из карточки, а ФИО
        # человек подтверждает глазами на следующем экране.
        phone = request.POST.get('phone', '').strip()
        matches = find_unactivated_accesses_by_phone(
            phone,
            role_codes=allowed_role_codes or ([role_app.role_code] if role_app else None),
        )
        matches = [
            candidate
            for candidate in matches
            if employee_has_effective_access_role(
                candidate.employee,
                candidate.role.code,
                allow_pending_access=True,
            )
        ]
        # У сотрудника может быть несколько записей доступа на один номер —
        # например, доступ переоформляли. Если хоть одна уже активирована, у
        # человека есть рабочий пинкод, и вести его на регистрацию нельзя:
        # заведёт второй и запутается, каким входить.
        already_registered = any(
            candidate.status == EmployeeAccess.Status.ACTIVATED
            for candidate in matches
        )
        pending = [
            candidate
            for candidate in matches
            if candidate.status == EmployeeAccess.Status.NOT_ACTIVATED
        ]
        if pending and not already_registered:
            access = pending[0]
            request.session.cycle_key()
            accept_current_privacy_policy(
                request,
                submitted_privacy_consent,
                access=access,
            )
            request.session['pending_activation_access_id'] = access.id
            request.session['pending_activation_role_code'] = access.role.code
            if target_role_app:
                request.session['pending_activation_target_app_code'] = target_role_app.role_code
            else:
                request.session.pop('pending_activation_target_app_code', None)
            if next_url:
                request.session['post_activation_next'] = next_url
            set_session_device_kind(request, selected_device_kind)
            return _login_redirect_response(request, reverse('activate_access'))
        if matches:
            # Пинкод у человека уже есть — просим именно его, вторым шагом.
            return render(
                request,
                'users/login.html',
                {
                    'selected_device_kind': selected_device_kind,
                    'next_url': next_url,
                    'login_role_app': login_role_app,
                    # Старый cached-клиент action=continue заменяет только
                    # <main>, но не может подхватить новый CSS из <head>.
                    # Возвращаем ему прежний самостоятельный PIN-шаг.
                    'combined_mobile_login': False,
                    'submitted_phone': phone,
                    'login_step': 'pin',
                },
            )
        else:
            # Номер может быть в базе, но за другой должностью: человек открыл
            # чужое приложение. Писать ему «номер не найден» — врать и сбивать с
            # толку, он пойдёт менять номер, который менять не нужно.
            elsewhere = find_unactivated_accesses_by_phone(phone)
            elsewhere = [
                candidate
                for candidate in elsewhere
                if employee_has_effective_access_role(
                    candidate.employee,
                    candidate.role.code,
                    allow_pending_access=True,
                )
            ]
            if elsewhere:
                return render(
                    request,
                    'users/login_phone_not_found.html',
                    {
                        'login_role_app': login_role_app,
                        'submitted_phone': format_phone_for_display(phone),
                        'support_chat_url': getattr(settings, 'SUPPORT_CHAT_URL', ''),
                        'support_chat_label': getattr(settings, 'SUPPORT_CHAT_LABEL', ''),
                        'wrong_app': True,
                        'own_role_names': sorted({
                            candidate.role.name for candidate in elsewhere
                        }),
                    },
                )
            return render(
                request,
                'users/login_phone_not_found.html',
                {
                    'login_role_app': login_role_app,
                    'submitted_phone': format_phone_for_display(phone),
                    'support_chat_url': getattr(settings, 'SUPPORT_CHAT_URL', ''),
                    'support_chat_label': getattr(settings, 'SUPPORT_CHAT_LABEL', ''),
                },
            )

    if request.method == 'POST' and login_action == 'login':
        phone = request.POST.get('phone', '').strip()
        access_code = request.POST.get('access_code', '').strip()
        access = find_employee_access_by_credentials(
            phone,
            access_code,
            role_code=(
                role_app.role_code
                if role_app and not allowed_role_codes
                else None
            ),
            role_codes=allowed_role_codes,
        )
        if access and not employee_has_effective_access_role(
            access.employee,
            access.role.code,
            allow_pending_access=True,
        ):
            access = None
        if access:
            if access.status == EmployeeAccess.Status.NOT_ACTIVATED:
                if access.primary_code_issued_at:
                    request.session.cycle_key()
                    accept_current_privacy_policy(
                        request,
                        submitted_privacy_consent,
                        access=access,
                    )
                    request.session['pending_activation_access_id'] = access.id
                    request.session['pending_activation_role_code'] = access.role.code
                    if target_role_app:
                        request.session['pending_activation_target_app_code'] = (
                            target_role_app.role_code
                        )
                    else:
                        request.session.pop(
                            'pending_activation_target_app_code',
                            None,
                        )
                    if next_url:
                        request.session['post_activation_next'] = next_url
                    set_session_device_kind(request, selected_device_kind)
                    return _login_redirect_response(request, reverse('activate_access'))
            try:
                with transaction.atomic():
                    locked_employee = Employee.objects.select_for_update().get(pk=access.employee_id)
                    locked_access = (
                        EmployeeAccess.objects
                        .select_for_update()
                        .select_related('employee', 'role')
                        .get(pk=access.pk)
                    )
                    if locked_access.status == EmployeeAccess.Status.NOT_ACTIVATED:
                        locked_access.status = EmployeeAccess.Status.ACTIVATED
                        locked_access.activated_at = timezone.now()
                        locked_access.save(update_fields=['status', 'activated_at'])
                        if locked_employee.status == Employee.Status.NOT_ACTIVATED:
                            locked_employee.status = Employee.Status.ACTIVE
                            locked_employee.is_active = True
                            locked_employee.save(update_fields=['status', 'is_active', 'updated_at'])
                    activate_role_session(request, locked_access)
            except ValidationError as error:
                messages.error(request, '; '.join(error.messages))
                return render(
                    request,
                    'users/login.html',
                    {
                        'selected_device_kind': selected_device_kind,
                        'next_url': next_url,
                        'login_role_app': login_role_app,
                        'combined_mobile_login': combined_mobile_login,
                        'submitted_phone': submitted_phone,
                        'login_step': 'pin' if submitted_phone else '',
                    },
                )
            request.session.cycle_key()
            set_session_device_kind(request, selected_device_kind)
            accept_current_privacy_policy(
                request,
                submitted_privacy_consent,
                access=locked_access,
            )
            return _login_redirect_response(
                request,
                next_url or _role_landing_url(access),
            )
        # Пинкода у номера ещё нет — человек первый раз в приложении. Отбивать
        # его «неверный пинкод» бессмысленно: вводить ему нечего. Ведём на экран,
        # где он увидит своё ФИО и придумает пинкод.
        phone_accesses = [
            candidate
            for candidate in find_unactivated_accesses_by_phone(
                phone,
                role_codes=allowed_role_codes or ([role_app.role_code] if role_app else None),
            )
            if employee_has_effective_access_role(
                candidate.employee,
                candidate.role.code,
                allow_pending_access=True,
            )
        ]
        # То же правило, что и на первом шаге: пинкод уже заведён — значит вход
        # обычный, а не повторная регистрация.
        first_time = [] if any(
            candidate.status == EmployeeAccess.Status.ACTIVATED
            for candidate in phone_accesses
        ) else [
            candidate
            for candidate in phone_accesses
            if candidate.status == EmployeeAccess.Status.NOT_ACTIVATED
        ]
        if first_time:
            pending_access = first_time[0]
            request.session.cycle_key()
            accept_current_privacy_policy(
                request,
                submitted_privacy_consent,
                access=pending_access,
            )
            request.session['pending_activation_access_id'] = pending_access.id
            request.session['pending_activation_role_code'] = pending_access.role.code
            if target_role_app:
                request.session['pending_activation_target_app_code'] = target_role_app.role_code
            else:
                request.session.pop('pending_activation_target_app_code', None)
            if next_url:
                request.session['post_activation_next'] = next_url
            set_session_device_kind(request, selected_device_kind)
            return _login_redirect_response(request, reverse('activate_access'))

        other_access = None
        if allowed_role_codes:
            other_access = find_employee_access_by_credentials(phone, access_code)
        if other_access:
            messages.error(
                request,
                f'У этой учетной записи нет доступа к приложению «{login_role_app.short_name}».',
            )
        elif login_role_app:
            messages.error(
                request,
                f'Телефон или пинкод указаны неверно для приложения «{login_role_app.short_name}».',
            )
        else:
            messages.error(request, 'Телефон или пинкод указаны неверно.')
    return render(
        request,
        'users/login.html',
        {
            'selected_device_kind': selected_device_kind,
            'next_url': next_url,
            'login_role_app': login_role_app,
            'combined_mobile_login': combined_mobile_login,
            'submitted_phone': submitted_phone or prefilled_phone,
            'login_step': 'pin' if submitted_phone else '',
            # Отличаем «номер пришёл в ссылке со /start/» от «человек уже
            # пробовал войти»: в первом случае экран установки должен
            # показаться как обычно, во втором — форма уже открыта на JS,
            # и это единственный случай, кроме постоянного GET, что доходит
            # досюда без реальной попытки входа.
            'phone_is_prefill_only': bool(prefilled_phone) and not submitted_phone,
        },
    )


@require_POST
def reclaim_role_session_view(request):
    """Return the active session to this device.

    Only one session per employee may act at a time, so logging in elsewhere
    leaves this screen read-only with no way back except retyping the PIN. This
    lets the person already authenticated here take the session back in one tap.
    It re-activates the very access this session is already signed in as, so it
    grants nothing new; switching between different roles still goes through
    activate_role_session's checks for open shifts, trips and downtimes.
    """
    access_id = request.session.get('employee_access_id')
    if not access_id:
        return JsonResponse(
            {'ok': False, 'error': 'Сессия не найдена. Войдите заново.'},
            status=401,
        )
    access = (
        EmployeeAccess.objects
        .select_related('employee', 'role')
        .filter(
            id=access_id,
            is_active=True,
            status=EmployeeAccess.Status.ACTIVATED,
            role__is_active=True,
            employee__is_active=True,
            employee__status=Employee.Status.ACTIVE,
        )
        .first()
    )
    if not access:
        return JsonResponse(
            {'ok': False, 'error': 'Доступ отключен. Обратитесь к администратору.'},
            status=403,
        )
    try:
        activate_role_session(request, access)
    except ValidationError as error:
        return JsonResponse(
            {'ok': False, 'error': '; '.join(error.messages)},
            status=409,
        )
    return JsonResponse({'ok': True})


def activate_access_view(request):
    access_id = request.session.get('pending_activation_access_id')
    if not access_id:
        return redirect('login')
    pending_role_code = request.session.get('pending_activation_role_code')
    target_app = get_role_app(
        request.session.get('pending_activation_target_app_code', '')
    )
    access_queryset = (
        EmployeeAccess.objects
        .select_related('employee', 'role')
        .filter(id=access_id, is_active=True, status=EmployeeAccess.Status.NOT_ACTIVATED)
    )
    host_role_app = get_role_app_for_request(request)
    if pending_role_code:
        access_queryset = access_queryset.filter(role__code=pending_role_code)
    elif host_role_app:
        access_queryset = access_queryset.filter(role__code=host_role_app.role_code)
    access = access_queryset.first()
    if access and not employee_has_effective_access_role(
        access.employee,
        access.role.code,
        allow_pending_access=True,
    ):
        access = None
    if not access:
        for key in (
            'pending_activation_access_id',
            'pending_activation_role_code',
            'pending_activation_target_app_code',
            'post_activation_next',
        ):
            request.session.pop(key, None)
        return redirect('login')

    activation_role_app = target_app or host_role_app
    if (
        activation_role_app
        and activation_role_app.role_code in {'driver', 'excavator_operator'}
        and not privacy_consent_matches_access(request.session, access)
    ):
        for key in (
            'pending_activation_access_id',
            'pending_activation_role_code',
            'pending_activation_target_app_code',
            'post_activation_next',
        ):
            request.session.pop(key, None)
        messages.error(request, PRIVACY_CONSENT_REQUIRED_MESSAGE)
        return redirect('login')

    if request.method == 'POST':
        form = AccessActivationForm(request.POST, access=access)
        if form.is_valid():
            try:
                with transaction.atomic():
                    locked_employee = Employee.objects.select_for_update().get(pk=access.employee_id)
                    locked_access = (
                        EmployeeAccess.objects
                        .select_for_update()
                        .select_related('employee', 'role')
                        .get(
                            pk=access.pk,
                            is_active=True,
                            status=EmployeeAccess.Status.NOT_ACTIVATED,
                        )
                    )
                    locked_access.access_code = form.cleaned_data['new_access_code']
                    locked_access.status = EmployeeAccess.Status.ACTIVATED
                    locked_access.activated_at = timezone.now()
                    locked_access.save(update_fields=['access_code', 'status', 'activated_at'])
                    if locked_employee.status == Employee.Status.NOT_ACTIVATED:
                        locked_employee.status = Employee.Status.ACTIVE
                        locked_employee.is_active = True
                        locked_employee.save(update_fields=['status', 'is_active', 'updated_at'])
                    access = activate_role_session(request, locked_access)
            except ValidationError as error:
                form.add_error(None, '; '.join(error.messages))
                return render(
                    request,
                    'users/activate_access.html',
                    {
                        'access': access,
                        'activation_phone': _masked_activation_phone(access.employee.phone),
                        'form': form,
                        'activation_role_app': activation_role_app,
                    },
                )
            for key in (
                'pending_activation_access_id',
                'pending_activation_role_code',
                'pending_activation_target_app_code',
            ):
                request.session.pop(key, None)
            request.session.cycle_key()
            set_session_device_kind(request, get_session_device_kind(request))
            if access.role.code != 'oup':
                messages.success(
                    request,
                    'Постоянный пинкод создан. Первичный пинкод больше не действует.',
                )
            next_url = _validated_next_url(request, request.session.pop('post_activation_next', ''))
            return redirect(next_url or _role_landing_url(access))
    else:
        form = AccessActivationForm(access=access)

    return render(
        request,
        'users/activate_access.html',
        {
            'access': access,
            'activation_phone': _masked_activation_phone(access.employee.phone),
            'form': form,
            'activation_role_app': activation_role_app,
        },
    )


def logout_view(request):
    request.session.flush()
    return redirect('login')


def role_home_view(request):
    access = get_current_access(request)
    if not access:
        request.session.flush()
        return redirect('login')
    if access.role.code == 'driver':
        return redirect('driver_work')
    if access.role.code == 'mining_master':
        return redirect('mining_master_assignments')
    if access.role.code == 'deputy_mining_manager':
        return redirect('deputy_mining_manager_placement')
    if access.role.code == 'oup':
        return redirect('oup_home')
    if access.role.code == 'timekeeper':
        return redirect('rotation_timekeeper_dashboard')
    if access.role.code == 'site_manager':
        return redirect('rotation_site_manager_queue')
    if access.role.code == 'employee_portal':
        return redirect('rotation_employee_home')
    if access.role.code == 'excavator_operator':
        return redirect('excavator_work')
    if access.role.code == 'dispatcher':
        return redirect('dispatcher_control')
    if access.role.code == 'mechanic':
        return redirect('mechanic_dashboard')
    if access.role.code == 'manager':
        return redirect('management_dashboard')
    if access.role.code == 'settlement_clerk':
        return redirect('clerk_home')
    if access.role.code == 'admin':
        return redirect('system_admin_dashboard')
    interface_name = ROLE_INTERFACE_NAMES.get(access.role.code, f'Интерфейс роли: {access.role.name}')
    return render(
        request,
        'users/role_home.html',
        {
            'access': access,
            'interface_name': interface_name,
        },
    )


def system_admin_dashboard_view(request):
    access = require_admin_access(request)
    if not access:
        return redirect('role_home')

    employee_status_counts = {
        item['status']: item['total']
        for item in Employee.objects.values('status').annotate(total=Count('id'))
    }
    access_status_counts = {
        item['status']: item['total']
        for item in EmployeeAccess.objects.values('status').annotate(total=Count('id'))
    }
    reference_counts = [
        ('Подразделения', PersonnelDepartment.objects.count(), '/system-admin/references/personnel-departments/'),
        ('Графики работы', WorkSchedule.objects.count(), '/system-admin/references/work-schedules/'),
        ('Утверждённые составы вахт', WatchComposition.objects.count(), '/system-admin/references/watch-compositions/'),
        ('Периоды рейтинга', RatingPeriod.objects.count(), '/system-admin/references/rating-periods/'),
        ('Кадровые должности', PersonnelPosition.objects.count(), '/system-admin/references/personnel-positions/'),
        ('Производственные специализации', ProductionSpecialization.objects.count(), '/system-admin/references/production-specializations/'),
        ('Виды техники', EquipmentType.objects.count(), '/admin/references/equipmenttype/'),
        ('Техника', Equipment.objects.count(), '/admin/references/equipment/'),
        ('Состояния техники', EquipmentState.objects.count(), '/admin/references/equipmentstate/'),
        ('Причины простоев', DowntimeReason.objects.count(), '/admin/downtimes/downtimereason/'),
        ('Породы', RockType.objects.count(), '/admin/references/rocktype/'),
        ('Точки разгрузки', DumpPoint.objects.count(), '/admin/references/dumppoint/'),
        ('Общежития', Dormitory.objects.count(), '/admin/references/dormitory/'),
        ('Секции общежитий', DormitorySection.objects.count(), '/admin/references/dormitorysection/'),
        ('Шаблоны отчетов', ReportTemplate.objects.count(), '/reports/templates/'),
    ]

    return render(
        request,
        'users/system_admin_dashboard.html',
        {
            'access': access,
            'employee_total': Employee.objects.count(),
            'active_total': employee_status_counts.get(Employee.Status.ACTIVE, 0),
            'not_activated_total': access_status_counts.get(EmployeeAccess.Status.NOT_ACTIVATED, 0),
            'blocked_total': access_status_counts.get(EmployeeAccess.Status.BLOCKED, 0),
            'deactivated_total': access_status_counts.get(EmployeeAccess.Status.DEACTIVATED, 0),
            'recent_employees': Employee.objects.order_by('-created_at')[:5],
            'recent_accesses': EmployeeAccess.objects.select_related('employee', 'role').order_by('-last_login_at', '-created_at')[:5],
            'recent_logs': AdminActionLog.objects.select_related('actor')[:8],
            'open_conflicts': AdminConflict.objects.select_related('employee', 'role').filter(status=AdminConflict.Status.OPEN)[:8],
            'reference_counts': reference_counts,
            'shift_fact_total': Trip.objects.count() + DowntimeEvent.objects.count(),
            'shift_reading_corrections': recent_shift_reading_corrections(),
        },
    )


@require_POST
def system_admin_reset_shift_test_data_view(request):
    access = require_admin_access(request)
    if not access:
        return redirect('role_home')

    deleted_counts = {
        'рейсы': Trip.objects.count(),
        'простои': DowntimeEvent.objects.count(),
        'оперативные события': OperationalStateEvent.objects.count(),
        'клиентские действия рейсов': TripClientAction.objects.count(),
        'диспетчерские журналы действий': DispatcherActionLog.objects.count(),
    }

    with transaction.atomic():
        TripClientAction.objects.all().delete()
        DispatcherActionLog.objects.all().delete()
        Trip.objects.all().delete()
        DowntimeEvent.objects.all().delete()
        OperationalStateEvent.objects.all().delete()
        bump_operational_state(
            'SystemAdmin:test_shift_data_reset',
            event_type='test_shift_data_reset',
            object_type='SystemAdmin',
            payload={'action': 'test_shift_data_reset', 'deleted_counts': deleted_counts},
        )
        log_admin_action(
            access.employee,
            'Сброшены тестовые показатели смены',
            new_value=json.dumps(deleted_counts, ensure_ascii=False),
            comment='Удалены только рейсы, простои, оперативные события и журналы действий. Справочники, сотрудники, техника и планы сохранены.',
        )

    deleted_total = sum(deleted_counts.values())
    messages.success(request, f'Тестовые показатели смены сброшены. Удалено записей: {deleted_total}.')
    return redirect('system_admin_dashboard')


def system_admin_references_view(request):
    access = require_admin_access(request)
    if not access:
        return redirect('role_home')

    reference_configs = get_system_admin_reference_configs()
    reference_sections = [
        {
            'title': 'Сотрудники и доступы',
            'items': [
                {'name': 'Сотрудники', 'count': Employee.objects.count(), 'url': 'system_admin_employees', 'external_url': ''},
                {'name': 'Подразделения', 'count': PersonnelDepartment.objects.count(), 'url': '', 'external_url': '/admin/users/personneldepartment/', 'detail_code': 'personnel-departments'},
                {'name': 'Графики работы', 'count': WorkSchedule.objects.count(), 'url': '', 'external_url': '/admin/users/workschedule/', 'detail_code': 'work-schedules'},
                {'name': 'Утверждённые составы вахт', 'count': WatchComposition.objects.count(), 'url': '', 'external_url': '/admin/users/watchcomposition/', 'detail_code': 'watch-compositions'},
                {'name': 'Кадровые должности', 'count': PersonnelPosition.objects.count(), 'url': '', 'external_url': '/admin/users/personnelposition/', 'detail_code': 'personnel-positions'},
                {'name': 'Производственные специализации', 'count': ProductionSpecialization.objects.count(), 'url': '', 'external_url': '/admin/users/productionspecialization/', 'detail_code': 'production-specializations'},
                {'name': 'Роли', 'count': Role.objects.count(), 'url': '', 'external_url': '/admin/users/role/'},
                {'name': 'Доступы', 'count': EmployeeAccess.objects.count(), 'url': '', 'external_url': '/admin/users/employeeaccess/'},
            ],
        },
        {
            'title': 'Техника',
            'items': [
                {'name': 'Виды техники', 'count': EquipmentType.objects.count(), 'url': '', 'external_url': '/admin/references/equipmenttype/', 'detail_code': 'equipment-types'},
                {'name': 'Техника', 'count': Equipment.objects.count(), 'url': '', 'external_url': '/admin/references/equipment/', 'detail_code': 'equipment'},
                {'name': 'Состояния техники', 'count': EquipmentState.objects.count(), 'url': '', 'external_url': '/admin/references/equipmentstate/', 'detail_code': 'equipment-states'},
            ],
        },
        {
            'title': 'Производственные справочники',
            'items': [
                {'name': 'Породы', 'count': RockType.objects.count(), 'url': '', 'external_url': '/admin/references/rocktype/', 'detail_code': 'rocks'},
                {'name': 'Точки разгрузки', 'count': DumpPoint.objects.count(), 'url': '', 'external_url': '/admin/references/dumppoint/', 'detail_code': 'dump-points'},
                {'name': 'Шаблоны отчетов', 'count': ReportTemplate.objects.count(), 'url': '', 'external_url': '/reports/templates/'},
                {'name': 'Ежесменные планы техники', 'count': EquipmentPlanGroup.objects.count(), 'url': '', 'external_url': '/admin/shifts/equipmentplangroup/', 'detail_code': 'equipment-plan-groups'},
                {'name': 'Приз за 100% плана', 'count': AchievementPrize.objects.count(), 'url': '', 'external_url': '/admin/shifts/achievementprize/', 'detail_code': 'achievement-prizes'},
                {'name': 'Сменные планы (история)', 'count': ShiftPlan.objects.count(), 'url': '', 'external_url': '/admin/shifts/shiftplan/', 'detail_code': 'shift-plans'},
                {'name': 'Планы техники (история)', 'count': EquipmentShiftPlan.objects.count(), 'url': '', 'external_url': '/admin/shifts/equipmentshiftplan/', 'detail_code': 'equipment-shift-plans'},
            ],
        },
        {
            'title': 'Рейтинг',
            'items': [
                {
                    'name': 'Периоды рейтинга',
                    'count': RatingPeriod.objects.count(),
                    'url': '',
                    'external_url': '/admin/reports/ratingperiod/',
                    'detail_code': 'rating-periods',
                },
            ],
        },
        {
            'title': 'Простои',
            'items': [
                {'name': 'Общий список простоев', 'count': DowntimeReason.objects.count(), 'url': '', 'external_url': '/admin/downtimes/downtimereason/', 'detail_code': 'downtime-reasons'},
                {'name': 'Простои водителя самосвала', 'count': DowntimeReason.objects.filter(show_for_truck_driver=True).count(), 'url': '', 'external_url': '/admin/downtimes/downtimereason/', 'detail_code': 'truck-driver-downtimes'},
                {'name': 'Простои машиниста экскаватора', 'count': DowntimeReason.objects.filter(show_for_excavator_operator=True).count(), 'url': '', 'external_url': '/admin/downtimes/downtimereason/', 'detail_code': 'excavator-operator-downtimes'},
                {'name': 'Детальные простои механика', 'count': DowntimeReason.objects.filter(show_for_mechanic=True).count(), 'url': '', 'external_url': '/admin/downtimes/downtimereason/', 'detail_code': 'mechanic-downtimes'},
            ],
        },
        {
            'title': 'Проживание',
            'items': [
                {'name': 'Общежития', 'count': Dormitory.objects.count(), 'url': '', 'external_url': '/admin/references/dormitory/', 'detail_code': 'dormitories'},
                {'name': 'Секции общежитий', 'count': DormitorySection.objects.count(), 'url': '', 'external_url': '/admin/references/dormitorysection/', 'detail_code': 'dormitory-sections'},
            ],
        },
    ]
    reference_total = 0
    empty_total = 0
    for section in reference_sections:
        section_count = 0
        empty_count = 0
        for item in section['items']:
            count = item['count']
            section_count += count
            reference_total += count
            if count:
                item['status_label'] = 'Заполнен'
                item['status_class'] = 'ok'
            else:
                empty_count += 1
                empty_total += 1
                item['status_label'] = 'Пусто'
                item['status_class'] = 'warning'
            if item.get('detail_code') in reference_configs:
                item['target_label'] = 'Рабочий экран'
            else:
                item['target_label'] = 'Админка' if item['external_url'].startswith('/admin/') else 'Рабочий экран'
        section['count'] = section_count
        section['empty_count'] = empty_count
        section['status_label'] = 'Требует заполнения' if empty_count else 'Готов'
        section['status_class'] = 'warning' if empty_count else 'ok'

    return render(
        request,
        'users/system_admin_references.html',
        {
            'access': access,
            'reference_sections': reference_sections,
            'reference_total': reference_total,
            'empty_total': empty_total,
        },
    )


def get_system_admin_reference_configs():
    return {
        'personnel-departments': {
            'title': 'Подразделения',
            'section': 'Сотрудники и доступы',
            'model': PersonnelDepartment,
            'description': 'Официальные подразделения компании из 1С. В карточке сотрудника выбирается готовое значение без ручного ввода.',
            'fields': ['name', 'code', 'is_active'],
            'search_fields': ['name', 'code'],
            'preview_fields': ['code', 'is_active'],
            'initial': {'is_active': True},
            'admin_url': '/admin/users/personneldepartment/',
        },
        'work-schedules': {
            'title': 'Графики работы',
            'section': 'Сотрудники и доступы',
            'model': WorkSchedule,
            'description': 'Стандартные кадровые графики из 1С. Номер бригады выбирается отдельно и ограничивается количеством бригад в графике.',
            'fields': ['name', 'code', 'brigade_count', 'is_active'],
            'search_fields': ['name', 'code'],
            'preview_fields': ['code', 'brigade_count', 'is_active'],
            'initial': {'is_active': True, 'brigade_count': 2},
            'admin_url': '/admin/users/workschedule/',
        },
        'watch-compositions': {
            'title': 'Утверждённые составы вахт',
            'section': 'Сотрудники и доступы',
            'model': WatchComposition,
            'description': 'Уже утверждённые составы вахт. ОУП выбирает состав в карточке сотрудника; график и бригада не подставляют его автоматически.',
            'fields': ['name', 'code', 'is_active'],
            'search_fields': ['name', 'code'],
            'preview_fields': ['code', 'is_active'],
            'initial': {'is_active': True},
            'admin_url': '/admin/users/watchcomposition/',
        },
        'rating-periods': {
            'title': 'Периоды рейтинга',
            'section': 'Рейтинг',
            'model': RatingPeriod,
            'form_class': RatingPeriodReferenceForm,
            'description': (
                'Обычные периоды создаются автоматически по правилу '
                '14-е → 14-е: текущий и 12 следующих. Существующие записи '
                'система не изменяет. Вручную корректируйте только исключения '
                'и обязательно указывайте причину изменения дат.'
            ),
            'fields': [
                'name',
                'starts_on',
                'ends_before',
                'comment',
                'is_active',
            ],
            'search_fields': ['name', 'comment'],
            'preview_fields': [
                'starts_on',
                'ends_before',
                'generation_source_label',
                'manual_override_label',
                'is_active',
            ],
            'preview_labels': {
                'generation_source_label': 'Создание',
                'manual_override_label': 'Режим дат',
            },
            'initial': {'is_active': True},
            'admin_url': '/admin/reports/ratingperiod/',
        },
        'personnel-positions': {
            'title': 'Кадровые должности',
            'section': 'Сотрудники и доступы',
            'model': PersonnelPosition,
            'form_class': PersonnelPositionReferenceForm,
            'description': 'Официальные должности из 1С. Одна должность действует у сотрудника одновременно; она определяет допустимые базовые специализации.',
            'fields': ['name', 'code', 'requires_specialization', 'allowed_specializations', 'default_specialization', 'is_active'],
            'search_fields': ['name', 'code', 'allowed_specializations__name', 'default_specialization__name'],
            'preview_fields': ['code', 'requires_specialization', 'allowed_specializations', 'default_specialization', 'is_active'],
            'prefetch_related': ['allowed_specializations'],
            'select_related': ['default_specialization'],
            'initial': {'is_active': True},
            'admin_url': '/admin/users/personnelposition/',
        },
        'production-specializations': {
            'title': 'Производственные специализации',
            'section': 'Сотрудники и доступы',
            'model': ProductionSpecialization,
            'description': 'Определяют доступность сотрудника для расстановки и соответствующее приложение. Базовая специализация может быть временно заменена переводом до конца вахты.',
            'fields': ['name', 'code', 'equipment_type', 'access_role', 'is_active'],
            'search_fields': ['name', 'code', 'equipment_type__name', 'access_role__name'],
            'preview_fields': ['code', 'equipment_type', 'access_role', 'is_active'],
            'select_related': ['equipment_type', 'access_role'],
            'initial': {'is_active': True},
            'admin_url': '/admin/users/productionspecialization/',
        },
        'equipment-types': {
            'title': 'Виды техники',
            'section': 'Техника',
            'model': EquipmentType,
            'search_fields': ['name'],
            'preview_fields': ['name', 'is_active'],
            'admin_url': '/admin/references/equipmenttype/',
        },
        'equipment': {
            'title': 'Техника',
            'section': 'Техника',
            'model': Equipment,
            'search_fields': ['garage_number', 'vin', 'equipment_type__name', 'model__name'],
            'preview_fields': ['equipment_type', 'garage_number', 'model', 'vin'],
            'select_related': ['equipment_type', 'model'],
            'admin_url': '/admin/references/equipment/',
        },
        'equipment-states': {
            'title': 'Состояния техники',
            'section': 'Техника',
            'model': EquipmentState,
            'search_fields': ['code', 'name', 'short_label', 'description'],
            'preview_fields': ['code', 'name', 'short_label', 'color_group', 'semantic_group'],
            'admin_url': '/admin/references/equipmentstate/',
        },
        'downtime-reasons': {
            'title': 'Общий список простоев',
            'section': 'Простои',
            'model': DowntimeReason,
            'fields': [
                'name',
                'short_label',
                'equipment_type',
                'equipment_state',
                'is_critical',
                'show_for_truck_driver',
                'show_for_excavator_operator',
                'show_for_mechanic',
                'sort_order',
                'is_active',
            ],
            'search_fields': ['name', 'short_label', 'equipment_type__name', 'equipment_state__name'],
            'preview_fields': ['short_label', 'equipment_type', 'equipment_state', 'show_for_truck_driver', 'show_for_excavator_operator', 'show_for_mechanic'],
            'select_related': ['equipment_type', 'equipment_state'],
            'admin_url': '/admin/downtimes/downtimereason/',
        },
        'truck-driver-downtimes': {
            'title': 'Простои водителя самосвала',
            'section': 'Простои',
            'model': DowntimeReason,
            'fields': ['name', 'short_label', 'equipment_type', 'equipment_state', 'is_critical', 'show_for_truck_driver', 'sort_order', 'is_active'],
            'search_fields': ['name', 'short_label', 'equipment_type__name', 'equipment_state__name'],
            'preview_fields': ['short_label', 'equipment_type', 'equipment_state', 'is_critical', 'show_for_truck_driver'],
            'select_related': ['equipment_type', 'equipment_state'],
            'base_filter': {'show_for_truck_driver': True},
            'initial': {'show_for_truck_driver': True},
            'admin_url': '/admin/downtimes/downtimereason/',
        },
        'excavator-operator-downtimes': {
            'title': 'Простои машиниста экскаватора',
            'section': 'Простои',
            'model': DowntimeReason,
            'fields': ['name', 'short_label', 'equipment_type', 'equipment_state', 'is_critical', 'show_for_excavator_operator', 'sort_order', 'is_active'],
            'search_fields': ['name', 'short_label', 'equipment_type__name', 'equipment_state__name'],
            'preview_fields': ['short_label', 'equipment_type', 'equipment_state', 'is_critical', 'show_for_excavator_operator'],
            'select_related': ['equipment_type', 'equipment_state'],
            'base_filter': {'show_for_excavator_operator': True},
            'initial': {'show_for_excavator_operator': True},
            'admin_url': '/admin/downtimes/downtimereason/',
        },
        'mechanic-downtimes': {
            'title': 'Детальные простои механика',
            'section': 'Простои',
            'model': DowntimeReason,
            'fields': ['name', 'short_label', 'equipment_type', 'equipment_state', 'is_critical', 'show_for_mechanic', 'sort_order', 'is_active'],
            'search_fields': ['name', 'short_label', 'equipment_type__name', 'equipment_state__name'],
            'preview_fields': ['short_label', 'equipment_type', 'equipment_state', 'is_critical', 'show_for_mechanic'],
            'select_related': ['equipment_type', 'equipment_state'],
            'base_filter': {'show_for_mechanic': True},
            'initial': {'show_for_mechanic': True},
            'admin_url': '/admin/downtimes/downtimereason/',
        },
        'rocks': {
            'title': 'Породы',
            'section': 'Производство',
            'model': RockType,
            'search_fields': ['name'],
            'preview_fields': ['name', 'density', 'loosening_factor'],
            'admin_url': '/admin/references/rocktype/',
        },
        'dump-points': {
            'title': 'Точки разгрузки',
            'section': 'Производство',
            'model': DumpPoint,
            'search_fields': ['name'],
            'preview_fields': ['name', 'is_active'],
            'admin_url': '/admin/references/dumppoint/',
        },
        'equipment-plan-groups': {
            'title': 'Ежесменные планы техники',
            'section': 'Производство',
            'model': EquipmentPlanGroup,
            'form_class': EquipmentPlanGroupForm,
            'description': 'Один активный план задается на группу техники и автоматически фиксируется snapshot при открытии смены.',
            'fields': ['name', 'code', 'calculation_mode', 'plan_value', 'equipment', 'is_active', 'active_from', 'comment'],
            'search_fields': ['name', 'code', 'comment', 'equipment__garage_number', 'equipment__equipment_type__name', 'equipment__model__name'],
            'preview_fields': ['calculation_mode', 'plan_value', 'equipment', 'is_active', 'active_from', 'updated_by', 'updated_at'],
            'select_related': ['updated_by'],
            'prefetch_related': ['equipment', 'equipment__equipment_type', 'equipment__model'],
            'initial': {'is_active': True},
            'field_choices': {
                'calculation_mode': [
                    (PlanCalculationMode.TRIPS, 'По рейсам'),
                    (PlanCalculationMode.VOLUME, 'По объему, м3'),
                ],
            },
            'admin_url': '/admin/shifts/equipmentplangroup/',
        },
        'achievement-prizes': {
            'title': 'Приз за 100% плана',
            'section': 'Производство',
            'model': AchievementPrize,
            'description': 'Одна активная призовая картинка для водителей самосвалов и машинистов экскаваторов. Активная картинка выдается только после выполнения 100% сменного плана.',
            'fields': ['title', 'image', 'is_active'],
            'search_fields': ['title'],
            'preview_fields': ['title', 'image', 'is_active', 'updated_at'],
            'initial': {'title': 'План выполнен', 'is_active': True},
            'admin_url': '/admin/shifts/achievementprize/',
        },
        'shift-plans': {
            'title': 'Сменные планы (история)',
            'section': 'Производство',
            'model': ShiftPlan,
            'description': 'Старая схема планов по дате и смене сохранена для истории и совместимости. Основной способ - ежесменные планы техники.',
            'fields': ['plan_scope', 'name', 'plan_volume_m3', 'is_active', 'comment'],
            'search_fields': ['name', 'comment'],
            'preview_fields': ['plan_scope', 'plan_volume_m3'],
            'select_related': ['created_by'],
            'initial': {'plan_scope': ShiftPlanScope.DAY_SHIFT, 'name': 'Дневной сменный план', 'is_active': True},
            'hide_actions_card': True,
            'admin_url': '/admin/shifts/shiftplan/',
        },
        'equipment-shift-plans': {
            'title': 'Планы техники (история)',
            'section': 'Производство',
            'model': EquipmentShiftPlan,
            'description': 'Старая схема планов по конкретной технике на дату/смену. Для новых смен используйте ежесменные планы техники.',
            'fields': ['shift_plan', 'equipment', 'employee', 'calculation_mode', 'plan_trips', 'plan_volume_m3', 'is_active', 'comment'],
            'search_fields': ['shift_plan__name', 'equipment__garage_number', 'equipment__equipment_type__name', 'employee__full_name', 'comment'],
            'preview_fields': ['shift_plan', 'equipment', 'employee', 'calculation_mode', 'plan_trips', 'plan_volume_m3'],
            'select_related': ['shift_plan', 'equipment', 'equipment__equipment_type', 'employee'],
            'initial': {'is_active': True},
            'field_choices': {
                'calculation_mode': [
                    (PlanCalculationMode.TRIPS, 'По рейсам'),
                    (PlanCalculationMode.VOLUME, 'По объему, м3'),
                ],
            },
            'admin_url': '/admin/shifts/equipmentshiftplan/',
        },
        'dormitories': {
            'title': 'Общежития',
            'section': 'Проживание',
            'model': Dormitory,
            'search_fields': ['number'],
            'preview_fields': ['number', 'is_active'],
            'admin_url': '/admin/references/dormitory/',
        },
        'dormitory-sections': {
            'title': 'Секции общежитий',
            'section': 'Проживание',
            'model': DormitorySection,
            'search_fields': ['name', 'block__name', 'block__dormitory__number'],
            'preview_fields': ['block', 'name', 'day_capacity', 'night_capacity'],
            'select_related': ['block', 'block__dormitory'],
            'admin_url': '/admin/references/dormitorysection/',
        },
    }


def build_reference_form(model, config=None):
    config = config or {}
    if config.get('form_class'):
        form_class = config['form_class']
        field_choices = config.get('field_choices') or {}
        if not field_choices:
            return form_class

        class ReferenceForm(form_class):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                for field_name, choices in field_choices.items():
                    if field_name in self.fields:
                        self.fields[field_name].choices = choices

        return ReferenceForm

    editable_fields = config.get('fields') or [
        field.name
        for field in model._meta.fields
        if field.name != 'id' and getattr(field, 'editable', True)
    ]
    form_class = modelform_factory(model, fields=editable_fields)
    field_choices = config.get('field_choices') or {}
    if not field_choices:
        return form_class

    class ReferenceForm(form_class):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            for field_name, choices in field_choices.items():
                if field_name in self.fields:
                    self.fields[field_name].choices = choices

    return ReferenceForm


def prepare_reference_record_for_save(reference_code, record, access):
    if reference_code == 'shift-plans':
        if not record.date:
            record.date = timezone.localdate()
        record.plan_trips = None
        record.plan_tonnage = None
        if not record.created_by_id:
            record.created_by = access.employee
    elif reference_code == 'equipment-shift-plans':
        record.plan_tonnage = None
    elif reference_code == 'equipment-plan-groups':
        record.updated_by = access.employee
    return record


def build_reference_queryset(config):
    queryset = config['model'].objects.all()
    base_filter = config.get('base_filter') or {}
    if base_filter:
        queryset = queryset.filter(**base_filter)
    select_related = config.get('select_related') or []
    if select_related:
        queryset = queryset.select_related(*select_related)
    prefetch_related = config.get('prefetch_related') or []
    if prefetch_related:
        queryset = queryset.prefetch_related(*prefetch_related)
    return queryset


def build_reference_search_filter(search_fields, query):
    search_filter = Q()
    for field_name in search_fields:
        search_filter |= Q(**{f'{field_name}__icontains': query})
    return search_filter


def get_reference_status(record):
    if hasattr(record, 'is_active') and not record.is_active:
        return 'Отключен', 'neutral'
    return 'Активен', 'ok'


def get_reference_record_preview(record, config):
    preview = []
    for field_name in config.get('preview_fields', []):
        try:
            field = record._meta.get_field(field_name)
            label = config.get('preview_labels', {}).get(
                field_name,
                field.verbose_name,
            )
            value = getattr(record, field_name)
            if getattr(field, 'many_to_many', False):
                value = ', '.join(str(item) for item in value.all()) or 'Не указано'
            elif field.get_internal_type() == 'BooleanField':
                value = 'Да' if value else 'Нет'
            elif getattr(field, 'choices', None):
                value = getattr(record, f'get_{field_name}_display')()
            elif value in (None, ''):
                value = 'Не указано'
        except FieldDoesNotExist:
            label = config.get('preview_labels', {}).get(
                field_name,
                field_name.replace('_', ' '),
            )
            value = getattr(record, field_name, '')
            if callable(value):
                value = value()
            if value in (None, ''):
                value = 'Не указано'
        preview.append({'label': label, 'value': value})
    return preview


def _add_validation_error_to_form(form, error):
    if hasattr(error, 'message_dict'):
        for field_name, field_messages in error.message_dict.items():
            target_field = field_name if field_name in form.fields else None
            for field_message in field_messages:
                form.add_error(target_field, field_message)
        return
    for message in error.messages:
        form.add_error(None, message)


def _rating_period_automation_context():
    inspection = inspect_rating_period_calendar()
    prepared = inspection.prepared_through
    current_start = inspection.current_nominal_start
    return {
        'rule_label': '14-е → 14-е',
        'horizon_label': (
            f'{current_start:%d.%m.%Y}–'
            f'{inspection.horizon_end - timedelta(days=1):%d.%m.%Y}'
        ),
        'prepared_through_label': (
            f'{prepared - timedelta(days=1):%d.%m.%Y}'
            if prepared > current_start
            else 'не сформирован'
        ),
        'automatic_count': inspection.automatic_count,
        'manual_count': inspection.manual_count,
        'override_count': inspection.override_count,
        'gap_count': len(inspection.gap_ranges),
        'overlap_count': len(inspection.overlap_pairs),
        'is_ready': inspection.is_ready,
        'gaps': [
            (
                f'{starts_on:%d.%m.%Y}–'
                f'{ends_before - timedelta(days=1):%d.%m.%Y}'
            )
            for starts_on, ends_before in inspection.gap_ranges[:3]
        ],
    }


def system_admin_reference_detail_view(request, reference_code):
    access = require_admin_access(request)
    if not access:
        return redirect('role_home')

    configs = get_system_admin_reference_configs()
    config = configs.get(reference_code)
    if not config:
        messages.error(request, 'Справочник не найден.')
        return redirect('system_admin_references')

    model = config['model']
    form_class = build_reference_form(model, config)
    query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', '').strip()
    edit_id = request.GET.get('edit', '').strip()
    selected_record = None
    if edit_id.isdigit():
        selected_record = get_object_or_404(build_reference_queryset(config), id=edit_id)

    def reference_detail_redirect_url(record_id=None):
        params = []
        if query:
            params.append(('q', query))
        if status_filter:
            params.append(('status', status_filter))
        if record_id:
            params.append(('edit', record_id))
        query_string = urlencode(params)
        url = reverse('system_admin_reference_detail', kwargs={'reference_code': reference_code})
        return f'{url}?{query_string}' if query_string else url

    if request.method == 'POST':
        action = request.POST.get('action', 'save')
        record_id = request.POST.get('record_id', '').strip()
        record = None
        if record_id.isdigit():
            record = get_object_or_404(model, id=record_id)

        if action in {'disable', 'enable'} and record and hasattr(record, 'is_active'):
            try:
                if reference_code == 'rating-periods':
                    with transaction.atomic():
                        RatingPeriod.lock_catalog()
                        record = (
                            RatingPeriod._base_manager
                            .select_for_update()
                            .get(pk=record.pk)
                        )
                        old_value = record.audit_value()
                        record.is_active = action == 'enable'
                        record.save(update_fields=['is_active'])
                        new_value = record.audit_value()
                        action_label = (
                            'Период рейтинга включён'
                            if record.is_active
                            else 'Период рейтинга отключён'
                        )
                        log_admin_action(
                            access.employee,
                            action_label,
                            record,
                            old_value,
                            new_value,
                        )
                else:
                    old_value = (
                        'Активен' if record.is_active else 'Отключен'
                    )
                    record.is_active = action == 'enable'
                    record.save(update_fields=['is_active'])
                    new_value = (
                        'Активен' if record.is_active else 'Отключен'
                    )
                    log_admin_action(
                        access.employee,
                        f'Справочник: {config["title"]}',
                        record,
                        old_value,
                        new_value,
                    )
            except ValidationError as error:
                error_messages = []
                if hasattr(error, 'message_dict'):
                    for field_messages in error.message_dict.values():
                        error_messages.extend(field_messages)
                else:
                    error_messages.extend(error.messages)
                messages.error(
                    request,
                    'Состояние записи не изменено. '
                    + ' '.join(str(message) for message in error_messages),
                )
                return redirect(reference_detail_redirect_url(record.id))
            messages.success(request, 'Состояние записи обновлено.')
            return redirect(reference_detail_redirect_url(record.id))

        form = form_class(request.POST, request.FILES, instance=record)
        if form.is_valid():
            saved_record = form.save(commit=False)
            saved_record = prepare_reference_record_for_save(reference_code, saved_record, access)
            try:
                if reference_code == 'rating-periods':
                    with transaction.atomic():
                        RatingPeriod.lock_catalog()
                        old_audit_value = ''
                        if saved_record.pk:
                            stored_record = (
                                RatingPeriod._base_manager
                                .select_for_update()
                                .get(pk=saved_record.pk)
                            )
                            old_audit_value = stored_record.audit_value()
                        saved_record.save()
                        form.save_m2m()
                        log_admin_action(
                            access.employee,
                            (
                                'Период рейтинга изменён'
                                if record is not None
                                else 'Период рейтинга создан вручную'
                            ),
                            saved_record,
                            old_audit_value,
                            saved_record.audit_value(),
                        )
                else:
                    saved_record.save()
                    form.save_m2m()
                    log_admin_action(
                        access.employee,
                        f'Справочник: {config["title"]}',
                        saved_record,
                        '',
                        'Сохранено',
                    )
            except ValidationError as error:
                _add_validation_error_to_form(form, error)
            else:
                messages.success(request, 'Запись справочника сохранена.')
                return redirect(reference_detail_redirect_url(saved_record.id))
    else:
        form_initial = None if selected_record else config.get('initial')
        form = form_class(instance=selected_record, initial=form_initial)

    records_queryset = build_reference_queryset(config)
    if query:
        records_queryset = records_queryset.filter(build_reference_search_filter(config.get('search_fields', []), query)).distinct()
    if status_filter and hasattr(model, 'is_active'):
        records_queryset = records_queryset.filter(is_active=status_filter == 'active')

    records = []
    for record in records_queryset[:300]:
        status_label, status_class = get_reference_status(record)
        records.append({
            'object': record,
            'title': str(record),
            'status_label': status_label,
            'status_class': status_class,
            'preview': get_reference_record_preview(record, config),
        })

    count_queryset = build_reference_queryset(config)
    active_total = count_queryset.filter(is_active=True).count() if hasattr(model, 'is_active') else count_queryset.count()
    inactive_total = count_queryset.filter(is_active=False).count() if hasattr(model, 'is_active') else 0

    return render(
        request,
        'users/system_admin_reference_detail.html',
        {
            'access': access,
            'reference_code': reference_code,
            'reference_config': config,
            'form': form,
            'selected_record': selected_record,
            'records': records,
            'records_total': count_queryset.count(),
            'active_total': active_total,
            'inactive_total': inactive_total,
            'query': query,
            'status_filter': status_filter,
            'has_active_status': hasattr(model, 'is_active'),
            'rating_period_automation': (
                _rating_period_automation_context()
                if reference_code == 'rating-periods'
                else None
            ),
        },
    )


def system_admin_conflicts_view(request):
    access = require_admin_access(request)
    if not access:
        return redirect('role_home')

    status = request.GET.get('status', '').strip()
    query = request.GET.get('q', '').strip()
    conflicts = AdminConflict.objects.select_related('employee', 'role').order_by('-created_at')
    if status:
        conflicts = conflicts.filter(status=status)
    if query:
        conflicts = conflicts.filter(
            Q(conflict_type__icontains=query)
            | Q(process__icontains=query)
            | Q(description__icontains=query)
            | Q(comment__icontains=query)
            | Q(employee__full_name__icontains=query)
            | Q(role__name__icontains=query)
        )

    conflict_status_counts = {
        item['status']: item['total']
        for item in AdminConflict.objects.values('status').annotate(total=Count('id'))
    }
    conflicts = list(conflicts[:200])
    for conflict in conflicts:
        if conflict.status == AdminConflict.Status.OPEN:
            conflict.status_class = 'danger'
        elif conflict.status == AdminConflict.Status.IN_PROGRESS:
            conflict.status_class = 'warning'
        elif conflict.status == AdminConflict.Status.RESOLVED:
            conflict.status_class = 'ok'
        else:
            conflict.status_class = 'neutral'

    return render(
        request,
        'users/system_admin_conflicts.html',
        {
            'access': access,
            'conflicts': conflicts,
            'statuses': AdminConflict.Status.choices,
            'selected_status': status,
            'query': query,
            'open_total': conflict_status_counts.get(AdminConflict.Status.OPEN, 0),
            'in_progress_total': conflict_status_counts.get(AdminConflict.Status.IN_PROGRESS, 0),
            'resolved_total': conflict_status_counts.get(AdminConflict.Status.RESOLVED, 0),
            'rejected_total': conflict_status_counts.get(AdminConflict.Status.REJECTED, 0),
            'conflict_total': sum(conflict_status_counts.values()),
        },
    )


def system_admin_conflict_action_view(request, conflict_id, action):
    access = require_admin_access(request)
    if not access:
        return redirect('role_home')

    conflict = get_object_or_404(AdminConflict, id=conflict_id)
    if request.method == 'POST':
        status_by_action = {
            'in-progress': AdminConflict.Status.IN_PROGRESS,
            'resolved': AdminConflict.Status.RESOLVED,
            'rejected': AdminConflict.Status.REJECTED,
        }
        new_status = status_by_action.get(action)
        if new_status:
            old_status = conflict.get_status_display()
            conflict.status = new_status
            conflict.resolved_by = access.employee
            conflict.resolved_at = timezone.now()
            conflict.save(update_fields=['status', 'resolved_by', 'resolved_at'])
            log_admin_action(
                access.employee,
                'Изменен статус административного конфликта',
                conflict,
                old_value=old_status,
                new_value=conflict.get_status_display(),
            )
            messages.success(request, 'Статус конфликта обновлен.')

    redirect_url = request.POST.get('next') or 'system_admin_conflicts'
    if redirect_url == 'dashboard':
        return redirect('system_admin_dashboard')
    return redirect('system_admin_conflicts')


def system_admin_logs_view(request):
    access = require_admin_access(request)
    if not access:
        return redirect('role_home')

    query = request.GET.get('q', '').strip()
    log_type = request.GET.get('type', '').strip()
    logs = AdminActionLog.objects.select_related('actor', 'reversal_of').order_by('-created_at')
    if query:
        logs = logs.filter(
            Q(action__icontains=query)
            | Q(object_type__icontains=query)
            | Q(object_repr__icontains=query)
            | Q(comment__icontains=query)
            | Q(actor__full_name__icontains=query)
        )
    if log_type:
        if log_type == 'access':
            logs = logs.filter(Q(action__icontains='доступ') | Q(action__icontains='пинкод') | Q(object_type__icontains='Access'))
        elif log_type == 'employee':
            logs = logs.filter(Q(action__icontains='сотрудник') | Q(object_type__icontains='Employee'))
        elif log_type == 'conflict':
            logs = logs.filter(Q(action__icontains='конфликт') | Q(object_type__icontains='AdminConflict'))
        elif log_type == 'reference':
            logs = logs.filter(Q(action__icontains='Справочник') | Q(object_type__icontains='references'))
        elif log_type == 'oup':
            logs = logs.filter(Q(action__startswith='ОУП:') | Q(action_code='admin_oup_action_reversed'))

    total_logs = AdminActionLog.objects.count()
    access_total = AdminActionLog.objects.filter(Q(action__icontains='доступ') | Q(action__icontains='пинкод') | Q(object_type__icontains='Access')).count()
    employee_total = AdminActionLog.objects.filter(Q(action__icontains='сотрудник') | Q(object_type__icontains='Employee')).count()
    conflict_total = AdminActionLog.objects.filter(Q(action__icontains='конфликт') | Q(object_type__icontains='AdminConflict')).count()
    oup_total = AdminActionLog.objects.filter(action__startswith='ОУП:').count()
    logs = list(logs[:200])
    for log in logs:
        action_text = f'{log.action} {log.object_type}'.lower()
        if log.action.startswith('ОУП:'):
            log.type_label = 'ОУП'
            log.type_class = 'info'
            log.undo_state = get_oup_action_undo_state(log)
        elif log.action_code == 'admin_oup_action_reversed':
            log.type_label = 'Отмена ОУП'
            log.type_class = 'ok'
            log.undo_state = None
        elif 'конфликт' in action_text or 'adminconflict' in action_text:
            log.type_label = 'Конфликт'
            log.type_class = 'danger'
        elif 'доступ' in action_text or 'пинкод' in action_text or 'access' in action_text:
            log.type_label = 'Доступ'
            log.type_class = 'warning'
        elif 'сотрудник' in action_text or 'employee' in action_text:
            log.type_label = 'Сотрудник'
            log.type_class = 'ok'
        elif 'справочник' in action_text:
            log.type_label = 'Справочник'
            log.type_class = 'neutral'
        else:
            log.type_label = 'Действие'
            log.type_class = 'neutral'
            log.undo_state = None

    return render(
        request,
        'users/system_admin_logs.html',
        {
            'access': access,
            'logs': logs,
            'query': query,
            'selected_log_type': log_type,
            'total_logs': total_logs,
            'access_log_total': access_total,
            'employee_log_total': employee_total,
            'conflict_log_total': conflict_total,
            'oup_log_total': oup_total,
            'return_url': request.get_full_path(),
        },
    )


@require_POST
def system_admin_undo_oup_action_view(request, log_id):
    access = require_admin_access(request)
    if not access:
        return redirect('role_home')
    try:
        result, _reversal = undo_oup_action(
            log_id=log_id,
            actor_access_id=access.pk,
            comment=request.POST.get('comment', '').strip(),
        )
    except ValidationError as error:
        messages.error(request, '; '.join(error.messages))
    else:
        messages.success(request, result)
    return redirect_after_admin_action(request, 'system_admin_logs')


def system_admin_exports_view(request):
    access = require_admin_access(request)
    if not access:
        return redirect('role_home')

    export_groups = [
        {
            'title': 'Администрирование',
            'items': [
                {
                    'title': 'Сотрудники',
                    'description': 'Кадровая карточка, статус, телефон, вахта и проживание.',
                    'url_name': 'system_admin_employee_export',
                    'count': Employee.objects.count(),
                    'status_label': 'готово',
                    'status_class': 'ok',
                },
                {
                    'title': 'Доступы',
                    'description': 'Роли, статусы входа, даты выдачи, активации и последнего входа.',
                    'url_name': 'system_admin_access_export',
                    'count': EmployeeAccess.objects.count(),
                    'status_label': 'готово',
                    'status_class': 'ok',
                },
                {
                    'title': 'Журнал действий',
                    'description': 'История административных действий для сверки и аудита.',
                    'url_name': 'system_admin_log_export',
                    'count': AdminActionLog.objects.count(),
                    'status_label': 'готово',
                    'status_class': 'ok',
                },
                {
                    'title': 'Конфликты',
                    'description': 'Заблокированные рискованные действия и статусы разбора.',
                    'url_name': 'system_admin_conflict_export',
                    'count': AdminConflict.objects.count(),
                    'status_label': 'готово',
                    'status_class': 'warning' if AdminConflict.objects.filter(status=AdminConflict.Status.OPEN).exists() else 'ok',
                },
            ],
        },
        {
            'title': 'Рабочие отчеты MVP',
            'items': [
                {
                    'title': 'Объемы',
                    'description': 'Производственный отчет по рейсам, группировкам и шаблонам.',
                    'external_url': '/reports/volume/export/',
                    'count': Trip.objects.count(),
                    'status_label': 'отчет',
                    'status_class': 'neutral',
                },
                {
                    'title': 'Суточный отчет заказчику',
                    'description': 'Суточная форма по дате отчета для внешней сверки.',
                    'external_url': '/reports/customer-daily/export/',
                    'count': Trip.objects.count(),
                    'status_label': 'отчет',
                    'status_class': 'neutral',
                },
                {
                    'title': 'Витрина руководства',
                    'description': 'Excel-срез руководителя: сводка, динамика и сравнение смен.',
                    'external_url': '/reports/management/export/',
                    'count': Trip.objects.count(),
                    'status_label': 'отчет',
                    'status_class': 'neutral',
                },
                {
                    'title': 'Механические простои',
                    'description': 'Отчет по простоям техники с фильтрами механической службы.',
                    'external_url': '/reports/downtimes/export/',
                    'count': 0,
                    'status_label': 'отчет',
                    'status_class': 'neutral',
                },
            ],
        },
    ]
    export_total = sum(len(group['items']) for group in export_groups)
    ready_total = sum(1 for group in export_groups for item in group['items'] if item['status_class'] == 'ok')
    warning_total = sum(1 for group in export_groups for item in group['items'] if item['status_class'] == 'warning')

    return render(
        request,
        'users/system_admin_exports.html',
        {
            'access': access,
            'export_groups': export_groups,
            'export_total': export_total,
            'ready_total': ready_total,
            'warning_total': warning_total,
        },
    )


EMPLOYEES_WITHOUT_POSITION = 'none'


def system_admin_employees_view(request):
    access = require_admin_access(request)
    if not access:
        return redirect('role_home')

    employees = (
        Employee.objects
        .select_related('personnel_position')
        .prefetch_related('accesses__role')
        .order_by('full_name')
    )
    status = request.GET.get('status', '').strip()
    access_status = request.GET.get('access_status', '').strip()
    role_id = request.GET.get('role', '').strip()
    personnel_position = request.GET.get('personnel_position', '').strip()
    query = request.GET.get('q', '').strip()
    if status:
        employees = employees.filter(status=status)
    if access_status:
        employees = employees.filter(accesses__status=access_status).distinct()
    if role_id.isdigit():
        employees = employees.filter(accesses__role_id=int(role_id)).distinct()
    # Разметка фильтра по должности была на месте, а данные в неё не приходили:
    # список открывался пустым, и выбор в нём ничего не менял.
    if personnel_position == EMPLOYEES_WITHOUT_POSITION:
        employees = employees.filter(personnel_position__isnull=True)
    elif personnel_position.isdigit():
        employees = employees.filter(personnel_position_id=int(personnel_position))
    if query:
        employees = employees.filter(full_name__icontains=query)

    return render(
        request,
        'users/system_admin_employees.html',
        {
            'access': access,
            'employees': employees,
            'statuses': Employee.Status.choices,
            'access_statuses': EmployeeAccess.Status.choices,
            'roles': Role.objects.filter(is_active=True).order_by('name'),
            'personnel_positions': (
                PersonnelPosition.objects.filter(is_active=True).order_by('name')
            ),
            # Отдельной строкой — те, у кого должность не проставлена: при разборе
            # выгрузки из отдела кадров их надо находить в первую очередь.
            'personnel_position_groups': [
                (EMPLOYEES_WITHOUT_POSITION, 'Без кадровой должности'),
            ],
            'selected_status': status,
            'selected_access_status': access_status,
            'selected_role': role_id,
            'selected_personnel_position': personnel_position,
            'query': query,
        },
    )


def system_admin_employee_create_view(request):
    access = require_admin_access(request)
    if not access:
        return redirect('role_home')

    if request.method == 'POST':
        form = AdminEmployeeForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                with transaction.atomic():
                    role = form.cleaned_data.get('role')
                    employee = form.save(commit=False)
                    employee.status = Employee.Status.ACTIVE
                    employee.is_active = True
                    employee.dismissed_at = None
                    employee.save()
                    code = ''
                    if role and form.cleaned_data['generate_access']:
                        code = generate_unique_access_code()
                        EmployeeAccess.objects.create(
                            employee=employee,
                            role=role,
                            access_code=code,
                            status=EmployeeAccess.Status.NOT_ACTIVATED,
                            primary_code_issued_at=timezone.now(),
                        )
                    elif role:
                        EmployeeAccess.objects.create(
                            employee=employee,
                            role=role,
                            access_code='',
                            status=EmployeeAccess.Status.NOT_ACTIVATED,
                        )
                    work_assignment = form.save_work_assignment(
                        employee=employee,
                        assigned_by=access.employee,
                    )
            except ValidationError as error:
                form.add_error('assignment_equipment', error)
            else:
                assignment_label = (
                    f'{work_assignment.work_shift_label}; {work_assignment.equipment}'
                    if work_assignment else 'не задано'
                )
                if code:
                    log_admin_action(
                        access.employee,
                        'Создан сотрудник и выдан первичный пинкод',
                        employee,
                        new_value=(
                            f'Доступ: {role or "не требуется"}; '
                            f'назначение: {assignment_label}; пинкод: {code}'
                        ),
                    )
                    messages.success(
                        request,
                        f'Сотрудник создан. Первичный пинкод: {code}',
                        extra_tags='employee-card-silent',
                    )
                else:
                    log_admin_action(
                        access.employee,
                        'Создан сотрудник без пинкода',
                        employee,
                        new_value=f'Доступ: {role or "не требуется"}; назначение: {assignment_label}',
                    )
                    messages.success(request, 'Сотрудник создан.', extra_tags='employee-card-silent')
                return redirect_after_admin_action(request, 'system_admin_employee_detail', employee_id=employee.id)
    else:
        form = AdminEmployeeForm(initial={
            'hired_at': timezone.localdate(),
            'status': Employee.Status.ACTIVE,
        })

    return render(
        request,
        'users/employee_card.html',
        {
            'access': access,
            'form': form,
            'title': 'Создать сотрудника',
            'page_mode': 'create',
            'employee_card_context': 'admin',
            'can_submit_employee_card': True,
        },
    )


def editing_own_protected_card(access, employee):
    """Защищённую карточку правит только её владелец — и только свою.

    Остальным путь закрыт на уровне модели, здесь мы лишь решаем, открывать ли
    дверь: администратор, вошедший под собой, редактирует сам себя.
    """
    if employee.is_protected and access.employee_id == employee.id:
        return allow_protected_card_write()
    return nullcontext()


def system_admin_employee_detail_view(request, employee_id):
    access = require_admin_access(request)
    if not access:
        return redirect('role_home')

    employee = get_object_or_404(Employee, id=employee_id)
    if request.method == 'POST':
        initial_status = employee.status
        old_photo_name = employee.photo.name if employee.photo else ''
        if request.POST.get('remove_photo') == '1':
            if old_photo_name:
                employee.photo.storage.delete(old_photo_name)
                employee.photo = ''
                with editing_own_protected_card(access, employee):
                    employee.save(update_fields=['photo', 'updated_at'])
                log_admin_action(access.employee, 'Удалено фото сотрудника', employee)
                messages.success(request, 'Фото сотрудника удалено.')
            return redirect_after_admin_action(request, 'system_admin_employee_detail', employee_id=employee.id)
        form = AdminEmployeeEditForm(request.POST, request.FILES, instance=employee)
        if form.is_valid():
            try:
                with transaction.atomic(), editing_own_protected_card(access, employee):
                    locked_employee = Employee.objects.select_for_update().get(pk=employee.pk)
                    if locked_employee.status != initial_status:
                        raise ValidationError(
                            'Статус сотрудника уже изменился. Обновите страницу.',
                            code='stale_employee_status',
                        )
                    form.validate_protected_profile(locked_employee)
                    excluded_fields = set(form._meta.exclude or ())
                    excluded_fields.update(AdminEmployeeEditForm.PROTECTED_EXISTING_PROFILE_FIELDS)
                    form.instance = construct_instance(
                        form,
                        locked_employee,
                        form._meta.fields,
                        tuple(excluded_fields),
                    )
                    previous_specialization_id = locked_employee.base_specialization_id
                    saved_employee = form.save()
                    if (
                        previous_specialization_id != saved_employee.base_specialization_id
                        or production_access_is_out_of_sync(saved_employee)
                    ):
                        sync_employee_production_access(employee=saved_employee)
                    work_assignment = form.save_work_assignment(assigned_by=access.employee)
            except ValidationError as error:
                error_code = getattr(error, 'code', '')
                form.add_error(
                    None
                    if error_code in {'stale_employee_status', 'admin_watch_profile_forbidden', PROTECTED_WRITE_CODE}
                    else 'assignment_equipment',
                    error,
                )
            else:
                if request.FILES.get('photo') and old_photo_name and old_photo_name != saved_employee.photo.name:
                    saved_employee.photo.storage.delete(old_photo_name)
                assignment_label = (
                    f'{work_assignment.work_shift_label}; {work_assignment.equipment}'
                    if work_assignment else 'назначение снято'
                )
                log_admin_action(
                    access.employee,
                    'Изменена карточка сотрудника',
                    saved_employee,
                    new_value=f'Рабочее назначение: {assignment_label}',
                )
                messages.success(
                    request,
                    'Карточка сотрудника и рабочее назначение сохранены.',
                    extra_tags='employee-card-silent',
                )
                return redirect_after_admin_action(request, 'system_admin_employee_detail', employee_id=employee.id)
    else:
        form = AdminEmployeeEditForm(instance=employee)

    employee_accesses = employee.accesses.select_related('role').order_by('role__name')
    current_role_access = (
        employee_accesses
        .filter(is_active=True)
        .exclude(status=EmployeeAccess.Status.DEACTIVATED)
        .order_by('status', 'role__name')
        .first()
        or employee_accesses.first()
    )
    role_form_initial = {'role': current_role_access.role_id} if current_role_access else None
    active_equipment_assignment = get_active_equipment_assignment(employee)
    effective_employee_specialization = effective_specialization(employee)
    work_assignment_role = active_equipment_assignment.role if active_equipment_assignment else None
    if not work_assignment_role and effective_employee_specialization:
        work_assignment_role = effective_employee_specialization.access_role
    if not work_assignment_role and current_role_access:
        work_assignment_role = current_role_access.role

    return render(
        request,
        'users/employee_card.html',
        {
            'access': access,
            'employee': employee,
            'form': form,
            'title': employee.full_name,
            'page_mode': 'detail',
            'employee_card_context': 'admin',
            'can_submit_employee_card': True,
            'role_form': AdminAccessRoleForm(initial=role_form_initial),
            'block_form': AdminAccessBlockForm(),
            'employee_accesses': employee_accesses,
            'current_role_access': current_role_access,
            'active_equipment_assignment': active_equipment_assignment,
            'work_assignment_role': work_assignment_role,
            'work_assignment_supports_equipment': bool(
                work_assignment_role
                and work_assignment_role.code in WORK_ASSIGNMENT_ROLE_EQUIPMENT_TYPES
            ),
            'effective_specialization': effective_employee_specialization,
            'temporary_work_transfers': (
                employee.temporary_work_transfers
                .select_related(
                    'source_specialization',
                    'target_specialization',
                    'watch_period',
                    'requested_by',
                    'reviewed_by',
                )
                .order_by('-requested_at')[:8]
            ),
            'can_restore_employee': (
                employee.status in ADMIN_RESTORABLE_EMPLOYEE_STATUSES
                or not employee.is_active
            ),
            'logs': AdminActionLog.objects.filter(
                Q(object_type='Employee', object_id=str(employee.id))
                | Q(object_id='', object_repr=str(employee))
            )[:10],
        },
    )


def system_admin_generate_access_view(request, employee_id):
    access = require_admin_access(request)
    if not access:
        return redirect('role_home')
    employee = get_object_or_404(Employee, id=employee_id)
    if request.method == 'POST':
        form = AdminAccessRoleForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                employee = Employee.objects.select_for_update().get(pk=employee.pk)
                if EmployeeShift.objects.filter(employee=employee, closed_at__isnull=True).exists():
                    messages.error(request, 'Сначала закройте текущую смену сотрудника, затем сбросьте PIN.')
                    return redirect_after_admin_action(
                        request,
                        'system_admin_employee_detail',
                        employee_id=employee.id,
                    )
                role = form.cleaned_data['role']
                if (
                    employee.personnel_position_id
                    and role.code in {'driver', 'excavator_operator'}
                    and not employee_has_effective_access_role(employee, role.code)
                ):
                    messages.error(
                        request,
                        'Сначала назначьте сотруднику подходящую производственную специализацию.',
                    )
                    return redirect_after_admin_action(
                        request,
                        'system_admin_employee_detail',
                        employee_id=employee.id,
                    )
                code = generate_unique_access_code()
                employee_access, _created = EmployeeAccess.objects.update_or_create(
                    employee=employee,
                    role=role,
                    defaults={
                        'access_code': code,
                        'status': EmployeeAccess.Status.NOT_ACTIVATED,
                        'is_active': True,
                        'primary_code_issued_at': timezone.now(),
                        'activated_at': None,
                        'deactivated_at': None,
                        'blocked_at': None,
                        'block_reason': '',
                    },
                )
                log_admin_action(access.employee, 'Выдан новый первичный пинкод', employee_access, new_value=code)
            messages.success(request, f'Новый первичный пинкод: {code}')
    return redirect_after_admin_action(request, 'system_admin_employee_detail', employee_id=employee.id)


@require_POST
def system_admin_change_access_role_view(request, access_id):
    admin_access = require_admin_access(request)
    if not admin_access:
        return redirect('role_home')

    form = AdminAccessRoleForm(request.POST)
    employee_access = get_object_or_404(
        EmployeeAccess.objects.select_related('employee', 'role'),
        id=access_id,
    )
    employee_id = employee_access.employee_id
    if not form.is_valid():
        messages.error(request, 'Выберите новую роль сотрудника.')
        return redirect_after_admin_action(request, 'system_admin_employee_detail', employee_id=employee_id)

    new_role = form.cleaned_data['role']
    if employee_access.role_id == new_role.id:
        messages.info(request, 'У сотрудника уже назначена эта роль.')
        return redirect_after_admin_action(request, 'system_admin_employee_detail', employee_id=employee_id)
    if employee_access.id == admin_access.id:
        messages.error(request, 'Нельзя изменить собственную роль администратора.')
        return redirect_after_admin_action(request, 'system_admin_employee_detail', employee_id=employee_id)

    with transaction.atomic():
        locked_employee = Employee.objects.select_for_update().get(id=employee_id)
        employee_access = (
            EmployeeAccess.objects
            .select_for_update()
            .select_related('role')
            .get(id=access_id, employee_id=employee_id)
        )
        if EmployeeShift.objects.filter(employee_id=employee_id, closed_at__isnull=True).exists():
            messages.error(request, 'Сначала закройте текущую смену сотрудника, затем измените его роль.')
            return redirect_after_admin_action(request, 'system_admin_employee_detail', employee_id=employee_id)
        if (
            locked_employee.personnel_position_id
            and (
                employee_access.role.code in {'driver', 'excavator_operator'}
                or new_role.code in {'driver', 'excavator_operator'}
            )
        ):
            expected_specialization = effective_specialization(locked_employee)
            expected_role_id = getattr(expected_specialization, 'access_role_id', None)
            if expected_role_id != new_role.id:
                messages.error(
                    request,
                    'Производственное приложение меняется через кадровую должность и специализацию, а не вручную.',
                )
                return redirect_after_admin_action(request, 'system_admin_employee_detail', employee_id=employee_id)
        if (
            EmployeeAccess.objects
            .filter(employee_id=employee_id, role=new_role)
            .exclude(id=employee_access.id)
            .exists()
        ):
            messages.error(request, 'У сотрудника уже есть отдельный доступ с выбранной ролью.')
            return redirect_after_admin_action(request, 'system_admin_employee_detail', employee_id=employee_id)

        old_role = employee_access.role
        cleared_assignments = clear_active_equipment_assignment(
            employee=locked_employee,
            assigned_by=admin_access.employee,
            role_code=old_role.code,
        )
        employee_access.role = new_role
        employee_access.save(update_fields=['role'])

    log_admin_action(
        admin_access.employee,
        'Изменена роль доступа сотрудника',
        employee_access,
        old_value=old_role.name,
        new_value=new_role.name,
        comment='PIN, пароль и статус доступа сохранены.',
    )
    assignment_note = ' Старое назначение на технику снято.' if cleared_assignments else ''
    messages.success(
        request,
        f'Роль изменена: {old_role.name} → {new_role.name}. PIN и пароль сохранены.{assignment_note}',
    )
    return redirect_after_admin_action(request, 'system_admin_employee_detail', employee_id=employee_id)


@require_POST
def system_admin_cancel_temporary_work_transfer_view(request, transfer_id):
    admin_access = require_admin_access(request)
    if not admin_access:
        return redirect('role_home')

    transfer = get_object_or_404(
        TemporaryWorkTransfer.objects.select_related(
            'employee',
            'source_specialization',
            'target_specialization',
        ),
        id=transfer_id,
    )
    employee_id = transfer.employee_id
    try:
        with transaction.atomic():
            previous_specialization = transfer.target_specialization
            transfer, restored_access = cancel_temporary_work_transfer(
                transfer=transfer,
                cancelled_by=admin_access.employee,
                comment=request.POST.get('comment', ''),
            )
            AdminActionLog.objects.create(
                actor=admin_access.employee,
                action='Администратор: отменен временный производственный перевод',
                action_code='admin_temporary_transfer_cancelled',
                object_type='TemporaryWorkTransfer',
                object_id=str(transfer.id),
                object_repr=f'{transfer.employee.full_name} — {previous_specialization}',
                old_value=f'Специализация: {previous_specialization}; статус: Одобрен',
                new_value=(
                    f'Специализация: {transfer.source_specialization or "не выбрана"}; '
                    'статус: Отменен'
                ),
                comment=transfer.review_comment,
            )
            bump_operational_state(
                'Employee:admin_temporary_transfer_cancelled',
                event_type='personnel_changed',
                object_type='Employee',
                object_id=employee_id,
                payload={
                    'action': 'admin_temporary_transfer_cancelled',
                    'employee_ids': [employee_id],
                    'access_id': restored_access.id if restored_access else None,
                },
            )
    except ValidationError as error:
        messages.error(request, '; '.join(error.messages))
    else:
        messages.success(request, 'Временный перевод отменен. Сотрудник возвращен на базовую специализацию.')
    return redirect_after_admin_action(request, 'system_admin_employee_detail', employee_id=employee_id)


def system_admin_access_action_view(request, access_id, action):
    admin_access = require_admin_access(request)
    if not admin_access:
        return redirect('role_home')
    employee_access = get_object_or_404(EmployeeAccess.objects.select_related('employee'), id=access_id)
    if request.method == 'POST':
        with transaction.atomic():
            employee = Employee.objects.select_for_update().get(pk=employee_access.employee_id)
            employee_access = (
                EmployeeAccess.objects.select_for_update()
                .select_related('role')
                .get(pk=employee_access.pk, employee=employee)
            )
            if employee_access.id == admin_access.id and action in {'block', 'deactivate'}:
                messages.error(request, 'Нельзя заблокировать или деактивировать собственный доступ администратора.')
                return redirect_after_admin_action(request, 'system_admin_employee_detail', employee_id=employee.id)
            if action in {'block', 'deactivate'} and EmployeeShift.objects.filter(
                employee=employee,
                closed_at__isnull=True,
            ).exists():
                messages.error(request, 'Сначала закройте текущую смену сотрудника, затем измените его доступ.')
                return redirect_after_admin_action(request, 'system_admin_employee_detail', employee_id=employee.id)
            if action == 'block':
                form = AdminAccessBlockForm(request.POST)
                if form.is_valid():
                    employee_access.status = EmployeeAccess.Status.BLOCKED
                    employee_access.is_active = False
                    employee_access.blocked_at = timezone.now()
                    employee_access.block_reason = form.cleaned_data['reason']
                    employee_access.save(update_fields=['status', 'is_active', 'blocked_at', 'block_reason'])
                    log_admin_action(admin_access.employee, 'Заблокирован доступ', employee_access, comment=employee_access.block_reason)
                    messages.success(request, 'Доступ заблокирован.')
            elif action == 'unblock':
                if employee.status in {
                    Employee.Status.ARCHIVED,
                    Employee.Status.DISMISSED,
                    Employee.Status.DELETED,
                }:
                    messages.error(
                        request,
                        'Сначала восстановите сотрудника отдельным действием, затем разблокируйте доступ.',
                    )
                    return redirect_after_admin_action(request, 'system_admin_employee_detail', employee_id=employee.id)
                employee_access.status = EmployeeAccess.Status.ACTIVATED
                employee_access.is_active = True
                employee_access.blocked_at = None
                employee_access.block_reason = ''
                employee_access.deactivated_at = None
                employee_access.save(update_fields=['status', 'is_active', 'blocked_at', 'block_reason', 'deactivated_at'])
                if employee.status in {
                    Employee.Status.NOT_ACTIVATED,
                    Employee.Status.DEACTIVATED,
                }:
                    employee.status = Employee.Status.ACTIVE
                    employee.is_active = True
                    employee.save(update_fields=['status', 'is_active'])
                log_admin_action(admin_access.employee, 'Разблокирован доступ', employee_access)
                messages.success(request, 'Доступ разблокирован.')
            elif action == 'deactivate':
                employee_access.status = EmployeeAccess.Status.DEACTIVATED
                employee_access.is_active = False
                employee_access.deactivated_at = timezone.now()
                employee_access.save(update_fields=['status', 'is_active', 'deactivated_at'])
                clear_active_equipment_assignment(
                    employee=employee,
                    assigned_by=admin_access.employee,
                    role_code=employee_access.role.code,
                )
                log_admin_action(admin_access.employee, 'Доступ деактивирован', employee_access)
                messages.success(request, 'Доступ деактивирован.')
    return redirect_after_admin_action(request, 'system_admin_employee_detail', employee_id=employee_access.employee.id)


def restore_employee_access(employee, requested_access_id=None):
    accesses = (
        EmployeeAccess.objects.select_for_update()
        .select_related('role')
        .filter(employee=employee)
    )
    employee_access = None
    if requested_access_id and str(requested_access_id).isdigit():
        employee_access = accesses.filter(id=int(requested_access_id)).first()
    if not employee_access:
        employee_access = (
            accesses.filter(status=EmployeeAccess.Status.DEACTIVATED)
            .order_by('-created_at', '-id')
            .first()
            or accesses.order_by('-is_active', '-created_at', '-id').first()
        )
    if not employee_access:
        return None, 'missing'

    if (
        employee_access.status == EmployeeAccess.Status.BLOCKED
        or employee_access.blocked_at
        or employee_access.block_reason
    ):
        if (
            employee_access.status != EmployeeAccess.Status.BLOCKED
            or employee_access.is_active
            or employee_access.deactivated_at
        ):
            employee_access.status = EmployeeAccess.Status.BLOCKED
            employee_access.is_active = False
            employee_access.deactivated_at = None
            employee_access.save(
                update_fields=['status', 'is_active', 'deactivated_at']
            )
        return employee_access, 'blocked'

    if employee_access.is_active and employee_access.status != EmployeeAccess.Status.DEACTIVATED:
        return employee_access, 'already_active'

    if employee_access.activated_at:
        employee_access.status = EmployeeAccess.Status.ACTIVATED
    elif employee_access.primary_code_issued_at:
        employee_access.status = EmployeeAccess.Status.NOT_ACTIVATED
    elif employee_access.access_code and employee_access.last_login_at:
        employee_access.status = EmployeeAccess.Status.ACTIVATED
    else:
        employee_access.status = EmployeeAccess.Status.NOT_ACTIVATED
    employee_access.is_active = True
    employee_access.deactivated_at = None
    employee_access.blocked_at = None
    employee_access.block_reason = ''
    employee_access.save(
        update_fields=[
            'status',
            'is_active',
            'deactivated_at',
            'blocked_at',
            'block_reason',
        ]
    )
    return employee_access, 'restored'


def system_admin_employee_status_action_view(request, employee_id, action):
    access = require_admin_access(request)
    if not access:
        return redirect('role_home')
    employee = get_object_or_404(Employee, id=employee_id)
    # Деактивация и удаление защищённой карточки запрещены моделью. Без этой
    # проверки администратор получил бы вместо объяснения страницу с ошибкой.
    if (
        request.method == 'POST'
        and employee.is_protected
        and action != 'restore'
        and access.employee_id != employee.id
    ):
        messages.error(
            request,
            'Карточка защищена: изменить или закрыть её может только её владелец.',
        )
        return redirect_after_admin_action(
            request, 'system_admin_employee_detail', employee_id=employee.id,
        )
    if request.method == 'POST':
        with transaction.atomic(), editing_own_protected_card(access, employee):
            employee = Employee.objects.select_for_update().get(pk=employee.pk)
            if action == 'restore':
                if (
                    employee.status not in ADMIN_RESTORABLE_EMPLOYEE_STATUSES
                    and employee.is_active
                ):
                    messages.info(request, 'Сотрудник уже находится в рабочем состоянии.')
                    return redirect_after_admin_action(
                        request,
                        'system_admin_employee_detail',
                        employee_id=employee.id,
                    )

                old_status = employee.get_status_display()
                employee.status = Employee.Status.ACTIVE
                employee.is_active = True
                employee.dismissed_at = None
                employee.save(
                    update_fields=['status', 'is_active', 'dismissed_at', 'updated_at']
                )
                employee_access, access_result = restore_employee_access(
                    employee,
                    request.POST.get('access_id'),
                )

                if access_result == 'restored':
                    if employee_access.status == EmployeeAccess.Status.ACTIVATED:
                        access_note = (
                            f' Доступ «{employee_access.role.name}» включен; '
                            'действующий PIN/пароль сохранен.'
                        )
                    elif employee_access.access_code:
                        access_note = (
                            f' Доступ «{employee_access.role.name}» возвращен в ожидание '
                            'первого входа; первичный PIN сохранен.'
                        )
                    else:
                        access_note = (
                            f' Доступ «{employee_access.role.name}» включен, но PIN еще '
                            'не выдан.'
                        )
                elif access_result == 'blocked':
                    access_note = (
                        f' Доступ «{employee_access.role.name}» остался заблокированным; '
                        'разблокируйте его отдельно.'
                    )
                elif access_result == 'already_active':
                    access_note = (
                        f' Доступ «{employee_access.role.name}» уже был активен; '
                        'PIN/пароль не изменялся.'
                    )
                else:
                    access_note = ' Доступ не найден; назначьте роль и выдайте PIN отдельно.'

                log_admin_action(
                    access.employee,
                    'Сотрудник восстановлен администратором',
                    employee,
                    old_value=old_status,
                    new_value=employee.get_status_display(),
                    comment=(
                        access_note.strip()
                        + ' Смена и техника автоматически не восстанавливались.'
                    ),
                )
                bump_operational_state(
                    'Employee:admin_restored',
                    event_type='personnel_changed',
                    object_type='Employee',
                    object_id=employee.id,
                    payload={
                        'action': 'admin_restored',
                        'employee_ids': [employee.id],
                        'status': employee.status,
                        'is_active': employee.is_active,
                        'access_id': employee_access.id if employee_access else None,
                        'access_result': access_result,
                    },
                )
                messages.success(
                    request,
                    'Сотрудник восстановлен.' + access_note
                    + ' Смена и техника не назначались автоматически.',
                )
                return redirect_after_admin_action(
                    request,
                    'system_admin_employee_detail',
                    employee_id=employee.id,
                )

            if employee.id == access.employee.id and action in {'deactivate', 'archive', 'delete'}:
                messages.error(request, 'Нельзя деактивировать, архивировать или удалить собственную учетную запись администратора.')
                return redirect_after_admin_action(request, 'system_admin_employee_detail', employee_id=employee.id)
            if action in {'deactivate', 'archive', 'delete'} and EmployeeShift.objects.filter(
                employee=employee,
                closed_at__isnull=True,
            ).exists():
                messages.error(request, 'Сначала закройте текущую смену сотрудника, затем измените его статус.')
                return redirect_after_admin_action(request, 'system_admin_employee_detail', employee_id=employee.id)
            if action == 'deactivate':
                employee.status = Employee.Status.DEACTIVATED
                employee.is_active = False
                employee.accesses.update(status=EmployeeAccess.Status.DEACTIVATED, is_active=False, deactivated_at=timezone.now())
                clear_active_equipment_assignment(employee=employee, assigned_by=access.employee)
                messages.success(request, 'Сотрудник деактивирован.')
                log_admin_action(access.employee, 'Сотрудник деактивирован', employee)
            elif action == 'archive':
                employee.status = Employee.Status.ARCHIVED
                employee.is_active = False
                employee.accesses.update(status=EmployeeAccess.Status.DEACTIVATED, is_active=False, deactivated_at=timezone.now())
                clear_active_equipment_assignment(employee=employee, assigned_by=access.employee)
                messages.success(request, 'Сотрудник отправлен в архив.')
                log_admin_action(access.employee, 'Сотрудник отправлен в архив', employee)
            elif action == 'delete':
                if employee.has_production_history():
                    AdminConflict.objects.create(
                        employee=employee,
                        role=employee.accesses.select_related('role').first().role if employee.accesses.exists() else None,
                        conflict_type='Попытка удаления сотрудника с историей',
                        process='Админка MVP',
                        description='Полное удаление заблокировано: у сотрудника есть смены, рейсы, простои, назначения, перевахта или другие учетные действия.',
                    )
                    messages.error(request, 'Удаление запрещено: у сотрудника есть учетная история. Используйте архив.')
                    log_admin_action(access.employee, 'Удаление сотрудника заблокировано', employee)
                    return redirect_after_admin_action(request, 'system_admin_employee_detail', employee_id=employee.id)
                employee_name = employee.full_name
                log_admin_action(access.employee, 'Сотрудник полностью удален', employee, old_value=employee_name)
                employee.delete()
                messages.success(request, f'Сотрудник {employee_name} удален.')
                return redirect('system_admin_employees')
            else:
                return redirect_after_admin_action(request, 'system_admin_employee_detail', employee_id=employee.id)
            employee.save(update_fields=['status', 'is_active', 'updated_at'])
    return redirect_after_admin_action(request, 'system_admin_employee_detail', employee_id=employee.id)


def system_admin_employee_export_view(request):
    access = require_admin_access(request)
    if not access:
        return redirect('role_home')
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Сотрудники'
    sheet.append([
        'ФИО', 'Табельный номер', 'Телефон', 'Пол', 'Статус', 'Подразделение',
        'Дата приема', 'Дата увольнения', 'График работы', 'Бригада',
        'Утверждённый состав вахты', 'Место проживания',
    ])
    for employee in Employee.objects.select_related(
        'personnel_department',
        'work_schedule',
        'watch_composition',
    ).order_by('full_name'):
        sheet.append([
            employee.full_name,
            employee.personnel_number,
            employee.phone,
            employee.get_sex_display(),
            employee.get_status_display(),
            employee.department_label,
            excel_value(employee.hired_at),
            excel_value(employee.dismissed_at),
            employee.work_schedule_label,
            employee.get_brigade_number_display() if employee.brigade_number else '',
            str(employee.watch_composition or ''),
            employee.residence_text,
        ])
    return build_workbook_response(workbook, 'admin_employees.xlsx')


def system_admin_access_export_view(request):
    access = require_admin_access(request)
    if not access:
        return redirect('role_home')
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Доступы'
    sheet.append(['Сотрудник', 'Роль', 'Статус доступа', 'Дата выдачи', 'Дата активации', 'Последний вход'])
    for employee_access in EmployeeAccess.objects.select_related('employee', 'role').order_by('employee__full_name'):
        sheet.append([
            employee_access.employee.full_name,
            employee_access.role.name,
            employee_access.get_status_display(),
            excel_value(employee_access.primary_code_issued_at),
            excel_value(employee_access.activated_at),
            excel_value(employee_access.last_login_at),
        ])
    return build_workbook_response(workbook, 'admin_accesses.xlsx')


def system_admin_log_export_view(request):
    access = require_admin_access(request)
    if not access:
        return redirect('role_home')
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Журнал действий'
    sheet.append([
        'Дата', 'Кто', 'Действие', 'Код действия', 'Тип объекта', 'Объект',
        'Комментарий', 'Отменяет запись', 'Дата отмены', 'Кто отменил',
    ])
    logs = AdminActionLog.objects.select_related('actor', 'reversal_of').order_by('-created_at')
    reversals = {
        item.reversal_of_id: item
        for item in logs
        if item.reversal_of_id
    }
    for log in logs:
        reversal = reversals.get(log.id)
        sheet.append([
            excel_value(log.created_at),
            log.actor.full_name if log.actor else '',
            log.action,
            log.action_code,
            log.object_type,
            log.object_repr,
            log.comment,
            log.reversal_of_id or '',
            excel_value(reversal.created_at) if reversal else '',
            reversal.actor.full_name if reversal and reversal.actor else '',
        ])
    return build_workbook_response(workbook, 'admin_action_log.xlsx')


def system_admin_conflict_export_view(request):
    access = require_admin_access(request)
    if not access:
        return redirect('role_home')
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Конфликты'
    sheet.append(['Дата', 'Сотрудник', 'Роль', 'Тип', 'Процесс', 'Статус', 'Описание'])
    for conflict in AdminConflict.objects.select_related('employee', 'role').order_by('-created_at'):
        sheet.append([
            excel_value(conflict.created_at),
            conflict.employee.full_name if conflict.employee else '',
            conflict.role.name if conflict.role else '',
            conflict.conflict_type,
            conflict.process,
            conflict.get_status_display(),
            conflict.description,
        ])
    return build_workbook_response(workbook, 'admin_conflicts.xlsx')


def driver_registration_view(request):
    access_id = request.session.get('employee_access_id')
    if not access_id:
        return redirect('login')
    access = EmployeeAccess.objects.select_related('employee', 'role').filter(id=access_id, is_active=True).first()
    if not access or access.role.code != 'driver':
        return redirect('role_home')

    registration = getattr(access.employee, 'driver_registration', None)
    if registration:
        return redirect('role_home')

    if request.method == 'POST':
        form = DriverPrimaryRegistrationForm(request.POST, employee=access.employee)
        if form.is_valid():
            DriverPrimaryRegistration.objects.create(employee=access.employee, **form.cleaned_data)
            messages.success(request, 'Первичная регистрация сохранена.')
            return redirect('role_home')
    else:
        form = DriverPrimaryRegistrationForm(employee=access.employee)

    return render(request, 'users/driver_registration.html', {'form': form, 'access': access})


def driver_format_duration_label(seconds):
    seconds = max(0, int(seconds or 0))
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    return f'{hours:02d}:{minutes:02d}:{seconds % 60:02d}'


def driver_report_duration_label(seconds, *, total=False):
    rounded_minutes = max(0, int((max(0, int(seconds or 0)) + 30) // 60))
    hours, minutes = divmod(rounded_minutes, 60)
    if total and hours:
        return f'{hours}:{minutes:02d} мин.'
    if hours and minutes:
        return f'{hours} ч. {minutes} мин.'
    if hours:
        hour_word = 'час' if hours % 10 == 1 and hours % 100 != 11 else ('часа' if hours % 10 in {2, 3, 4} and hours % 100 not in {12, 13, 14} else 'часов')
        return f'{hours} {hour_word}.'
    return f'{rounded_minutes} мин.'


def driver_shift_downtime_seconds(equipment, shift, *, until=None):
    if not equipment or not shift or not shift.opened_at:
        return 0
    calculation_end = until or timezone.now()
    source_period_end = shift.closed_at or calculation_end
    events = DowntimeEvent.objects.filter(
        equipment=equipment,
        started_at__gte=shift.opened_at,
        started_at__lt=source_period_end,
    )
    total_seconds = 0
    for event in events.only('started_at', 'ended_at'):
        event_end = min(event.ended_at or calculation_end, calculation_end)
        total_seconds += max(0, int((event_end - event.started_at).total_seconds()))
    return total_seconds


def driver_downtime_reason_status_key(reason):
    if reason:
        return reason.effective_color_group
    return 'yellow'


def driver_downtime_event_payload(event, *, action='', closed=False, shift=None):
    now = timezone.now()
    started_at = event.started_at or now
    ended_at = event.ended_at
    elapsed_until = ended_at or now
    elapsed_seconds = max(0, int((elapsed_until - started_at).total_seconds()))
    reason = event.reason if event.reason_id else None
    workflow = driver_downtime_flow(reason)
    shift_total_seconds = driver_shift_downtime_seconds(event.equipment, shift)
    return {
        'ok': True,
        'action': action,
        'active': not bool(ended_at),
        'closed': bool(closed),
        'event_id': event.id,
        'reason_id': event.reason_id,
        'reason': str(reason) if reason else '',
        'reason_label': reason.button_label if reason else '',
        'workflow': workflow,
        'requires_loaded_trip': workflow == DRIVER_DOWNTIME_FLOW_WAITING_UNLOAD,
        'requires_empty_truck': driver_downtime_requires_empty_truck(reason),
        'started_at': started_at.isoformat(),
        'ended_at': ended_at.isoformat() if ended_at else '',
        'elapsed_seconds': elapsed_seconds,
        'elapsed_label': driver_format_duration_label(elapsed_seconds),
        'shift_total_seconds': shift_total_seconds,
        'shift_total_label': driver_format_duration_label(shift_total_seconds),
        'status_key': driver_downtime_reason_status_key(reason),
    }


def driver_json_payload(request):
    if 'application/json' not in (request.headers.get('Content-Type') or ''):
        return request.POST
    try:
        return json.loads(request.body.decode('utf-8') or '{}')
    except (TypeError, ValueError, UnicodeDecodeError):
        return {}


def driver_wants_json(request):
    return (
        request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        or 'application/json' in (request.headers.get('Accept') or '')
    )


def driver_employee_short_name(employee):
    parts = [part for part in (getattr(employee, 'full_name', '') or '').split() if part]
    if not parts:
        return 'Водитель'
    if len(parts) == 1:
        return parts[0]
    initials = ''.join(f'{part[0]}.' for part in parts[1:3] if part)
    return f'{parts[0]} {initials}'.strip()


def driver_equipment_number(equipment):
    return str(getattr(equipment, 'garage_number', '') or equipment or '').strip()


def driver_excavator_short_label(equipment):
    number = driver_equipment_number(equipment)
    if not number:
        return '—'
    upper_number = number.upper()
    if upper_number.startswith(('ЭКС', 'ЭКГ', 'EX')):
        return number
    return f'ЭКС-{number}'


def driver_complex_label_for_excavator(equipment):
    number = driver_equipment_number(equipment)
    if not number:
        return 'К-—'
    upper_number = number.upper()
    if upper_number.startswith('К-'):
        return number
    return f'К-{number}'


def driver_prefixed_context_value(prefix, value):
    value = str(value or '').strip()
    if not value:
        return f'{prefix} —'
    if value.lower().startswith(prefix.lower()):
        return value
    return f'{prefix} {value}'


def driver_compact_context_value(prefix, compact_prefix, value):
    value = str(value or '').strip()
    if not value:
        return f'{compact_prefix}—'
    for candidate in (prefix, compact_prefix):
        if value.lower().startswith(candidate.lower()):
            value = value[len(candidate):].strip()
            break
    return f'{compact_prefix}{value}'


def driver_open_shift_queryset(employee):
    return (
        EmployeeShift.objects
        .filter(employee=employee, closed_at__isnull=True)
        .filter(
            Q(workplace_code='driver')
            | Q(workplace_code='', equipment__equipment_type__name='Самосвал')
        )
    )


def driver_assignment_countdown_label(assignment):
    if not assignment:
        return '05:00'
    if assignment.effective_at:
        remaining_seconds = max(0, int((assignment.effective_at - timezone.now()).total_seconds()))
    elif assignment.assigned_at:
        elapsed_seconds = max(0, int((timezone.now() - assignment.assigned_at).total_seconds()))
        remaining_seconds = max(0, (5 * 60) - elapsed_seconds)
    else:
        remaining_seconds = 5 * 60
    minutes, seconds = divmod(remaining_seconds, 60)
    return f'{minutes:02d}:{seconds:02d}'


def shift_plan_display_context(progress):
    status = progress.get('plan_status') if progress else ''
    percent_value = progress.get('progress_percent') if progress else None
    has_plan = percent_value is not None
    visual = progress_cycle_visual_context(percent_value if has_plan else 0)
    plan_value = progress.get('plan_value') if progress else None
    calculation_mode = progress.get('calculation_mode') if progress else ''
    return {
        'percent': visual['percent'] if has_plan else 0,
        'status': status or PlanAssignmentStatus.NO_PLAN_GROUP,
        'status_label': plan_status_label(status),
        'short_label': 'Нет группы' if status == PlanAssignmentStatus.NO_PLAN_GROUP else 'Нет плана' if status == PlanAssignmentStatus.NO_ACTIVE_PLAN else plan_status_label(status),
        'has_plan': has_plan,
        'value': plan_value,
        'unit': plan_unit_label(calculation_mode),
        'group_name': progress.get('plan_group_name') if progress else '',
        'visual': visual,
    }


def driver_shift_view(request):
    requested_fragment = request.GET.get('_operational_fragment', '').strip()
    access_id = request.session.get('employee_access_id')
    if not access_id:
        if requested_fragment == 'driver':
            return JsonResponse({'authenticated': False}, status=401)
        return redirect('login')
    access = EmployeeAccess.objects.select_related('employee', 'role').filter(id=access_id, is_active=True).first()
    if not access or access.role.code != 'driver':
        return redirect('role_home')
    registration = getattr(access.employee, 'driver_registration', None)

    posted_client_action_id = request.POST.get('client_action_id', '').strip() if request.method == 'POST' else ''
    if posted_client_action_id and ShiftClientAction.objects.filter(
        action_type='driver_shift_opened',
        client_action_id=posted_client_action_id,
        employee=access.employee,
    ).exists():
        messages.success(request, 'Смена открыта.')
        return redirect('driver_work')

    open_shift = driver_open_shift_queryset(access.employee).order_by('-opened_at').first()
    last_closed_shift = (
        EmployeeShift.objects
        .select_related('equipment', 'equipment__equipment_type')
        .filter(employee=access.employee, closed_at__isnull=False)
        .filter(
            Q(workplace_code='driver')
            | Q(workplace_code='', equipment__equipment_type__name='Самосвал')
        )
        .order_by('-closed_at')
        .first()
    )
    report_shift = open_shift or last_closed_shift
    report_truck = report_shift.equipment if report_shift else None
    work_assignment = get_active_equipment_assignment(access.employee, 'driver')
    assignment_state = work_assignment_state(access.employee, work_assignment)
    shift_start_conflict_message = ''
    if work_assignment and assignment_state == 'assignment_conflict':
        busy_shift = (
            EmployeeShift.objects
            .select_related('employee', 'equipment', 'equipment__equipment_type')
            .filter(
                equipment=work_assignment.equipment,
                closed_at__isnull=True,
            )
            .exclude(employee=access.employee)
            .order_by('id')
            .first()
        )
        shift_start_conflict_message = open_shift_conflict_message(
            busy_shift,
            equipment=work_assignment.equipment,
        )
    current_truck = open_shift.equipment if open_shift else None
    assigned_truck = work_assignment.equipment if work_assignment and assignment_state == 'assigned' else None
    assignment_truck = (
        work_assignment.equipment
        if work_assignment and assignment_state in {'assigned', 'assignment_conflict'}
        else None
    )
    header_truck = current_truck or assignment_truck
    if current_truck:
        reconcile_due_haul_assignments(truck_id=current_truck.id)
    current_assignment = None
    pending_assignment_action = None
    active_trip = None
    active_downtime = None
    shift_trips = []
    if current_truck:
        open_assignments = list(
            HaulAssignment.objects
            .filter(
                truck=current_truck,
                ended_at__isnull=True,
            )
            .exclude(status=AssignmentStatus.CANCELLED)
            .select_related('truck', 'excavator')
            .order_by('-assigned_at')
        )
        accepted_assignment = next(
            (assignment for assignment in open_assignments if assignment.status == AssignmentStatus.ACCEPTED),
            None,
        )
        pending_assignment = next(
            (assignment for assignment in open_assignments if assignment.status == AssignmentStatus.PENDING),
            None,
        )
        current_assignment = accepted_assignment
        pending_assignment_action = pending_assignment
        active_trip = Trip.objects.filter(
            truck=current_truck,
            status__in=OPEN_TRIP_STATUSES,
        ).select_related(
            'truck',
            'excavator',
            'rock_type',
            'dump_point',
            'assigned_dump_point',
            'actual_dump_point',
        ).order_by('-created_at').first()
        active_downtime = (
            DowntimeEvent.objects
            .select_related('reason', 'reason__equipment_state')
            .filter(equipment=current_truck, ended_at__isnull=True)
            .order_by('-started_at')
            .first()
        )

    if report_shift and report_truck:
        legacy_driver_trip_filter = Q(
            driver=access.employee,
            completed_at__gte=report_shift.opened_at,
        )
        if report_shift.closed_at:
            legacy_driver_trip_filter &= Q(completed_at__lte=report_shift.closed_at)
        shift_trips = list(
            Trip.objects
            .select_related('excavator', 'rock_type', 'dump_point', 'assigned_dump_point', 'actual_dump_point')
            .filter(
                Q(unloading_shift=report_shift)
                | Q(loading_shift=report_shift)
                | legacy_driver_trip_filter
            )
            .distinct()
            .order_by('created_at')[:30]
        )

    for trip in shift_trips:
        started_at = timezone.localtime(trip.created_at) if trip.created_at else None
        completed_at = timezone.localtime(trip.completed_at) if trip.completed_at else None
        duration_end = report_shift.closed_at if report_shift and report_shift.closed_at else timezone.now()
        finish_for_duration = completed_at or timezone.localtime(duration_end)
        duration_seconds = 0
        if started_at:
            duration_seconds = max(0, int((finish_for_duration - started_at).total_seconds()))
        duration_minutes = max(1, round(duration_seconds / 60)) if duration_seconds else 0
        trip.driver_excavator_label = trip.excavator.garage_number if trip.excavator_id else '—'
        driver_dump_point = trip.actual_dump_point or trip.dump_point or trip.assigned_dump_point
        trip.driver_dump_point_label = str(driver_dump_point) if driver_dump_point else '—'
        started_label = started_at.strftime('%H:%M') if started_at else '—'
        completed_label = completed_at.strftime('%H:%M') if completed_at else '...'
        trip.driver_time_range_label = f'{started_label}–{completed_label}'
        trip.driver_duration_label = f'{duration_minutes}м' if completed_at else 'в рейсе'

    completed_shift_trips = [trip for trip in shift_trips if trip.status == TripStatus.COMPLETED]
    shift_trip_count = len(completed_shift_trips)
    report_trip_map = {}
    for trip in completed_shift_trips:
        report_key = (trip.driver_excavator_label, trip.driver_dump_point_label)
        report_trip_map.setdefault(report_key, 0)
        report_trip_map[report_key] += 1
    driver_shift_report_trip_rows = [
        {'excavator': key[0], 'dump_point': key[1], 'count': count}
        for key, count in report_trip_map.items()
    ]

    driver_shift_downtime_events = []
    driver_shift_downtime_rows = []
    driver_shift_timeline = []
    if report_truck and report_shift:
        shift_period_end = report_shift.closed_at or timezone.now()
        driver_shift_downtime_events = list(
            DowntimeEvent.objects
            .select_related('reason')
            .filter(
                equipment=report_truck,
                employee=access.employee,
                started_at__lt=shift_period_end,
            )
            .filter(Q(ended_at__isnull=True) | Q(ended_at__gt=report_shift.opened_at))
            .order_by('started_at')
        )
        downtime_totals = {}
        for event in driver_shift_downtime_events:
            overlap_start = max(event.started_at, report_shift.opened_at)
            overlap_end = min(event.ended_at or shift_period_end, shift_period_end)
            duration_seconds = max(0, int((overlap_end - overlap_start).total_seconds()))
            reason_label = event.reason.button_label
            downtime_totals.setdefault(reason_label, 0)
            downtime_totals[reason_label] += duration_seconds
            driver_shift_timeline.append({
                'at': overlap_start,
                'time': timezone.localtime(overlap_start).strftime('%H:%M'),
                'kind': 'downtime-start',
                'title': f'Начат простой: {reason_label}',
                'meta': '',
            })
            if event.ended_at:
                driver_shift_timeline.append({
                    'at': overlap_end,
                    'time': timezone.localtime(overlap_end).strftime('%H:%M'),
                    'kind': 'downtime-end',
                    'title': f'Завершён простой: {reason_label}',
                    'meta': driver_format_duration_label(duration_seconds),
                })
        driver_shift_downtime_rows = [
            {
                'reason': reason,
                'seconds': seconds,
                'duration': driver_report_duration_label(seconds),
            }
            for reason, seconds in downtime_totals.items()
        ]

    for index, trip in enumerate(shift_trips, start=1):
        driver_shift_timeline.append({
            'at': trip.created_at,
            'time': timezone.localtime(trip.created_at).strftime('%H:%M'),
            'kind': 'trip',
            'title': f'Рейс {index:02d} · {trip.driver_excavator_label} → {trip.driver_dump_point_label}',
            'meta': trip.driver_time_range_label,
        })
    driver_shift_timeline.sort(key=lambda item: item['at'])
    shift_progress = calculate_truck_shift_progress(current_truck, reference_shift=open_shift)
    shift_plan = shift_plan_display_context(shift_progress)
    shift_plan_percent = shift_plan['percent']
    active_tab = request.GET.get('tab', 'work' if open_shift else 'shift')
    if active_tab not in {'work', 'shift', 'downtimes', 'manifest'}:
        active_tab = 'work'
    driver_status = 'ПУСТОЙ'
    driver_status_class = 'is-empty'
    driver_target_label = '—'
    driver_trip_context_source = active_trip
    if active_trip:
        driver_status = 'ЗАГРУЖЕН'
        driver_status_class = 'is-loaded'
        driver_target_label = active_trip.actual_dump_point or active_trip.dump_point
    elif active_downtime:
        driver_status = active_downtime.reason.button_label
        driver_status_class = 'is-downtime'

    driver_has_open_trip = bool(active_trip)
    driver_has_loaded_trip = bool(
        active_trip
        and active_trip.status == TripStatus.LOADED_WAITING_UNLOAD
    )
    active_downtime_flow = driver_downtime_flow(
        active_downtime.reason if active_downtime else None
    )
    driver_waiting_operation_active = driver_downtime_opens_work(
        active_downtime.reason if active_downtime else None
    )
    driver_loading_wait_active = bool(
        not driver_has_open_trip
        and active_downtime_flow == DRIVER_DOWNTIME_FLOW_WAITING_LOADING
    )
    driver_unloading_wait_active = bool(
        driver_has_loaded_trip
        and active_downtime_flow == DRIVER_DOWNTIME_FLOW_WAITING_UNLOAD
    )

    driver_work_excavator = active_trip.excavator if active_trip else (current_assignment.excavator if current_assignment else None)
    driver_work_context_placement = None
    if driver_work_excavator:
        driver_work_context_placement = (
            ExcavatorPlacement.objects
            .select_related('work_rock_type', 'work_dump_point')
            .filter(excavator=driver_work_excavator)
            .first()
        )
    if not driver_trip_context_source and driver_work_context_placement:
        has_work_context = any([
            driver_work_context_placement.loading_horizon,
            driver_work_context_placement.loading_block,
            driver_work_context_placement.work_rock_type_id,
        ])
        if has_work_context:
            driver_trip_context_source = driver_work_context_placement
    if not driver_trip_context_source and current_assignment:
        driver_trip_context_source = (
            Trip.objects
            .filter(excavator=current_assignment.excavator)
            .select_related('rock_type', 'dump_point', 'actual_dump_point', 'assigned_dump_point')
            .order_by('-created_at')
            .first()
        )
    driver_header_truck_label = (
        f'Самосвал {driver_equipment_number(header_truck)}'
        if header_truck
        else 'Самосвал'
    )
    driver_header_person_label = driver_employee_short_name(access.employee)
    driver_header_label = (
        f'{driver_header_truck_label} · {driver_header_person_label}'
    )
    driver_context_rock = (
        getattr(driver_trip_context_source, 'rock_type', None)
        or getattr(driver_trip_context_source, 'work_rock_type', None)
    )
    driver_excavator_label = driver_excavator_short_label(driver_work_excavator)
    driver_complex_label = driver_complex_label_for_excavator(driver_work_excavator)
    driver_geology_parts = [
        driver_prefixed_context_value('Горизонт', getattr(driver_trip_context_source, 'loading_horizon', '')),
        driver_prefixed_context_value('Блок', getattr(driver_trip_context_source, 'loading_block', '')),
        str(driver_context_rock or '—'),
    ]
    driver_context_parts = [driver_complex_label, *driver_geology_parts]
    driver_context_label = ' · '.join(driver_context_parts)
    if active_trip:
        driver_dial_label = str(driver_target_label)
        driver_dial_note = (
            active_downtime.reason.button_label.upper()
            if driver_unloading_wait_active
            else 'ТОЧКА РАЗГРУЗКИ'
        )
    elif driver_loading_wait_active:
        # Во время ожидания сохраняем полезный ориентир погрузки в крупной
        # строке, а само состояние показываем в подписи. Так server render и
        # мгновенное AJAX-состояние выглядят одинаково.
        driver_dial_label = driver_excavator_short_label(driver_work_excavator)
        driver_dial_note = active_downtime.reason.button_label.upper()
    elif active_downtime:
        driver_dial_label = active_downtime.reason.button_label
        driver_dial_note = 'ПРИЧИНА ПРОСТОЯ'
    else:
        driver_dial_label = driver_excavator_short_label(driver_work_excavator)
        driver_dial_note = 'НА ЗАГРУЗКУ'
    driver_new_assignment_label = ''
    driver_assignment_action_label = ''
    driver_assignment_effective_at = ''
    driver_assignment_countdown = '05:00'
    if pending_assignment_action:
        if pending_assignment_action.action == HaulAssignmentAction.RELEASE:
            action_label = 'НАЗНАЧЕНИЕ СНЯТО'
        else:
            action_label = f'ВЫ НАЗНАЧЕНЫ НА {driver_excavator_short_label(pending_assignment_action.excavator)}'
        driver_assignment_action_label = action_label
        driver_assignment_countdown = driver_assignment_countdown_label(pending_assignment_action)
        driver_new_assignment_label = (
            f'{action_label} · ПРИНЯТЬ · {driver_assignment_countdown}'
        )
        if pending_assignment_action.effective_at:
            driver_assignment_effective_at = pending_assignment_action.effective_at.isoformat()

    downtime_equipment_type = current_truck.equipment_type if current_truck else None
    downtime_reasons = list(
        DowntimeReason.for_workplace('truck_driver', downtime_equipment_type)
    )
    for reason in downtime_reasons:
        reason.driver_workflow = driver_downtime_flow(reason)
        reason.driver_requires_loaded_trip = driver_downtime_requires_loaded_trip(reason)
        reason.driver_requires_empty_truck = driver_downtime_requires_empty_truck(reason)
        reason.driver_unavailable_message = ''
        if reason.driver_requires_loaded_trip and not driver_has_loaded_trip:
            reason.driver_unavailable_message = 'Доступно только после погрузки'
        elif reason.driver_requires_empty_truck and driver_has_open_trip:
            reason.driver_unavailable_message = 'Самосвал уже загружен'
    unload_points = DumpPoint.objects.filter(is_active=True).order_by('name')[:10]
    active_trip_assigned_dump_point = None
    active_trip_actual_dump_point_id = None
    if active_trip:
        active_trip_assigned_dump_point = active_trip.assigned_dump_point or active_trip.dump_point
        active_trip_actual_dump_point_id = (active_trip.actual_dump_point_id or active_trip.dump_point_id)
    active_downtime_elapsed_seconds = 0
    active_downtime_elapsed_label = '00:00:00'
    shift_downtime_total_seconds = driver_shift_downtime_seconds(
        open_shift.equipment if open_shift else current_truck,
        open_shift,
    )
    shift_downtime_total_label = driver_format_duration_label(shift_downtime_total_seconds)
    shift_downtime_report_total_seconds = sum(
        row['seconds'] for row in driver_shift_downtime_rows
    )
    shift_downtime_report_total_label = driver_report_duration_label(
        shift_downtime_report_total_seconds,
        total=True,
    )
    active_downtime_started_at = ''
    active_downtime_status_key = 'yellow'
    if active_downtime and active_downtime.started_at:
        active_downtime_elapsed_seconds = max(0, int((timezone.now() - active_downtime.started_at).total_seconds()))
        active_downtime_elapsed_label = driver_format_duration_label(active_downtime_elapsed_seconds)
        active_downtime_started_at = active_downtime.started_at.isoformat()
        active_downtime_status_key = driver_downtime_reason_status_key(active_downtime.reason)

    if request.method == 'POST' and not open_shift:
        form = DriverOpenShiftForm(request.POST, employee=access.employee, work_assignment=work_assignment) if assignment_state == 'assigned' else None
        if form and form.is_valid():
            current_work_assignment = get_active_equipment_assignment(access.employee, 'driver')
            if work_assignment_state(access.employee, current_work_assignment) != 'assigned':
                form.add_error(None, 'Назначение изменилось. Обновите экран перед началом смены.')
            else:
                try:
                    with transaction.atomic():
                        Employee.objects.select_for_update().get(pk=access.employee_id)
                        if not role_session_state(request, access)['is_active']:
                            raise ValidationError('Роль неактивна — доступен только просмотр.')
                        open_driver_shift(
                            employee=access.employee,
                            work_assignment=current_work_assignment,
                            readings={
                                'start_fuel': form.cleaned_data['start_fuel'],
                                'start_mileage': form.cleaned_data['start_mileage'],
                                'start_engine_hours': form.cleaned_data['start_engine_hours'],
                            },
                            client_action_id=form.cleaned_data.get('client_action_id') or secrets.token_urlsafe(24),
                        )
                except ValidationError as error:
                    form.add_error(None, error)
                else:
                    messages.success(request, 'Смена открыта.')
                    return redirect('driver_work')
    else:
        form_initial = {}
        if last_closed_shift:
            form_initial = {
                'start_fuel': last_closed_shift.end_fuel,
                'start_mileage': last_closed_shift.end_mileage,
                'start_engine_hours': last_closed_shift.end_engine_hours,
            }
        form_initial['client_action_id'] = secrets.token_urlsafe(24)
        form = (
            DriverOpenShiftForm(initial=form_initial, employee=access.employee, work_assignment=work_assignment)
            if not open_shift and assignment_state == 'assigned'
            else None
        )

    close_form = getattr(request, '_driver_close_form', None)
    if close_form is None and open_shift:
        close_form = DriverCloseShiftForm(
            instance=open_shift,
            initial={'client_action_id': secrets.token_urlsafe(24)},
        )

    operational_state_version = (
        OperationalStateVersion.objects
        .filter(key='production')
        .values_list('version', flat=True)
        .first()
        or 0
    )
    response = render(
        request,
        'users/driver_shift.html',
        {
            'access': access,
            'registration': registration,
            'current_truck': current_truck,
            'header_truck': header_truck,
            'open_shift': open_shift,
            'work_assignment': work_assignment,
            'work_assignment_state': assignment_state,
            'shift_start_conflict_message': shift_start_conflict_message,
            'work_assignment_shift_label': work_assignment.work_shift_label if work_assignment else '',
            'work_assignment_equipment': assignment_truck,
            'current_assignment': current_assignment,
            'active_trip': active_trip,
            'form': form,
            'close_form': close_form,
            'close_review': getattr(request, '_driver_close_review', None),
            'last_closed_shift': last_closed_shift,
            'active_tab': active_tab,
            'active_downtime': active_downtime,
            'active_downtime_started_at': active_downtime_started_at,
            'active_downtime_elapsed_seconds': active_downtime_elapsed_seconds,
            'active_downtime_elapsed_label': active_downtime_elapsed_label,
            'shift_downtime_total_seconds': shift_downtime_total_seconds,
            'shift_downtime_total_label': shift_downtime_total_label,
            'shift_downtime_report_total_label': shift_downtime_report_total_label,
            'active_downtime_status_key': active_downtime_status_key,
            'downtime_reasons': downtime_reasons,
            'shift_trips': shift_trips,
            'shift_trip_count': shift_trip_count,
            'driver_shift_report_trip_rows': driver_shift_report_trip_rows,
            'driver_shift_downtime_rows': driver_shift_downtime_rows,
            'driver_shift_timeline': driver_shift_timeline,
            'driver_shift_report_date': timezone.localtime(report_shift.opened_at).strftime('%d.%m.%Y') if report_shift else '—',
            'driver_shift_report_shift': report_shift.get_shift_type_display() if report_shift else 'Смена не открыта',
            'driver_shift_report_driver': driver_employee_short_name(access.employee),
            'driver_shift_report_truck': driver_equipment_number(report_truck) if report_truck else '—',
            'driver_shift_report_end_fuel': report_shift.end_fuel if report_shift and report_shift.closed_at else None,
            'driver_shift_report_end_mileage': report_shift.end_mileage if report_shift and report_shift.closed_at else None,
            'driver_shift_report_end_engine_hours': report_shift.end_engine_hours if report_shift and report_shift.closed_at else None,
            'shift_plan_percent': shift_plan_percent,
            'shift_plan_status': shift_plan['status'],
            'shift_plan_status_label': shift_plan['status_label'],
            'shift_plan_short_label': shift_plan['short_label'],
            'shift_plan_has_plan': shift_plan['has_plan'],
            'shift_plan_value': shift_plan['value'],
            'shift_plan_unit': shift_plan['unit'],
            'shift_plan_group_name': shift_plan['group_name'],
            'shift_plan_visual': shift_plan['visual'],
            'driver_status': driver_status,
            'driver_status_class': driver_status_class,
            'driver_has_open_trip': driver_has_open_trip,
            'driver_has_loaded_trip': driver_has_loaded_trip,
            'driver_waiting_operation_active': driver_waiting_operation_active,
            'driver_loading_wait_active': driver_loading_wait_active,
            'driver_unloading_wait_active': driver_unloading_wait_active,
            'active_downtime_flow': active_downtime_flow,
            'driver_target_label': driver_target_label,
            'driver_header_label': driver_header_label,
            'driver_header_truck_label': driver_header_truck_label,
            'driver_header_person_label': driver_header_person_label,
            'driver_excavator_label': driver_excavator_label,
            'driver_complex_label': driver_complex_label,
            'driver_geology_parts': driver_geology_parts,
            'driver_context_parts': driver_context_parts,
            'driver_context_label': driver_context_label,
            'driver_dial_label': driver_dial_label,
            'driver_dial_note': driver_dial_note,
            'driver_new_assignment_label': driver_new_assignment_label,
            'driver_assignment_action_label': driver_assignment_action_label,
            'driver_assignment_effective_at': driver_assignment_effective_at,
            'driver_assignment_countdown': driver_assignment_countdown,
            'pending_assignment_action': pending_assignment_action,
            'unload_points': unload_points,
            'active_trip_assigned_dump_point': active_trip_assigned_dump_point,
            'active_trip_actual_dump_point_id': active_trip_actual_dump_point_id,
            'trip_status_loaded': TripStatus.LOADED_WAITING_UNLOAD,
            'driver_shell_version': DRIVER_SHELL_VERSION,
            'operational_state_version': operational_state_version,
        },
    )
    if requested_fragment == 'driver':
        return operational_fragment_response(
            response,
            screen='driver',
            selector='[data-driver-shell]',
            version=operational_state_version,
        )
    response['Cache-Control'] = 'no-cache'
    return response


@require_POST
@transaction.atomic
def driver_accept_assignment_view(request, assignment_id):
    access_id = request.session.get('employee_access_id')
    access = (
        EmployeeAccess.objects.select_related('employee', 'role')
        .filter(id=access_id, is_active=True, role__code='driver')
        .first()
    )
    if not access:
        return JsonResponse({'ok': False, 'error': 'Нет доступа к приложению водителя.'}, status=403)
    Employee.objects.select_for_update().get(pk=access.employee_id)
    if not role_session_state(request, access)['is_active']:
        return JsonResponse(
            {
                'ok': False,
                'error': 'Роль неактивна — доступен только просмотр',
                'code': 'inactive_role',
            },
            status=409,
        )
    open_shift = (
        driver_open_shift_queryset(access.employee)
        .select_for_update(of=('self',))
        .select_related('equipment')
        .order_by('-opened_at')
        .first()
    )
    if not open_shift or not open_shift.equipment_id:
        return JsonResponse({'ok': False, 'error': 'Открытая смена водителя не найдена.'}, status=409)
    assignment = get_object_or_404(
        HaulAssignment.objects.select_for_update(),
        id=assignment_id,
        truck_id=open_shift.equipment_id,
        status=AssignmentStatus.PENDING,
        ended_at__isnull=True,
    )
    applied = apply_pending_haul_assignment(assignment.id)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'ok': bool(applied), 'action': assignment.action})
    return redirect('driver_work')


def driver_close_shift_view(request):
    access_id = request.session.get('employee_access_id')
    if not access_id:
        return redirect('login')
    access = EmployeeAccess.objects.select_related('employee', 'role').filter(id=access_id, is_active=True).first()
    if not access or access.role.code != 'driver':
        return redirect('role_home')

    client_action_id = request.POST.get('client_action_id', '').strip()
    completed_action = lambda: ShiftClientAction.objects.filter(
        action_type='driver_shift_closed',
        client_action_id=client_action_id,
        employee=access.employee,
    ).exists()
    if client_action_id and completed_action():
        messages.success(request, 'Смена закрыта.')
        return redirect(f"{reverse('driver_work')}?tab=manifest")

    open_shift = driver_open_shift_queryset(access.employee).order_by('-opened_at').first()
    if not open_shift:
        if client_action_id and completed_action():
            messages.success(request, 'Смена закрыта.')
            return redirect(f"{reverse('driver_work')}?tab=manifest")
        messages.error(request, 'Открытая смена не найдена.')
        return redirect('driver_work')

    form = DriverCloseShiftForm(request.POST, instance=open_shift)
    request._driver_close_form = form
    if form.is_valid():
        readings = {
            'end_fuel': form.cleaned_data['end_fuel'],
            'end_mileage': form.cleaned_data['end_mileage'],
            'end_engine_hours': form.cleaned_data['end_engine_hours'],
        }
        try:
            with transaction.atomic():
                Employee.objects.select_for_update().get(pk=access.employee_id)
                if not role_session_state(request, access)['is_active']:
                    raise ValidationError('Роль неактивна — доступен только просмотр.')
                close_driver_shift(
                    shift=open_shift,
                    employee=access.employee,
                    readings=readings,
                    client_action_id=form.cleaned_data.get('client_action_id') or secrets.token_urlsafe(24),
                )
        except ValidationError as error:
            form.add_error(None, error)
        else:
            messages.success(request, 'Смена закрыта.')
            return redirect(f"{reverse('driver_work')}?tab=manifest")
    request.GET = request.GET.copy()
    request.GET['tab'] = 'shift'
    return driver_shift_view(request)


@transaction.atomic
def driver_downtime_action_view(request):
    wants_json = driver_wants_json(request)
    access_id = request.session.get('employee_access_id')
    if not access_id:
        if wants_json:
            return JsonResponse({'ok': False, 'error': 'Нет доступа к экрану водителя.'}, status=403)
        return redirect('login')
    access = EmployeeAccess.objects.select_related('employee', 'role').filter(id=access_id, is_active=True).first()
    if not access or access.role.code != 'driver':
        if wants_json:
            return JsonResponse({'ok': False, 'error': 'Нет доступа к экрану водителя.'}, status=403)
        return redirect('role_home')


    if request.method != 'POST':
        if wants_json:
            return JsonResponse({'ok': False, 'error': 'Некорректный метод действия простоя.'}, status=405)
        return redirect(f'{reverse("driver_work")}?tab=downtimes')

    Employee.objects.select_for_update().get(pk=access.employee_id)
    if not role_session_state(request, access)['is_active']:
        if wants_json:
            return JsonResponse(
                {
                    'ok': False,
                    'error': 'Роль неактивна — доступен только просмотр',
                    'code': 'inactive_role',
                },
                status=409,
            )
        messages.error(request, 'Роль неактивна — доступен только просмотр.')
        return redirect(f'{reverse("driver_work")}?tab=downtimes')

    open_shift = (
        driver_open_shift_queryset(access.employee)
        .select_for_update(of=('self',))
        .select_related('equipment', 'equipment__equipment_type')
        .order_by('-opened_at')
        .first()
    )
    if not open_shift or not open_shift.equipment:
        if wants_json:
            return JsonResponse({'ok': False, 'error': 'Нельзя зафиксировать простой: открытая смена с самосвалом не найдена.'}, status=409)
        messages.error(request, 'Нельзя зафиксировать простой: открытая смена с самосвалом не найдена.')
        return redirect(f'{reverse("driver_work")}?tab=downtimes')

    payload = driver_json_payload(request)
    action = (payload.get('action') or '').strip()
    locked_equipment = Equipment.objects.select_for_update().get(pk=open_shift.equipment_id)
    open_shift.equipment = locked_equipment
    if action == 'close':
        active_event = (
            DowntimeEvent.objects
            .select_for_update(of=('self',))
            .select_related('reason', 'reason__equipment_state')
            .filter(equipment=open_shift.equipment, ended_at__isnull=True)
            .order_by('-started_at')
            .first()
        )
        if active_event:
            active_event.ended_at = timezone.now()
            active_event.save(update_fields=['ended_at'])
            if wants_json:
                return JsonResponse(driver_downtime_event_payload(active_event, action='downtime_closed', closed=True, shift=open_shift))
        else:
            if wants_json:
                return JsonResponse({
                    'ok': True,
                    'active': False,
                    'closed': False,
                    'elapsed_seconds': 0,
                    'elapsed_label': '00:00:00',
                    'shift_total_seconds': driver_shift_downtime_seconds(open_shift.equipment, open_shift),
                    'shift_total_label': driver_format_duration_label(driver_shift_downtime_seconds(open_shift.equipment, open_shift)),
                })
            messages.error(request, 'Активный простой не найден.')
        return redirect(f'{reverse("driver_work")}?tab=downtimes')

    reason_id = payload.get('reason_id')
    reason = DowntimeReason.for_workplace('truck_driver', open_shift.equipment.equipment_type).filter(id=reason_id).first()
    if not reason:
        if wants_json:
            return JsonResponse({'ok': False, 'error': 'Причина простоя не найдена.'}, status=400)
        messages.error(request, 'Причина простоя не найдена.')
        return redirect(f'{reverse("driver_work")}?tab=downtimes')
    workflow = driver_downtime_flow(reason)
    if driver_downtime_requires_empty_truck(reason):
        open_trip = (
            Trip.objects
            .select_for_update()
            .filter(
                truck=open_shift.equipment,
                status__in=OPEN_TRIP_STATUSES,
            )
            .order_by('-created_at')
            .first()
        )
        if open_trip:
            error = 'Ожидание погрузки нельзя начать: самосвал уже загружен.'
            if wants_json:
                return JsonResponse(
                    {
                        'ok': False,
                        'error': error,
                        'code': 'empty_truck_required',
                        'workflow': workflow,
                    },
                    status=409,
                )
            messages.error(request, error)
            return redirect(f'{reverse("driver_work")}?tab=downtimes')
    if driver_downtime_requires_loaded_trip(reason):
        loaded_trip = (
            Trip.objects
            .select_for_update()
            .filter(
                truck=open_shift.equipment,
                status=TripStatus.LOADED_WAITING_UNLOAD,
            )
            .order_by('-created_at')
            .first()
        )
        if not loaded_trip:
            error = 'Этот простой доступен только после погрузки самосвала.'
            if wants_json:
                return JsonResponse(
                    {
                        'ok': False,
                        'error': error,
                        'code': 'loaded_trip_required',
                        'workflow': workflow,
                    },
                    status=409,
                )
            messages.error(request, error)
            return redirect(f'{reverse("driver_work")}?tab=downtimes')
    active_event = (
        DowntimeEvent.objects
        .select_for_update(of=('self',))
        .select_related('reason', 'reason__equipment_state')
        .filter(equipment=open_shift.equipment, ended_at__isnull=True)
        .order_by('-started_at')
        .first()
    )
    if active_event:
        if active_event.employee_id != access.employee_id:
            error = (
                'Этот непрерывный простой уже начат другим сотрудником. '
                'Сменщик может завершить его, но не менять причину или автора.'
            )
            if wants_json:
                return JsonResponse(
                    {'ok': False, 'error': error, 'code': 'transferred_downtime_read_only'},
                    status=409,
                )
            messages.error(request, error)
            return redirect(f'{reverse("driver_work")}?tab=downtimes')
        active_event.reason = reason
        active_event.save(update_fields=['reason'])
        event = active_event
        action_label = 'downtime_updated'
    else:
        event = DowntimeEvent.objects.create(
            equipment=open_shift.equipment,
            employee=access.employee,
            reason=reason,
            started_at=timezone.now(),
            comment='Зафиксировано водителем самосвала',
        )
        action_label = 'downtime_started'
    if wants_json:
        return JsonResponse(driver_downtime_event_payload(event, action=action_label, shift=open_shift))
    if driver_downtime_opens_work(reason):
        return redirect(f'{reverse("driver_work")}?tab=work')
    return redirect(f'{reverse("driver_work")}?tab=downtimes')

# Create your views here.

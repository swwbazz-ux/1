import hashlib
import inspect
import io
import json
import tempfile
import zipfile
from dataclasses import replace
from datetime import date, timedelta
from pathlib import Path
from unittest.mock import patch

from django.apps import apps
from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import IntegrityError, connection, models, transaction
from django.db.migrations.executor import MigrationExecutor
from django.db.models import QuerySet
from django.test import Client, TestCase, TransactionTestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from settlement.models import EmployeeBedOccupancy, SettlementCohort, SettlementResident
from shifts.models import WatchPeriod
from users.models import Employee, EmployeeAccess, Role, WatchComposition

from .arrival_roster_parser import UnsafeArrivalWorkbook, parse_arrival_workbook
from .arrival_roster_approvals import (
    _is_confirmed_period_collision,
    build_arrival_roster_confirmation_proposal,
    confirm_arrival_roster_version,
    create_arrival_roster_correction_revision,
)
from .arrival_rosters import (
    _VerifiedTimekeeperContext,
    _access_snapshot,
    _lock_timekeeper_access,
    _trusted_create_arrival_roster_issue,
    _trusted_create_arrival_roster_match,
    _trusted_create_arrival_roster_event,
    _trusted_confirm_arrival_roster_version,
    _trusted_supersede_arrival_roster_version,
    arrival_roster_issue_policy,
    arrival_roster_match_readiness,
    clear_arrival_roster_resident,
    reopen_arrival_roster_issue,
    resolve_arrival_roster_issue,
    search_arrival_roster_residents,
    select_arrival_roster_resident,
    set_arrival_roster_dates,
    set_arrival_roster_notes,
    set_arrival_roster_participation,
    upload_arrival_roster,
)
from .arrival_roster_pool import (
    _employee_snapshot,
    _resident_snapshot,
    _trusted_create_pool_arrival_roster_version,
    _trusted_finalize_pool_arrival_roster_version,
    _trusted_create_pool_row,
    add_employee_to_arrival_roster,
    add_external_resident_to_arrival_roster,
    confirm_unambiguous_arrival_roster_rows,
    create_arrival_roster_from_employee_pool,
)
from .models import (
    ArrivalRosterEvent,
    ArrivalRosterIssue,
    ArrivalRosterIssueResolution,
    ArrivalRosterMatch,
    ArrivalRosterMatchCandidate,
    ArrivalRosterMatchRow,
    ArrivalRosterNormalizedRow,
    ArrivalRosterParserProfile,
    ArrivalRosterPoolRow,
    ArrivalRosterRowReview,
    ArrivalRosterSourceFile,
    ArrivalRosterSourceRow,
    ArrivalRosterVersion,
)


def _workbook_bytes(*, full_name='Иванов Иван Иванович', position='Водитель',
                    phone='+7 999 123-45-67', arrival_value='14.08.2026',
                    primary_name=None, formula_name=False, omit_sheet='',
                    shift_hint=1):
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = {
        'билеты': ['№', 'ФИО', 'Должность', 'Смена', 'Дата прилёта в Хабаровск', 'Маршрут', 'Телефон'],
        'трансфер,сам': ['№', 'ФИО сотрудника', 'Должность', 'сам/трансфер', 'Смена', 'Комментарий'],
        'продление 14.04.26-14.05.26': ['№', 'ФИО', 'Должность', 'Согласование', 'Решение', 'Дата окончания'],
        'Список сотрудников': ['№', 'ФИО', 'Должность', 'Комментарий'],
        'Числ': ['Сверка', 'Значение', '', '', '', '', '', '', '', ''],
    }
    for sheet_name, headers in sheets.items():
        if sheet_name == omit_sheet:
            continue
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)

    tickets = workbook['билеты']
    tickets.append([1, full_name, position, shift_hint, arrival_value, 'Хабаровск', phone])
    primary = workbook['Список сотрудников']
    primary.append([1, primary_name or full_name, position, ''])
    if formula_name:
        primary['B2'] = '=CONCAT("Иванов"," Иван")'
    if 'Числ' in workbook.sheetnames:
        workbook['Числ'].append(['Всего', '=COUNTA(\'Список сотрудников\'!B:B)-1'])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _with_zip_entry(payload, name, content):
    source = zipfile.ZipFile(io.BytesIO(payload))
    output = io.BytesIO()
    with source, zipfile.ZipFile(output, 'w', compression=zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            target.writestr(item, source.read(item.filename))
        target.writestr(name, content)
    return output.getvalue()


def _mark_first_zip_entry_encrypted(payload):
    data = bytearray(payload)
    local = data.find(b'PK\x03\x04')
    central = data.find(b'PK\x01\x02')
    if local < 0 or central < 0:
        raise AssertionError('Не найдены служебные заголовки ZIP.')
    local_flags = int.from_bytes(data[local + 6:local + 8], 'little') | 1
    central_flags = int.from_bytes(data[central + 8:central + 10], 'little') | 1
    data[local + 6:local + 8] = local_flags.to_bytes(2, 'little')
    data[central + 8:central + 10] = central_flags.to_bytes(2, 'little')
    return bytes(data)


def _with_content_outside_profile(payload):
    workbook = load_workbook(io.BytesIO(payload))
    workbook['Список сотрудников']['B250'] = 'Скрытая строка'
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _with_cell_value(payload, sheet_name, coordinate, value):
    workbook = load_workbook(io.BytesIO(payload))
    workbook[sheet_name][coordinate] = value
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _source_row_sha256(raw_values, raw_styles):
    payload = json.dumps(
        {'values': raw_values, 'styles': raw_styles},
        ensure_ascii=False,
        sort_keys=True,
        separators=(',', ':'),
    ).encode('utf-8')
    return hashlib.sha256(payload).hexdigest()


class ArrivalRosterT11Tests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.timekeeper_role = Role.objects.get(code='timekeeper')
        cls.other_role = Role.objects.create(
            code='driver-test-t11', name='Водитель тест', is_active=True,
        )
        cls.actor = Employee.objects.create(
            full_name='Табельщик Тестовый',
            phone='+79990000001',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        cls.access = EmployeeAccess.objects.create(
            employee=cls.actor,
            role=cls.timekeeper_role,
            access_code='710001',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        cls.other_access = EmployeeAccess.objects.create(
            employee=cls.actor,
            role=cls.other_role,
            access_code='710002',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        cls.period = WatchPeriod.objects.create(
            name='Заезд 14.08.2026',
            starts_on=date(2026, 8, 14),
            ends_on=date(2026, 9, 13),
            is_active=True,
        )
        cls.employee = Employee.objects.create(
            full_name='Иванов Иван Иванович',
            position='Водитель',
            phone='+7 999 123-45-67',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        cls.resident = SettlementResident(
            employee=cls.employee,
            resident_type=SettlementResident.ResidentType.EMPLOYEE,
            status=SettlementResident.Status.ACTIVE,
        )
        cls.resident.save()

    def setUp(self):
        self.private_directory = tempfile.TemporaryDirectory(prefix='arrival-roster-test-')
        self.settings_override = override_settings(
            ROTATIONS_PRIVATE_MEDIA_ROOT=Path(self.private_directory.name),
        )
        self.settings_override.enable()
        self.addCleanup(self.settings_override.disable)
        self.addCleanup(self.private_directory.cleanup)

    def _upload(self, payload=None, *, access=None, name='реестр.xlsx'):
        payload = payload if payload is not None else _workbook_bytes()
        uploaded = SimpleUploadedFile(
            name,
            payload,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        return upload_arrival_roster(
            uploaded_file=uploaded,
            watch_period_id=self.period.pk,
            actor_access_id=(access or self.access).pk,
        )

    def _verified_context(self, access=None):
        access = access or self.access
        return _lock_timekeeper_access(_access_snapshot(access.pk))

    def _login(self, client, access=None):
        session = client.session
        session['employee_access_id'] = (access or self.access).pk
        session.save()

    def _source_file_kwargs(self, marker):
        payload = f'публичная подмена {marker}'.encode('utf-8')
        return {
            'sha256': hashlib.sha256(payload).hexdigest(),
            'original_name': f'подмена-{marker}.xlsx',
            'byte_size': len(payload),
            'content_type': 'application/octet-stream',
            'file': ContentFile(payload, name=f'подмена-{marker}.xlsx'),
            'uploaded_by_access': self.access,
        }

    def _source_row(self, version, row_number):
        raw_values = [{'coordinate': f'A{row_number}', 'value': 'подмена'}]
        raw_styles = [{'coordinate': f'A{row_number}', 'style_id': 0}]
        return ArrivalRosterSourceRow(
            version=version,
            sheet_name='Список сотрудников',
            row_number=row_number,
            row_kind=ArrivalRosterSourceRow.RowKind.PERSON,
            raw_values=raw_values,
            raw_styles=raw_styles,
            row_sha256=_source_row_sha256(raw_values, raw_styles),
        )

    def _assert_public_write_forbidden(self, operation):
        with self.assertRaises(ValidationError) as caught:
            operation()
        self.assertEqual(
            caught.exception.code,
            'rotations.arrival_roster.public_write_forbidden',
        )

    def test_upload_saves_private_immutable_version_and_exact_match(self):
        before = {
            'employees': Employee.objects.count(),
            'accesses': EmployeeAccess.objects.count(),
            'residents': SettlementResident.objects.count(),
            'cohorts': SettlementCohort.objects.count(),
            'occupancies': EmployeeBedOccupancy.objects.count(),
        }

        version, created = self._upload()

        self.assertTrue(created)
        self.assertEqual(version.status, ArrivalRosterVersion.Status.REVIEW_REQUIRED)
        self.assertEqual(version.watch_period, self.period)
        self.assertEqual(version.created_by_access, self.access)
        self.assertEqual(version.source_kind, ArrivalRosterVersion.SourceKind.EXCEL)
        self.assertEqual(version.source_fingerprint, version.source_file.sha256)
        self.assertRegex(version.source_file.sha256, r'^[0-9a-f]{64}$')
        self.assertTrue(version.snapshot_sha256)
        self.assertTrue(version.source_file.file.storage.exists(version.source_file.file.name))
        with self.assertRaises(ValueError):
            version.source_file.file.storage.url(version.source_file.file.name)
        match = ArrivalRosterMatch.objects.get(version=version)
        self.assertEqual(match.status, ArrivalRosterMatch.Status.EXACT)
        self.assertEqual(match.matched_resident, self.resident)
        self.assertEqual(match.row_links.count(), 2)
        self.assertEqual(
            before,
            {
                'employees': Employee.objects.count(),
                'accesses': EmployeeAccess.objects.count(),
                'residents': SettlementResident.objects.count(),
                'cohorts': SettlementCohort.objects.count(),
                'occupancies': EmployeeBedOccupancy.objects.count(),
            },
        )
        self.assertEqual(
            set(version.events.values_list('action', flat=True)),
            {ArrivalRosterEvent.Action.UPLOADED, ArrivalRosterEvent.Action.PARSED},
        )
        self.assertTrue(version.issues.filter(code='summary_requires_review').exists())

    def test_same_file_and_period_is_idempotent(self):
        payload = _workbook_bytes()
        first, first_created = self._upload(payload)
        second, second_created = self._upload(payload)
        self.assertTrue(first_created)
        self.assertFalse(second_created)
        self.assertEqual(first.pk, second.pk)
        self.assertEqual(ArrivalRosterVersion.objects.count(), 1)
        self.assertEqual(ArrivalRosterSourceFile.objects.count(), 1)
        self.assertTrue(
            ArrivalRosterEvent.objects.filter(
                version=first, action=ArrivalRosterEvent.Action.REUSED,
            ).exists()
        )

    def test_formula_in_content_is_blocking_but_formula_in_number_is_allowed(self):
        payload = _workbook_bytes(formula_name=True)
        version, _created = self._upload(payload)
        self.assertTrue(
            version.issues.filter(code='formula_in_content', severity='error').exists()
        )
        formula_columns = [
            issue.details['columns']
            for issue in version.issues.filter(code='formula_in_content')
        ]
        self.assertNotIn([1], formula_columns)

    def test_excel_date_and_unambiguous_text_are_parsed(self):
        version, _created = self._upload(_workbook_bytes(arrival_value=date(2026, 8, 14)))
        ticket = ArrivalRosterNormalizedRow.objects.get(
            source_row__version=version,
            source_row__sheet_name='билеты',
        )
        self.assertEqual(ticket.arrival_date_candidate, date(2026, 8, 14))
        self.assertEqual(ticket.date_comment, '')

    def test_ambiguous_date_is_preserved_without_guess(self):
        version, _created = self._upload(
            _workbook_bytes(arrival_value='примерно 13 или 14.08.2026'),
        )
        ticket = ArrivalRosterNormalizedRow.objects.get(
            source_row__version=version,
            source_row__sheet_name='билеты',
        )
        self.assertIsNone(ticket.arrival_date_candidate)
        self.assertEqual(ticket.raw_date, 'примерно 13 или 14.08.2026')
        self.assertTrue(version.issues.filter(code='date_requires_review').exists())

    def test_name_and_position_only_is_probable_not_exact(self):
        version, _created = self._upload(_workbook_bytes(phone=''))
        match = version.matches.filter(
            row_links__normalized_row__source_row__sheet_name='Список сотрудников',
        ).get()
        self.assertEqual(match.status, ArrivalRosterMatch.Status.PROBABLE)
        self.assertIsNone(match.matched_resident)
        self.assertEqual(match.candidates.get().resident, self.resident)

    def test_unique_phone_with_disagreeing_name_is_conflict(self):
        version, _created = self._upload(
            _workbook_bytes(full_name='Петров Пётр Петрович', primary_name='Петров Пётр Петрович'),
        )
        match = version.matches.get(status=ArrivalRosterMatch.Status.CONFLICT)
        self.assertEqual(match.status, ArrivalRosterMatch.Status.CONFLICT)
        self.assertIsNone(match.matched_resident)

    def test_missing_person_is_not_created(self):
        resident_count = SettlementResident.objects.count()
        employee_count = Employee.objects.count()
        version, _created = self._upload(
            _workbook_bytes(
                full_name='Неизвестный Человек Тестовый',
                primary_name='Неизвестный Человек Тестовый',
                phone='+79998887766',
            ),
        )
        self.assertTrue(
            version.matches.filter(status=ArrivalRosterMatch.Status.UNMATCHED).exists()
        )
        self.assertEqual(SettlementResident.objects.count(), resident_count)
        self.assertEqual(Employee.objects.count(), employee_count)

    def test_wrong_or_inactive_exact_access_has_no_fallback(self):
        with self.assertRaises(ValidationError):
            self._upload(access=self.other_access)
        self.access.status = EmployeeAccess.Status.DEACTIVATED
        self.access.save(update_fields=['status'])
        replacement = EmployeeAccess.objects.create(
            employee=self.actor,
            role=self.timekeeper_role,
            access_code='710003',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        with self.assertRaises(ValidationError):
            self._upload(access=self.access)
        self.assertTrue(replacement.is_active)
        self.assertEqual(ArrivalRosterVersion.objects.count(), 0)

    def test_unsafe_archives_are_rejected_without_partial_rows_or_files(self):
        cases = (
            _with_zip_entry(_workbook_bytes(), 'xl/vbaProject.bin', b'macro'),
            _with_zip_entry(_workbook_bytes(), 'xl/externalLinks/externalLink1.xml', b'<externalLink/>'),
            _with_zip_entry(
                _workbook_bytes(),
                '_rels/unsafe.rels',
                b"<Relationships><Relationship TargetMode='External'/></Relationships>",
            ),
            _with_zip_entry(
                _workbook_bytes(),
                'customXml/content.xml',
                b'<Override ContentType="application/vnd.ms-office.vbaProject"/>',
            ),
            _with_zip_entry(_workbook_bytes(), 'xl/media/bomb.bin', b'0' * 1024 * 1024),
            _mark_first_zip_entry_encrypted(_workbook_bytes()),
            b'EncryptedPackage is not an Open XML archive',
        )
        for payload in cases:
            with self.subTest(size=len(payload)):
                with self.assertRaises((ValidationError, UnsafeArrivalWorkbook)):
                    self._upload(payload)
                self.assertEqual(ArrivalRosterVersion.objects.count(), 0)
                self.assertEqual(ArrivalRosterSourceFile.objects.count(), 0)
                self.assertEqual(list(Path(self.private_directory.name).rglob('*')), [])

    def test_archive_entry_and_unpacked_size_limits_are_enforced(self):
        payload = _workbook_bytes()
        for setting_name in ('MAX_ARCHIVE_ENTRIES', 'MAX_UNCOMPRESSED_SIZE'):
            with self.subTest(setting_name=setting_name):
                with patch(f'rotations.arrival_roster_parser.{setting_name}', 1):
                    with self.assertRaises(UnsafeArrivalWorkbook):
                        self._upload(payload)
                self.assertEqual(ArrivalRosterVersion.objects.count(), 0)

    def test_only_xlsx_and_size_limit_are_accepted(self):
        for filename in ('реестр.xls', 'реестр.xlsm'):
            with self.subTest(filename=filename):
                with self.assertRaises(ValidationError):
                    self._upload(_workbook_bytes(), name=filename)
        oversized = SimpleUploadedFile(
            'реестр.xlsx',
            b'x' * (10 * 1024 * 1024 + 1),
            content_type='application/octet-stream',
        )
        with self.assertRaises(ValidationError):
            upload_arrival_roster(
                uploaded_file=oversized,
                watch_period_id=self.period.pk,
                actor_access_id=self.access.pk,
            )
        self.assertEqual(ArrivalRosterVersion.objects.count(), 0)

    def test_result_rows_and_final_version_are_immutable(self):
        version, _created = self._upload()
        row = version.source_rows.filter(row_kind='person').first()
        row.sheet_name = 'подмена'
        with self.assertRaises(ValidationError):
            row.save()
        with self.assertRaises(ValidationError):
            ArrivalRosterSourceRow.objects.filter(pk=row.pk).update(sheet_name='подмена')
        with self.assertRaises(ValidationError):
            ArrivalRosterSourceRow.objects.bulk_update([row], ['sheet_name'])
        with self.assertRaises(ValidationError):
            ArrivalRosterSourceRow.objects.bulk_create([])
        with self.assertRaises(ValidationError):
            row.delete()
        version.status = ArrivalRosterVersion.Status.DRAFT
        with self.assertRaises(ValidationError):
            version.save()

    def test_public_source_file_creation_paths_are_forbidden(self):
        initial_count = ArrivalRosterSourceFile.objects.count()
        operations = [
            lambda: ArrivalRosterSourceFile(**self._source_file_kwargs('save')).save(),
            lambda: ArrivalRosterSourceFile.objects.create(
                **self._source_file_kwargs('create')
            ),
            lambda: ArrivalRosterSourceFile.objects.get_or_create(
                sha256=self._source_file_kwargs('get-or-create')['sha256'],
                defaults={
                    key: value
                    for key, value in self._source_file_kwargs('get-or-create').items()
                    if key != 'sha256'
                },
            ),
            lambda: ArrivalRosterSourceFile.objects.update_or_create(
                sha256=self._source_file_kwargs('update-or-create')['sha256'],
                defaults={
                    key: value
                    for key, value in self._source_file_kwargs('update-or-create').items()
                    if key != 'sha256'
                },
            ),
            lambda: ArrivalRosterSourceFile.objects.bulk_create([
                ArrivalRosterSourceFile(**self._source_file_kwargs('bulk-create')),
            ]),
        ]
        for operation in operations:
            with self.subTest(operation=operation):
                self._assert_public_write_forbidden(operation)
                self.assertEqual(ArrivalRosterSourceFile.objects.count(), initial_count)
        self.assertEqual(
            [path for path in Path(self.private_directory.name).rglob('*') if path.is_file()],
            [],
        )

    def test_public_source_row_creation_paths_are_forbidden(self):
        version, _created = self._upload()
        initial_count = ArrivalRosterSourceRow.objects.count()
        operations = [
            lambda: self._source_row(version, 901).save(),
            lambda: ArrivalRosterSourceRow.objects.create(
                **{
                    field: value
                    for field, value in self._source_row(version, 902).__dict__.items()
                    if field not in {'_state', 'id'}
                }
            ),
            lambda: ArrivalRosterSourceRow.objects.get_or_create(
                version=version,
                sheet_name='Список сотрудников',
                row_number=903,
                defaults={
                    'row_kind': ArrivalRosterSourceRow.RowKind.PERSON,
                    'raw_values': self._source_row(version, 903).raw_values,
                    'raw_styles': self._source_row(version, 903).raw_styles,
                    'row_sha256': self._source_row(version, 903).row_sha256,
                },
            ),
            lambda: ArrivalRosterSourceRow.objects.update_or_create(
                version=version,
                sheet_name='Список сотрудников',
                row_number=904,
                defaults={
                    'row_kind': ArrivalRosterSourceRow.RowKind.PERSON,
                    'raw_values': self._source_row(version, 904).raw_values,
                    'raw_styles': self._source_row(version, 904).raw_styles,
                    'row_sha256': self._source_row(version, 904).row_sha256,
                },
            ),
            lambda: ArrivalRosterSourceRow.objects.bulk_create([
                self._source_row(version, 905),
            ]),
        ]
        for operation in operations:
            with self.subTest(operation=operation):
                self._assert_public_write_forbidden(operation)
                self.assertEqual(ArrivalRosterSourceRow.objects.count(), initial_count)

    def test_trusted_service_recomputes_file_and_row_hashes(self):
        payload = _workbook_bytes()
        version, _created = self._upload(payload)
        self.assertEqual(version.source_file.sha256, hashlib.sha256(payload).hexdigest())
        self.assertEqual(version.source_file.byte_size, len(payload))
        for source_row in version.source_rows.all():
            self.assertEqual(
                source_row.row_sha256,
                _source_row_sha256(source_row.raw_values, source_row.raw_styles),
            )

    def test_tampered_prepared_file_hash_rolls_back_everything(self):
        payload = _workbook_bytes()
        prepared = {
            'payload': payload,
            'sha256': '0' * 64,
            'byte_size': len(payload),
            'original_name': 'реестр.xlsx',
            'content_type': 'application/octet-stream',
        }
        with patch('rotations.arrival_rosters._read_uploaded_xlsx', return_value=prepared):
            with self.assertRaises(ValidationError) as caught:
                self._upload(payload)
        self.assertEqual(
            caught.exception.code,
            'rotations.arrival_roster.source_file_integrity_error',
        )
        self.assertEqual(ArrivalRosterSourceFile.objects.count(), 0)
        self.assertEqual(ArrivalRosterVersion.objects.count(), 0)
        self.assertEqual(ArrivalRosterSourceRow.objects.count(), 0)
        self.assertEqual(
            [path for path in Path(self.private_directory.name).rglob('*') if path.is_file()],
            [],
        )

    def test_one_tampered_row_hash_rolls_back_file_and_all_rows(self):
        payload = _workbook_bytes()
        parsed = parse_arrival_workbook(payload)
        tampered_rows = list(parsed.source_rows)
        tampered_rows[-1] = replace(tampered_rows[-1], row_sha256='0' * 64)
        tampered = replace(parsed, source_rows=tampered_rows)
        with patch('rotations.arrival_rosters.parse_arrival_workbook', return_value=tampered):
            with self.assertRaises(ValidationError) as caught:
                self._upload(payload)
        self.assertEqual(
            caught.exception.code,
            'rotations.arrival_roster.source_row_integrity_error',
        )
        self.assertEqual(ArrivalRosterSourceFile.objects.count(), 0)
        self.assertEqual(ArrivalRosterVersion.objects.count(), 0)
        self.assertEqual(ArrivalRosterSourceRow.objects.count(), 0)
        self.assertEqual(
            [path for path in Path(self.private_directory.name).rglob('*') if path.is_file()],
            [],
        )

    def test_failure_after_private_write_removes_file_and_database_rows(self):
        with patch(
            'rotations.arrival_rosters._persist_parse_result',
            side_effect=RuntimeError('synthetic persistence failure'),
        ):
            with self.assertRaisesRegex(RuntimeError, 'synthetic persistence failure'):
                self._upload()
        self.assertEqual(ArrivalRosterSourceFile.objects.count(), 0)
        self.assertEqual(ArrivalRosterVersion.objects.count(), 0)
        self.assertEqual(
            [path for path in Path(self.private_directory.name).rglob('*') if path.is_file()],
            [],
        )

    def test_shift_values_remain_hints_and_unknown_value_requires_review(self):
        version, _created = self._upload(_workbook_bytes(shift_hint='неясно'))
        ticket = ArrivalRosterNormalizedRow.objects.get(
            source_row__version=version,
            source_row__sheet_name='билеты',
        )
        self.assertEqual(ticket.raw_shift_hint, 'неясно')
        self.assertTrue(version.issues.filter(code='unknown_shift_hint').exists())
        self.assertEqual(version.status, ArrivalRosterVersion.Status.REVIEW_REQUIRED)

    def test_content_outside_profile_range_is_not_silently_ignored(self):
        version, _created = self._upload(
            _with_content_outside_profile(_workbook_bytes()),
        )
        issue = version.issues.get(code='content_outside_profile')
        self.assertEqual(issue.severity, ArrivalRosterIssue.Severity.ERROR)
        self.assertIn('B250', issue.details['sample_coordinates'])

    def test_different_file_creates_next_version(self):
        first, _created = self._upload()
        second, _created = self._upload(_workbook_bytes(arrival_value='15.08.2026'))
        self.assertEqual(first.version_number, 1)
        self.assertEqual(second.version_number, 2)
        self.assertEqual(ArrivalRosterVersion.objects.count(), 2)

    def test_upload_endpoint_is_post_only_csrf_protected_and_hides_phone(self):
        client = Client(enforce_csrf_checks=True)
        self._login(client)
        upload_url = reverse('arrival_roster_upload')
        self.assertEqual(client.get(upload_url).status_code, 405)
        without_csrf = client.post(
            upload_url,
            {'watch_period': self.period.pk, 'workbook': SimpleUploadedFile('реестр.xlsx', _workbook_bytes())},
        )
        self.assertEqual(without_csrf.status_code, 403)
        form_response = client.get(reverse('arrival_roster_upload_form'))
        token = client.cookies['csrftoken'].value
        response = client.post(
            upload_url,
            {
                'csrfmiddlewaretoken': token,
                'watch_period': self.period.pk,
                'workbook': SimpleUploadedFile(
                    'реестр.xlsx',
                    _workbook_bytes(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                ),
                'employee_access_id': self.other_access.pk,
            },
        )
        self.assertEqual(response.status_code, 302)
        version = ArrivalRosterVersion.objects.get()
        self.assertEqual(version.created_by_access, self.access)
        review = client.get(response['Location'])
        self.assertEqual(review.status_code, 200)
        self.assertNotContains(review, '+7 999 123-45-67')
        self.assertContains(review, 'Результат проверки')
        self.assertContains(form_response, 'Загрузка реестра заезда')

    def test_missing_required_sheet_becomes_review_issue(self):
        version, _created = self._upload(_workbook_bytes(omit_sheet='Числ'))
        self.assertEqual(version.status, ArrivalRosterVersion.Status.REVIEW_REQUIRED)
        issue = version.issues.get(code='missing_sheet')
        self.assertEqual(issue.severity, ArrivalRosterIssue.Severity.ERROR)
        self.assertEqual(issue.details, {})

    def test_header_issue_without_normalized_row_is_visible_on_review_screen(self):
        version, _created = self._upload(
            _with_cell_value(_workbook_bytes(), 'билеты', 'B1', 'Неизвестный столбец'),
        )
        client = Client()
        self._login(client)
        response = client.get(reverse('arrival_roster_review', args=[version.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            'Заголовок листа не соответствует выбранному профилю.',
        )

    def test_profile_snapshot_and_source_sha_are_reproducible(self):
        payload = _workbook_bytes()
        version, _created = self._upload(payload)
        self.assertEqual(version.source_file.sha256, hashlib.sha256(payload).hexdigest())
        profile = ArrivalRosterParserProfile.objects.get()
        self.assertRegex(profile.configuration_sha256, r'^[0-9a-f]{64}$')
        parsed = parse_arrival_workbook(payload)
        self.assertEqual(version.source_row_count, len(parsed.source_rows))
        self.assertEqual(version.normalized_row_count, len(parsed.normalized_rows))
        self.assertEqual(ArrivalRosterMatchCandidate.objects.filter(match__version=version).count(), 0)

    def test_t12_selects_and_clears_resident_with_linear_revision_history(self):
        version, _created = self._upload()
        match = version.matches.get()
        review = select_arrival_roster_resident(
            match_id=match.pk,
            resident_id=self.resident.pk,
            expected_revision=0,
            actor_access_id=self.access.pk,
        )
        self.assertEqual(review.revision, 1)
        self.assertEqual(review.selected_resident, self.resident)
        review = clear_arrival_roster_resident(
            match_id=match.pk,
            expected_revision=1,
            actor_access_id=self.access.pk,
        )
        self.assertEqual(review.revision, 2)
        self.assertEqual(review.resident_resolution, ArrivalRosterRowReview.ResidentResolution.CLEARED)
        self.assertIsNone(review.selected_resident_id)
        self.assertEqual(
            list(match.review_events.values_list('action', 'review_revision')),
            [('resident_selected', 1), ('resident_cleared', 2)],
        )

    def test_t12_rejects_duplicate_resident_and_rolls_back_second_review(self):
        version, _created = self._upload()
        first = version.matches.get()
        second = _trusted_create_arrival_roster_match(
            version=version,
            status=ArrivalRosterMatch.Status.UNMATCHED,
            method='manual_test',
            quality='unmatched',
            evidence={'test': True},
        )
        select_arrival_roster_resident(
            match_id=first.pk,
            resident_id=self.resident.pk,
            expected_revision=0,
            actor_access_id=self.access.pk,
        )
        with self.assertRaises(ValidationError) as caught:
            select_arrival_roster_resident(
                match_id=second.pk,
                resident_id=self.resident.pk,
                expected_revision=0,
                actor_access_id=self.access.pk,
            )
        self.assertEqual(caught.exception.code, 'arrival_roster.duplicate_resident')
        self.assertFalse(ArrivalRosterRowReview._base_manager.filter(match=second).exists())
        self.assertFalse(second.review_events.exists())

    def test_t12_participation_mode_dates_and_notes_are_validated(self):
        version, _created = self._upload()
        match = version.matches.get()
        review = set_arrival_roster_participation(
            match_id=match.pk,
            participation_status='arriving',
            arrival_mode='transfer',
            expected_revision=0,
            actor_access_id=self.access.pk,
        )
        review = set_arrival_roster_dates(
            match_id=match.pk,
            arrival_on=date(2026, 8, 14),
            departure_on=date(2026, 9, 13),
            expected_revision=review.revision,
            actor_access_id=self.access.pk,
        )
        review = set_arrival_roster_notes(
            match_id=match.pk,
            basis='Заявка табельщика',
            comment='Дата сверена.',
            expected_revision=review.revision,
            actor_access_id=self.access.pk,
        )
        self.assertEqual(review.revision, 3)
        self.assertEqual(review.arrival_mode, 'transfer')
        self.assertEqual(review.arrival_on, date(2026, 8, 14))
        with self.assertRaises(ValidationError):
            set_arrival_roster_participation(
                match_id=match.pk,
                participation_status='extended',
                arrival_mode='self',
                expected_revision=3,
                actor_access_id=self.access.pk,
            )
        with self.assertRaises(ValidationError):
            set_arrival_roster_dates(
                match_id=match.pk,
                arrival_on=date(2026, 9, 13),
                departure_on=date(2026, 8, 14),
                expected_revision=3,
                actor_access_id=self.access.pk,
            )
        self.assertEqual(match.row_review.revision, 3)

    def test_t12_stale_revision_and_wrong_access_are_fail_closed(self):
        version, _created = self._upload()
        match = version.matches.get()
        select_arrival_roster_resident(
            match_id=match.pk,
            resident_id=self.resident.pk,
            expected_revision=0,
            actor_access_id=self.access.pk,
        )
        with self.assertRaises(ValidationError) as stale:
            clear_arrival_roster_resident(
                match_id=match.pk,
                expected_revision=0,
                actor_access_id=self.access.pk,
            )
        self.assertEqual(stale.exception.code, 'arrival_roster.stale_review_revision')
        with self.assertRaises(ValidationError) as denied:
            clear_arrival_roster_resident(
                match_id=match.pk,
                expected_revision=1,
                actor_access_id=self.other_access.pk,
            )
        self.assertEqual(denied.exception.code, 'arrival_roster.access_denied')
        match.row_review.refresh_from_db()
        self.assertEqual(match.row_review.revision, 1)

    def test_t12_search_is_exact_access_bounded_and_contains_no_phone(self):
        version, _created = self._upload()
        results = search_arrival_roster_residents(
            version_id=version.pk,
            query='Иванов',
            actor_access_id=self.access.pk,
        )
        self.assertEqual(len(results), 1)
        self.assertNotIn('phone', results[0])
        self.assertNotIn('+7 999 123-45-67', json.dumps(results, ensure_ascii=False))
        with self.assertRaises(ValidationError):
            search_arrival_roster_residents(
                version_id=version.pk,
                query='Ив',
                actor_access_id=self.access.pk,
            )
        with self.assertRaises(ValidationError):
            search_arrival_roster_residents(
                version_id=version.pk,
                query='Иванов',
                actor_access_id=self.other_access.pk,
            )

    def test_t12_resolves_and_reopens_nonblocking_warning(self):
        version, _created = self._upload()
        issue = version.issues.get(code='summary_requires_review')
        before_readiness = arrival_roster_match_readiness(match_id=version.matches.get().pk)
        resolution = resolve_arrival_roster_issue(
            issue_id=issue.pk,
            expected_revision=0,
            resolution_note='Сверочный лист проверен вручную.',
            actor_access_id=self.access.pk,
        )
        self.assertTrue(resolution.is_resolved)
        self.assertEqual(resolution.revision, 1)
        client = Client()
        self._login(client)
        response = client.get(reverse('arrival_roster_review', args=[version.pk]))
        self.assertEqual(response.context['open_blocking_count'], 1)
        self.assertEqual(
            arrival_roster_match_readiness(match_id=version.matches.get().pk),
            before_readiness,
        )
        resolution = reopen_arrival_roster_issue(
            issue_id=issue.pk,
            expected_revision=1,
            resolution_note='Нужна повторная проверка сверочного листа.',
            actor_access_id=self.access.pk,
        )
        self.assertFalse(resolution.is_resolved)
        self.assertEqual(resolution.revision, 2)
        self.assertEqual(
            list(issue.review_events.values_list('action', 'review_revision')),
            [('issue_resolved', 1), ('issue_reopened', 2)],
        )

    def test_t12_public_projection_writes_and_deletes_are_forbidden(self):
        version, _created = self._upload(_workbook_bytes(omit_sheet='Числ'))
        match = version.matches.first()
        issue = version.issues.get(code='missing_sheet')
        review = ArrivalRosterRowReview(
            version=version,
            match=match,
            updated_by_access=self.access,
        )
        resolution = ArrivalRosterIssueResolution(
            issue=issue,
            is_resolved=True,
            resolution_note='Проверено вручную.',
            updated_by_access=self.access,
        )
        self._assert_public_write_forbidden(review.save)
        self._assert_public_write_forbidden(lambda: ArrivalRosterRowReview.objects.create(
            version=version, match=match, updated_by_access=self.access,
        ))
        self._assert_public_write_forbidden(
            lambda: ArrivalRosterRowReview.objects.bulk_create([review]),
        )
        self._assert_public_write_forbidden(resolution.save)
        self._assert_public_write_forbidden(
            lambda: ArrivalRosterIssueResolution.objects.bulk_create([resolution]),
        )
        created = select_arrival_roster_resident(
            match_id=match.pk,
            resident_id=self.resident.pk,
            expected_revision=0,
            actor_access_id=self.access.pk,
        )
        self._assert_public_write_forbidden(
            lambda: ArrivalRosterRowReview.objects.filter(pk=created.pk).update(comment='подмена'),
        )
        self._assert_public_write_forbidden(lambda: ArrivalRosterRowReview.objects.bulk_update(
            [created], ['comment'],
        ))
        self._assert_public_write_forbidden(created.delete)
        self._assert_public_write_forbidden(
            lambda: ArrivalRosterRowReview.objects.filter(pk=created.pk).delete(),
        )

    def test_t12_http_commands_are_post_only_and_ignore_client_actor(self):
        version, _created = self._upload()
        match = version.matches.get()
        client = Client(enforce_csrf_checks=True)
        self._login(client)
        url = reverse('arrival_roster_resident_select', args=[version.pk, match.pk])
        self.assertEqual(client.get(url).status_code, 405)
        self.assertEqual(client.post(url, {
            'expected_revision': 0,
            'resident_id': self.resident.pk,
        }).status_code, 403)
        page = client.get(reverse('arrival_roster_review', args=[version.pk]))
        token = page.cookies['csrftoken'].value
        response = client.post(url, {
            'csrfmiddlewaretoken': token,
            'expected_revision': 0,
            'resident_id': self.resident.pk,
            'actor_access_id': self.other_access.pk,
            'employee_access_id': self.other_access.pk,
        })
        self.assertEqual(response.status_code, 302)
        review = match.row_review
        self.assertEqual(review.updated_by_access, self.access)

    def test_t12_does_not_modify_t11_or_create_business_entities(self):
        version, _created = self._upload()
        match = version.matches.first()
        issue = version.issues.get(code='summary_requires_review')
        before = {
            'employees': Employee.objects.count(),
            'accesses': EmployeeAccess.objects.count(),
            'residents': SettlementResident.objects.count(),
            'cohorts': SettlementCohort.objects.count(),
            'occupancies': EmployeeBedOccupancy.objects.count(),
            'snapshot': version.snapshot_sha256,
            'source_rows': list(version.source_rows.values_list('row_sha256', flat=True)),
        }
        select_arrival_roster_resident(
            match_id=match.pk,
            resident_id=self.resident.pk,
            expected_revision=0,
            actor_access_id=self.access.pk,
        )
        resolve_arrival_roster_issue(
            issue_id=issue.pk,
            expected_revision=0,
            resolution_note='Проверено табельщиком.',
            actor_access_id=self.access.pk,
        )
        version.refresh_from_db()
        self.assertEqual(Employee.objects.count(), before['employees'])
        self.assertEqual(EmployeeAccess.objects.count(), before['accesses'])
        self.assertEqual(SettlementResident.objects.count(), before['residents'])
        self.assertEqual(SettlementCohort.objects.count(), before['cohorts'])
        self.assertEqual(EmployeeBedOccupancy.objects.count(), before['occupancies'])
        self.assertEqual(version.snapshot_sha256, before['snapshot'])
        self.assertEqual(
            list(version.source_rows.values_list('row_sha256', flat=True)),
            before['source_rows'],
        )

    def test_t12_public_event_creation_paths_are_forbidden(self):
        version, _created = self._upload()
        event = ArrivalRosterEvent(
            version=version,
            actor_access=self.access,
            action=ArrivalRosterEvent.Action.REUSED,
            details={'sha256': version.source_file.sha256},
        )
        self._assert_public_write_forbidden(event.save)
        self._assert_public_write_forbidden(lambda: ArrivalRosterEvent.objects.create(
            version=version,
            actor_access=self.access,
            action=ArrivalRosterEvent.Action.REUSED,
            details={'sha256': version.source_file.sha256},
        ))
        self._assert_public_write_forbidden(lambda: ArrivalRosterEvent.objects.get_or_create(
            version=version,
            actor_access=self.access,
            action=ArrivalRosterEvent.Action.REUSED,
            details={'sha256': '0' * 64},
        ))
        self._assert_public_write_forbidden(lambda: ArrivalRosterEvent.objects.update_or_create(
            version=version,
            actor_access=self.access,
            action=ArrivalRosterEvent.Action.REUSED,
            defaults={'details': {'sha256': '1' * 64}},
        ))
        self._assert_public_write_forbidden(
            lambda: ArrivalRosterEvent.objects.bulk_create([event]),
        )

    def test_t12_upload_and_commands_use_only_trusted_event_path(self):
        version, _created = self._upload()
        self.assertEqual(
            list(version.events.values_list('action', flat=True)),
            ['uploaded', 'parsed'],
        )
        match = version.matches.get()
        warning = version.issues.get(code='summary_requires_review')
        with patch(
            'rotations.arrival_rosters._trusted_create_arrival_roster_event',
            wraps=_trusted_create_arrival_roster_event,
        ) as trusted:
            review = select_arrival_roster_resident(
                match_id=match.pk,
                resident_id=self.resident.pk,
                expected_revision=0,
                actor_access_id=self.access.pk,
            )
            review = set_arrival_roster_participation(
                match_id=match.pk,
                participation_status='arriving',
                arrival_mode='transfer',
                expected_revision=review.revision,
                actor_access_id=self.access.pk,
            )
            review = set_arrival_roster_dates(
                match_id=match.pk,
                arrival_on=date(2026, 8, 14),
                departure_on=date(2026, 9, 13),
                expected_revision=review.revision,
                actor_access_id=self.access.pk,
            )
            review = set_arrival_roster_notes(
                match_id=match.pk,
                basis='Реестр заезда',
                comment='Проверено.',
                expected_revision=review.revision,
                actor_access_id=self.access.pk,
            )
            clear_arrival_roster_resident(
                match_id=match.pk,
                expected_revision=review.revision,
                actor_access_id=self.access.pk,
            )
            resolution = resolve_arrival_roster_issue(
                issue_id=warning.pk,
                expected_revision=0,
                resolution_note='Предупреждение проверено вручную.',
                actor_access_id=self.access.pk,
            )
            reopen_arrival_roster_issue(
                issue_id=warning.pk,
                expected_revision=resolution.revision,
                resolution_note='Предупреждение возвращено на проверку.',
                actor_access_id=self.access.pk,
            )
        self.assertEqual(trusted.call_count, 8)
        self.assertTrue(all(
            type(call.kwargs['actor_context']) is _VerifiedTimekeeperContext
            for call in trusted.call_args_list
        ))

    def test_t12_trusted_event_rejects_cross_version_and_personal_details(self):
        first, _created = self._upload()
        second_payload = _with_cell_value(_workbook_bytes(), 'билеты', 'F2', 'Другой маршрут')
        second, _created = self._upload(second_payload, name='другой.xlsx')
        second_match = second.matches.get()
        actor_context = self._verified_context()
        with self.assertRaises(ValidationError) as mismatch:
            _trusted_create_arrival_roster_event(
                version=first,
                actor_context=actor_context,
                match=second_match,
                review_revision=1,
                action=ArrivalRosterEvent.Action.NOTES_CHANGED,
                details={},
            )
        self.assertEqual(mismatch.exception.code, 'arrival_roster.event_version_mismatch')
        with self.assertRaises(ValidationError) as unsafe:
            _trusted_create_arrival_roster_event(
                version=first,
                actor_context=actor_context,
                match=first.matches.get(),
                review_revision=1,
                action=ArrivalRosterEvent.Action.NOTES_CHANGED,
                details={'phone': '+79990000000'},
            )
        self.assertEqual(unsafe.exception.code, 'arrival_roster.unsafe_event_details')

    def test_t12_trusted_event_rejects_plain_access_and_forged_context(self):
        version, _created = self._upload()
        event_count = version.events.count()
        kwargs = {
            'version': version,
            'action': ArrivalRosterEvent.Action.REUSED,
            'details': {'sha256': version.source_file.sha256},
        }
        with self.assertRaises(ValidationError) as plain_access:
            _trusted_create_arrival_roster_event(
                actor_context=self.access,
                **kwargs,
            )
        self.assertEqual(
            plain_access.exception.code,
            'arrival_roster.verified_context_required',
        )
        forged = replace(self._verified_context(), _marker=object())
        with self.assertRaises(ValidationError) as forged_context:
            _trusted_create_arrival_roster_event(
                actor_context=forged,
                **kwargs,
            )
        self.assertEqual(
            forged_context.exception.code,
            'arrival_roster.verified_context_required',
        )
        mismatched = replace(self._verified_context(), access_id=self.other_access.pk)
        with self.assertRaises(ValidationError) as mismatched_context:
            _trusted_create_arrival_roster_event(
                actor_context=mismatched,
                **kwargs,
            )
        self.assertEqual(
            mismatched_context.exception.code,
            'arrival_roster.verified_context_invalid',
        )
        self.assertEqual(version.events.count(), event_count)

    def test_t12_verified_context_rejects_inactive_access_employee_role_and_no_fallback(self):
        event_count = ArrivalRosterEvent.objects.count()

        EmployeeAccess.objects.filter(pk=self.access.pk).update(is_active=False)
        with self.assertRaises(ValidationError) as inactive_access:
            _lock_timekeeper_access(_access_snapshot(self.access.pk))
        self.assertEqual(inactive_access.exception.code, 'arrival_roster.access_denied')
        EmployeeAccess.objects.filter(pk=self.access.pk).update(is_active=True)

        Employee.objects.filter(pk=self.actor.pk).update(is_active=False)
        with self.assertRaises(ValidationError) as inactive_employee:
            _lock_timekeeper_access(_access_snapshot(self.access.pk))
        self.assertEqual(inactive_employee.exception.code, 'arrival_roster.access_denied')
        Employee.objects.filter(pk=self.actor.pk).update(is_active=True)

        Role.objects.filter(pk=self.timekeeper_role.pk).update(is_active=False)
        with self.assertRaises(ValidationError) as inactive_role:
            _lock_timekeeper_access(_access_snapshot(self.access.pk))
        self.assertEqual(inactive_role.exception.code, 'arrival_roster.access_denied')
        Role.objects.filter(pk=self.timekeeper_role.pk).update(is_active=True)

        with self.assertRaises(ValidationError) as wrong_access:
            _lock_timekeeper_access(_access_snapshot(self.other_access.pk))
        self.assertEqual(wrong_access.exception.code, 'arrival_roster.access_denied')
        self.assertEqual(ArrivalRosterEvent.objects.count(), event_count)

    def test_t12_trusted_event_uses_context_without_late_employee_or_access_lock(self):
        version, _created = self._upload()
        actor_context = self._verified_context()
        with (
            patch.object(
                Employee.objects,
                'select_for_update',
                side_effect=AssertionError('late Employee lock'),
            ),
            patch.object(
                EmployeeAccess.objects,
                'select_for_update',
                side_effect=AssertionError('late Access lock'),
            ),
        ):
            event = _trusted_create_arrival_roster_event(
                version=version,
                actor_context=actor_context,
                action=ArrivalRosterEvent.Action.REUSED,
                details={'sha256': version.source_file.sha256},
            )
        self.assertEqual(event.actor_access_id, self.access.pk)

    def test_t12_generic_resolution_rejects_protected_and_unknown_blockers(self):
        version, _created = self._upload()
        protected = {
            'match_unmatched': ArrivalRosterIssue.Severity.ERROR,
            'conflicting_shift_hints': ArrivalRosterIssue.Severity.ERROR,
            'unknown_shift_hint': ArrivalRosterIssue.Severity.WARNING,
            'formula_in_content': ArrivalRosterIssue.Severity.ERROR,
            'missing_sheet': ArrivalRosterIssue.Severity.ERROR,
            'content_outside_profile': ArrivalRosterIssue.Severity.ERROR,
            'future_blocking_code': ArrivalRosterIssue.Severity.ERROR,
        }
        for code, severity in protected.items():
            issue = _trusted_create_arrival_roster_issue(
                version=version,
                severity=severity,
                code=code,
                message='Тестовый блокирующий вопрос.',
                details={},
            )
            with self.assertRaises(ValidationError) as caught:
                resolve_arrival_roster_issue(
                    issue_id=issue.pk,
                    expected_revision=0,
                    resolution_note='Попытка ручного закрытия.',
                    actor_access_id=self.access.pk,
                )
            self.assertEqual(
                caught.exception.code,
                'arrival_roster.blocking_issue_requires_action',
            )
            self.assertFalse(ArrivalRosterIssueResolution._base_manager.filter(issue=issue).exists())
            self.assertFalse(issue.review_events.exists())

    def test_t12_factual_actions_drive_readiness(self):
        version, _created = self._upload(_workbook_bytes(arrival_value='неверная дата'))
        match = version.matches.get()
        self.assertEqual(
            arrival_roster_match_readiness(match_id=match.pk)['code'],
            'timekeeper',
        )
        review = select_arrival_roster_resident(
            match_id=match.pk,
            resident_id=self.resident.pk,
            expected_revision=0,
            actor_access_id=self.access.pk,
        )
        review = set_arrival_roster_participation(
            match_id=match.pk,
            participation_status='arriving',
            arrival_mode='self',
            expected_revision=review.revision,
            actor_access_id=self.access.pk,
        )
        self.assertEqual(
            arrival_roster_match_readiness(match_id=match.pk)['code'],
            'timekeeper',
        )
        review = set_arrival_roster_dates(
            match_id=match.pk,
            arrival_on=date(2026, 8, 14),
            departure_on=date(2026, 9, 13),
            expected_revision=review.revision,
            actor_access_id=self.access.pk,
        )
        readiness = arrival_roster_match_readiness(match_id=match.pk)
        self.assertTrue(readiness['ready'])
        self.assertEqual(readiness['blocking_codes'], [])
        clear_arrival_roster_resident(
            match_id=match.pk,
            expected_revision=review.revision,
            actor_access_id=self.access.pk,
        )
        self.assertFalse(arrival_roster_match_readiness(match_id=match.pk)['ready'])

    def test_t12_structural_error_cannot_be_hidden_by_resolution(self):
        version, _created = self._upload(_workbook_bytes(omit_sheet='Числ'))
        match = version.matches.get()
        issue = version.issues.get(code='missing_sheet')
        ArrivalRosterIssueResolution._base_manager.bulk_create([
            ArrivalRosterIssueResolution(
                issue=issue,
                is_resolved=True,
                resolution_note='Техническая подмена решения.',
                revision=1,
                updated_by_access=self.access,
            ),
        ])
        review = select_arrival_roster_resident(
            match_id=match.pk,
            resident_id=self.resident.pk,
            expected_revision=0,
            actor_access_id=self.access.pk,
        )
        review = set_arrival_roster_participation(
            match_id=match.pk,
            participation_status='arriving',
            arrival_mode='transfer',
            expected_revision=review.revision,
            actor_access_id=self.access.pk,
        )
        set_arrival_roster_dates(
            match_id=match.pk,
            arrival_on=date(2026, 8, 14),
            departure_on=date(2026, 9, 13),
            expected_revision=review.revision,
            actor_access_id=self.access.pk,
        )
        readiness = arrival_roster_match_readiness(match_id=match.pk)
        self.assertEqual(readiness['code'], 'corrected_file')
        self.assertIn('missing_sheet', readiness['blocking_codes'])

    def test_t12_event_failure_rolls_back_domain_change(self):
        version, _created = self._upload()
        match = version.matches.get()
        event_count = version.events.count()
        with patch(
            'rotations.arrival_rosters._trusted_create_arrival_roster_event',
            side_effect=ValidationError('Событие не записано.'),
        ):
            with self.assertRaises(ValidationError):
                select_arrival_roster_resident(
                    match_id=match.pk,
                    resident_id=self.resident.pk,
                    expected_revision=0,
                    actor_access_id=self.access.pk,
                )
        self.assertFalse(ArrivalRosterRowReview._base_manager.filter(match=match).exists())
        self.assertEqual(version.events.count(), event_count)


class ArrivalRosterT13aTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.timekeeper_role = Role.objects.get(code='timekeeper')
        cls.other_role = Role.objects.create(
            code='driver-test-t13a', name='Водитель T1.3a', is_active=True,
        )
        cls.actor = Employee.objects.create(
            full_name='Табельщик T1.3a',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        cls.access = EmployeeAccess.objects.create(
            employee=cls.actor,
            role=cls.timekeeper_role,
            access_code='713001',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        cls.other_access = EmployeeAccess.objects.create(
            employee=cls.actor,
            role=cls.other_role,
            access_code='713002',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        cls.composition = WatchComposition.objects.create(
            code='watch-t13a', name='Вахта T1.3a', is_active=True,
        )
        cls.other_composition = WatchComposition.objects.create(
            code='watch-other-t13a', name='Другая вахта T1.3a', is_active=True,
        )
        cls.period = WatchPeriod.objects.create(
            name='Период T1.3a',
            watch_composition=cls.composition,
            starts_on=date(2026, 8, 14),
            ends_on=date(2026, 9, 13),
            is_active=True,
        )
        cls.employee = Employee.objects.create(
            full_name='Сотрудник Однозначный T1.3a',
            position='Водитель',
            status=Employee.Status.ACTIVE,
            is_active=True,
            hired_at=date(2025, 1, 1),
            watch_composition=cls.composition,
        )
        cls.resident = SettlementResident(
            employee=cls.employee,
            resident_type=SettlementResident.ResidentType.EMPLOYEE,
            status=SettlementResident.Status.ACTIVE,
        )
        cls.resident.save()
        cls.other_employee = Employee.objects.create(
            full_name='Сотрудник Другой Вахты T1.3a',
            position='Механик',
            status=Employee.Status.ACTIVE,
            is_active=True,
            hired_at=date(2025, 2, 1),
            watch_composition=cls.other_composition,
        )
        cls.other_resident = SettlementResident(
            employee=cls.other_employee,
            resident_type=SettlementResident.ResidentType.EMPLOYEE,
            status=SettlementResident.Status.ACTIVE,
        )
        cls.other_resident.save()
        cls.external = SettlementResident(
            resident_type=SettlementResident.ResidentType.CONTRACTOR,
            full_name='Внешний Жилец T1.3a',
            position_title='Специалист',
            organization='Подрядчик T1.3a',
            phone='+79990001122',
            external_sex='male',
            status=SettlementResident.Status.ACTIVE,
            created_by_access=cls.access,
        )
        cls.external.save()

    def _create_pool(self, *, access=None, period=None):
        return create_arrival_roster_from_employee_pool(
            watch_period_id=(period or self.period).pk,
            actor_access_id=(access or self.access).pk,
        )

    def _lock_trace(self, operation):
        trace = []
        original = QuerySet.select_for_update

        def traced(queryset, *args, **kwargs):
            trace.append((queryset.model.__name__, kwargs.get('of')))
            return original(queryset, *args, **kwargs)

        with patch.object(QuerySet, 'select_for_update', new=traced):
            operation()
        return trace

    def _assert_lock_prefix(self, trace):
        models = [model for model, _of in trace]
        expected = [
            'Employee',
            'EmployeeAccess',
            'WatchPeriod',
            'ArrivalRosterVersion',
            'SettlementResident',
        ]
        positions = [models.index(model) for model in expected]
        self.assertEqual(positions, sorted(positions))
        self.assertNotIn('Role', models)
        access_position = models.index('EmployeeAccess')
        self.assertNotIn('Employee', models[access_position + 1:])
        for model, of in trace[:access_position + 1]:
            if model in {'Employee', 'EmployeeAccess'}:
                self.assertEqual(of, ('self',))

    def test_t13a_automatic_pool_uses_exact_composition_and_existing_resident(self):
        version = self._create_pool()

        self.assertEqual(version.source_kind, ArrivalRosterVersion.SourceKind.EMPLOYEE_POOL)
        self.assertIsNone(version.source_file_id)
        self.assertIsNone(version.parser_profile_id)
        self.assertEqual(version.created_by_access_id, self.access.pk)
        self.assertRegex(version.source_fingerprint, r'^[0-9a-f]{64}$')
        row = version.pool_rows.get()
        self.assertEqual(row.employee_id, self.employee.pk)
        self.assertEqual(row.resident_id, self.resident.pk)
        self.assertEqual(row.watch_composition_id, self.composition.pk)
        self.assertEqual(row.origin_kind, ArrivalRosterPoolRow.OriginKind.AUTOMATIC_EMPLOYEE)
        self.assertEqual(row.suggested_participation, 'arriving')
        self.assertNotIn('phone', row.employee_snapshot)
        self.assertNotIn('phone', row.resident_snapshot)
        self.assertFalse(version.matches.filter(matched_resident=self.other_resident).exists())
        review = row.match.row_review
        self.assertEqual(review.selected_resident_id, self.resident.pk)
        self.assertIsNone(review.participation_status)
        self.assertTrue(version.events.filter(action='pool_created').exists())

        self.assertEqual(row.employee_snapshot, _employee_snapshot(self.employee))
        self.assertEqual(row.resident_snapshot, _resident_snapshot(self.resident))
        trusted_parameters = inspect.signature(_trusted_create_pool_row).parameters
        self.assertNotIn('employee_snapshot', trusted_parameters)
        self.assertNotIn('resident_snapshot', trusted_parameters)

    def test_t13a_personnel_dates_and_missing_resident_are_blocking_without_wrapper(self):
        late_employee = Employee.objects.create(
            full_name='Новый После Начала T1.3a',
            status=Employee.Status.ACTIVE,
            is_active=True,
            hired_at=date(2026, 8, 20),
            watch_composition=self.composition,
        )
        missing_employee = Employee.objects.create(
            full_name='Без Карточки Жильца T1.3a',
            status=Employee.Status.ACTIVE,
            is_active=True,
            hired_at=date(2025, 1, 1),
            watch_composition=self.composition,
        )
        resident_count = SettlementResident.objects.count()

        version = self._create_pool()

        self.assertEqual(SettlementResident.objects.count(), resident_count)
        self.assertTrue(version.issues.filter(
            code='employee_hired_after_period_start',
            details__employee_id=late_employee.pk,
        ).exists())
        self.assertTrue(version.issues.filter(
            code='employee_resident_missing',
            details__employee_id=missing_employee.pk,
        ).exists())
        missing_row = version.pool_rows.get(employee=missing_employee)
        self.assertIsNone(missing_row.resident_id)
        self.assertEqual(missing_row.employee_snapshot['employee_id'], missing_employee.pk)
        self.assertRegex(missing_row.snapshot_sha256, r'^[0-9a-f]{64}$')
        unresolved = version.matches.get(evidence__employee_id=missing_employee.pk)
        self.assertEqual(missing_row.match_id, unresolved.pk)
        self.assertEqual(unresolved.status, ArrivalRosterMatch.Status.UNMATCHED)
        self.assertIsNone(unresolved.row_review.selected_resident_id)
        self.assertEqual(version.status, ArrivalRosterVersion.Status.REVIEW_REQUIRED)

    def test_t13a_missing_resident_can_be_selected_later_only_for_exact_employee(self):
        employee = Employee.objects.create(
            full_name='Сотрудник с поздней карточкой T1.3a',
            status=Employee.Status.ACTIVE,
            is_active=True,
            hired_at=date(2025, 1, 1),
            watch_composition=self.composition,
        )
        version = self._create_pool()
        row = version.pool_rows.get(employee=employee)
        self.assertIsNone(row.resident_id)

        with self.assertRaises(ValidationError) as foreign:
            select_arrival_roster_resident(
                match_id=row.match_id,
                resident_id=self.other_resident.pk,
                expected_revision=1,
                actor_access_id=self.access.pk,
            )
        self.assertEqual(
            foreign.exception.code,
            'arrival_roster.pool_employee_resident_mismatch',
        )

        resident = SettlementResident(
            employee=employee,
            resident_type=SettlementResident.ResidentType.EMPLOYEE,
            status=SettlementResident.Status.ACTIVE,
        )
        resident.save()
        review = select_arrival_roster_resident(
            match_id=row.match_id,
            resident_id=resident.pk,
            expected_revision=1,
            actor_access_id=self.access.pk,
        )
        self.assertEqual(review.selected_resident_id, resident.pk)
        row.refresh_from_db()
        self.assertIsNone(row.resident_id)
        self.assertEqual(row.employee_id, employee.pk)

    def test_t13a_inactive_archived_and_dismissed_subjects_are_not_silently_accepted(self):
        dismissed = Employee.objects.create(
            full_name='Уволенный До Периода T1.3a',
            status=Employee.Status.DISMISSED,
            is_active=False,
            hired_at=date(2024, 1, 1),
            dismissed_at=date(2026, 8, 1),
            watch_composition=self.composition,
        )
        dismissed_resident = SettlementResident(
            employee=dismissed,
            resident_type=SettlementResident.ResidentType.EMPLOYEE,
            status=SettlementResident.Status.ACTIVE,
        )
        dismissed_resident.save()
        archived_employee = Employee.objects.create(
            full_name='Архивный Жилец T1.3a',
            status=Employee.Status.ACTIVE,
            is_active=True,
            hired_at=date(2024, 1, 1),
            watch_composition=self.composition,
        )
        archived_resident = SettlementResident(
            employee=archived_employee,
            resident_type=SettlementResident.ResidentType.EMPLOYEE,
            status=SettlementResident.Status.ARCHIVED,
            archived_at=timezone.now(),
        )
        archived_resident.save()

        version = self._create_pool()

        dismissed_codes = set(version.issues.filter(
            details__employee_id=dismissed.pk,
        ).values_list('code', flat=True))
        self.assertIn('employee_inactive', dismissed_codes)
        self.assertIn('employee_dismissed_before_period', dismissed_codes)
        self.assertTrue(version.issues.filter(
            code='employee_resident_unavailable',
            details__employee_id=archived_employee.pk,
        ).exists())
        archived_row = version.pool_rows.get(employee=archived_employee)
        self.assertIsNone(archived_row.resident_id)

    def test_t13a_requires_active_period_composition(self):
        missing = WatchPeriod.objects.create(
            name='Без состава T1.3a',
            starts_on=date(2026, 10, 1),
            ends_on=date(2026, 10, 31),
            is_active=True,
        )
        with self.assertRaises(ValidationError) as absent:
            self._create_pool(period=missing)
        self.assertEqual(absent.exception.code, 'arrival_roster.watch_composition_required')
        WatchComposition.objects.filter(pk=self.composition.pk).update(is_active=False)
        with self.assertRaises(ValidationError) as inactive:
            self._create_pool()
        self.assertEqual(inactive.exception.code, 'arrival_roster.watch_composition_required')

    def test_t13a_repeat_creates_new_version_with_deterministic_fingerprint(self):
        first = self._create_pool()
        second = self._create_pool()

        self.assertNotEqual(first.pk, second.pk)
        self.assertEqual(second.version_number, first.version_number + 1)
        self.assertEqual(first.source_fingerprint, second.source_fingerprint)
        self.assertRegex(first.pool_rows.get().snapshot_sha256, r'^[0-9a-f]{64}$')
        self.assertRegex(second.pool_rows.get().snapshot_sha256, r'^[0-9a-f]{64}$')

    def test_t13a_manual_employee_can_come_from_entire_database(self):
        version = self._create_pool()
        row = add_employee_to_arrival_roster(
            version_id=version.pk,
            employee_id=self.other_employee.pk,
            basis='Дополнительный сотрудник для заезда.',
            actor_access_id=self.access.pk,
        )

        self.assertEqual(row.origin_kind, ArrivalRosterPoolRow.OriginKind.MANUAL_EMPLOYEE)
        self.assertEqual(row.employee_id, self.other_employee.pk)
        self.assertEqual(row.watch_composition_id, self.other_composition.pk)
        self.assertTrue(version.issues.filter(
            match=row.match,
            code='employee_watch_composition_mismatch',
        ).exists())
        self.assertTrue(version.events.filter(action='pool_employee_added').exists())

    def test_t13a_manual_employee_without_composition_remains_visible_with_oup_issue(self):
        employee = Employee.objects.create(
            full_name='Без Вахты T1.3a',
            status=Employee.Status.ACTIVE,
            is_active=True,
            hired_at=date(2025, 1, 1),
        )
        resident = SettlementResident(
            employee=employee,
            resident_type=SettlementResident.ResidentType.EMPLOYEE,
            status=SettlementResident.Status.ACTIVE,
        )
        resident.save()
        version = self._create_pool()

        row = add_employee_to_arrival_roster(
            version_id=version.pk,
            employee_id=employee.pk,
            basis='Добавлен табельщиком.',
            actor_access_id=self.access.pk,
        )

        self.assertIsNone(row.watch_composition_id)
        self.assertTrue(version.issues.filter(
            match=row.match,
            code='employee_watch_composition_missing',
        ).exists())

    def test_t13a_manual_external_uses_existing_card_without_business_side_effects(self):
        version = self._create_pool()
        before = {
            'employees': Employee.objects.count(),
            'accesses': EmployeeAccess.objects.count(),
            'residents': SettlementResident.objects.count(),
            'cohorts': SettlementCohort.objects.count(),
            'occupancies': EmployeeBedOccupancy.objects.count(),
        }

        row = add_external_resident_to_arrival_roster(
            version_id=version.pk,
            resident_id=self.external.pk,
            basis='Подтверждённая заявка подрядчика.',
            actor_access_id=self.access.pk,
        )

        self.assertEqual(row.origin_kind, ArrivalRosterPoolRow.OriginKind.MANUAL_EXTERNAL)
        self.assertIsNone(row.employee_id)
        self.assertIsNone(row.watch_composition_id)
        self.assertEqual(row.created_by_access_id, self.access.pk)
        self.assertEqual(before, {
            'employees': Employee.objects.count(),
            'accesses': EmployeeAccess.objects.count(),
            'residents': SettlementResident.objects.count(),
            'cohorts': SettlementCohort.objects.count(),
            'occupancies': EmployeeBedOccupancy.objects.count(),
        })
        self.assertTrue(version.events.filter(action='pool_external_added').exists())

    def test_t13a_duplicates_and_missing_basis_are_controlled(self):
        version = self._create_pool()
        with self.assertRaises(ValidationError) as duplicate:
            add_employee_to_arrival_roster(
                version_id=version.pk,
                employee_id=self.employee.pk,
                basis='Повтор.',
                actor_access_id=self.access.pk,
            )
        self.assertEqual(duplicate.exception.code, 'arrival_roster.pool_duplicate')
        with self.assertRaises(ValidationError) as no_basis:
            add_external_resident_to_arrival_roster(
                version_id=version.pk,
                resident_id=self.external.pk,
                basis=' ',
                actor_access_id=self.access.pk,
            )
        self.assertEqual(no_basis.exception.code, 'arrival_roster.pool_basis_required')

    def test_t13a_exact_timekeeper_access_has_no_fallback(self):
        with self.assertRaises(ValidationError) as wrong_role:
            self._create_pool(access=self.other_access)
        self.assertEqual(wrong_role.exception.code, 'arrival_roster.access_denied')
        EmployeeAccess.objects.filter(pk=self.access.pk).update(is_active=False)
        replacement = EmployeeAccess.objects.create(
            employee=self.actor,
            role=self.timekeeper_role,
            access_code='713003',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        with self.assertRaises(ValidationError) as inactive:
            self._create_pool(access=self.access)
        self.assertEqual(inactive.exception.code, 'arrival_roster.access_denied')
        self.assertNotEqual(replacement.pk, self.access.pk)
        self.assertEqual(ArrivalRosterVersion.objects.count(), 0)

    def test_t13a_automatic_pool_lock_order_has_no_late_employee_lock(self):
        trace = self._lock_trace(self._create_pool)
        self._assert_lock_prefix(trace)

    def test_t13a_manual_employee_lock_order_has_no_late_employee_lock(self):
        version = self._create_pool()
        trace = self._lock_trace(lambda: add_employee_to_arrival_roster(
            version_id=version.pk,
            employee_id=self.other_employee.pk,
            basis='Проверка порядка блокировок.',
            actor_access_id=self.access.pk,
        ))
        self._assert_lock_prefix(trace)

    def test_t13a_external_lock_order_has_no_late_employee_lock(self):
        version = self._create_pool()
        trace = self._lock_trace(lambda: add_external_resident_to_arrival_roster(
            version_id=version.pk,
            resident_id=self.external.pk,
            basis='Проверка порядка блокировок.',
            actor_access_id=self.access.pk,
        ))
        self._assert_lock_prefix(trace)

    def test_t13a_pool_rows_are_append_only_and_public_orm_writes_are_closed(self):
        version = self._create_pool()
        row = version.pool_rows.get()
        clone = ArrivalRosterPoolRow(
            version=version,
            resident=self.other_resident,
            employee=self.other_employee,
            watch_composition=self.other_composition,
            match=row.match,
            origin_kind=ArrivalRosterPoolRow.OriginKind.MANUAL_EMPLOYEE,
            suggested_participation='additional',
            employee_snapshot={},
            resident_snapshot={},
            snapshot_sha256='0' * 64,
            created_by_access=self.access,
            basis='Подмена.',
        )
        operations = [
            clone.save,
            lambda: ArrivalRosterPoolRow.objects.create(
                version=version,
                resident=self.other_resident,
                employee=self.other_employee,
                watch_composition=self.other_composition,
                match=row.match,
                origin_kind='manual_employee',
                suggested_participation='additional',
                employee_snapshot={},
                resident_snapshot={},
                snapshot_sha256='0' * 64,
                created_by_access=self.access,
                basis='Подмена.',
            ),
            lambda: ArrivalRosterPoolRow.objects.bulk_create([clone]),
            lambda: ArrivalRosterPoolRow.objects.filter(pk=row.pk).update(basis='Подмена.'),
            lambda: ArrivalRosterPoolRow.objects.bulk_update([row], ['basis']),
            row.delete,
            lambda: ArrivalRosterPoolRow.objects.filter(pk=row.pk).delete(),
        ]
        for operation in operations:
            with self.assertRaises(ValidationError) as caught:
                operation()
            self.assertEqual(
                caught.exception.code,
                'rotations.arrival_roster.public_write_forbidden',
            )
        row.refresh_from_db()
        self.assertEqual(row.basis, '')

        original_created_at = row.created_at
        row.created_at = original_created_at + timedelta(seconds=1)
        with self.assertRaises(ValidationError) as caught:
            row.save()
        self.assertIn('created_at', caught.exception.message_dict)
        row.refresh_from_db()
        self.assertEqual(row.created_at, original_created_at)


class ArrivalRosterT14aApprovalTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        role = Role.objects.get(code='timekeeper')
        composition = WatchComposition.objects.create(code='watch-t14a', name='Вахта T1.4a')
        cls.actor = Employee.objects.create(
            full_name='Табельщик T1.4a', status=Employee.Status.ACTIVE,
            is_active=True, hired_at=date(2025, 1, 1),
        )
        cls.access = EmployeeAccess.objects.create(
            employee=cls.actor, role=role, access_code='t14a',
            status=EmployeeAccess.Status.ACTIVATED, is_active=True,
        )
        cls.employee = Employee.objects.create(
            full_name='Участник T1.4a', status=Employee.Status.ACTIVE,
            is_active=True, hired_at=date(2025, 1, 1), watch_composition=composition,
        )
        cls.resident = SettlementResident.objects.create(
            employee=cls.employee, resident_type=SettlementResident.ResidentType.EMPLOYEE,
            status=SettlementResident.Status.ACTIVE,
        )
        cls.period = WatchPeriod.objects.create(
            name='Период T1.4a', watch_composition=composition,
            starts_on=date(2026, 8, 14), ends_on=date(2026, 9, 13), is_active=True,
        )
        cls.other_role = Role.objects.create(code='t14a-other', name='Другая роль T1.4a')
        cls.other_access = EmployeeAccess.objects.create(
            employee=cls.actor, role=cls.other_role, access_code='t14a-other',
            status=EmployeeAccess.Status.ACTIVATED, is_active=True,
        )

    def _ready_version(self):
        version = create_arrival_roster_from_employee_pool(
            watch_period_id=self.period.pk, actor_access_id=self.access.pk,
        )
        for match in version.matches.order_by('pk'):
            review = match.row_review
            if review.selected_resident_id is None:
                continue
            set_arrival_roster_participation(
                match_id=match.pk, participation_status='arriving', arrival_mode='transfer',
                expected_revision=review.revision, actor_access_id=self.access.pk,
            )
            review.refresh_from_db()
            set_arrival_roster_dates(
                match_id=match.pk, arrival_on=self.period.starts_on,
                departure_on=self.period.ends_on, expected_revision=review.revision,
                actor_access_id=self.access.pk,
            )
        return version

    def _login(self, client=None, access=None):
        client = client or self.client
        session = client.session
        session['employee_access_id'] = (access or self.access).pk
        session.save()
        return client

    def _approval_token(self, client, version):
        response = client.get(reverse('arrival_roster_review', args=[version.pk]))
        self.assertEqual(response.status_code, 200)
        return response, response.context['approval_form']['expected_sha256'].value()

    def _ready_excel(self, *, shift_hint=1):
        uploaded = SimpleUploadedFile(
            't14a.xlsx',
            _workbook_bytes(
                full_name=self.employee.full_name,
                primary_name=self.employee.full_name,
                phone='', shift_hint=shift_hint,
            ),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        version, created = upload_arrival_roster(
            uploaded_file=uploaded, watch_period_id=self.period.pk,
            actor_access_id=self.access.pk,
        )
        self.assertTrue(created)
        for index, match in enumerate(version.matches.order_by('pk')):
            try:
                review = match.row_review
            except ArrivalRosterRowReview.DoesNotExist:
                review = None
            if review is None:
                resident = match.matched_resident
                if resident is None:
                    resident = SettlementResident.objects.create(
                        resident_type=SettlementResident.ResidentType.CONTRACTOR,
                        full_name=f'Внешний T1.4a {version.pk}-{match.pk}',
                        position_title='Специалист', organization='Подрядчик',
                        phone='+79990000000', external_sex='male',
                        created_by_access=self.access,
                        status=SettlementResident.Status.ACTIVE,
                    )
                review = select_arrival_roster_resident(
                    match_id=match.pk, resident_id=resident.pk,
                    expected_revision=0, actor_access_id=self.access.pk,
                )
            if review.resident_resolution != ArrivalRosterRowReview.ResidentResolution.SELECTED:
                resident = SettlementResident.objects.create(
                    resident_type=SettlementResident.ResidentType.CONTRACTOR,
                    full_name=f'Внешний T1.4a {version.pk}-{match.pk}',
                    position_title='Специалист', organization='Подрядчик',
                    phone='+79990000000', external_sex='male',
                    created_by_access=self.access,
                    status=SettlementResident.Status.ACTIVE,
                )
                select_arrival_roster_resident(
                    match_id=match.pk, resident_id=resident.pk,
                    expected_revision=review.revision,
                    actor_access_id=self.access.pk,
                )
                review.refresh_from_db()
            participation = 'arriving' if index == 0 else 'not_arriving'
            set_arrival_roster_participation(
                match_id=match.pk, participation_status=participation,
                arrival_mode='transfer' if index == 0 else '',
                expected_revision=review.revision, actor_access_id=self.access.pk,
            )
            review.refresh_from_db()
            if index == 0:
                set_arrival_roster_dates(
                    match_id=match.pk, arrival_on=self.period.starts_on,
                    departure_on=self.period.ends_on, expected_revision=review.revision,
                    actor_access_id=self.access.pk,
                )
        return version

    def _proposal(self, version):
        return build_arrival_roster_confirmation_proposal(
            version_id=version.pk, actor_access_id=self.access.pk,
        )

    def _confirm(self, version):
        proposal = self._proposal(version)
        return confirm_arrival_roster_version(
            version_id=version.pk, expected_sha256=proposal['confirmation_sha256'],
            actor_access_id=self.access.pk,
        )

    def _force_legacy_watch_composition_drift_for_test(self, *, composition_id):
        # Имитирует historical/pre-guard либо внешнее повреждение baseline в БД.
        with connection.cursor() as cursor:
            cursor.execute(
                'UPDATE users_employee SET watch_composition_id = %s WHERE id = %s',
                [composition_id, self.employee.pk],
            )
        self.employee.refresh_from_db()

    def test_t14a_confirmation_is_sha_guarded_idempotent_and_contains_no_pii(self):
        version = self._ready_version()
        proposal = build_arrival_roster_confirmation_proposal(
            version_id=version.pk, actor_access_id=self.access.pk,
        )
        with self.assertRaises(ValidationError) as caught:
            confirm_arrival_roster_version(
                version_id=version.pk, expected_sha256='0' * 64,
                actor_access_id=self.access.pk,
            )
        self.assertEqual(caught.exception.code, 'arrival_roster.stale_confirmation_sha256')
        confirmed = confirm_arrival_roster_version(
            version_id=version.pk, expected_sha256=proposal['confirmation_sha256'],
            actor_access_id=self.access.pk,
        )
        first_time = confirmed.confirmed_at
        first_events = confirmed.events.filter(action=ArrivalRosterEvent.Action.CONFIRMED).count()
        repeated = confirm_arrival_roster_version(
            version_id=version.pk, expected_sha256=proposal['confirmation_sha256'],
            actor_access_id=self.access.pk,
        )
        self.assertEqual(repeated.confirmed_at, first_time)
        self.assertEqual(repeated.events.filter(action=ArrivalRosterEvent.Action.CONFIRMED).count(), first_events)
        serialized = json.dumps(repeated.confirmation_snapshot, ensure_ascii=False)
        self.assertNotIn(self.employee.full_name, serialized)
        self.assertEqual(repeated.status, ArrivalRosterVersion.Status.CONFIRMED)

    def test_t14a_version_public_orm_writes_are_closed(self):
        version = self._ready_version()
        operations = [
            lambda: version.save(),
            lambda: ArrivalRosterVersion.objects.filter(pk=version.pk).update(status='confirmed'),
            lambda: ArrivalRosterVersion.objects.bulk_create([version]),
            lambda: ArrivalRosterVersion.objects.bulk_update([version], ['status']),
            version.delete,
        ]
        for operation in operations:
            with self.assertRaises(ValidationError) as caught:
                operation()
            self.assertEqual(caught.exception.code, 'rotations.arrival_roster.public_write_forbidden')

    def test_t14a_excel_and_pool_happy_paths_and_shift_hints(self):
        pool = self._ready_version()
        self.assertEqual(self._confirm(pool).source_kind, ArrivalRosterVersion.SourceKind.EMPLOYEE_POOL)

        second_period = WatchPeriod.objects.create(
            name='Период Excel T1.4a', watch_composition=self.period.watch_composition,
            starts_on=self.period.starts_on, ends_on=self.period.ends_on,
        )
        original_period = self.period
        self.period = second_period
        try:
            excel = self._ready_excel(shift_hint=2)
            confirmed = self._confirm(excel)
        finally:
            self.period = original_period
        self.assertEqual(confirmed.source_kind, ArrivalRosterVersion.SourceKind.EXCEL)
        serialized = json.dumps(confirmed.confirmation_snapshot, ensure_ascii=False)
        self.assertNotIn('official_shift', serialized)
        self.assertFalse(hasattr(confirmed, 'official_shift'))

    def test_t14a_readiness_rejects_empty_resident_participation_and_dates(self):
        empty_composition = WatchComposition.objects.create(
            code='empty-t14a', name='Пустая вахта T1.4a',
        )
        empty_period = WatchPeriod.objects.create(
            name='Пустой период T1.4a', watch_composition=empty_composition,
            starts_on=self.period.starts_on, ends_on=self.period.ends_on,
        )
        empty = create_arrival_roster_from_employee_pool(
            watch_period_id=empty_period.pk, actor_access_id=self.access.pk,
        )
        with self.assertRaises(ValidationError) as empty_error:
            self._proposal(empty)
        self.assertEqual(empty_error.exception.code, 'arrival_roster.empty_or_incomplete')
        cases = [
            ({'resident_resolution': 'unreviewed', 'selected_resident_id': None}, 'arrival_roster.resident_required'),
            ({'participation_status': None}, 'arrival_roster.participation_required'),
            ({'arrival_on': None, 'departure_on': None}, 'arrival_roster.dates_required'),
        ]
        for changes, code in cases:
            version = self._ready_version()
            target = version.row_reviews.get()
            ArrivalRosterRowReview._base_manager.filter(pk=target.pk).update(**changes)
            with self.assertRaises(ValidationError) as caught:
                self._proposal(version)
            self.assertEqual(caught.exception.code, code)

    def test_t14a_stale_employee_fields_resident_and_source_snapshot(self):
        mutations = [
            {'status': Employee.Status.DISMISSED},
            {'is_active': False},
            {'hired_at': date(2026, 8, 15)},
            {'dismissed_at': date(2026, 8, 1)},
            {'watch_composition_id': None},
        ]
        for mutation in mutations:
            version = self._ready_version()
            proposal = self._proposal(version)
            before = Employee.objects.filter(pk=self.employee.pk).values().get()
            if 'watch_composition_id' in mutation:
                self._force_legacy_watch_composition_drift_for_test(
                    composition_id=mutation['watch_composition_id'],
                )
            else:
                Employee.objects.filter(pk=self.employee.pk).update(**mutation)
            with self.assertRaises(ValidationError) as caught:
                confirm_arrival_roster_version(
                    version_id=version.pk, expected_sha256=proposal['confirmation_sha256'],
                    actor_access_id=self.access.pk,
                )
            self.assertEqual(caught.exception.code, 'arrival_roster.stale_confirmation_sha256')
            Employee.objects.filter(pk=self.employee.pk).update(
                status=before['status'], is_active=before['is_active'],
                hired_at=before['hired_at'], dismissed_at=before['dismissed_at'],
                updated_at=before['updated_at'],
            )
            if 'watch_composition_id' in mutation:
                self._force_legacy_watch_composition_drift_for_test(
                    composition_id=before['watch_composition_id'],
                )
            else:
                self.employee.refresh_from_db()

        version = self._ready_version()
        proposal = self._proposal(version)
        SettlementResident._base_manager.filter(pk=self.resident.pk).update(revision=models.F('revision') + 1)
        with self.assertRaises(ValidationError) as resident_stale:
            confirm_arrival_roster_version(
                version_id=version.pk, expected_sha256=proposal['confirmation_sha256'],
                actor_access_id=self.access.pk,
            )
        self.assertEqual(resident_stale.exception.code, 'arrival_roster.stale_confirmation_sha256')

        version = self._ready_version()
        proposal = self._proposal(version)
        models.QuerySet.update(
            ArrivalRosterVersion._base_manager.filter(pk=version.pk), snapshot={'tampered': True},
        )
        with self.assertRaises(ValidationError) as source_stale:
            confirm_arrival_roster_version(
                version_id=version.pk, expected_sha256=proposal['confirmation_sha256'],
                actor_access_id=self.access.pk,
            )
        self.assertEqual(source_stale.exception.code, 'arrival_roster.stale_confirmation_sha256')

    def test_t14a_inactive_and_wrong_resident_duplicates_and_invalid_interval(self):
        version = self._ready_version()
        proposal = self._proposal(version)
        SettlementResident._base_manager.filter(pk=self.resident.pk).update(
            status=SettlementResident.Status.ARCHIVED, archived_at=timezone.now(),
        )
        with self.assertRaises(ValidationError) as inactive:
            confirm_arrival_roster_version(
                version_id=version.pk, expected_sha256=proposal['confirmation_sha256'],
                actor_access_id=self.access.pk,
            )
        self.assertEqual(inactive.exception.code, 'arrival_roster.stale_confirmation_sha256')
        SettlementResident._base_manager.filter(pk=self.resident.pk).update(
            status=SettlementResident.Status.ACTIVE, archived_at=None,
        )

        version = self._ready_version()
        proposal = self._proposal(version)
        other_employee = Employee.objects.create(
            full_name='Другой Employee T1.4a', status=Employee.Status.ACTIVE,
            is_active=True, hired_at=date(2025, 1, 1),
            watch_composition=self.period.watch_composition,
        )
        SettlementResident._base_manager.filter(pk=self.resident.pk).update(
            employee_id=other_employee.pk, revision=models.F('revision') + 1,
        )
        with self.assertRaises(ValidationError) as wrong:
            confirm_arrival_roster_version(
                version_id=version.pk, expected_sha256=proposal['confirmation_sha256'],
                actor_access_id=self.access.pk,
            )
        self.assertEqual(wrong.exception.code, 'arrival_roster.stale_confirmation_sha256')
        SettlementResident._base_manager.filter(pk=self.resident.pk).update(
            employee_id=self.employee.pk, revision=models.F('revision') + 1,
        )

        version = self._ready_version()
        match = version.matches.filter(pool_row__employee=self.employee).get()
        with self.assertRaises(ValidationError) as dates:
            set_arrival_roster_dates(
                match_id=match.pk, arrival_on=self.period.ends_on,
                departure_on=self.period.starts_on,
                expected_revision=match.row_review.revision,
                actor_access_id=self.access.pk,
            )
        self.assertEqual(dates.exception.code, 'arrival_roster.invalid_dates')

        second_employee = Employee.objects.create(
            full_name='Дубль T1.4a', status=Employee.Status.ACTIVE,
            is_active=True, hired_at=date(2025, 1, 1),
            watch_composition=self.period.watch_composition,
        )
        second_resident = SettlementResident.objects.create(
            employee=second_employee,
            resident_type=SettlementResident.ResidentType.EMPLOYEE,
            status=SettlementResident.Status.ACTIVE,
        )
        second_version = self._ready_version()
        first_match = second_version.matches.filter(pool_row__employee=self.employee).get()
        second_match = second_version.matches.filter(pool_row__employee=second_employee).get()
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                ArrivalRosterRowReview._base_manager.filter(pk=second_match.row_review.pk).update(
                    selected_resident_id=first_match.row_review.selected_resident_id,
                )

    def test_t14a_question_policy_is_closed_and_unknown_error_cannot_be_resolved(self):
        expected = {
            'conflicting_shift_hints': ('deputy', False),
            'employee_watch_composition_mismatch': ('oup', False),
            'employee_resident_missing': ('clerk', True),
            'participation_requires_review': ('timekeeper', False),
            'equipment_missing': ('deputy', False),
            'official_shift_missing': ('deputy', False),
            'external_card_incomplete': ('clerk', False),
        }
        for code, (role, blocks) in expected.items():
            self.assertEqual(arrival_roster_issue_policy(code, 'error'), {
                'role': role, 'blocks_confirmation': blocks,
            })
        self.assertTrue(arrival_roster_issue_policy('future_unknown', 'error')['blocks_confirmation'])

        version = self._ready_version()
        match = version.matches.get()
        context = _lock_timekeeper_access(_access_snapshot(self.access.pk))
        issue = _trusted_create_arrival_roster_issue(
            version=version, match=match, severity='error', code='future_unknown',
            message='Будущая неизвестная ошибка.', details={},
        )
        ArrivalRosterIssueResolution._base_manager.bulk_create([
            ArrivalRosterIssueResolution(
                issue=issue, is_resolved=True, resolution_note='Ручная отметка.',
                revision=1, updated_by_access=self.access,
            )
        ])
        with self.assertRaises(ValidationError) as caught:
            self._proposal(version)
        self.assertEqual(caught.exception.code, 'arrival_roster.blocking_issue')

    def test_t14a_allowed_next_role_questions_and_missing_operational_fields_do_not_block(self):
        version = self._ready_version()
        match = version.matches.get()
        for code in [
            'employee_watch_composition_mismatch', 'external_card_incomplete',
            'conflicting_shift_hints', 'equipment_missing', 'official_shift_missing',
        ]:
            _trusted_create_arrival_roster_issue(
                version=version, match=match, severity='error', code=code,
                message=f'Проверка политики {code}.', details={},
            )
        proposal = self._proposal(version)
        roles = {
            row['code']: row['responsible_role']
            for row in proposal['confirmation_snapshot']['issues']
        }
        self.assertEqual(roles['employee_watch_composition_mismatch'], 'oup')
        self.assertEqual(roles['external_card_incomplete'], 'clerk')
        self.assertEqual(roles['conflicting_shift_hints'], 'deputy')
        self.assertEqual(self._confirm(version).status, ArrivalRosterVersion.Status.CONFIRMED)

    def test_t14a_exact_access_no_fallback_and_lock_order(self):
        version = self._ready_version()
        with self.assertRaises(ValidationError) as wrong_role:
            build_arrival_roster_confirmation_proposal(
                version_id=version.pk, actor_access_id=self.other_access.pk,
            )
        self.assertEqual(wrong_role.exception.code, 'arrival_roster.access_denied')

        fallback = EmployeeAccess.objects.create(
            employee=self.actor, role=self.access.role, access_code='fallback-t14a',
            status=EmployeeAccess.Status.ACTIVATED, is_active=True,
        )
        EmployeeAccess.objects.filter(pk=self.access.pk).update(is_active=False)
        with self.assertRaises(ValidationError) as inactive:
            build_arrival_roster_confirmation_proposal(
                version_id=version.pk, actor_access_id=self.access.pk,
            )
        self.assertEqual(inactive.exception.code, 'arrival_roster.access_denied')
        self.assertTrue(fallback.is_active)
        EmployeeAccess.objects.filter(pk=self.access.pk).update(is_active=True)

        Employee.objects.filter(pk=self.actor.pk).update(is_active=False)
        with self.assertRaises(ValidationError) as inactive_employee:
            build_arrival_roster_confirmation_proposal(
                version_id=version.pk, actor_access_id=self.access.pk,
            )
        self.assertEqual(inactive_employee.exception.code, 'arrival_roster.access_denied')
        Employee.objects.filter(pk=self.actor.pk).update(is_active=True)

        trace = []
        original = QuerySet.select_for_update
        def traced(queryset, *args, **kwargs):
            trace.append(queryset.model.__name__)
            return original(queryset, *args, **kwargs)
        with patch.object(QuerySet, 'select_for_update', new=traced):
            self._proposal(version)
        expected = ['Employee', 'EmployeeAccess', 'WatchPeriod', 'ArrivalRosterVersion',
                    'SettlementResident', 'ArrivalRosterMatch', 'ArrivalRosterRowReview',
                    'ArrivalRosterIssue', 'ArrivalRosterIssueResolution']
        positions = [trace.index(name) for name in expected]
        self.assertEqual(positions, sorted(positions))
        access_position = trace.index('EmployeeAccess')
        self.assertNotIn('Employee', trace[access_position + 1:])
        self.assertNotIn('Role', trace)

    def test_t14a_confirmed_and_superseded_metadata_are_immutable(self):
        confirmed = self._confirm(self._ready_version())
        before = (confirmed.confirmed_by_access_id, confirmed.confirmed_at,
                  confirmed.confirmation_snapshot, confirmed.confirmation_sha256)
        with self.assertRaises(ValidationError):
            _trusted_confirm_arrival_roster_version(
                version_id=confirmed.pk, actor_access_id=self.access.pk,
                confirmation_snapshot={'rewrite': True}, confirmation_sha256='1' * 64,
                confirmed_at=timezone.now(),
            )
        superseded = _trusted_supersede_arrival_roster_version(
            version_id=confirmed.pk, superseded_at=timezone.now(),
        )
        self.assertEqual(
            (superseded.confirmed_by_access_id, superseded.confirmed_at,
             superseded.confirmation_snapshot, superseded.confirmation_sha256), before,
        )
        with self.assertRaises(ValidationError):
            _trusted_supersede_arrival_roster_version(
                version_id=superseded.pk, superseded_at=timezone.now(),
            )

    def test_t14a_replacement_lineage_and_atomic_rollback(self):
        first = self._confirm(self._ready_version())
        replacement = self._ready_version()
        proposal = self._proposal(replacement)
        with self.assertRaises(ValidationError) as missing_lineage:
            confirm_arrival_roster_version(
                version_id=replacement.pk, expected_sha256=proposal['confirmation_sha256'],
                actor_access_id=self.access.pk,
            )
        self.assertEqual(missing_lineage.exception.code, 'arrival_roster.replacement_lineage_required')
        models.QuerySet.update(
            ArrivalRosterVersion._base_manager.filter(pk=replacement.pk),
            based_on_version_id=first.pk,
        )
        proposal = self._proposal(replacement)
        with patch(
            'rotations.arrival_roster_approvals._trusted_create_arrival_roster_event',
            side_effect=RuntimeError('event failure'),
        ):
            with self.assertRaises(RuntimeError):
                confirm_arrival_roster_version(
                    version_id=replacement.pk, expected_sha256=proposal['confirmation_sha256'],
                    actor_access_id=self.access.pk,
                )
        first.refresh_from_db(); replacement.refresh_from_db()
        self.assertEqual(first.status, ArrivalRosterVersion.Status.CONFIRMED)
        self.assertNotEqual(replacement.status, ArrivalRosterVersion.Status.CONFIRMED)
        self.assertFalse(replacement.events.filter(action=ArrivalRosterEvent.Action.CONFIRMED).exists())

        confirmed_replacement = confirm_arrival_roster_version(
            version_id=replacement.pk, expected_sha256=proposal['confirmation_sha256'],
            actor_access_id=self.access.pk,
        )
        first.refresh_from_db()
        self.assertEqual(first.status, ArrivalRosterVersion.Status.SUPERSEDED)
        self.assertEqual(confirmed_replacement.status, ArrivalRosterVersion.Status.CONFIRMED)

    def test_t14a_wrong_period_lineage_and_second_confirmed_are_rejected(self):
        first = self._confirm(self._ready_version())
        other_composition = WatchComposition.objects.create(code='other-base-t14a', name='Другая база T1.4a')
        other_period = WatchPeriod.objects.create(
            name='Другой период T1.4a', watch_composition=other_composition,
            starts_on=self.period.starts_on, ends_on=self.period.ends_on,
        )
        other_version = create_arrival_roster_from_employee_pool(
            watch_period_id=other_period.pk, actor_access_id=self.access.pk,
        )
        replacement = self._ready_version()
        models.QuerySet.update(
            ArrivalRosterVersion._base_manager.filter(pk=replacement.pk),
            based_on_version_id=other_version.pk,
        )
        with self.assertRaises(ValidationError) as wrong_period:
            self._proposal(replacement)
        self.assertEqual(wrong_period.exception.code, 'arrival_roster.replacement_base_invalid')

        clean_second = self._ready_version()
        proposal = self._proposal(clean_second)
        with self.assertRaises(ValidationError) as one_confirmed:
            confirm_arrival_roster_version(
                version_id=clean_second.pk,
                expected_sha256=proposal['confirmation_sha256'],
                actor_access_id=self.access.pk,
            )
        self.assertEqual(one_confirmed.exception.code, 'arrival_roster.replacement_lineage_required')
        first.refresh_from_db()
        self.assertEqual(first.status, ArrivalRosterVersion.Status.CONFIRMED)

    def test_t14a_pool_trusted_snapshot_ignores_caller_snapshots(self):
        plans = [{
            'employee': self.employee,
            'resident': self.resident,
            'employee_snapshot': {'forged_employee': 'must-not-be-stored'},
            'resident_snapshot': {'forged_resident': 'must-not-be-stored'},
            'issues': [],
        }]
        version = _trusted_create_pool_arrival_roster_version(
            period=self.period,
            version_number=997,
            actor_access=self.access,
            plans=plans,
        )
        version = _trusted_finalize_pool_arrival_roster_version(
            version=version,
            period=self.period,
            plans=plans,
            pool_rows=[],
        )
        actual_source = {
            'watch_period_id': self.period.pk,
            'watch_composition_id': self.period.watch_composition_id,
            'employees': [{
                'employee': _employee_snapshot(
                    Employee._base_manager.get(pk=self.employee.pk)
                ),
                'resident': _resident_snapshot(
                    SettlementResident._base_manager.get(pk=self.resident.pk)
                ),
                'issue_codes': [],
            }],
        }
        expected_fingerprint = hashlib.sha256(
            json.dumps(
                actual_source,
                ensure_ascii=False,
                sort_keys=True,
                separators=(',', ':'),
            ).encode('utf-8')
        ).hexdigest()
        version.refresh_from_db()
        self.assertEqual(version.snapshot['source'], actual_source)
        self.assertEqual(version.source_fingerprint, expected_fingerprint)
        self.assertEqual(version.snapshot['source_fingerprint'], expected_fingerprint)
        serialized = json.dumps(version.snapshot, ensure_ascii=False)
        self.assertNotIn('forged_employee', serialized)
        self.assertNotIn('forged_resident', serialized)
        self.assertNotIn('must-not-be-stored', serialized)

    def test_t14a_database_enforces_one_confirmed_version_per_period(self):
        first = self._confirm(self._ready_version())
        second = self._ready_version()
        collision = None
        try:
            with transaction.atomic():
                models.QuerySet.update(
                    ArrivalRosterVersion._base_manager.filter(pk=second.pk),
                    status=ArrivalRosterVersion.Status.CONFIRMED,
                    confirmed_by_access_id=self.access.pk,
                    confirmed_at=timezone.now(),
                    confirmation_snapshot={'test': 'second-confirmed'},
                    confirmation_sha256='2' * 64,
                )
        except IntegrityError as error:
            collision = error
        self.assertIsNotNone(collision)
        self.assertTrue(_is_confirmed_period_collision(collision))
        self.assertEqual(
            ArrivalRosterVersion._base_manager.filter(
                watch_period=self.period,
                status=ArrivalRosterVersion.Status.CONFIRMED,
            ).count(),
            1,
        )
        first.refresh_from_db()
        second.refresh_from_db()
        self.assertEqual(first.status, ArrivalRosterVersion.Status.CONFIRMED)
        self.assertNotEqual(second.status, ArrivalRosterVersion.Status.CONFIRMED)

    def test_t14b_ready_button_and_unready_blockers_are_server_rendered(self):
        client = self._login()
        ready_response, _token = self._approval_token(client, self._ready_version())
        self.assertContains(ready_response, 'Утвердить список заезда')
        self.assertContains(ready_response, 'Версия готова к утверждению')

        unready = create_arrival_roster_from_employee_pool(
            watch_period_id=self.period.pk,
            actor_access_id=self.access.pk,
        )
        blocked = client.get(reverse('arrival_roster_review', args=[unready.pk]))
        self.assertNotContains(blocked, 'Утвердить список заезда')
        self.assertContains(blocked, 'Утверждение пока недоступно')
        self.assertTrue(blocked.context['approval_error'])
        self.assertContains(blocked, blocked.context['approval_error'])
        self.assertContains(blocked, 'Требуют решения табельщика')

    def test_t14b_approval_is_post_csrf_session_access_and_payload_safe(self):
        version = self._ready_version()
        csrf_client = self._login(Client(enforce_csrf_checks=True))
        url = reverse('arrival_roster_approval_confirm', args=[version.pk])
        self.assertEqual(csrf_client.get(url).status_code, 405)
        self.assertEqual(csrf_client.post(url, {}).status_code, 403)

        client = self._login()
        _response, token = self._approval_token(client, version)
        confirmed = client.post(url, {
            'expected_sha256': token,
            'actor_access_id': self.other_access.pk,
            'employee_access_id': self.other_access.pk,
            'employee_id': 999999,
            'role_id': 999999,
            'confirmed_at': '2000-01-01T00:00:00Z',
            'session_id': 'forged',
        })
        self.assertEqual(confirmed.status_code, 302)
        version.refresh_from_db()
        self.assertEqual(version.status, ArrivalRosterVersion.Status.CONFIRMED)
        self.assertEqual(version.confirmed_by_access_id, self.access.pk)

    def test_t14b_wrong_or_inactive_session_access_is_rejected(self):
        version = self._ready_version()
        url = reverse('arrival_roster_approval_confirm', args=[version.pk])
        other_client = self._login(Client(), self.other_access)
        self.assertEqual(other_client.post(url, {'expected_sha256': '0' * 64}).status_code, 302)
        version.refresh_from_db()
        self.assertNotEqual(version.status, ArrivalRosterVersion.Status.CONFIRMED)

        EmployeeAccess.objects.filter(pk=self.access.pk).update(is_active=False)
        inactive_client = self._login(Client(), self.access)
        self.assertIn(
            inactive_client.post(url, {'expected_sha256': '0' * 64}).status_code,
            {302, 409},
        )
        version.refresh_from_db()
        self.assertNotEqual(version.status, ArrivalRosterVersion.Status.CONFIRMED)

    def test_t14b_stale_success_repeat_and_read_only_html(self):
        Employee.objects.filter(pk=self.employee.pk).update(phone='+79990001122')
        version = self._ready_version()
        client = self._login()
        url = reverse('arrival_roster_approval_confirm', args=[version.pk])
        stale = client.post(url, {'expected_sha256': '0' * 64}, follow=True)
        self.assertContains(stale, 'Предложение утверждения устарело.')
        version.refresh_from_db()
        self.assertNotEqual(version.status, ArrivalRosterVersion.Status.CONFIRMED)

        _response, token = self._approval_token(client, version)
        successful = client.post(url, {'expected_sha256': token}, follow=True)
        self.assertContains(successful, 'Список заезда утверждён.')
        version.refresh_from_db()
        confirmed_at = version.confirmed_at
        event_count = version.events.filter(action=ArrivalRosterEvent.Action.CONFIRMED).count()

        repeated = client.post(url, {'expected_sha256': token}, follow=True)
        self.assertContains(repeated, 'Список уже утверждён.')
        version.refresh_from_db()
        self.assertEqual(version.confirmed_at, confirmed_at)
        self.assertEqual(version.events.filter(action=ArrivalRosterEvent.Action.CONFIRMED).count(), event_count)

        page = client.get(reverse('arrival_roster_review', args=[version.pk]))
        html = page.content.decode('utf-8')
        self.assertContains(page, 'Список заезда утверждён')
        self.assertContains(page, 'Табельщик:')
        self.assertNotContains(page, 'Утвердить список заезда')
        self.assertNotIn('Сохранить участие', html)
        self.assertNotIn('+79990001122', html)
        self.assertNotIn(version.confirmation_sha256, html)
        self.assertNotIn('confirmation_snapshot', html)
        self.assertNotIn('employee_access_id', html)
        self.assertNotIn('PIN', html)

    def test_t14b_index_uses_russian_states_and_current_confirmed_marker(self):
        first = self._confirm(self._ready_version())
        replacement = self._ready_version()
        models.QuerySet.update(
            ArrivalRosterVersion._base_manager.filter(pk=replacement.pk),
            based_on_version_id=first.pk,
        )
        confirmed = self._confirm(replacement)
        draft = create_arrival_roster_from_employee_pool(
            watch_period_id=self.period.pk,
            actor_access_id=self.access.pk,
        )
        review_required = create_arrival_roster_from_employee_pool(
            watch_period_id=self.period.pk,
            actor_access_id=self.access.pk,
        )
        models.QuerySet.update(
            ArrivalRosterVersion._base_manager.filter(pk=review_required.pk),
            status=ArrivalRosterVersion.Status.REVIEW_REQUIRED,
        )
        response = self._login().get(reverse('arrival_roster_index'))
        self.assertContains(response, 'Утверждена · действующая версия')
        self.assertContains(response, 'Заменена новой версией')
        self.assertContains(response, 'Подготовка')
        self.assertContains(response, 'Требуется проверка')
        self.assertContains(response, 'Из базы сотрудников')
        self.assertEqual(confirmed.status, ArrivalRosterVersion.Status.CONFIRMED)
        self.assertEqual(draft.status, ArrivalRosterVersion.Status.DRAFT)

    def test_t14c_pool_revision_is_independent_and_reuses_effective_decisions(self):
        parent = self._confirm(self._ready_version())
        parent_pool = list(parent.pool_rows.order_by('pk'))
        parent_matches = list(parent.matches.order_by('pk'))
        parent_reviews = list(parent.row_reviews.order_by('match_id'))
        child = create_arrival_roster_correction_revision(
            version_id=parent.pk, actor_access_id=self.access.pk,
        )
        child.refresh_from_db()
        self.assertEqual(child.watch_period_id, parent.watch_period_id)
        self.assertEqual(child.based_on_version_id, parent.pk)
        self.assertEqual(child.status, ArrivalRosterVersion.Status.REVIEW_REQUIRED)
        self.assertFalse(child.confirmed_by_access_id)
        self.assertFalse(child.confirmed_at)
        self.assertEqual(child.confirmation_snapshot, {})
        self.assertEqual(child.confirmation_sha256, '')
        self.assertEqual(child.version_number, parent.version_number + 1)
        self.assertEqual(child.pool_rows.count(), len(parent_pool))
        self.assertEqual(child.matches.count(), len(parent_matches))
        self.assertEqual(child.row_reviews.count(), len(parent_reviews))
        self.assertFalse(set(parent.pool_rows.values_list('pk', flat=True)) & set(child.pool_rows.values_list('pk', flat=True)))
        self.assertFalse(set(parent.matches.values_list('pk', flat=True)) & set(child.matches.values_list('pk', flat=True)))
        self.assertTrue(all(review.revision == 1 for review in child.row_reviews.all()))
        self.assertTrue(all(review.updated_by_access_id == self.access.pk for review in child.row_reviews.all()))
        self.assertEqual(parent.events.filter(action='revision_created').count(), 0)
        self.assertEqual(child.events.filter(action='revision_created').count(), 1)
        parent.refresh_from_db()
        self.assertEqual(parent.status, ArrivalRosterVersion.Status.CONFIRMED)

    def test_t14c_excel_revision_copies_rows_links_issues_and_resolutions(self):
        editable = self._ready_excel()
        issue = _trusted_create_arrival_roster_issue(
            version=editable,
            match=editable.matches.order_by('pk').first(),
            severity=ArrivalRosterIssue.Severity.WARNING,
            code='external_card_incomplete',
            message='Требуется уточнение карточки.',
        )
        resolve_arrival_roster_issue(
            issue_id=issue.pk, expected_revision=0,
            resolution_note='Карточка проверена табельщиком.', actor_access_id=self.access.pk,
        )
        parent = self._confirm(editable)
        parent_source_ids = set(parent.source_rows.values_list('pk', flat=True))
        parent_normalized_ids = set(
            ArrivalRosterNormalizedRow.objects.filter(source_row__version=parent).values_list('pk', flat=True)
        )
        child = create_arrival_roster_correction_revision(
            version_id=parent.pk, actor_access_id=self.access.pk,
        )
        child.refresh_from_db()
        self.assertEqual(child.source_kind, ArrivalRosterVersion.SourceKind.EXCEL)
        self.assertEqual(child.source_file_id, parent.source_file_id)
        self.assertEqual(child.parser_profile_id, parent.parser_profile_id)
        self.assertEqual(child.source_rows.count(), parent.source_rows.count())
        child_normalized = ArrivalRosterNormalizedRow.objects.filter(source_row__version=child)
        self.assertEqual(child_normalized.count(), len(parent_normalized_ids))
        self.assertFalse(parent_source_ids & set(child.source_rows.values_list('pk', flat=True)))
        self.assertFalse(parent_normalized_ids & set(child_normalized.values_list('pk', flat=True)))
        self.assertTrue(all(link.normalized_row.source_row.version_id == child.pk for link in ArrivalRosterMatchRow.objects.filter(match__version=child).select_related('normalized_row__source_row')))
        self.assertEqual(child.issues.count(), parent.issues.count())
        copied_resolution = child.issues.get(code='external_card_incomplete').resolution
        self.assertTrue(copied_resolution.is_resolved)
        self.assertEqual(copied_resolution.resolution_note, 'Карточка проверена табельщиком.')
        self.assertEqual(copied_resolution.revision, 1)
        self.assertEqual(copied_resolution.updated_by_access_id, self.access.pk)
        self.assertEqual(child.events.count(), 1)
        self.assertEqual(child.events.get().action, 'revision_created')

    def test_t14c_http_is_post_csrf_session_bound_and_idempotent(self):
        parent = self._confirm(self._ready_version())
        self.employee.phone = '+79995550123'
        self.employee.save(update_fields=['phone'])
        self.employee.refresh_from_db()
        self.assertTrue(self.employee.phone)
        url = reverse('arrival_roster_create_revision', args=[parent.pk])
        csrf_client = self._login(Client(enforce_csrf_checks=True))
        self.assertEqual(csrf_client.get(url).status_code, 405)
        self.assertEqual(csrf_client.post(url, {}).status_code, 403)
        other_client = self._login(Client(), self.other_access)
        self.assertEqual(other_client.post(url).status_code, 302)
        self.assertFalse(ArrivalRosterVersion.objects.filter(based_on_version=parent).exists())
        client = self._login()
        response = client.post(url, {
            'actor_access_id': self.other_access.pk,
            'employee_id': 999999,
            'role_id': 999999,
            'watch_period_id': 999999,
            'based_on_version': 999999,
            'created_at': '2000-01-01T00:00:00Z',
        })
        self.assertEqual(response.status_code, 302)
        child = ArrivalRosterVersion.objects.get(based_on_version=parent)
        self.assertEqual(child.created_by_access_id, self.access.pk)
        repeated = client.post(url)
        self.assertEqual(repeated.status_code, 302)
        self.assertEqual(ArrivalRosterVersion.objects.filter(based_on_version=parent).count(), 1)
        page = client.get(reverse('arrival_roster_review', args=[parent.pk]))
        self.assertContains(page, 'Перейти к версии для исправления')
        child_page = client.get(reverse('arrival_roster_review', args=[child.pk]))
        self.assertContains(child_page, f'Создана на основании утверждённой версии № {parent.version_number}')
        self.assertNotContains(child_page, parent.confirmation_sha256)
        self.assertNotContains(child_page, 'employee_access_id')
        self.assertNotContains(child_page, self.employee.phone)
        self.assertNotContains(child_page, 'snapshot_sha256')

    def test_t14c_rejects_invalid_parent_conflicts_and_rolls_back(self):
        draft = self._ready_version()
        with self.assertRaises(ValidationError) as invalid:
            create_arrival_roster_correction_revision(version_id=draft.pk, actor_access_id=self.access.pk)
        self.assertEqual(invalid.exception.code, 'arrival_roster.revision_parent_not_confirmed')
        parent = self._confirm(self._ready_version())
        before = ArrivalRosterVersion.objects.count()
        with patch(
            'rotations.arrival_roster_approvals._trusted_create_arrival_roster_event',
            side_effect=RuntimeError('event failure'),
        ):
            with self.assertRaises(RuntimeError):
                create_arrival_roster_correction_revision(version_id=parent.pk, actor_access_id=self.access.pk)
        self.assertEqual(ArrivalRosterVersion.objects.count(), before)
        first = create_arrival_roster_correction_revision(version_id=parent.pk, actor_access_id=self.access.pk)
        duplicate = ArrivalRosterVersion._base_manager.get(pk=first.pk)
        models.QuerySet.update(
            ArrivalRosterVersion._base_manager.filter(pk=duplicate.pk),
            version_number=duplicate.version_number + 100,
        )
        # Test-only base manager bypass proves the controlled ambiguity branch.
        clone = ArrivalRosterVersion._base_manager.get(pk=first.pk)
        clone.pk = None
        clone.id = None
        clone.version_number = first.version_number + 101
        clone.snapshot = {}
        clone.snapshot_sha256 = ''
        clone.created_at = None
        clone.updated_at = None
        ArrivalRosterVersion._base_manager.bulk_create([clone])
        with self.assertRaises(ValidationError) as conflict:
            create_arrival_roster_correction_revision(version_id=parent.pk, actor_access_id=self.access.pk)
        self.assertEqual(conflict.exception.code, 'arrival_roster.revision_child_conflict')

    def test_t14a_controlled_unique_collision_and_unknown_integrity_error(self):
        version = self._ready_version()
        proposal = self._proposal(version)
        collision = IntegrityError(
            'UNIQUE constraint failed: rotations_arrivalrosterversion.watch_period_id'
        )
        with patch(
            'rotations.arrival_roster_approvals._confirm_arrival_roster_version_once',
            side_effect=collision,
        ):
            with self.assertRaises(ValidationError) as caught:
                confirm_arrival_roster_version(
                    version_id=version.pk, expected_sha256=proposal['confirmation_sha256'],
                    actor_access_id=self.access.pk,
                )
        self.assertEqual(caught.exception.code, 'arrival_roster.confirmed_period_conflict')
        with patch(
            'rotations.arrival_roster_approvals._confirm_arrival_roster_version_once',
            side_effect=IntegrityError('unknown constraint'),
        ):
            with self.assertRaises(IntegrityError):
                confirm_arrival_roster_version(
                    version_id=version.pk, expected_sha256=proposal['confirmation_sha256'],
                    actor_access_id=self.access.pk,
                )

    def test_t14a_public_create_shortcuts_queryset_delete_and_db_shape_are_closed(self):
        version = self._ready_version()
        kwargs = {
            'watch_period': self.period, 'version_number': 999,
            'status': 'draft', 'source_kind': 'employee_pool',
            'created_by_access': self.access, 'source_fingerprint': '1' * 64,
        }
        operations = [
            lambda: ArrivalRosterVersion.objects.create(**kwargs),
            lambda: ArrivalRosterVersion.objects.get_or_create(version_number=998, defaults=kwargs),
            lambda: ArrivalRosterVersion.objects.update_or_create(version_number=997, defaults=kwargs),
            lambda: ArrivalRosterVersion.objects.filter(pk=version.pk).delete(),
        ]
        for operation in operations:
            with self.assertRaises(ValidationError) as caught:
                operation()
            self.assertEqual(caught.exception.code, 'rotations.arrival_roster.public_write_forbidden')
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                models.QuerySet.update(
                    ArrivalRosterVersion._base_manager.filter(pk=version.pk),
                    status='confirmed', confirmed_by_access_id=self.access.pk,
                    confirmed_at=timezone.now(), confirmation_snapshot={},
                    confirmation_sha256='1' * 64,
                )

    def test_t14a_snapshot_is_deterministic_complete_safe_and_has_no_side_effects(self):
        version = self._ready_version()
        counts = {
            'employee': Employee.objects.count(),
            'access': EmployeeAccess.objects.count(),
            'resident': SettlementResident.objects.count(),
            'cohort': SettlementCohort.objects.count(),
            'occupancy': EmployeeBedOccupancy.objects.count(),
            'assignment': apps.get_model('assignments', 'EquipmentAssignment').objects.count(),
        }
        first = self._proposal(version)
        second = self._proposal(version)
        self.assertEqual(first, second)
        row = first['confirmation_snapshot']['rows'][0]
        self.assertEqual(set(row['employee']), {
            'employee_id', 'status', 'is_active', 'hired_at', 'dismissed_at',
            'watch_composition_id', 'updated_at',
        })
        serialized = json.dumps(first['confirmation_snapshot'], ensure_ascii=False).lower()
        for secret in [self.employee.full_name, 'phone', 'pin', 'password', 'passport']:
            self.assertNotIn(secret.lower(), serialized)
        self.assertEqual(counts, {
            'employee': Employee.objects.count(),
            'access': EmployeeAccess.objects.count(),
            'resident': SettlementResident.objects.count(),
            'cohort': SettlementCohort.objects.count(),
            'occupancy': EmployeeBedOccupancy.objects.count(),
            'assignment': apps.get_model('assignments', 'EquipmentAssignment').objects.count(),
        })


class ArrivalRosterT14aMigrationTests(TransactionTestCase):
    reset_sequences = True

    def _migrate(self, target):
        executor = MigrationExecutor(connection)
        executor.migrate([('rotations', target)])
        return executor.loader.project_state([('rotations', target)]).apps

    def tearDown(self):
        self._migrate('0006_arrival_roster_excel_revision')
        super().tearDown()

    def _historical_version(self, apps, *, confirmed=False):
        RoleModel = apps.get_model('users', 'Role')
        EmployeeModel = apps.get_model('users', 'Employee')
        AccessModel = apps.get_model('users', 'EmployeeAccess')
        CompositionModel = apps.get_model('users', 'WatchComposition')
        PeriodModel = apps.get_model('shifts', 'WatchPeriod')
        VersionModel = apps.get_model('rotations', 'ArrivalRosterVersion')
        role, _created = RoleModel.objects.get_or_create(
            code='timekeeper', defaults={'name': 'Табельщик', 'is_active': True},
        )
        employee = EmployeeModel.objects.create(
            full_name='Migration T1.4a', status='active', is_active=True,
        )
        access = AccessModel.objects.create(
            employee_id=employee.pk, role_id=role.pk, access_code='migration-t14a',
            status='activated', is_active=True,
        )
        composition = CompositionModel.objects.create(
            code='migration-t14a', name='Migration T1.4a',
        )
        period = PeriodModel.objects.create(
            name='Migration T1.4a', watch_composition_id=composition.pk,
            starts_on=date(2026, 8, 1), ends_on=date(2026, 9, 1), is_active=True,
        )
        kwargs = {
            'watch_period_id': period.pk, 'version_number': 1,
            'status': 'confirmed' if confirmed else 'draft',
            'source_kind': 'employee_pool', 'created_by_access_id': access.pk,
            'source_fingerprint': '1' * 64,
            'snapshot': {'source': 'migration'}, 'snapshot_sha256': '2' * 64,
        }
        if confirmed:
            kwargs.update(
                confirmed_by_access_id=access.pk, confirmed_at=timezone.now(),
                confirmation_snapshot={'schema': 1}, confirmation_sha256='3' * 64,
            )
        return VersionModel.objects.create(**kwargs)

    def test_t14a_migration_cycle_and_safe_reverse_preserve_draft_version(self):
        apps_0004 = self._migrate('0004_arrival_roster_employee_pool')
        version = self._historical_version(apps_0004)
        apps_0005 = self._migrate('0005_arrival_roster_confirmation')
        Version0005 = apps_0005.get_model('rotations', 'ArrivalRosterVersion')
        migrated = Version0005.objects.get(pk=version.pk)
        self.assertEqual(migrated.status, 'draft')
        self.assertEqual(migrated.confirmation_snapshot, {})
        apps_reversed = self._migrate('0004_arrival_roster_employee_pool')
        self.assertTrue(
            apps_reversed.get_model('rotations', 'ArrivalRosterVersion').objects.filter(pk=version.pk).exists()
        )
        apps_forward = self._migrate('0005_arrival_roster_confirmation')
        self.assertTrue(
            apps_forward.get_model('rotations', 'ArrivalRosterVersion').objects.filter(pk=version.pk).exists()
        )

    def test_t14a_migration_reverse_fails_closed_with_count_and_pk(self):
        apps_0005 = self._migrate('0005_arrival_roster_confirmation')
        version = self._historical_version(apps_0005, confirmed=True)
        with self.assertRaisesRegex(
            RuntimeError, rf'count=1; PK=\[{version.pk}\]',
        ):
            self._migrate('0004_arrival_roster_employee_pool')


class ArrivalRosterT14cMigrationTests(TransactionTestCase):
    reset_sequences = True

    def _migrate(self, target):
        executor = MigrationExecutor(connection)
        executor.migrate([('rotations', target)])
        return executor.loader.project_state([('rotations', target)]).apps

    def tearDown(self):
        self._migrate('0006_arrival_roster_excel_revision')
        super().tearDown()

    def _excel_version(self, apps, *, version_number=1, based_on_version_id=None):
        RoleModel = apps.get_model('users', 'Role')
        EmployeeModel = apps.get_model('users', 'Employee')
        AccessModel = apps.get_model('users', 'EmployeeAccess')
        CompositionModel = apps.get_model('users', 'WatchComposition')
        PeriodModel = apps.get_model('shifts', 'WatchPeriod')
        SourceModel = apps.get_model('rotations', 'ArrivalRosterSourceFile')
        ProfileModel = apps.get_model('rotations', 'ArrivalRosterParserProfile')
        VersionModel = apps.get_model('rotations', 'ArrivalRosterVersion')
        role, _ = RoleModel.objects.get_or_create(code='timekeeper', defaults={'name': 'Табельщик'})
        employee, _ = EmployeeModel.objects.get_or_create(
            full_name='Migration T1.4c', defaults={'status': 'active', 'is_active': True},
        )
        access, _ = AccessModel.objects.get_or_create(
            employee_id=employee.pk, role_id=role.pk, access_code='migration-t14c',
            defaults={'status': 'activated', 'is_active': True},
        )
        composition, _ = CompositionModel.objects.get_or_create(code='migration-t14c', defaults={'name': 'Migration T1.4c'})
        period, _ = PeriodModel.objects.get_or_create(
            name='Migration T1.4c',
            defaults={
                'watch_composition_id': composition.pk,
                'starts_on': date(2026, 8, 1), 'ends_on': date(2026, 9, 1), 'is_active': True,
            },
        )
        source, _ = SourceModel.objects.get_or_create(
            sha256='a' * 64,
            defaults={
                'original_name': 'migration.xlsx', 'byte_size': 1,
                'content_type': 'application/octet-stream', 'file': 'migration.xlsx',
                'uploaded_by_access_id': access.pk,
            },
        )
        profile, _ = ProfileModel.objects.get_or_create(
            code='migration-t14c', version=1,
            defaults={'configuration': {}, 'configuration_sha256': 'b' * 64},
        )
        return VersionModel.objects.create(
            watch_period_id=period.pk, version_number=version_number,
            status='draft', source_kind='excel', source_file_id=source.pk,
            parser_profile_id=profile.pk, created_by_access_id=access.pk,
            source_fingerprint='c' * 64, snapshot={}, snapshot_sha256='',
            based_on_version_id=based_on_version_id,
        )

    def test_t14c_migration_primary_unique_revision_lineage_and_safe_cycle(self):
        apps_0005 = self._migrate('0005_arrival_roster_confirmation')
        primary = self._excel_version(apps_0005)
        apps_0006 = self._migrate('0006_arrival_roster_excel_revision')
        Version = apps_0006.get_model('rotations', 'ArrivalRosterVersion')
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                self._excel_version(apps_0006, version_number=2)
        revision = self._excel_version(apps_0006, version_number=2, based_on_version_id=primary.pk)
        grandchild = self._excel_version(apps_0006, version_number=3, based_on_version_id=revision.pk)
        self.assertEqual(Version.objects.filter(source_kind='excel').count(), 3)
        self.assertEqual(grandchild.based_on_version_id, revision.pk)
        with self.assertRaisesRegex(RuntimeError, rf'count=3; PK=\[{primary.pk}, {revision.pk}, {grandchild.pk}\]'):
            self._migrate('0005_arrival_roster_confirmation')

    def test_t14c_migration_reverse_without_revisions_and_employee_pool_is_unchanged(self):
        apps_0005 = self._migrate('0005_arrival_roster_confirmation')
        version = self._excel_version(apps_0005)
        apps_0006 = self._migrate('0006_arrival_roster_excel_revision')
        Version = apps_0006.get_model('rotations', 'ArrivalRosterVersion')
        pool = Version.objects.create(
            watch_period_id=version.watch_period_id, version_number=2,
            status='draft', source_kind='employee_pool', source_file_id=None,
            parser_profile_id=None, created_by_access_id=version.created_by_access_id,
            source_fingerprint='d' * 64, snapshot={}, snapshot_sha256='',
        )
        self.assertTrue(Version.objects.filter(pk=version.pk).exists())
        apps_reversed = self._migrate('0005_arrival_roster_confirmation')
        self.assertTrue(apps_reversed.get_model('rotations', 'ArrivalRosterVersion').objects.filter(pk=version.pk).exists())
        apps_forward = self._migrate('0006_arrival_roster_excel_revision')
        self.assertTrue(apps_forward.get_model('rotations', 'ArrivalRosterVersion').objects.filter(pk=pool.pk).exists())


class ArrivalRosterT13bTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.timekeeper_role = Role.objects.get(code='timekeeper')
        cls.other_role = Role.objects.create(
            code='driver-test-t13b', name='Водитель T1.3b', is_active=True,
        )
        cls.actor = Employee.objects.create(
            full_name='Табельщик T1.3b',
            status=Employee.Status.ACTIVE,
            is_active=True,
        )
        cls.access = EmployeeAccess.objects.create(
            employee=cls.actor,
            role=cls.timekeeper_role,
            access_code='713101',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        cls.other_access = EmployeeAccess.objects.create(
            employee=cls.actor,
            role=cls.other_role,
            access_code='713102',
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        )
        cls.composition = WatchComposition.objects.create(
            code='watch-t13b', name='Вахта T1.3b', is_active=True,
        )
        cls.other_composition = WatchComposition.objects.create(
            code='watch-other-t13b', name='Другая вахта T1.3b', is_active=True,
        )
        cls.period = WatchPeriod.objects.create(
            name='Период T1.3b',
            watch_composition=cls.composition,
            starts_on=date(2026, 8, 14),
            ends_on=date(2026, 9, 13),
            is_active=True,
        )
        cls.employee = Employee.objects.create(
            full_name='Сотрудник Основной T1.3b',
            position='Водитель',
            phone='+79991110001',
            status=Employee.Status.ACTIVE,
            is_active=True,
            hired_at=date(2025, 1, 1),
            watch_composition=cls.composition,
        )
        cls.resident = SettlementResident(
            employee=cls.employee,
            resident_type=SettlementResident.ResidentType.EMPLOYEE,
            status=SettlementResident.Status.ACTIVE,
        )
        cls.resident.save()
        cls.other_employee = Employee.objects.create(
            full_name='Сотрудник Другой Вахты T1.3b',
            position='Механик',
            phone='+79992220002',
            status=Employee.Status.ACTIVE,
            is_active=True,
            hired_at=date(2025, 2, 1),
            watch_composition=cls.other_composition,
        )
        cls.other_resident = SettlementResident(
            employee=cls.other_employee,
            resident_type=SettlementResident.ResidentType.EMPLOYEE,
            status=SettlementResident.Status.ACTIVE,
        )
        cls.other_resident.save()
        cls.dismissed = Employee.objects.create(
            full_name='Уволенный Сотрудник T1.3b',
            position='Слесарь',
            phone='+79993330003',
            status=Employee.Status.DISMISSED,
            is_active=False,
            hired_at=date(2024, 1, 1),
            dismissed_at=date(2026, 8, 1),
            watch_composition=cls.other_composition,
        )
        cls.external = SettlementResident(
            resident_type=SettlementResident.ResidentType.CONTRACTOR,
            full_name='Внешний Специалист T1.3b',
            position_title='Наладчик',
            organization='Подрядчик T1.3b',
            phone='+79994440004',
            external_sex='male',
            status=SettlementResident.Status.ACTIVE,
            created_by_access=cls.access,
        )
        cls.external.save()

    def _login(self, client=None, access=None):
        client = client or self.client
        session = client.session
        session['employee_access_id'] = (access or self.access).pk
        session.save()
        return client

    def _pool(self):
        return create_arrival_roster_from_employee_pool(
            watch_period_id=self.period.pk,
            actor_access_id=self.access.pk,
        )

    def _create_pool(self):
        return self._pool()

    def _assert_public_write_forbidden(self, operation):
        with self.assertRaises(ValidationError) as caught:
            operation()
        self.assertEqual(
            caught.exception.code,
            'rotations.arrival_roster.public_write_forbidden',
        )

    def test_t13b_workplace_lists_period_versions_and_sources(self):
        version = self._pool()
        self._login()
        response = self.client.get(reverse('arrival_roster_index'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Рабочее место табельщика')
        self.assertContains(response, 'Перевахта и состав заезда')
        self.assertContains(response, self.period.name)
        self.assertContains(response, self.composition.name)
        self.assertContains(response, 'Из базы сотрудников')
        self.assertContains(response, reverse('arrival_roster_review', args=[version.pk]))

    def test_t13b_create_from_employees_is_post_only_uses_session_and_keeps_history(self):
        self._login()
        url = reverse('arrival_roster_pool_create')
        self.assertEqual(self.client.get(url).status_code, 405)
        first = self.client.post(url, {
            'watch_period': self.period.pk,
            'actor_access_id': self.other_access.pk,
            'employee_access_id': self.other_access.pk,
            'created_at': '2000-01-01',
        })
        second = self.client.post(url, {'watch_period': self.period.pk})
        self.assertEqual(first.status_code, 302)
        self.assertEqual(second.status_code, 302)
        versions = list(ArrivalRosterVersion.objects.order_by('version_number'))
        self.assertEqual([item.version_number for item in versions], [1, 2])
        self.assertTrue(all(item.created_by_access_id == self.access.pk for item in versions))
        self.assertNotEqual(versions[0].snapshot_sha256, '')

    def test_t13b_excel_path_remains_available_as_verification_source(self):
        self._login()
        response = self.client.get(reverse('arrival_roster_upload_form'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Загрузить Excel для сверки')
        self.assertContains(response, 'Excel используется для сверки')

    def test_t13b_employee_search_covers_all_watches_filters_and_shows_phone(self):
        version = self._pool()
        self._login()
        response = self.client.get(reverse('arrival_roster_review', args=[version.pk]), {
            'employee_search': '1',
            'query': 'Механик',
            'watch_composition': self.other_composition.pk,
            'employment_status': 'active',
        })
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.other_employee.full_name)
        self.assertContains(response, self.other_employee.phone)
        self.assertEqual(response['Cache-Control'], 'private, no-store')
        self.assertEqual(
            [item.pk for item in response.context['employee_results']],
            [self.other_employee.pk],
        )

    def test_t13b_search_is_limited_deterministic_and_short_query_is_rejected(self):
        for index in range(35):
            Employee.objects.create(
                full_name=f'Поиск Сотрудник {index:02d}',
                position='Поисковая должность',
                status=Employee.Status.ACTIVE,
                is_active=True,
            )
        version = self._pool()
        self._login()
        response = self.client.get(reverse('arrival_roster_review', args=[version.pk]), {
            'employee_search': '1',
            'query': 'Поиск',
            'employment_status': 'active',
        })
        self.assertEqual(len(response.context['employee_results']), 30)
        names = [item.full_name for item in response.context['employee_results']]
        self.assertEqual(names, sorted(names))
        short = self.client.get(reverse('arrival_roster_review', args=[version.pk]), {
            'employee_search': '1', 'query': 'П', 'employment_status': 'all',
        })
        self.assertEqual(short.context['employee_results'], [])
        self.assertContains(short, 'Убедитесь, что это значение содержит не менее 2 символов')
        blank = self.client.get(reverse('arrival_roster_review', args=[version.pk]), {
            'employee_search': '1', 'query': '', 'employment_status': 'active',
        })
        self.assertEqual(blank.context['employee_results'], [])

    def test_t13b_other_role_gets_no_search_result_or_phone(self):
        version = self._pool()
        client = self._login(Client(), self.other_access)
        response = client.get(reverse('arrival_roster_review', args=[version.pk]), {
            'employee_search': '1', 'query': 'Механик', 'employment_status': 'all',
        })
        self.assertEqual(response.status_code, 302)
        self.assertNotIn(self.other_employee.phone, response.content.decode('utf-8', errors='ignore'))

    def test_t13b_adds_employee_from_other_watch_without_hidden_business_entities(self):
        version = self._pool()
        self._login()
        before = (Employee.objects.count(), EmployeeAccess.objects.count(), SettlementResident.objects.count())
        response = self.client.post(
            reverse('arrival_roster_employee_add', args=[version.pk]),
            {'employee_id': self.other_employee.pk, 'actor_access_id': self.other_access.pk},
        )
        self.assertEqual(response.status_code, 302)
        row = version.pool_rows.get(employee=self.other_employee)
        self.assertEqual(row.origin_kind, ArrivalRosterPoolRow.OriginKind.MANUAL_EMPLOYEE)
        self.assertTrue(row.match.issues.filter(code='employee_watch_composition_mismatch').exists())
        self.assertEqual(before, (Employee.objects.count(), EmployeeAccess.objects.count(), SettlementResident.objects.count()))
        event = version.events.get(action=ArrivalRosterEvent.Action.POOL_EMPLOYEE_ADDED)
        self.assertEqual(event.actor_access_id, self.access.pk)
        self.assertNotIn('phone', json.dumps(event.details))

    def test_t13b_dismissed_before_period_is_rejected_without_row(self):
        version = self._pool()
        self._login()
        response = self.client.post(
            reverse('arrival_roster_employee_add', args=[version.pk]),
            {'employee_id': self.dismissed.pk},
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(version.pool_rows.filter(employee=self.dismissed).exists())

    def test_t13b_duplicate_employee_is_controlled(self):
        version = self._pool()
        self._login()
        url = reverse('arrival_roster_employee_add', args=[version.pk])
        first = self.client.post(url, {'employee_id': self.other_employee.pk})
        second = self.client.post(url, {'employee_id': self.other_employee.pk})
        self.assertEqual(first.status_code, 302)
        self.assertEqual(second.status_code, 302)
        self.assertEqual(version.pool_rows.filter(employee=self.other_employee).count(), 1)

    def test_t13b_external_search_and_add_use_existing_card_and_safe_event(self):
        version = self._pool()
        self._login()
        search = self.client.get(reverse('arrival_roster_review', args=[version.pk]), {
            'external_search': '1', 'external_query': 'Подрядчик',
        })
        self.assertContains(search, self.external.full_name)
        self.assertContains(search, self.external.phone)
        self.assertEqual(search['Cache-Control'], 'private, no-store')
        before = (Employee.objects.count(), EmployeeAccess.objects.count(), SettlementResident.objects.count())
        added = self.client.post(reverse('arrival_roster_external_add', args=[version.pk]), {
            'resident_id': self.external.pk,
            'basis': 'Приглашён для выполнения пусконаладочных работ.',
            'actor_access_id': self.other_access.pk,
        })
        self.assertEqual(added.status_code, 302)
        self.assertEqual(before, (Employee.objects.count(), EmployeeAccess.objects.count(), SettlementResident.objects.count()))
        event = version.events.get(action=ArrivalRosterEvent.Action.POOL_EXTERNAL_ADDED)
        self.assertEqual(event.actor_access_id, self.access.pk)
        self.assertNotIn('phone', json.dumps(event.details))

    def test_t13b_review_has_all_russian_groups_and_no_final_confirmation(self):
        version = self._pool()
        self._login()
        response = self.client.get(reverse('arrival_roster_review', args=[version.pk]))
        for label in (
            'Ожидаются к заезду', 'Требуют решения табельщика', 'Продлеваются',
            'Не заезжают', 'Новые сотрудники', 'Требуют действий ОУП',
            'Требуют действий делопроизводителя',
            'Требуют назначения заместителем начальника участка',
        ):
            self.assertContains(response, label)
        self.assertNotContains(response, 'Утвердить версию')
        self.assertNotEqual(version.status, 'confirmed')

    def test_t13b_bulk_confirms_only_unambiguous_rows(self):
        unresolved_employee = Employee.objects.create(
            full_name='Без карточки Жильца T1.3b',
            position='Водитель',
            status=Employee.Status.ACTIVE,
            is_active=True,
            hired_at=date(2025, 1, 1),
            watch_composition=self.composition,
        )
        version = self._pool()
        exact = version.pool_rows.get(employee=self.employee).match.row_review
        unresolved = version.pool_rows.get(employee=unresolved_employee).match.row_review
        self._login()
        response = self.client.post(reverse('arrival_roster_confirm_unambiguous', args=[version.pk]))
        self.assertEqual(response.status_code, 302)
        exact.refresh_from_db()
        unresolved.refresh_from_db()
        self.assertEqual(exact.revision, 2)
        self.assertEqual(exact.participation_status, ArrivalRosterRowReview.ParticipationStatus.ARRIVING)
        self.assertEqual(unresolved.revision, 1)
        self.assertIsNone(unresolved.participation_status)
        version.refresh_from_db()
        self.assertIn(version.status, {ArrivalRosterVersion.Status.DRAFT, ArrivalRosterVersion.Status.REVIEW_REQUIRED})

    def test_t13b_bulk_confirmation_is_idempotent(self):
        version = self._pool()
        review = version.pool_rows.get(employee=self.employee).match.row_review
        revision_before = review.revision
        events_before = version.events.count()
        resident_events_before = version.events.filter(
            action=ArrivalRosterEvent.Action.RESIDENT_SELECTED,
        ).count()
        arrival_mode_events_before = version.events.filter(
            action=ArrivalRosterEvent.Action.ARRIVAL_MODE_CHANGED,
        ).count()

        first = confirm_unambiguous_arrival_roster_rows(
            version_id=version.pk,
            actor_access_id=self.access.pk,
        )
        review.refresh_from_db()
        self.assertEqual((first.changed, first.already_confirmed, first.skipped), (1, 0, 0))
        self.assertEqual(review.revision, revision_before + 1)
        self.assertEqual(version.events.count(), events_before + 1)
        self.assertEqual(
            version.events.filter(action=ArrivalRosterEvent.Action.RESIDENT_SELECTED).count(),
            resident_events_before,
        )
        self.assertEqual(
            version.events.filter(action=ArrivalRosterEvent.Action.ARRIVAL_MODE_CHANGED).count(),
            arrival_mode_events_before,
        )

        revision_after_first = review.revision
        events_after_first = version.events.count()
        second = confirm_unambiguous_arrival_roster_rows(
            version_id=version.pk,
            actor_access_id=self.access.pk,
        )
        review.refresh_from_db()
        self.assertEqual((second.changed, second.already_confirmed, second.skipped), (0, 1, 0))
        self.assertEqual(review.revision, revision_after_first)
        self.assertEqual(version.events.count(), events_after_first)

    def test_t13b_bulk_confirmation_changes_only_missing_state_in_mixed_batch(self):
        second_employee = Employee.objects.create(
            full_name='Второй Однозначный Смешанный T1.3b',
            position='Водитель',
            status=Employee.Status.ACTIVE,
            is_active=True,
            hired_at=date(2025, 1, 1),
            watch_composition=self.composition,
        )
        second_resident = SettlementResident(
            employee=second_employee,
            resident_type=SettlementResident.ResidentType.EMPLOYEE,
            status=SettlementResident.Status.ACTIVE,
        )
        second_resident.save()
        version = self._pool()
        first_review = version.pool_rows.get(employee=self.employee).match.row_review
        second_review = version.pool_rows.get(employee=second_employee).match.row_review
        set_arrival_roster_participation(
            match_id=first_review.match_id,
            participation_status=ArrivalRosterRowReview.ParticipationStatus.ARRIVING,
            arrival_mode='',
            expected_revision=first_review.revision,
            actor_access_id=self.access.pk,
        )
        first_review.refresh_from_db()
        first_revision = first_review.revision
        second_revision = second_review.revision
        events_before = version.events.count()

        result = confirm_unambiguous_arrival_roster_rows(
            version_id=version.pk,
            actor_access_id=self.access.pk,
        )
        first_review.refresh_from_db()
        second_review.refresh_from_db()
        self.assertEqual((result.changed, result.already_confirmed, result.skipped), (1, 1, 0))
        self.assertEqual(first_review.revision, first_revision)
        self.assertEqual(second_review.revision, second_revision + 1)
        self.assertEqual(version.events.count(), events_before + 1)
        self.assertEqual(
            version.events.filter(
                match_id=second_review.match_id,
                action=ArrivalRosterEvent.Action.PARTICIPATION_CHANGED,
            ).count(),
            1,
        )

    def test_t13b_bulk_confirmation_does_not_replace_conflicting_manual_resident(self):
        version = self._pool()
        review = version.pool_rows.get(employee=self.employee).match.row_review
        ArrivalRosterRowReview._base_manager.filter(pk=review.pk).update(
            selected_resident_id=self.other_resident.pk,
        )
        review.refresh_from_db()
        revision_before = review.revision
        events_before = version.events.count()

        result = confirm_unambiguous_arrival_roster_rows(
            version_id=version.pk,
            actor_access_id=self.access.pk,
        )
        review.refresh_from_db()
        self.assertEqual((result.changed, result.already_confirmed, result.skipped), (0, 0, 1))
        self.assertEqual(review.selected_resident_id, self.other_resident.pk)
        self.assertEqual(review.revision, revision_before)
        self.assertEqual(version.events.count(), events_before)

    def test_t13b_bulk_confirmation_does_not_replace_manual_participation(self):
        version = self._pool()
        review = version.pool_rows.get(employee=self.employee).match.row_review
        review = set_arrival_roster_participation(
            match_id=review.match_id,
            participation_status=ArrivalRosterRowReview.ParticipationStatus.NOT_ARRIVING,
            arrival_mode='',
            expected_revision=review.revision,
            actor_access_id=self.access.pk,
        )
        revision_before = review.revision
        events_before = version.events.count()

        result = confirm_unambiguous_arrival_roster_rows(
            version_id=version.pk,
            actor_access_id=self.access.pk,
        )
        review.refresh_from_db()
        self.assertEqual((result.changed, result.already_confirmed, result.skipped), (0, 0, 1))
        self.assertEqual(
            review.participation_status,
            ArrivalRosterRowReview.ParticipationStatus.NOT_ARRIVING,
        )
        self.assertEqual(review.revision, revision_before)
        self.assertEqual(version.events.count(), events_before)

    def test_t13b_bulk_confirmation_rolls_back_atomically(self):
        second_employee = Employee.objects.create(
            full_name='Второй Однозначный T1.3b',
            position='Водитель',
            status=Employee.Status.ACTIVE,
            is_active=True,
            hired_at=date(2025, 1, 1),
            watch_composition=self.composition,
        )
        resident = SettlementResident(
            employee=second_employee,
            resident_type=SettlementResident.ResidentType.EMPLOYEE,
            status=SettlementResident.Status.ACTIVE,
        )
        resident.save()
        version = self._pool()
        revisions_before = list(version.row_reviews.order_by('pk').values_list('revision', flat=True))
        events_before = version.events.count()
        original = _trusted_create_arrival_roster_event
        calls = {'count': 0}

        def fail_second_event(**kwargs):
            calls['count'] += 1
            if calls['count'] == 2:
                raise ValidationError('Проверочный отказ второй строки.')
            return original(**kwargs)

        with patch(
            'rotations.arrival_roster_pool._trusted_create_arrival_roster_event',
            side_effect=fail_second_event,
        ):
            with self.assertRaises(ValidationError):
                confirm_unambiguous_arrival_roster_rows(
                    version_id=version.pk,
                    actor_access_id=self.access.pk,
                )
        self.assertEqual(
            list(version.row_reviews.order_by('pk').values_list('revision', flat=True)),
            revisions_before,
        )
        self.assertEqual(version.events.count(), events_before)

    def test_t13b_bulk_confirmation_result_is_independent_of_row_names(self):
        employees = []
        for name in ('Я Последний По Имени T1.3b', 'А Первый По Имени T1.3b'):
            employee = Employee.objects.create(
                full_name=name,
                position='Водитель',
                status=Employee.Status.ACTIVE,
                is_active=True,
                hired_at=date(2025, 1, 1),
                watch_composition=self.composition,
            )
            resident = SettlementResident(
                employee=employee,
                resident_type=SettlementResident.ResidentType.EMPLOYEE,
                status=SettlementResident.Status.ACTIVE,
            )
            resident.save()
            employees.append(employee)
        version = self._pool()

        result = confirm_unambiguous_arrival_roster_rows(
            version_id=version.pk,
            actor_access_id=self.access.pk,
        )
        self.assertEqual((result.changed, result.already_confirmed, result.skipped), (3, 0, 0))
        self.assertEqual(
            list(
                version.row_reviews
                .order_by('match__pool_row__employee__full_name')
                .values_list('participation_status', flat=True)
            ),
            [ArrivalRosterRowReview.ParticipationStatus.ARRIVING] * 3,
        )

    def test_t13b_mutations_are_post_and_csrf_protected(self):
        version = self._pool()
        client = self._login(Client(enforce_csrf_checks=True))
        urls = (
            reverse('arrival_roster_employee_add', args=[version.pk]),
            reverse('arrival_roster_external_add', args=[version.pk]),
            reverse('arrival_roster_confirm_unambiguous', args=[version.pk]),
        )
        for url in urls:
            self.assertEqual(client.get(url).status_code, 405)
            self.assertEqual(client.post(url, {}).status_code, 403)

    def test_t13a_match_issue_and_links_reject_public_writes(self):
        version = self._create_pool()
        row = version.pool_rows.get(employee=self.employee)
        match = row.match
        issue = _trusted_create_arrival_roster_issue(
            version=version,
            match=match,
            severity=ArrivalRosterIssue.Severity.WARNING,
            code='trusted_test_warning',
            message='Проверочное предупреждение.',
            details={},
        )
        new_match = ArrivalRosterMatch(
            version=version,
            status=ArrivalRosterMatch.Status.UNMATCHED,
            method='public_test',
            quality='unmatched',
            evidence={},
        )
        new_issue = ArrivalRosterIssue(
            version=version,
            severity=ArrivalRosterIssue.Severity.WARNING,
            code='public_test',
            message='Подмена.',
            details={},
        )
        operations = [
            new_match.save,
            lambda: ArrivalRosterMatch.objects.create(
                version=version,
                status='unmatched',
                method='public_test',
                quality='unmatched',
                evidence={},
            ),
            lambda: ArrivalRosterMatch.objects.bulk_create([new_match]),
            lambda: ArrivalRosterMatch.objects.get_or_create(
                version=version,
                method='public_get_or_create',
                defaults={
                    'status': 'unmatched',
                    'quality': 'unmatched',
                    'evidence': {},
                },
            ),
            lambda: ArrivalRosterMatch.objects.update_or_create(
                version=version,
                method='public_update_or_create',
                defaults={
                    'status': 'unmatched',
                    'quality': 'unmatched',
                    'evidence': {},
                },
            ),
            new_issue.save,
            lambda: ArrivalRosterIssue.objects.create(
                version=version,
                severity='warning',
                code='public_test',
                message='Подмена.',
                details={},
            ),
            lambda: ArrivalRosterIssue.objects.bulk_create([new_issue]),
            lambda: ArrivalRosterIssue.objects.get_or_create(
                version=version,
                code='public_get_or_create',
                defaults={
                    'severity': 'warning',
                    'message': 'Подмена.',
                    'details': {},
                },
            ),
            lambda: ArrivalRosterIssue.objects.update_or_create(
                version=version,
                code='public_update_or_create',
                defaults={
                    'severity': 'warning',
                    'message': 'Подмена.',
                    'details': {},
                },
            ),
            lambda: ArrivalRosterMatch.objects.filter(pk=match.pk).update(method='changed'),
            lambda: ArrivalRosterIssue.objects.filter(pk=issue.pk).update(message='Подмена.'),
            match.delete,
            issue.delete,
        ]
        for operation in operations:
            with self.assertRaises(ValidationError) as caught:
                operation()
            self.assertEqual(
                caught.exception.code,
                'rotations.arrival_roster.public_write_forbidden',
            )

    def test_t13a_trusted_pool_writer_rejects_tampered_employee_snapshot(self):
        version = self._create_pool()
        employee = Employee.objects.create(
            full_name='Неподменённый сотрудник T1.3a',
            status=Employee.Status.ACTIVE,
            is_active=True,
            hired_at=date(2025, 1, 1),
            watch_composition=self.composition,
        )
        resident = SettlementResident(
            employee=employee,
            resident_type=SettlementResident.ResidentType.EMPLOYEE,
            status=SettlementResident.Status.ACTIVE,
        )
        resident.save()
        match = _trusted_create_arrival_roster_match(
            version=version,
            status=ArrivalRosterMatch.Status.EXACT,
            method='trusted_snapshot_test',
            quality='exact',
            matched_resident=resident,
            evidence={'employee_id': employee.pk},
        )
        actor_context = _lock_timekeeper_access(_access_snapshot(self.access.pk))
        employee.full_name = 'Подменённое имя'
        with self.assertRaises(ValidationError) as caught:
            _trusted_create_pool_row(
                version=version,
                match=match,
                resident=resident,
                employee=employee,
                watch_composition=self.composition,
                origin_kind=ArrivalRosterPoolRow.OriginKind.MANUAL_EMPLOYEE,
                suggested_participation=ArrivalRosterPoolRow.SuggestedParticipation.ADDITIONAL,
                basis='Проверка защиты снимка.',
                actor_context=actor_context,
            )
        self.assertEqual(caught.exception.code, 'arrival_roster.pool_snapshot_mismatch')
        self.assertFalse(version.pool_rows.filter(employee=employee).exists())

    def test_t13a_pool_generation_rolls_back_all_rows_on_one_failure(self):
        second_employee = Employee.objects.create(
            full_name='Второй сотрудник для отката T1.3a',
            status=Employee.Status.ACTIVE,
            is_active=True,
            hired_at=date(2025, 1, 1),
            watch_composition=self.composition,
        )
        second_resident = SettlementResident(
            employee=second_employee,
            resident_type=SettlementResident.ResidentType.EMPLOYEE,
            status=SettlementResident.Status.ACTIVE,
        )
        second_resident.save()
        from . import arrival_roster_pool
        original = arrival_roster_pool._trusted_create_pool_row
        call_count = 0

        def fail_second(*args, **kwargs):
            nonlocal call_count
            call_count += 1
            if call_count == 2:
                raise ValidationError('Проверочная ошибка строки.', code='test.row_failure')
            return original(*args, **kwargs)

        with patch.object(arrival_roster_pool, '_trusted_create_pool_row', side_effect=fail_second):
            with self.assertRaises(ValidationError) as caught:
                self._create_pool()
        self.assertEqual(caught.exception.code, 'test.row_failure')
        self.assertEqual(ArrivalRosterVersion.objects.count(), 0)
        self.assertEqual(ArrivalRosterPoolRow.objects.count(), 0)
        self.assertEqual(ArrivalRosterMatch.objects.count(), 0)
        self.assertEqual(ArrivalRosterIssue.objects.count(), 0)

    def test_t13a_excel_path_remains_independent_of_pool_rows(self):
        uploaded = SimpleUploadedFile(
            'совместимость.xlsx',
            _workbook_bytes(
                full_name=self.employee.full_name,
                primary_name=self.employee.full_name,
                phone='',
            ),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        version, created = upload_arrival_roster(
            uploaded_file=uploaded,
            watch_period_id=self.period.pk,
            actor_access_id=self.access.pk,
        )

        self.assertTrue(created)
        self.assertEqual(version.source_kind, ArrivalRosterVersion.SourceKind.EXCEL)
        self.assertIsNotNone(version.source_file_id)
        self.assertIsNotNone(version.parser_profile_id)
        self.assertEqual(version.source_fingerprint, version.source_file.sha256)
        self.assertEqual(version.pool_rows.count(), 0)
        link = ArrivalRosterMatchRow.objects.order_by('pk').first()
        self.assertIsNotNone(link)
        candidate = ArrivalRosterMatchCandidate(
            match=link.match,
            resident=self.other_resident,
            evidence={},
        )
        operations = [
            lambda: ArrivalRosterMatchRow(
                match=link.match,
                normalized_row=link.normalized_row,
            ).save(),
            candidate.save,
            lambda: ArrivalRosterMatchCandidate.objects.bulk_create([candidate]),
            lambda: ArrivalRosterMatchRow.objects.filter(pk=link.pk).update(
                match_id=link.match_id,
            ),
            link.delete,
        ]
        for operation in operations:
            with self.assertRaises(ValidationError) as caught:
                operation()
            self.assertEqual(
                caught.exception.code,
                'rotations.arrival_roster.public_write_forbidden',
            )

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from numbers import Number

from django.core.exceptions import ObjectDoesNotExist
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from .models import RotationCollectionCycle, RotationResponse, WatchExtensionCase


XLSX_DARK = "263D38"
XLSX_GREEN = "4D7A69"
XLSX_GREEN_SOFT = "E8F1ED"
XLSX_LINE = "CBD8D2"
XLSX_MUTED = "66756F"
XLSX_WHITE = "FFFFFF"
XLSX_WARNING = "FFF4D6"

DATA_HEADER_ROW = 4


def _safe_text(value):
    """Return plain Excel text and neutralize formula-like user input."""
    if value is None:
        return ""
    text = str(value)
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _excel_value(value):
    if isinstance(value, datetime) and timezone.is_aware(value):
        return timezone.localtime(value).replace(tzinfo=None)
    if isinstance(value, (date, datetime)):
        return value
    if isinstance(value, Number):
        return value
    return _safe_text(value)


def _display(instance, field_name, *, default=""):
    if instance is None:
        return default
    raw_value = getattr(instance, field_name, None)
    if raw_value in (None, ""):
        return default
    display_method = getattr(instance, f"get_{field_name}_display", None)
    if callable(display_method):
        displayed = display_method()
        if displayed not in (None, ""):
            return displayed
    return raw_value


def _person_name(person):
    if person is None:
        return ""
    return _safe_text(getattr(person, "full_name", "") or str(person))


def _brigade_label(response: RotationResponse):
    brigade_number = getattr(
        response,
        "snapshot_brigade_number",
        getattr(response, "brigade_number", None),
    )
    if not brigade_number:
        return ""
    return f"Бригада №{brigade_number}"


def _response_snapshot(response: RotationResponse, field_name):
    snapshot_name = {
        "full_name": "snapshot_full_name",
        "personnel_number": "snapshot_personnel_number",
        "position": "snapshot_position",
        "department": "snapshot_department",
        "work_schedule": "snapshot_work_schedule",
    }[field_name]
    value = getattr(response, snapshot_name, None)
    if value in (None, ""):
        legacy_name = {
            "full_name": "employee_name",
            "personnel_number": "personnel_number",
            "position": "position_name",
            "department": "department_name",
            "work_schedule": "work_schedule_name",
        }[field_name]
        value = getattr(response, legacy_name, "")
    return value


def _shift_label(response: RotationResponse):
    if not response.next_shift_type:
        return "Не определена"
    return _safe_text(_display(response, "next_shift_type", default="Не определена"))


def _extension_case(response: RotationResponse):
    try:
        return response.extension_case
    except ObjectDoesNotExist:
        return None


def _cycle_responses(cycle: RotationCollectionCycle):
    queryset = cycle.responses.select_related(
        "employee",
        "submitted_by",
        "extension_case__decision_by",
        "extension_case__documentation_by",
    )
    return list(queryset.order_by("snapshot_full_name", "id"))


def _is_submitted(response: RotationResponse):
    return response.state == "submitted"


def _is_travel_response(response: RotationResponse):
    return _is_submitted(response) and response.intent in {
        "arrival",
        "departure",
        "not_travelling",
        "travel",
    }


def _is_extension_response(response: RotationResponse):
    return _is_submitted(response) and response.intent == "extension"


def _is_approved(case: WatchExtensionCase | None):
    return bool(case and case.decision_status == "approved")


def _write_cell(cell, value):
    value = _excel_value(value)
    cell.value = value
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    if isinstance(value, datetime):
        cell.number_format = "dd.mm.yyyy hh:mm"
    elif isinstance(value, date):
        cell.number_format = "dd.mm.yyyy"


def _append_row(sheet, values):
    row_index = sheet.max_row + 1
    for column_index, value in enumerate(values, start=1):
        cell = sheet.cell(row=row_index, column=column_index)
        _write_cell(cell, value)
        cell.border = Border(bottom=Side(style="thin", color=XLSX_LINE))
        if row_index % 2 == 0:
            cell.fill = PatternFill("solid", fgColor="F7FAF8")
    return row_index


def _fit_columns(sheet, headers):
    for column_index, header in enumerate(headers, start=1):
        width = len(str(header)) + 2
        for row_index in range(DATA_HEADER_ROW + 1, sheet.max_row + 1):
            value = sheet.cell(row=row_index, column=column_index).value
            if isinstance(value, datetime):
                visible_length = 16
            elif isinstance(value, date):
                visible_length = 10
            else:
                visible_length = max((len(part) for part in str(value or "").splitlines()), default=0)
            width = max(width, visible_length + 2)
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max(width, 10), 38)


def _configure_print(sheet, *, header_row=DATA_HEADER_ROW):
    last_column = max(sheet.max_column, 1)
    last_row = max(sheet.max_row, header_row)
    last_column_letter = get_column_letter(last_column)
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:{last_column_letter}{last_row}"
    sheet.print_area = f"A1:{last_column_letter}{last_row}"
    sheet.print_title_rows = f"{header_row}:{header_row}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_options.horizontalCentered = True
    sheet.page_margins = PageMargins(
        left=0.25,
        right=0.25,
        top=0.35,
        bottom=0.45,
        header=0.15,
        footer=0.2,
    )
    sheet.oddFooter.center.text = "Страница &P из &N"
    sheet.oddFooter.center.size = 9
    sheet.oddFooter.center.color = XLSX_MUTED


def _add_data_sheet(workbook, title, description, headers, rows):
    sheet = workbook.create_sheet(title)
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet["A1"] = _safe_text(title)
    sheet["A1"].fill = PatternFill("solid", fgColor=XLSX_DARK)
    sheet["A1"].font = Font(name="Arial", size=15, bold=True, color=XLSX_WHITE)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    sheet["A2"] = _safe_text(description)
    sheet["A2"].fill = PatternFill("solid", fgColor=XLSX_GREEN_SOFT)
    sheet["A2"].font = Font(name="Arial", size=9, color=XLSX_MUTED)
    sheet["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 24

    header_fill = PatternFill("solid", fgColor=XLSX_GREEN)
    header_border = Border(
        left=Side(style="thin", color=XLSX_LINE),
        right=Side(style="thin", color=XLSX_LINE),
        top=Side(style="thin", color=XLSX_LINE),
        bottom=Side(style="thin", color=XLSX_LINE),
    )
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=DATA_HEADER_ROW, column=column_index, value=_safe_text(header))
        cell.fill = header_fill
        cell.font = Font(name="Arial", size=9, bold=True, color=XLSX_WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border
    sheet.row_dimensions[DATA_HEADER_ROW].height = 32

    for values in rows:
        _append_row(sheet, values)

    _fit_columns(sheet, headers)
    _configure_print(sheet)
    return sheet


def _build_summary_sheet(
    workbook,
    cycle: RotationCollectionCycle,
    responses,
    cases,
    *,
    generated_by=None,
):
    sheet = workbook.active
    sheet.title = "Сводка"
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:B1")
    sheet["A1"] = "ПЕРЕВАХТА · СВОДКА СБОРА"
    sheet["A1"].fill = PatternFill("solid", fgColor=XLSX_DARK)
    sheet["A1"].font = Font(name="Arial", size=16, bold=True, color=XLSX_WHITE)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 30

    target_watch = cycle.target_watch_period
    submitted = [response for response in responses if _is_submitted(response)]
    travel = [response for response in responses if _is_travel_response(response)]
    extension = [response for response in responses if _is_extension_response(response)]
    approved = [case for case in cases if _is_approved(case)]
    rejected = [case for case in cases if case and case.decision_status == "rejected"]
    pending_decisions = [
        case
        for case in cases
        if case and case.decision_status not in {"approved", "rejected"}
    ]
    shift_sources = sorted(
        {
            str(_display(response, "shift_source"))
            for response in submitted
            if response.shift_source
        }
    )

    rows = [
        ("Сбор", cycle.name),
        ("Целевая вахта", target_watch.name),
        ("Начало вахты", target_watch.starts_on),
        ("Окончание вахты", target_watch.ends_on),
        ("Срок ответа", cycle.response_deadline),
        ("Статус", _display(cycle, "status")),
        ("Ревизия", cycle.revision),
        ("Сформировал", _person_name(generated_by)),
        ("Сформировано", timezone.now()),
        ("Адресатов", len(responses)),
        ("Ответили", len(submitted)),
        ("Не ответили", len(responses) - len(submitted)),
        ("Маршруты", len(travel)),
        ("Запросы на продление", len(extension)),
        ("Продление одобрено", len(approved)),
        ("Продление отклонено", len(rejected)),
        ("Ожидают решения", len(pending_decisions)),
        ("Источники смены", ", ".join(shift_sources)),
    ]
    for row_index, (label, value) in enumerate(rows, start=3):
        label_cell = sheet.cell(row=row_index, column=1, value=_safe_text(label))
        label_cell.fill = PatternFill("solid", fgColor=XLSX_GREEN_SOFT)
        label_cell.font = Font(name="Arial", size=10, bold=True, color=XLSX_DARK)
        label_cell.alignment = Alignment(vertical="top", wrap_text=True)
        label_cell.border = Border(bottom=Side(style="thin", color=XLSX_LINE))
        value_cell = sheet.cell(row=row_index, column=2)
        _write_cell(value_cell, value)
        value_cell.font = Font(name="Arial", size=10, color=XLSX_DARK)
        value_cell.border = Border(bottom=Side(style="thin", color=XLSX_LINE))

    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 52
    sheet.freeze_panes = "A3"
    sheet.print_area = f"A1:B{sheet.max_row}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.page_margins = PageMargins(left=0.35, right=0.35, top=0.45, bottom=0.45)
    return sheet


def build_cycle_workbook(cycle: RotationCollectionCycle, generated_by=None):
    responses = _cycle_responses(cycle)
    response_cases = [(response, _extension_case(response)) for response in responses]
    cases = [case for _response, case in response_cases if case is not None]

    workbook = Workbook()
    _build_summary_sheet(
        workbook,
        cycle,
        responses,
        cases,
        generated_by=generated_by,
    )

    generated_label = _person_name(generated_by) or "не указан"
    description = (
        f"Сбор: {cycle.name} · ревизия {cycle.revision} · "
        f"сформировал: {generated_label} · {timezone.localtime():%d.%m.%Y %H:%M}"
    )

    route_rows = []
    for number, response in enumerate(
        (item for item in responses if _is_travel_response(item)),
        start=1,
    ):
        route_rows.append(
            (
                number,
                _response_snapshot(response, "full_name"),
                _response_snapshot(response, "personnel_number"),
                _response_snapshot(response, "position"),
                _response_snapshot(response, "department"),
                _response_snapshot(response, "work_schedule"),
                _brigade_label(response),
                _shift_label(response),
                _display(response, "shift_source"),
                _display(response, "intent"),
                response.departure_on,
                response.arrival_on,
                response.route_text,
                _display(response, "travel_mode"),
                _display(response, "transfer_mode"),
                response.transport_details,
                response.comment,
                response.submitted_at,
            )
        )
    _add_data_sheet(
        workbook,
        "Маршруты",
        description,
        (
            "№",
            "ФИО",
            "Табельный номер",
            "Должность",
            "Подразделение",
            "График работы",
            "Бригада",
            "Смена",
            "Источник смены",
            "Вариант поездки",
            "Дата выезда",
            "Дата заезда",
            "Маршрут",
            "Способ проезда",
            "Трансфер",
            "Детали транспорта",
            "Комментарий",
            "Ответ отправлен",
        ),
        route_rows,
    )

    approved_rows = []
    approved_pairs = [
        (response, case)
        for response, case in response_cases
        if _is_approved(case)
    ]
    for number, (response, case) in enumerate(approved_pairs, start=1):
        approved_rows.append(
            (
                number,
                _response_snapshot(response, "full_name"),
                _response_snapshot(response, "personnel_number"),
                _response_snapshot(response, "position"),
                _response_snapshot(response, "department"),
                _response_snapshot(response, "work_schedule"),
                _brigade_label(response),
                _shift_label(response),
                _display(response, "shift_source"),
                cycle.target_watch_period.name,
                case.extension_start,
                case.extension_end,
                _display(case, "decision_status"),
                _person_name(case.decision_by),
                case.decision_at,
                case.decision_comment,
                _display(case, "documentation_status"),
                _person_name(case.documentation_by),
                case.documentation_at,
                case.documentation_note,
                response.submitted_at,
            )
        )
    _add_data_sheet(
        workbook,
        "Продление к оформлению",
        f"Только заявки со статусом «Одобрено». {description}",
        (
            "№",
            "ФИО",
            "Табельный номер",
            "Должность",
            "Подразделение",
            "График работы",
            "Бригада",
            "Смена",
            "Источник смены",
            "Целевая вахта",
            "Продление с",
            "Продление по",
            "Решение",
            "Согласовал",
            "Дата решения",
            "Комментарий решения",
            "Статус документов",
            "Оформил",
            "Дата оформления",
            "Примечание к документам",
            "Ответ отправлен",
        ),
        approved_rows,
    )

    control_rows = []
    for number, (response, case) in enumerate(response_cases, start=1):
        control_rows.append(
            (
                number,
                _response_snapshot(response, "full_name"),
                _response_snapshot(response, "personnel_number"),
                _response_snapshot(response, "position"),
                _response_snapshot(response, "department"),
                _brigade_label(response),
                _display(response, "state"),
                _display(response, "intent"),
                _shift_label(response),
                _display(response, "shift_source"),
                response.submitted_at,
                response.updated_at,
                _display(case, "decision_status") if case else "",
                _display(case, "documentation_status") if case else "",
            )
        )
    control_sheet = _add_data_sheet(
        workbook,
        "Контроль ответов",
        description,
        (
            "№",
            "ФИО",
            "Табельный номер",
            "Должность",
            "Подразделение",
            "Бригада",
            "Статус ответа",
            "Вариант ответа",
            "Смена",
            "Источник смены",
            "Ответ отправлен",
            "Последнее изменение",
            "Решение по продлению",
            "Статус документов",
        ),
        control_rows,
    )
    for row_index, response in enumerate(responses, start=DATA_HEADER_ROW + 1):
        if not _is_submitted(response):
            for cell in control_sheet[row_index]:
                cell.fill = PatternFill("solid", fgColor=XLSX_WARNING)

    return workbook


def workbook_bytes(workbook: Workbook):
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()

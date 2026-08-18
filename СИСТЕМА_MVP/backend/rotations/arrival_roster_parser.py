import hashlib
import io
import json
import re
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import PurePosixPath

from django.core.exceptions import ValidationError
from openpyxl import load_workbook

from users.forms import normalize_phone


MAX_FILE_SIZE = 10 * 1024 * 1024
MAX_UNCOMPRESSED_SIZE = 50 * 1024 * 1024
MAX_ARCHIVE_ENTRIES = 2000
MAX_COMPRESSION_RATIO = 100

PROFILE_CODE = 'arrival-roster-14-08-26'
PROFILE_VERSION = 1
PROFILE_CONFIGURATION = {
    'sheets': {
        'билеты': {
            'range': 'A1:G194',
            'max_row': 194,
            'max_column': 7,
            'kind': 'supporting',
            'columns': {
                'name': 2, 'position': 3, 'shift': 4, 'date': 5,
                'route': 6, 'phone': 7,
            },
            'headers': {2: 'фио', 3: 'должность', 4: 'смена', 5: 'дата', 6: 'маршрут', 7: 'телефон'},
        },
        'трансфер,сам': {
            'range': 'A1:F29',
            'max_row': 29,
            'max_column': 6,
            'kind': 'supporting',
            'columns': {
                'name': 2, 'position': 3, 'route': 4, 'shift': 5, 'comments': 6,
            },
            'headers': {2: 'сотрудника', 3: 'должность', 4: 'трансфер', 5: 'смена'},
        },
        'продление 14.04.26-14.05.26': {
            'range': 'A1:F27',
            'max_row': 27,
            'max_column': 6,
            'kind': 'supporting',
            'columns': {
                'name': 2, 'position': 3, 'comments': [4, 5], 'date': 6,
            },
            'headers': {2: 'фио', 3: 'должность', 4: 'согласование', 6: 'дата'},
        },
        'Список сотрудников': {
            'range': 'A1:D249',
            'max_row': 249,
            'max_column': 4,
            'kind': 'primary',
            'columns': {'name': 2, 'position': 3, 'comments': 4},
            'headers': {2: 'фио', 3: 'должность'},
        },
        'Числ': {
            'range': 'A1:J63',
            'max_row': 63,
            'max_column': 10,
            'kind': 'summary',
            'columns': {},
            'headers': {},
        },
    },
    'primary_sheet': 'Список сотрудников',
    'formula_columns_allowed': {'person_sheets': [1], 'summary_sheet': 'all'},
    'color_rules': {},
    'shift_values_are_hints_only': True,
}


class UnsafeArrivalWorkbook(ValidationError):
    pass


@dataclass(frozen=True)
class ParsedIssue:
    severity: str
    code: str
    message: str
    sheet_name: str = ''
    row_number: int | None = None
    details: dict | None = None


@dataclass(frozen=True)
class ParsedSourceRow:
    sheet_name: str
    row_number: int
    row_kind: str
    raw_values: list
    raw_styles: list
    row_sha256: str


@dataclass(frozen=True)
class ParsedNormalizedRow:
    sheet_name: str
    row_number: int
    raw_full_name: str
    normalized_full_name: str
    normalized_name_key: str
    name_comment: str
    source_position: str
    normalized_position_key: str
    raw_shift_hint: str
    raw_date: str
    arrival_date_candidate: date | None
    date_comment: str
    route_text: str
    raw_phone: str
    normalized_phones: list[str]
    comments: str
    participation_hint: str
    color_hint: dict


@dataclass(frozen=True)
class ParsedArrivalWorkbook:
    source_rows: list[ParsedSourceRow]
    normalized_rows: list[ParsedNormalizedRow]
    issues: list[ParsedIssue]
    workbook_summary: dict


def parser_profile_snapshot():
    configuration = json.loads(json.dumps(PROFILE_CONFIGURATION, ensure_ascii=False))
    encoded = json.dumps(
        configuration,
        ensure_ascii=False,
        sort_keys=True,
        separators=(',', ':'),
    ).encode('utf-8')
    return configuration, hashlib.sha256(encoded).hexdigest()


def _unsafe(message, code):
    return UnsafeArrivalWorkbook(message, code=code)


def validate_xlsx_archive(payload):
    try:
        archive = zipfile.ZipFile(io.BytesIO(payload))
    except (zipfile.BadZipFile, OSError) as error:
        raise _unsafe('Файл не является корректной книгой .xlsx.', 'unsafe_xlsx') from error

    with archive:
        entries = archive.infolist()
        if len(entries) > MAX_ARCHIVE_ENTRIES:
            raise _unsafe('В книге слишком много внутренних элементов.', 'xlsx_too_many_entries')
        names_seen = set()
        total_uncompressed = 0
        for entry in entries:
            normalized_name = entry.filename.replace('\\', '/')
            path = PurePosixPath(normalized_name)
            if path.is_absolute() or '..' in path.parts:
                raise _unsafe('Книга содержит небезопасный путь.', 'xlsx_unsafe_path')
            lowered_name = normalized_name.casefold()
            if lowered_name in names_seen:
                raise _unsafe('Книга содержит повторяющиеся внутренние элементы.', 'xlsx_duplicate_entry')
            names_seen.add(lowered_name)
            if entry.flag_bits & 0x1:
                raise _unsafe('Зашифрованные книги не поддерживаются.', 'xlsx_encrypted')
            total_uncompressed += entry.file_size
            if total_uncompressed > MAX_UNCOMPRESSED_SIZE:
                raise _unsafe('Распакованный размер книги превышает 50 МиБ.', 'xlsx_unpacked_too_large')
            if entry.file_size:
                if not entry.compress_size:
                    raise _unsafe('Книга содержит опасно сжатый элемент.', 'xlsx_compression_ratio')
                if entry.file_size / entry.compress_size > MAX_COMPRESSION_RATIO:
                    raise _unsafe('Степень сжатия книги превышает безопасный предел.', 'xlsx_compression_ratio')
            if any(
                marker in lowered_name
                for marker in ('vbaproject', 'activex', 'oleobject', 'embeddings/')
            ):
                raise _unsafe('Книга содержит макросы или встроенные исполняемые объекты.', 'xlsx_active_content')
            if lowered_name.startswith('xl/externallinks/'):
                raise _unsafe('Книга содержит внешние связи.', 'xlsx_external_links')

        required_parts = {'[content_types].xml', 'xl/workbook.xml'}
        if not required_parts.issubset(names_seen):
            raise _unsafe('В книге отсутствуют обязательные части .xlsx.', 'xlsx_missing_parts')

        for entry in entries:
            lowered_name = entry.filename.casefold()
            if not (lowered_name.endswith('.xml') or lowered_name.endswith('.rels')):
                continue
            content = archive.read(entry)
            lowered_content = content.lower()
            if b'<!doctype' in lowered_content or b'<!entity' in lowered_content:
                raise _unsafe('Книга содержит запрещённые XML-объявления.', 'xlsx_unsafe_xml')
            if lowered_name.endswith('.rels') and re.search(
                rb'targetmode\s*=\s*[\'\"]external[\'\"]',
                lowered_content,
            ):
                raise _unsafe('Книга содержит внешние связи.', 'xlsx_external_links')
            if any(
                marker in lowered_content
                for marker in (b'vbaproject', b'activex', b'oleobject')
            ):
                raise _unsafe(
                    'Книга содержит макросы или встроенные исполняемые объекты.',
                    'xlsx_active_content',
                )

        bad_entry = archive.testzip()
        if bad_entry:
            raise _unsafe('Повреждена внутренняя структура книги.', 'xlsx_bad_crc')


def _plain_text(value):
    if value is None:
        return ''
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def _normalized_text(value):
    return ' '.join(_plain_text(value).replace('\xa0', ' ').split())


def _text_key(value):
    return _normalized_text(value).casefold().replace('ё', 'е')


_COMMENT_RE = re.compile(r'\(([^()]*)\)')


def _split_name_and_comment(value):
    raw = _plain_text(value)
    comments = [item.strip() for item in _COMMENT_RE.findall(raw) if item.strip()]
    without_parentheses = _COMMENT_RE.sub(' ', raw)
    lines = [item.strip() for item in without_parentheses.splitlines() if item.strip()]
    name = lines[0] if lines else ''
    if len(lines) > 1:
        comments.extend(lines[1:])
    normalized_name = _normalized_text(name)
    return normalized_name, '; '.join(comments)


_PHONE_CANDIDATE_RE = re.compile(r'(?:\+?\d[\d\s()\-]{8,}\d)')


def _normalized_phones(value):
    result = []
    for candidate in _PHONE_CANDIDATE_RE.findall(_plain_text(value)):
        digits = normalize_phone(candidate)
        if len(digits) == 10:
            digits = f'7{digits}'
        if len(digits) == 11 and digits.startswith('7') and digits not in result:
            result.append(digits)
    return result


_DATE_RE = re.compile(r'(?<!\d)(\d{1,2})[.]([01]?\d)[.](\d{4})(?!\d)')
_AMBIGUOUS_DATE_WORDS = ('примерно', 'ориентировочно', 'около', 'утром', 'вечером')


def _parse_date_candidate(value):
    if isinstance(value, datetime):
        return value.date(), ''
    if isinstance(value, date):
        return value, ''
    raw = _plain_text(value)
    if not raw:
        return None, ''
    matches = list(_DATE_RE.finditer(raw))
    lowered = raw.casefold()
    if len(matches) != 1 or any(word in lowered for word in _AMBIGUOUS_DATE_WORDS):
        return None, raw
    match = matches[0]
    try:
        candidate = date(int(match.group(3)), int(match.group(2)), int(match.group(1)))
    except ValueError:
        return None, raw
    comment = _normalized_text(f'{raw[:match.start()]} {raw[match.end():]}')
    return candidate, comment


def _json_cell(cell):
    value = cell.value
    if cell.data_type == 'f':
        value_type = 'formula'
        serialized = str(value or '')
    elif isinstance(value, datetime):
        value_type = 'datetime'
        serialized = value.isoformat()
    elif isinstance(value, date):
        value_type = 'date'
        serialized = value.isoformat()
    elif isinstance(value, bool):
        value_type = 'boolean'
        serialized = value
    elif isinstance(value, (int, float)):
        value_type = 'number'
        serialized = value
    elif value is None:
        value_type = 'blank'
        serialized = None
    else:
        value_type = 'text'
        serialized = str(value)
    return {
        'column': cell.column,
        'coordinate': cell.coordinate,
        'type': value_type,
        'value': serialized,
    }


def _json_style(cell):
    color = cell.fill.fgColor
    return {
        'column': cell.column,
        'coordinate': cell.coordinate,
        'style_id': cell.style_id,
        'number_format': cell.number_format,
        'fill': {
            'type': cell.fill.fill_type or '',
            'color_type': color.type or '',
            'rgb': color.rgb if color.type == 'rgb' else '',
            'indexed': color.indexed if color.type == 'indexed' else None,
            'theme': color.theme if color.type == 'theme' else None,
            'tint': color.tint,
        },
    }


def _row_hash(raw_values, raw_styles):
    payload = json.dumps(
        {'values': raw_values, 'styles': raw_styles},
        ensure_ascii=False,
        sort_keys=True,
        separators=(',', ':'),
    ).encode('utf-8')
    return hashlib.sha256(payload).hexdigest()


def _cell_value(row, column):
    if not column:
        return None
    return row[column - 1].value


def _combined_values(row, columns):
    if not columns:
        return ''
    if isinstance(columns, int):
        columns = [columns]
    return '\n'.join(
        value
        for value in (_plain_text(_cell_value(row, column)) for column in columns)
        if value
    )


def _row_color_hint(raw_styles):
    fills = []
    for style in raw_styles:
        fill = style['fill']
        if fill['type'] == 'solid' and any(
            fill.get(key) not in ('', None) for key in ('rgb', 'indexed', 'theme')
        ):
            fills.append(fill)
    if not fills:
        return {}
    unique = []
    for fill in fills:
        if fill not in unique:
            unique.append(fill)
    return {'fills': unique}


def parse_arrival_workbook(payload):
    validate_xlsx_archive(payload)
    try:
        workbook = load_workbook(
            io.BytesIO(payload),
            read_only=False,
            data_only=False,
            keep_links=False,
        )
    except Exception as error:
        raise _unsafe('Книгу не удалось безопасно открыть.', 'xlsx_open_failed') from error

    source_rows = []
    normalized_rows = []
    issues = []
    expected_sheets = PROFILE_CONFIGURATION['sheets']
    actual_sheet_names = set(workbook.sheetnames)

    for missing_sheet in sorted(set(expected_sheets) - actual_sheet_names):
        issues.append(ParsedIssue(
            severity='error',
            code='missing_sheet',
            message='В книге отсутствует обязательный лист.',
            sheet_name=missing_sheet,
        ))
    for extra_sheet in sorted(actual_sheet_names - set(expected_sheets)):
        issues.append(ParsedIssue(
            severity='warning',
            code='unexpected_sheet',
            message='Обнаружен лист вне выбранного профиля.',
            sheet_name=extra_sheet,
        ))

    for sheet_name, sheet_config in expected_sheets.items():
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        max_row = sheet_config['max_row']
        max_column = sheet_config['max_column']
        outside_coordinates = sorted(
            cell.coordinate
            for cell in sheet._cells.values()
            if cell.value not in (None, '')
            and (cell.row > max_row or cell.column > max_column)
        )
        if outside_coordinates:
            issues.append(ParsedIssue(
                severity='error',
                code='content_outside_profile',
                message='За пределами проверяемого диапазона обнаружены данные.',
                sheet_name=sheet_name,
                details={
                    'cell_count': len(outside_coordinates),
                    'sample_coordinates': outside_coordinates[:10],
                },
            ))
        for column, expected_text in sheet_config['headers'].items():
            actual = _text_key(sheet.cell(1, column).value)
            if expected_text not in actual:
                issues.append(ParsedIssue(
                    severity='error',
                    code='invalid_header',
                    message='Заголовок листа не соответствует выбранному профилю.',
                    sheet_name=sheet_name,
                    row_number=1,
                    details={'column': column},
                ))

        for row_number in range(1, max_row + 1):
            row = [sheet.cell(row_number, column) for column in range(1, max_column + 1)]
            if not any(cell.value not in (None, '') for cell in row):
                continue
            raw_values = [_json_cell(cell) for cell in row]
            raw_styles = [_json_style(cell) for cell in row]
            if row_number == 1:
                row_kind = 'header'
            elif sheet_config['kind'] == 'summary':
                row_kind = 'summary'
            else:
                row_kind = 'person'
            source_rows.append(ParsedSourceRow(
                sheet_name=sheet_name,
                row_number=row_number,
                row_kind=row_kind,
                raw_values=raw_values,
                raw_styles=raw_styles,
                row_sha256=_row_hash(raw_values, raw_styles),
            ))
            if row_kind != 'person':
                continue

            content_columns = {
                column
                for value in sheet_config['columns'].values()
                for column in (value if isinstance(value, list) else [value])
                if column
            }
            formula_columns = [cell.column for cell in row if cell.data_type == 'f']
            forbidden_formula_columns = sorted(set(formula_columns).intersection(content_columns))
            if forbidden_formula_columns:
                issues.append(ParsedIssue(
                    severity='error',
                    code='formula_in_content',
                    message='Формула обнаружена в содержательном поле.',
                    sheet_name=sheet_name,
                    row_number=row_number,
                    details={'columns': forbidden_formula_columns},
                ))

            columns = sheet_config['columns']
            raw_name = _plain_text(_cell_value(row, columns.get('name')))
            if not raw_name:
                if any(
                    _cell_value(row, column) not in (None, '')
                    for column in content_columns - {columns.get('name')}
                ):
                    issues.append(ParsedIssue(
                        severity='error',
                        code='missing_name',
                        message='В заполненной строке отсутствует ФИО.',
                        sheet_name=sheet_name,
                        row_number=row_number,
                    ))
                continue

            normalized_name, name_comment = _split_name_and_comment(raw_name)
            if not normalized_name:
                issues.append(ParsedIssue(
                    severity='error',
                    code='invalid_name',
                    message='ФИО не удалось распознать.',
                    sheet_name=sheet_name,
                    row_number=row_number,
                ))
                continue
            raw_date_value = _cell_value(row, columns.get('date'))
            date_candidate, date_comment = _parse_date_candidate(raw_date_value)
            raw_date = _plain_text(raw_date_value)
            if raw_date and date_candidate is None:
                issues.append(ParsedIssue(
                    severity='warning',
                    code='date_requires_review',
                    message='Дата сохранена без догадки и требует проверки.',
                    sheet_name=sheet_name,
                    row_number=row_number,
                ))
            raw_shift_hint = _plain_text(_cell_value(row, columns.get('shift')))
            if raw_shift_hint and raw_shift_hint not in {'1', '2', '1.0', '2.0'}:
                issues.append(ParsedIssue(
                    severity='warning',
                    code='unknown_shift_hint',
                    message='Значение смены сохранено только как подсказка и требует проверки.',
                    sheet_name=sheet_name,
                    row_number=row_number,
                ))
            color_hint = _row_color_hint(raw_styles)
            participation_hint = 'review_required'
            if sheet_config['kind'] == 'primary':
                issues.append(ParsedIssue(
                    severity='error',
                    code='participation_requires_review',
                    message='Участие в заезде требует проверки табельщиком.',
                    sheet_name=sheet_name,
                    row_number=row_number,
                    details={'has_color_hint': bool(color_hint)},
                ))

            position = _combined_values(row, columns.get('position'))
            raw_phone = _combined_values(row, columns.get('phone'))
            normalized_rows.append(ParsedNormalizedRow(
                sheet_name=sheet_name,
                row_number=row_number,
                raw_full_name=raw_name,
                normalized_full_name=normalized_name,
                normalized_name_key=_text_key(normalized_name),
                name_comment=name_comment,
                source_position=position,
                normalized_position_key=_text_key(position),
                raw_shift_hint=raw_shift_hint,
                raw_date=raw_date,
                arrival_date_candidate=date_candidate,
                date_comment=date_comment,
                route_text=_combined_values(row, columns.get('route')),
                raw_phone=raw_phone,
                normalized_phones=_normalized_phones(raw_phone),
                comments=_combined_values(row, columns.get('comments')),
                participation_hint=participation_hint,
                color_hint=color_hint,
            ))

    primary_rows = sum(
        1 for row in normalized_rows if row.sheet_name == PROFILE_CONFIGURATION['primary_sheet']
    )
    if primary_rows == 0:
        issues.append(ParsedIssue(
            severity='error',
            code='empty_primary_sheet',
            message='Основной список сотрудников не содержит строк.',
            sheet_name=PROFILE_CONFIGURATION['primary_sheet'],
        ))

    summary_rows = sum(1 for row in source_rows if row.sheet_name == 'Числ')
    if summary_rows:
        issues.append(ParsedIssue(
            severity='warning',
            code='summary_requires_review',
            message='Сверочный лист сохранён без вычисления формул и требует проверки.',
            sheet_name='Числ',
        ))

    return ParsedArrivalWorkbook(
        source_rows=source_rows,
        normalized_rows=normalized_rows,
        issues=issues,
        workbook_summary={
            'sheet_names': workbook.sheetnames,
            'source_row_count': len(source_rows),
            'normalized_row_count': len(normalized_rows),
            'primary_row_count': primary_rows,
            'summary_row_count': summary_rows,
        },
    )

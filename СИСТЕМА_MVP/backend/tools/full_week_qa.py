#!/usr/bin/env python
"""Изолированный сквозной QA-прогон производственной недели.

Сценарий намеренно запускается только против отдельной локальной PostgreSQL-БД.
Он использует реальные Django URL, middleware, формы и доменные сервисы. ORM
используется для bootstrap первого администратора, чтения созданных объектов и
независимой итоговой сверки.
"""

from __future__ import annotations

import argparse
import hashlib
import io
import json
import os
import sys
import time
import traceback
from collections import Counter, defaultdict
from contextlib import contextmanager
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlparse
from unittest.mock import patch
from zoneinfo import ZoneInfo


BACKEND_DIR = Path(__file__).resolve().parents[1]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")

import django  # noqa: E402

django.setup()

from django.conf import settings  # noqa: E402
from django.core.exceptions import ValidationError  # noqa: E402
from django.db import connection  # noqa: E402
from django.db.models import Count, Max, Sum  # noqa: E402
from django.test import Client  # noqa: E402
from django.utils import timezone  # noqa: E402

from assignments.models import (  # noqa: E402
    AssignmentStatus,
    CrewPlan,
    CrewPlanSlot,
    CrewPlanStatus,
    EquipmentAssignment,
    ExcavatorPlacement,
    HaulAssignment,
    WorkShiftType,
)
from core.production_time import (  # noqa: E402
    BUSINESS_TIME_ZONE,
    production_shift_context,
    production_work_date,
)
from downtimes.models import DowntimeEvent, DowntimeReason  # noqa: E402
from references.models import (  # noqa: E402
    DormitorySection,
    DumpPoint,
    Equipment,
    RockType,
    TruckCapacityRule,
)
from reports.driver_watch_observation import (  # noqa: E402
    build_driver_watch_linkage_audit,
    build_driver_watch_observation,
)
from reports.models import (  # noqa: E402
    DriverShiftPassportCaptureRequest,
    DriverShiftPassportRequestStatus,
    DriverShiftPassportSnapshot,
)
from shifts.models import (  # noqa: E402
    EmployeeShift,
    ShiftClientAction,
    ShiftReadingCorrection,
    WatchPeriod,
)
from trips.models import OPEN_TRIP_STATUSES, Trip, TripClientAction, TripStatus  # noqa: E402
from users.models import (  # noqa: E402
    DriverPrimaryRegistration,
    Employee,
    EmployeeAccess,
    PersonnelDepartment,
    PersonnelPosition,
    ProductionSpecialization,
    Role,
    WatchComposition,
    WorkSchedule,
)


LOCAL_QA_DB_NAME = os.environ.get(
    "WEEK_QA_DB_NAME",
    "copper_week_qa_20260727",
)
LOCAL_QA_DB_HOST = os.environ.get("WEEK_QA_DB_HOST", "127.0.0.1")
LOCAL_QA_DB_PORT = os.environ.get("WEEK_QA_DB_PORT", "55432")
DEFAULT_START_DATE = date(2026, 7, 20)
DEFAULT_RUN_ID = "QA-WEEK-20260727"
DEFAULT_MARKER = "ТЕСТ_НЕДЕЛЯ_20260727"
DEFAULT_ARTIFACT_DIR = Path(
    r"C:\Users\swwba\AppData\Local\Temp\copper-week-qa-20260727"
)

ROLE_HOSTS = {
    "admin": "admin.localhost",
    "oup": "oup.localhost",
    "deputy_mining_manager": "deputy.localhost",
    "dispatcher": "dispatcher.localhost",
    "mining_master": "mining-master.localhost",
    "excavator_operator": "excavator.localhost",
    "driver": "driver.localhost",
    "manager": "management.localhost",
}

# Реалистичный фрагмент четырёхбригадного цикла: две дневные, сутки отдыха,
# две ночные. На горизонте семи суток каждая из четырёх бригад участвует.
DAY_BRIGADES = (1, 1, 4, 4, 2, 2, 3)
NIGHT_BRIGADES = (3, 3, 1, 1, 4, 4, 2)

POSITION_HINTS = {
    "oup": ("ведущий специалист", "специалист"),
    "deputy_mining_manager": ("заместитель начальника участка",),
    "dispatcher": ("горный диспетчер",),
    "mining_master": ("горный мастер",),
    "manager": ("начальник участка", "начальник отдела"),
    "driver": ("водитель автомобиля, занятый на транспортировании горной массы",),
    "excavator_operator": ("машинист экскаватора",),
}

DEPARTMENT_HINTS = {
    "oup": ("отдел управления персоналом",),
    "deputy_mining_manager": ("горный участок", "участок"),
    "dispatcher": ("диспетчерская служба",),
    "mining_master": ("горный участок", "участок"),
    "manager": ("горный участок", "участок"),
    "driver": ("горный участок", "участок"),
    "excavator_operator": ("горный участок", "участок"),
}


class QAError(RuntimeError):
    """Проверяемая ошибка подготовки или исполнения QA-сценария."""


@dataclass(frozen=True)
class RunConfig:
    run_id: str = DEFAULT_RUN_ID
    marker: str = DEFAULT_MARKER
    start_date: date = DEFAULT_START_DATE
    artifact_dir: Path = DEFAULT_ARTIFACT_DIR
    expected_trucks: int = 53
    expected_excavators: int = 8

    @property
    def end_date(self) -> date:
        return self.start_date + timedelta(days=6)


@dataclass
class StaffMember:
    employee_id: int
    access_id: int
    role_code: str
    phone: str
    permanent_pin: str
    brigade: int | None
    ordinal: int
    equipment_id: int | None = None
    client: "RoleHttpClient | None" = None


@dataclass
class ShiftResult:
    index: int
    production_date: str
    shift_type: str
    driver_shifts: int = 0
    operator_shifts: int = 0
    loaded_trips: int = 0
    unloaded_trips: int = 0
    carryover_in: int = 0
    carryover_out: int = 0
    trip_counts_by_truck: dict[str, int] = field(default_factory=dict)
    volume_m3: str = "0.00"
    tonnage: str = "0.00"
    duration_seconds: float = 0.0


def q2(value: Decimal | int | str) -> Decimal:
    return Decimal(value).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def json_default(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, Path):
        return str(value)
    raise TypeError(f"Unsupported JSON value: {type(value)!r}")


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=json_default) + "\n",
        encoding="utf-8",
    )


def append_jsonl(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8", newline="\n") as stream:
        stream.write(json.dumps(payload, ensure_ascii=False, default=json_default))
        stream.write("\n")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


@contextmanager
def at_time(value: datetime):
    """Freeze application-level Django time at one Vladivostok instant."""

    if value.tzinfo is None:
        value = value.replace(tzinfo=BUSINESS_TIME_ZONE)
    with patch("django.utils.timezone.now", return_value=value):
        yield value


def local_dt(day: date, hour: int, minute: int = 0, second: int = 0) -> datetime:
    return datetime(
        day.year,
        day.month,
        day.day,
        hour,
        minute,
        second,
        tzinfo=BUSINESS_TIME_ZONE,
    )


def verify_isolated_database(config: RunConfig, *, require_empty: bool) -> None:
    db = settings.DATABASES["default"]
    engine = str(db.get("ENGINE") or "")
    name = str(db.get("NAME") or "")
    host = str(db.get("HOST") or "")
    port = str(db.get("PORT") or "")
    if not engine.endswith("postgresql"):
        raise QAError("QA-прогон разрешён только на PostgreSQL.")
    if (name, host, port) != (
        LOCAL_QA_DB_NAME,
        LOCAL_QA_DB_HOST,
        LOCAL_QA_DB_PORT,
    ):
        raise QAError(
            "Защитная остановка: разрешена только отдельная локальная БД "
            f"{LOCAL_QA_DB_NAME}@{LOCAL_QA_DB_HOST}:{LOCAL_QA_DB_PORT}; "
            f"получено {name}@{host}:{port}."
        )
    with connection.cursor() as cursor:
        cursor.execute(
            "select current_database(), inet_server_addr()::text, "
            "inet_server_port()::text"
        )
        actual_name, actual_host, actual_port = cursor.fetchone()
    if actual_name != LOCAL_QA_DB_NAME or actual_port != LOCAL_QA_DB_PORT:
        raise QAError(
            "Защитная остановка: фактическое соединение не соответствует "
            "изолированной QA-БД."
        )
    if actual_host not in {LOCAL_QA_DB_HOST, "127.0.0.1/32", "::1"}:
        raise QAError(
            f"Защитная остановка: сервер БД не локальный ({actual_host})."
        )
    if require_empty:
        business_counts = {
            "employees": Employee.objects.count(),
            "shifts": EmployeeShift.objects.count(),
            "watch_compositions": WatchComposition.objects.count(),
            "watch_periods": WatchPeriod.objects.count(),
            "trips": Trip.objects.count(),
            "haul_assignments": HaulAssignment.objects.count(),
        }
        if any(business_counts.values()):
            raise QAError(
                "Изолированная QA-БД уже содержит рабочие записи. "
                f"Нужна новая чистая БД: {business_counts}."
            )
    config.artifact_dir.mkdir(parents=True, exist_ok=True)


class RoleHttpClient:
    """CSRF-strict Django client bound to one role host and one employee."""

    def __init__(self, role_code: str):
        try:
            self.host = ROLE_HOSTS[role_code]
        except KeyError as error:
            raise QAError(f"Неизвестный ролевой host: {role_code}") from error
        self.role_code = role_code
        self.client = Client(enforce_csrf_checks=True)

    def _csrf_token(self) -> str:
        cookie = self.client.cookies.get(settings.CSRF_COOKIE_NAME)
        if cookie:
            return cookie.value
        response = self.client.get("/", HTTP_HOST=self.host)
        if response.status_code not in {200, 302}:
            raise QAError(
                f"{self.role_code}: GET / для CSRF вернул {response.status_code}."
            )
        cookie = self.client.cookies.get(settings.CSRF_COOKIE_NAME)
        if not cookie:
            raise QAError(f"{self.role_code}: сервер не выдал CSRF cookie.")
        return cookie.value

    @staticmethod
    def _response_excerpt(response) -> str:
        try:
            raw = response.content.decode("utf-8", errors="replace")
        except Exception:
            return ""
        return " ".join(raw.split())[:800]

    def get(
        self,
        path: str,
        *,
        expected: Iterable[int] = (200,),
        label: str = "",
        **extra,
    ):
        response = self.client.get(path, HTTP_HOST=self.host, **extra)
        if response.status_code not in set(expected):
            raise QAError(
                f"{label or path}: GET вернул {response.status_code}; "
                f"{self._response_excerpt(response)}"
            )
        return response

    def post_form(
        self,
        path: str,
        data: dict[str, Any],
        *,
        expected: Iterable[int] = (302,),
        label: str = "",
        ajax: bool = False,
        **extra,
    ):
        payload = dict(data)
        payload.setdefault("csrfmiddlewaretoken", self._csrf_token())
        if ajax:
            extra.setdefault("HTTP_X_REQUESTED_WITH", "XMLHttpRequest")
        response = self.client.post(
            path,
            payload,
            HTTP_HOST=self.host,
            **extra,
        )
        if response.status_code not in set(expected):
            raise QAError(
                f"{label or path}: POST form вернул {response.status_code}; "
                f"{self._response_excerpt(response)}"
            )
        return response

    def post_json(
        self,
        path: str,
        payload: dict[str, Any],
        *,
        expected: Iterable[int] = (200,),
        label: str = "",
        **extra,
    ):
        response = self.client.post(
            path,
            data=json.dumps(payload, ensure_ascii=False),
            content_type="application/json",
            HTTP_HOST=self.host,
            HTTP_X_CSRFTOKEN=self._csrf_token(),
            **extra,
        )
        if response.status_code not in set(expected):
            raise QAError(
                f"{label or path}: POST JSON вернул {response.status_code}; "
                f"{self._response_excerpt(response)}"
            )
        try:
            body = response.json()
        except Exception as error:
            raise QAError(f"{label or path}: ответ не является JSON.") from error
        if not body.get("ok", False):
            raise QAError(f"{label or path}: сервер вернул ошибку {body!r}.")
        return response, body

    def post_json_raw(
        self,
        path: str,
        payload: dict[str, Any],
        *,
        expected: Iterable[int],
        label: str = "",
        **extra,
    ):
        response = self.client.post(
            path,
            data=json.dumps(payload, ensure_ascii=False),
            content_type="application/json",
            HTTP_HOST=self.host,
            HTTP_X_CSRFTOKEN=self._csrf_token(),
            **extra,
        )
        if response.status_code not in set(expected):
            raise QAError(
                f"{label or path}: POST JSON вернул {response.status_code}; "
                f"{self._response_excerpt(response)}"
            )
        try:
            return response, response.json()
        except Exception as error:
            raise QAError(f"{label or path}: ответ не является JSON.") from error

    def login(self, phone: str, pin: str, *, device_kind: str = "personal"):
        response = self.post_form(
            "/",
            {
                "phone": phone,
                "access_code": pin,
                "device_kind": device_kind,
            },
            expected=(302,),
            label=f"{self.role_code}: вход",
        )
        return response

    def activate(
        self,
        phone: str,
        primary_pin: str,
        permanent_pin: str,
        *,
        device_kind: str = "personal",
    ):
        login_response = self.login(
            phone,
            primary_pin,
            device_kind=device_kind,
        )
        location = urlparse(login_response["Location"]).path
        if location != "/activate-access/":
            raise QAError(
                f"{self.role_code}: первичный PIN не направил на активацию "
                f"(Location={login_response['Location']!r})."
            )
        self.get(
            "/activate-access/",
            label=f"{self.role_code}: экран активации",
        )
        response = self.post_form(
            "/activate-access/",
            {
                "new_access_code": permanent_pin,
                "confirm_access_code": permanent_pin,
            },
            expected=(302,),
            label=f"{self.role_code}: постоянный PIN",
        )
        if urlparse(response["Location"]).path not in {"/home/", "/"}:
            raise QAError(
                f"{self.role_code}: неожиданный переход после активации "
                f"{response['Location']!r}."
            )
        return response


class ReferenceCatalog:
    def __init__(self, config: RunConfig):
        self.config = config
        self.roles = {
            role.code: role
            for role in Role.objects.filter(
                code__in=ROLE_HOSTS,
                is_active=True,
            )
        }
        missing_roles = sorted(set(ROLE_HOSTS) - set(self.roles))
        if missing_roles:
            raise QAError(f"Нет активных ролей: {missing_roles}.")

        self.trucks = list(
            Equipment.objects.filter(
                is_active=True,
                equipment_type__is_active=True,
                equipment_type__name__iexact="Самосвал",
            )
            .select_related("equipment_type", "model")
            .order_by("garage_number")
        )
        self.excavators = list(
            Equipment.objects.filter(
                is_active=True,
                equipment_type__is_active=True,
                equipment_type__name__iexact="Экскаватор",
            )
            .select_related("equipment_type", "model")
            .order_by("garage_number")
        )
        if len(self.trucks) != config.expected_trucks:
            raise QAError(
                f"Ожидалось {config.expected_trucks} активных самосвалов, "
                f"в справочнике {len(self.trucks)}."
            )
        if len(self.excavators) != config.expected_excavators:
            raise QAError(
                f"Ожидалось {config.expected_excavators} активных экскаваторов, "
                f"в справочнике {len(self.excavators)}."
            )

        self.rocks = list(RockType.objects.filter(is_active=True).order_by("name"))
        self.dump_points = list(
            DumpPoint.objects.filter(is_active=True).order_by("name")
        )
        self.dormitory_sections = list(
            DormitorySection.objects.select_related("block__dormitory").order_by(
                "block__dormitory__number",
                "block__name",
                "name",
            )
        )
        if not self.rocks:
            raise QAError("Справочник активных пород пуст.")
        if not self.dump_points:
            raise QAError("Справочник активных точек разгрузки пуст.")
        if not self.dormitory_sections:
            raise QAError("Справочник секций общежития пуст.")

        self.shift_schedule = (
            WorkSchedule.objects.filter(is_active=True, brigade_count=4)
            .order_by("id")
            .first()
        )
        self.individual_schedule = (
            WorkSchedule.objects.filter(is_active=True, brigade_count=0)
            .order_by("id")
            .first()
        )
        if not self.shift_schedule:
            raise QAError("Нет активного четырёхбригадного графика.")
        if not self.individual_schedule:
            raise QAError("Нет активного индивидуального графика.")

        self.specializations = {}
        for role_code in ("driver", "excavator_operator"):
            specialization = (
                ProductionSpecialization.objects.filter(
                    is_active=True,
                    access_role__code=role_code,
                    access_role__is_active=True,
                )
                .select_related("access_role", "equipment_type")
                .order_by("id")
                .first()
            )
            if not specialization:
                raise QAError(
                    f"Нет активной производственной специализации {role_code}."
                )
            self.specializations[role_code] = specialization

        self.positions = {
            code: self._position_for_role(code)
            for code in POSITION_HINTS
        }
        self.departments = {
            code: self._department_for_role(code)
            for code in DEPARTMENT_HINTS
        }
        self.operational_rocks: list[RockType] = []
        self.unresolved_capacity_pairs: list[dict[str, Any]] = []
        self._validate_equipment_measurements()

    @staticmethod
    def _match_by_hints(items, hints: tuple[str, ...], attr: str = "name"):
        for hint in hints:
            hint_folded = hint.casefold()
            for item in items:
                if hint_folded in str(getattr(item, attr, "") or "").casefold():
                    return item
        return None

    def _position_for_role(self, role_code: str) -> PersonnelPosition:
        items = list(
            PersonnelPosition.objects.filter(is_active=True)
            .prefetch_related("allowed_specializations")
            .order_by("id")
        )
        position = self._match_by_hints(items, POSITION_HINTS[role_code])
        if not position:
            raise QAError(
                f"В реальном справочнике не найдена кадровая должность для "
                f"{role_code}; подсказки={POSITION_HINTS[role_code]}."
            )
        specialization = self.specializations.get(role_code)
        if specialization and not position.allowed_specializations.filter(
            pk=specialization.pk
        ).exists():
            raise QAError(
                f"Должность «{position.name}» не разрешает специализацию "
                f"«{specialization.name}»."
            )
        if not specialization and position.default_specialization_id:
            default_role = getattr(
                getattr(position.default_specialization, "access_role", None),
                "code",
                "",
            )
            if default_role and default_role != role_code:
                raise QAError(
                    f"Должность «{position.name}» конфликтует с ролью {role_code}."
                )
        return position

    def _department_for_role(self, role_code: str) -> PersonnelDepartment:
        items = list(
            PersonnelDepartment.objects.filter(is_active=True).order_by("id")
        )
        department = self._match_by_hints(items, DEPARTMENT_HINTS[role_code])
        if not department:
            raise QAError(
                f"В реальном справочнике не найдено подразделение для "
                f"{role_code}; подсказки={DEPARTMENT_HINTS[role_code]}."
            )
        return department

    def _validate_equipment_measurements(self) -> None:
        invalid = []
        for equipment in [*self.trucks, *self.excavators]:
            if not equipment.model_id:
                invalid.append(f"{equipment.garage_number}: нет модели")
                continue
            if not equipment.model.fuel_capacity_limit_l:
                invalid.append(
                    f"{equipment.garage_number}: нет лимита топлива у "
                    f"{equipment.model.name}"
                )
        if invalid:
            raise QAError(
                "Нельзя штатно открыть/закрыть смены из-за справочника: "
                + "; ".join(invalid[:20])
            )

        unresolved_by_rock: dict[int, list[Equipment]] = defaultdict(list)
        for rock in self.rocks:
            for truck in self.trucks:
                has_rule = TruckCapacityRule.objects.filter(
                    equipment_model=truck.model,
                    rock_type=rock,
                ).exists()
                if not has_rule and not truck.model.body_volume_m3:
                    unresolved_by_rock[rock.id].append(truck)
                    self.unresolved_capacity_pairs.append(
                        {
                            "truck_id": truck.id,
                            "truck": truck.garage_number,
                            "model_id": truck.model_id,
                            "model": truck.model.name,
                            "rock_type_id": rock.id,
                            "rock_type": rock.name,
                        }
                    )
        self.operational_rocks = [
            rock
            for rock in self.rocks
            if rock.id not in unresolved_by_rock
        ]
        if not self.operational_rocks:
            raise QAError(
                "Ни одна активная порода не имеет кубатуры для всех "
                "действующих моделей самосвалов."
            )

    def trip_measurements(
        self,
        truck: Equipment,
        rock: RockType,
    ) -> tuple[Decimal, Decimal | None]:
        rule = TruckCapacityRule.objects.filter(
            equipment_model=truck.model,
            rock_type=rock,
        ).first()
        volume = rule.volume_m3 if rule else truck.model.body_volume_m3
        if volume is None:
            raise QAError(
                f"Нет кубатуры для {truck.garage_number}/{rock.name}."
            )
        tonnage = q2(Decimal(volume) * Decimal(rock.density)) if rock.density else None
        return q2(volume), tonnage


def base_employee_payload(
    catalog: ReferenceCatalog,
    *,
    role_code: str,
    full_name: str,
    phone: str,
    personnel_number: str,
    hired_at: date,
    brigade: int | None,
    watch_composition: WatchComposition | None = None,
) -> dict[str, Any]:
    specialization = catalog.specializations.get(role_code)
    schedule = (
        catalog.shift_schedule
        if brigade is not None
        else catalog.individual_schedule
    )
    return {
        "full_name": full_name,
        "birth_date": "1985-01-01",
        "personnel_number": personnel_number,
        "phone": phone,
        "personnel_position": str(catalog.positions[role_code].id),
        "base_specialization": str(specialization.id) if specialization else "",
        "position": "",
        "personnel_department": str(catalog.departments[role_code].id),
        "work_category": "",
        "hired_at": hired_at.isoformat(),
        "dismissed_at": "",
        "work_schedule": str(schedule.id),
        "brigade_number": str(brigade or ""),
        "watch_composition": (
            str(watch_composition.id)
            if watch_composition
            else ""
        ),
        "residence_text": "",
        "comment": f"{catalog.config.marker}: изолированный недельный QA",
        "hr_data": "",
    }


def permanent_pin_for(ordinal: int, primary_pin: str = "") -> str:
    candidate = 630000 + ordinal
    while True:
        pin = f"{candidate:06d}"
        weak = len(set(pin)) == 1 or pin in {
            "012345",
            "123456",
            "234567",
            "345678",
            "456789",
            "987654",
            "876543",
            "765432",
            "654321",
            "543210",
        }
        if not weak and pin != primary_pin:
            return pin
        candidate += 1


def phone_for(ordinal: int) -> str:
    return f"+7999200{ordinal:04d}"


ROLE_NAME_TOKENS = {
    "admin": "АДМИНИСТРАТОР",
    "oup": "ОУП",
    "deputy_mining_manager": "ЗАМЕСТИТЕЛЬ",
    "dispatcher": "ДИСПЕТЧЕР",
    "mining_master": "ГОРНЫЙ_МАСТЕР",
    "excavator_operator": "МАШИНИСТ",
    "driver": "ВОДИТЕЛЬ",
    "manager": "РУКОВОДИТЕЛЬ",
}


class WeekOnboarding:
    def __init__(self, config: RunConfig, catalog: ReferenceCatalog):
        self.config = config
        self.catalog = catalog
        self.watch_composition = WatchComposition.objects.create(
            code=f"qa-week-kpi-{config.start_date:%Y%m%d}",
            name=f"{config.marker} УТВЕРЖДЁННЫЙ СОСТАВ",
            is_active=True,
        )
        self.watch_period = WatchPeriod.objects.create(
            name=f"{config.marker} КАЛЕНДАРНАЯ ВАХТА",
            watch_composition=self.watch_composition,
            starts_on=config.start_date,
            ends_on=config.end_date,
            is_active=True,
        )
        self.staff: list[StaffMember] = []
        self.by_role: dict[str, list[StaffMember]] = defaultdict(list)
        self.drivers_by_brigade: dict[int, dict[int, StaffMember]] = {
            brigade: {} for brigade in range(1, 5)
        }
        self.operators_by_brigade: dict[int, dict[int, StaffMember]] = {
            brigade: {} for brigade in range(1, 5)
        }
        self.shift_roles_by_brigade: dict[
            str, dict[int, StaffMember]
        ] = {
            "dispatcher": {},
            "mining_master": {},
        }
        self.next_ordinal = 0
        self.admin: StaffMember | None = None
        self.oup: StaffMember | None = None
        self.deputy: StaffMember | None = None
        self.manager: StaffMember | None = None

    def _new_ordinal(self) -> int:
        ordinal = self.next_ordinal
        self.next_ordinal += 1
        return ordinal

    def _register_member(self, member: StaffMember) -> StaffMember:
        self.staff.append(member)
        self.by_role[member.role_code].append(member)
        return member

    def _full_name(
        self,
        role_code: str,
        *,
        brigade: int | None = None,
        equipment: Equipment | None = None,
        ordinal: int,
    ) -> str:
        parts = [self.config.marker, ROLE_NAME_TOKENS[role_code]]
        if brigade is not None:
            parts.append(f"БРИГАДА_{brigade}")
        if equipment is not None:
            parts.append(f"ТЕХНИКА_{equipment.garage_number}")
        parts.append(f"N{ordinal:03d}")
        return " ".join(parts)

    def _personnel_number(
        self,
        role_code: str,
        *,
        brigade: int | None,
        ordinal: int,
    ) -> str:
        short_role = {
            "oup": "OUP",
            "deputy_mining_manager": "DEP",
            "dispatcher": "DSP",
            "mining_master": "MM",
            "excavator_operator": "EO",
            "driver": "DRV",
            "manager": "MGR",
        }[role_code]
        brigade_part = f"B{brigade}" if brigade is not None else "BI"
        return f"QAWEEK-20260727-{short_role}-{brigade_part}-{ordinal:03d}"

    @staticmethod
    def _created_access(employee: Employee, role_code: str) -> EmployeeAccess:
        try:
            return employee.accesses.select_related("role").get(
                role__code=role_code,
                is_active=True,
            )
        except EmployeeAccess.DoesNotExist as error:
            raise QAError(
                f"После создания {employee.full_name} не найден доступ "
                f"{role_code}."
            ) from error

    def bootstrap_admin(self, when: datetime) -> StaffMember:
        ordinal = self._new_ordinal()
        phone = phone_for(ordinal)
        permanent_pin = permanent_pin_for(ordinal)
        with at_time(when):
            employee = Employee.objects.create(
                full_name=self._full_name("admin", ordinal=ordinal),
                phone=phone,
                personnel_number=(
                    f"QAWEEK-20260727-ADM-BI-{ordinal:03d}"
                ),
                status=Employee.Status.ACTIVE,
                is_active=True,
                hired_at=self.config.start_date,
                watch_composition=self.watch_composition,
                comment=f"{self.config.marker}: служебный bootstrap",
            )
            access = EmployeeAccess.objects.create(
                employee=employee,
                role=self.catalog.roles["admin"],
                access_code=permanent_pin,
                status=EmployeeAccess.Status.ACTIVATED,
                is_active=True,
                activated_at=when,
            )
        client = RoleHttpClient("admin")
        response = client.login(phone, permanent_pin)
        if urlparse(response["Location"]).path not in {"/home/", "/"}:
            raise QAError(
                f"Bootstrap-администратор не вошёл: {response['Location']!r}."
            )
        member = self._register_member(
            StaffMember(
                employee_id=employee.id,
                access_id=access.id,
                role_code="admin",
                phone=phone,
                permanent_pin=permanent_pin,
                brigade=None,
                ordinal=ordinal,
                client=client,
            )
        )
        self.admin = member
        return member

    def create_oup(self, when: datetime) -> StaffMember:
        if not self.admin or not self.admin.client:
            raise QAError("Сначала нужен bootstrap-администратор.")
        ordinal = self._new_ordinal()
        phone = phone_for(ordinal)
        payload = base_employee_payload(
            self.catalog,
            role_code="oup",
            full_name=self._full_name("oup", ordinal=ordinal),
            phone=phone,
            personnel_number=self._personnel_number(
                "oup",
                brigade=None,
                ordinal=ordinal,
            ),
            hired_at=self.config.start_date,
            brigade=None,
            watch_composition=self.watch_composition,
        )
        payload.update(
            {
                "role": str(self.catalog.roles["oup"].id),
                "generate_access": "on",
                "assignment_shift_type": "",
                "assignment_equipment": "",
            }
        )
        with at_time(when):
            response = self.admin.client.post_form(
                "/system-admin/employees/create/",
                payload,
                label="Администратор создаёт ОУП",
            )
            location = urlparse(response["Location"]).path
            if not location.startswith("/system-admin/employees/"):
                raise QAError(
                    f"Неожиданный redirect создания ОУП: {location!r}."
                )
            self.admin.client.get(
                location,
                label="Карточка созданного ОУП",
            )
            employee = Employee.objects.get(phone=phone)
            access = self._created_access(employee, "oup")
            if (
                access.status != EmployeeAccess.Status.NOT_ACTIVATED
                or not access.primary_code_issued_at
                or len(access.access_code) != 6
            ):
                raise QAError("Администратор не выдал корректный первичный PIN ОУП.")
            permanent_pin = permanent_pin_for(ordinal, access.access_code)
            client = RoleHttpClient("oup")
            client.activate(phone, access.access_code, permanent_pin)
            access.refresh_from_db()
        member = self._register_member(
            StaffMember(
                employee_id=employee.id,
                access_id=access.id,
                role_code="oup",
                phone=phone,
                permanent_pin=permanent_pin,
                brigade=None,
                ordinal=ordinal,
                client=client,
            )
        )
        self.oup = member
        return member

    def start_oup_period(self, when: datetime) -> None:
        if not self.oup or not self.oup.client:
            raise QAError("ОУП ещё не создан.")
        with at_time(when):
            self.oup.client.post_form(
                "/oup/shift/start/",
                {"next": "/oup/employees/"},
                label="ОУП начинает кадровый период",
            )
        open_periods = EmployeeShift.objects.filter(
            employee_id=self.oup.employee_id,
            closed_at__isnull=True,
            workplace_code="oup",
        ).count()
        if open_periods != 1:
            raise QAError(
                f"После старта ОУП открыто периодов: {open_periods}, ожидался 1."
            )

    def close_oup_period(self, when: datetime) -> None:
        if not self.oup or not self.oup.client:
            raise QAError("ОУП ещё не создан.")
        with at_time(when):
            self.oup.client.post_form(
                "/oup/shift/close/",
                {"next": "/oup/employees/"},
                label="ОУП завершает кадровый период",
            )
        if EmployeeShift.objects.filter(
            employee_id=self.oup.employee_id,
            closed_at__isnull=True,
            workplace_code="oup",
        ).exists():
            raise QAError("Кадровый период ОУП не завершился.")

    def create_employee(
        self,
        *,
        role_code: str,
        when: datetime,
        brigade: int | None = None,
        equipment: Equipment | None = None,
    ) -> StaffMember:
        if not self.oup or not self.oup.client:
            raise QAError("Сотрудников должен создавать авторизованный ОУП.")
        ordinal = self._new_ordinal()
        phone = phone_for(ordinal)
        payload = base_employee_payload(
            self.catalog,
            role_code=role_code,
            full_name=self._full_name(
                role_code,
                brigade=brigade,
                equipment=equipment,
                ordinal=ordinal,
            ),
            phone=phone,
            personnel_number=self._personnel_number(
                role_code,
                brigade=brigade,
                ordinal=ordinal,
            ),
            hired_at=self.config.start_date,
            brigade=brigade,
            watch_composition=self.watch_composition,
        )
        payload.update(
            {
                "issue_access": "on",
                "access_role": str(self.catalog.roles[role_code].id),
            }
        )
        with at_time(when):
            response = self.oup.client.post_form(
                "/oup/employees/new/",
                payload,
                label=f"ОУП создаёт {role_code} N{ordinal}",
            )
            location = urlparse(response["Location"]).path
            if not location.startswith("/oup/employees/"):
                raise QAError(
                    f"Неожиданный redirect создания {role_code}: {location!r}."
                )
            self.oup.client.get(
                location,
                label=f"Карточка {role_code} N{ordinal}",
            )
            employee = Employee.objects.get(phone=phone)
            access = self._created_access(employee, role_code)
            if (
                access.status != EmployeeAccess.Status.NOT_ACTIVATED
                or not access.primary_code_issued_at
                or len(access.access_code) != 6
            ):
                raise QAError(
                    f"ОУП не выдал корректный первичный PIN {role_code} N{ordinal}."
                )
            permanent_pin = permanent_pin_for(ordinal, access.access_code)
            client = RoleHttpClient(role_code)
            client.activate(
                phone,
                access.access_code,
                permanent_pin,
                device_kind=(
                    "shared"
                    if role_code in {"dispatcher", "mining_master"}
                    else "personal"
                ),
            )
            access.refresh_from_db()
            if access.status != EmployeeAccess.Status.ACTIVATED:
                raise QAError(f"Доступ {role_code} N{ordinal} не активирован.")
            if role_code == "driver":
                section = self.catalog.dormitory_sections[
                    ordinal % len(self.catalog.dormitory_sections)
                ]
                registration_response = client.post_form(
                    "/driver/registration/",
                    {"dormitory_section": str(section.id)},
                    label=f"Регистрация водителя N{ordinal}",
                )
                if urlparse(registration_response["Location"]).path != "/home/":
                    raise QAError(
                        f"Водитель N{ordinal}: неожиданный redirect регистрации."
                    )
        member = self._register_member(
            StaffMember(
                employee_id=employee.id,
                access_id=access.id,
                role_code=role_code,
                phone=phone,
                permanent_pin=permanent_pin,
                brigade=brigade,
                ordinal=ordinal,
                equipment_id=equipment.id if equipment else None,
                client=client,
            )
        )
        return member

    def run(self) -> "WeekOnboarding":
        onboarding_time = local_dt(self.config.start_date, 6, 0)
        self.bootstrap_admin(onboarding_time)
        self.create_oup(onboarding_time + timedelta(minutes=1))
        self.start_oup_period(onboarding_time + timedelta(minutes=2))

        self.deputy = self.create_employee(
            role_code="deputy_mining_manager",
            when=onboarding_time + timedelta(minutes=3),
        )
        self.manager = self.create_employee(
            role_code="manager",
            when=onboarding_time + timedelta(minutes=4),
        )

        creation_minute = 5
        for brigade in range(1, 5):
            dispatcher = self.create_employee(
                role_code="dispatcher",
                brigade=brigade,
                when=onboarding_time + timedelta(
                    minutes=creation_minute,
                    seconds=brigade,
                ),
            )
            mining_master = self.create_employee(
                role_code="mining_master",
                brigade=brigade,
                when=onboarding_time + timedelta(
                    minutes=creation_minute,
                    seconds=10 + brigade,
                ),
            )
            self.shift_roles_by_brigade["dispatcher"][brigade] = dispatcher
            self.shift_roles_by_brigade["mining_master"][
                brigade
            ] = mining_master
            creation_minute += 1

        for brigade in range(1, 5):
            for index, excavator in enumerate(self.catalog.excavators):
                member = self.create_employee(
                    role_code="excavator_operator",
                    brigade=brigade,
                    equipment=excavator,
                    when=onboarding_time
                    + timedelta(
                        minutes=creation_minute,
                        seconds=index,
                    ),
                )
                self.operators_by_brigade[brigade][excavator.id] = member
            creation_minute += 1

        for brigade in range(1, 5):
            for index, truck in enumerate(self.catalog.trucks):
                member = self.create_employee(
                    role_code="driver",
                    brigade=brigade,
                    equipment=truck,
                    when=onboarding_time
                    + timedelta(
                        minutes=creation_minute,
                        seconds=index,
                    ),
                )
                self.drivers_by_brigade[brigade][truck.id] = member
            creation_minute += 1

        self.close_oup_period(
            onboarding_time + timedelta(minutes=creation_minute + 1)
        )
        self._verify()
        return self

    def _verify(self) -> None:
        expected_by_role = {
            "admin": 1,
            "oup": 1,
            "deputy_mining_manager": 1,
            "manager": 1,
            "dispatcher": 4,
            "mining_master": 4,
            "excavator_operator": len(self.catalog.excavators) * 4,
            "driver": len(self.catalog.trucks) * 4,
        }
        actual_by_role = {
            role_code: len(self.by_role.get(role_code, []))
            for role_code in expected_by_role
        }
        if actual_by_role != expected_by_role:
            raise QAError(
                f"Состав тестового штата неверен: {actual_by_role}, "
                f"ожидался {expected_by_role}."
            )
        marker_count = Employee.objects.filter(
            full_name__startswith=self.config.marker
        ).count()
        expected_total = sum(expected_by_role.values())
        if marker_count != expected_total or len(self.staff) != expected_total:
            raise QAError(
                f"Маркированных сотрудников {marker_count}, "
                f"в памяти {len(self.staff)}, ожидалось {expected_total}."
            )
        composition_members = Employee.objects.filter(
            full_name__startswith=self.config.marker,
            watch_composition=self.watch_composition,
        ).count()
        if composition_members != expected_total:
            raise QAError(
                f"В утверждённом составе {composition_members} сотрудников, "
                f"ожидалось {expected_total}."
            )
        activated = EmployeeAccess.objects.filter(
            employee__full_name__startswith=self.config.marker,
            status=EmployeeAccess.Status.ACTIVATED,
            is_active=True,
        ).count()
        if activated != expected_total:
            raise QAError(
                f"Активировано доступов {activated}, ожидалось {expected_total}."
            )
        registrations = DriverPrimaryRegistration.objects.filter(
            employee__full_name__startswith=self.config.marker
        ).count()
        if registrations != expected_by_role["driver"]:
            raise QAError(
                f"Регистраций водителей {registrations}, "
                f"ожидалось {expected_by_role['driver']}."
            )


class FullWeekRunner:
    def __init__(
        self,
        config: RunConfig,
        catalog: ReferenceCatalog,
        onboarding: WeekOnboarding,
    ):
        self.config = config
        self.catalog = catalog
        self.onboarding = onboarding
        if not onboarding.deputy or not onboarding.deputy.client:
            raise QAError("Нет авторизованного заместителя для расстановки.")
        self.deputy = onboarding.deputy
        self.manifest_path = config.artifact_dir / "trip_manifest.jsonl"
        self.action_log_path = config.artifact_dir / "action_log.jsonl"
        self.shift_results: list[ShiftResult] = []
        self.current_haul_map: dict[int, int] = {}
        self.carryover_trip_id: int | None = None
        self.carryover_truck_id: int | None = None
        self.transferred_downtime_ids: dict[str, int] = {}
        self.eto_to_checks: list[dict[str, Any]] = []
        self.stale_version_checked = False
        self.idempotency_load_checks = 0
        self.idempotency_unload_checks = 0
        self.idempotency_shift_checks = 0
        self.total_load_events = 0
        self.total_unload_events = 0
        self.reference_capacity_probe: dict[str, Any] | None = None
        self.reference_context_by_excavator = self._reference_work_contexts()
        self.manifest_path.unlink(missing_ok=True)
        self.action_log_path.unlink(missing_ok=True)

    def _reference_work_contexts(self) -> dict[int, dict[str, Any]]:
        contexts = {}
        placements = {
            placement.excavator_id: placement
            for placement in ExcavatorPlacement.objects.filter(
                excavator_id__in=[item.id for item in self.catalog.excavators]
            ).select_related("work_rock_type", "work_dump_point")
        }
        for excavator in self.catalog.excavators:
            placement = placements.get(excavator.id)
            contexts[excavator.id] = {
                "loading_horizon": (
                    placement.loading_horizon if placement else ""
                ),
                "loading_block": placement.loading_block if placement else "",
                "source_rock_type_id": (
                    placement.work_rock_type_id if placement else None
                ),
                "source_dump_point_id": (
                    placement.work_dump_point_id if placement else None
                ),
            }
        return contexts

    @staticmethod
    def shift_type_for_index(shift_index: int) -> str:
        return (
            WorkShiftType.SHIFT_1
            if shift_index % 2 == 0
            else WorkShiftType.SHIFT_2
        )

    def date_for_shift(self, shift_index: int) -> date:
        return self.config.start_date + timedelta(days=shift_index // 2)

    def brigade_for_shift(self, shift_index: int) -> int:
        day_index = shift_index // 2
        if self.shift_type_for_index(shift_index) == WorkShiftType.SHIFT_1:
            return DAY_BRIGADES[day_index]
        return NIGHT_BRIGADES[day_index]

    def shift_bounds(self, shift_index: int) -> tuple[datetime, datetime]:
        production_date = self.date_for_shift(shift_index)
        if self.shift_type_for_index(shift_index) == WorkShiftType.SHIFT_1:
            return (
                local_dt(production_date, 7, 5),
                local_dt(production_date, 18, 55),
            )
        return (
            local_dt(production_date, 19, 5),
            local_dt(production_date + timedelta(days=1), 6, 55),
        )

    def target_trip_count(self, truck_index: int, shift_index: int) -> int:
        return 17 + ((truck_index + shift_index) % 7)

    def carry_truck_for_boundary(self, shift_index: int) -> Equipment:
        if not 0 <= shift_index < 13:
            raise QAError("Переходящий рейс возможен только между сменами.")
        start_index = (shift_index * 7 + 5) % len(self.catalog.trucks)
        for offset in range(len(self.catalog.trucks)):
            truck_index = (start_index + offset) % len(self.catalog.trucks)
            current_target = self.target_trip_count(truck_index, shift_index)
            next_target = self.target_trip_count(truck_index, shift_index + 1)
            if current_target <= 22 and next_target >= 18:
                return self.catalog.trucks[truck_index]
        raise QAError("Не удалось выбрать самосвал для безопасного carryover.")

    def _plan_one_role(
        self,
        *,
        production_date: date,
        role_code: str,
        day_brigade: int,
        night_brigade: int,
        when: datetime,
    ) -> CrewPlan:
        assert self.deputy.client is not None
        path = (
            "/deputy-mining-manager/"
            f"?role={role_code}&date={production_date.isoformat()}"
        )
        with at_time(when):
            self.deputy.client.get(
                path,
                label=f"Расстановка {role_code} {production_date}",
            )
        plan = (
            CrewPlan.objects.filter(
                work_date=production_date,
                role__code=role_code,
                status=CrewPlanStatus.DRAFT,
            )
            .select_related("role")
            .order_by("-revision")
            .first()
        )
        if not plan:
            raise QAError(
                f"GET расстановки не создал draft {role_code}/{production_date}."
            )

        equipment_items = (
            self.catalog.trucks
            if role_code == "driver"
            else self.catalog.excavators
        )
        staff_map = (
            self.onboarding.drivers_by_brigade
            if role_code == "driver"
            else self.onboarding.operators_by_brigade
        )
        for shift_type, brigade in (
            (WorkShiftType.SHIFT_1, day_brigade),
            (WorkShiftType.SHIFT_2, night_brigade),
        ):
            for equipment in equipment_items:
                employee = staff_map[brigade][equipment.id]
                expected_version = plan.version
                payload = {
                    "plan_id": plan.id,
                    "expected_version": expected_version,
                    "equipment_id": equipment.id,
                    "shift_type": shift_type,
                    "employee_id": employee.employee_id,
                }
                with at_time(when):
                    _, body = self.deputy.client.post_json(
                        "/deputy-mining-manager/slot/",
                        payload,
                        label=(
                            f"Слот {role_code}/{equipment.garage_number}/"
                            f"{shift_type}"
                        ),
                    )
                new_version = int(body["payload"]["plan"]["version"])
                if new_version not in {
                    expected_version,
                    expected_version + 1,
                }:
                    raise QAError(
                        f"Версия плана выросла {expected_version}->{new_version}, "
                        "ожидался no-op или шаг +1."
                    )
                plan.version = new_version

                if not self.stale_version_checked:
                    stale_payload = dict(payload)
                    stale_payload["employee_id"] = employee.employee_id
                    with at_time(when):
                        _, stale_body = self.deputy.client.post_json_raw(
                            "/deputy-mining-manager/slot/",
                            stale_payload,
                            expected=(409,),
                            label="Защита stale_version",
                        )
                    if stale_body.get("code") != "stale_version":
                        raise QAError(
                            f"Ожидался stale_version, получено {stale_body!r}."
                        )
                    self.stale_version_checked = True

        with at_time(when):
            _, publish_body = self.deputy.client.post_json(
                "/deputy-mining-manager/publish/",
                {
                    "plan_id": plan.id,
                    "expected_version": plan.version,
                },
                label=f"Публикация {role_code}/{production_date}",
            )
        if not publish_body.get("published"):
            raise QAError("Сервер не подтвердил публикацию расстановки.")
        plan.refresh_from_db()
        if plan.status != CrewPlanStatus.PUBLISHED:
            raise QAError(f"План {plan.id} не получил статус PUBLISHED.")
        return plan

    def publish_daily_plans(self, day_index: int) -> tuple[CrewPlan, CrewPlan]:
        production_date = self.config.start_date + timedelta(days=day_index)
        when = local_dt(production_date, 7, 1)
        day_brigade = DAY_BRIGADES[day_index]
        night_brigade = NIGHT_BRIGADES[day_index]
        driver_plan = self._plan_one_role(
            production_date=production_date,
            role_code="driver",
            day_brigade=day_brigade,
            night_brigade=night_brigade,
            when=when,
        )
        operator_plan = self._plan_one_role(
            production_date=production_date,
            role_code="excavator_operator",
            day_brigade=day_brigade,
            night_brigade=night_brigade,
            when=when + timedelta(seconds=30),
        )
        return driver_plan, operator_plan

    def _last_closed_shift(self, equipment: Equipment) -> EmployeeShift | None:
        return (
            EmployeeShift.objects.filter(
                equipment=equipment,
                closed_at__isnull=False,
            )
            .order_by("-closed_at", "-id")
            .first()
        )

    def driver_start_readings(
        self,
        truck: Equipment,
        truck_index: int,
        shift_index: int,
    ) -> dict[str, Decimal]:
        previous = self._last_closed_shift(truck)
        if not previous:
            limit = Decimal(truck.model.fuel_capacity_limit_l)
            return {
                "start_fuel": q2(limit * Decimal("0.75")),
                "start_mileage": q2(10000 + truck_index * 100),
                "start_engine_hours": q2(2000 + truck_index * 20),
            }
        values = {
            "start_fuel": q2(previous.end_fuel),
            "start_mileage": q2(previous.end_mileage),
            "start_engine_hours": q2(previous.end_engine_hours),
        }
        if shift_index > 0 and truck_index == (shift_index * 11) % len(
            self.catalog.trucks
        ):
            values["start_mileage"] = q2(
                values["start_mileage"] + Decimal("0.50")
            )
        return values

    def operator_start_readings(
        self,
        excavator: Equipment,
        excavator_index: int,
        shift_index: int,
    ) -> dict[str, Decimal]:
        previous = self._last_closed_shift(excavator)
        if not previous:
            limit = Decimal(excavator.model.fuel_capacity_limit_l)
            return {
                "fuel": q2(limit * Decimal("0.75")),
                "engine_hours": q2(5000 + excavator_index * 100),
            }
        fuel = q2(previous.end_fuel)
        engine_hours = q2(previous.end_engine_hours)
        if shift_index > 0 and excavator_index == shift_index % len(
            self.catalog.excavators
        ):
            engine_hours = q2(engine_hours + Decimal("0.25"))
        return {
            "fuel": fuel,
            "engine_hours": engine_hours,
        }

    def open_shift_roles(
        self,
        *,
        shift_index: int,
        open_time: datetime,
    ) -> tuple[StaffMember, StaffMember]:
        brigade = self.brigade_for_shift(shift_index)
        dispatcher = self.onboarding.shift_roles_by_brigade["dispatcher"][
            brigade
        ]
        mining_master = self.onboarding.shift_roles_by_brigade["mining_master"][
            brigade
        ]
        assert dispatcher.client is not None
        assert mining_master.client is not None
        with at_time(open_time):
            dispatcher.client.post_form(
                "/dispatcher/shift/toggle/",
                {
                    "shift_action": "start",
                    "reauth_phone": dispatcher.phone,
                    "reauth_access_code": dispatcher.permanent_pin,
                    "device_kind": "shared",
                },
                label=f"Старт диспетчера смены {shift_index}",
            )
            mining_master.client.post_form(
                "/mining-master/assignments/",
                {
                    "action": "start_shift",
                    "reauth_phone": mining_master.phone,
                    "reauth_access_code": mining_master.permanent_pin,
                    "device_kind": "shared",
                },
                label=f"Старт горного мастера смены {shift_index}",
            )
        for member, workplace in (
            (dispatcher, "dispatcher"),
            (mining_master, "mining_master"),
        ):
            if not EmployeeShift.objects.filter(
                employee_id=member.employee_id,
                workplace_code=workplace,
                closed_at__isnull=True,
            ).exists():
                raise QAError(f"Не открылась смена {workplace}.")
        return dispatcher, mining_master

    def open_equipment_shifts(
        self,
        *,
        shift_index: int,
        open_time: datetime,
    ) -> tuple[dict[int, EmployeeShift], dict[int, EmployeeShift]]:
        brigade = self.brigade_for_shift(shift_index)
        driver_shifts: dict[int, EmployeeShift] = {}
        operator_shifts: dict[int, EmployeeShift] = {}

        for truck_index, truck in enumerate(self.catalog.trucks):
            member = self.onboarding.drivers_by_brigade[brigade][truck.id]
            assert member.client is not None
            readings = self.driver_start_readings(
                truck,
                truck_index,
                shift_index,
            )
            action_id = (
                f"{self.config.run_id}-S{shift_index:02d}-"
                f"T{truck.id}-DRIVER-OPEN"
            )
            with at_time(open_time + timedelta(seconds=truck_index % 10)):
                response = member.client.post_form(
                    "/driver/shift/",
                    {
                        **{key: str(value) for key, value in readings.items()},
                        "client_action_id": action_id,
                    },
                    label=f"Открытие смены водителя {truck.garage_number}",
                )
            if urlparse(response["Location"]).path not in {
                "/driver/",
                "/driver/shift/",
            }:
                raise QAError(
                    f"Водитель {truck.garage_number}: неожиданный redirect "
                    f"{response['Location']!r}."
                )
            shift = (
                EmployeeShift.objects.filter(
                    employee_id=member.employee_id,
                    equipment=truck,
                    workplace_code="driver",
                    closed_at__isnull=True,
                )
                .order_by("-id")
                .first()
            )
            if not shift:
                raise QAError(
                    f"Не открылась смена водителя {truck.garage_number}."
                )
            if shift.watch_period_id != self.onboarding.watch_period.id:
                raise QAError(
                    f"Смена водителя {truck.garage_number} не получила "
                    "доказанный snapshot календарной вахты."
                )
            driver_shifts[truck.id] = shift

        for excavator_index, excavator in enumerate(self.catalog.excavators):
            member = self.onboarding.operators_by_brigade[brigade][excavator.id]
            assert member.client is not None
            readings = self.operator_start_readings(
                excavator,
                excavator_index,
                shift_index,
            )
            action_id = (
                f"{self.config.run_id}-S{shift_index:02d}-"
                f"E{excavator.id}-OPERATOR-OPEN"
            )
            with at_time(open_time + timedelta(seconds=20 + excavator_index)):
                _, body = member.client.post_json(
                    "/excavator/shift/",
                    {
                        "action": "open",
                        "client_action_id": action_id,
                        **{key: str(value) for key, value in readings.items()},
                    },
                    label=(
                        f"Открытие смены машиниста "
                        f"{excavator.garage_number}"
                    ),
                )
            if not body.get("shift_id"):
                raise QAError(
                    f"Машинист {excavator.garage_number}: нет shift_id."
                )
            shift = EmployeeShift.objects.get(pk=body["shift_id"])
            operator_shifts[excavator.id] = shift

        if len(driver_shifts) != len(self.catalog.trucks):
            raise QAError("Открыты не все водительские смены.")
        if len(operator_shifts) != len(self.catalog.excavators):
            raise QAError("Открыты не все смены машинистов.")
        return driver_shifts, operator_shifts

    def _active_haul_map(self) -> dict[int, int]:
        result = {}
        assignments = (
            HaulAssignment.objects.filter(
                status=AssignmentStatus.ACCEPTED,
                ended_at__isnull=True,
            )
            .order_by("truck_id", "-accepted_at", "-id")
            .values_list("truck_id", "excavator_id")
        )
        for truck_id, excavator_id in assignments:
            result.setdefault(truck_id, excavator_id)
        return result

    def establish_initial_complexes(
        self,
        *,
        shift_index: int,
        dispatcher: StaffMember,
        mining_master: StaffMember,
        when: datetime,
    ) -> None:
        assert dispatcher.client is not None
        assert mining_master.client is not None
        brigade = self.brigade_for_shift(shift_index)
        for excavator in self.catalog.excavators:
            with at_time(when):
                dispatcher.client.post_json(
                    "/dispatcher/control/excavator/move/",
                    {"excavator_id": excavator.id, "zone": "active"},
                    label=f"Активация экскаватора {excavator.garage_number}",
                )

        delayed_assignment_ids = []
        for truck_index, truck in enumerate(self.catalog.trucks):
            excavator = self.catalog.excavators[
                truck_index % len(self.catalog.excavators)
            ]
            actor = dispatcher if truck_index % 5 else mining_master
            assert actor.client is not None
            endpoint = (
                "/dispatcher/control/truck/assign/"
                if actor.role_code == "dispatcher"
                else "/mining-master/assignments/truck/assign/"
            )
            payload = {
                "action": "assign",
                "truck_id": truck.id,
                "excavator_id": excavator.id,
            }
            if actor.role_code == "mining_master":
                payload.update(
                    {
                        "expected_source_excavator_id": "",
                        "client_action_id": (
                            f"{self.config.run_id}-INITIAL-ASSIGN-{truck.id}"
                        ),
                    }
                )
            with at_time(when + timedelta(seconds=truck_index % 20)):
                _, body = actor.client.post_json(
                    endpoint,
                    payload,
                    label=f"Назначение {truck.garage_number}",
                )
            assignment_id = int(body["assignment_id"])
            if truck_index % 5:
                driver = self.onboarding.drivers_by_brigade[brigade][truck.id]
                assert driver.client is not None
                with at_time(when + timedelta(minutes=1)):
                    _, accept_body = driver.client.post_json_raw(
                        f"/driver/assignment/{assignment_id}/accept/",
                        {},
                        expected=(200,),
                        label=f"Принятие назначения {truck.garage_number}",
                        HTTP_X_REQUESTED_WITH="XMLHttpRequest",
                    )
                if not accept_body.get("ok"):
                    raise QAError(
                        f"Водитель не принял назначение {assignment_id}."
                    )
            else:
                delayed_assignment_ids.append(assignment_id)

        if delayed_assignment_ids:
            with at_time(when + timedelta(minutes=6)):
                dispatcher.client.get(
                    "/realtime/state/?after=0&limit=1&include_events=0",
                    label="Автоприменение назначений через realtime",
                )
            unresolved = HaulAssignment.objects.filter(
                id__in=delayed_assignment_ids,
                status=AssignmentStatus.PENDING,
                ended_at__isnull=True,
            ).count()
            if unresolved:
                raise QAError(
                    f"Через realtime не применено назначений: {unresolved}."
                )

        self.current_haul_map = self._active_haul_map()
        if set(self.current_haul_map) != {
            truck.id for truck in self.catalog.trucks
        }:
            raise QAError("Не все самосвалы получили действующий комплекс.")

    def rotate_daily_complexes(
        self,
        *,
        shift_index: int,
        mining_master: StaffMember,
        when: datetime,
    ) -> None:
        if shift_index == 0 or shift_index % 2:
            return
        assert mining_master.client is not None
        brigade = self.brigade_for_shift(shift_index)
        carry_in_id = self.carryover_truck_id
        moved = 0
        day_index = shift_index // 2
        used_trucks = set()
        for excavator_index, source_excavator in enumerate(
            self.catalog.excavators
        ):
            candidates = [
                truck
                for truck in self.catalog.trucks
                if self.current_haul_map.get(truck.id) == source_excavator.id
                and truck.id != carry_in_id
                and truck.id not in used_trucks
            ]
            if not candidates:
                continue
            truck = candidates[(day_index + excavator_index) % len(candidates)]
            target_excavator = self.catalog.excavators[
                (excavator_index + 1) % len(self.catalog.excavators)
            ]
            payload = {
                "action": "assign",
                "truck_id": truck.id,
                "excavator_id": target_excavator.id,
                "expected_source_excavator_id": str(source_excavator.id),
                "client_action_id": (
                    f"{self.config.run_id}-D{day_index}-MOVE-{truck.id}"
                ),
            }
            with at_time(when + timedelta(seconds=moved)):
                _, body = mining_master.client.post_json(
                    "/mining-master/assignments/truck/assign/",
                    payload,
                    label=f"Суточная перестановка {truck.garage_number}",
                )
            assignment_id = int(body["assignment_id"])
            driver = self.onboarding.drivers_by_brigade[brigade][truck.id]
            assert driver.client is not None
            with at_time(when + timedelta(seconds=20 + moved)):
                _, accept_body = driver.client.post_json_raw(
                    f"/driver/assignment/{assignment_id}/accept/",
                    {},
                    expected=(200,),
                    label=f"Принятие перестановки {truck.garage_number}",
                    HTTP_X_REQUESTED_WITH="XMLHttpRequest",
                )
            if not accept_body.get("ok"):
                raise QAError(
                    f"Не принята перестановка {truck.garage_number}."
                )
            used_trucks.add(truck.id)
            moved += 1
        self.current_haul_map = self._active_haul_map()
        if len(self.current_haul_map) != len(self.catalog.trucks):
            raise QAError("После перестановки потеряны комплексы.")

    def apply_excavator_settings(
        self,
        *,
        shift_index: int,
        when: datetime,
    ) -> dict[int, dict[str, Any]]:
        brigade = self.brigade_for_shift(shift_index)
        contexts = {}
        for excavator_index, excavator in enumerate(self.catalog.excavators):
            member = self.onboarding.operators_by_brigade[brigade][excavator.id]
            assert member.client is not None
            source = self.reference_context_by_excavator[excavator.id]
            rock = self.catalog.operational_rocks[
                (shift_index * len(self.catalog.excavators) + excavator_index)
                % len(self.catalog.operational_rocks)
            ]
            dump_primary = self.catalog.dump_points[
                (shift_index + excavator_index) % len(self.catalog.dump_points)
            ]
            dump_secondary = self.catalog.dump_points[
                (shift_index + excavator_index + 1)
                % len(self.catalog.dump_points)
            ]
            dump_points = [dump_primary]
            if dump_secondary.id != dump_primary.id:
                dump_points.append(dump_secondary)
            payload = {
                "client_action_id": (
                    f"{self.config.run_id}-S{shift_index:02d}-"
                    f"E{excavator.id}-SETTINGS"
                ),
                "rock_type_id": rock.id,
                "dump_point_ids": [point.id for point in dump_points],
                "loading_horizon": source["loading_horizon"],
                "loading_block": source["loading_block"],
            }
            with at_time(when + timedelta(seconds=excavator_index)):
                _, body = member.client.post_json(
                    "/excavator/work/settings/",
                    payload,
                    label=f"Настройки {excavator.garage_number}",
                )
            contexts[excavator.id] = {
                "rock": rock,
                "dump_points": dump_points,
                "loading_horizon": body["loading_horizon"],
                "loading_block": body["loading_block"],
            }
        return contexts

    def probe_missing_capacity_rule(
        self,
        *,
        shift_index: int,
        when: datetime,
        contexts: dict[int, dict[str, Any]],
    ) -> None:
        """Prove the real unresolved reference combination through HTTP."""

        if not self.catalog.unresolved_capacity_pairs:
            return
        unresolved = self.catalog.unresolved_capacity_pairs[0]
        truck = next(
            item
            for item in self.catalog.trucks
            if item.id == unresolved["truck_id"]
        )
        rock = next(
            item
            for item in self.catalog.rocks
            if item.id == unresolved["rock_type_id"]
        )
        excavator_id = self.current_haul_map[truck.id]
        context = contexts[excavator_id]
        brigade = self.brigade_for_shift(shift_index)
        operator = self.onboarding.operators_by_brigade[brigade][
            excavator_id
        ]
        assert operator.client is not None
        trip_count_before = Trip.objects.count()
        action_id = f"{self.config.run_id}-CAPACITY-MISSING-PROBE"
        with at_time(when):
            response, body = operator.client.post_json_raw(
                "/excavator/truck-loaded/",
                {
                    "client_action_id": action_id,
                    "truck_id": truck.id,
                    "excavator_id": excavator_id,
                    "dump_point_id": context["dump_points"][0].id,
                    "rock_type_id": rock.id,
                    "loading_horizon": context["loading_horizon"],
                    "loading_block": context["loading_block"],
                },
                expected=(200, 400, 409),
                label="Проверка отсутствующей кубатуры справочника",
            )
        trip_created = bool(body.get("ok") and body.get("trip_id"))
        probe_trip = None
        if trip_created:
            probe_trip = Trip.objects.get(pk=body["trip_id"])
            if Trip.objects.count() != trip_count_before + 1:
                raise QAError(
                    "Проверка отсутствующей кубатуры создала неожиданное "
                    "число рейсов."
                )
            dispatcher = self.onboarding.shift_roles_by_brigade[
                "dispatcher"
            ][brigade]
            assert dispatcher.client is not None
            with at_time(when + timedelta(seconds=1)):
                dispatcher.client.post_form(
                    f"/dispatcher/trips/{probe_trip.id}/cancel/",
                    {
                        "reason": (
                            "QA: активная порода не имеет кубатуры "
                            "для модели самосвала"
                        )
                    },
                    label="Штатная отмена диагностического рейса",
                )
            probe_trip.refresh_from_db()
            if probe_trip.status != TripStatus.CANCELLED:
                raise QAError(
                    "Диагностический рейс без кубатуры не отменён."
                )
        elif Trip.objects.count() != trip_count_before:
            raise QAError(
                "Отклонённая комбинация без кубатуры изменила число рейсов."
            )

        self.reference_capacity_probe = {
            **unresolved,
            "http_status": response.status_code,
            "server_error": body.get("error", ""),
            "trip_created": trip_created,
            "trip_id": probe_trip.id if probe_trip else None,
            "trip_status": probe_trip.status if probe_trip else None,
            "volume_m3": probe_trip.volume_m3 if probe_trip else None,
            "tonnage": probe_trip.tonnage if probe_trip else None,
        }
        append_jsonl(
            self.action_log_path,
            {
                "action": "missing_capacity_rule_probe",
                **self.reference_capacity_probe,
            },
        )

    def close_transferred_downtimes(
        self,
        *,
        shift_index: int,
        when: datetime,
    ) -> None:
        brigade = self.brigade_for_shift(shift_index)
        for event in list(
            DowntimeEvent.objects.filter(ended_at__isnull=True)
            .select_related(
                "equipment",
                "equipment__equipment_type",
                "employee",
                "reason",
            )
            .order_by("id")
        ):
            original_employee_id = event.employee_id
            equipment_type = event.equipment.equipment_type.name.casefold()
            if "самосвал" in equipment_type:
                member = self.onboarding.drivers_by_brigade[brigade].get(
                    event.equipment_id
                )
                if not member or not member.client:
                    continue
                with at_time(when):
                    _, body = member.client.post_json(
                        "/driver/downtime/",
                        {"action": "close"},
                        label=(
                            f"Сменщик закрывает простой "
                            f"{event.equipment.garage_number}"
                        ),
                        HTTP_X_REQUESTED_WITH="XMLHttpRequest",
                    )
                if not body.get("closed"):
                    raise QAError(
                        f"Простой {event.id} самосвала не был закрыт сменщиком."
                    )
            elif "экскаватор" in equipment_type:
                member = self.onboarding.operators_by_brigade[brigade].get(
                    event.equipment_id
                )
                if not member or not member.client:
                    continue
                with at_time(when):
                    _, body = member.client.post_json(
                        "/excavator/downtime/",
                        {"action": "close"},
                        label=(
                            f"Сменщик закрывает простой "
                            f"{event.equipment.garage_number}"
                        ),
                    )
                if not body.get("closed"):
                    raise QAError(
                        f"Простой {event.id} экскаватора не был закрыт сменщиком."
                    )
            else:
                continue
            event.refresh_from_db()
            if not event.ended_at or event.employee_id != original_employee_id:
                raise QAError(
                    f"Переданный простой {event.id} потерял время или автора."
                )
            append_jsonl(
                self.action_log_path,
                {
                    "action": "transferred_downtime_closed",
                    "shift_index": shift_index,
                    "downtime_id": event.id,
                    "equipment_id": event.equipment_id,
                    "original_employee_id": original_employee_id,
                    "closed_at": event.ended_at,
                },
            )

    def start_handoff_downtime(
        self,
        *,
        shift_index: int,
        when: datetime,
        carry_out_truck_id: int | None,
    ) -> set[int]:
        """Create two explicit handoff cases during the week."""

        excluded_from_maintenance_check: set[int] = set()
        brigade = self.brigade_for_shift(shift_index)
        if shift_index == 3:
            truck = next(
                item
                for item in self.catalog.trucks
                if item.id != carry_out_truck_id
            )
            member = self.onboarding.drivers_by_brigade[brigade][truck.id]
            assert member.client is not None
            reason = (
                DowntimeReason.for_workplace(
                    "truck_driver",
                    truck.equipment_type,
                )
                .filter(name__icontains="Заправ")
                .first()
                or DowntimeReason.for_workplace(
                    "truck_driver",
                    truck.equipment_type,
                ).first()
            )
            if not reason:
                raise QAError("Нет причины простоя для Водителя.")
            with at_time(when):
                _, body = member.client.post_json(
                    "/driver/downtime/",
                    {"action": "start", "reason_id": reason.id},
                    label="Водитель передаёт открытый простой сменщику",
                    HTTP_X_REQUESTED_WITH="XMLHttpRequest",
                )
            if not body.get("active"):
                raise QAError("Не удалось открыть передаваемый простой Водителя.")
            event = DowntimeEvent.objects.get(
                equipment=truck,
                ended_at__isnull=True,
            )
            self.transferred_downtime_ids["driver"] = event.id
            excluded_from_maintenance_check.add(truck.id)

        if shift_index == 7:
            excavator = self.catalog.excavators[0]
            member = self.onboarding.operators_by_brigade[brigade][
                excavator.id
            ]
            assert member.client is not None
            reason = (
                DowntimeReason.for_workplace(
                    "excavator_operator",
                    excavator.equipment_type,
                )
                .filter(name__icontains="Подготов")
                .first()
                or DowntimeReason.for_workplace(
                    "excavator_operator",
                    excavator.equipment_type,
                ).first()
            )
            if not reason:
                raise QAError("Нет причины простоя для Машиниста.")
            with at_time(when):
                _, body = member.client.post_json(
                    "/excavator/downtime/",
                    {
                        "action": "start",
                        "reason_id": reason.id,
                        "comment": "",
                    },
                    label="Машинист передаёт открытый простой сменщику",
                )
            if not body.get("active"):
                raise QAError(
                    "Не удалось открыть передаваемый простой Машиниста."
                )
            event = DowntimeEvent.objects.get(
                equipment=excavator,
                ended_at__isnull=True,
            )
            self.transferred_downtime_ids["excavator_operator"] = event.id
            excluded_from_maintenance_check.add(excavator.id)
        return excluded_from_maintenance_check

    def _unload_trip(
        self,
        *,
        shift_index: int,
        truck: Equipment,
        trip_id: int,
        member: StaffMember,
        when: datetime,
        sequence: str,
    ) -> None:
        assert member.client is not None
        action_id = (
            f"{self.config.run_id}-S{shift_index:02d}-"
            f"T{truck.id}-{sequence}-UNLOAD"
        )
        with at_time(when):
            member.client.post_form(
                f"/driver/trip/{trip_id}/complete/",
                {"client_action_id": action_id},
                label=f"Разгрузка {truck.garage_number}/{trip_id}",
            )
        trip = Trip.objects.select_related("unloading_shift").get(pk=trip_id)
        if (
            trip.status != TripStatus.COMPLETED
            or not trip.completed_at
            or not trip.unloading_shift_id
        ):
            raise QAError(f"Рейс {trip_id} не завершён штатной разгрузкой.")
        expected_volume, expected_tonnage = self.catalog.trip_measurements(
            truck,
            trip.rock_type,
        )
        if q2(trip.volume_m3) != expected_volume:
            raise QAError(
                f"Рейс {trip_id}: после разгрузки кубатура "
                f"{trip.volume_m3}, ожидалась {expected_volume}."
            )
        if expected_tonnage is None:
            if trip.tonnage is not None:
                raise QAError(
                    f"Рейс {trip_id}: тоннаж появился без плотности."
                )
        elif q2(trip.tonnage) != expected_tonnage:
            raise QAError(
                f"Рейс {trip_id}: после разгрузки тоннаж "
                f"{trip.tonnage}, ожидался {expected_tonnage}."
            )
        self.total_unload_events += 1
        if self.total_unload_events % 997 == 0:
            before_actions = TripClientAction.objects.filter(
                action_type="trip_unloaded",
                client_action_id=action_id,
            ).count()
            with at_time(when + timedelta(milliseconds=100)):
                member.client.post_form(
                    f"/driver/trip/{trip_id}/complete/",
                    {"client_action_id": action_id},
                    label=f"Идемпотентная разгрузка {trip_id}",
                )
            after_actions = TripClientAction.objects.filter(
                action_type="trip_unloaded",
                client_action_id=action_id,
            ).count()
            if before_actions != 1 or after_actions != 1:
                raise QAError(
                    f"Дубликат разгрузки {trip_id} изменил ledger."
                )
            self.idempotency_unload_checks += 1

    def execute_trip_cycle(
        self,
        *,
        shift_index: int,
        open_time: datetime,
        close_time: datetime,
        contexts: dict[int, dict[str, Any]],
        driver_shifts: dict[int, EmployeeShift],
        operator_shifts: dict[int, EmployeeShift],
    ) -> tuple[int | None, int | None]:
        del driver_shifts  # Связь проверяется через unloading_shift.
        brigade = self.brigade_for_shift(shift_index)
        carry_out_truck = (
            self.carry_truck_for_boundary(shift_index)
            if shift_index < 13
            else None
        )
        carry_out_truck_id = carry_out_truck.id if carry_out_truck else None
        carry_in_trip_id = self.carryover_trip_id
        carry_in_truck_id = self.carryover_truck_id

        events: list[tuple[datetime, int, str, dict[str, Any]]] = []
        if carry_in_trip_id and carry_in_truck_id:
            carry_truck = next(
                item
                for item in self.catalog.trucks
                if item.id == carry_in_truck_id
            )
            events.append(
                (
                    open_time + timedelta(minutes=6),
                    0,
                    "carry_unload",
                    {
                        "truck": carry_truck,
                        "trip_id": carry_in_trip_id,
                    },
                )
            )

        total_window_seconds = int(
            (
                (close_time - timedelta(minutes=25))
                - (open_time + timedelta(minutes=15))
            ).total_seconds()
        )
        open_trip_keys: dict[tuple[int, int], int] = {}
        new_load_count_expected = 0
        for truck_index, truck in enumerate(self.catalog.trucks):
            target = self.target_trip_count(truck_index, shift_index)
            has_carry_in = truck.id == carry_in_truck_id
            has_carry_out = truck.id == carry_out_truck_id
            new_loads = target - int(has_carry_in) + int(has_carry_out)
            if not 17 <= new_loads <= 23:
                raise QAError(
                    f"Нереалистичное число загрузок {new_loads} для "
                    f"{truck.garage_number}/S{shift_index}."
                )
            new_load_count_expected += new_loads
            spacing = total_window_seconds / max(new_loads, 1)
            for trip_index in range(new_loads):
                load_time = (
                    open_time
                    + timedelta(minutes=15)
                    + timedelta(seconds=round(trip_index * spacing))
                    + timedelta(seconds=(truck_index % 8) * 2)
                )
                key = (truck.id, trip_index)
                events.append(
                    (
                        load_time,
                        1,
                        "load",
                        {
                            "key": key,
                            "truck": truck,
                            "truck_index": truck_index,
                            "trip_index": trip_index,
                            "is_carry_out": (
                                has_carry_out and trip_index == new_loads - 1
                            ),
                        },
                    )
                )
                if not (has_carry_out and trip_index == new_loads - 1):
                    unload_delay = 12 + ((truck_index + trip_index) % 7)
                    events.append(
                        (
                            load_time + timedelta(minutes=unload_delay),
                            2,
                            "unload",
                            {
                                "key": key,
                                "truck": truck,
                                "trip_index": trip_index,
                            },
                        )
                    )

        events.sort(key=lambda item: (item[0], item[1], item[3]["truck"].id))
        next_carry_trip_id = None
        next_carry_truck_id = None
        for event_time, _priority, action, payload in events:
            truck = payload["truck"]
            member = self.onboarding.drivers_by_brigade[brigade][truck.id]
            if action == "carry_unload":
                self._unload_trip(
                    shift_index=shift_index,
                    truck=truck,
                    trip_id=payload["trip_id"],
                    member=member,
                    when=event_time,
                    sequence="CARRYIN",
                )
                continue
            if action == "load":
                excavator_id = self.current_haul_map.get(truck.id)
                if not excavator_id:
                    raise QAError(
                        f"У {truck.garage_number} нет принятого комплекса."
                    )
                context = contexts[excavator_id]
                operator = self.onboarding.operators_by_brigade[brigade][
                    excavator_id
                ]
                assert operator.client is not None
                dump_points = context["dump_points"]
                dump_point = dump_points[
                    (payload["truck_index"] + payload["trip_index"])
                    % len(dump_points)
                ]
                rock = context["rock"]
                action_id = (
                    f"{self.config.run_id}-S{shift_index:02d}-"
                    f"T{truck.id}-R{payload['trip_index']:02d}-LOAD"
                )
                request_payload = {
                    "client_action_id": action_id,
                    "truck_id": truck.id,
                    "excavator_id": excavator_id,
                    "dump_point_id": dump_point.id,
                    "rock_type_id": rock.id,
                    "loading_horizon": context["loading_horizon"],
                    "loading_block": context["loading_block"],
                }
                with at_time(event_time):
                    _, body = operator.client.post_json(
                        "/excavator/truck-loaded/",
                        request_payload,
                        label=(
                            f"Погрузка {truck.garage_number}/"
                            f"{payload['trip_index']}"
                        ),
                    )
                trip_id = int(body["trip_id"])
                open_trip_keys[payload["key"]] = trip_id
                trip = Trip.objects.select_related(
                    "loading_shift",
                    "rock_type",
                    "dump_point",
                    "assigned_dump_point",
                ).get(pk=trip_id)
                expected_volume, expected_tonnage = (
                    self.catalog.trip_measurements(truck, rock)
                )
                if (
                    trip.volume_m3 is not None
                    and q2(trip.volume_m3) != expected_volume
                ):
                    raise QAError(
                        f"Рейс {trip_id}: кубатура {trip.volume_m3}, "
                        f"ожидалась {expected_volume}."
                    )
                if expected_tonnage is None:
                    if trip.tonnage is not None:
                        raise QAError(
                            f"Рейс {trip_id}: тоннаж появился без плотности."
                        )
                elif (
                    trip.tonnage is not None
                    and q2(trip.tonnage) != expected_tonnage
                ):
                    raise QAError(
                        f"Рейс {trip_id}: тоннаж {trip.tonnage}, "
                        f"ожидался {expected_tonnage}."
                    )
                manifest_row = {
                    "run_id": self.config.run_id,
                    "shift_index": shift_index,
                    "production_date": self.date_for_shift(
                        shift_index
                    ).isoformat(),
                    "shift_type": self.shift_type_for_index(shift_index),
                    "truck_id": truck.id,
                    "truck": truck.garage_number,
                    "excavator_id": excavator_id,
                    "operator_id": operator.employee_id,
                    "driver_id_at_load": member.employee_id,
                    "trip_sequence": payload["trip_index"],
                    "trip_id": trip_id,
                    "load_action_id": action_id,
                    "rock_type_id": rock.id,
                    "rock_type": rock.name,
                    "dump_point_id": dump_point.id,
                    "dump_point": dump_point.name,
                    "loading_horizon": context["loading_horizon"],
                    "loading_block": context["loading_block"],
                    "expected_volume_m3": expected_volume,
                    "expected_tonnage": expected_tonnage,
                    "loaded_at": event_time,
                    "is_planned_carryover": payload["is_carry_out"],
                }
                append_jsonl(self.manifest_path, manifest_row)

                self.total_load_events += 1
                if self.total_load_events % 997 == 0:
                    before_trip_count = Trip.objects.count()
                    with at_time(event_time + timedelta(milliseconds=100)):
                        _, duplicate_body = operator.client.post_json(
                            "/excavator/truck-loaded/",
                            request_payload,
                            label=f"Идемпотентная погрузка {trip_id}",
                        )
                    if (
                        not duplicate_body.get("deduplicated")
                        or int(duplicate_body["trip_id"]) != trip_id
                        or Trip.objects.count() != before_trip_count
                    ):
                        raise QAError(
                            f"Дубликат погрузки создал второй рейс {trip_id}."
                        )
                    self.idempotency_load_checks += 1
                if payload["is_carry_out"]:
                    next_carry_trip_id = trip_id
                    next_carry_truck_id = truck.id
                continue

            trip_id = open_trip_keys.get(payload["key"])
            if not trip_id:
                raise QAError(
                    f"Событие разгрузки не нашло погрузку {payload['key']}."
                )
            self._unload_trip(
                shift_index=shift_index,
                truck=truck,
                trip_id=trip_id,
                member=member,
                when=event_time,
                sequence=f"R{payload['trip_index']:02d}",
            )

        actual_loaded = Trip.objects.filter(
            loading_shift_id__in=[
                shift.id for shift in operator_shifts.values()
            ]
        ).exclude(status=TripStatus.CANCELLED).count()
        if actual_loaded != new_load_count_expected:
            raise QAError(
                f"Смена {shift_index}: загрузок {actual_loaded}, "
                f"ожидалось {new_load_count_expected}."
            )
        return next_carry_trip_id, next_carry_truck_id

    def close_equipment_shifts(
        self,
        *,
        shift_index: int,
        close_time: datetime,
        driver_shifts: dict[int, EmployeeShift],
        operator_shifts: dict[int, EmployeeShift],
        carry_out_truck_id: int | None,
    ) -> None:
        """Close every equipment shift through its real workplace endpoint."""

        brigade = self.brigade_for_shift(shift_index)
        for truck_index, truck in enumerate(self.catalog.trucks):
            member = self.onboarding.drivers_by_brigade[brigade][truck.id]
            assert member.client is not None
            shift = driver_shifts[truck.id]
            shift.refresh_from_db()
            target = self.target_trip_count(truck_index, shift_index)
            limit = Decimal(truck.model.fuel_capacity_limit_l)
            end_fuel = max(
                q2(
                    shift.start_fuel
                    - Decimal(target * 3)
                ),
                q2(limit * Decimal("0.20")),
            )
            readings = {
                "end_fuel": end_fuel,
                "end_mileage": q2(
                    shift.start_mileage + Decimal(target * 8)
                ),
                "end_engine_hours": q2(
                    shift.start_engine_hours
                    + Decimal("9.75")
                    + Decimal((truck_index + shift_index) % 5)
                    * Decimal("0.25")
                ),
            }
            action_id = (
                f"{self.config.run_id}-S{shift_index:02d}-"
                f"T{truck.id}-DRIVER-CLOSE"
            )
            # Переходящий водитель завершает смену раньше остальных:
            # самосвал остаётся загруженным и передаётся сменщику.
            effective_close_time = (
                close_time - timedelta(minutes=35)
                if truck.id == carry_out_truck_id
                else close_time + timedelta(seconds=truck_index % 10)
            )
            payload = {
                **{key: str(value) for key, value in readings.items()},
                "client_action_id": action_id,
            }
            with at_time(effective_close_time):
                review_response = member.client.post_form(
                    "/driver/shift/close/",
                    {**payload, "shift_action": "review"},
                    expected=(200,),
                    label=f"Проверка показаний {truck.garage_number}",
                )
                if "data-driver-shift-review" not in review_response.content.decode(
                    "utf-8",
                    errors="replace",
                ):
                    raise QAError(
                        f"Водитель {truck.garage_number}: не показано "
                        "подтверждение конечных показаний."
                    )
                member.client.post_form(
                    "/driver/shift/close/",
                    {**payload, "shift_action": "close"},
                    label=f"Закрытие смены {truck.garage_number}",
                )
            shift.refresh_from_db()
            if (
                not shift.closed_at
                or shift.closed_by_id != member.employee_id
                or q2(shift.end_fuel) != end_fuel
                or q2(shift.end_mileage) != readings["end_mileage"]
                or q2(shift.end_engine_hours)
                != readings["end_engine_hours"]
            ):
                raise QAError(
                    f"Смена Водителя {truck.garage_number} закрыта неверно."
                )

            if truck_index == shift_index % len(self.catalog.trucks):
                before_actions = ShiftClientAction.objects.filter(
                    action_type="driver_shift_closed",
                    client_action_id=action_id,
                ).count()
                with at_time(effective_close_time + timedelta(milliseconds=100)):
                    member.client.post_form(
                        "/driver/shift/close/",
                        {**payload, "shift_action": "close"},
                        label=(
                            f"Идемпотентное закрытие "
                            f"{truck.garage_number}"
                        ),
                    )
                after_actions = ShiftClientAction.objects.filter(
                    action_type="driver_shift_closed",
                    client_action_id=action_id,
                ).count()
                if before_actions != 1 or after_actions != 1:
                    raise QAError(
                        f"Повтор закрытия смены {shift.id} изменил ledger."
                    )
                self.idempotency_shift_checks += 1

        for excavator_index, excavator in enumerate(self.catalog.excavators):
            member = self.onboarding.operators_by_brigade[brigade][
                excavator.id
            ]
            assert member.client is not None
            shift = operator_shifts[excavator.id]
            shift.refresh_from_db()
            limit = Decimal(excavator.model.fuel_capacity_limit_l)
            end_fuel = max(
                q2(shift.start_fuel - Decimal("180.00")),
                q2(limit * Decimal("0.20")),
            )
            end_engine_hours = q2(
                shift.start_engine_hours
                + Decimal("10.00")
                + Decimal((excavator_index + shift_index) % 4)
                * Decimal("0.25")
            )
            action_id = (
                f"{self.config.run_id}-S{shift_index:02d}-"
                f"E{excavator.id}-OPERATOR-CLOSE"
            )
            payload = {
                "action": "close",
                "client_action_id": action_id,
                "fuel": str(end_fuel),
                "engine_hours": str(end_engine_hours),
            }
            with at_time(
                close_time + timedelta(seconds=20 + excavator_index)
            ):
                _, body = member.client.post_json(
                    "/excavator/shift/",
                    payload,
                    label=(
                        f"Закрытие смены машиниста "
                        f"{excavator.garage_number}"
                    ),
                )
            if body.get("shift_open") is not False:
                raise QAError(
                    f"Машинист {excavator.garage_number}: сервер не "
                    "подтвердил закрытие."
                )
            shift.refresh_from_db()
            if (
                not shift.closed_at
                or shift.closed_by_id != member.employee_id
                or q2(shift.end_fuel) != end_fuel
                or q2(shift.end_engine_hours) != end_engine_hours
            ):
                raise QAError(
                    f"Смена Машиниста {excavator.garage_number} "
                    "закрыта неверно."
                )

            if excavator_index == shift_index % len(self.catalog.excavators):
                before_actions = ShiftClientAction.objects.filter(
                    action_type="excavator_shift_closed",
                    client_action_id=action_id,
                ).count()
                with at_time(
                    close_time
                    + timedelta(seconds=30 + excavator_index)
                ):
                    _, duplicate_body = member.client.post_json(
                        "/excavator/shift/",
                        payload,
                        label=(
                            f"Идемпотентное закрытие "
                            f"{excavator.garage_number}"
                        ),
                    )
                after_actions = ShiftClientAction.objects.filter(
                    action_type="excavator_shift_closed",
                    client_action_id=action_id,
                ).count()
                if (
                    not duplicate_body.get("deduplicated")
                    or before_actions != 1
                    or after_actions != 1
                ):
                    raise QAError(
                        f"Повтор закрытия смены {shift.id} изменил ledger."
                    )
                self.idempotency_shift_checks += 1

    def close_shift_roles(
        self,
        *,
        shift_index: int,
        close_time: datetime,
        dispatcher: StaffMember,
        mining_master: StaffMember,
    ) -> None:
        assert dispatcher.client is not None
        assert mining_master.client is not None
        with at_time(close_time + timedelta(minutes=1)):
            dispatcher.client.post_form(
                "/dispatcher/shift/toggle/",
                {"shift_action": "end"},
                label=f"Закрытие смены Диспетчера {shift_index}",
            )
            mining_master.client.post_form(
                "/mining-master/assignments/",
                {"action": "end_shift"},
                label=f"Закрытие смены Горного мастера {shift_index}",
            )
        for member, workplace in (
            (dispatcher, "dispatcher"),
            (mining_master, "mining_master"),
        ):
            if EmployeeShift.objects.filter(
                employee_id=member.employee_id,
                workplace_code=workplace,
                closed_at__isnull=True,
            ).exists():
                raise QAError(f"Не закрылась смена {workplace}.")

    def record_maintenance_handoff_check(
        self,
        *,
        shift_index: int,
        excluded_equipment_ids: set[int],
        checked_at: datetime,
    ) -> None:
        """Record, but do not mask, the owner's expected ETO/TO rule."""

        all_equipment = [*self.catalog.trucks, *self.catalog.excavators]
        expected_ids = {
            equipment.id
            for equipment in all_equipment
            if equipment.id not in excluded_equipment_ids
        }
        active_events = list(
            DowntimeEvent.objects.filter(
                equipment_id__in=expected_ids,
                ended_at__isnull=True,
            )
            .select_related("reason", "equipment")
            .order_by("equipment_id", "-started_at", "-id")
        )
        maintenance_ids = set()
        unexpected = []
        for event in active_events:
            reason_text = " ".join(
                filter(
                    None,
                    [
                        event.reason.name,
                        event.reason.short_label,
                    ],
                )
            ).casefold()
            if (
                "ето" in reason_text
                or reason_text.strip() == "то"
                or "техническ" in reason_text
            ):
                maintenance_ids.add(event.equipment_id)
            else:
                unexpected.append(
                    {
                        "equipment_id": event.equipment_id,
                        "equipment": event.equipment.garage_number,
                        "reason": event.reason.name,
                    }
                )
        missing_ids = sorted(expected_ids - maintenance_ids)
        check = {
            "shift_index": shift_index,
            "checked_at": checked_at,
            "expected_equipment_count": len(expected_ids),
            "maintenance_event_count": len(maintenance_ids),
            "missing_equipment_count": len(missing_ids),
            "missing_equipment_ids": missing_ids,
            "unexpected_open_events": unexpected,
            "passed": not missing_ids and not unexpected,
        }
        self.eto_to_checks.append(check)
        append_jsonl(
            self.action_log_path,
            {"action": "eto_to_after_shift_close_check", **check},
        )

    def verify_shift(
        self,
        *,
        shift_index: int,
        driver_shifts: dict[int, EmployeeShift],
        operator_shifts: dict[int, EmployeeShift],
        carry_in_trip_id: int | None,
        carry_out_trip_id: int | None,
        started_at: float,
    ) -> ShiftResult:
        expected_by_truck = {
            truck.id: self.target_trip_count(truck_index, shift_index)
            for truck_index, truck in enumerate(self.catalog.trucks)
        }
        unloading_shift_ids = [shift.id for shift in driver_shifts.values()]
        loading_shift_ids = [shift.id for shift in operator_shifts.values()]
        actual_by_truck = dict(
            Trip.objects.filter(
                unloading_shift_id__in=unloading_shift_ids,
                status=TripStatus.COMPLETED,
            )
            .values_list("truck_id")
            .annotate(total=Count("id"))
        )
        if actual_by_truck != expected_by_truck:
            differences = {
                truck_id: {
                    "expected": expected_by_truck[truck_id],
                    "actual": actual_by_truck.get(truck_id, 0),
                }
                for truck_id in expected_by_truck
                if actual_by_truck.get(truck_id, 0)
                != expected_by_truck[truck_id]
            }
            raise QAError(
                f"Смена {shift_index}: неверные рейсы по самосвалам "
                f"{dict(list(differences.items())[:10])}."
            )

        unloaded_queryset = Trip.objects.filter(
            unloading_shift_id__in=unloading_shift_ids,
            status=TripStatus.COMPLETED,
        )
        loaded_queryset = Trip.objects.filter(
            loading_shift_id__in=loading_shift_ids,
        ).exclude(status=TripStatus.CANCELLED)
        unloaded_count = unloaded_queryset.count()
        loaded_count = loaded_queryset.count()
        expected_unloaded = sum(expected_by_truck.values())
        expected_loaded = (
            expected_unloaded
            - int(carry_in_trip_id is not None)
            + int(carry_out_trip_id is not None)
        )
        if unloaded_count != expected_unloaded or loaded_count != expected_loaded:
            raise QAError(
                f"Смена {shift_index}: loaded/unloaded="
                f"{loaded_count}/{unloaded_count}, ожидалось "
                f"{expected_loaded}/{expected_unloaded}."
            )

        open_trips = Trip.objects.filter(status__in=OPEN_TRIP_STATUSES)
        expected_open = 1 if carry_out_trip_id else 0
        if open_trips.count() != expected_open:
            raise QAError(
                f"Смена {shift_index}: открытых рейсов "
                f"{open_trips.count()}, ожидалось {expected_open}."
            )
        if carry_out_trip_id:
            open_trip = open_trips.get()
            if (
                open_trip.id != carry_out_trip_id
                or not open_trip.is_carryover
            ):
                raise QAError(
                    f"Смена {shift_index}: переходящий рейс потерян."
                )

        if EmployeeShift.objects.filter(
            closed_at__isnull=True,
            equipment__isnull=False,
        ).exists():
            raise QAError(
                f"Смена {shift_index}: после закрытия остались смены техники."
            )

        totals = unloaded_queryset.aggregate(
            volume=Sum("volume_m3"),
            tonnage=Sum("tonnage"),
        )
        result = ShiftResult(
            index=shift_index,
            production_date=self.date_for_shift(shift_index).isoformat(),
            shift_type=self.shift_type_for_index(shift_index),
            driver_shifts=len(driver_shifts),
            operator_shifts=len(operator_shifts),
            loaded_trips=loaded_count,
            unloaded_trips=unloaded_count,
            carryover_in=int(carry_in_trip_id is not None),
            carryover_out=int(carry_out_trip_id is not None),
            trip_counts_by_truck={
                next(
                    truck.garage_number
                    for truck in self.catalog.trucks
                    if truck.id == truck_id
                ): count
                for truck_id, count in expected_by_truck.items()
            },
            volume_m3=str(q2(totals["volume"] or 0)),
            tonnage=str(q2(totals["tonnage"] or 0)),
            duration_seconds=round(time.perf_counter() - started_at, 3),
        )
        self.shift_results.append(result)
        append_jsonl(
            self.action_log_path,
            {"action": "shift_verified", **asdict(result)},
        )
        return result

    def run(self) -> list[ShiftResult]:
        run_started = time.perf_counter()
        for day_index in range(7):
            self.publish_daily_plans(day_index)
            for shift_in_day in range(2):
                shift_index = day_index * 2 + shift_in_day
                shift_started = time.perf_counter()
                open_time, close_time = self.shift_bounds(shift_index)
                carry_in_trip_id = self.carryover_trip_id
                dispatcher, mining_master = self.open_shift_roles(
                    shift_index=shift_index,
                    open_time=open_time,
                )
                driver_shifts, operator_shifts = self.open_equipment_shifts(
                    shift_index=shift_index,
                    open_time=open_time,
                )
                self.close_transferred_downtimes(
                    shift_index=shift_index,
                    when=open_time + timedelta(minutes=2),
                )
                if shift_index == 0:
                    self.establish_initial_complexes(
                        shift_index=shift_index,
                        dispatcher=dispatcher,
                        mining_master=mining_master,
                        when=open_time + timedelta(minutes=3),
                    )
                else:
                    self.rotate_daily_complexes(
                        shift_index=shift_index,
                        mining_master=mining_master,
                        when=open_time + timedelta(minutes=3),
                    )
                contexts = self.apply_excavator_settings(
                    shift_index=shift_index,
                    when=open_time + timedelta(minutes=5),
                )
                if shift_index == 0:
                    self.probe_missing_capacity_rule(
                        shift_index=shift_index,
                        when=open_time + timedelta(minutes=10),
                        contexts=contexts,
                    )
                carry_out_trip_id, carry_out_truck_id = (
                    self.execute_trip_cycle(
                        shift_index=shift_index,
                        open_time=open_time,
                        close_time=close_time,
                        contexts=contexts,
                        driver_shifts=driver_shifts,
                        operator_shifts=operator_shifts,
                    )
                )
                excluded_ids = self.start_handoff_downtime(
                    shift_index=shift_index,
                    when=close_time - timedelta(minutes=10),
                    carry_out_truck_id=carry_out_truck_id,
                )
                self.close_equipment_shifts(
                    shift_index=shift_index,
                    close_time=close_time,
                    driver_shifts=driver_shifts,
                    operator_shifts=operator_shifts,
                    carry_out_truck_id=carry_out_truck_id,
                )
                self.record_maintenance_handoff_check(
                    shift_index=shift_index,
                    excluded_equipment_ids=excluded_ids,
                    checked_at=close_time + timedelta(seconds=40),
                )
                self.close_shift_roles(
                    shift_index=shift_index,
                    close_time=close_time,
                    dispatcher=dispatcher,
                    mining_master=mining_master,
                )
                self.verify_shift(
                    shift_index=shift_index,
                    driver_shifts=driver_shifts,
                    operator_shifts=operator_shifts,
                    carry_in_trip_id=carry_in_trip_id,
                    carry_out_trip_id=carry_out_trip_id,
                    started_at=shift_started,
                )
                self.carryover_trip_id = carry_out_trip_id
                self.carryover_truck_id = carry_out_truck_id
                result = self.shift_results[-1]
                print(
                    "SHIFT_OK "
                    f"{shift_index + 1:02d}/14 "
                    f"{result.production_date} {result.shift_type} "
                    f"trips={result.unloaded_trips} "
                    f"range={min(result.trip_counts_by_truck.values())}-"
                    f"{max(result.trip_counts_by_truck.values())} "
                    f"seconds={result.duration_seconds}",
                    flush=True,
                )

        if self.carryover_trip_id or self.carryover_truck_id:
            raise QAError("После 14-й смены остался переходящий рейс.")
        summary = {
            "run_id": self.config.run_id,
            "start_date": self.config.start_date,
            "end_date": self.config.end_date,
            "shifts": [asdict(item) for item in self.shift_results],
            "total_unloaded_trips": sum(
                item.unloaded_trips for item in self.shift_results
            ),
            "total_loaded_trips": sum(
                item.loaded_trips for item in self.shift_results
            ),
            "idempotency": {
                "load": self.idempotency_load_checks,
                "unload": self.idempotency_unload_checks,
                "shift": self.idempotency_shift_checks,
            },
            "eto_to_checks": self.eto_to_checks,
            "duration_seconds": round(time.perf_counter() - run_started, 3),
        }
        write_json(self.config.artifact_dir / "generation_summary.json", summary)
        return self.shift_results


class WeekVerifier:
    """Independent DB, report, HTTP and Excel oracle for the generated week."""

    def __init__(
        self,
        config: RunConfig,
        catalog: ReferenceCatalog,
        onboarding: WeekOnboarding,
        runner: FullWeekRunner,
    ):
        self.config = config
        self.catalog = catalog
        self.onboarding = onboarding
        self.runner = runner
        self.checks: list[dict[str, Any]] = []
        self.excel_dir = config.artifact_dir / "excel"
        self.excel_dir.mkdir(parents=True, exist_ok=True)

    def record(
        self,
        check_id: str,
        passed: bool,
        *,
        actual: Any = None,
        expected: Any = None,
        detail: str = "",
        blocking: bool = True,
    ) -> None:
        row = {
            "check_id": check_id,
            "passed": bool(passed),
            "blocking": blocking,
            "actual": actual,
            "expected": expected,
            "detail": detail,
        }
        self.checks.append(row)
        if blocking and not passed:
            raise QAError(
                f"{check_id}: {detail or 'проверка не пройдена'}; "
                f"actual={actual!r}, expected={expected!r}."
            )

    @staticmethod
    def _workbook_from_response(response):
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise QAError("Для сверки Excel не установлен openpyxl.") from error
        return load_workbook(io.BytesIO(response.content), data_only=True)

    def verify_database(self) -> None:
        self.record(
            "REFERENCE-CAPACITY-COVERAGE",
            not self.catalog.unresolved_capacity_pairs,
            actual={
                "unresolved_pairs": len(
                    self.catalog.unresolved_capacity_pairs
                ),
                "affected_rocks": sorted(
                    {
                        row["rock_type"]
                        for row in self.catalog.unresolved_capacity_pairs
                    }
                ),
                "http_probe": self.runner.reference_capacity_probe,
            },
            expected={
                "unresolved_pairs": 0,
                "affected_rocks": [],
            },
            detail=(
                "Для каждой активной породы и модели самосвала должна "
                "существовать кубатура в правиле или модели."
            ),
            blocking=False,
        )
        expected_week_total = sum(
            self.runner.target_trip_count(truck_index, shift_index)
            for shift_index in range(14)
            for truck_index in range(len(self.catalog.trucks))
        )
        self.record(
            "DB-COMPLETED-TRIPS",
            Trip.objects.filter(status=TripStatus.COMPLETED).count()
            == expected_week_total,
            actual=Trip.objects.filter(status=TripStatus.COMPLETED).count(),
            expected=expected_week_total,
        )
        self.record(
            "DB-OPEN-TRIPS",
            not Trip.objects.filter(status__in=OPEN_TRIP_STATUSES).exists(),
            actual=Trip.objects.filter(status__in=OPEN_TRIP_STATUSES).count(),
            expected=0,
        )
        self.record(
            "DB-OPEN-SHIFTS",
            not EmployeeShift.objects.filter(closed_at__isnull=True).exists(),
            actual=EmployeeShift.objects.filter(closed_at__isnull=True).count(),
            expected=0,
        )

        per_truck = dict(
            Trip.objects.filter(status=TripStatus.COMPLETED)
            .values_list("truck_id")
            .annotate(total=Count("id"))
        )
        expected_per_truck = 280
        self.record(
            "DB-PER-TRUCK-WEEK",
            len(per_truck) == len(self.catalog.trucks)
            and set(per_truck.values()) == {expected_per_truck},
            actual={
                "truck_count": len(per_truck),
                "min": min(per_truck.values(), default=0),
                "max": max(per_truck.values(), default=0),
            },
            expected={
                "truck_count": len(self.catalog.trucks),
                "min": expected_per_truck,
                "max": expected_per_truck,
            },
        )

        invalid_measurements = []
        invalid_links = []
        for trip in (
            Trip.objects.filter(status=TripStatus.COMPLETED)
            .select_related(
                "truck__model",
                "rock_type",
                "loading_shift",
                "unloading_shift",
            )
            .iterator(chunk_size=1000)
        ):
            expected_volume, expected_tonnage = (
                self.catalog.trip_measurements(trip.truck, trip.rock_type)
            )
            if q2(trip.volume_m3) != expected_volume or (
                expected_tonnage is not None
                and q2(trip.tonnage) != expected_tonnage
            ):
                invalid_measurements.append(trip.id)
            if (
                not trip.loading_shift_id
                or not trip.unloading_shift_id
                or not trip.assigned_dump_point_id
                or not trip.actual_dump_point_id
                or trip.actual_dump_point_id != trip.assigned_dump_point_id
            ):
                invalid_links.append(trip.id)
        self.record(
            "DB-REFERENCE-MEASUREMENTS",
            not invalid_measurements,
            actual=invalid_measurements[:20],
            expected=[],
        )
        self.record(
            "DB-TRIP-LINKS",
            not invalid_links,
            actual=invalid_links[:20],
            expected=[],
        )

        duplicate_trip_actions = list(
            TripClientAction.objects.values(
                "action_type",
                "client_action_id",
            )
            .annotate(total=Count("id"))
            .filter(total__gt=1)[:20]
        )
        duplicate_shift_actions = list(
            ShiftClientAction.objects.values(
                "action_type",
                "client_action_id",
            )
            .annotate(total=Count("id"))
            .filter(total__gt=1)[:20]
        )
        self.record(
            "DB-IDEMPOTENCY-LEDGER",
            not duplicate_trip_actions and not duplicate_shift_actions,
            actual={
                "trip": duplicate_trip_actions,
                "shift": duplicate_shift_actions,
            },
            expected={"trip": [], "shift": []},
        )

        corrections = ShiftReadingCorrection.objects.filter(
            new_shift__employee__full_name__startswith=self.config.marker
        )
        correction_metrics = dict(
            corrections.values_list("metric").annotate(total=Count("id"))
        )
        self.record(
            "DB-READING-CORRECTIONS",
            corrections.count() == 26
            and correction_metrics
            == {
                ShiftReadingCorrection.Metric.MILEAGE: 13,
                ShiftReadingCorrection.Metric.ENGINE_HOURS: 13,
            },
            actual={
                "count": corrections.count(),
                "metrics": correction_metrics,
            },
            expected={
                "count": 26,
                "metrics": {"mileage": 13, "engine_hours": 13},
            },
        )

        published_plans = CrewPlan.objects.filter(
            status=CrewPlanStatus.PUBLISHED,
            work_date__range=(self.config.start_date, self.config.end_date),
            role__code__in={"driver", "excavator_operator"},
        )
        expected_slots = 7 * (
            len(self.catalog.trucks) * 2
            + len(self.catalog.excavators) * 2
        )
        self.record(
            "DB-PUBLISHED-PLANS",
            published_plans.count() == 14
            and CrewPlanSlot.objects.filter(plan__in=published_plans).count()
            == expected_slots,
            actual={
                "plans": published_plans.count(),
                "slots": CrewPlanSlot.objects.filter(
                    plan__in=published_plans
                ).count(),
            },
            expected={"plans": 14, "slots": expected_slots},
        )

        for key, downtime_id in self.runner.transferred_downtime_ids.items():
            event = DowntimeEvent.objects.get(pk=downtime_id)
            self.record(
                f"DB-DOWNTIME-HANDOFF-{key.upper()}",
                bool(event.ended_at)
                and event.employee_id
                != (
                    self.onboarding.drivers_by_brigade[
                        self.runner.brigade_for_shift(4)
                    ][event.equipment_id].employee_id
                    if key == "driver"
                    else self.onboarding.operators_by_brigade[
                        self.runner.brigade_for_shift(8)
                    ][event.equipment_id].employee_id
                ),
                actual={
                    "ended_at": event.ended_at,
                    "author_id": event.employee_id,
                },
                expected="закрыт сменщиком без переписывания автора",
            )

        failed_eto = [
            check
            for check in self.runner.eto_to_checks
            if not check["passed"]
        ]
        self.record(
            "BUSINESS-ETO-TO-AFTER-CLOSE",
            not failed_eto,
            actual={
                "failed_shifts": len(failed_eto),
                "missing_equipment_total": sum(
                    row["missing_equipment_count"] for row in failed_eto
                ),
            },
            expected={
                "failed_shifts": 0,
                "missing_equipment_total": 0,
            },
            detail=(
                "Штатное закрытие смены должно автоматически выставлять "
                "ЕТО или ТО."
            ),
            blocking=False,
        )

    def verify_reports_and_excel(self) -> None:
        from reports.shift_analytics import (
            build_excavator_dynamics,
            build_shift_analytics,
        )
        from reports.views import build_customer_daily_report

        dispatcher = self.onboarding.shift_roles_by_brigade["dispatcher"][
            NIGHT_BRIGADES[-1]
        ]
        manager = self.onboarding.manager
        if not dispatcher.client or not manager or not manager.client:
            raise QAError("Нет авторизованных клиентов для отчётов.")

        week_loaded_total = 0
        week_unloaded_total = 0
        for shift_index, result in enumerate(self.runner.shift_results):
            selected_date = self.runner.date_for_shift(shift_index)
            shift_type = self.runner.shift_type_for_index(shift_index)
            analytics = build_shift_analytics(selected_date, shift_type)
            totals = analytics["totals"]
            self.record(
                f"REPORT-SHIFT-{shift_index:02d}",
                totals["loaded_trip_count"] == result.loaded_trips
                and totals["unloaded_trip_count"] == result.unloaded_trips
                and len(analytics["truck_rows"]) == len(self.catalog.trucks)
                and sorted(
                    row["trip_count"] for row in analytics["truck_rows"]
                )
                == sorted(result.trip_counts_by_truck.values()),
                actual={
                    "loaded": totals["loaded_trip_count"],
                    "unloaded": totals["unloaded_trip_count"],
                    "truck_rows": len(analytics["truck_rows"]),
                },
                expected={
                    "loaded": result.loaded_trips,
                    "unloaded": result.unloaded_trips,
                    "truck_rows": len(self.catalog.trucks),
                },
            )
            week_loaded_total += totals["loaded_trip_count"]
            week_unloaded_total += totals["unloaded_trip_count"]

            query = (
                f"?date={selected_date.isoformat()}&shift_type={shift_type}"
            )
            for path in (
                "/dispatcher/mining-volumes/",
                "/dispatcher/transport/",
                "/dispatcher/reports/",
                "/reports/shift-analytics/",
            ):
                dispatcher.client.get(
                    path + query,
                    label=f"HTTP отчёт {path} смена {shift_index}",
                )

            excel_response = dispatcher.client.get(
                "/reports/shift-analytics/export/" + query,
                label=f"Excel сменной аналитики {shift_index}",
            )
            excel_path = (
                self.excel_dir
                / f"shift_{shift_index:02d}_{selected_date}_{shift_type}.xlsx"
            )
            excel_path.write_bytes(excel_response.content)
            workbook = self._workbook_from_response(excel_response)
            sheet = workbook["Сменная аналитика"]
            excel_loaded = int(sheet["B7"].value)
            excel_unloaded = int(sheet["B8"].value)
            self.record(
                f"EXCEL-SHIFT-{shift_index:02d}",
                excel_loaded == result.loaded_trips
                and excel_unloaded == result.unloaded_trips,
                actual={
                    "loaded": excel_loaded,
                    "unloaded": excel_unloaded,
                },
                expected={
                    "loaded": result.loaded_trips,
                    "unloaded": result.unloaded_trips,
                },
            )

        self.record(
            "REPORT-WEEK-TOTALS",
            week_loaded_total == 14840 and week_unloaded_total == 14840,
            actual={
                "loaded": week_loaded_total,
                "unloaded": week_unloaded_total,
            },
            expected={"loaded": 14840, "unloaded": 14840},
        )

        for day_index in range(7):
            selected_date = self.config.start_date + timedelta(days=day_index)
            daily = build_customer_daily_report(selected_date)
            expected_day = self.runner.shift_results[day_index * 2].loaded_trips
            expected_night = self.runner.shift_results[
                day_index * 2 + 1
            ].loaded_trips
            self.record(
                f"REPORT-DAILY-{day_index}",
                daily["day_trip_count"] == expected_day
                and daily["night_trip_count"] == expected_night
                and daily["total_trip_count"] == expected_day + expected_night,
                actual={
                    "day": daily["day_trip_count"],
                    "night": daily["night_trip_count"],
                    "total": daily["total_trip_count"],
                },
                expected={
                    "day": expected_day,
                    "night": expected_night,
                    "total": expected_day + expected_night,
                },
            )
            query = f"?date={selected_date.isoformat()}"
            manager.client.get(
                "/reports/customer-daily/" + query,
                label=f"Суточный отчёт {selected_date}",
            )
            manager.client.get(
                "/reports/management/" + query,
                label=f"Отчёт руководства {selected_date}",
            )
            excel_response = manager.client.get(
                "/reports/customer-daily/export/" + query,
                label=f"Excel суточного отчёта {selected_date}",
            )
            excel_path = (
                self.excel_dir / f"daily_{selected_date.isoformat()}.xlsx"
            )
            excel_path.write_bytes(excel_response.content)
            workbook = self._workbook_from_response(excel_response)
            sheet = workbook["Суточный отчет"]
            self.record(
                f"EXCEL-DAILY-{day_index}",
                int(sheet["B10"].value) == expected_day
                and int(sheet["C10"].value) == expected_night
                and int(sheet["D10"].value)
                == expected_day + expected_night,
                actual={
                    "day": sheet["B10"].value,
                    "night": sheet["C10"].value,
                    "total": sheet["D10"].value,
                },
                expected={
                    "day": expected_day,
                    "night": expected_night,
                    "total": expected_day + expected_night,
                },
            )

        dynamics = build_excavator_dynamics(
            self.config.start_date,
            self.config.end_date,
            granularity="day",
            excavator_ids=[item.id for item in self.catalog.excavators],
            shift_type="day",
            chart_mode="trips",
        )
        self.record(
            "REPORT-MANAGEMENT-DYNAMICS",
            bool(dynamics),
            actual=type(dynamics).__name__,
            expected="непустой результат",
        )
        dynamics_query = (
            f"?date_from={self.config.start_date.isoformat()}"
            f"&date_to={self.config.end_date.isoformat()}"
            "&granularity=day&shift_type=day&chart_mode=trips"
        )
        manager.client.get(
            "/reports/management/dynamics/" + dynamics_query,
            label="Недельная динамика руководства",
        )

        volume_query = (
            f"?date_from={self.config.start_date.isoformat()}"
            f"&date_to={self.config.end_date.isoformat()}&group_by=truck"
        )
        manager.client.get(
            "/reports/volume/" + volume_query,
            label="Недельный отчёт по объёмам",
        )
        volume_excel_response = manager.client.get(
            "/reports/volume/export/" + volume_query,
            label="Excel недельного отчёта по объёмам",
        )
        volume_path = self.excel_dir / "week_volume_by_truck.xlsx"
        volume_path.write_bytes(volume_excel_response.content)
        workbook = self._workbook_from_response(volume_excel_response)
        sheet = workbook["Объемы"]
        truck_rows = []
        for row in sheet.iter_rows(values_only=True):
            if len(row) >= 4 and row[0] not in {
                None,
                "Самосвал",
                "Итого",
            } and isinstance(row[3], (int, float)):
                truck_rows.append(row)
        self.record(
            "EXCEL-WEEK-BY-TRUCK",
            len(truck_rows) == len(self.catalog.trucks)
            and {int(row[3]) for row in truck_rows} == {280},
            actual={
                "rows": len(truck_rows),
                "min_trips": min(
                    (int(row[3]) for row in truck_rows),
                    default=0,
                ),
                "max_trips": max(
                    (int(row[3]) for row in truck_rows),
                    default=0,
                ),
            },
            expected={
                "rows": len(self.catalog.trucks),
                "min_trips": 280,
                "max_trips": 280,
            },
        )

    @staticmethod
    def _diagnostic_payload_violations(payload: Any) -> list[str]:
        violations = []

        def walk(value: Any, path: str = "$") -> None:
            if isinstance(value, dict):
                for key, item in value.items():
                    normalized = str(key).strip().lower()
                    item_path = f"{path}.{key}"
                    if normalized in {"score", "place", "weight"}:
                        violations.append(item_path)
                    if normalized == "official" and item is not False:
                        violations.append(item_path)
                    if normalized.startswith("official_") and item is not False:
                        violations.append(item_path)
                    walk(item, item_path)
            elif isinstance(value, (list, tuple)):
                for index, item in enumerate(value):
                    walk(item, f"{path}[{index}]")

        walk(payload)
        return violations

    def verify_driver_watch_and_passports(self) -> None:
        expected_shift_count = len(self.catalog.trucks) * 14
        watch_period = self.onboarding.watch_period
        driver_shifts = EmployeeShift.objects.filter(
            employee__full_name__startswith=self.config.marker,
            workplace_code="driver",
            closed_at__isnull=False,
        )
        linked_shift_count = driver_shifts.filter(
            watch_period=watch_period,
        ).count()
        self.record(
            "KPI-DRIVER-SHIFT-WATCH-LINKAGE",
            driver_shifts.count() == expected_shift_count
            and linked_shift_count == expected_shift_count,
            actual={
                "closed_driver_shifts": driver_shifts.count(),
                "linked_to_selected_watch": linked_shift_count,
            },
            expected={
                "closed_driver_shifts": expected_shift_count,
                "linked_to_selected_watch": expected_shift_count,
            },
        )

        linkage_audit = build_driver_watch_linkage_audit(watch_period)
        self.record(
            "KPI-WATCH-LINKAGE-AUDIT",
            linkage_audit["candidate_closed_shift_count"]
            == expected_shift_count
            and linkage_audit["linked_to_selected_watch_count"]
            == expected_shift_count
            and linkage_audit["unlinked_shift_count"] == 0
            and linkage_audit["linked_to_other_watch_count"] == 0
            and linkage_audit["selected_watch_outside_period_count"] == 0
            and linkage_audit["linkage_ready"],
            actual=linkage_audit,
            expected={
                "candidate_closed_shift_count": expected_shift_count,
                "linked_to_selected_watch_count": expected_shift_count,
                "unlinked_shift_count": 0,
                "linked_to_other_watch_count": 0,
                "selected_watch_outside_period_count": 0,
                "linkage_ready": True,
            },
        )

        requests = DriverShiftPassportCaptureRequest.objects.filter(
            shift__in=driver_shifts,
        )
        request_statuses = dict(
            requests.values_list("status")
            .annotate(total=Count("id"))
        )
        completed_request_count = requests.filter(
            status=DriverShiftPassportRequestStatus.COMPLETED,
            snapshot__isnull=False,
        ).count()
        self.record(
            "KPI-PASSPORT-OUTBOX",
            requests.count() == expected_shift_count
            and completed_request_count == expected_shift_count
            and request_statuses
            == {
                DriverShiftPassportRequestStatus.COMPLETED:
                    expected_shift_count,
            },
            actual={
                "total": requests.count(),
                "completed_with_snapshot": completed_request_count,
                "statuses": request_statuses,
            },
            expected={
                "total": expected_shift_count,
                "completed_with_snapshot": expected_shift_count,
                "statuses": {
                    DriverShiftPassportRequestStatus.COMPLETED:
                        expected_shift_count,
                },
            },
        )

        snapshots = DriverShiftPassportSnapshot.objects.filter(
            shift__in=driver_shifts,
        )
        snapshot_revision_audit = list(
            snapshots.values("shift_id")
            .annotate(total=Count("id"), max_revision=Max("revision"))
            .exclude(total=1, max_revision=1)[:20]
        )
        self.record(
            "KPI-PASSPORT-REVISIONS",
            snapshots.count() == expected_shift_count
            and not snapshot_revision_audit,
            actual={
                "snapshot_count": snapshots.count(),
                "invalid_shift_revisions": snapshot_revision_audit,
            },
            expected={
                "snapshot_count": expected_shift_count,
                "invalid_shift_revisions": [],
            },
        )

        total_completed_trips = 0
        total_volume = Decimal("0")
        total_tonnage = Decimal("0")
        total_m3_km_missing = 0
        total_t_km_missing = 0
        usable_shift_count = 0
        official_payload_violations = []
        source_manifest_violations = []
        fingerprint_violations = []
        capture_violations = []
        quality_flag_counts: Counter[str] = Counter()
        coverage_values = []

        for row in snapshots.values(
            "shift_id",
            "revision",
            "source_fingerprint",
            "payload_fingerprint",
            "payload",
            "trigger",
            "captured_late",
            "captured_by_id",
        ).iterator(chunk_size=50):
            payload = row["payload"]
            passport = payload.get("passport", {})
            production = passport.get("production", {})
            quality = passport.get("quality", {})
            source_watch = (
                payload.get("source_manifest", {})
                .get("shift", {})
                .get("watch_period")
            )
            source_composition = (
                source_watch.get("watch_composition")
                if source_watch
                else None
            )
            if (
                not source_watch
                or source_watch.get("id") != watch_period.id
                or not source_composition
                or source_composition.get("id")
                != self.onboarding.watch_composition.id
            ):
                source_manifest_violations.append(row["shift_id"])

            violations = self._diagnostic_payload_violations(payload)
            if violations:
                official_payload_violations.append({
                    "shift_id": row["shift_id"],
                    "paths": violations[:10],
                })
            if (
                payload.get("official") is not False
                or quality.get("official_rating_eligible") is not False
            ):
                official_payload_violations.append({
                    "shift_id": row["shift_id"],
                    "paths": ["$.official_contract"],
                })
            if (
                len(row["source_fingerprint"] or "") != 64
                or len(row["payload_fingerprint"] or "") != 64
            ):
                fingerprint_violations.append(row["shift_id"])
            if (
                row["trigger"] != "driver_close"
                or row["captured_late"]
                or not row["captured_by_id"]
            ):
                capture_violations.append(row["shift_id"])

            total_completed_trips += int(
                production.get("completed_trip_count", 0)
            )
            volume = production.get("volume_m3", {})
            tonnage = production.get("tonnage_t", {})
            m3_km = production.get("m3_km", {})
            t_km = production.get("t_km", {})
            total_volume += Decimal(str(volume.get("known_value") or "0"))
            total_tonnage += Decimal(
                str(tonnage.get("known_value") or "0")
            )
            total_m3_km_missing += int(
                m3_km.get("missing_trip_count", 0)
            )
            total_t_km_missing += int(
                t_km.get("missing_trip_count", 0)
            )
            usable_shift_count += int(
                bool(quality.get("data_usable_for_formula_review"))
            )
            quality_flag_counts.update(quality.get("flags", []))
            coverage_values.append(
                Decimal(str(quality.get("coverage_percent") or "0"))
            )

        trip_totals = Trip.objects.filter(
            status=TripStatus.COMPLETED,
        ).aggregate(
            volume=Sum("volume_m3"),
            tonnage=Sum("tonnage"),
        )
        expected_trip_total = 14840
        expected_volume = Decimal(str(trip_totals["volume"] or "0"))
        expected_tonnage = Decimal(str(trip_totals["tonnage"] or "0"))
        self.record(
            "KPI-PASSPORT-PRODUCTION-TOTALS",
            total_completed_trips == expected_trip_total
            and total_volume == expected_volume
            and total_tonnage == expected_tonnage,
            actual={
                "completed_trips": total_completed_trips,
                "volume_m3": total_volume,
                "tonnage_t": total_tonnage,
            },
            expected={
                "completed_trips": expected_trip_total,
                "volume_m3": expected_volume,
                "tonnage_t": expected_tonnage,
            },
        )
        self.record(
            "KPI-PASSPORT-DISTANCE-HONEST-HOLD",
            total_m3_km_missing == expected_trip_total
            and total_t_km_missing == expected_trip_total
            and usable_shift_count == 0,
            actual={
                "m3_km_missing_trip_count": total_m3_km_missing,
                "t_km_missing_trip_count": total_t_km_missing,
                "formula_usable_shift_count": usable_shift_count,
            },
            expected={
                "m3_km_missing_trip_count": expected_trip_total,
                "t_km_missing_trip_count": expected_trip_total,
                "formula_usable_shift_count": 0,
            },
            detail=(
                "В справочниках нет подтверждённого плеча маршрута; "
                "симулятор не должен выдумывать расстояния или объявлять "
                "формулу готовой."
            ),
        )
        self.record(
            "KPI-PASSPORT-DIAGNOSTIC-ONLY",
            not official_payload_violations
            and not source_manifest_violations
            and not fingerprint_violations
            and not capture_violations,
            actual={
                "official_payload_violations":
                    official_payload_violations[:20],
                "source_manifest_violations":
                    source_manifest_violations[:20],
                "fingerprint_violations": fingerprint_violations[:20],
                "capture_violations": capture_violations[:20],
            },
            expected={
                "official_payload_violations": [],
                "source_manifest_violations": [],
                "fingerprint_violations": [],
                "capture_violations": [],
            },
        )

        observation = build_driver_watch_observation(watch_period)
        observation_summary = observation["summary"]
        self.record(
            "KPI-WATCH-OBSERVATION",
            observation.get("official_rating_eligible") is False
            and observation_summary["closed_shift_count"]
            == expected_shift_count
            and observation_summary["usable_shift_count"] == 0
            and observation_summary["withheld_shift_count"]
            == expected_shift_count
            and observation_summary["data_ready_for_formula_review"] is False,
            actual={
                "row_count": observation["row_count"],
                **observation_summary,
                "official_rating_eligible":
                    observation.get("official_rating_eligible"),
            },
            expected={
                "closed_shift_count": expected_shift_count,
                "usable_shift_count": 0,
                "withheld_shift_count": expected_shift_count,
                "data_ready_for_formula_review": False,
                "official_rating_eligible": False,
            },
        )

        passport_audit = {
            "watch_period": {
                "id": watch_period.id,
                "name": watch_period.name,
                "watch_composition_id":
                    self.onboarding.watch_composition.id,
                "starts_on": watch_period.starts_on,
                "ends_on": watch_period.ends_on,
            },
            "linkage_audit": linkage_audit,
            "driver_shift_count": driver_shifts.count(),
            "capture_request_statuses": request_statuses,
            "snapshot_count": snapshots.count(),
            "production_totals": {
                "completed_trips": total_completed_trips,
                "volume_m3": total_volume,
                "tonnage_t": total_tonnage,
            },
            "distance_completeness": {
                "m3_km_missing_trip_count": total_m3_km_missing,
                "t_km_missing_trip_count": total_t_km_missing,
                "reason":
                    "Подтверждённое плечо маршрута отсутствует в справочниках.",
            },
            "quality": {
                "usable_shift_count": usable_shift_count,
                "withheld_shift_count":
                    expected_shift_count - usable_shift_count,
                "flag_counts": dict(sorted(quality_flag_counts.items())),
                "coverage_percent": {
                    "min": min(coverage_values, default=Decimal("0")),
                    "average": (
                        sum(coverage_values, Decimal("0"))
                        / len(coverage_values)
                        if coverage_values
                        else Decimal("0")
                    ).quantize(Decimal("0.01")),
                    "max": max(coverage_values, default=Decimal("0")),
                },
            },
        }
        write_json(
            self.config.artifact_dir / "driver_passport_audit.json",
            passport_audit,
        )
        write_json(
            self.config.artifact_dir / "driver_watch_observation.json",
            observation,
        )

    def run(self) -> dict[str, Any]:
        started_at = time.perf_counter()
        self.verify_database()
        self.verify_driver_watch_and_passports()
        self.verify_reports_and_excel()
        failures = [check for check in self.checks if not check["passed"]]
        blocking_failures = [
            check
            for check in failures
            if check["blocking"]
        ]
        report = {
            "run_id": self.config.run_id,
            "checks": self.checks,
            "passed": len(self.checks) - len(failures),
            "failed": len(failures),
            "blocking_failed": len(blocking_failures),
            "non_blocking_failed": len(failures) - len(blocking_failures),
            "duration_seconds": round(time.perf_counter() - started_at, 3),
        }
        write_json(self.config.artifact_dir / "verification_report.json", report)
        return report


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Прогнать 7 дневных и 7 ночных смен в отдельной локальной "
            "PostgreSQL-БД."
        )
    )
    parser.add_argument(
        "--start-date",
        type=date.fromisoformat,
        default=DEFAULT_START_DATE,
    )
    parser.add_argument(
        "--run-id",
        default=DEFAULT_RUN_ID,
    )
    parser.add_argument(
        "--marker",
        default=DEFAULT_MARKER,
    )
    parser.add_argument(
        "--artifact-dir",
        type=Path,
        default=DEFAULT_ARTIFACT_DIR,
    )
    return parser.parse_args(argv)


def build_run_metadata(
    config: RunConfig,
    catalog: ReferenceCatalog,
    onboarding: WeekOnboarding,
    runner: FullWeekRunner,
    verification: dict[str, Any],
) -> dict[str, Any]:
    artifact_hashes = {}
    for path in sorted(
        item
        for item in config.artifact_dir.rglob("*")
        if item.is_file()
    ):
        artifact_hashes[str(path.relative_to(config.artifact_dir))] = (
            sha256_file(path)
        )
    return {
        "run_id": config.run_id,
        "marker": config.marker,
        "database": {
            "engine": settings.DATABASES["default"]["ENGINE"],
            "name": settings.DATABASES["default"]["NAME"],
            "host": settings.DATABASES["default"]["HOST"],
            "port": settings.DATABASES["default"]["PORT"],
            "postgresql_version": connection.pg_version,
        },
        "production_calendar": {
            "timezone": str(BUSINESS_TIME_ZONE),
            "start_date": config.start_date,
            "end_date": config.end_date,
            "day_shift": "07:00-19:00",
            "night_shift": "19:00-07:00",
        },
        "references": {
            "trucks": [
                {
                    "id": item.id,
                    "garage_number": item.garage_number,
                    "model": item.model.name,
                    "body_volume_m3": item.model.body_volume_m3,
                    "fuel_capacity_limit_l": (
                        item.model.fuel_capacity_limit_l
                    ),
                }
                for item in catalog.trucks
            ],
            "excavators": [
                {
                    "id": item.id,
                    "garage_number": item.garage_number,
                    "model": item.model.name,
                    "fuel_capacity_limit_l": (
                        item.model.fuel_capacity_limit_l
                    ),
                }
                for item in catalog.excavators
            ],
            "rocks": [
                {
                    "id": item.id,
                    "name": item.name,
                    "density": item.density,
                    "operational_for_all_trucks": (
                        item in catalog.operational_rocks
                    ),
                }
                for item in catalog.rocks
            ],
            "unresolved_capacity_pairs": (
                catalog.unresolved_capacity_pairs
            ),
            "dump_points": [
                {"id": item.id, "name": item.name}
                for item in catalog.dump_points
            ],
        },
        "staff": {
            role_code: len(members)
            for role_code, members in onboarding.by_role.items()
        },
        "trip_targets": {
            "per_shift_per_truck_min": 17,
            "per_shift_per_truck_max": 23,
            "per_truck_week": 280,
            "week_total": 14840,
        },
        "generation": {
            "shifts": len(runner.shift_results),
            "loaded": sum(item.loaded_trips for item in runner.shift_results),
            "unloaded": sum(
                item.unloaded_trips for item in runner.shift_results
            ),
        },
        "verification": verification,
        "artifact_sha256": artifact_hashes,
    }


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    config = RunConfig(
        run_id=args.run_id,
        marker=args.marker,
        start_date=args.start_date,
        artifact_dir=args.artifact_dir.resolve(),
    )
    started_at = time.perf_counter()
    try:
        verify_isolated_database(config, require_empty=True)
        catalog = ReferenceCatalog(config)
        write_json(
            config.artifact_dir / "preflight.json",
            {
                "run_id": config.run_id,
                "marker": config.marker,
                "reference_counts": {
                    "trucks": len(catalog.trucks),
                    "excavators": len(catalog.excavators),
                    "rocks": len(catalog.rocks),
                    "operational_rocks": len(catalog.operational_rocks),
                    "dump_points": len(catalog.dump_points),
                    "capacity_rules": TruckCapacityRule.objects.count(),
                    "unresolved_capacity_pairs": len(
                        catalog.unresolved_capacity_pairs
                    ),
                },
                "timezone": str(BUSINESS_TIME_ZONE),
                "start_date": config.start_date,
                "end_date": config.end_date,
            },
        )
        print(
            "PREFLIGHT_OK "
            f"trucks={len(catalog.trucks)} "
            f"excavators={len(catalog.excavators)} "
            f"rocks={len(catalog.rocks)} "
            f"dumps={len(catalog.dump_points)}",
            flush=True,
        )
        onboarding = WeekOnboarding(config, catalog).run()
        print(
            f"ONBOARDING_OK staff={len(onboarding.staff)}",
            flush=True,
        )
        runner = FullWeekRunner(config, catalog, onboarding)
        runner.run()
        verification = WeekVerifier(
            config,
            catalog,
            onboarding,
            runner,
        ).run()
        metadata = build_run_metadata(
            config,
            catalog,
            onboarding,
            runner,
            verification,
        )
        metadata["total_duration_seconds"] = round(
            time.perf_counter() - started_at,
            3,
        )
        write_json(config.artifact_dir / "run_metadata.json", metadata)
        print(
            "FULL_WEEK_QA_COMPLETE "
            f"trips={metadata['generation']['unloaded']} "
            f"checks_passed={verification['passed']} "
            f"checks_failed={verification['failed']} "
            f"seconds={metadata['total_duration_seconds']}",
            flush=True,
        )
        return 0 if not verification["blocking_failed"] else 1
    except Exception as error:
        failure = {
            "run_id": config.run_id,
            "error_type": type(error).__name__,
            "error": str(error),
            "traceback": traceback.format_exc(),
            "duration_seconds": round(time.perf_counter() - started_at, 3),
        }
        write_json(config.artifact_dir / "failure.json", failure)
        print(
            f"FULL_WEEK_QA_FAILED {type(error).__name__}: {error}",
            file=sys.stderr,
            flush=True,
        )
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
